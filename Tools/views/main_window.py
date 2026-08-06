"""
views/main_window.py — BOMToolApp: cửa sổ chính và toàn bộ tab UI.
"""
import logging
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk
import re
import os
import sys
import threading
import datetime
import unicodedata
import io
from openpyxl import load_workbook
try:
    import msoffcrypto
    _HAS_MSOFFCRYPTO = True
except ImportError:
    _HAS_MSOFFCRYPTO = False
try:
    import tksheet
    _HAS_TKSHEET = True
except ImportError:
    _HAS_TKSHEET = False

import json
_log = logging.getLogger(__name__)

from services.utils import (
    APP_VERSION, TAB_IMPORT, TAB_THDM, TAB_CATALOG,
    BASE_DIR, CONFIG_DIR, MAPPING_FILE, DB_CONFIG_FILE, SETTINGS_FILE,
    DEFAULT_CREATOR_USER_ID, C, THEMES,
    _norm_vn, _nan_str, _MAPPING_ENG_COLS, _SECTION_ENG_COLS,
    guess_col_align, resource_path,
    FONT_BODY, FONT_BODY_B, FONT_SMALL, FONT_SMALL_B, FONT_SMALL_I,
    FONT_MD, FONT_MD_B, FONT_LABEL, FONT_LABEL_N, FONT_TITLE, FONT_HERO, FONT_ICON,
    PAD_XS, PAD_SM, PAD_MD, PAD_LG,
)
from services.mapping_loader import (
    load_mapping, build_reverse_map, match_col_to_sql,
    build_meta_keys_from_mapping, _load_section_rows, _load_config,
)
from services.bom_parser import (
    parse_bom_file, _parse_sheet, _parse_section_excel_rows,
    _is_roman_numeral, _resolve_row_mapping, _run_row_sp_hooks,
    SECTION_STT_PATTERN, NUMERIC_STT_PATTERN, _ROMAN_SIMPLE_RE,
    _thdm_parse_thvt_sheet, _thdm_load_bom_qty_dict,
    _thdm_apply_row_filter, _thdm_get_detail_col_defs,
    _thdm_find_thvt_sheet, _thdm_open_workbook,
    _thdm_find_header_row, _norm_muc_key, _thdm_expand_muc_rows,
    _load_sheet_config, _detect_sheet_type,
    _THDM_COL_ALIASES, _THDM_HEADER_ANCHORS_FALLBACK,
    _extract_meta, _build_excel_col_map, _merge_meta_rows,
    _build_headers, _norm_col, _find_header_row,
    _resolve_formula, _is_data_row,
    _is_encrypted_excel, _decrypt_excel, _ask_excel_password,
)
from services.validators import (
    validate_layer1, count_errors,
    REQUIRED_HEADER_FIELDS, NUMERIC_RE,
)
from views.widgets import (
    CLabel, CButton, _SearchCombo, Tooltip, SheetTable,
)
class BOMToolApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # An ngay - tranh flash den khi khoi dong
        self.title(f"BOHO IMPORT BOM/THDM v{APP_VERSION}")
        self.geometry("1360x820")
        # Set app icon
        try:
            ico_path = resource_path("icon.ico")
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
        except Exception:
            pass

        self.tables        = {}
        self.global_meta   = {}
        self.val_errors    = {}
        self.mapping       = load_mapping()
        self._current_file = ""
        self._batch_files  = []
        self._loading_dlg  = None
        self._last_import_bom_id = None   # BOM Id của lần import gần nhất (cho Hoàn tác)
        self._last_import_fname  = ''
        self.db_cfg        = self._load_db_config()
        self._skipped_sheets = []
        self._current_creator_user_id = None   # UserId (B00UserList.Id) được chọn ở tab BOM
        self._thdm_creator_user_id    = None   # UserId riêng cho tab THDM
        self._thdm_creator_employee_id = None  # EmployeeId (B20Employee.Id) riêng cho tab THDM

        # Theme hiện tại — đọc từ settings trước khi build UI
        self._current_theme = "Dark"
        self._current_scaling_pct = "100%"
        _settings = self._load_settings_data()
        _init_theme = _settings.get("theme", "Dark")
        _init_scale = _settings.get("scaling", "100%")
        self._current_theme = _init_theme
        self._current_scaling_pct = _init_scale
        try:
            ctk.set_widget_scaling(int(_init_scale.replace("%", "")) / 100)
        except Exception:
            ctk.set_widget_scaling(1.0)
        self._setup_styles()
        self._build_ui()

        # Hiện startup overlay, show window, rồi maximize sau 100ms
        self._show_startup_overlay()
        self.deiconify()
        self.update()
        self.after(100, lambda: self.wm_state('zoomed'))

    def _setup_styles(self):
        """TTK styles — dùng THEMES để đồng bộ Dark/Light."""
        mode = ctk.get_appearance_mode()  # "Dark" or "Light"
        t = THEMES.get(mode, THEMES["Dark"])

        s = ttk.Style()
        s.theme_use("clam")
        s.configure("BOM.Treeview",
            rowheight=28, font=FONT_SMALL,
            background=t["tv_sep"], foreground=t["tv_text"],
            fieldbackground=t["tv_sep"], borderwidth=0, indent=12)
        s.configure("BOM.Treeview.Heading",
            font=FONT_SMALL_B,
            background=t["tv_heading_bg"], foreground=t["tv_heading_fg"],
            borderwidth=0, relief="flat",
            padding=(8, 6))
        s.configure("BOM.Treeview.Cell",
            padding=2, borderwidth=1, relief="solid")
        s.map("BOM.Treeview",
            background=[("selected", t["tv_sel"])],
            foreground=[("selected", t["tv_text"])])
        s.map("BOM.Treeview.Heading",
            background=[("active", t["tv_heading_act"])])

        self._tv_bg    = t["bg_main"]
        self._tv_panel = t["bg_card"]
        self._tv_text  = t["tv_text"]
        self._tv_sel   = t["tv_sel"]

        # TCombobox style (đồng bộ Dark/Light)
        s.configure("TCombobox",
            fieldbackground=t["bg_card"],
            background=t["bg_card"],
            foreground=t["text_main"],
            selectbackground=t["sel_bg"],
            selectforeground="#FFFFFF",
            bordercolor=t["border"],
            darkcolor=t["bg_card"],
            lightcolor=t["bg_card"],
            arrowcolor=t["text_main"],
            arrowsize=15,
            padding=(6, 4))
        s.map("TCombobox",
            fieldbackground=[("readonly", t["bg_card"]), ("disabled", t["bg_panel"])],
            foreground=[("disabled", t["text_muted"])],
            background=[("active", t["bg_panel"])],
            bordercolor=[("focus", t["sel_bg"]), ("active", t["border"])])
        self.option_add("*TCombobox*Listbox.background", t["bg_card"])
        self.option_add("*TCombobox*Listbox.foreground", t["text_main"])
        self.option_add("*TCombobox*Listbox.selectBackground", t["sel_bg"])
        self.option_add("*TCombobox*Listbox.selectForeground", "#FFFFFF")
        self.option_add("*TCombobox*Listbox.relief", "flat")
        self.option_add("*TCombobox*Listbox.font", "{Segoe UI} 12")

    # ── CTk appearance callbacks (từ sidebar OptionMenu) ──────────────────────
    def _setup_scrollbar_style(self):
        """Scrollbar mỏng, phẳng kiểu VS Code."""
        mode = ctk.get_appearance_mode()
        is_dark = mode == "Dark"
        sb_bg     = "#3E3E42" if is_dark else "#C8C8C8"
        sb_trough = "#1E1E1E" if is_dark else "#F0F0F0"
        sb_active = "#686868" if is_dark else "#A0A0A0"
        sb_arrow  = "#858585" if is_dark else "#787878"
        s = ttk.Style()
        for orient in ("Vertical", "Horizontal"):
            name = f"{orient}.TScrollbar"
            s.configure(name,
                gripcount=0,
                relief="flat",
                background=sb_bg,
                darkcolor=sb_trough,
                lightcolor=sb_trough,
                troughcolor=sb_trough,
                bordercolor=sb_trough,
                arrowcolor=sb_arrow,
                arrowsize=12,
                width=10 if orient == "Vertical" else 10)
            s.map(name,
                background=[("active", sb_active), ("pressed", sb_active)],
                arrowcolor=[("active", "#CCCCCC" if is_dark else "#555555")])
        # Slim scrollbar variant (no arrows) for main content
        for orient in ("Vertical", "Horizontal"):
            name = f"Slim.{orient}.TScrollbar"
            s.configure(name,
                gripcount=0,
                relief="flat",
                background=sb_bg,
                darkcolor=sb_trough,
                lightcolor=sb_trough,
                troughcolor=sb_trough,
                bordercolor=sb_trough,
                arrowcolor=sb_trough,   # arrows same color = hidden
                arrowsize=0,
                width=8 if orient == "Vertical" else 8)
            s.map(name,
                background=[("active", sb_active)])

    # ── Settings & Theme ────────────────────────────────────────────────────────

    def _load_settings_data(self) -> dict:
        """Đọc settings.json; trả về {} nếu không tồn tại hoặc lỗi."""
        try:
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_settings(self):
        """Ghi theme + scaling vào settings.json."""
        try:
            data = {
                "theme":   getattr(self, "_current_theme",       "Dark"),
                "scaling": getattr(self, "_current_scaling_pct", "100%"),
            }
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _recolor_tk_widgets(self, widget, color_map: dict, fg_map: dict = None):
        """Đổi bg (và optionally fg) của tk.Frame/Label/Listbox theo color_map/fg_map."""
        for w in widget.winfo_children():
            try:
                cls = w.winfo_class()
                if cls in ("Frame", "Label"):
                    bg = w.cget("bg").lower()
                    if bg in color_map:
                        w.configure(bg=color_map[bg])
                    if cls == "Label" and fg_map:
                        fg = w.cget("fg").lower()
                        if fg in fg_map:
                            w.configure(fg=fg_map[fg])
                elif cls == "Listbox":
                    bg = w.cget("background").lower()
                    if bg in color_map:
                        w.configure(background=color_map[bg])
            except Exception:
                pass
            try:
                self._recolor_tk_widgets(w, color_map, fg_map)
            except Exception:
                pass

    @property
    def _text_fg(self) -> str:
        """Màu chữ chính theo theme hiện tại — dùng thay C["text"] trong .config(fg=)."""
        return THEMES.get(ctk.get_appearance_mode(), THEMES["Dark"])["text_main"]

    @property
    def _muted_fg(self) -> str:
        """Màu chữ phụ theo theme hiện tại — dùng thay C["muted"] trong .config(fg=)."""
        return THEMES.get(ctk.get_appearance_mode(), THEMES["Dark"])["text_muted"]

    def apply_theme(self, mode: str):
        """Chuyển theme toàn cục. mode: 'Dark' | 'Light' | 'System'."""
        # 1. Cập nhật CTk appearance
        ctk.set_appearance_mode(mode)
        actual = ctk.get_appearance_mode()
        # Robust resolve: CTk 6.x có thể trả về "System" thay vì "Dark"/"Light"
        if actual not in ("Dark", "Light"):
            try:
                import winreg
                _k = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                    r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
                _v, _ = winreg.QueryValueEx(_k, "AppsUseLightTheme")
                winreg.CloseKey(_k)
                actual = "Light" if _v else "Dark"
            except Exception:
                actual = "Dark"
        t = THEMES.get(actual, THEMES["Dark"])
        self._current_theme = mode

        # 2. Re-run TTK styles (Treeview + Combobox + Scrollbar)
        self._setup_styles()
        self._setup_scrollbar_style()

        # 3. Tất cả SheetTable instances — màu nền + tag rows
        def _update_sheet_tags(obj: SheetTable):
            obj.apply_theme(t)
            obj.tag_configure("sql_names",
                background=t.get("field_row_bg"),
                foreground=t.get("field_row_fg"))
            obj.tag_configure("row_normal", foreground=t.get("row_normal_fg"))
            obj.tag_configure("row_alt",    foreground=t.get("row_alt_fg"))
            obj.tag_configure("err_row",
                background=t.get("badge_err_bg"),
                foreground=t.get("badge_err_fg"))
            obj.tag_configure("warn_row",
                background=t.get("badge_warn_bg"),
                foreground=t.get("badge_warn_fg"))
            obj.rehighlight()

        for _attr in ("tree", "map_tree", "log_tree", "thdm_raw_tree", "_thdm_header_tree",
                      "catalog_import_sheet"):
            _obj = getattr(self, _attr, None)
            if isinstance(_obj, SheetTable):
                _update_sheet_tags(_obj)
        for _tr in getattr(self, "_thdm_sec_trees", {}).values():
            if isinstance(_tr, SheetTable):
                _update_sheet_tags(_tr)

        # 4. thdm_bom_sheet (raw tksheet.Sheet)
        _sh = getattr(self, "thdm_bom_sheet", None)
        if _sh is not None:
            _st  = t["scroll_track"]
            _sfg = t["scroll_thumb"]
            _sac = t["scroll_thumb_active"]
            try:
                _sh.set_options(
                    table_bg=t["sheet_table_bg"],
                    table_fg=t["sheet_table_fg"],
                    header_bg=t["sheet_header_bg"],
                    header_fg=t["sheet_header_fg"],
                    index_bg=t["sheet_index_bg"],
                    index_fg=t["sheet_index_fg"],
                    top_left_bg=t["sheet_header_bg"],
                    show_horizontal_grid=True,
                    show_vertical_grid=True,
                    table_grid_fg=t["sheet_grid"],
                    header_grid_fg=t["sheet_grid"],
                    index_grid_fg=t["sheet_grid"],
                    outline_color=t["sheet_grid"],
                    # Scrollbar
                    vertical_scroll_troughcolor=_st,
                    vertical_scroll_not_active_bg=_sfg,
                    vertical_scroll_not_active_fg=_sfg,
                    vertical_scroll_active_bg=_sac,
                    vertical_scroll_active_fg=_sac,
                    vertical_scroll_pressed_bg=_sac,
                    vertical_scroll_pressed_fg=_sac,
                    vertical_scroll_bordercolor=_st,
                    vertical_scroll_lightcolor=_st,
                    vertical_scroll_darkcolor=_st,
                    horizontal_scroll_troughcolor=_st,
                    horizontal_scroll_not_active_bg=_sfg,
                    horizontal_scroll_not_active_fg=_sfg,
                    horizontal_scroll_active_bg=_sac,
                    horizontal_scroll_active_fg=_sac,
                    horizontal_scroll_pressed_bg=_sac,
                    horizontal_scroll_pressed_fg=_sac,
                    horizontal_scroll_bordercolor=_st,
                    horizontal_scroll_lightcolor=_st,
                    horizontal_scroll_darkcolor=_st,
                )
                _sh.redraw()
            except Exception:
                pass

        # 5. SearchCombo popup colors
        _SearchCombo.update_theme(t)

        # 6. Recolor tk.Frame + tk.Label + tk.Listbox — map mọi bg/fg đã biết → màu mới
        _D = THEMES["Dark"]
        _L = THEMES["Light"]
        _is_light = (actual == "Light")
        _color_map = {
            _D["bg_main"].lower():  t["bg_main"],
            _D["bg_card"].lower():  t["bg_card"],
            _D["bg_panel"].lower(): t["bg_panel"],
            _L["bg_main"].lower():  t["bg_main"],
            _L["bg_card"].lower():  t["bg_card"],
            _L["bg_panel"].lower(): t["bg_panel"],
            _L["border"].lower():   t["border"],    # "#e2e8f0" reverse → Dark border
            "#000000":              t["bg_main"],   # empty-state frame
            # Sidebar/separator colors (không có trong THEMES palette)
            "#161616": t["bg_panel"],               # sidebar bg rất tối
            "#2a2a2a": t["border"],                 # dải phân cách dọc
            "#333333": t["border"],                 # đường kẻ ngang trên bảng
            # Warning panel dark-red ↔ light-red (cả hai chiều Light↔Dark)
            "#251515": "#FFF5F5" if _is_light else "#251515",
            "#1a1212": "#FFF5F5" if _is_light else "#1A1212",
            "#7b1a1a": "#FECACA" if _is_light else "#7B1A1A",
            "#fff5f5": "#251515" if not _is_light else "#FFF5F5",
            "#fecaca": "#7B1A1A" if not _is_light else "#FECACA",
        }
        # fg_map: đổi màu chữ tk.Label theo theme (neutral colors, không đụng semantic)
        _fg_map = {
            "#8a8a8a": t["text_muted"],   # C["muted"] dark
            "#64748b": t["text_muted"],   # Light text_muted
            "#e8e8e8": t["text_main"],    # C["text"] dark
            "#cccccc": t["tv_text"],      # common dark text
            "#1e293b": t["tv_text"],      # Light tv_text
            "#0f172a": t["text_main"],    # Light text_main
        }
        self._recolor_tk_widgets(self, _color_map, _fg_map)

        # Cập nhật fg/selectbackground của Listbox (color_map chỉ đổi bg)
        _listbox = getattr(self, "listbox", None)
        if _listbox:
            try:
                _listbox.configure(
                    fg=t["tv_text"],
                    selectbackground=t["tv_sel"],
                    selectforeground=t["tv_text"],
                    highlightbackground=t["border"],
                )
            except Exception:
                pass

        # Cập nhật fg của warning panel labels (đảm bảo đọc được trong Light mode)
        _warn_fg  = "#DC2626" if _is_light else "#CC8888"
        _warn_tog = "#888888"
        _bdg_fg   = "#7F1D1D" if _is_light else "#FFCCCC"
        for _attr, _cfg in [
            ("_warn_hdr_lbl", {"fg": _warn_fg}),
            ("_warn_toggle",  {"fg": _warn_tog}),
            ("_warn_badge",   {"fg": _bdg_fg}),
        ]:
            _w = getattr(self, _attr, None)
            if _w:
                try:
                    _w.configure(**_cfg)
                except Exception:
                    pass

        # Cập nhật tag màu của _warn_err_tree (Treeview lỗi/cảnh báo)
        try:
            self._warn_err_tree.tag_configure("err_item",
                foreground=t["badge_err_fg"], background=t["badge_err_bg"])
            self._warn_err_tree.tag_configure("warn_item",
                foreground=t["badge_warn_fg"], background=t["badge_warn_bg"])
        except Exception:
            pass

        # 7. Cập nhật tag màu chữ các Treeview (catalog, log)
        try:
            self.catalog_tree.tag_configure("leaf",
                foreground=t["cat_leaf"])
            self.catalog_tree.tag_configure("group_inactive",
                foreground=t["cat_inactive"],
                font=FONT_SMALL_I)
            self.catalog_tree.tag_configure("orphan_group",
                foreground=t["cat_orphan"],
                font=FONT_SMALL_I)
        except Exception:
            pass
        try:
            self.log_tree.tag_configure("ok",   foreground=t["log_ok"])
            self.log_tree.tag_configure("warn", foreground=t["log_warn"])
            self.log_tree.tag_configure("err",  foreground=t["log_err"])
        except Exception:
            pass

        # 8. Force Tk repaint để CTk widgets hiển thị ngay màu mới
        self.update_idletasks()

        # 9. Lưu settings
        self._save_settings()

    def _open_settings(self):
        """Popup Cài đặt redesign: Theme + Tỷ lệ hiển thị."""
        # Toggle: bấm lại thì đóng
        try:
            if hasattr(self, '_settings_dlg') and self._settings_dlg.winfo_exists():
                self._settings_dlg.destroy()
                return
        except Exception:
            pass

        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        dlg.title("Cài đặt")
        dlg.geometry("400x270")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        self._settings_dlg = dlg

        # ── Header ──────────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(dlg, fg_color="transparent")
        hdr.pack(fill="x", padx=PAD_LG, pady=(16, 0))
        ctk.CTkLabel(hdr, text="⚙️  Cài đặt",
            font=ctk.CTkFont("Segoe UI", 15, "bold")).pack(side="left")

        # ── Giao diện ────────────────────────────────────────────────────────────
        ctk.CTkLabel(dlg, text="Giao diện",
            font=ctk.CTkFont(*FONT_MD_B),
            anchor="w").pack(fill="x", padx=PAD_LG, pady=(14, 4))

        _theme_labels = {"Dark": "🌙 Tối", "Light": "☀️ Sáng", "System": "💻 Hệ thống"}
        _label_to_mode = {v: k for k, v in _theme_labels.items()}
        _current_label = _theme_labels.get(
            getattr(self, "_current_theme", "Dark"), "🌙 Tối")

        theme_seg = ctk.CTkSegmentedButton(
            dlg,
            values=["🌙 Tối", "☀️ Sáng", "💻 Hệ thống"],
            font=ctk.CTkFont(*FONT_MD),
            command=lambda lbl: self.apply_theme(_label_to_mode.get(lbl, "Dark")),
        )
        theme_seg.set(_current_label)
        theme_seg.pack(fill="x", padx=PAD_LG)

        # ── Tỷ lệ hiển thị ───────────────────────────────────────────────────────
        ctk.CTkLabel(dlg, text="Tỷ lệ hiển thị",
            font=ctk.CTkFont(*FONT_MD_B),
            anchor="w").pack(fill="x", padx=PAD_LG, pady=(16, 4))

        scale_seg = ctk.CTkSegmentedButton(
            dlg,
            values=["80%", "90%", "100%", "110%", "120%"],
            font=ctk.CTkFont(*FONT_BODY),
            command=self.change_scaling_event,
        )
        scale_seg.set(getattr(self, "_current_scaling_pct", "100%"))
        scale_seg.pack(fill="x", padx=PAD_LG)

        # ── Divider + Đóng ───────────────────────────────────────────────────────
        ctk.CTkFrame(dlg, height=1, fg_color=("gray80", "gray30")).pack(
            fill="x", padx=PAD_LG, pady=(18, 10))
        ctk.CTkButton(dlg, text="Đóng", width=100,
            fg_color="transparent",
            border_width=1,
            command=dlg.destroy).pack()

        # Neo popup góc phải màn hình, ngay dưới nút ⚙️
        def _place_and_show():
            try:
                sw = self.winfo_screenwidth()
                by = self._btn_settings.winfo_rooty() + self._btn_settings.winfo_height() + 4
                dlg.geometry(f"+{sw - 415}+{by}")
            except Exception:
                dlg.geometry(f"+{self.winfo_screenwidth() - 415}+{55}")
            dlg.deiconify()
            dlg.lift()

        self.after(10, _place_and_show)

    def change_scaling_event(self, value: str):
        try:
            scale = int(value.replace("%", "")) / 100
            ctk.set_widget_scaling(scale)
        except Exception:
            pass
        self._current_scaling_pct = value
        # ctk.set_widget_scaling() → CTkTabview._set_scaling() → _configure_grid()
        # tự reset thanh tab về sticky="ns" (căn giữa) — áp lại canh trái.
        try:
            self.nb._segmented_button.grid_configure(sticky="w", padx=PAD_SM, pady=(4, 0))
        except Exception:
            pass
        self._save_settings()

    # ── Startup overlay ─────────────────────────────────────────────────────────

    def _show_startup_overlay(self):
        """Overlay kiểm tra kết nối DB khi khởi động. Che toàn bộ main UI."""
        dt = THEMES[ctk.get_appearance_mode()]
        overlay = ctk.CTkFrame(self, fg_color=dt["overlay_bg"], corner_radius=0)
        overlay.place(x=0, y=0, relwidth=1, relheight=1)
        overlay.lift()
        self._startup_overlay = overlay

        # Khung nội dung căn giữa
        center = ctk.CTkFrame(overlay, fg_color="transparent")
        center.place(relx=0.5, rely=0.46, anchor="center")

        # Logo circle
        logo_wrap = ctk.CTkFrame(center, fg_color=dt["bg_main"],
                                 width=88, height=88, corner_radius=44)
        logo_wrap.pack(pady=(0, 22))
        logo_wrap.pack_propagate(False)
        ctk.CTkLabel(logo_wrap, text="B",
                     font=ctk.CTkFont(*FONT_ICON),
                     text_color=dt["btn_primary"]).place(relx=0.5, rely=0.48, anchor="center")

        # Tên app
        ctk.CTkLabel(center, text="BOHO IMPORT BOM/THDM",
                     font=ctk.CTkFont("Segoe UI", 24, "bold"),
                     text_color=dt["text_main"]).pack()
        ctk.CTkLabel(center, text="v1  —  BOHO",
                     font=ctk.CTkFont(*FONT_MD),
                     text_color=dt["text_muted"]).pack(pady=(3, 22))

        # Separator
        ctk.CTkFrame(center, fg_color=dt["border"], height=1, width=340).pack(pady=(0, 22))

        # Thông tin kết nối
        cfg = self.db_cfg or {}
        server   = cfg.get("server",   "Chưa cấu hình")
        database = cfg.get("database", "—")
        db_card = ctk.CTkFrame(center, fg_color=dt["bg_main"], corner_radius=10)
        db_card.pack(pady=(0, 26), ipadx=24, ipady=4)
        ctk.CTkLabel(db_card, text=f"🖧  {server}",
                     font=ctk.CTkFont(*FONT_LABEL_N),
                     text_color=dt["tv_text"]).pack(padx=28, pady=(12, 3))
        ctk.CTkLabel(db_card, text=f"📋  {database}",
                     font=ctk.CTkFont(*FONT_MD),
                     text_color=dt["text_muted"]).pack(padx=28, pady=(0, 12))

        # Status label
        self._startup_status_lbl = ctk.CTkLabel(
            center, text="⏳  Đang kiểm tra kết nối...",
            font=ctk.CTkFont(*FONT_LABEL_N),
            text_color=dt["text_muted"])
        self._startup_status_lbl.pack()

        # Button frame (ẩn, chỉ hiện khi có lỗi)
        self._startup_btn_frame = ctk.CTkFrame(center, fg_color="transparent")

        self.update()
        threading.Thread(target=self._startup_check_worker, daemon=True).start()

    def _startup_check_worker(self):
        """Background thread: thử kết nối DB. Delay tối thiểu 0.8s để overlay hiện đủ lâu."""
        import time
        t0 = time.time()
        if not self.db_cfg:
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda: self._startup_check_done(False,
                "Chưa có cấu hình DB (db_config.json)."))
            return
        try:
            conn = self._get_db_conn(timeout_sec=15)
            conn.close()
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda: self._startup_check_done(True, None))
        except Exception as e:
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda err=e: self._startup_check_done(False, str(err)))

    def _startup_check_done(self, ok, error, attempt=1):
        """Main thread: cập nhật UI startup sau khi có kết quả kiểm tra."""
        dt = THEMES[ctk.get_appearance_mode()]
        if ok:
            self._startup_status_lbl.configure(
                text="✅  Kết nối thành công — Đang mở ứng dụng...",
                text_color=dt["log_ok"])
            self.after(1200, self._dismiss_startup_overlay)
        else:
            if attempt < 3:
                # Tự động thử lại tối đa 3 lần sau 2s — cho server/VPN kịp ổn định
                self._startup_status_lbl.configure(
                    text=f"⏳  Đang thử lại ({attempt}/3)...",
                    text_color=dt["text_muted"])
                def _retry_worker(att=attempt):
                    import time
                    time.sleep(2.0)
                    try:
                        conn = self._get_db_conn(timeout_sec=15)
                        conn.close()
                        self.after(0, lambda: self._startup_check_done(True, None, att + 1))
                    except Exception as ex:
                        self.after(0, lambda e=ex: self._startup_check_done(False, str(e), att + 1))
                threading.Thread(target=_retry_worker, daemon=True).start()
            else:
                # Sau 3 lần thất bại → hiện nút cho user
                self._startup_status_lbl.configure(
                    text="❌  Không kết nối được DB",
                    text_color=dt["log_err"])
                bf = self._startup_btn_frame
                bf.pack(pady=(18, 0))
                # Hiện chi tiết lỗi để dễ debug
                err_short = str(error)[:200] if error else "Không rõ nguyên nhân"
                ctk.CTkLabel(bf,
                    text=err_short,
                    font=ctk.CTkFont(*FONT_SMALL),
                    text_color=dt["text_muted"],
                    wraplength=420,
                    justify="left").pack(pady=(0, 8))
                ctk.CTkLabel(bf,
                    text="⚠️  Kiểm tra VPN hoặc cấu hình DB rồi thử lại",
                    font=ctk.CTkFont(*FONT_BODY),
                    text_color=dt["text_muted"]).pack(pady=(0, 14))
                row = ctk.CTkFrame(bf, fg_color="transparent")
                row.pack()
                ctk.CTkButton(row, text="🔄  Thử lại",
                    command=self._startup_retry,
                    fg_color=dt["btn_primary"],
                    hover_color=dt["btn_primary_hover"],
                    width=120, height=36, corner_radius=8).pack(side="left", padx=PAD_SM)
                ctk.CTkButton(row, text="Bỏ qua →",
                    command=self._dismiss_startup_overlay,
                    fg_color="transparent", border_width=1,
                    border_color=dt["border"], text_color=dt["text_muted"],
                    hover_color=dt["border"],
                    width=100, height=36, corner_radius=8).pack(side="left", padx=PAD_SM)

    def _startup_retry(self):
        """Xóa nút, reset status, thử kết nối lại."""
        dt = THEMES[ctk.get_appearance_mode()]
        for w in self._startup_btn_frame.winfo_children():
            w.destroy()
        self._startup_btn_frame.pack_forget()
        self._startup_status_lbl.configure(
            text="⏳  Đang thử lại...", text_color=dt["text_muted"])
        threading.Thread(target=self._startup_check_worker, daemon=True).start()

    def _dismiss_startup_overlay(self):
        """Đóng overlay, để lộ main UI."""
        try:
            self._startup_overlay.destroy()
        except Exception:
            pass
        # Auto-load Người lập + Dự án THDM — cố ý đợi tới khi overlay kiểm
        # tra kết nối khởi động đã có kết quả (thành công hoặc user bấm Bỏ
        # qua), thay vì bắn theo timer cố định 600ms/800ms độc lập với kết
        # quả kiểm tra (như trước đây). Race condition cũ: nếu VPN vừa mới
        # kết nối lúc app khởi động, 2 lệnh timer cũ fire trước khi VPN kịp
        # lên, báo lỗi và KHÔNG tự thử lại — dù startup overlay retry sau
        # đó thành công, 2 combo này vẫn kẹt ở trạng thái lỗi cũ.
        if not getattr(self, '_auto_loaded_after_startup', False):
            self._auto_loaded_after_startup = True
            self.after(200, self._load_creator_combo)
            self.after(400, self._thdm_load_products)


    def _build_ui(self):
        # ── Root grid ────────────────────────────────────────────────────────
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.var_export_sql = tk.BooleanVar(value=False)

        # ── Top bar ──────────────────────────────────────────────────────────
        topbar = ctk.CTkFrame(self, corner_radius=0, height=46,
                              fg_color=("gray83", "gray13"))
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(1, weight=1)

        # Trái: App name
        ctk.CTkLabel(topbar,
            text="🏭  BOHO  BOM / THDM",
            font=ctk.CTkFont(*FONT_LABEL),
            text_color=("#16A34A", "#4EC9B0")).grid(
            row=0, column=0, padx=(14, 16), pady=PAD_MD)

        # Giữa: DB info
        if self.db_cfg:
            srv = self.db_cfg.get("server", "?")
            db  = self.db_cfg.get("database", "?")
            db_txt   = f"🗄  {srv} / {db}"
            db_color = ("gray40", "gray55")
        else:
            db_txt   = "⚠️  Chưa có db_config.json"
            db_color = ("#D97706", "#D7BA7D")
        ctk.CTkLabel(topbar, text=db_txt,
            font=ctk.CTkFont(*FONT_BODY),
            text_color=db_color).grid(row=0, column=1, sticky="w")

        # Phải: status + mapping + settings
        right_bar = ctk.CTkFrame(topbar, fg_color="transparent")
        right_bar.grid(row=0, column=2, padx=(0, 10))

        self.lbl_status = CLabel(right_bar, text="Sẵn sàng...",
            font=ctk.CTkFont("Segoe UI", 11, slant="italic"),
            text_color=("gray40", "gray55"),
            fg_color="transparent")
        self.lbl_status.pack(side=tk.LEFT, padx=(0, 10))
        Tooltip(self.lbl_status,
            lambda: ("Bỏ qua: " + ", ".join(self._skipped_sheets))
                    if self._skipped_sheets else "")

        mapping_ok = bool(self.mapping)
        _mapping_text_color = ("#16A34A","#4ADE80") if mapping_ok else ("#D97706","#D7BA7D")
        ctk.CTkButton(right_bar, text="⚙️  Mapping",
            command=self._open_mapping_window,
            fg_color="transparent", border_width=1,
            text_color=_mapping_text_color,
            border_color=_mapping_text_color,
            hover_color=("gray90","gray25"),
            font=ctk.CTkFont(*FONT_BODY),
            width=110, height=28, corner_radius=6).pack(side=tk.LEFT, padx=(0, 4))

        self._btn_settings = ctk.CTkButton(right_bar, text="🎨",
            command=self._open_settings,
            fg_color="transparent", border_width=1,
            text_color=("gray50","gray60"),
            border_color=("gray50","gray60"),
            hover_color=("gray90","gray25"),
            font=ctk.CTkFont(*FONT_BODY),
            width=34, height=28, corner_radius=6)
        self._btn_settings.pack(side=tk.LEFT)

        # ── Tabview ──────────────────────────────────────────────────────────
        self.nb = ctk.CTkTabview(self,
            fg_color=("gray95", C["bg"]),
            segmented_button_fg_color=("gray82", C["panel"]),
            segmented_button_selected_color=("#007ACC", "#007ACC"),
            segmented_button_selected_hover_color=("#005999", "#005999"),
            segmented_button_unselected_color=("gray82", C["panel"]),
            segmented_button_unselected_hover_color=("gray76", "#2D2D2D"),
            text_color=("gray20", C["text"]),
            border_width=0, corner_radius=0)
        self.nb.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        self.nb.add(TAB_IMPORT)
        self.nb.add(TAB_THDM)
        self.nb.add(TAB_CATALOG)
        try:
            self.nb._segmented_button.configure(
                font=ctk.CTkFont(*FONT_MD_B), height=38)
            self.nb._segmented_button.grid_configure(sticky="w", padx=PAD_MD, pady=(6, 0))
        except Exception:
            pass

        self._build_tab_import()
        self._build_tab_catalog()
        self._build_tab_thdm()
        self._build_log_panel()

        # Auto-load Người lập + Dự án THDM giờ được kích từ
        # _dismiss_startup_overlay() (chỉ chạy sau khi biết kết quả kiểm
        # tra kết nối khởi động — xem ghi chú tại đó).
        # Reload khi user switch sang tab THDM (bắt sự kiện tab change)
        self.nb.configure(command=self._on_tab_changed)

    def _build_tab_import(self):
        tab = self.nb.tab(TAB_IMPORT)
        dt = THEMES[ctk.get_appearance_mode()]

        # ── Action bar — Card phẳng nhóm toàn bộ công cụ ───────────────────────
        bar = ctk.CTkFrame(tab, fg_color=("gray88", "#252526"),
                           corner_radius=0, height=52)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        # Ghost Button — style mặc định cho mọi nút phụ (Flat/Minimal)
        _G = dict(font=ctk.CTkFont(*FONT_BODY_B),
                  fg_color=("#EBEBEB", "#2D2D2D"),
                  text_color=("#1A1A1A", "#E1E1E1"),
                  hover_color=("#D8D8D8", "#3A3A3A"),
                  border_width=1, border_color=("#CCCCCC", "#3C3C3C"),
                  corner_radius=8, height=32)

        def _ab_btn(text, cmd, state="normal", text_color_disabled=None, **extra):
            kw = dict(_G, state=state, **extra)
            if text_color_disabled:
                kw["text_color_disabled"] = text_color_disabled
            b = CButton(bar, text=text, command=cmd, **kw)
            b.pack(side=tk.LEFT, padx=(0, 4), pady=10)
            return b

        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_open = _ab_btn("📂  ①  Chọn file Excel", self._open_file)

        # Batch — ghost style, ẩn mặc định
        self.btn_batch = CButton(bar,
            text="📚  Batch nhiều file", command=self._open_multiple_files, **_G)

        ctk.CTkFrame(bar, fg_color=("gray65","#3C3C3C"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=PAD_MD)

        self.btn_validate = _ab_btn("🔍  ②  Kiểm tra", self._run_validate,
                                    state="disabled")
        self.btn_report   = _ab_btn("📋  Báo cáo lỗi",
                                    self._export_validate_report, state="disabled")

        ctk.CTkFrame(bar, fg_color=("gray65","#3C3C3C"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=PAD_MD)

        ctk.CTkLabel(bar, text="Nhân viên:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray55"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0, 4))

        self.cmb_creator = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=200, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._on_creator_change,
            placeholder="— Chọn nhân viên —")
        self.cmb_creator.pack(side=tk.LEFT, padx=(0, 6), pady=10)

        ctk.CTkFrame(bar, fg_color=("gray65","#3C3C3C"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=PAD_MD)

        # ── Primary: nút hành động chính duy nhất — Accent Blue ──────────────
        self.btn_import = CButton(bar, text="🚀  ③  Import vào BRAVO",
            command=self._start_import, state="disabled",
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color=("#0066CC", "#0066CC"),
            hover_color=("#0055AA", "#0055AA"),
            text_color="#FFFFFF",
            text_color_disabled="#6B9CD6",
            corner_radius=8, height=32)
        self.btn_import.pack(side=tk.LEFT, padx=(0, 4), pady=10)

        self.btn_undo_import = _ab_btn("↩  Hoàn tác import", self._undo_last_import,
                                       state="disabled")
        Tooltip(self.btn_undo_import,
                lambda: "Xóa BOM vừa import khỏi BRAVO (gọi SP usp_BOMTool_DeleteBOM"
                        " — khách hàng cần deploy SP trước)")

        # Kẻ ngang phân cách toolbar → bảng
        tk.Frame(tab, bg=dt["border"], height=1).pack(fill=tk.X)

        # ── Body: sheet list (trái) | treeview (phải) ────────────────────────
        body = tk.Frame(tab, bg=dt["bg_main"])
        body.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # Sheet list
        sb_frame = tk.Frame(body, bg=dt["bg_deep"], width=170)
        sb_frame.pack(side=tk.LEFT, fill=tk.Y)
        sb_frame.pack_propagate(False)

        tk.Label(sb_frame, text="BẢNG DỮ LIỆU",
            bg=dt["bg_deep"], fg=C["muted"],
            font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=PAD_MD, pady=(10, 4))

        lb_wrap = tk.Frame(sb_frame, bg=dt["bg_deep"])
        lb_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self.listbox = tk.Listbox(lb_wrap,
            bg=dt["bg_deep"], fg=dt["tv_text"],
            selectbackground=dt["tv_sel"], selectforeground=dt["tv_text"],
            font=FONT_BODY,
            relief=tk.FLAT, bd=0, highlightthickness=1,
            highlightcolor=dt["accent"], highlightbackground=dt["border"],
            activestyle="none")
        self.listbox.pack(fill=tk.BOTH, expand=True)
        self.listbox.bind("<<ListboxSelect>>", self._on_select)

        # Separator
        tk.Frame(body, bg=dt["border"], width=1).pack(side=tk.LEFT, fill=tk.Y)

        # Content phải — card với nền panel và viền mảnh (fix horizontal scrollbar)
        right = ctk.CTkFrame(body, fg_color=("#FFFFFF", C["panel"]),
                             corner_radius=6, border_width=1,
                             border_color=("#E2E8F0", C["border"]))
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6), pady=6)
        right.rowconfigure(0, weight=0)   # lbl_table
        right.rowconfigure(1, weight=0)   # _warn_outer (ẩn/hiện)
        right.rowconfigure(2, weight=1)   # _tf_outer — luôn chiếm phần còn lại
        right.columnconfigure(0, weight=1)

        self.lbl_table = CLabel(right,
            text="— Chưa có dữ liệu —",
            font=ctk.CTkFont(*FONT_LABEL),
            text_color=("gray50","gray60"),
            fg_color="transparent", anchor="w")
        self.lbl_table.grid(row=0, column=0, sticky="ew", pady=(0, 2))

        # ── Warning panel (thay thế lbl_errors) ─────────────────────────────
        self._warn_outer = tk.Frame(right, bg=dt["bg_main"])
        # Đặt vào grid row=1, ẩn ngay bằng grid_remove() (giữ cấu hình grid)
        self._warn_outer.grid(row=1, column=0, sticky="ew")
        self._warn_outer.grid_remove()   # ẩn khi chưa có lỗi

        # header: badge + text tóm tắt + toggle
        _wh = tk.Frame(self._warn_outer, bg=dt["danger_hdr_bg"], pady=2)
        _wh.pack(fill=tk.X)

        self._warn_badge = tk.Label(_wh, text="", bg=dt["danger_badge_bg"],
            fg=dt["danger_badge_fg"],
            font=FONT_SMALL_B, padx=PAD_SM, pady=0, cursor="hand2")
        self._warn_badge.pack(side=tk.LEFT, padx=(4, 6))

        self._warn_hdr_lbl = tk.Label(_wh, text="", bg=dt["danger_hdr_bg"],
            fg=dt["danger_text"],
            font=FONT_SMALL, anchor="w")
        self._warn_hdr_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self._warn_expanded = True
        self._warn_toggle = tk.Label(_wh, text="▾", bg=dt["danger_hdr_bg"],
            fg=dt["text_muted"],
            font=FONT_BODY, cursor="hand2", padx=6)
        self._warn_toggle.pack(side=tk.RIGHT)
        self._warn_toggle.bind("<Button-1>", lambda e: self._toggle_warn_panel())
        self._warn_badge.bind("<Button-1>",  lambda e: self._toggle_warn_panel())

        # list area (collapsible)
        self._warn_list_outer = tk.Frame(self._warn_outer, bg=dt["danger_list_bg"])
        self._warn_list_outer.pack(fill=tk.X)

        self._warn_err_tree = ttk.Treeview(
            self._warn_list_outer,
            columns=("rownum", "field", "msg"),
            show="", height=4, selectmode="browse")
        self._warn_err_tree.column("rownum", width=70,  minwidth=60,  anchor="w", stretch=False)
        self._warn_err_tree.column("field",  width=90,  minwidth=70,  anchor="w", stretch=False)
        self._warn_err_tree.column("msg",    width=500, minwidth=200, anchor="w", stretch=True)

        _wsb = ttk.Scrollbar(self._warn_list_outer, orient="vertical",
            command=self._warn_err_tree.yview)
        self._warn_err_tree.configure(yscrollcommand=_wsb.set)
        self._warn_err_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
        _wsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._warn_err_tree.tag_configure("err_item",  foreground=C["badge_err_fg"],  background=C["badge_err_bg"])
        self._warn_err_tree.tag_configure("warn_item", foreground=C["badge_warn_fg"], background=C["badge_warn_bg"])
        self._warn_err_tree.bind("<ButtonRelease-1>", self._on_warn_item_click)

        # map: warn_tree iid → main tree iid (populated trong _show_table)
        self._warn_iid_to_tree_iid = {}

        # Empty state overlay
        self._tf_outer = tk.Frame(right, bg=dt["bg_main"])
        self._tf_outer.grid(row=2, column=0, sticky="nsew", pady=(4, 0))

        # Match màu nền sheet_table_bg: SheetTable giờ dùng dt["sheet_table_bg"]
        # (không còn hardcode #000000) nên overlay cũng phải đồng bộ theo theme.
        _EMPTY_BG = dt["sheet_table_bg"]
        # fill toàn bộ _tf_outer để che hoàn toàn tksheet bên dưới
        self._empty_frame = tk.Frame(self._tf_outer, bg=_EMPTY_BG)
        self._empty_frame.place(x=0, y=0, relwidth=1, relheight=1)
        # sub-frame chứa nội dung, căn giữa bên trong _empty_frame
        _ec = tk.Frame(self._empty_frame, bg=_EMPTY_BG)
        _ec.place(relx=0.5, rely=0.45, anchor="center")
        tk.Label(_ec, text="📂", bg=_EMPTY_BG, fg="#3E3E42",
                 font=FONT_HERO).pack()
        tk.Label(_ec,
                 text="①  Chọn file Excel ở sidebar để bắt đầu",
                 bg=_EMPTY_BG, fg="#555555",
                 font=FONT_LABEL_N).pack(pady=(4, 0))
        tk.Label(_ec,
                 text="Chọn Nhân viên  →  ②  Kiểm tra  →  ③  Import vào BRAVO",
                 bg=_EMPTY_BG, fg="#3E3E42",
                 font=FONT_BODY).pack(pady=(2, 0))

        tf = self._tf_outer
        if _HAS_TKSHEET:
            self.tree = SheetTable(tf)
            self.tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
            hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
            self.tree = ttk.Treeview(tf, style="BOM.Treeview", show="headings",
                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.configure(command=self.tree.yview)
            hsb.configure(command=self.tree.xview)
            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)

        # self.tree được tạo SAU _empty_frame nên theo thứ tự chồng lớp mặc
        # định của Tk, nó nằm TRÊN _empty_frame và che khuất hoàn toàn màn
        # hình hướng dẫn (dù _empty_frame vẫn còn đó, chỉ là bị đè). Nâng nó
        # lên trên cùng — khi có dữ liệu, nơi khác đã gọi place_forget() để
        # ẩn hẳn empty_frame nên không xung đột.
        self._empty_frame.lift()

        self.tree.tag_configure("sql_names",
            background=C["field_row_bg"], foreground=C["field_row_fg"],
            font=FONT_MD)
        self.tree.tag_configure("oddrow",  background=dt["bg_main"])
        self.tree.tag_configure("evenrow", background="#2A2D2E")
        self.tree.tag_configure("err_row",  background=C["badge_err_bg"])
        self.tree.tag_configure("warn_row", background=C["badge_warn_bg"])
        # Căn giữa riêng hàng phụ SQL (giống THDM); dữ liệu vẫn căn trái
        if hasattr(self.tree, 'center_rows_by_tag'):
            self.tree.center_rows_by_tag("sql_names")

        # Cell tooltip
        self._tree_tip = None
        def _motion(event):
            item = self.tree.identify_row(event.y)
            col  = self.tree.identify_column(event.x)
            if not item or not col:
                _hide(); return
            idx = int(col[1:]) - 1
            vals = self.tree.item(item, "values")
            text = str(vals[idx]) if idx < len(vals) else ""
            if not text or text == "—": _hide(); return
            if self._tree_tip:
                try: self._tree_tip.destroy()
                except Exception: pass
            tip = tk.Toplevel(self)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{event.x_root+16}+{event.y_root+12}")
            tk.Label(tip, text=text, bg=dt["bg_card"], fg=dt["tv_text"],
                     font=FONT_MD, relief="solid", bd=1,
                     padx=PAD_SM, pady=PAD_XS).pack()
            self._tree_tip = tip
        def _hide(*_):
            if getattr(self, "_tree_tip", None):
                try: self._tree_tip.destroy()
                except Exception: pass
                self._tree_tip = None
        if isinstance(self.tree, ttk.Treeview):
            # Tooltip + copy thủ công chỉ cần cho Treeview;
            # SheetTable (tksheet) có sẵn copy Ctrl+C / chuột phải.
            self.tree.bind("<Motion>", _motion)
            self.tree.bind("<Leave>",  _hide)
            self._bind_copy(self.tree, "Data")

    # ── Copy helper ──────────────────────────────────────────────────────────
    def _bind_copy(self, tree, title=""):
        """
        Gắn copy cho bất kỳ ttk.Treeview nào:
          - Click trái vào ô  → copy giá trị ô đó + tooltip "Đã copy"
          - Ctrl+C            → copy dòng đang select (tab-separated, kèm header)
          - Ctrl+A            → select all
          - Right-click       → context menu
        """
        dt = THEMES[ctk.get_appearance_mode()]
        if not isinstance(tree, ttk.Treeview):
            return   # SheetTable (tksheet) đã có copy tích hợp
        _tip_job  = [None]   # after() job id
        _tip_wnd  = [None]   # toplevel tooltip window

        # ── Tooltip "Đã copy" ────────────────────────────────────────────────
        def _show_copied_tip(x_root, y_root, text):
            # Huỷ tooltip cũ nếu có
            if _tip_wnd[0]:
                try: _tip_wnd[0].destroy()
                except Exception: pass
            if _tip_job[0]:
                tree.after_cancel(_tip_job[0])

            tip = tk.Toplevel(tree)
            tip.wm_overrideredirect(True)          # không có title bar
            tip.wm_attributes("-topmost", True)
            tip.configure(bg="#094771")

            # Cắt text dài
            display = text if len(text) <= 60 else text[:57] + "..."
            lbl = tk.Label(tip, text=f"✓ Đã copy: {display}",
                           bg="#094771", fg="white",
                           font=FONT_SMALL, padx=PAD_SM, pady=PAD_XS)
            lbl.pack()

            tip.update_idletasks()
            tip.wm_geometry(f"+{x_root + 12}+{y_root - 30}")
            _tip_wnd[0] = tip

            # Tự đóng sau 1.4 giây
            def _destroy():
                try: tip.destroy()
                except Exception: pass
                _tip_wnd[0] = None
            _tip_job[0] = tree.after(1400, _destroy)

        # ── Click trái vào ô ────────────────────────────────────────────────
        def _on_click(event):
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)   # "#1", "#2", ...
            if not row_id or not col_id:
                return
            try:
                col_idx = int(col_id.lstrip("#")) - 1
                vals = tree.item(row_id, "values")
                if col_idx < 0 or col_idx >= len(vals):
                    return
                cell_val = str(vals[col_idx])
                tree.clipboard_clear()
                tree.clipboard_append(cell_val)
                _show_copied_tip(event.x_root, event.y_root, cell_val)
                # Cell highlight — flash ngắn rồi tự ẩn (tránh để vết khi scroll/click khác)
                try:
                    bbox = tree.bbox(row_id, col_id)
                    if bbox:
                        x, y, w, h = bbox
                        if not hasattr(tree, '_cell_hl'):
                            tree._cell_hl    = tk.Frame(tree,
                                highlightbackground="#007ACC",
                                highlightthickness=2, bd=0, bg="")
                            tree._cell_hl_job = [None]
                        # Hủy job ẩn cũ nếu có
                        if tree._cell_hl_job[0]:
                            tree.after_cancel(tree._cell_hl_job[0])
                        tree._cell_hl.place(x=x, y=y, width=w, height=h)
                        tree._cell_hl.lift()
                        # Tự ẩn sau 500ms
                        def _hide_hl(hl=tree._cell_hl, job=tree._cell_hl_job):
                            try: hl.place_forget()
                            except Exception: pass
                            job[0] = None
                        tree._cell_hl_job[0] = tree.after(500, _hide_hl)
                except Exception:
                    pass
            except Exception:
                pass

        # ── Copy dòng được chọn ─────────────────────────────────────────────
        def _rows_to_text(iids):
            cols = tree["columns"]
            lines = ["\t".join(str(tree.heading(c)["text"]) for c in cols)]
            for iid in iids:
                vals = tree.item(iid, "values")
                lines.append("\t".join(str(v) for v in vals))
            return "\n".join(lines)

        def _copy_selected(event=None):
            sel = tree.selection()
            if not sel:
                return
            text = _rows_to_text(sel)
            tree.clipboard_clear()
            tree.clipboard_append(text)

        def _copy_all(event=None):
            all_iids = tree.get_children()
            if not all_iids:
                return
            text = _rows_to_text(all_iids)
            tree.clipboard_clear()
            tree.clipboard_append(text)

        def _select_all(event=None):
            tree.selection_set(tree.get_children())
            return "break"

        # ── Right-click menu ─────────────────────────────────────────────────
        def _show_menu(event):
            menu = tk.Menu(tree, tearoff=0,
                           bg=dt["bg_panel"], fg=dt["tv_text"],
                           activebackground="#094771", activeforeground="white",
                           relief="flat", bd=0)
            menu.add_command(label="Copy ô đang click        (click trái)", command=lambda: None)
            menu.add_separator()
            menu.add_command(label="Copy dòng được chọn  Ctrl+C", command=_copy_selected)
            menu.add_command(label="Copy tất cả             Ctrl+A → Ctrl+C", command=_copy_all)
            menu.add_separator()
            menu.add_command(label="Chọn tất cả              Ctrl+A", command=_select_all)
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        tree.bind("<Button-1>",  _on_click)
        tree.bind("<Control-c>", _copy_selected)
        tree.bind("<Control-C>", _copy_selected)
        tree.bind("<Control-a>", _select_all)
        tree.bind("<Control-A>", _select_all)
        tree.bind("<Button-3>",  _show_menu)

    def _open_mapping_window(self):
        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        dlg.title("Mapping Config")
        dlg.geometry("1200x680")
        dlg.transient(self)
        dlg.grab_set()
        self._build_mapping_content(dlg)
        dlg.update_idletasks()
        mx = self.winfo_rootx() + (self.winfo_width()  - dlg.winfo_width())  // 2
        my = self.winfo_rooty() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"1200x680+{mx}+{my}")
        dlg.deiconify()
        dlg.focus_force()

    def _build_mapping_content(self, parent):
        dt = THEMES[ctk.get_appearance_mode()]
        bar = ctk.CTkFrame(parent, corner_radius=0, height=48)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.pack(fill=tk.X, padx=PAD_MD, pady=PAD_SM)

        ctk.CTkLabel(inner, text="Phần:",
            font=ctk.CTkFont(*FONT_MD_B)).pack(side=tk.LEFT, padx=(0,8))

        _sect_names = list(self.mapping.get('_CONFIG', {}).keys())
        self._map_section_var = tk.StringVar(
            value=_sect_names[0] if _sect_names else "")
        section_cb = ttk.Combobox(inner, textvariable=self._map_section_var,
            values=_sect_names, state="readonly", width=22,
            font=FONT_MD)
        section_cb.pack(side=tk.LEFT)
        section_cb.bind("<<ComboboxSelected>>", lambda _: self._refresh_mapping_tab())

        for txt, color in [("✅ OK","#4EC9B0"), ("⚠️ Thiếu","#D7BA7D"),
                            ("❓ Cần xác nhận","#F97316"), ("🔧 System","gray")]:
            ctk.CTkLabel(inner, text=txt,
                font=ctk.CTkFont(*FONT_MD),
                text_color=color).pack(side=tk.RIGHT, padx=6)

        MAP_COLS   = ["SQL Column","Tên trên Excel","Kiểu DL","Bắt buộc",
                      "Nguồn DL","Bảng Master","Điều kiện","Kiểu Lookup","Ghi chú"]
        MAP_WIDTHS = [170,220,90,80,90,130,220,130,280]

        tf = tk.Frame(parent, bg=dt["bg_main"])
        tf.pack(fill=tk.BOTH, expand=True, padx=PAD_SM, pady=PAD_SM)
        if _HAS_TKSHEET:
            self.map_tree = SheetTable(tf, columns=MAP_COLS)
            self.map_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL)
            hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL)
            self.map_tree = ttk.Treeview(tf, style="BOM.Treeview", show="headings",
                columns=MAP_COLS, yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.configure(command=self.map_tree.yview)
            hsb.configure(command=self.map_tree.xview)
            self.map_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)
        for col, w in zip(MAP_COLS, MAP_WIDTHS):
            _al = guess_col_align(col)
            self.map_tree.heading(col, text=col, anchor=_al)
            self.map_tree.column(col, width=w, anchor=_al, stretch=(col=="Ghi chú"))
        self.map_tree.tag_configure("sys",  foreground="gray")
        self.map_tree.tag_configure("miss", foreground=dt["log_warn"])
        self.map_tree.tag_configure("conf", foreground="#F97316")
        self.map_tree.tag_configure("ok",   foreground=dt["tv_text"])
        self._bind_copy(self.map_tree, "Mapping")

        self._map_info_lbl = tk.Label(parent, text="",
            bg=dt["bg_main"], fg="gray", font=FONT_MD, anchor="w")
        self._map_info_lbl.pack(fill=tk.X, padx=10, pady=(0,4))
        self._refresh_mapping_tab()

    def _refresh_mapping_tab(self):
        section = self._map_section_var.get()
        records = self.mapping.get(section, [])

        self.map_tree.delete(*self.map_tree.get_children())
        n_ok = n_miss = n_conf = n_sys = 0

        for r in records:
            nguon  = r["nguon_dl"]
            kl     = r.get("kieu_lookup", "")
            kieu   = r["kieu_dl"] + (f"({r['do_dai']})" if r["do_dai"] else "")

            # Assign tag theo Nguon_DL
            if nguon in ("HeThong", "CoDinh"):
                tag = "sys";  n_sys  += 1
            elif nguon == "TinhToan":
                tag = "conf"; n_conf += 1
            elif nguon == "SP":
                tag = "miss"; n_miss += 1
            else:
                tag = "ok";   n_ok   += 1

            vals = (r["sql_col"], r["ten_excel"], kieu, r["bat_buoc"],
                    nguon, r["bang_master"], r["dieu_kien_master"],
                    kl, r["ghi_chu"])
            self.map_tree.insert("", tk.END, values=vals, tags=(tag,))

        total = len(records)
        self._map_info_lbl.config(
            text=("Section: " + section + "  |  Tong: " + str(total) + " cot"
                  + "  |  OK: " + str(n_ok)
                  + "  |  Thieu info: " + str(n_miss)
                  + "  |  Can xac nhan: " + str(n_conf)
                  + "  |  System: " + str(n_sys)))

    # ─ Tab 3: Tra cứu Danh mục ──────────────────────────────────────────────────────
    def _build_tab_catalog(self):
        tab = self.nb.tab(TAB_CATALOG)

        # Shared state
        self._catalog_all_rows              = []
        self._catalog_mode                  = "all"
        self._catalog_template_rows         = []
        self._catalog_current_col_defs      = []
        self._catalog_group_filter          = None
        self._catalog_import_filepath       = None   # path file đã chọn ở Import tab
        self._catalog_import_sheet_headers  = []     # row 0 của Excel (dùng làm col headers)
        self._catalog_import_sheet_data     = []     # rows 1+ của Excel
        self._template_id_map               = {}     # display_str → TemplateId (int)
        self.catalog_import_sheet           = None

        # ── Sub-tab container ─────────────────────────────────────────────
        _SUB_TAB_VIEW   = "📋  Danh sách vật tư"
        _SUB_TAB_IMPORT = "📥  Import dữ liệu"
        self._catalog_subtabs = ctk.CTkTabview(tab,
            fg_color="transparent",
            segmented_button_fg_color=("gray85", "#252526"),
            segmented_button_selected_color=("#1D4ED8", "#2563EB"),
            segmented_button_selected_hover_color=("#1E40AF", "#1D4ED8"),
            segmented_button_unselected_color=("gray85", "#252526"),
            segmented_button_unselected_hover_color=("gray75", "#333333"),
            text_color=("gray10", "gray90"),
            border_width=0, corner_radius=8)
        self._catalog_subtabs.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self._catalog_subtabs.add(_SUB_TAB_VIEW)
        self._catalog_subtabs.add(_SUB_TAB_IMPORT)
        try:
            self._catalog_subtabs._segmented_button.grid_configure(
                sticky="w", padx=PAD_SM, pady=(4, 0))
        except Exception:
            pass

        self._setup_catalog_view_tab(self._catalog_subtabs.tab(_SUB_TAB_VIEW))
        self._setup_catalog_import_tab(self._catalog_subtabs.tab(_SUB_TAB_IMPORT))
        self._init_catalog_icons()

    def _setup_catalog_view_tab(self, tab):
        """Sub-tab 1: Treeview xem danh mục + Export."""
        dt = THEMES[ctk.get_appearance_mode()]
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=48, corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)
        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_catalog_load = ctk.CTkButton(bar, text="🔄  Tải dữ liệu",
            command=self._catalog_load,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color=("#1D4ED8","#1D4ED8"),
            hover_color=("#1E40AF","#1E40AF"),
            height=32, corner_radius=6)
        self.btn_catalog_load.pack(side=tk.LEFT, padx=(0,4), pady=PAD_SM)

        self.btn_catalog_export = CButton(bar, text="📤  Export Excel",
            command=self._catalog_export_excel,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#60A5FA"),
            border_color=("#3B82F6","#60A5FA"),
            hover_color=("#DBEAFE","#1E3A5F"),
            height=32, corner_radius=6, state="disabled")
        self.btn_catalog_export.pack(side=tk.LEFT, padx=(0,8), pady=PAD_SM)

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=(0,8), pady=10)

        self.catalog_search_var = tk.StringVar()
        self.catalog_search_var.trace_add("write", lambda *_: self._catalog_filter())
        ctk.CTkEntry(bar, textvariable=self.catalog_search_var,
            font=ctk.CTkFont(*FONT_BODY),
            placeholder_text="🔍  Tìm mã / tên...",
            width=220, height=32, corner_radius=6).pack(side=tk.LEFT, pady=PAD_SM)

        self.lbl_catalog_status = CLabel(bar, text="—  Chưa tải",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_catalog_status.pack(side=tk.RIGHT, padx=(0,12))

        content_card = ctk.CTkFrame(tab,
            fg_color=("gray95", C["panel"]),
            corner_radius=6, border_width=1,
            border_color=("gray80", C["border"]))
        content_card.pack(fill=tk.BOTH, expand=True, padx=6, pady=(4, 6))
        content_card.rowconfigure(0, weight=0)
        content_card.rowconfigure(1, weight=1)
        content_card.columnconfigure(0, weight=1)

        self.lbl_catalog_table = CLabel(content_card,
            text="— Chưa tải dữ liệu —",
            font=ctk.CTkFont(*FONT_LABEL),
            text_color=("gray50", "gray60"),
            fg_color="transparent", anchor="w")
        self.lbl_catalog_table.grid(row=0, column=0, sticky="ew", padx=PAD_SM, pady=(2, 2))

        tf = tk.Frame(content_card, bg=C["panel"])
        tf.grid(row=1, column=0, sticky="nsew", padx=1, pady=(0, 1))
        COLS = (
            "Mã Vật Tư", "Tên Vật Tư", "ĐVT",
            "Quy Cách (mm)", "Bộ Định Khoản", "Mã Ngành Hàng",
            "Cốt Ván", "Chất Liệu",
            "S.Mặt Chính", "Lớp Phủ Chính",
            "S.Mặt Phụ",  "Lớp Phủ Phụ",
            "Độ Dày",
        )
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
        self.catalog_tree = ttk.Treeview(tf, style="BOM.Treeview", show="tree headings",
            columns=COLS, yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.configure(command=self.catalog_tree.yview)
        hsb.configure(command=self.catalog_tree.xview)
        self.catalog_tree.column("#0",             width=72,  minwidth=36, stretch=False)
        self.catalog_tree.heading("#0",            text="")
        for _c in COLS:
            self.catalog_tree.heading(_c, text=_c, anchor="center")
        self.catalog_tree.column("Mã Vật Tư",      width=160, anchor="w",      stretch=False)
        self.catalog_tree.column("Tên Vật Tư",     width=200, anchor="w",      stretch=False)
        self.catalog_tree.column("ĐVT",            width=55,  anchor="center", stretch=False)
        self.catalog_tree.column("Quy Cách (mm)",  width=130, anchor="w",      stretch=False)
        self.catalog_tree.column("Bộ Định Khoản",  width=140, anchor="w",      stretch=False)
        self.catalog_tree.column("Mã Ngành Hàng",  width=120, anchor="w",      stretch=False)
        self.catalog_tree.column("Cốt Ván",        width=110, anchor="w",      stretch=False)
        self.catalog_tree.column("Chất Liệu",      width=110, anchor="w",      stretch=False)
        self.catalog_tree.column("S.Mặt Chính",    width=75,  anchor="center", stretch=False)
        self.catalog_tree.column("Lớp Phủ Chính",  width=120, anchor="w",      stretch=False)
        self.catalog_tree.column("S.Mặt Phụ",      width=75,  anchor="center", stretch=False)
        self.catalog_tree.column("Lớp Phủ Phụ",    width=120, anchor="w",      stretch=False)
        self.catalog_tree.column("Độ Dày",         width=65,  anchor="center", stretch=False)
        self.catalog_tree.tag_configure("group",          foreground=dt["log_ok"],       font=FONT_SMALL_B)
        self.catalog_tree.tag_configure("group_inactive", foreground=dt["cat_inactive"], font=FONT_SMALL_I)
        self.catalog_tree.tag_configure("leaf",           foreground=dt["cat_leaf"])
        self.catalog_tree.tag_configure("orphan_group",   foreground=dt["cat_orphan"],   font=FONT_SMALL_I)
        self.catalog_tree.tag_configure("sql_names",
            background=C["field_row_bg"], foreground=C["field_row_fg"],
            font=FONT_SMALL_I)
        self.catalog_tree.tag_configure("oddrow",  background=dt["bg_main"])
        self.catalog_tree.tag_configure("evenrow", background="#2A2D2E")
        self.catalog_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)
        self.catalog_tree.bind("<Double-1>", self._catalog_copy_code)
        self.catalog_tree.bind("<<TreeviewOpen>>",  self._catalog_on_expand)
        self.catalog_tree.bind("<<TreeviewClose>>", self._catalog_on_collapse)
        self._bind_copy(self.catalog_tree, "Catalog")

    def _setup_catalog_import_tab(self, tab):
        """Sub-tab 2: Template selector, chọn file Excel, SheetTable preview editable, Import."""
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=48, corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)
        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        CLabel(bar, text="Template NVL:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))

        self.cmb_item_template = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=220, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._on_catalog_template_change,
            placeholder="— Chọn Template —")
        self.cmb_item_template.pack(side=tk.LEFT, padx=(0,6), pady=PAD_SM)

        self.btn_export_template = CButton(bar, text="📋  Xuất Template",
            command=self._catalog_export_template_excel,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color="transparent", border_width=1,
            text_color=("#059669","#34D399"),
            border_color=("#059669","#34D399"),
            hover_color=("#D1FAE5","#064E3B"),
            height=32, corner_radius=6, state="disabled")
        self.btn_export_template.pack(side=tk.LEFT, padx=(0,4), pady=PAD_SM)

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=(4,8), pady=10)

        self.btn_catalog_file_select = CButton(bar, text="📂  Chọn file Excel",
            command=self._catalog_select_import_file,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color="transparent", border_width=1,
            text_color=("#D97706","#FCD34D"),
            border_color=("#D97706","#FCD34D"),
            hover_color=("#FEF3C7","#3D2A00"),
            height=32, corner_radius=6)
        self.btn_catalog_file_select.pack(side=tk.LEFT, padx=(0,4), pady=PAD_SM)

        self.btn_import_template = CButton(bar, text="📥  Import",
            command=self._catalog_import_template_excel,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color="transparent", border_width=1,
            text_color=("#7C3AED","#A78BFA"),
            border_color=("#7C3AED","#A78BFA"),
            hover_color=("#EDE9FE","#2D1B69"),
            height=32, corner_radius=6, state="disabled")
        self.btn_import_template.pack(side=tk.LEFT, padx=(0,4), pady=PAD_SM)

        self.btn_preview_sql = CButton(bar, text="📋  Preview SQL",
            command=self._catalog_preview_sql,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color="transparent", border_width=1,
            text_color=("#0369A1","#38BDF8"),
            border_color=("#0369A1","#38BDF8"),
            hover_color=("#E0F2FE","#0C4A6E"),
            height=32, corner_radius=6, state="disabled")
        # btn_preview_sql: tạm ẩn, giữ widget để state logic không bị lỗi

        self.lbl_import_status = CLabel(bar, text="—  Chưa chọn file",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_import_status.pack(side=tk.RIGHT, padx=(0,12))

        # Info bar hiển thị tên file đã chọn
        info_bar = ctk.CTkFrame(tab, fg_color=("gray85","gray18"), height=26, corner_radius=0)
        info_bar.pack(fill=tk.X)
        info_bar.pack_propagate(False)
        self.lbl_import_filename = CLabel(info_bar,
            text="📄  Chưa chọn file",
            font=ctk.CTkFont(*FONT_SMALL),
            text_color=("gray50","gray60"),
            fg_color="transparent", anchor="w")
        self.lbl_import_filename.pack(side=tk.LEFT, padx=PAD_MD, pady=3)

        # SheetTable editable preview
        content_card = ctk.CTkFrame(tab,
            fg_color=("gray95", C["panel"]),
            corner_radius=6, border_width=1,
            border_color=("gray80", C["border"]))
        content_card.pack(fill=tk.BOTH, expand=True, padx=6, pady=(2, 6))

        if _HAS_TKSHEET:
            self.catalog_import_sheet = SheetTable(content_card, columns=[])
            self.catalog_import_sheet.tag_configure(
                "sql_names",
                background=C["field_row_bg"], foreground=C["field_row_fg"])
            self.catalog_import_sheet.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        else:
            CLabel(content_card,
                text="⚠️  tksheet chưa được cài — không thể hiển thị preview.",
                font=ctk.CTkFont(*FONT_BODY),
                fg_color="transparent").pack(expand=True)

    # ── Catalog helper methods ──────────────────────────────────────────────────────────────────

    def _init_catalog_icons(self):
        """Tạo icon 16×16 cho catalog tree bằng tk.PhotoImage thuần —
        KHÔNG cần PIL (bản cũ dùng PIL nên máy thiếu Pillow là mất icon)."""
        self._icons_loaded = False
        try:
            def _img(rects):
                im = tk.PhotoImage(width=16, height=16)
                for (x0, y0, x1, y1, color) in rects:
                    im.put(color, to=(x0, y0, x1 + 1, y1 + 1))
                return im

            # Folder đóng (amber đậm)
            self._icon_folder_closed = _img([
                (0, 2, 5, 4,  "#C8881A"),   # tab góc trên-trái
                (0, 5, 15, 13, "#9A6010"),  # viền
                (1, 6, 14, 12, "#E8A828"),  # thân
                (1, 6, 14, 6,  "#F0C060"),  # viền trên sáng
            ])
            # Folder mở (amber sáng)
            self._icon_folder_open = _img([
                (0, 2, 5, 4,  "#E0A030"),
                (0, 5, 15, 13, "#9A6010"),
                (1, 6, 14, 12, "#F0C050"),
                (1, 6, 14, 6,  "#F8D878"),
            ])
            # Chồng tài liệu (nhóm "Mã lẻ không thuộc nhóm")
            self._icon_orphan = _img([
                (3, 3, 13, 13, "#555555"),   # bóng sau
                (1, 1, 11, 12, "#888888"),   # thân
                (3, 3, 9, 3,   "#BBBBBB"),   # dòng kẻ
                (3, 6, 9, 6,   "#BBBBBB"),
                (3, 9, 9, 9,   "#BBBBBB"),
            ])
            self._icons_loaded = True
        except Exception:
            pass  # môi trường lạ → chạy bình thường không icon

    def _catalog_on_collapse(self, event=None):
        """Đổi icon thư mục từ mở → đóng khi user collapse node."""
        if not getattr(self, '_icons_loaded', False):
            return
        iid = self.catalog_tree.focus()
        if not iid or iid == "_orphan_":
            return
        try:
            int(iid)
        except ValueError:
            return
        row = self._catalog_id_to_row.get(int(iid))
        if row and row['is_group']:
            self.catalog_tree.item(iid, image=self._icon_folder_closed)

    def _catalog_read_filter_ids(self):
        """
        Đọc sheet CATALOG_FILTER từ CK_Mapping_v2.xlsx.
        Trả về list[int] các RootId được khai báo.
        Nếu sheet không có hoặc rỗng → trả về [] (không lọc = load tất cả).
        Dòng bắt đầu bằng '#' bị bỏ qua (comment).
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True, data_only=True)
            if 'CATALOG_FILTER' not in wb.sheetnames:
                return []
            ws = wb['CATALOG_FILTER']
            ids = []
            for row in ws.iter_rows(min_row=4, values_only=True):  # skip title/header rows
                val = row[0] if row else None
                if val is None:
                    continue
                s = str(val).strip()
                if not s or s.startswith('#'):
                    continue
                try:
                    ids.append(int(float(s)))
                except (ValueError, TypeError):
                    pass
            return ids
        except Exception:
            return []

    def _catalog_load(self):
        """Trigger: disable nút, hiện loading, bắt thread worker."""
        self.lbl_catalog_status.config(text="⏳  Đang tải danh mục...", fg=C["yellow"])
        try:
            self.btn_catalog_load.configure(state="disabled", text="⏳  Đang tải...")
        except Exception:
            pass
        self.update_idletasks()
        threading.Thread(target=self._catalog_load_worker, daemon=True).start()

    def _catalog_load_worker(self):
        """Background thread: kết nối DB, chạy CTE, trả về data qua after()."""
        try:
            filter_ids = self._catalog_read_filter_ids()
            conn = self._get_db_conn()
            cur  = conn.cursor()

            try:
                cur.execute("SELECT TOP 1 Unit FROM B20Item")
                has_unit = True
            except Exception:
                has_unit = False

            unit_anchor    = "Unit"    if has_unit else "N'' AS Unit"
            unit_recursive = "b.Unit"  if has_unit else "N''"

            sort_anchor    = "CAST(RIGHT(REPLICATE('0',10)+CAST(Id AS VARCHAR(10)),10) AS VARCHAR(MAX))"
            sort_recursive = "t.SortPath+'/'+RIGHT(REPLICATE('0',10)+CAST(b.Id AS VARCHAR(10)),10)"

            # Extra fields từ vB20Item (JOIN sau CTE để lấy thêm thông tin vật tư)
            _extra_sel = (
                ", ISNULL(v.Specification,N'') AS Specification"
                ", ISNULL(v.ItemGroupName,N'') AS ItemGroupName"
                ", ISNULL(v.ItemCatgName,N'') AS ItemCatgName"
                ", ISNULL(v.BoardCore,N'') AS BoardCore"
                ", ISNULL(v.Material,N'') AS Material"
                ", ISNULL(CAST(v.MainFaceCount AS NVARCHAR(20)),N'') AS MainFaceCount"
                ", ISNULL(v.MainFinish,N'') AS MainFinish"
                ", ISNULL(CAST(v.SecondaryFaceCount AS NVARCHAR(20)),N'') AS SecFaceCount"
                ", ISNULL(v.SecondaryFinish,N'') AS SecondaryFinish"
                ", ISNULL(CAST(v.NormRate AS NVARCHAR(20)),N'') AS NormRate"
            )
            _final_sel = (
                f" SELECT ct.Id,ct.ParentId,ct.IsGroup,ct.Code,ct.Name,ct.Unit,ct.IsActive,ct.Lvl"
                f"{_extra_sel}"
                f" FROM CatalogTree ct"
                f" LEFT JOIN [BOMTool].[dbo].[vB20Item] v ON ct.Id=v.Id"
                f" ORDER BY ct.SortPath OPTION(MAXRECURSION 0)"
            )

            if filter_ids:
                if len(filter_ids) == 1:
                    sql = (
                        "WITH CatalogTree AS ("
                        f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                        f" {sort_anchor} AS SortPath"
                        " FROM B20Item"
                        " WHERE ParentId=? AND (IsActive=1 OR IsGroup=1)"
                        " UNION ALL"
                        f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                        " t.Lvl+1,"
                        f" {sort_recursive}"
                        " FROM B20Item b"
                        " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                        " WHERE b.IsActive=1 OR b.IsGroup=1"
                        f"){_final_sel}"
                    )
                    cur.execute(sql, filter_ids)
                else:
                    placeholders = ','.join('?' * len(filter_ids))
                    sql = (
                        "WITH CatalogTree AS ("
                        f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                        f" {sort_anchor} AS SortPath"
                        " FROM B20Item"
                        f" WHERE Id IN ({placeholders}) AND (IsActive=1 OR IsGroup=1)"
                        " UNION ALL"
                        f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                        " t.Lvl+1,"
                        f" {sort_recursive}"
                        " FROM B20Item b"
                        " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                        " WHERE b.IsActive=1 OR b.IsGroup=1"
                        f"){_final_sel}"
                    )
                    cur.execute(sql, filter_ids)
            else:
                sql = (
                    "WITH CatalogTree AS ("
                    f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                    f" {sort_anchor} AS SortPath"
                    " FROM B20Item"
                    " WHERE (ParentId IS NULL OR ParentId=0) AND (IsActive=1 OR IsGroup=1)"
                    " UNION ALL"
                    f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                    " t.Lvl+1,"
                    f" {sort_recursive}"
                    " FROM B20Item b"
                    " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                    " WHERE b.IsActive=1 OR b.IsGroup=1"
                    f"){_final_sel}"
                )
                cur.execute(sql)

            rows = cur.fetchall()
            conn.close()

            all_rows = [
                {
                    'id'              : r[0],
                    'parent_id'       : r[1] or 0,
                    'is_group'        : bool(r[2]),
                    'code'            : (r[3]  or '').strip(),
                    'name'            : (r[4]  or '').strip(),
                    'unit'            : (r[5]  or '').strip(),
                    'active'          : bool(r[6]),
                    'lvl'             : r[7],
                    'specification'   : (r[8]  or '').strip(),
                    'item_group_name' : (r[9]  or '').strip(),
                    'item_catg_name'  : (r[10] or '').strip(),
                    'board_core'      : (r[11] or '').strip(),
                    'material'        : (r[12] or '').strip(),
                    'main_face_count' : (r[13] or '').strip(),
                    'main_finish'     : (r[14] or '').strip(),
                    'sec_face_count'  : (r[15] or '').strip(),
                    'sec_finish'      : (r[16] or '').strip(),
                    'norm_rate'       : (r[17] or '').strip(),
                }
                for r in rows
            ]
            self.after(0, lambda d=all_rows, f=filter_ids: self._catalog_load_done(d, f, None))
        except Exception as e:
            self.after(0, lambda err=e: self._catalog_load_done(None, [], err))

    def _catalog_load_done(self, all_rows, filter_ids, error):
        """Main thread callback: cập nhật UI sau khi worker hoàn tất."""
        try:
            self.btn_catalog_load.configure(state="normal", text="🔄  Tải dữ liệu")
        except Exception:
            pass

        if error:
            if isinstance(error, ConnectionError):
                self.lbl_catalog_status.config(text="❌  VPN/mạng lỗi", fg=C["red"])
                self._show_msg("Lỗi kết nối DB", str(error))
            else:
                self.lbl_catalog_status.config(text=f"❌  Lỗi: {error}", fg=C["red"])
                self._show_msg("Lỗi tải danh mục", str(error))
            return

        self._catalog_all_rows              = all_rows
        self._catalog_mode                  = "all"
        self._catalog_group_filter          = None
        self._catalog_active_template_name  = None
        self._catalog_template_data         = {}
        self._catalog_template_col_defs     = []
        try:
            self.lbl_catalog_table.config(text="— Chưa tải dữ liệu —", fg=self._text_fg)
        except Exception:
            pass
        try:
            self.cmb_item_template.clear()
        except Exception:
            pass
        self._catalog_reset_default_cols()
        self._catalog_filter()
        total = sum(1 for r in all_rows if r['active'])
        filter_note = f"  (lọc {len(filter_ids)} nhóm)" if filter_ids else ""
        self.lbl_catalog_status.config(
            text=f"✅  {total:,} vật tư{filter_note}", fg=C["green"])
        try:
            self.btn_catalog_export.configure(state="normal")
        except Exception:
            pass
    # Placeholder 13 cột cho lazy-load nodes
    _CATALOG_PH_VALS = ("", "…", "", "", "", "", "", "", "", "", "", "", "")

    def _catalog_row_vals(self, row, name_disp=None):
        """Build tuple 13 giá trị cột cho một row dict của catalog."""
        return (
            row.get('code', ''),
            name_disp if name_disp is not None else row.get('name', ''),
            row.get('unit', ''),
            row.get('specification', ''),
            row.get('item_group_name', ''),
            row.get('item_catg_name', ''),
            row.get('board_core', ''),
            row.get('material', ''),
            row.get('main_face_count', ''),
            row.get('main_finish', ''),
            row.get('sec_face_count', ''),
            row.get('sec_finish', ''),
            row.get('norm_rate', ''),
        )

    def _catalog_filter(self):
        """Cây lazy (cha trước con). Có keyword → flat. Group filter → chỉ hiện subtree nhóm."""
        if not hasattr(self, '_catalog_all_rows'):
            return
        q    = self.catalog_search_var.get().strip().lower() if hasattr(self, 'catalog_search_var') else ''
        tree = self.catalog_tree
        tree.delete(*tree.get_children())

        # ── Tính subset rows dựa trên group filter (template selection) ──────
        grp_code = getattr(self, '_catalog_group_filter', None)
        if grp_code and self._catalog_all_rows:
            root = next((r for r in self._catalog_all_rows
                         if r['code'] == grp_code and r['is_group']), None)
            if root:
                subtree_ids = {root['id']}
                all_rows = []
                for r in self._catalog_all_rows:
                    if r['id'] in subtree_ids or r['parent_id'] in subtree_ids:
                        subtree_ids.add(r['id'])
                        all_rows.append(r)
                grp_suffix = f"  [{grp_code}]"
            else:
                # Mã nhóm không tồn tại trong tree → show all + cảnh báo
                all_rows = self._catalog_all_rows
                self.lbl_catalog_status.config(
                    text=f"⚠️  Không tìm thấy nhóm [{grp_code}] trong danh mục — kiểm tra ImportTemplate.ItemGroupCode",
                    fg=C.get("yellow", "orange"))
                grp_suffix = ""
        else:
            all_rows = self._catalog_all_rows
            grp_suffix = ""
        total = sum(1 for r in all_rows if r['active'])

        if q:
            # Flat search: hiện active nodes khớp keyword
            is_tmpl    = getattr(self, '_catalog_mode', 'all') == 'template'
            tmpl_data  = getattr(self, '_catalog_template_data', {})  if is_tmpl else {}
            col_keys_t = [cd[0] for cd in getattr(self, '_catalog_template_col_defs', [])]
            shown      = 0
            for row in all_rows:
                if not row.get('active', True):
                    continue
                if q not in row['code'].lower() and q not in row['name'].lower():
                    continue
                tag = "group" if row['is_group'] else "leaf"
                _fic = {}
                if row['is_group'] and getattr(self, '_icons_loaded', False):
                    _fic = {"image": self._icon_folder_closed}
                if is_tmpl and col_keys_t:
                    if row['is_group']:
                        vals = tuple(
                            row['code'] if ck == 'Code' else
                            row['name'] if ck == 'Name' else
                            (row.get('unit', '') or '') if ck == 'Unit' else ''
                            for ck in col_keys_t
                        )
                    else:
                        iv   = tmpl_data.get(row['id'], {})
                        vals = tuple(str(iv.get(ck, '') or '') for ck in col_keys_t)
                else:
                    vals = self._catalog_row_vals(row)
                tree.insert("", "end", values=vals, tags=(tag,), **_fic)
                shown += 1
                if shown >= 3000:
                    break
            self.lbl_catalog_status.config(
                text=f"\U0001f50d  {shown:,} / {total:,} kết quả{grp_suffix}", fg=self._text_fg)
        else:
            # Tree view: data SQL đã sort đúng thứ tự → build cây 1 lần
            id_set = {row['id'] for row in all_rows}
            children_map: dict = {}
            for row in all_rows:
                pid = row['parent_id']
                if pid and pid in id_set:
                    children_map.setdefault(pid, []).append(row)
            # Nhóm (IsGroup) luôn hiện trước mã lẻ trong mỗi cấp
            def _sort_key(r):
                return (0 if r['is_group'] else 1, r['code'])
            for lst in children_map.values():
                lst.sort(key=_sort_key)
            self._catalog_children_map = children_map
            self._catalog_id_to_row    = {row['id']: row for row in all_rows}

            # Bottom-up leaf count: SQL sort parent-trước-con
            _leaf_count: dict = {}
            for row in reversed(all_rows):
                if not row['is_group']:
                    _leaf_count[row['id']] = 1
                else:
                    _leaf_count[row['id']] = sum(
                        _leaf_count.get(c['id'], 0)
                        for c in children_map.get(row['id'], [])
                    )
            self._catalog_leaf_count = _leaf_count

            # Root = node có parent_id không nằm trong id_set
            _roots = [r for r in all_rows
                      if not r['parent_id'] or r['parent_id'] not in id_set]
            _roots.sort(key=_sort_key)

            if getattr(self, '_catalog_mode', 'all') == 'template':
                # Template mode: eager full tree + dynamic column values từ ImportTemplateDetail
                tmpl_data  = getattr(self, '_catalog_template_data', {})
                col_defs_t = getattr(self, '_catalog_template_col_defs', [])
                col_keys_t = [cd[0] for cd in col_defs_t]

                def _tmpl_grp_vals(row, name_disp):
                    return tuple(
                        row['code'] if ck == 'Code' else
                        name_disp   if ck == 'Name' else
                        (row.get('unit', '') or '') if ck == 'Unit' else ''
                        for ck in col_keys_t
                    )

                for row in all_rows:
                    rid        = row['id']
                    iid        = str(rid)
                    pid        = row['parent_id']
                    parent_iid = str(pid) if (pid and pid in id_set) else ""
                    if row['is_group']:
                        cnt       = _leaf_count.get(rid, 0)
                        nm        = f"{row['name']}  ({cnt:,})" if cnt else row['name']
                        tag       = "group" if row.get('active', True) else "group_inactive"
                        _ik       = {"image": self._icon_folder_open} if getattr(self, '_icons_loaded', False) else {}
                        tree.insert(parent_iid, "end", iid=iid, text="",
                                    values=_tmpl_grp_vals(row, nm),
                                    open=True, tags=(tag,), **_ik)
                    else:
                        iv   = tmpl_data.get(rid, {})
                        vals = tuple(str(iv.get(ck, '') or '') for ck in col_keys_t)
                        tree.insert(parent_iid, "end", iid=iid, text="",
                                    values=vals, open=False, tags=("leaf",))

                n_leaves = sum(1 for r in all_rows if not r['is_group'])
                self.lbl_catalog_status.config(
                    text=f"✅  {n_leaves:,} vật tư{grp_suffix}", fg=C["green"])
                _tmpl_name = getattr(self, '_catalog_active_template_name', '') or 'Template'
                _n_cols    = len(col_keys_t)
                try:
                    self.lbl_catalog_table.config(
                        text=f"Bảng: {_tmpl_name}   —   {n_leaves:,} vật tư × {_n_cols} cột",
                        fg=self._text_fg)
                except Exception:
                    pass
                self._catalog_autosize_cols(all_rows)
                if col_keys_t:
                    try:
                        tree.insert("", 0, iid="_catalog_sql_names_",
                                    text="", values=tuple(str(ck) for ck in col_keys_t),
                                    tags=("sql_names",))
                    except Exception:
                        pass
                return

            orphan_leaves = []
            group_count   = 0
            for row in _roots:
                if not row['is_group']:
                    # Mã lẻ không thuộc nhóm nào → gom lại cuối danh sách
                    orphan_leaves.append(row)
                    continue
                group_count += 1
                iid   = str(row['id'])
                cnt   = _leaf_count.get(row['id'], 0)
                name_disp = f"{row['name']}  ({cnt:,})" if cnt else row['name']
                tag   = "group" if row.get('active', True) else "group_inactive"
                _icon_kw = {"image": self._icon_folder_closed} if getattr(self, '_icons_loaded', False) else {}
                tree.insert("", "end", iid=iid, text="",
                            values=self._catalog_row_vals(row, name_disp=name_disp),
                            open=False, tags=(tag,), **_icon_kw)
                if row['id'] in children_map:
                    tree.insert(iid, "end", iid=f"_ph_{iid}",
                                values=self._CATALOG_PH_VALS, tags=())

            # Nhóm ảo cuối cùng: gom mã lẻ không thuộc nhóm nào
            self._catalog_orphan_leaves = orphan_leaves
            if orphan_leaves:
                _orphan_icon_kw = {"image": self._icon_orphan} if getattr(self, '_icons_loaded', False) else {}
                tree.insert("", "end", iid="_orphan_", text="",
                            values=("—",
                                    f"Mã lẻ (không thuộc nhóm)  ({len(orphan_leaves):,})",
                                    "", "", "", "", "", "", "", "", "", "", ""),
                            open=False, tags=("orphan_group",), **_orphan_icon_kw)
                tree.insert("_orphan_", "end", iid="_ph__orphan_",
                            values=self._CATALOG_PH_VALS, tags=())

            n_leaves = sum(1 for r in all_rows if not r['is_group'])
            self.lbl_catalog_status.config(
                text=(f"\U0001f4c2  {group_count} nhóm  \u2022  {n_leaves:,} vật tư" + grp_suffix),
                fg=self._text_fg)
            _n_cols    = len(getattr(self, '_catalog_current_col_defs', None) or []) or 3
            _grp_label = f" [{grp_code}]" if grp_code else ""
            try:
                self.lbl_catalog_table.config(
                    text=(f"Bảng: Danh mục Vật tư{_grp_label}"
                          f"   —   {n_leaves:,} vật tư × {_n_cols} cột"),
                    fg=self._text_fg)
            except Exception:
                pass
            self._catalog_autosize_cols(all_rows)

    def _catalog_on_expand(self, event=None):
        """Lazy-load children khi user click mở rộng một node trong catalog tree."""
        if getattr(self, '_catalog_mode', 'all') == 'template':
            return  # template mode: tree đã được build đầy đủ, không cần lazy load
        iid = self.catalog_tree.focus()
        if not iid:
            return
        children = self.catalog_tree.get_children(iid)
        # Chỉ xử lý nếu child đầu tiên là placeholder
        if not children or not str(children[0]).startswith("_ph_"):
            return
        # Xóa placeholder, insert real children
        self.catalog_tree.delete(children[0])

        # Swap icon parent: đóng → mở
        if getattr(self, '_icons_loaded', False) and iid != "_orphan_":
            try:
                int(iid)
                self.catalog_tree.item(iid, image=self._icon_folder_open)
            except ValueError:
                pass

        # Xử lý nhóm ảo "Mã lẻ không thuộc nhóm"
        if iid == "_orphan_":
            for leaf_row in getattr(self, '_catalog_orphan_leaves', []):
                leaf_iid = str(leaf_row['id'])
                self.catalog_tree.insert("_orphan_", "end", iid=leaf_iid, text="",
                    values=self._catalog_row_vals(leaf_row),
                    open=False, tags=("leaf",))
            return

        try:
            row_id = int(iid)
        except ValueError:
            return
        cmap = getattr(self, '_catalog_children_map', {})
        for child_row in cmap.get(row_id, []):
            child_iid = str(child_row['id'])
            if child_row['is_group']:
                tag = "group" if child_row.get('active', True) else "group_inactive"
                _cicon = {"image": self._icon_folder_closed} if getattr(self, '_icons_loaded', False) else {}
            else:
                tag = "leaf"
                _cicon = {}
            self.catalog_tree.insert(iid, "end", iid=child_iid, text="",
                values=self._catalog_row_vals(child_row),
                open=False, tags=(tag,), **_cicon)
            # Nếu child cũng có con → thêm placeholder
            if child_row['id'] in cmap:
                self.catalog_tree.insert(child_iid, "end",
                    iid=f"_ph_{child_iid}",
                    values=self._CATALOG_PH_VALS, tags=())

    def _catalog_export_excel(self):
        """Xuất toàn bộ danh mục ra file .xlsx."""
        if not hasattr(self, '_catalog_all_rows') or not self._catalog_all_rows:
            self._show_msg("Chưa có dữ liệu", "Hãy bấm 'Tải dữ liệu' trước.", 'warning')
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="DanhMucVatTu.xlsx")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "B20Item"
            _HEADERS = [
                "Mã Vật Tư", "Tên Vật Tư", "ĐVT",
                "Quy Cách (mm)", "Bộ Định Khoản", "Mã Ngành Hàng",
                "Cốt Ván", "Chất Liệu",
                "S.Mặt Chính", "Lớp Phủ Chính",
                "S.Mặt Phụ", "Lớp Phủ Phụ",
                "Độ Dày",
            ]
            _FIELDS = [
                'code', 'name', 'unit',
                'specification', 'item_group_name', 'item_catg_name',
                'board_core', 'material',
                'main_face_count', 'main_finish',
                'sec_face_count', 'sec_finish',
                'norm_rate',
            ]
            _COL_W = [22, 60, 10, 18, 20, 18, 14, 14, 12, 18, 12, 18, 10]
            ws.append(_HEADERS)
            ws.row_dimensions[1].height = 20
            hdr_fill = PatternFill("solid", fgColor="2B7A78")
            hdr_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for row in self._catalog_all_rows:
                ws.append([row.get(f, '') for f in _FIELDS])
            from openpyxl.utils import get_column_letter
            for ci, w in enumerate(_COL_W, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            wb.save(path)
            self._show_export_success(
                "Xuất Excel thành công",
                f"Đã xuất {len(self._catalog_all_rows):,} vật tư\n\u2192 {path}",
                path)
        except Exception as e:
            self._show_msg("Lỗi xuất Excel", str(e))

    def _catalog_copy_code(self, event):
        """Double-click dòng -> copy Mã Vật Tư vào clipboard."""
        sel = self.catalog_tree.selection()
        if not sel or sel[0] == "_catalog_sql_names_":
            return
        code = self.catalog_tree.item(sel[0], "values")[0]
        self.clipboard_clear()
        self.clipboard_append(code)
        self._toast(f"Đã copy: {code}", kind="success")
        self.lbl_catalog_status.config(
            text=f"📋  Đã copy: {code}", fg=C["accent"])

    # ── Import Template NVL ────────────────────────────────────────────────────

    DB_TEMPLATE_TABLE        = '[BOMTool].dbo.ImportTemplate'
    DB_TEMPLATE_DETAIL_TABLE = '[BOMTool].dbo.ImportTemplateDetail'

    def _catalog_load_templates(self):
        """Tải danh sách Template từ [BOMTool].dbo.ImportTemplate vào combobox."""
        try:
            self.cmb_item_template.configure(state="disabled")
        except Exception:
            pass
        threading.Thread(target=self._catalog_load_templates_worker, daemon=True).start()

    def _catalog_load_templates_worker(self):
        try:
            conn = self._get_db_conn(timeout_sec=5)
            cur  = conn.cursor()
            cur.execute(
                f"SELECT Id, TemplateCode, TemplateName, ItemGroupCode, NameGenAuto "
                f"FROM {self.DB_TEMPLATE_TABLE} "
                f"WHERE IsActive = 1 ORDER BY TemplateCode"
            )
            rows = cur.fetchall()
            conn.close()
            self.after(0, lambda r=rows: self._catalog_load_templates_done(r, None))
        except Exception as e:
            self.after(0, lambda err=str(e): self._catalog_load_templates_done(None, err))

    def _catalog_load_templates_done(self, rows, error):
        try:
            self.cmb_item_template.configure(state="normal")
        except Exception:
            pass
        if error:
            self.cmb_item_template.set_values(["— Chưa có template —"])
            return
        if not rows:
            self.cmb_item_template.set_values(["— Chưa có template —"])
            return
        self._template_id_map    = {}   # display_str → Id
        self._template_group_map = {}   # Id → ItemGroupCode
        display_list = []
        for tid, code, name, igc, nga in rows:
            disp = f"[{code}]  {name}"
            self._template_id_map[disp]    = tid
            self._template_group_map[tid]  = igc or ''
            display_list.append(disp)
        self.cmb_item_template.set_values(display_list)

    def _on_catalog_template_change(self, value):
        has = bool(value and value in self._template_id_map)
        try:
            self.btn_export_template.configure(state="normal" if has else "disabled")
        except Exception:
            pass
        try:
            self.btn_import_template.configure(state="normal" if has else "disabled")
        except Exception:
            pass
        try:
            self.btn_preview_sql.configure(state="normal" if has else "disabled")
        except Exception:
            pass
        # Tính năng lọc danh mục theo template tạm thời tắt (Import tab đã tách riêng)
        # if has:
        #     template_id     = self._template_id_map[value]
        #     item_group_code = getattr(self, '_template_group_map', {}).get(template_id, '')
        #     self._catalog_group_filter         = item_group_code or None
        #     self._catalog_active_template_name = value
        #     self._catalog_load_by_template(template_id, item_group_code)
        # else:
        #     self._catalog_group_filter         = None
        #     self._catalog_active_template_name = None
        #     self._catalog_mode = "all"
        #     self._catalog_reset_default_cols()
        #     self._catalog_filter()

    # ── Cột B20Item hợp lệ để SELECT động ──────────────────────────────────
    # Cột được phép SELECT từ vB20Item (whitelist chống injection cho ColKey thô)
    # ColView được validate bằng regex riêng (_SAFE_IDENT), không cần whitelist cứng
    _VALID_B20ITEM_COLS = frozenset({
        # ── B20Item ───────────────────────────────────────────────────────────
        'Code', 'Name', 'Unit', 'Name2', 'Specification',
        'ItemType', 'ItemGroupId', 'ItemCatgId',
        'ItemId0', 'ProductId1', 'BranchCode',
        'ParentId0', 'ParentId',
        'ItemClassPurchaseId', 'CustomFieldId1', 'QCItemGroupId',
        'BoardCore', 'MainFaceCount', 'MainFinish',
        'SecondaryFaceCount', 'SecondaryFinish',
        'EdgeBanding', 'Material', 'WoodType',
        # ── B20ItemInfo (vB20Item đã JOIN sẵn — không cần prefix) ────────────
        'Height', 'Width', 'Length', 'Weight',
        'ProductId', 'UnitOfLength', 'UnitOfHeight', 'UnitOfWidth', 'UnitOfWeight',
        'ProductClassId1', 'ProductClassId2',
        'SalesTaxCode', 'ExciseTaxId', 'EnvTaxId',
        'ProductItemId', 'ManufacturerId', 'ItemPurchasePriceTypeId',
    })
    # ColKey → tên cột thực trong B20Item (khi khác nhau)
    _COL_KEY_ALIAS = {
        'WoodType2': 'WoodType',
    }
    # ColKey → (width, anchor)
    _COL_WIDTH = {
        'Code':                (160, 'w'),
        'Name':                (360, 'w'),
        'Unit':                (60,  'center'),
        'ItemType':            (100, 'w'),
        'ItemCatgId':          (100, 'w'),
        'ItemGroupId':         (100, 'w'),
        'Specification':       (110, 'w'),
        'Name2':               (200, 'w'),
        'ParentId0':           (120, 'w'),
        'ParentId':            (150, 'w'),
        'BoardCore':           (120, 'w'),
        'MainFaceCount':       (80,  'center'),
        'MainFinish':          (130, 'w'),
        'SecondaryFaceCount':  (80,  'center'),
        'SecondaryFinish':     (130, 'w'),
        'EdgeBanding':         (130, 'w'),
        'Material':            (130, 'w'),
        'WoodType':            (130, 'w'),
        'B20ItemInfo.Height':  (80,  'center'),
    }

    def _catalog_load_by_template(self, template_id: int, item_group_code: str):
        """Trigger: status loading + bắt background thread load dữ liệu theo template."""
        self.lbl_catalog_status.config(
            text=f"⏳  Đang tải dữ liệu [{item_group_code}]...", fg=C["yellow"])
        try:
            self.btn_catalog_load.configure(state="disabled")
        except Exception:
            pass
        threading.Thread(
            target=self._catalog_template_worker,
            args=(template_id, item_group_code),
            daemon=True
        ).start()

    def _catalog_template_worker(self, template_id: int, item_group_code: str):
        """Background: query ImportTemplateDetail → build SELECT → query vB20Item."""
        import re
        _SAFE_IDENT = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')

        try:
            conn = self._get_db_conn(timeout_sec=10)
            cur  = conn.cursor()

            # 1. Lấy cấu hình cột từ ImportTemplateDetail (bao gồm ColView)
            cur.execute(
                f"SELECT ColKey, ColName, DataType, ColGroup, "
                f"ISNULL(ColView, N'') AS ColView "
                f"FROM {self.DB_TEMPLATE_DETAIL_TABLE} "
                f"WHERE TemplateId = ? ORDER BY ColIndex ASC",
                (template_id,)
            )
            detail_rows = cur.fetchall()   # [(col_key, col_name, dtype, col_group, col_view)]

            # 1b. Introspect các cột thực tế trong vB20Item, B20Item, B20ItemInfo
            #     TABLE_NAME trong INFORMATION_SCHEMA chỉ là tên object (không có db/schema prefix)
            cur.execute(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME = N'vB20Item' AND TABLE_SCHEMA = N'dbo'"
            )
            _view_cols = {r[0] for r in cur.fetchall()}

            cur.execute(
                "SELECT COLUMN_NAME, TABLE_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME IN (N'B20Item', N'B20ItemInfo') "
                "AND TABLE_SCHEMA = N'dbo'"
            )
            _base_col_src = {}   # col_name → 'B20Item' | 'B20ItemInfo'
            for _cn, _tn in cur.fetchall():
                _base_col_src.setdefault(_cn, _tn)   # B20Item wins nếu cùng tên

            # helper: chọn nguồn cho 1 tên cột → ('fragment', 'v'|'b'|'bi') hoặc None
            def _pick_src(cname, alias):
                if cname in _view_cols:
                    return (f"ISNULL(CAST(b.[{cname}] AS NVARCHAR(500)), N'') AS [{alias}]", 'v')
                if _base_col_src.get(cname) == 'B20Item':
                    return (f"ISNULL(CAST(braw.[{cname}] AS NVARCHAR(500)), N'') AS [{alias}]", 'b')
                if _base_col_src.get(cname) == 'B20ItemInfo':
                    return (f"ISNULL(CAST(bi.[{cname}] AS NVARCHAR(500)), N'') AS [{alias}]", 'bi')
                return None

            # 2. Resolve alias, chọn cột SELECT, fallback B20Item/B20ItemInfo
            select_parts        = []
            disp_col_defs       = []
            seen_db_cols        = set()
            need_b20item_join   = False
            need_b20iteminfo_join = False

            for col_key, col_name, dtype, col_group, col_view in detail_rows:
                db_col = self._COL_KEY_ALIAS.get(col_key, col_key)
                if db_col in seen_db_cols:
                    continue
                seen_db_cols.add(db_col)

                safe_alias = col_key.replace("'", "")
                w, a = self._COL_WIDTH.get(col_key, (100, 'w'))

                # Xác định raw_candidate (tên cột thực tế trong B20Item/B20ItemInfo)
                if db_col.startswith('B20ItemInfo.'):
                    raw_candidate = db_col[len('B20ItemInfo.'):]
                elif db_col in self._VALID_B20ITEM_COLS:
                    raw_candidate = db_col
                else:
                    continue   # ColKey không hợp lệ

                # Candidate ưu tiên: ColView (lookup name) → raw ColKey làm fallback
                candidates = []
                if col_view and _SAFE_IDENT.match(col_view):
                    candidates.append(col_view)
                if _SAFE_IDENT.match(raw_candidate):
                    candidates.append(raw_candidate)

                resolved = None
                for cand in candidates:
                    resolved = _pick_src(cand, safe_alias)
                    if resolved:
                        break

                if resolved is None:
                    continue   # cột không tìm thấy ở bất kỳ nguồn nào

                frag, src = resolved
                select_parts.append(frag)
                if src == 'b':
                    need_b20item_join = True
                elif src == 'bi':
                    need_b20iteminfo_join = True

                disp_col_defs.append((col_key, col_name or col_key, w, a))

            if not select_parts:
                select_parts  = [
                    "ISNULL(CAST(b.[Code] AS NVARCHAR(500)),N'') AS [Code]",
                    "ISNULL(CAST(b.[Name] AS NVARCHAR(500)),N'') AS [Name]",
                    "ISNULL(CAST(b.[Unit] AS NVARCHAR(500)),N'') AS [Unit]",
                ]
                disp_col_defs = [
                    ("Code", "Mã Vật Tư", 160, "w"),
                    ("Name", "Tên Vật Tư", 360, "w"),
                    ("Unit", "DVT",         60, "center"),
                ]

            sql_select = ",\n    ".join(["b.Id AS [_item_id]"] + select_parts)

            join_clause = ""
            if need_b20item_join:
                join_clause += "LEFT JOIN B20Item braw ON braw.Id = b.Id\n"
            if need_b20iteminfo_join:
                join_clause += "LEFT JOIN B20ItemInfo bi ON bi.ItemId = b.Id\n"

            # 3. CTE dùng B20Item (cấu trúc cây); SELECT chính dùng vB20Item
            if item_group_code:
                sql_data = (
                    f"WITH GrpTree AS (\n"
                    f"    SELECT Id FROM B20Item WHERE Code = ? AND IsGroup = 1\n"
                    f"    UNION ALL\n"
                    f"    SELECT b2.Id FROM B20Item b2\n"
                    f"    INNER JOIN GrpTree g ON b2.ParentId = g.Id\n"
                    f"    WHERE b2.IsGroup = 1\n"
                    f")\n"
                    f"SELECT {sql_select}\n"
                    f"FROM [BOMTool].[dbo].[vB20Item] b\n"
                    f"{join_clause}"
                    f"WHERE b.ParentId IN (SELECT Id FROM GrpTree)\n"
                    f"  AND (b.IsGroup IS NULL OR b.IsGroup = 0)\n"
                    f"  AND b.IsActive = 1\n"
                    f"ORDER BY b.Code\n"
                    f"OPTION (MAXRECURSION 0)"
                )
                cur.execute(sql_data, (item_group_code,))
            else:
                sql_data = (
                    f"SELECT {sql_select}\n"
                    f"FROM vB20Item b\n"
                    f"{join_clause}"
                    f"WHERE (b.IsGroup IS NULL OR b.IsGroup = 0)\n"
                    f"  AND b.IsActive = 1\n"
                    f"ORDER BY b.Code"
                )
                cur.execute(sql_data)

            raw_rows  = cur.fetchall()
            col_keys  = [cd[0] for cd in disp_col_defs]
            all_dicts = []
            for r in raw_rows:
                d = {'_item_id': r[0]}            # r[0] = b.Id (prepended above)
                for i, ck in enumerate(col_keys):
                    d[ck] = r[i + 1] or ''        # r[1..n] = template columns
                all_dicts.append(d)
            conn.close()

            self.after(0, lambda cd=disp_col_defs, rd=all_dicts, igc=item_group_code:
                       self._catalog_template_done(cd, rd, igc, None))
        except Exception as e:
            self.after(0, lambda err=str(e):
                       self._catalog_template_done(None, None, item_group_code, err))

    def _catalog_template_done(self, col_defs, row_dicts, item_group_code, error):
        """UI thread: store template data + rebuild cols + call _catalog_filter."""
        try:
            self.btn_catalog_load.configure(state="normal")
        except Exception:
            pass

        if error:
            self.lbl_catalog_status.config(text=f"❌  Lỗi: {error}", fg=C["red"])
            self._show_msg("Lỗi tải dữ liệu Template", error, "error")
            return

        # Key by B20Item.Id so _catalog_filter can look up values per leaf node
        self._catalog_template_data    = {rd['_item_id']: rd for rd in row_dicts}
        self._catalog_template_col_defs = col_defs
        self._catalog_mode             = "template"

        # Rebuild column headers (keep #0 visible for tree expand icons)
        self._catalog_rebuild_treeview_cols(col_defs)

        # Render the tree using _catalog_filter (handles both subtree filter + eager build)
        self._catalog_filter()

        try:
            self.btn_catalog_export.configure(state="normal")
        except Exception:
            pass

    def _catalog_rebuild_treeview_cols(self, col_defs):
        """Xóa & xây lại cột Treeview theo col_defs = [(col_key, col_name, w, anchor)]."""
        tree = self.catalog_tree
        tree.delete(*tree.get_children())

        col_ids = tuple(cd[0] for cd in col_defs)
        tree["columns"] = col_ids

        # Cột #0 — giữ để hiện icon expand/collapse
        tree.column("#0", width=72, minwidth=36, stretch=False)
        tree.heading("#0", text="")

        for col_key, col_name, w, a in col_defs:
            stretch = (col_key == 'Name')
            _ca = guess_col_align(col_name)
            tree.column(col_key,  width=w, minwidth=50, anchor=_ca, stretch=stretch)
            tree.heading(col_key, text=col_name, anchor="center")

        self._catalog_current_col_defs = col_defs

    def _catalog_reset_default_cols(self):
        """Khôi phục 13 cột mặc định sau khi rời template mode."""
        tree = self.catalog_tree
        tree.delete(*tree.get_children())
        DEFAULT = (
            "Mã Vật Tư", "Tên Vật Tư", "ĐVT",
            "Quy Cách (mm)", "Bộ Định Khoản", "Mã Ngành Hàng",
            "Cốt Ván", "Chất Liệu",
            "S.Mặt Chính", "Lớp Phủ Chính",
            "S.Mặt Phụ",  "Lớp Phủ Phụ",
            "Độ Dày",
        )
        tree["columns"] = DEFAULT
        tree.column("#0",            width=72,  minwidth=36, stretch=False)
        tree.heading("#0",           text="")
        for _c in DEFAULT:
            tree.heading(_c, text=_c, anchor="center")
        tree.column("Mã Vật Tư",     width=160, anchor="w",      stretch=False)
        tree.column("Tên Vật Tư",    width=200, anchor="w",      stretch=False)
        tree.column("ĐVT",           width=55,  anchor="center", stretch=False)
        tree.column("Quy Cách (mm)", width=130, anchor="w",      stretch=False)
        tree.column("Bộ Định Khoản", width=140, anchor="w",      stretch=False)
        tree.column("Mã Ngành Hàng", width=120, anchor="w",      stretch=False)
        tree.column("Cốt Ván",       width=110, anchor="w",      stretch=False)
        tree.column("Chất Liệu",     width=110, anchor="w",      stretch=False)
        tree.column("S.Mặt Chính",   width=75,  anchor="center", stretch=False)
        tree.column("Lớp Phủ Chính", width=120, anchor="w",      stretch=False)
        tree.column("S.Mặt Phụ",     width=75,  anchor="center", stretch=False)
        tree.column("Lớp Phủ Phụ",   width=120, anchor="w",      stretch=False)
        tree.column("Độ Dày",        width=65,  anchor="center", stretch=False)
        self._catalog_current_col_defs = [
            ("Mã Vật Tư",     "Mã Vật Tư",     160, "w"),
            ("Tên Vật Tư",    "Tên Vật Tư",     200, "w"),
            ("ĐVT",           "ĐVT",             55,  "center"),
            ("Quy Cách (mm)", "Quy Cách (mm)",  130, "w"),
            ("Bộ Định Khoản", "Bộ Định Khoản",  140, "w"),
            ("Mã Ngành Hàng", "Mã Ngành Hàng",  120, "w"),
            ("Cốt Ván",       "Cốt Ván",         110, "w"),
            ("Chất Liệu",     "Chất Liệu",       110, "w"),
            ("S.Mặt Chính",   "S.Mặt Chính",     75,  "center"),
            ("Lớp Phủ Chính", "Lớp Phủ Chính",  120, "w"),
            ("S.Mặt Phụ",     "S.Mặt Phụ",       75,  "center"),
            ("Lớp Phủ Phụ",   "Lớp Phủ Phụ",    120, "w"),
            ("Độ Dày",        "Độ Dày",           65,  "center"),
        ]

    def _catalog_autosize_cols(self, all_rows):
        """Tự động đặt chiều rộng cột = max(header, dữ liệu dài nhất), tối đa 300px."""
        import tkinter.font as tkfont
        tree     = self.catalog_tree
        col_defs = getattr(self, '_catalog_current_col_defs', [])
        if not col_defs:
            return
        data_font   = tkfont.Font(family="Segoe UI", size=10)
        header_font = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        PAD   = 28    # khớp với BOM tab
        MAX_W = 300   # giới hạn tối đa

        is_tmpl   = getattr(self, '_catalog_mode', 'all') == 'template'
        tmpl_data = getattr(self, '_catalog_template_data', {})

        # col_key → tên trường trong all_rows (cho cả 2 mode)
        _row_field = {
            'Mã Vật Tư':     'code',            'Code': 'code',
            'Tên Vật Tư':    'name',            'Name': 'name',
            'ĐVT':           'unit',            'Unit': 'unit',
            'DVT':           'unit',
            'Quy Cách (mm)': 'specification',  'Specification': 'specification',
            'Bộ Định Khoản': 'item_group_name', 'ItemGroupName': 'item_group_name',
            'Mã Ngành Hàng': 'item_catg_name',  'ItemCatgName':  'item_catg_name',
            'Cốt Ván':       'board_core',       'BoardCore':     'board_core',
            'Chất Liệu':     'material',         'Material':      'material',
            'S.Mặt Chính':   'main_face_count',  'MainFaceCount': 'main_face_count',
            'Lớp Phủ Chính': 'main_finish',      'MainFinish':    'main_finish',
            'S.Mặt Phụ':     'sec_face_count',   'SecFaceCount':  'sec_face_count',
            'Lớp Phủ Phụ':   'sec_finish',       'SecondaryFinish': 'sec_finish',
            'Độ Dày':        'norm_rate',         'NormRate':        'norm_rate',
        }

        for col_key, col_name, _, _ in col_defs:
            # heading + sql_names row (ColKey) — lấy cái rộng hơn làm min_w
            min_w = max(header_font.measure(col_name),
                        data_font.measure(col_key)) + PAD

            candidates = []
            field = _row_field.get(col_key)
            if field:
                candidates = [str(r.get(field, '') or '') for r in all_rows]
            if is_tmpl:
                candidates += [str(rd.get(col_key, '') or '') for rd in tmpl_data.values()]

            if candidates:
                longest_val = max(candidates, key=len)
                max_data_w  = data_font.measure(longest_val)
            else:
                max_data_w = 0

            col_w = min(MAX_W, max(min_w, max_data_w + PAD))
            try:
                tree.column(col_key, width=col_w)
            except Exception:
                pass

    def _catalog_export_template_excel(self):
        """Xuất file Excel mẫu 3 dòng header dựa trên ImportTemplateDetail."""
        selected = getattr(self.cmb_item_template, '_current_value', None)
        if not selected:
            try:
                selected = self.cmb_item_template.get()
            except Exception:
                selected = None
        template_id = self._template_id_map.get(selected) if selected else None
        if not template_id:
            self._show_msg("Chưa chọn Template",
                           "Hãy chọn Template trên Combobox trước.", "warning")
            return

        # Lấy tên template để dùng làm tên file gợi ý
        template_label = selected.replace('[', '').replace(']', '').strip()
        safe_name = "".join(c if c.isalnum() or c in ' _-' else '_'
                            for c in template_label)[:40].strip('_ ')

        try:
            conn = self._get_db_conn(timeout_sec=5)
            cur  = conn.cursor()
            cur.execute(
                f"SELECT NameGenAuto FROM {self.DB_TEMPLATE_TABLE} WHERE Id = ?",
                (template_id,)
            )
            tpl_row      = cur.fetchone()
            name_gen_auto = bool(tpl_row[0]) if tpl_row else False

            cur.execute(
                f"SELECT ColIndex, ColGroup, ColName, ColKey, DataType, IsNamePart, "
                f"NameConcatOrder, ISNULL(ColView, ColKey) AS ColView "
                f"FROM {self.DB_TEMPLATE_DETAIL_TABLE} "
                f"WHERE TemplateId = ? ORDER BY ColIndex ASC",
                (template_id,)
            )
            all_cols = cur.fetchall()

            # Giữ tất cả cột (kể cả Code/Name) để ColIndex không bị lệch khi import.
            # Các cột tự sinh sẽ bị ẩn trong Excel thay vì xóa.
            cols_data = list(all_cols)

            # ── Fetch lookup values cho FK fields (ColView != ColKey) ──────────
            import re as _re
            _SAFE = _re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
            cur.execute(
                "SELECT COLUMN_NAME FROM [BOMTool].INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME = N'vB20Item' AND TABLE_SCHEMA = N'dbo'"
            )
            _view_col_set = {r[0] for r in cur.fetchall()}

            col_lookup_values = {}   # col_key → [display_str, ...]
            _seen_lk = set()
            for _row in cols_data:
                _ck  = _row[3]   # ColKey
                _cv  = _row[7]   # ColView
                if (_ck not in _seen_lk
                        and _cv != _ck
                        and _cv in _view_col_set
                        and _SAFE.match(_ck)
                        and _SAFE.match(_cv)):
                    _seen_lk.add(_ck)
                    try:
                        cur.execute(
                            f"SELECT DISTINCT [{_cv}] FROM [BOMTool].dbo.vB20Item "
                            f"WHERE [{_cv}] IS NOT NULL "
                            f"  AND LTRIM(RTRIM(CAST([{_cv}] AS NVARCHAR(500)))) <> N'' "
                            f"ORDER BY [{_cv}]"
                        )
                        _vals = [str(r[0]).strip() for r in cur.fetchall() if r[0] is not None]
                        if _vals:
                            col_lookup_values[_ck] = _vals
                    except Exception:
                        pass
            conn.close()
        except Exception as e:
            self._show_msg("Lỗi tải cấu trúc Template", str(e), "error")
            return

        if not cols_data:
            self._show_msg("Template rỗng",
                           "Template này chưa có cột nào được cấu hình.", "warning")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"Template_{safe_name}.xlsx",
            title="Lưu file Template Excel")
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template NVL"

            # ── Màu sắc ──────────────────────────────────────────────────────
            FILL_GROUP   = PatternFill("solid", fgColor="C5D9F1")   # xanh nhạt (nhóm)
            FILL_HEADER  = PatternFill("solid", fgColor="17375E")   # xanh đậm (tên cột)
            FILL_HINT    = PatternFill("solid", fgColor="FFFACD")   # vàng nhạt (gợi ý)
            FILL_AUTOKEY = PatternFill("solid", fgColor="E2EFDA")   # xanh lá nhạt (tự động)

            FONT_GROUP  = Font(name="Segoe UI", size=10, bold=True,  color="17375E")
            FONT_HEADER = Font(name="Segoe UI", size=10, bold=True,  color="FFFFFF")
            FONT_HINT   = Font(name="Segoe UI", size=9,  italic=True, color="595959")
            FONT_AUTO   = Font(name="Segoe UI", size=9,  italic=True, color="375623")

            ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_side = Side(style="thin", color="AAAAAA")
            BORDER    = Border(left=thin_side, right=thin_side,
                               top=thin_side, bottom=thin_side)

            # ── Dựng 3 dòng header ────────────────────────────────────────────
            row1_group  = []   # ColGroup  (dòng 1)
            row2_name   = []   # ColName   (dòng 2)
            row3_hint   = []   # Hint text (dòng 3)
            col_widths  = []

            for col_idx, col_group, col_name, col_key, data_type, is_name_part, *_ in cols_data:
                row1_group.append(col_group or "")
                row2_name.append(col_name or "")

                # Dòng 3: gợi ý kiểu dữ liệu / vai trò
                if col_key == 'Code':
                    hint = "TỰ ĐỘNG SINH"
                elif col_key == 'Name' and name_gen_auto:
                    hint = "TỰ ĐỘNG SINH"
                elif is_name_part:
                    hint = "Thành phần ghép tên"
                else:
                    hint = data_type or "NVARCHAR"
                row3_hint.append(hint)

                max_len = max(len(col_group or ""), len(col_name or ""), len(hint))
                col_widths.append(min(max(max_len + 2, 14), 36))

            ws.append(row1_group)
            ws.append(row2_name)
            ws.append(row3_hint)

            n_cols = len(cols_data)

            # ── Áp style từng ô ──────────────────────────────────────────────
            _hidden_cols = set()   # excel col numbers bị ẩn
            for c in range(1, n_cols + 1):
                _ck     = cols_data[c - 1][3]   # ColKey
                is_auto = (bool(cols_data[c - 1][5])          # IsNamePart
                           or _ck == 'Code'
                           or (_ck == 'Name' and name_gen_auto))

                cell1 = ws.cell(row=1, column=c)
                cell1.fill   = FILL_GROUP
                cell1.font   = FONT_GROUP
                cell1.alignment = ALIGN_C
                cell1.border = BORDER

                cell2 = ws.cell(row=2, column=c)
                cell2.fill   = FILL_HEADER
                cell2.font   = FONT_HEADER
                cell2.alignment = ALIGN_C
                cell2.border = BORDER

                cell3 = ws.cell(row=3, column=c)
                cell3.fill   = FILL_AUTOKEY if is_auto else FILL_HINT
                cell3.font   = FONT_AUTO    if is_auto else FONT_HINT
                cell3.alignment = ALIGN_C
                cell3.border = BORDER

                _col_letter = get_column_letter(c)
                if _ck == 'Code' or (_ck == 'Name' and name_gen_auto):
                    ws.column_dimensions[_col_letter].hidden = True
                    _hidden_cols.add(c)
                else:
                    ws.column_dimensions[_col_letter].width = col_widths[c - 1]

            # ── Merge cells dòng 1: cùng ColGroup liên tiếp → gộp ───────────
            merge_start = 1
            for i in range(1, n_cols + 1):
                cur_group  = row1_group[i - 1]
                next_group = row1_group[i] if i < n_cols else None
                if next_group != cur_group:
                    if i > merge_start:
                        ws.merge_cells(
                            start_row=1, start_column=merge_start,
                            end_row=1,   end_column=i)
                        # Sau merge chỉ cell góc trên-trái giữ value
                        ws.cell(row=1, column=merge_start).value   = cur_group
                        ws.cell(row=1, column=merge_start).fill    = FILL_GROUP
                        ws.cell(row=1, column=merge_start).font    = FONT_GROUP
                        ws.cell(row=1, column=merge_start).alignment = ALIGN_C
                    merge_start = i + 1

            # ── Chiều cao dòng ────────────────────────────────────────────────
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 36

            # ── Freeze panes dưới 3 dòng header ─────────────────────────────
            ws.freeze_panes = "A4"

            # ── Sheet ẩn _Lookup + Data Validation dropdown ───────────────────
            if col_lookup_values:
                from openpyxl.worksheet.datavalidation import DataValidation as _DV

                # col_key → index cột trong _Lookup sheet (1-based, theo thứ tự xuất hiện)
                _lk_col_map = {ck: i + 1
                               for i, ck in enumerate(col_lookup_values.keys())}
                ws_lk = wb.create_sheet("_Lookup")
                ws_lk.sheet_state = 'hidden'

                for _ck, _vals in col_lookup_values.items():
                    _lc = _lk_col_map[_ck]
                    ws_lk.cell(row=1, column=_lc, value=_ck)
                    for _ri, _v in enumerate(_vals, start=2):
                        ws_lk.cell(row=_ri, column=_lc, value=_v)

                # Gán DataValidation cho từng cột trong sheet data
                for _c_pos, _row in enumerate(cols_data, start=1):
                    _ck = _row[3]
                    if _ck not in col_lookup_values:
                        continue
                    _lc       = _lk_col_map[_ck]
                    _n_vals   = len(col_lookup_values[_ck])
                    _lk_letter = get_column_letter(_lc)
                    _formula  = f"_Lookup!${_lk_letter}$2:${_lk_letter}${_n_vals + 1}"
                    _dv = _DV(
                        type="list", formula1=_formula,
                        showDropDown=False,
                        showErrorMessage=True,
                        errorStyle="warning",
                        errorTitle="Giá trị không có trong danh sách",
                        error="Giá trị bạn nhập không khớp với bất kỳ lựa chọn nào. "
                              "Import vẫn sẽ cố resolve — kiểm tra lại nếu bị lỗi."
                    )
                    _col_letter = get_column_letter(_c_pos)
                    _dv.sqref   = f"{_col_letter}4:{_col_letter}1048576"
                    ws.add_data_validation(_dv)

            wb.save(path)
            n_visible = n_cols - len(_hidden_cols)
            self._show_export_success(
                "Xuất Template thành công",
                f"Template  {selected}\n"
                f"→  {n_visible} cột hiển thị ({len(_hidden_cols)} cột ẩn tự sinh)\n"
                f"→  {path}",
                path)
        except Exception as e:
            self._show_msg("Lỗi xuất Template Excel", str(e), "error")

    # ── Import Template NVL — helpers ─────────────────────────────────────────

    @staticmethod
    def _catalog_generate_next_code(code_pattern: str, cur) -> tuple:
        """Parse CodePattern (e.g. 'NVL01_XXXXXX') → query MAX → (next_code, prefix, digits, seq)."""
        if '_' not in code_pattern:
            raise ValueError(f"CodePattern không hợp lệ (thiếu dấu _): {code_pattern!r}")
        prefix, x_part = code_pattern.split('_', 1)
        digits = len(x_part)
        # Escape '_' trong LIKE bằng [_] để tránh wildcard
        like_pat = f"{prefix}[_]{'_' * digits}"
        cur.execute(
            "SELECT MAX(Code) FROM [B10_Boho_Data].dbo.B20Item WHERE Code LIKE ?",
            (like_pat,)
        )
        row = cur.fetchone()
        max_code = row[0] if row and row[0] else None
        if max_code:
            try:
                seq = int(max_code[-digits:]) + 1
            except (ValueError, IndexError):
                seq = 1
        else:
            seq = 1
        next_code = f"{prefix}_{seq:0{digits}d}"
        return next_code, prefix, digits, seq

    @staticmethod
    def _catalog_build_lookup_dicts(details: list, view_col_set: set, cur) -> dict:
        """
        Với mỗi detail có ColView != ColKey và ColView tồn tại trong vB20Item,
        build dict: lookup[ColKey] = {UPPER(display_value): id_value}.
        Keys được normalize UPPER để so sánh case-insensitive.
        Dual lookup: nếu col_key kết thúc bằng 'Id', cũng thêm Code → id.
        """
        import re
        _SAFE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
        lookup = {}
        seen = set()
        for d in details:
            col_key  = d['ColKey']
            col_view = d['ColView']
            if col_key in seen:
                continue
            if (col_view and col_view != col_key
                    and col_view in view_col_set
                    and _SAFE.match(col_key)
                    and _SAFE.match(col_view)):
                seen.add(col_key)
                try:
                    cur.execute(
                        f"SELECT DISTINCT [{col_key}], [{col_view}] "
                        f"FROM [BOMTool].dbo.vB20Item "
                        f"WHERE [{col_key}] IS NOT NULL "
                        f"  AND [{col_view}] IS NOT NULL "
                        f"  AND LTRIM(RTRIM(CAST([{col_view}] AS NVARCHAR(500)))) <> N''"
                    )
                    # Keys → UPPER+strip để so khớp case-insensitive từ Excel
                    lookup[col_key] = {
                        str(r[1]).strip().upper(): r[0]
                        for r in cur.fetchall()
                        if r[0] is not None
                    }
                except Exception:
                    pass

                # Dual lookup: ItemCatgId → cũng map ItemCatgCode → id (setdefault: Name thắng)
                if col_key.endswith('Id') and col_key in lookup:
                    code_col = col_key[:-2] + 'Code'
                    if code_col in view_col_set and _SAFE.match(code_col):
                        try:
                            cur.execute(
                                f"SELECT DISTINCT [{col_key}], [{code_col}] "
                                f"FROM [BOMTool].dbo.vB20Item "
                                f"WHERE [{col_key}] IS NOT NULL "
                                f"  AND [{code_col}] IS NOT NULL "
                                f"  AND LTRIM(RTRIM(CAST([{code_col}] AS NVARCHAR(500)))) <> N''"
                            )
                            for r in cur.fetchall():
                                if r[0] is not None and r[1] is not None:
                                    lookup[col_key].setdefault(
                                        str(r[1]).strip().upper(), r[0])
                        except Exception:
                            pass
        return lookup

    def _catalog_select_import_file(self):
        """Mở file dialog, load file Excel vào SheetTable preview."""
        path = filedialog.askopenfilename(
            filetypes=[("Excel Workbook", "*.xlsx")],
            title="Chọn file Excel để import")
        if not path:
            return
        self._catalog_import_filepath = path
        import os as _os
        short_name = _os.path.basename(path)
        try:
            self.lbl_import_filename.configure(text=f"📄  {short_name}")
        except Exception:
            pass
        self.lbl_import_status.config(text="⏳  Đang đọc file...", fg=C["yellow"])
        threading.Thread(
            target=self._catalog_load_excel_preview_worker,
            args=(path,), daemon=True).start()

    def _catalog_load_excel_preview_worker(self, filepath: str):
        """Background: đọc Excel → trả headers + data rows + hidden cols về UI thread."""
        try:
            import openpyxl as _xl
            from openpyxl.utils import get_column_letter
            wb = _xl.load_workbook(filepath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self.after(0, lambda: self._catalog_load_excel_preview_done([], [], [], None))
                return
            headers = [str(c) if c is not None else "" for c in rows[0]]
            data    = [[str(c) if c is not None else "" for c in r] for r in rows[1:]]
            # Phát hiện cột bị ẩn trong Excel (0-based index cho tksheet)
            hidden_cols = [
                ci - 1
                for ci in range(1, len(headers) + 1)
                if ws.column_dimensions.get(get_column_letter(ci), None)
                and ws.column_dimensions[get_column_letter(ci)].hidden
            ]
            self.after(0, lambda h=headers, d=data, hc=hidden_cols:
                       self._catalog_load_excel_preview_done(h, d, hc, None))
        except Exception as e:
            self.after(0, lambda err=str(e):
                       self._catalog_load_excel_preview_done([], [], [], err))

    def _catalog_load_excel_preview_done(self, headers, data, hidden_cols, error):
        """UI thread: nạp data vào SheetTable.
        Excel structure:
          row 0 = headers (group header row, e.g. 'NHÓM THÔNG TIN, B, C...')
          row 1 = field names (e.g. 'MÃ, MÃ NHÓM VẬT TƯ, TÊN...')
          row 2 = data types (NVARCHAR...)
          row 3+ = actual data
        Hiển thị row 1 làm column headers, row 2+ làm data rows.
        """
        if error:
            self.lbl_import_status.config(text=f"❌  {error}", fg=C["red"])
            return
        if not headers:
            self.lbl_import_status.config(text="⚠️  File rỗng hoặc không đọc được", fg=C["yellow"])
            return
        # _catalog_import_row0 → Excel row 1 (lưu cho worker cần đúng offset)
        self._catalog_import_row0 = headers
        # Dùng row 2 làm display headers (tên field cụ thể)
        display_headers = data[0] if data else headers
        # data rows bắt đầu từ row 3 (data types + actual data)
        display_data    = data[1:] if len(data) > 1 else []
        self._catalog_import_sheet_headers = display_headers
        self._catalog_import_sheet_data    = display_data
        if self.catalog_import_sheet is not None:
            try:
                self.catalog_import_sheet.load(display_headers, display_data)
                # Highlight dòng 0 (kiểu dữ liệu / hint) như row sql_names ở BOM tab
                if display_data:
                    self.catalog_import_sheet.mark_row(0, "sql_names")
                    self.catalog_import_sheet.italic_rows_by_tag("sql_names")
                # Ẩn các cột bị hide trong Excel — data vẫn giữ nguyên vị trí
                # để import worker map theo col_idx không bị lệch
                if hidden_cols:
                    self.catalog_import_sheet.hide_columns_by_index(hidden_cols)
                # Auto-fit width theo nội dung, giới hạn 90px–230px
                self.catalog_import_sheet.autofit_columns(min_w=90, max_w=230)
            except Exception as e:
                self.lbl_import_status.config(text=f"❌  Lỗi hiển thị: {e}", fg=C["red"])
                return
        n_rows = len(display_data)
        self.lbl_import_status.config(
            text=f"✅  Đã tải {n_rows} dòng — có thể chỉnh sửa trước khi Import",
            fg=C["green"])
        # Enable Import/Preview nếu đã có template được chọn
        selected = getattr(self.cmb_item_template, '_current_value', None)
        if not selected:
            try:
                selected = self.cmb_item_template.get()
            except Exception:
                selected = None
        if selected and selected in self._template_id_map:
            try:
                self.btn_import_template.configure(state="normal")
            except Exception:
                pass
            try:
                self.btn_preview_sql.configure(state="normal")
            except Exception:
                pass

    def _catalog_import_template_excel(self):
        """Entry point: validate selection, lấy dữ liệu từ SheetTable, start background import."""
        selected = getattr(self.cmb_item_template, '_current_value', None)
        if not selected:
            try:
                selected = self.cmb_item_template.get()
            except Exception:
                selected = None
        template_id = self._template_id_map.get(selected) if selected else None
        if not template_id:
            self._show_msg("Chưa chọn Template",
                           "Hãy chọn Template trên Combobox trước.", "warning")
            return

        filepath = getattr(self, '_catalog_import_filepath', None)
        if not filepath:
            self._show_msg("Chưa chọn file",
                           "Hãy bấm 'Chọn file Excel' trước.", "warning")
            return

        # Lấy dữ liệu hiện tại từ SheetTable (người dùng có thể đã chỉnh sửa)
        sheet_rows = None
        if self.catalog_import_sheet is not None:
            try:
                updated_data = self.catalog_import_sheet.sheet.get_sheet_data()
                # sheet_rows layout: [row0(Excel1)] + [row1(Excel2/fieldnames)] + [row2+(Excel3+)]
                # worker reads sheet_rows[3:] as actual data (Excel row 4+)
                sheet_rows = ([self._catalog_import_row0]
                              + [self._catalog_import_sheet_headers]
                              + [list(r) for r in updated_data])
            except Exception:
                sheet_rows = None

        try:
            self.btn_import_template.configure(state="disabled")
        except Exception:
            pass
        try:
            self.btn_preview_sql.configure(state="disabled")
        except Exception:
            pass
        self.lbl_import_status.config(text="⏳  Đang import...", fg=C["yellow"])

        threading.Thread(
            target=self._catalog_import_worker,
            args=(filepath, template_id, False, sheet_rows),
            daemon=True
        ).start()

    def _catalog_import_worker(self, filepath: str, template_id: int, dry_run: bool = False, sheet_rows=None):
        """Background thread: Phase 0-7 — load config, read Excel, INSERT B20Item + B20ItemInfo."""
        import openpyxl as _xl

        results = []  # [(row_num, code, status, message)]

        try:
            conn = self._get_db_conn(timeout_sec=10)
            cur  = conn.cursor()

            # ── Phase 0: Load template info ──────────────────────────────────
            cur.execute(
                f"SELECT TemplateCode, NameGenAuto, CodePattern, ItemGroupCode "
                f"FROM {self.DB_TEMPLATE_TABLE} WHERE Id = ?",
                (template_id,)
            )
            tpl_row = cur.fetchone()
            if not tpl_row:
                raise RuntimeError(f"Không tìm thấy template Id={template_id}")
            tpl_code, name_gen_auto, code_pattern, item_group_code = tpl_row
            name_gen_auto = bool(name_gen_auto)

            # Resolve ParentId từ ItemGroupCode → B20Item.Code → B20Item.Id
            tpl_parent_id = None
            if item_group_code:
                cur.execute(
                    "SELECT TOP 1 Id FROM [B10_Boho_Data].dbo.B20Item "
                    "WHERE Code = ? AND IsGroup = 1",
                    (item_group_code,)
                )
                _prow = cur.fetchone()
                if _prow:
                    tpl_parent_id = _prow[0]

            cur.execute(
                f"SELECT ColIndex, ColGroup, ColName, ColKey, "
                f"ISNULL(ColView, ColKey) AS ColView, "
                f"DataType, ISNULL(IsNamePart,0) AS IsNamePart, "
                f"ISNULL(NameConcatOrder,9999) AS NameConcatOrder, "
                f"ISNULL(TargetTable, 'B20Item') AS TargetTable "
                f"FROM {self.DB_TEMPLATE_DETAIL_TABLE} "
                f"WHERE TemplateId = ? ORDER BY ColIndex ASC",
                (template_id,)
            )
            details = [
                {
                    'ColIndex': r[0], 'ColGroup': r[1], 'ColName': r[2],
                    'ColKey':   r[3], 'ColView':  r[4],
                    'DataType': r[5], 'IsNamePart': bool(r[6]),
                    'NameConcatOrder': r[7], 'TargetTable': r[8]
                }
                for r in cur.fetchall()
            ]

            # ── Phase 1: Build lookup dicts ───────────────────────────────────
            # Phải dùng [BOMTool].INFORMATION_SCHEMA vì vB20Item nằm trong BOMTool
            cur.execute(
                "SELECT COLUMN_NAME FROM [BOMTool].INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME = N'vB20Item' AND TABLE_SCHEMA = N'dbo'"
            )
            view_col_set = {r[0] for r in cur.fetchall()}
            lookup = self._catalog_build_lookup_dicts(details, view_col_set, cur)

            # ── Phase 2: Table routing — đọc từ TargetTable trong ImportTemplateDetail
            # ColKey → bảng đích; higher ColIndex ghi đè nếu cùng ColKey
            target_map = {d['ColKey']: d['TargetTable'] for d in details}

            # ── Phase 4 (pre-loop): Sinh code đầu tiên ───────────────────────
            next_code, prefix, digits, seq = self._catalog_generate_next_code(
                code_pattern, cur)

            # ── Phase 3: Đọc Excel (hoặc lấy từ SheetTable nếu sheet_rows có sẵn) ─
            col_map = {d['ColIndex']: d for d in details}
            name_parts_sorted = sorted(
                [d for d in details if d['IsNamePart']],
                key=lambda d: d['NameConcatOrder']
            )

            if sheet_rows is not None:
                # sheet_rows[0] = headers row (Excel row 1), data bắt đầu từ index 3 (row 4)
                _data_rows = sheet_rows[3:] if len(sheet_rows) > 3 else []
                _row_iter  = [(i + 4, tuple(r)) for i, r in enumerate(_data_rows)]
            else:
                try:
                    wb = _xl.load_workbook(filepath, data_only=True)
                    ws = wb.active
                except Exception as e:
                    raise RuntimeError(f"Lỗi mở file Excel: {e}") from e
                _row_iter = list(enumerate(ws.iter_rows(min_row=4, values_only=True), start=4))

            for row_num, row_values in _row_iter:

                if all(v is None or str(v).strip() == '' for v in row_values):
                    continue

                row_data   = {}   # ColKey → resolved value (higher ColIndex wins)
                row_errors = []

                for col_idx, value in enumerate(row_values, start=1):
                    detail = col_map.get(col_idx)
                    if detail is None:
                        continue
                    col_key  = detail['ColKey']
                    col_view = detail['ColView']

                    # Skip auto-generated fields
                    if col_key == 'Code':
                        continue
                    if col_key == 'Name' and name_gen_auto:
                        continue

                    if value is None or str(value).strip() == '':
                        # Higher ColIndex wins → chỉ ghi None nếu chưa có giá trị
                        if col_key not in row_data:
                            row_data[col_key] = None
                        continue

                    str_val = str(value).strip()

                    # Lookup resolution cho FK fields — key normalize UPPER
                    lookup_key = str_val.upper()
                    if col_key in lookup:
                        resolved = lookup[col_key].get(lookup_key)
                        if resolved is None:
                            row_errors.append(
                                f"Giá trị '{str_val}' ở cột [{col_key}] "
                                f"không tồn tại trong danh mục tra cứu")
                            row_data[col_key] = None
                        else:
                            row_data[col_key] = resolved  # higher ColIndex wins
                    else:
                        if col_view != col_key and col_view in view_col_set:
                            # ColView tồn tại trong view nhưng lookup không build được → lỗi thật
                            row_errors.append(
                                f"Cột [{col_key}] là FK (ColView={col_view}) "
                                f"nhưng không build được lookup — kiểm tra BOMTool/vB20Item")
                            row_data[col_key] = None
                        elif col_view != col_key:
                            # ColView không có trong view → insert text trực tiếp (ví dụ Unit là text)
                            row_data[col_key] = str_val
                        elif col_key == 'Name' and not name_gen_auto:
                            row_data[col_key] = str_val.upper()
                        elif detail['DataType'] == 'DECIMAL':
                            try:
                                row_data[col_key] = float(str_val)
                            except ValueError:
                                row_data[col_key] = str_val
                        else:
                            row_data[col_key] = str_val  # higher ColIndex wins

                # ── Phase 5: Sinh Name ────────────────────────────────────────
                if name_gen_auto:
                    parts = []
                    for np in name_parts_sorted:
                        v = row_data.get(np['ColKey'])
                        if v is not None and str(v).strip():
                            parts.append(str(v).strip())
                    name = ' '.join(parts).strip()
                else:
                    name = str(row_data.get('Name') or '').strip()

                if not name:
                    results.append((row_num, '?', 'ERROR', 'Tên vật tư trống'))
                    continue

                if row_errors:
                    results.append((row_num, '?', 'ERROR', '; '.join(row_errors)))
                    continue

                current_code = next_code

                # ── Phase 6: Kiểm tra trùng code ─────────────────────────────
                cur.execute(
                    "SELECT 1 FROM [B10_Boho_Data].dbo.B20Item WHERE Code = ?",
                    (current_code,)
                )
                if cur.fetchone():
                    results.append((row_num, current_code, 'SKIP', 'Code đã tồn tại'))
                    seq      += 1
                    next_code = f"{prefix}_{seq:0{digits}d}"
                    continue

                # ── Phase 7: INSERT ───────────────────────────────────────────
                b20_data  = {}
                info_data = {}

                for col_key, val in row_data.items():
                    if col_key in ('Code', 'Name') or val is None:
                        continue
                    target = target_map.get(col_key)
                    if target == 'B20ItemInfo':
                        info_data[col_key] = val
                    elif target == 'B20Item':
                        b20_data[col_key] = val
                    # TargetTable=None hoặc không khai báo → bỏ qua

                # Id tự sinh bởi DB (DEFAULT NEWID()), không cần insert
                b20_data.update({
                    'ParentId': tpl_parent_id,
                    'Code':     current_code,
                    'Name':     name,
                    'IsGroup':  0,
                    'IsActive': 1,
                })

                b20_cols = list(b20_data.keys())
                b20_vals = [b20_data[c] for c in b20_cols]
                cols_str = ', '.join(f'[{c}]' for c in b20_cols)

                if dry_run:
                    def _fv(v):
                        if v is None:                   return 'NULL'
                        if isinstance(v, bool):         return '1' if v else '0'
                        if isinstance(v, (int, float)): return repr(v)
                        return "N'" + str(v).replace("'", "''") + "'"
                    sql = (
                        f"-- Row {row_num}: {current_code}\n"
                        f"INSERT INTO [B10_Boho_Data].dbo.B20Item\n"
                        f"  ({cols_str})\n"
                        f"OUTPUT INSERTED.[Id]\n"
                        f"VALUES ({', '.join(_fv(v) for v in b20_vals)});"
                    )
                    if info_data:
                        ic = ['ItemId'] + list(info_data.keys())
                        iv_preview = ['<<B20Item.Id>>'] + [
                            _fv(info_data[c]) for c in info_data
                        ]
                        sql += (
                            f"\n-- Dùng Id vừa OUTPUT để set ItemId\n"
                            f"INSERT INTO [B10_Boho_Data].dbo.B20ItemInfo\n"
                            f"  ({', '.join(f'[{c}]' for c in ic)})\n"
                            f"VALUES ({', '.join(iv_preview)});"
                        )
                    results.append((row_num, current_code, 'SQL', sql))
                else:
                    try:
                        ph_str = ', '.join('?' for _ in b20_cols)
                        cur.execute(
                            f"INSERT INTO [B10_Boho_Data].dbo.B20Item "
                            f"({cols_str}) "
                            f"OUTPUT INSERTED.[Id] "
                            f"VALUES ({ph_str})",
                            b20_vals
                        )
                        new_id_row = cur.fetchone()
                        new_id = new_id_row[0] if new_id_row else None

                        if info_data:
                            ic = ['ItemId'] + list(info_data.keys())
                            iv = [new_id]   + [info_data[c] for c in info_data]
                            cur.execute(
                                f"INSERT INTO [B10_Boho_Data].dbo.B20ItemInfo "
                                f"({', '.join(f'[{c}]' for c in ic)}) "
                                f"VALUES ({', '.join('?' for _ in ic)})",
                                iv
                            )

                        conn.commit()
                        results.append((row_num, current_code, 'OK', ''))

                    except Exception as db_err:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        results.append((row_num, current_code, 'ERROR',
                                        str(db_err)[:220]))

                seq      += 1
                next_code = f"{prefix}_{seq:0{digits}d}"

            conn.close()

        except Exception as fatal:
            self.after(0, lambda e=str(fatal), dr=dry_run: self._catalog_import_done([], e, dr))
            return

        self.after(0, lambda r=results, dr=dry_run: self._catalog_import_done(r, None, dr))

    def _catalog_import_done(self, results: list, fatal_error, dry_run: bool = False):
        """UI thread: re-enable button, show result dialog, reload tree nếu có insert."""
        try:
            self.btn_import_template.configure(state="normal")
        except Exception:
            pass
        try:
            self.btn_preview_sql.configure(state="normal")
        except Exception:
            pass

        if dry_run:
            if fatal_error:
                self.lbl_import_status.config(text="❌  Preview thất bại", fg=C["red"])
                self._show_msg("Lỗi Preview SQL", fatal_error, "error")
            else:
                self.lbl_import_status.config(text="✅  SQL Preview sẵn sàng", fg=C["green"])
                self._catalog_show_sql_preview(results)
            return

        if fatal_error:
            self.lbl_import_status.config(text="❌  Import thất bại", fg=C["red"])
            self._show_msg("Lỗi Import Template", fatal_error, "error")
            return

        if not results:
            self.lbl_import_status.config(text="⚠️  File không có dữ liệu", fg=C["yellow"])
            self._show_msg("Import Template",
                           "File không có dòng dữ liệu (dữ liệu bắt đầu từ Row 4).",
                           "warning")
            return

        n_ok   = sum(1 for r in results if r[2] == 'OK')
        n_skip = sum(1 for r in results if r[2] == 'SKIP')
        n_err  = sum(1 for r in results if r[2] == 'ERROR')

        status_color = C["green"] if n_err == 0 else C["yellow"]
        status_icon  = "✅" if n_err == 0 else "⚠️"
        self.lbl_import_status.config(
            text=f"{status_icon}  Import: {n_ok} OK / {n_skip} bỏ qua / {n_err} lỗi",
            fg=status_color
        )

        self._catalog_show_import_results(results, n_ok, n_skip, n_err)

        if n_ok > 0:
            try:
                val = (getattr(self.cmb_item_template, '_current_value', None)
                       or self.cmb_item_template.get())
                self._on_catalog_template_change(val)
            except Exception:
                pass

    def _catalog_preview_sql(self):
        """Entry point: giống Import nhưng dry_run=True — gen SQL và hiển thị, không INSERT."""
        selected = getattr(self.cmb_item_template, '_current_value', None)
        if not selected:
            try:
                selected = self.cmb_item_template.get()
            except Exception:
                selected = None
        template_id = self._template_id_map.get(selected) if selected else None
        if not template_id:
            self._show_msg("Chưa chọn Template",
                           "Hãy chọn Template trên Combobox trước.", "warning")
            return

        filepath = getattr(self, '_catalog_import_filepath', None)
        if not filepath:
            self._show_msg("Chưa chọn file",
                           "Hãy bấm 'Chọn file Excel' trước.", "warning")
            return

        # Lấy dữ liệu hiện tại từ SheetTable
        sheet_rows = None
        if self.catalog_import_sheet is not None:
            try:
                updated_data = self.catalog_import_sheet.sheet.get_sheet_data()
                # sheet_rows layout: [row0(Excel1)] + [row1(Excel2/fieldnames)] + [row2+(Excel3+)]
                # worker reads sheet_rows[3:] as actual data (Excel row 4+)
                sheet_rows = ([self._catalog_import_row0]
                              + [self._catalog_import_sheet_headers]
                              + [list(r) for r in updated_data])
            except Exception:
                sheet_rows = None

        try:
            self.btn_import_template.configure(state="disabled")
        except Exception:
            pass
        try:
            self.btn_preview_sql.configure(state="disabled")
        except Exception:
            pass
        self.lbl_import_status.config(text="⏳  Đang gen SQL...", fg=C["yellow"])

        threading.Thread(
            target=self._catalog_import_worker,
            args=(filepath, template_id, True, sheet_rows),
            daemon=True
        ).start()

    def _catalog_show_sql_preview(self, results: list):
        """Hiển thị SQL INSERT preview (dry-run) trong dialog có scroll + Copy All."""
        dt = THEMES[ctk.get_appearance_mode()]
        import tkinter as _tk

        dlg = ctk.CTkToplevel(self)
        dlg.title("Preview SQL — Import Template")
        dlg.resizable(True, True)
        dlg.transient(self)
        dlg.grab_set()

        # ── Header ───────────────────────────────────────────────────────────────
        n_sql  = sum(1 for _, _, s, _ in results if s == 'SQL')
        n_skip = sum(1 for _, _, s, _ in results if s == 'SKIP')
        n_err  = sum(1 for _, _, s, _ in results if s == 'ERROR')

        hdr = _tk.Frame(dlg, bg="#0F172A")
        hdr.pack(fill=_tk.X)
        _tk.Label(hdr,
                  text=f"📋  {n_sql} câu INSERT  |  {n_skip} bỏ qua  |  {n_err} lỗi",
                  font=FONT_MD_B,
                  bg="#0F172A", fg="#E2E8F0", pady=10).pack(side=_tk.LEFT, padx=16)

        def _copy_all():
            all_sql = '\n\n'.join(msg for _, _, s, msg in results if s == 'SQL')
            dlg.clipboard_clear()
            dlg.clipboard_append(all_sql)

        CButton(hdr, text="📋 Copy All", command=_copy_all,
                font=ctk.CTkFont(*FONT_SMALL),
                fg_color="#1E3A5F", hover_color=dt["btn_primary"],
                text_color="#DBEAFE", height=28, corner_radius=5,
                width=110).pack(side=_tk.RIGHT, padx=PAD_MD, pady=PAD_SM)

        # ── Text area ─────────────────────────────────────────────────────────────
        txt_frame = _tk.Frame(dlg, bg="#0D1117")
        txt_frame.pack(fill=_tk.BOTH, expand=True, padx=10, pady=(4, 0))

        vsb = _tk.Scrollbar(txt_frame, orient=_tk.VERTICAL)
        hsb = _tk.Scrollbar(dlg, orient=_tk.HORIZONTAL)
        txt = _tk.Text(
            txt_frame,
            font=("Consolas", 9),
            bg="#0D1117", fg="#E6EDF3",
            selectbackground="#264F78",
            insertbackground="#E6EDF3",
            wrap=_tk.NONE,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            relief=_tk.FLAT, bd=0, padx=10, pady=PAD_SM
        )
        vsb.configure(command=txt.yview)
        hsb.configure(command=txt.xview)
        vsb.pack(side=_tk.RIGHT, fill=_tk.Y)
        txt.pack(side=_tk.LEFT, fill=_tk.BOTH, expand=True)
        hsb.pack(fill=_tk.X, padx=10, pady=(0, 4))

        txt.tag_configure("comment", foreground="#8B949E")
        txt.tag_configure("kw",      foreground="#FF7B72",
                          font=("Consolas", 9, "bold"))
        txt.tag_configure("tbl",     foreground="#79C0FF")
        txt.tag_configure("val",     foreground="#A5D6FF")
        txt.tag_configure("col",     foreground="#D2A8FF")
        txt.tag_configure("skip",    foreground="#FCD34D")
        txt.tag_configure("err",     foreground=dt["log_err"])

        def _insert_sql_line(line):
            ls = line.strip()
            if ls.startswith('--'):
                txt.insert(_tk.END, line + '\n', "comment")
            elif ls.upper().startswith('INSERT'):
                idx = line.upper().find('INSERT')
                txt.insert(_tk.END, line[:idx] + 'INSERT', "kw")
                txt.insert(_tk.END, line[idx + 6:] + '\n', "tbl")
            elif ls.upper().startswith('VALUES'):
                idx = line.upper().find('VALUES')
                txt.insert(_tk.END, line[:idx] + 'VALUES', "kw")
                txt.insert(_tk.END, line[idx + 6:] + '\n', "val")
            elif ls.startswith('(') and (ls.endswith(');') or ls.endswith(')')):
                txt.insert(_tk.END, line + '\n', "col")
            else:
                txt.insert(_tk.END, line + '\n')

        for row_num, code, status, msg in results:
            if status == 'SQL':
                for line in msg.split('\n'):
                    _insert_sql_line(line)
                txt.insert(_tk.END, '\n')
            elif status == 'SKIP':
                txt.insert(_tk.END,
                            f"-- Row {row_num}: {code}  [SKIP] {msg}\n\n", "skip")
            elif status == 'ERROR':
                txt.insert(_tk.END,
                            f"-- Row {row_num}: {code}  [ERROR] {msg}\n\n", "err")

        txt.configure(state=_tk.DISABLED)

        # ── Footer ────────────────────────────────────────────────────────────────
        CButton(dlg, text="Đóng", command=dlg.destroy,
                font=ctk.CTkFont(*FONT_BODY_B),
                height=32, corner_radius=6,
                fg_color=("#3B82F6","#1D4ED8"),
                hover_color=("#2563EB","#1E40AF"),
                text_color="white").pack(pady=(4, 10), padx=PAD_MD, fill=_tk.X)

        # ── Size & center ─────────────────────────────────────────────────────────
        dlg.update_idletasks()
        w, h = 960, min(140 + n_sql * 90, 680)
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
        dlg.wait_window()

    def _catalog_show_import_results(self, results, n_ok, n_skip, n_err):
        """Hiển thị dialog cuộn kết quả import từng dòng."""
        dt = THEMES[ctk.get_appearance_mode()]
        dlg = ctk.CTkToplevel(self)
        dlg.title("Kết quả Import Template")
        dlg.resizable(True, True)
        dlg.transient(self)
        dlg.grab_set()

        # ── Summary bar ──────────────────────────────────────────────────────
        summary_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        summary_frame.pack(fill=tk.X, padx=PAD_MD, pady=(10, 4))

        def _chip(parent, text, fg, bg):
            ctk.CTkLabel(parent, text=text,
                         font=ctk.CTkFont(*FONT_MD_B),
                         text_color=fg, fg_color=bg,
                         corner_radius=6, padx=10, pady=PAD_XS).pack(
                side=tk.LEFT, padx=PAD_XS)

        _chip(summary_frame, f"✅  {n_ok} Thành công",  "#065F46", "#D1FAE5")
        _chip(summary_frame, f"⚠️  {n_skip} Bỏ qua",    "#92400E", "#FEF3C7")
        _chip(summary_frame, f"❌  {n_err} Lỗi",         "#7F1D1D", "#FEE2E2")

        # ── Scrollable text ───────────────────────────────────────────────────
        txt_frame = ctk.CTkFrame(dlg, fg_color=("gray90", "#1E1E1E"), corner_radius=6)
        txt_frame.pack(fill=tk.BOTH, expand=True, padx=PAD_MD, pady=PAD_XS)

        import tkinter as _tk
        txt = _tk.Text(txt_frame, wrap="none",
                       font=("Consolas", 10),
                       bg=dt["bg_main"], fg="#D4D4D4",
                       relief="flat", bd=0,
                       selectbackground="#264F78")
        sb_y = _tk.Scrollbar(txt_frame, orient="vertical",   command=txt.yview)
        sb_x = _tk.Scrollbar(txt_frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side=_tk.RIGHT, fill=_tk.Y)
        sb_x.pack(side=_tk.BOTTOM, fill=_tk.X)
        txt.pack(fill=_tk.BOTH, expand=True)

        txt.tag_configure("ok",   foreground="#4ADE80")
        txt.tag_configure("skip", foreground="#FCD34D")
        txt.tag_configure("err",  foreground=dt["log_err"])
        txt.tag_configure("hdr",  foreground="#94A3B8")

        txt.insert("end", f"{'Dòng':<6}  {'Code':<18}  Trạng thái\n", "hdr")
        txt.insert("end", "─" * 72 + "\n", "hdr")
        for row_num, code, status, msg in results:
            if status == 'OK':
                tag  = "ok"
                icon = "✅"
            elif status == 'SKIP':
                tag  = "skip"
                icon = "⚠️"
            else:
                tag  = "err"
                icon = "❌"
            line = f"{str(row_num):<6}  {code:<18}  {icon} {status}"
            if msg:
                line += f"  —  {msg}"
            txt.insert("end", line + "\n", tag)

        txt.configure(state="disabled")

        # ── Close button ──────────────────────────────────────────────────────
        CButton(dlg, text="Đóng", command=dlg.destroy,
                font=ctk.CTkFont(*FONT_BODY_B),
                height=32, corner_radius=6,
                fg_color=("#3B82F6","#1D4ED8"),
                hover_color=("#2563EB","#1E40AF"),
                text_color="white").pack(pady=(4, 10), padx=PAD_MD, fill=tk.X)

        # ── Size & center ─────────────────────────────────────────────────────
        dlg.update_idletasks()
        w, h = 640, min(120 + len(results) * 18 + 80, 560)
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
        dlg.wait_window()

    # ─ Tab 3: Tổng hợp THDM ───────────────────────────────────────────────────────────────────

    def _on_tab_changed(self):
        """Khi switch sang tab THDM mà chưa có dữ liệu → tự tải.
        Luôn đóng tất cả _SearchCombo popup để tránh ghost dropdown khi chuyển tab."""
        # Đóng tất cả popup đang mở
        for attr in ('cmb_creator', 'cmb_thdm_product', 'cmb_thdm_order',
                     'cmb_thdm_creator', 'cmb_thdm_period'):
            combo = getattr(self, attr, None)
            if combo is not None:
                try:
                    combo._close_popup()
                except Exception:
                    pass
        try:
            tab_name = self.nb.get()
        except Exception:
            return
        # So khớp CHÍNH XÁC với TAB_THDM (không dùng substring "THDM in ..."
        # — dễ vỡ mỗi khi đổi tên tab, xem ghi chú tại TAB_THDM đầu file)
        if tab_name == TAB_THDM and not self._thdm_product_map:
            self._thdm_load_products()
        if tab_name == TAB_CATALOG and not self._template_id_map:
            self._catalog_load_templates()

    def _build_tab_thdm(self):
        tab = self.nb.tab(TAB_THDM)
        dt = THEMES[ctk.get_appearance_mode()]

        # ── Action bar ngang ─────────────────────────────────────────────────
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=56,
                           corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_thdm_load = ctk.CTkButton(bar, text="🔄  ① Tải dữ liệu",
            command=self._thdm_load_products,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color=("#EBEBEB", "#2D2D2D"),
            text_color=("#1A1A1A", "#E1E1E1"),
            hover_color=("#D8D8D8", "#3A3A3A"),
            border_width=1, border_color=("#CCCCCC", "#3C3C3C"),
            height=32, corner_radius=8)
        self.btn_thdm_load.pack(side=tk.LEFT, padx=(0,10), pady=PAD_MD)

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,10), pady=10)

        # Dự án
        CLabel(bar, text="Dự án:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_product = _SearchCombo(bar,
            values=["— Chọn dự án —"],
            width=230, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._thdm_on_product_change,
            placeholder="— Chọn dự án —")
        self.cmb_thdm_product.set("— Chọn dự án —")
        self.cmb_thdm_product.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,10), pady=10)

        # Đơn hàng
        CLabel(bar, text="Đơn hàng:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_order = _SearchCombo(bar,
            values=["— Chọn đơn hàng —"],
            width=240, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._thdm_on_order_change,
            placeholder="— Chọn đơn hàng —",
            state="disabled")
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self.cmb_thdm_order.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        CLabel(bar, text="Nhân viên:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_creator = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=200, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._on_thdm_creator_change,
            placeholder="— Chọn nhân viên —")
        self.cmb_thdm_creator.set("— Đang tải... —")
        self.cmb_thdm_creator.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        CLabel(bar, text="Đợt:",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_period = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=180, height=32,
            font=ctk.CTkFont(*FONT_BODY),
            command=self._on_thdm_period_change,
            placeholder="— Chọn đợt —")
        self.cmb_thdm_period.set("— Đang tải... —")
        self.cmb_thdm_period.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        self.lbl_thdm_status = CLabel(bar, text="—  Chưa tải",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_status.pack(side=tk.LEFT)

        # ── Main area: PanedWindow kéo-thả điều chỉnh tỉ lệ trái/phải ──────
        main = tk.PanedWindow(tab, orient=tk.HORIZONTAL,
                              bg=dt["border"], bd=0,
                              sashwidth=5, sashrelief="flat",
                              opaqueresize=True)
        main.pack(fill=tk.BOTH, expand=True)

        # ── LEFT: BOM list ────────────────────────────────────────────────────
        lf = tk.Frame(main, bg=dt["bg_deep"])
        lf.rowconfigure(1, weight=1)
        lf.columnconfigure(0, weight=1)
        main.add(lf, minsize=150, stretch="always")

        # Filter
        fe = ctk.CTkFrame(lf, fg_color=("gray85","gray20"), height=32, corner_radius=0)
        fe.grid(row=0, column=0, sticky="ew")
        fe.pack_propagate(False)
        self.thdm_search_var = tk.StringVar()
        self.thdm_search_var.trace_add("write", lambda *_: self._thdm_filter_bom())
        ctk.CTkEntry(fe, textvariable=self.thdm_search_var,
            font=ctk.CTkFont(*FONT_BODY),
            placeholder_text="🔍  Tìm BOM...",
            height=30, corner_radius=0, border_width=0,
            fg_color="transparent").pack(fill=tk.X, padx=6, pady=1)

        # BOM treeview
        btf = tk.Frame(lf, bg=dt["bg_deep"])
        btf.grid(row=1, column=0, sticky="nsew")
        btf.rowconfigure(0, weight=1)
        btf.columnconfigure(0, weight=1)

        # ── tksheet (Excel-like grid) ─────────────────────────────────────────
        if _HAS_TKSHEET:
            _sheet_theme = "light" if ctk.get_appearance_mode() == "Light" else "dark"
            self.thdm_bom_sheet = tksheet.Sheet(
                btf,
                headers=["✓", "Mục số", "Mã BOM", "Tên SP", "Version"],
                data=[],
                theme=_sheet_theme,
                show_row_index=False,
                show_top_left=False,
                row_height=30,
                header_height=34,
                show_horizontal_grid=True,
                show_vertical_grid=True,
                font=FONT_SMALL,
                header_font=FONT_SMALL_B,
            )
            self.thdm_bom_sheet.enable_bindings(
                "single_select", "drag_select", "row_select",
                "column_width_resize", "arrowkeys",
            )
            try:
                self.thdm_bom_sheet.set_options(
                    table_bg=dt["sheet_table_bg"],
                    table_fg=dt["sheet_table_fg"],
                    header_bg=dt["sheet_header_bg"],
                    header_fg=dt["sheet_header_fg"],
                    index_bg=dt["sheet_index_bg"],
                    index_fg=dt["sheet_index_fg"],
                    top_left_bg=dt["sheet_header_bg"],
                    show_horizontal_grid=True,
                    show_vertical_grid=True,
                    table_grid_fg=dt["sheet_grid"],
                    header_grid_fg=dt["sheet_grid"],
                    index_grid_fg=dt["sheet_grid"],
                    outline_color=dt["sheet_grid"],
                )
            except Exception:
                pass
            # Căn chỉnh data cells theo loại dữ liệu (header đã center toàn bộ)
            _bom_cols = ["✓", "Mục số", "Mã BOM", "Tên SP", "Version"]
            for _bi, _bc in enumerate(_bom_cols):
                _bal = guess_col_align(_bc)
                if _bal != 'w':
                    try:
                        self.thdm_bom_sheet.align_columns(columns=[_bi], align=_bal)
                    except Exception:
                        pass
            # Header luôn căn giữa
            try:
                self.thdm_bom_sheet.align_header(
                    columns=list(range(len(_bom_cols))), align='center', redraw=False
                )
            except Exception:
                pass
            # Bind trên MT (main table canvas) — hoạt động độc lập với tksheet version
            self.thdm_bom_sheet.MT.bind(
                "<ButtonRelease-1>",
                lambda e: self.after(40, self._thdm_process_sheet_click)
            )
            self.thdm_bom_sheet.grid(row=0, column=0, columnspan=2, sticky="nsew")
            self.thdm_bom_sheet.column_width(column=0, width=30)
            self.thdm_bom_sheet.column_width(column=1, width=70)
            self.thdm_bom_sheet.column_width(column=2, width=120)
            self.thdm_bom_sheet.column_width(column=3, width=200)
            self.thdm_bom_sheet.column_width(column=4, width=70)
            self.thdm_bom_tree = None   # alias — fallback path sử dụng
            # Cột "Tên SP" tự giãn lấp hết chiều ngang khi resize cửa sổ
            self._thdm_bom_stretch_job = None
            def _bom_sheet_on_resize(_e=None):
                if self._thdm_bom_stretch_job:
                    try:
                        self.after_cancel(self._thdm_bom_stretch_job)
                    except Exception:
                        pass
                self._thdm_bom_stretch_job = self.after(
                    100, self._thdm_bom_autostretch)
            self.thdm_bom_sheet.bind("<Configure>", _bom_sheet_on_resize)
        else:
            # Fallback: ttk.Treeview khi tksheet chưa cài
            BOM_COLS = ("✓", "Mục số", "Mã BOM", "Tên SP", "Version")
            vsb_b = ttk.Scrollbar(btf, orient=tk.VERTICAL, style="Slim.Vertical.TScrollbar")
            self.thdm_bom_tree = ttk.Treeview(btf, style="BOM.Treeview", show="headings",
                columns=BOM_COLS, yscrollcommand=vsb_b.set, selectmode="extended")
            vsb_b.configure(command=self.thdm_bom_tree.yview)
            self.thdm_bom_tree.heading("✓", text="✓", anchor="center")
            self.thdm_bom_tree.column("✓", width=26, minwidth=26, anchor="center", stretch=False)
            for _bc, _ba, _bw, _bs in [
                ("Mục số", "w", 70, False), ("Mã BOM", "w", 120, False),
                ("Tên SP",  "w", 200, True),  ("Version", "w", 70, False),
            ]:
                self.thdm_bom_tree.heading(_bc, text=_bc, anchor=_ba)
                self.thdm_bom_tree.column(_bc, width=_bw, minwidth=50, anchor=_ba, stretch=_bs)
            self.thdm_bom_tree.tag_configure("checked",   foreground=dt["log_ok"])
            self.thdm_bom_tree.tag_configure("unchecked", foreground=dt["text_muted"])
            self.thdm_bom_tree.grid(row=0, column=0, sticky="nsew")
            vsb_b.grid(row=0, column=1, sticky="ns")
            self.thdm_bom_tree.bind("<ButtonRelease-1>", self._thdm_on_bom_click)
            self.thdm_bom_sheet = None

        # Bottom bar left
        lb = ctk.CTkFrame(lf, fg_color=("gray85","gray18"), height=36, corner_radius=0)
        lb.grid(row=2, column=0, sticky="ew")
        lb.pack_propagate(False)

        self.btn_thdm_sel_all = ctk.CTkButton(lb, text="☑ Tất cả",
            command=lambda: self._thdm_toggle_all(True),
            font=ctk.CTkFont(*FONT_SMALL),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            border_color=("gray50","gray55"),
            hover_color=("gray80","gray25"),
            width=80, height=26, corner_radius=8)
        self.btn_thdm_sel_all.pack(side=tk.LEFT, padx=(6,2), pady=5)

        self.btn_thdm_desel_all = ctk.CTkButton(lb, text="☐ Bỏ",
            command=lambda: self._thdm_toggle_all(False),
            font=ctk.CTkFont(*FONT_SMALL),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            border_color=("gray50","gray55"),
            hover_color=("gray80","gray25"),
            width=60, height=26, corner_radius=8)
        self.btn_thdm_desel_all.pack(side=tk.LEFT, padx=(0,6), pady=5)

        self.lbl_thdm_sel_count = CLabel(lb, text="0 BOM chọn",
            font=ctk.CTkFont(*FONT_SMALL),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_sel_count.pack(side=tk.LEFT)

        # Xanh dương (hành động xử lý) — xanh lá dành riêng cho nút INSERT
        self.btn_thdm_aggregate = ctk.CTkButton(lb, text="📊  ③ Tổng hợp",
            command=self._thdm_aggregate,
            font=ctk.CTkFont(*FONT_BODY_B),
            fg_color=("#0066CC", "#0066CC"),
            hover_color=("#0055AA", "#0055AA"),
            text_color="#FFFFFF",
            height=26, corner_radius=8, state="normal")
        self.btn_thdm_aggregate.pack(side=tk.RIGHT, padx=6, pady=5)

        # ── RIGHT: File picker + Preview ─────────────────────────────────────
        rf = tk.Frame(main, bg=dt["bg_main"])
        main.add(rf, minsize=200, stretch="always")
        self.after(80, lambda: main.sash_place(0, 350, 0))

        # File picker row — pack TRÊN đầu
        fp = ctk.CTkFrame(rf, fg_color=("gray85","gray20"), height=42, corner_radius=0)
        fp.pack(side=tk.TOP, fill=tk.X)
        fp.pack_propagate(False)

        self.btn_thdm_pick_excel = ctk.CTkButton(fp, text="📂  ② Chọn file Excel THDM",
            command=self._thdm_pick_excel,
            font=ctk.CTkFont(*FONT_BODY),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            border_color=("gray50","gray55"),
            hover_color=("gray80","gray25"),
            height=28, corner_radius=5)
        self.btn_thdm_pick_excel.pack(side=tk.LEFT, padx=(6,8), pady=7)

        ctk.CTkFrame(fp, fg_color=("gray65","gray30"),
                     width=1, height=24).pack(side=tk.LEFT, padx=(0,8), pady=9)

        self.lbl_thdm_excel_path = CLabel(fp, text="Chưa chọn file",
            font=ctk.CTkFont(*FONT_SMALL),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_excel_path.pack(side=tk.LEFT)

        # ── Content area — switches between pre-preview and result ────────────
        # (pack() được gọi SAU khi ab đã pack(BOTTOM) để không bị che khuất)
        ca = tk.Frame(rf, bg=dt["bg_main"])
        ca.rowconfigure(0, weight=1)
        ca.columnconfigure(0, weight=1)
        self._thdm_content_area = ca

        # ── PRE-PREVIEW frame (hiện sau khi chọn file, trước khi Tổng hợp) ──
        pref = tk.Frame(ca, bg=dt["bg_main"])
        pref.grid(row=0, column=0, sticky="nsew")
        pref.rowconfigure(1, weight=1)
        pref.columnconfigure(0, weight=1)
        self._thdm_pre_frame = pref

        self._lbl_thdm_pre_status = tk.Label(pref,
            text="⏳  Đang đọc file Excel...",
            font=FONT_BODY, fg=dt["text_muted"], bg=dt["bg_main"], anchor="w")
        self._lbl_thdm_pre_status.grid(row=0, column=0, sticky="ew", padx=PAD_MD, pady=(8, 4))

        self._thdm_pre_nb = ttk.Notebook(pref)
        self._thdm_pre_nb.grid(row=1, column=0, sticky="nsew", padx=PAD_XS, pady=(0, 4))

        # Tab 1 — Raw Excel
        raw_tab = tk.Frame(self._thdm_pre_nb, bg=dt["bg_panel"])
        self._thdm_pre_nb.add(raw_tab, text="  📄 Raw Excel  ")
        raw_tab.rowconfigure(0, weight=1)
        raw_tab.columnconfigure(0, weight=1)
        if _HAS_TKSHEET:
            self.thdm_raw_tree = SheetTable(raw_tab, columns=("_ph",))
            self.thdm_raw_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb_r = ttk.Scrollbar(raw_tab, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
            hsb_r = ttk.Scrollbar(raw_tab, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
            self.thdm_raw_tree = ttk.Treeview(raw_tab, style="BOM.Treeview", show="headings",
                columns=("_ph",), yscrollcommand=vsb_r.set, xscrollcommand=hsb_r.set)
            vsb_r.configure(command=self.thdm_raw_tree.yview)
            hsb_r.configure(command=self.thdm_raw_tree.xview)
            self.thdm_raw_tree.grid(row=0, column=0, sticky="nsew")
            vsb_r.grid(row=0, column=1, sticky="ns")
            hsb_r.grid(row=1, column=0, sticky="ew")

        # (Tab "Đã xử lý" đã bỏ hẳn — chỉ xem Raw trước khi Tổng hợp,
        #  dữ liệu đã xử lý xem ở màn hình kết quả đa tab sau khi Tổng hợp.)

        pref.grid_remove()   # ẩn cho đến khi file được chọn

        # ── RESULT frame (hiển thị sau khi Tổng hợp) ─────────────────────────
        ptf = tk.Frame(ca, bg=dt["bg_panel"], bd=1, relief="flat")
        ptf.grid(row=0, column=0, sticky="nsew")
        ptf.rowconfigure(0, weight=0)   # filter bar — không mở rộng
        ptf.rowconfigure(1, weight=1)   # notebook chiếm phần còn lại
        ptf.columnconfigure(0, weight=1)
        self._thdm_result_frame = ptf

        # ── Filter bar ─────────────────────────────────────────────────────────
        _fb = tk.Frame(ptf, bg=dt["tv_heading_bg"])
        _fb.grid(row=0, column=0, sticky="ew")
        self._thdm_filter_seg = ctk.CTkSegmentedButton(
            _fb,
            values=["Tất cả (0)", "⚠ Dòng lỗi (0)"],
            command=self._thdm_toggle_filter,
            font=ctk.CTkFont(*FONT_BODY),
            height=26, corner_radius=6,
        )
        self._thdm_filter_seg.pack(side=tk.LEFT, padx=PAD_SM, pady=2)
        self._thdm_filter_seg.set("Tất cả (0)")

        # ── Notebook kết quả: Đầu phiếu + mỗi section detail 1 tab (giống BOM)
        self._thdm_result_nb = ttk.Notebook(ptf)
        self._thdm_result_nb.grid(row=1, column=0, sticky="nsew")

        def _mk_result_grid(title, columns):
            tabf = tk.Frame(self._thdm_result_nb, bg=dt["bg_panel"])
            self._thdm_result_nb.add(tabf, text=title)
            tabf.rowconfigure(0, weight=1)
            tabf.columnconfigure(0, weight=1)
            if _HAS_TKSHEET:
                tr = SheetTable(tabf, columns=columns)
                tr.grid(row=0, column=0, sticky="nsew")
            else:
                _v = ttk.Scrollbar(tabf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
                _h = ttk.Scrollbar(tabf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
                tr = ttk.Treeview(tabf, style="BOM.Treeview", show="headings",
                    columns=columns, yscrollcommand=_v.set, xscrollcommand=_h.set)
                _v.configure(command=tr.yview)
                _h.configure(command=tr.xview)
                tr.grid(row=0, column=0, sticky="nsew")
                _v.grid(row=0, column=1, sticky="ns")
                _h.grid(row=1, column=0, sticky="ew")
            tr.tag_configure("row_normal", foreground=dt["text_main"])
            tr.tag_configure("row_alt",    foreground=dt["text_muted"])
            tr.tag_configure("err_row",  background=dt["badge_err_bg"],  foreground=dt["badge_err_fg"])
            tr.tag_configure("warn_row", background=dt["badge_warn_bg"], foreground=dt["badge_warn_fg"])
            tr.tag_configure("sql_names", background=dt["field_row_bg"], foreground=dt["field_row_fg"])
            return tr

        # Tab 0 — Header (THDM_HEADER, resolve display-only)
        # Hiển thị NGANG giống BOM: mỗi field 1 cột, cột gán động khi fill
        self._thdm_header_tree = _mk_result_grid("  📋 Header  ", ())

        # Tab 1..n — mỗi child section của THDM_HEADER một tab
        self._thdm_sec_trees    = {}   # sec → tree
        self._thdm_sec_col_defs = {}   # sec → col defs
        for _sec, _info in self.mapping.get('_CONFIG', {}).items():
            if _info.get('parent_section') != 'THDM_HEADER':
                continue
            # Cột định danh = SQL_Column (duy nhất), header text = Ten_Excel
            _defs = _thdm_get_detail_col_defs(self.mapping.get(_sec, []))
            _cols = tuple(d['sql_col'] for d in _defs)
            _lbl  = _info.get('label') or _sec
            _tr   = _mk_result_grid(f"  {_lbl}  ", _cols)
            for d in _defs:
                # header rỗng → tksheet hiện chữ cái cột (P,Q,R...);
                # cột hệ thống không có Ten_Excel thì dùng SQL_Column làm header
                _htxt = d['ten_excel'] or d['sql_col']
                _tr.heading(d['sql_col'], text=_htxt, anchor='center')
                _tr.column(d['sql_col'], width=d['width'], anchor=d['anchor'],
                           stretch=d['stretch'], minwidth=50)
            # Căn giữa riêng hàng phụ SQL_Column (2 dòng header nhìn cho gọn),
            # dữ liệu vẫn giữ căn lề theo kiểu cột
            if hasattr(_tr, 'center_rows_by_tag'):
                _tr.center_rows_by_tag("sql_names")
            self._thdm_sec_trees[_sec]    = _tr
            self._thdm_sec_col_defs[_sec] = _defs

        # Alias tương thích code cũ: tree chính = section đầu tiên (THDM_THVT)
        self.thdm_preview_tree = (next(iter(self._thdm_sec_trees.values()), None)
                                  or self._thdm_header_tree)

        # Bind click / double-click cho statusbar + Quick Map (Feature 1 & 2).
        #
        # "cell_select" / "rc_select" → extra_bindings API; event.selected.row
        #   là 0-based row index trong tksheet data.
        #
        # Double-click → "double_click_cell" không tồn tại trong BINDING_TO_ATTR
        #   của tksheet 7.6.0; phải đặt MT.extra_double_b1_func trực tiếp.
        #   Callback nhận raw tkinter event (.x, .y); dùng MT.identify_row(y=.y)
        #   để lấy row index. edit_cell bật để người dùng sửa trực tiếp trên ô;
        #   _thdm_mt_dblclick gọi hide_text_editor() trước khi mở Quick Map.
        for _sb, _tb in self._thdm_sec_trees.items():
            if isinstance(_tb, SheetTable):
                _tb.sheet.extra_bindings([
                    ("cell_select",  lambda ev, s=_sb: self._thdm_mt_click(ev, s)),
                    ("rc_select",    lambda ev, s=_sb: self._thdm_mt_click(ev, s)),
                    ("end_edit_cell", lambda ev, s=_sb: self._thdm_on_cell_edited(ev, s)),
                    ("ctrl_v",       lambda ev, s=_sb: self._thdm_on_paste(ev, s)),
                ])
                _tb.sheet.MT.extra_double_b1_func = (
                    lambda ev, s=_sb: self._thdm_mt_dblclick(ev, s))
            else:
                _tb.bind("<ButtonRelease-1>",
                         lambda ev, s=_sb: self._thdm_mt_click(ev, s))
                _tb.bind("<Double-1>",
                         lambda ev, s=_sb: self._thdm_mt_dblclick(ev, s))

        # Action bar bottom — pack TRƯỚC ca để không bị che khuất bởi sheet
        ab = ctk.CTkFrame(rf, fg_color="transparent", height=44)
        ab.pack(side=tk.BOTTOM, fill=tk.X, pady=(4, 0))
        ab.pack_propagate(False)

        # Tạm ẩn "Xuất Excel" / "Xem SQL" theo yêu cầu — user không cần,
        # gây rối vì không nằm trong luồng ①→⑤. Widget vẫn tạo (không .pack())
        # để các chỗ .configure(state=...) khác trong code không lỗi;
        # muốn hiện lại chỉ cần thêm .pack(...) như 2 nút bên dưới.
        self.btn_thdm_export_xl = CButton(ab, text="📤  Xuất Excel",
            command=self._thdm_export_excel,
            font=ctk.CTkFont(*FONT_MD),
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            border_color=("gray50", "gray55"),
            hover_color=("gray90", "gray25"),
            width=120, height=30, corner_radius=8, state="disabled")

        self.btn_thdm_view_sql = CButton(ab, text="📋  Xem SQL",
            command=self._thdm_view_sql,
            font=ctk.CTkFont(*FONT_MD),
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            border_color=("gray50", "gray55"),
            hover_color=("gray90", "gray25"),
            width=100, height=30, corner_radius=8, state="disabled")

        self.btn_thdm_validate = CButton(ab, text="🔍  ④ Kiểm tra",
            command=self._thdm_validate,
            font=ctk.CTkFont(*FONT_MD_B),
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6", "#60A5FA"),
            border_color=("#3B82F6", "#60A5FA"),
            hover_color=("#DBEAFE", "#1E3A5F"),
            width=110, height=30, corner_radius=8, state="disabled")
        self.btn_thdm_validate.pack(side=tk.LEFT, padx=(0, 4))
        Tooltip(self.btn_thdm_validate,
                lambda: "Soát dòng thiếu Mã VTTB hoặc mã không có trong danh mục"
                        " — phải đạt 0 lỗi mới INSERT được")

        # btn_thdm_view_sql ẩn theo yêu cầu — không pack()
        Tooltip(self.btn_thdm_view_sql,
                lambda: "Xem câu SQL INSERT sẽ thực thi — dùng để đối chiếu dữ liệu trước khi tạo THDM")

        # INSERT — Primary Blue, đồng nhất với nút Import bên tab BOM
        self.btn_thdm_insert = CButton(ab, text="💾  ⑤ Tạo THDM",
            command=self._thdm_insert_db,
            font=ctk.CTkFont(*FONT_MD_B),
            fg_color=("#0066CC", "#0066CC"),
            hover_color=("#0055AA", "#0055AA"),
            text_color="#FFFFFF",
            text_color_disabled="#6B9CD6",
            width=210, height=30, corner_radius=8, state="disabled")
        self.btn_thdm_insert.pack(side=tk.RIGHT)

        # Label kết quả tổng hợp/kiểm tra — đặt ở action bar dưới (đủ chỗ,
        # tránh tràn như khi để trên thanh lookup)
        self.lbl_thdm_result = CLabel(ab, text="",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray40", "gray55"),
            fg_color="transparent")
        self.lbl_thdm_result.pack(side=tk.LEFT, padx=(10, 6))

        # Content area pack SAU ab — đảm bảo ab luôn hiển thị dưới cùng
        ca.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # ── Empty state (hiện khi chưa chọn gì, giống tab Import BOM) ────────
        self._thdm_result_frame.grid_remove()
        _ef = tk.Frame(ca, bg=dt["bg_main"])
        _ef.place(relx=0.5, rely=0.45, anchor="center")
        tk.Label(_ef, text="📋", bg=dt["bg_main"], fg="#3E3E42",
                 font=FONT_HERO).pack()
        tk.Label(_ef, text="①  Tải dữ liệu và chọn Dự án / Đơn hàng / Nhân viên / Đợt",
                 bg=dt["bg_main"], fg="#555555", font=FONT_LABEL_N).pack(pady=(4, 0))
        tk.Label(_ef, text="Tick chọn BOM ở danh sách bên trái  →  ②  Chọn file Excel THDM",
                 bg=dt["bg_main"], fg="#3E3E42", font=FONT_BODY).pack(pady=(2, 0))
        tk.Label(_ef, text="③ Tổng hợp  →  ④ Kiểm tra  →  ⑤ Tạo THDM",
                 bg=dt["bg_main"], fg="#3E3E42", font=FONT_BODY).pack(pady=(2, 0))
        self._thdm_empty_frame = _ef

        # Internal state
        self._thdm_all_bom_rows      = []    # rows BOM từ DB
        self._thdm_checked_ids       = set() # BOM id đang checked
        self._thdm_preview_data      = []    # rows parsed từ Excel
        self._thdm_excel_path        = None  # đường dẫn file Excel đã chọn
        self._thdm_product_map       = {}    # ProductId → Name
        self._thdm_order_map         = {}    # display → BizDocId (string, dùng INSERT + BOM filter)
        self._thdm_period_map        = {}    # PeriodId → Name
        self._thdm_selected_product_id  = None
        self._thdm_selected_order_id    = None   # BizDocId string (ví dụ "11034510FO")
        self._thdm_selected_period_id   = None
        self._thdm_sheet_row_ids        = []    # row index → BOM Id (tksheet)
        self._thdm_period_active_map    = {}    # display → PeriodId (filtered by order)
        self._thdm_valid_builtin_orders = None  # set of BuiltinOrder0 cho đợt đang chọn
        self._thdm_validated            = False # đã Kiểm tra đạt chưa
        self._thdm_val_errors           = {}    # (section, idx) → [msg,...]
        self._thdm_validate_caches      = {}    # {sec: {cache_key: entries}}
        self._thdm_lk_recs              = {}    # {sec: [lk_rec,...]}
        self._thdm_val_lk_fails         = {}    # {(sec,idx): [(sql_col,raw,key,rec),...]}
        self._thdm_filter_mode          = "all" # "all" | "errors_only"
        self._thdm_sec_visible          = {}    # {sec: [data_idx,...]} mapping hiển thị
        self._fuzzy_resolutions         = {}    # {(raw_str, cache_key): db_id}

    # ── THDM: B00Lookup helper ─────────────────────────────────────────────────

    @staticmethod
    def _bravo_view_to_table(name: str) -> str:
        """Chuyển tên view BRAVO sang tên bảng gốc.

        Quy tắc BRAVO: v{TênBảng}_{HậuTố...}
          vB20Product_Edit      → B20Product
          vB30BizDocSO_Edit1    → B30BizDocSO
          vB20BOM_Edit          → B20BOM
          B20Product            → B20Product  (đã là bảng, giữ nguyên)
        """
        import re
        if name and re.match(r'^v[A-Z]', name):
            base = name[1:]                    # bỏ chữ 'v' đầu
            base = re.sub(r'_.*$', '', base)   # bỏ _HậuTố
            return base
        return name

    @staticmethod
    def _b00lookup_safe_disp_col(cur, table: str, disp_col: str,
                                 sort_str: str = '', extra_str: str = '') -> str:
        """Tìm cột display an toàn trong bảng gốc.

        DisplayMember trong B00Lookup đôi khi là cột computed chỉ có trong view
        (VD: DocInfo, ItemInfo). Hàm này kiểm tra cột bằng SELECT TOP 0 rồi
        fallback theo thứ tự:
          1. disp_col gốc
          2. Các cột common BRAVO hay dùng làm display
          3. Các cột trong Sort (lấy tên, bỏ ASC/DESC)
          4. Các cột trong ExtraMemberList
        """
        def col_exists(col_name: str) -> bool:
            try:
                cur.execute(f"SELECT TOP 0 [{col_name}] FROM {table}")
                return True
            except Exception:
                return False

        if col_exists(disp_col):
            return disp_col

        # Fallback 1: cột display phổ biến trong BRAVO
        for fb in ('DocNo', 'Code', 'Name', 'Description'):
            if col_exists(fb):
                return fb

        # Fallback 2: lấy tên cột từ Sort (VD: "DocDate DESC,DocNo DESC")
        import re
        for token in re.split(r'[,\s]+', sort_str or ''):
            token = token.strip()
            if token.upper() in ('ASC', 'DESC', ''):
                continue
            if col_exists(token):
                return token

        # Fallback 3: ExtraMemberList
        for token in (extra_str or '').split(','):
            token = token.strip()
            if token and col_exists(token):
                return token

        return disp_col  # trả về gốc, để query tự báo lỗi

    def _b00lookup_build_query(self, lookup_key, extra_where=None, extra_params=None):
        """Đọc B00Lookup → build câu SELECT động trên bảng gốc (không qua view).

        BRAVO rules được áp dụng:
        - HiddenValueMember ưu tiên hơn ValueMember (là id thực).
        - DisplayMember đôi khi là cột computed chỉ có trong view → tự động
          fallback sang cột thực gần nhất (DocNo, Name, Code, ...).
        - FilterKey dùng format BRAVO riêng → caller truyền extra_where thay thế.
        """
        conn = self._get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT LookupTable, ValueMember, DisplayMember,
                   HiddenValueMember, Sort, ExtraMemberList
            FROM   B10_Boho.dbo.B00Lookup
            WHERE  LookupKey = ?
        """, (lookup_key,))
        row = cur.fetchone()
        if not row:
            conn.close()
            raise ValueError(
                f"Không tìm thấy LookupKey='{lookup_key}' trong B10_Boho.dbo.B00Lookup")

        raw_table, val_col, disp_col, hidden_val, sort_str, extra_str = row

        # Ưu tiên HiddenValueMember (id thực); ValueMember đôi khi rỗng
        actual_val  = (hidden_val or '').strip() or (val_col  or '').strip()
        actual_disp = (disp_col  or '').strip()

        if not actual_val:
            conn.close()
            raise ValueError(
                f"B00Lookup '{lookup_key}': ValueMember và HiddenValueMember đều rỗng")
        if not raw_table:
            conn.close()
            raise ValueError(f"B00Lookup '{lookup_key}': LookupTable rỗng")

        # Chuyển view name → bảng gốc
        table = self._bravo_view_to_table(raw_table)

        # === TỰ ĐỘNG CHECK QUYỀN / TỒN TẠI ĐỂ FALLBACK DB ===
        try:
            cur.execute(f"SELECT TOP 0 1 FROM [{table}]")
        except Exception as e:
            err_msg = str(e).lower()
            # 229: Denied Permission, 208: Invalid Object Name
            if "229" in err_msg or "208" in err_msg or "denied" in err_msg:
                table = f"B10_Boho.dbo.{table}"
        # ====================================================

        # DisplayMember có thể là cột computed của view → tìm cột an toàn
        safe_disp = self._b00lookup_safe_disp_col(
            cur, table, actual_disp, sort_str or '', extra_str or '')

        conn.close()

        where  = f" WHERE {extra_where}" if extra_where else ""
        params = list(extra_params or [])
        sql = (f"SELECT {actual_val}, {safe_disp} "
               f"FROM {table}{where} "
               f"ORDER BY 2")
        return sql, params

    # ── BOM: Load Người lập ───────────────────────────────────────────────────

    def _load_creator_combo(self):
        """Load danh sách UserList từ B00Lookup vào cmb_creator (background thread)."""
        threading.Thread(target=self._load_creator_worker, daemon=True).start()

    def _load_creator_worker(self, _retry=0):
        try:
            sql, params = self._b00lookup_build_query(
                "Employee",
                extra_where="IsGroup = 0 AND IsActive = 1",
            )
            # Inject cột Code vào SELECT để hiển thị "Name — Code"
            # sql dạng: "SELECT Id, Name FROM <table> WHERE ... ORDER BY 2"
            sql = sql.replace(" FROM ", ", Code FROM ", 1)
            conn = self._get_db_conn()
            cur  = conn.cursor()
            try:
                cur.execute(sql, params)
                rows = cur.fetchall()   # [(EmployeeId, Name, Code), ...]
            except Exception:
                # Fallback: Code column không tồn tại → dùng 2 cột
                sql_fb = sql.replace(", Code FROM ", " FROM ", 1)
                cur.execute(sql_fb, params)
                rows = [(r[0], r[1], None) for r in cur.fetchall()]

            # Map EmployeeId → UserId thật (B00UserList.Id) — CreatedBy phải là
            # UserId, không phải EmployeeId. Nhân viên chưa có tài khoản sẽ
            # fallback về DEFAULT_CREATOR_USER_ID khi build _creator_map.
            emp_to_user = {}
            emp_ids = [r[0] for r in rows if r[0] is not None]
            if emp_ids:
                _ph = ','.join(['?'] * len(emp_ids))
                for _ul_tbl in ('B10_Boho.dbo.B00UserList', 'B00UserList'):
                    try:
                        cur.execute(
                            f"SELECT EmployeeId, Id FROM {_ul_tbl}"
                            f" WHERE EmployeeId IN ({_ph})", emp_ids)
                        emp_to_user = {r[0]: r[1] for r in cur.fetchall()}
                        if emp_to_user:
                            break
                    except Exception:
                        continue

            conn.close()
            self.after(0, lambda d=rows, m=emp_to_user: self._load_creator_done(d, None, m))
        except Exception as e:
            _MAX_RETRY = 10
            if _retry < _MAX_RETRY:
                # Auto-retry sau 5 giây (chờ VPN kết nối)
                self.after(5000, lambda r=_retry: threading.Thread(
                    target=self._load_creator_worker,
                    args=(r + 1,), daemon=True).start())
            else:
                self.after(0, lambda err=e: self._load_creator_done([], str(err), {}))

    def _load_creator_done(self, rows, err, emp_to_user=None):
        if err:
            _err_val = ["⚠  Lỗi tải danh sách — click để thử lại"]
            self.cmb_creator.configure(values=_err_val)
            self.cmb_creator.set(_err_val[0])
            if hasattr(self, 'cmb_thdm_creator'):
                self.cmb_thdm_creator.configure(values=_err_val)
                self.cmb_thdm_creator.set(_err_val[0])
            return
        # rows = [(EmployeeId, Name, Code), ...]
        emp_to_user = emp_to_user or {}
        # Hiển thị "Code | Name", map display → UserId thật (B00UserList.Id),
        # fallback DEFAULT_CREATOR_USER_ID nếu nhân viên chưa có tài khoản.
        display_list = [
            f"{code}  |  {name}" if code else str(name)
            for _, name, code in rows
        ]
        self._creator_map = {
            disp: emp_to_user.get(emp_id, DEFAULT_CREATOR_USER_ID)
            for (emp_id, _, __), disp in zip(rows, display_list)
        }
        # Map phụ: display_text → EmployeeId — dùng để live-query khi cache miss
        self._employee_id_map = {
            disp: emp_id
            for (emp_id, _, __), disp in zip(rows, display_list)
        }
        self.cmb_creator.configure(values=display_list)
        self.cmb_creator.set("— Chọn nhân viên —")
        self._current_creator_user_id = None
        if hasattr(self, 'cmb_thdm_creator'):
            self.cmb_thdm_creator.configure(values=display_list)
            self.cmb_thdm_creator.set("— Chọn nhân viên —")
            self._thdm_creator_user_id = None

    def _lookup_user_id_by_emp(self, emp_id):
        """Query B00UserList lấy UserId từ EmployeeId — chỉ gọi khi cache miss."""
        for sql in (
            "SELECT TOP 1 Id FROM B10_Boho.dbo.B00UserList WHERE EmployeeId = ?",
            "SELECT TOP 1 Id FROM B00UserList WHERE EmployeeId = ?",
        ):
            try:
                conn = self._get_db_conn()
                try:
                    cur = conn.cursor()
                    cur.execute(sql, (emp_id,))
                    row = cur.fetchone()
                    if row:
                        return row[0]
                finally:
                    conn.close()
            except Exception:
                continue
        return None

    def _on_creator_change(self, selected_name):
        if selected_name and ("Lỗi" in selected_name or "Đang tải" in selected_name):
            self._load_creator_combo()
            return
        user_id = getattr(self, '_creator_map', {}).get(selected_name)
        if user_id is None:
            # Cache miss (startup failure / account mới tạo): query trực tiếp DB
            emp_id = getattr(self, '_employee_id_map', {}).get(selected_name)
            if emp_id is not None:
                user_id = self._lookup_user_id_by_emp(emp_id)
            if user_id is None:
                user_id = DEFAULT_CREATOR_USER_ID
        self._current_creator_user_id = user_id
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if selected_name and "|" in selected_name:
            self.cmb_creator.set(selected_name.split("|")[0].strip())

    def _on_thdm_creator_change(self, selected_name):
        if selected_name and ("Lỗi" in selected_name or "Đang tải" in selected_name):
            self._load_creator_combo()
            return
        user_id = getattr(self, '_creator_map', {}).get(selected_name)
        if user_id is None:
            # Cache miss: query trực tiếp DB
            emp_id = getattr(self, '_employee_id_map', {}).get(selected_name)
            if emp_id is not None:
                user_id = self._lookup_user_id_by_emp(emp_id)
            if user_id is None:
                user_id = DEFAULT_CREATOR_USER_ID
        self._thdm_creator_user_id     = user_id
        # EmployeeId riêng: B20Employee.Id (khác UserId) dùng cho THDM_HEADER.EmployeeId
        self._thdm_creator_employee_id = getattr(self, '_employee_id_map', {}).get(selected_name)
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if selected_name and "|" in selected_name:
            self.cmb_thdm_creator.set(selected_name.split("|")[0].strip())

    # ── THDM: Load Dự án ──────────────────────────────────────────────────────

    def _thdm_load_products(self):
        self.btn_thdm_load.configure(state="disabled", text="⏳  Đang tải...")
        self.lbl_thdm_status.config(text="⏳  Đang tải dự án...", fg=self._text_fg)
        threading.Thread(target=self._thdm_load_products_worker, daemon=True).start()

    def _thdm_load_products_worker(self):
        try:
            sql, params = self._b00lookup_build_query("Product1",
                extra_where="IsActive = 1")
            # Inject Code vào SELECT để hiển thị "Code | Name [Id]"
            sql = sql.replace(" FROM ", ", Code FROM ", 1)
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute(sql, params)
            rows = cur.fetchall()   # (Id, Name, Code)
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_products_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_products_done([], str(err)))

    def _thdm_load_products_done(self, rows, error):
        dt = THEMES[ctk.get_appearance_mode()]
        self.btn_thdm_load.configure(state="normal", text="🔄  ① Tải dữ liệu")
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg=dt["log_err"])
            return
        # Hiển thị "Code  |  Name  [Id]", map display → Id
        def _prod_disp(r):
            row_id, name, code = r[0], r[1] or "", r[2] or ""
            return f"{code}  |  {name}  [{row_id}]"
        display_list = [_prod_disp(r) for r in rows]
        self._thdm_product_map = {disp: r[0] for r, disp in zip(rows, display_list)}
        values = ["— Chọn dự án —"] + display_list
        self.cmb_thdm_product.configure(values=values)
        self.cmb_thdm_product.set("— Chọn dự án —")
        self.cmb_thdm_order.configure(state="disabled", values=["— Chọn đơn hàng —"])
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self._thdm_selected_product_id = None
        self._thdm_selected_order_id   = None
        self._thdm_all_bom_rows = []
        self._thdm_checked_ids  = set()
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} dự án — chọn dự án để tiếp tục",
            fg=C["green"])

    def _thdm_on_product_change(self, value=None):
        selected = self.cmb_thdm_product.get()
        product_id = self._thdm_product_map.get(selected)
        if product_id is None:
            self.cmb_thdm_order.configure(state="disabled",
                                          values=["— Chọn đơn hàng —"])
            self.cmb_thdm_order.set("— Chọn đơn hàng —")
            self._thdm_selected_product_id = None
            return
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if "|" in selected:
            self.cmb_thdm_product.set(selected.split("|")[0].strip())
        self._thdm_selected_product_id = product_id
        self._thdm_load_orders(product_id)

    # ── THDM: Load Đơn hàng ───────────────────────────────────────────────────

    def _thdm_load_orders(self, product_id):
        self.cmb_thdm_order.configure(state="disabled")
        self.lbl_thdm_status.config(text="⏳  Đang tải đơn hàng...", fg=self._text_fg)
        threading.Thread(
            target=self._thdm_load_orders_worker,
            args=(product_id,), daemon=True).start()

    def _thdm_load_orders_worker(self, product_id):
        try:
            # Custom SQL: hiển thị DocNo2 | DocDate | Description
            # BizDocId = string dạng "11034510FO", dùng cho INSERT BizDocId_SO và BOM filter
            sql = """
                SELECT BizDocId,
                       ISNULL(DocNo2, DocNo)                        AS DocNo2,
                       CONVERT(varchar(10), DocDate, 103)           AS DocDate,
                       ISNULL(Description, '')                      AS Descr
                FROM   B30BizDocSO
                WHERE  DocCode = 'FO'
                  AND  ProductId = ?
                  AND  IsActive  = 1
                ORDER  BY DocDate DESC, DocNo2
            """
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute(sql, [product_id])
            rows = cur.fetchall()   # (BizDocId, DocNo2, DocDate, Descr)
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_orders_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_orders_done([], str(err)))

    def _thdm_load_orders_done(self, rows, error):
        dt = THEMES[ctk.get_appearance_mode()]
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg=dt["log_err"])
            return
        # Hiển thị "DocNo2  |  DocDate  |  Description"
        # r = (BizDocId, DocNo2, DocDate, Descr) — BizDocId string dạng "11034510FO"
        def _ord_disp(r):
            return f"{r[1]}  |  {r[2]}  |  {r[3]}"
        display_list = [_ord_disp(r) for r in rows]
        self._thdm_order_map = {disp: r[0] for r, disp in zip(rows, display_list)}
        values = ["— Chọn đơn hàng —"] + display_list
        self.cmb_thdm_order.configure(state="normal", values=values)
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self._thdm_selected_order_id = None
        self._thdm_all_bom_rows = []
        self._thdm_checked_ids  = set()
        # Reset Đợt combo khi đơn hàng list thay đổi
        self.cmb_thdm_period.configure(state="disabled", values=["— Chọn đợt —"])
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        self._thdm_period_active_map    = {}
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} đơn hàng",
            fg=C["green"])

    def _thdm_on_order_change(self, value=None):
        selected = self.cmb_thdm_order.get()
        order_id = self._thdm_order_map.get(selected)   # BizDocId string, ví dụ "11034510FO"
        self._thdm_selected_order_id = order_id
        # Reset Đợt và BOM filter khi đổi đơn hàng
        self.cmb_thdm_period.configure(state="disabled", values=["— Chọn đợt —"])
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        self._thdm_period_active_map    = {}
        if order_id is None:
            return
        # Sau khi chọn chỉ hiển thị DocNo2 (phần trước |)
        if "|" in selected:
            self.cmb_thdm_order.set(selected.split("|")[0].strip())
        self._thdm_load_bom_list()
        self._thdm_load_period_by_order(order_id)   # Task 10

    # ── THDM: Load Đợt (Period) — Task 10 ────────────────────────────────────
    # Đợt được lọc theo đơn hàng đang chọn qua B30BizDocDetailFactory

    def _thdm_load_period_by_order(self, order_id):
        """Bắt đầu load danh sách Đợt cho đơn hàng (background)."""
        self.cmb_thdm_period.configure(state="disabled", values=["⏳  Đang tải đợt..."])
        self.cmb_thdm_period.set("⏳  Đang tải đợt...")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        threading.Thread(
            target=self._thdm_period_by_order_worker,
            args=(order_id,), daemon=True).start()

    def _thdm_period_by_order_worker(self, order_id):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            # B00Lookup[Period]: LookupTable=B20Period, HiddenValueMember=Id, DisplayMember=Name
            cur.execute("""
                SELECT DISTINCT f.PeriodId,
                       ISNULL(p.Name, CAST(f.PeriodId AS NVARCHAR(50)))
                FROM   B30BizDocDetailFactory AS f
                LEFT JOIN B20Period AS p ON p.Id = f.PeriodId
                WHERE  f.BizDocId = ?
                  AND  f.PeriodId IS NOT NULL
                ORDER BY ISNULL(p.Name, CAST(f.PeriodId AS NVARCHAR(50)))
            """, [order_id])
            rows = cur.fetchall()   # [(PeriodId, DisplayName), ...]
            conn.close()
            self.after(0, lambda d=rows: self._thdm_period_by_order_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_period_by_order_done([], str(err)))

    def _thdm_period_by_order_done(self, rows, error):
        dt = THEMES[ctk.get_appearance_mode()]
        if error:
            self.cmb_thdm_period.configure(
                state="normal", values=["⚠  Lỗi tải đợt"])
            self.cmb_thdm_period.set("⚠  Lỗi tải đợt")
            # Hiện lỗi thật ra status để debug
            self.lbl_thdm_status.config(text=f"❌  Đợt: {error}", fg=dt["log_err"])
            return
        if not rows:
            self.cmb_thdm_period.configure(values=["— Không có đợt —"])
            self.cmb_thdm_period.set("— Không có đợt —")
            return
        self._thdm_period_active_map = {r[1]: r[0] for r in rows}  # Name → PeriodId
        values = ["— Chọn đợt —"] + [r[1] for r in rows]
        self.cmb_thdm_period.configure(state="normal", values=values)
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id = None

    def _on_thdm_period_change(self, selected_name=None):
        selected = self.cmb_thdm_period.get()
        self._thdm_selected_period_id = self._thdm_period_active_map.get(selected)
        if self._thdm_selected_period_id is None:
            self._thdm_valid_builtin_orders = None
            self._thdm_filter_bom()
            return
        # Task 11: load BuiltinOrder0 cho đợt này → filter BOM grid
        self._thdm_load_builtin_orders_for_period()

    # ── THDM: Filter BOM theo BuiltinOrder0 — Task 11 ─────────────────────────

    def _thdm_load_builtin_orders_for_period(self):
        """Query DISTINCT BuiltinOrder0 cho (order, period) → filter BOM grid."""
        order_id  = self._thdm_selected_order_id
        period_id = self._thdm_selected_period_id
        if order_id is None or period_id is None:
            return
        threading.Thread(
            target=self._thdm_builtin_orders_worker,
            args=(order_id, period_id), daemon=True).start()

    def _thdm_builtin_orders_worker(self, order_id, period_id):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute("""
                SELECT DISTINCT BuiltinOrder0
                FROM   B30BizDocDetailFactory
                WHERE  BizDocId    = ?
                  AND  PeriodId    = ?
                  AND  BuiltinOrder0 IS NOT NULL
            """, [order_id, period_id])
            valid = {str(r[0]) for r in cur.fetchall()}
            conn.close()
            self.after(0, lambda v=valid: self._thdm_builtin_orders_done(v, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_builtin_orders_done(set(), str(err)))

    def _thdm_builtin_orders_done(self, valid_orders, error):
        dt = THEMES[ctk.get_appearance_mode()]
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg=dt["log_err"])
            return
        self._thdm_valid_builtin_orders = valid_orders
        self._thdm_filter_bom()

    # ── THDM: File picker Excel ────────────────────────────────────────────────

    def _thdm_pick_excel(self):
        import tkinter.filedialog as fd
        import os
        path = fd.askopenfilename(
            title="Chọn file Excel THDM",
            filetypes=[("Excel", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")])
        if not path:
            return
        self._thdm_excel_path = path
        self.lbl_thdm_excel_path.config(text=os.path.basename(path))
        # Reset preview
        self._thdm_clear_preview_trees()
        self._thdm_preview_data = []
        self._thdm_validated  = False
        self._thdm_val_errors = {}
        self.btn_thdm_export_xl.configure(state="disabled")
        self.btn_thdm_view_sql.configure(state="disabled")
        self.btn_thdm_validate.configure(state="disabled")
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(text="")
        self._thdm_update_ui()
        self._thdm_start_pre_preview(path)

    # ── THDM: Pre-preview (đọc Excel ngay sau khi chọn file) ──────────────────

    def _thdm_start_pre_preview(self, path):
        """Hiện pre-preview frame + bắt đầu đọc file (background)."""
        if hasattr(self, '_thdm_empty_frame'):
            self._thdm_empty_frame.place_forget()
        self._thdm_result_frame.grid_remove()
        self._thdm_pre_frame.grid()
        self._thdm_pre_nb.grid_remove()
        self._lbl_thdm_pre_status.config(text="⏳  Đang đọc file Excel...")
        threading.Thread(
            target=self._thdm_pre_preview_worker, args=(path,), daemon=True).start()

    def _thdm_pre_preview_worker(self, path):
        try:
            mapping         = load_mapping()
            thdm_map        = mapping.get('THDM_THVT', [])
            mapping_anchors = [
                _norm_vn(r['ten_excel'])
                for r in thdm_map
                if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')
            ]

            # Mở workbook giống aggregate_worker (1 lần, tái dùng cho parse)
            wb_c, wb_l, wb_h = _thdm_open_workbook(path)
            sheet_name = _thdm_find_thvt_sheet(wb_c)
            if not sheet_name:
                raise ValueError(
                    "Không tìm thấy sheet 'TH VT'.\n"
                    "Tên hợp lệ: 'TH VT', 'THVT', 'TH_VT', 'Tổng hợp vật tư'.")

            # Đọc raw rows từ wb_c
            ws_c     = wb_c[sheet_name]
            raw_base = [tuple(c.value for c in row) for row in ws_c.iter_rows()]

            # Áp merge info từ wb_h (read_only=False → có merged_cells)
            # Mục đích: fill giá trị sang các ô bị merge để hiển thị như Excel
            merge_map = {}   # (row_0idx, col_0idx) → value
            if wb_h and sheet_name in wb_h.sheetnames:
                ws_h = wb_h[sheet_name]
                for mr in ws_h.merged_cells.ranges:
                    top_val = ws_h.cell(mr.min_row, mr.min_col).value
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            if not (r == mr.min_row and c == mr.min_col):
                                merge_map[(r - 1, c - 1)] = top_val

            if merge_map:
                raw_rows = []
                for ri, row in enumerate(raw_base):
                    lst = list(row)
                    for ci in range(len(lst)):
                        if (ri, ci) in merge_map:
                            lst[ci] = merge_map[(ri, ci)]
                    raw_rows.append(tuple(lst))
            else:
                raw_rows = raw_base

            # Dùng raw_base (trước merge) để detect hidx
            # — merge fill có thể làm row CHI TIẾT trông giống header hơn
            hidx = _thdm_find_header_row(raw_base, anchors=mapping_anchors or None)

            # Full parse (giống _thdm_aggregate_worker)
            rows_raw = _thdm_parse_thvt_sheet(wb_c, wb_l, wb_h, sheet_name, thdm_map)
            wb_c.close()
            if wb_l: wb_l.close()
            if wb_h: wb_h.close()

            rows_out = []
            for rd in rows_raw:
                row_norm = {}
                for k, v in rd.items():
                    row_norm[k] = v.strip() if isinstance(v, str) else v
                rows_out.append(row_norm)

            self.after(0, lambda rr=raw_rows, rb=raw_base, hi=hidx, po=rows_out:
                       self._thdm_pre_preview_done(rr, rb, hi, po, None))
        except Exception as exc:
            self.after(0, lambda e=exc:
                       self._thdm_pre_preview_done([], [], None, [], str(e)))

    def _thdm_pre_preview_done(self, raw_rows, raw_base, hidx, parsed_rows, error):
        dt = THEMES[ctk.get_appearance_mode()]
        if error:
            self._lbl_thdm_pre_status.config(text=f"❌  {error}")
            self._thdm_pre_nb.grid()   # hiện notebook dù lỗi (tabs rỗng)
            return

        MAX_RAW = 500

        # ── Raw Excel tab ─────────────────────────────────────────────────────
        # raw_rows = merged version (để hiển thị body),
        # nhưng col_names lấy từ raw_rows[hidx] (row STT/MãVT...)
        # + fallback sang raw_rows[hidx+1] (row Muc 1, Muc 1.2...) cho ô trống
        if hidx is not None and hidx < len(raw_rows):
            # Dùng raw_base (TRƯỚC merge) để build col_names
            # — merge có thể span nhiều row → raw_rows[hidx] bị fill "CHI TIẾT..." sai
            hrow_b  = raw_base[hidx]     if hidx < len(raw_base)     else ()
            hrow2_b = raw_base[hidx + 1] if hidx + 1 < len(raw_base) else ()
            n       = max(len(hrow_b), len(hrow2_b))
            col_names, seen = [], {}
            for i in range(n):
                v1 = hrow_b[i]  if i < len(hrow_b)  else None
                v2 = hrow2_b[i] if i < len(hrow2_b) else None
                # Ưu tiên v2 (Muc 1, NHÀ MÁY...) vì cụ thể hơn
                # v1 có thể là super-header chung ("CHI TIẾT GIAO CT") — chỉ dùng khi v2 trống
                nm = str(v2).strip() if v2 is not None and str(v2).strip() else ""
                if not nm:
                    nm = str(v1).strip() if v1 is not None and str(v1).strip() else ""
                if not nm:
                    nm = f"_C{i}"
                if nm in seen:
                    seen[nm] += 1
                    nm = f"{nm}_{seen[nm]}"
                else:
                    seen[nm] = 0
                col_names.append(nm)
            # Body: bắt đầu sau 2 dòng header (col headings đã hiển thị đủ thông tin)
            display_rows = raw_rows[hidx + 2:]
        else:
            n_cols       = max((len(r) for r in raw_rows[:5]), default=10)
            col_names    = [f"C{i}" for i in range(n_cols)]
            display_rows = raw_rows

        import datetime as _dt

        def _raw_fmt(v):
            """Format giá trị raw cell: datetime → chuỗi ngày, số → str, None → ''."""
            if v is None:
                return ""
            if isinstance(v, (_dt.datetime, _dt.date)):
                return v.strftime("%d/%m/%Y")
            return str(v)

        raw_tree = self.thdm_raw_tree
        for iid in raw_tree.get_children():
            raw_tree.delete(iid)
        raw_tree["columns"] = tuple(col_names)
        for nm in col_names:
            _ra = guess_col_align(nm)
            raw_tree.heading(nm, text=nm, anchor=_ra)
            raw_tree.column(nm, width=90, anchor=_ra, minwidth=50, stretch=False)
        raw_tree.tag_configure("row_header", foreground="#60A5FA", background="#1E3A5F")
        raw_tree.tag_configure("row_normal", foreground=dt["tv_text"])
        raw_tree.tag_configure("row_alt",    foreground=dt["text_muted"])
        raw_tree.tag_configure("row_blank",  foreground="#444444")

        for i, row in enumerate(display_rows[:MAX_RAW]):
            padded = list(row) + [None] * max(0, len(col_names) - len(row))
            vals   = [_raw_fmt(v) for v in padded[:len(col_names)]]
            is_blank = all(v == "" for v in vals)
            if is_blank:
                tag = "row_blank"
            else:
                tag = "row_normal" if i % 2 == 0 else "row_alt"
            raw_tree.insert("", "end", values=tuple(vals), tags=(tag,))

        # ── Parsed data (tab "Đã xử lý" đã bỏ — chỉ giữ data cho Tổng hợp) ────
        self._thdm_preview_data = parsed_rows   # pre-populate để Tổng hợp không cần parse lại

        # Hiện notebook + cập nhật status
        self._thdm_pre_nb.grid()
        n_total  = len(raw_rows)
        n_parsed = len(parsed_rows)
        suffix   = f" (hiển thị {MAX_RAW})" if len(display_rows) > MAX_RAW else ""
        self._lbl_thdm_pre_status.config(
            text=(f"✅  {n_total:,} dòng Raw{suffix}  ·  {n_parsed:,} dòng đã parse"
                  f"  —  Kiểm tra xong thì nhấn 📊 Tổng hợp"))

    # ── THDM: Load BOM list ────────────────────────────────────────────────────

    def _thdm_load_bom_list(self):
        self.lbl_thdm_status.config(text="⏳  Đang tải danh sách BOM...", fg=self._text_fg)
        threading.Thread(target=self._thdm_load_bom_list_worker, daemon=True).start()

    def _thdm_load_bom_list_worker(self):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            where_extra = ""
            params = []
            if self._thdm_selected_order_id is not None:
                where_extra = "  AND  BOM.ParentBizDocId = ?"
                params = [self._thdm_selected_order_id]
            # CTE: chỉ lấy Version mới nhất của mỗi BOM.Code (task 7)
            cur.execute(f"""
                WITH RankedBOM AS (
                    SELECT BOM.Id,
                           BOM.Code,
                           BOM.Description,
                           CONVERT(varchar(10), BOM.DocDate, 103) AS DocDate,
                           BOM.IsActive,
                           Prd.Name          AS ProductName1,
                           Dt0.BuiltinOrder0,
                           Dt.DocNo          AS DocInforFO,
                           BOM.Version,
                           ROW_NUMBER() OVER (
                               PARTITION BY BOM.Code
                               ORDER BY     ISNULL(TRY_CAST(BOM.Version AS FLOAT), -1) DESC,
                                            BOM.DocDate DESC
                           ) AS _rn
                    FROM   B20BOM AS BOM
                    LEFT JOIN B20Product        AS Prd ON Prd.Id             = BOM.ProductId1
                    LEFT JOIN B30BizDocDetailSO AS Dt0 ON Dt0.DetailRowId_SO = BOM.DetailRowId_SO
                    LEFT JOIN B30BizDocSO       AS Dt  ON Dt.BizDocId        = BOM.ParentBizDocId
                    WHERE  BOM.IsGroup  = 0
                      AND  BOM.IsActive = 1
                      {where_extra}
                )
                SELECT Id, Code, Description, DocDate, IsActive,
                       ProductName1, BuiltinOrder0, DocInforFO, Version
                FROM   RankedBOM
                WHERE  _rn = 1
                ORDER  BY TRY_CAST(BuiltinOrder0 AS FLOAT), Code
            """, params)
            cols = [c[0] for c in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_bom_list_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_bom_list_done([], str(err)))

    def _thdm_load_bom_list_done(self, rows, error):
        dt = THEMES[ctk.get_appearance_mode()]
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg=dt["log_err"])
            return
        self._thdm_all_bom_rows = rows
        self._thdm_checked_ids  = set()
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} BOM — chọn BOM rồi chọn file Excel",
            fg=C["green"])

    def _thdm_filter_bom(self):
        kw = self.thdm_search_var.get().strip().lower()
        valid_orders = self._thdm_valid_builtin_orders  # None = không filter

        if _HAS_TKSHEET and self.thdm_bom_sheet is not None:
            # ── tksheet path ─────────────────────────────────────────────────
            data = []
            self._thdm_sheet_row_ids = []
            for r in self._thdm_all_bom_rows:
                code      = r.get("Code",         "") or ""
                item_name = r.get("ProductName1", "") or ""
                version   = str(r.get("Version",  "") or "")
                muc_so    = str(r.get("BuiltinOrder0", "") or "")
                if valid_orders is not None and muc_so not in valid_orders:
                    continue
                if kw and kw not in code.lower() \
                       and kw not in item_name.lower() \
                       and kw not in version.lower():
                    continue
                chk = "☑" if r["Id"] in self._thdm_checked_ids else "☐"
                data.append([chk, muc_so, code, item_name, version])
                self._thdm_sheet_row_ids.append(r["Id"])
            self.thdm_bom_sheet.set_sheet_data(data, reset_col_positions=False)
            self._thdm_apply_sheet_highlights()
        else:
            # ── Fallback: ttk.Treeview path ──────────────────────────────────
            tree = self.thdm_bom_tree
            for iid in tree.get_children():
                tree.delete(iid)
            for r in self._thdm_all_bom_rows:
                code      = r.get("Code",         "") or ""
                item_name = r.get("ProductName1", "") or ""
                version   = str(r.get("Version",  "") or "")
                muc_so    = str(r.get("BuiltinOrder0", "") or "")
                if valid_orders is not None and muc_so not in valid_orders:
                    continue
                if kw and kw not in code.lower() \
                       and kw not in item_name.lower() \
                       and kw not in version.lower():
                    continue
                rid = str(r["Id"])
                chk = "☑" if r["Id"] in self._thdm_checked_ids else "☐"
                tag = "checked" if r["Id"] in self._thdm_checked_ids else "unchecked"
                tree.insert("", "end", iid=rid,
                    values=(chk, muc_so, code, item_name, version),
                    tags=(tag,))

    def _thdm_bom_autostretch(self):
        """Giãn cột 'Tên SP' (index 3) lấp hết chiều ngang trống của BOM list."""
        self._thdm_bom_stretch_job = None
        sh = getattr(self, 'thdm_bom_sheet', None)
        if not (_HAS_TKSHEET and sh is not None):
            return
        try:
            total  = sh.MT.winfo_width()
            others = sum(int(sh.column_width(column=c)) for c in (0, 1, 2, 4))
            w = total - others - 6
            if w > 200:
                sh.column_width(column=3, width=w)
        except Exception:
            pass

    def _thdm_apply_sheet_highlights(self):
        """Tô màu các row đang được check trong tksheet."""
        self.thdm_bom_sheet.dehighlight_all()
        checked_rows = [
            i for i, rid in enumerate(self._thdm_sheet_row_ids)
            if rid in self._thdm_checked_ids
        ]
        if checked_rows:
            self.thdm_bom_sheet.highlight_rows(
                rows=checked_rows, bg="#1e3a5f", fg="#90c8f8", redraw=True)
        else:
            self.thdm_bom_sheet.refresh()

    def _thdm_on_bom_click(self, event=None):
        if _HAS_TKSHEET and self.thdm_bom_sheet is not None:
            # tksheet path — được gọi từ fallback Treeview nếu sheet=None
            pass   # xử lý ở _thdm_process_sheet_click (gọi qua MT.bind)
        else:
            # ── Fallback: ttk.Treeview path ──────────────────────────────────
            iid = self.thdm_bom_tree.identify_row(event.y)
            if not iid:
                return
            iid_to_id = {str(r["Id"]): r["Id"] for r in self._thdm_all_bom_rows}
            clicked_id = iid_to_id.get(iid)
            if clicked_id is None:
                return
            new_checked  = clicked_id not in self._thdm_checked_ids
            selected_iids = set(self.thdm_bom_tree.selection())
            selected_iids.add(iid)
            for s_iid in selected_iids:
                s_id = iid_to_id.get(s_iid)
                if s_id is None:
                    continue
                if new_checked:
                    self._thdm_checked_ids.add(s_id)
                else:
                    self._thdm_checked_ids.discard(s_id)
                chk  = "☑" if new_checked else "☐"
                tag  = "checked" if new_checked else "unchecked"
                vals = list(self.thdm_bom_tree.item(s_iid, "values"))
                vals[0] = chk
                self.thdm_bom_tree.item(s_iid, values=vals, tags=(tag,))
            self._thdm_update_ui()

    def _thdm_process_sheet_click(self):
        """Xử lý toggle checkbox sau khi tksheet update selection (gọi qua after(40))."""
        if not (_HAS_TKSHEET and self.thdm_bom_sheet is not None):
            return
        sheet = self.thdm_bom_sheet
        selected_rows = []

        # Method 1: get_selected_rows — works khi row header được click
        try:
            r = sheet.get_selected_rows()
            selected_rows = sorted(r)
        except Exception:
            pass

        # Method 2: get_selected_cells — works khi enable_bindings("row_select") chọn cell
        if not selected_rows:
            try:
                cells = sheet.get_selected_cells()
                selected_rows = sorted({c[0] for c in cells})
            except Exception:
                pass

        # Method 3: get_currently_selected — fallback cuối
        if not selected_rows:
            try:
                sel = sheet.get_currently_selected()
                if sel is not None:
                    if hasattr(sel, 'row'):
                        selected_rows = [sel.row]
                    elif isinstance(sel, (list, tuple)) and len(sel) > 0:
                        first = sel[0]
                        selected_rows = [first[0] if isinstance(first, (list, tuple)) else first]
            except Exception:
                pass

        if not selected_rows:
            return
        row_idx = selected_rows[0]
        if row_idx >= len(self._thdm_sheet_row_ids):
            return

        clicked_id  = self._thdm_sheet_row_ids[row_idx]
        new_checked = clicked_id not in self._thdm_checked_ids

        for r in selected_rows:
            if r >= len(self._thdm_sheet_row_ids):
                continue
            row_id = self._thdm_sheet_row_ids[r]
            if new_checked:
                self._thdm_checked_ids.add(row_id)
            else:
                self._thdm_checked_ids.discard(row_id)
            chk = "☑" if new_checked else "☐"
            try:
                self.thdm_bom_sheet.set_cell_data(r, 0, chk, redraw=False)
            except Exception:
                try:
                    self.thdm_bom_sheet.set_cell_data(r, 0, chk)
                except Exception:
                    pass

        self._thdm_apply_sheet_highlights()
        self._thdm_update_ui()

    def _thdm_toggle_all(self, select: bool):
        for r in self._thdm_all_bom_rows:
            if select:
                self._thdm_checked_ids.add(r["Id"])
            else:
                self._thdm_checked_ids.discard(r["Id"])
        self._thdm_filter_bom()
        self._thdm_update_ui()

    def _thdm_update_ui(self):
        n = len(self._thdm_checked_ids)
        self.lbl_thdm_sel_count.config(
            text=f"{n} BOM được chọn",
            fg=C["accent"] if n else C["text"])
        self.btn_thdm_aggregate.configure(state="normal")

    # ── THDM: Đọc Excel và preview ────────────────────────────────────────────

    def _thdm_aggregate(self):
        if not self._thdm_excel_path:
            self._show_msg("Chưa chọn file Excel",
                "Vui lòng nhấn '📂 Chọn file Excel THDM' để chọn file.", kind="info")
            return
        self.btn_thdm_aggregate.configure(state="disabled", text="⏳  Đang đọc Excel...")
        self.btn_thdm_export_xl.configure(state="disabled")
        self.btn_thdm_view_sql.configure(state="disabled")
        self.btn_thdm_validate.configure(state="disabled")
        self.btn_thdm_insert.configure(state="disabled")
        self._thdm_clear_preview_trees()
        path = self._thdm_excel_path
        threading.Thread(target=self._thdm_aggregate_worker, args=(path,), daemon=True).start()

    def _thdm_aggregate_worker(self, path):
        try:
            mapping   = load_mapping()
            cfg       = mapping.get('_CONFIG', {})

            # Tìm tất cả child sections của THDM_HEADER (theo thứ tự khai báo trong _CONFIG)
            child_sections = [
                sec for sec, info in cfg.items()
                if info.get('parent_section') == 'THDM_HEADER'
            ]
            if not child_sections:
                raise ValueError("Không tìm thấy child section nào của THDM_HEADER trong _CONFIG.")

            # ── Query B20BOMDetail cho các BOM đã chọn ─────────────────────────
            bom_qty_dict = None
            bom_qty_warn = None
            if self._thdm_checked_ids:
                try:
                    conn_bom = self._get_db_conn()
                    bom_qty_dict = _thdm_load_bom_qty_dict(conn_bom, self._thdm_checked_ids)
                    conn_bom.close()
                except Exception as e_bom:
                    bom_qty_warn = str(e_bom)
                    bom_qty_dict = None

            # Mở file 3 lần (cached / live formula / hidden detection) — dùng chung
            wb_c, wb_l, wb_h = _thdm_open_workbook(path)

            # Detect sheet TH VT — tất cả THDM child sections đều parse sheet này
            sheet_name = _thdm_find_thvt_sheet(wb_c)
            if not sheet_name:
                raise ValueError(
                    "Không tìm thấy sheet 'TH VT' trong file Excel.\n"
                    "Tên sheet hợp lệ: 'TH VT', 'THVT', 'TH_VT', 'Tổng hợp vật tư'.")

            rows_out = []
            for sec in child_sections:
                sec_cfg    = cfg[sec]
                sec_map    = mapping.get(sec, [])
                expand_muc = sec_cfg.get('expand_muc', False)
                row_filter = sec_cfg.get('row_filter') or ''

                rows_raw = _thdm_parse_thvt_sheet(
                    wb_c, wb_l, wb_h, sheet_name, sec_map,
                    expand_muc=expand_muc,
                    bom_qty_dict=bom_qty_dict if expand_muc else None,
                )

                # Áp dụng RowFilter nếu có (vd: 'Quantity>0')
                if row_filter:
                    rows_raw = _thdm_apply_row_filter(rows_raw, row_filter)

                # Chuẩn hoá + gán tag _section cho từng row
                for rd in rows_raw:
                    row_norm = {'_section': sec}
                    for k, v in rd.items():
                        row_norm[k] = v.strip() if isinstance(v, str) else v
                    rows_out.append(row_norm)

            wb_c.close()
            if wb_l: wb_l.close()
            if wb_h: wb_h.close()

            self.after(0, lambda d=rows_out, w=bom_qty_warn: self._thdm_aggregate_done(d, None, bom_warn=w))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_aggregate_done([], str(err)))

    def _thdm_clear_preview_trees(self):
        """Xóa dữ liệu tab Đầu phiếu + tất cả tab section detail."""
        for tr in [self._thdm_header_tree] + list(self._thdm_sec_trees.values()):
            ch = tr.get_children()
            if ch:
                tr.delete(*ch)

    def _thdm_fill_header_preview(self):
        """Đổ preview THDM_HEADER — CÙNG cấu trúc với tab detail:
        header = Ten_Excel (rỗng → SQL_Column), hàng phụ xanh = SQL_Column,
        hàng giá trị = resolve nhẹ (display-only, không cần DB). Cột có
        Ten_Excel xếp trước. Giá trị SP/tự sinh là placeholder, thật khi INSERT."""
        tr = self._thdm_header_tree
        ch = tr.get_children()
        if ch:
            tr.delete(*ch)
        if not getattr(self, '_thdm_now_str', None):
            self._thdm_now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        defs = _thdm_get_detail_col_defs((self.mapping or {}).get('THDM_HEADER', []))
        self._thdm_header_col_defs = defs

        import tkinter.font as _tkf
        _hfont = _tkf.Font(family="Segoe UI", size=12, weight="bold")
        _font  = _tkf.Font(family="Segoe UI", size=11)
        tr["columns"] = tuple(d['sql_col'] for d in defs)
        for d in defs:
            htxt = d['ten_excel'] or d['sql_col']
            val  = self._thdm_col_display(d, 0)
            tr.heading(d['sql_col'], text=htxt, anchor='center')
            w = max(_hfont.measure(htxt), _font.measure(d['sql_col']),
                    _font.measure(str(val))) + 24
            # Cột căn trái; chỉ 2 dòng header (tiêu đề + hàng phụ SQL) căn giữa
            tr.column(d['sql_col'], width=max(min(w, 320), 80),
                      anchor='w', stretch=False)
        # Hàng phụ (xanh) = SQL_Column ; hàng giá trị = resolved
        tr.insert("", "end", values=tuple(d['sql_col'] for d in defs),
                  tags=("sql_names",))
        tr.insert("", "end",
                  values=tuple(self._thdm_col_display(d, 0) for d in defs),
                  tags=("row_normal",))
        # Căn giữa riêng hàng phụ SQL_Column (dòng giá trị vẫn căn trái)
        if hasattr(tr, 'center_rows_by_tag'):
            tr.center_rows_by_tag("sql_names")

    def _thdm_aggregate_done(self, rows, error, bom_warn=None):
        self.btn_thdm_aggregate.configure(state="normal", text="📊  ③ Tổng hợp")
        # Switch từ pre-preview → result frame
        if hasattr(self, '_thdm_pre_frame'):
            self._thdm_pre_frame.grid_remove()
            self._thdm_result_frame.grid()
        if error:
            self._show_msg("Lỗi đọc Excel", str(error))
            return
        if bom_warn:
            self._show_msg("Cảnh báo BOM",
                f"Không load được định mức từ B20BOMDetail:\n{bom_warn}\n\n"
                "QuantityFactory sẽ dùng tỉ lệ phân bổ thay vì ĐỊNH MỨC thực tế.",
                kind="warning")
        self._thdm_preview_data = rows
        # Tổng hợp mới → reset trạng thái kiểm tra + filter
        self._thdm_validated        = False
        self._thdm_val_errors       = {}
        self._thdm_val_lk_fails     = {}
        self._thdm_validate_caches  = {}
        self._thdm_lk_recs          = {}
        self._fuzzy_resolutions     = {}
        self._thdm_filter_mode      = "all"
        _sec_counts = self._thdm_fill_section_trees()
        self._thdm_update_filter_btns()

        # Tab Đầu phiếu
        self._thdm_fill_header_preview()

        if rows:
            self.btn_thdm_export_xl.configure(state="normal")
            self.btn_thdm_view_sql.configure(state="normal")
            self.btn_thdm_validate.configure(state="normal")
            # INSERT chỉ bật sau khi Kiểm tra đạt (giống luồng BOM)
            self.btn_thdm_insert.configure(state="disabled")
            _detail = "  ·  ".join(
                f"{(self.mapping.get('_CONFIG', {}).get(s, {}).get('label') or s)}: {n:,}"
                for s, n in _sec_counts.items())
            self.lbl_thdm_result.config(
                text=f"📊  {len(rows):,} dòng ({_detail}) — bấm ④ Kiểm tra",
                fg=C["yellow"])

    def _thdm_fill_section_trees(self, error_map=None, visible_only=False):
        """Đổ dữ liệu từng section vào tab riêng. error_map (optional):
        {(section, idx_trong_section): [msg,...]} → tô đỏ dòng lỗi.
        visible_only=True → chỉ render dòng có lỗi (filter ⚠).
        Trả về dict {section: số_dòng}."""
        error_map = error_map or {}
        # Chỉ xóa các section detail tabs — KHÔNG xóa header tab.
        # Header được fill riêng bởi _thdm_fill_header_preview() và phải
        # được giữ nguyên khi filter/validate để tab Header không trống.
        for _st in self._thdm_sec_trees.values():
            _ch = _st.get_children()
            if _ch:
                _st.delete(*_ch)
        self._thdm_sec_visible = {}

        def _fmt(v, is_num=False):
            if v is None:
                return ""
            if is_num:
                try:
                    return f"{float(v):,.4f}"
                except (TypeError, ValueError):
                    pass
            return str(v) if v != "" else ""

        self._thdm_now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = self._thdm_preview_data
        counts = {}
        for _tab_i, (sec, tr) in enumerate(self._thdm_sec_trees.items(), start=1):
            defs     = self._thdm_sec_col_defs.get(sec, [])
            sec_rows = [r for r in rows if r.get('_section') == sec]
            counts[sec] = len(sec_rows)
            # Số lỗi thực tế luôn tính trên toàn section (không phụ thuộc filter)
            n_err_sec = sum(1 for i in range(len(sec_rows)) if (sec, i) in error_map)
            # Chỉ hiện dòng lỗi hay tất cả
            if visible_only:
                vis_indices = [i for i in range(len(sec_rows))
                               if (sec, i) in error_map]
            else:
                vis_indices = list(range(len(sec_rows)))
            self._thdm_sec_visible[sec] = vis_indices
            # Hàng phụ (xanh) hiển thị SQL_Column tương ứng mỗi cột
            tr.insert("", "end",
                      values=tuple(d['sql_col'] for d in defs),
                      tags=("sql_names",))
            for i in vis_indices:
                r = sec_rows[i]
                if (sec, i) in error_map:
                    tag = "err_row"
                else:
                    tag = "row_normal" if i % 2 == 0 else "row_alt"
                vals = []
                for d in defs:
                    v = r.get(d['sql_col'])
                    # Cột không phải Excel & chưa có sẵn trong row → hiện giá
                    # trị sẽ resolve khi INSERT (CoDinh/UILookup/HeThong...)
                    if v is None and (d.get('nguon_dl') or 'Excel') != 'Excel':
                        vals.append(self._thdm_col_display(d, i))
                    else:
                        vals.append(_fmt(v, is_num=d.get('anchor') == 'e'))
                tr.insert("", "end", values=tuple(vals), tags=(tag,))
            lbl = (self.mapping.get('_CONFIG', {}).get(sec, {}).get('label')
                   or sec)
            _cnt = (f"{len(sec_rows):,}" if not n_err_sec
                    else f"{len(sec_rows):,}, ⚠{n_err_sec}")
            try:
                self._thdm_result_nb.tab(_tab_i, text=f"  {lbl} ({_cnt})  ")
            except Exception:
                pass
        return counts

    # ── THDM Error UX: click / double-click / filter / quick-map ──────────────

    def _thdm_mt_click(self, event, sec):
        """Feature 1: cell_select / rc_select callback từ tksheet extra_bindings.
        event là EventDataDict; row index nằm ở event.selected.row."""
        try:
            vis_row = event.selected.row
        except (AttributeError, TypeError):
            return
        self.after(20, lambda r=vis_row: self._thdm_on_row_select(sec, r))

    def _thdm_on_row_select(self, sec, vis_row):
        """Feature 1: hiện lỗi dòng đang chọn lên thanh chân app (lbl_thdm_err_bar)."""
        if vis_row == 0:
            return   # dòng sql_names (row header phụ)
        vis_indices = self._thdm_sec_visible.get(sec, [])
        data_idx = vis_indices[vis_row - 1] if vis_row - 1 < len(vis_indices) else None
        if data_idx is None:
            return
        msgs = self._thdm_val_errors.get((sec, data_idx))
        if msgs:
            self.lbl_thdm_err_bar.configure(
                text=f"⚠ Dòng {data_idx + 1}: {' | '.join(msgs)}")
        else:
            self.lbl_thdm_err_bar.configure(text="")

    def _thdm_mt_dblclick(self, event, sec):
        """Feature 2: extra_double_b1_func callback — raw tkinter event (.x, .y).
        Dùng MT.identify_row(y=event.y) để lấy row index."""
        tr = self._thdm_sec_trees.get(sec)
        if tr is None:
            return
        try:
            vis_row = tr.sheet.MT.identify_row(y=event.y)
        except Exception:
            return
        if vis_row is None or vis_row == 0:
            return
        vis_indices = self._thdm_sec_visible.get(sec, [])
        data_idx = vis_indices[vis_row - 1] if vis_row - 1 < len(vis_indices) else None
        if data_idx is None:
            return
        if (sec, data_idx) not in self._thdm_val_lk_fails:
            return   # dòng không có lỗi lookup → để edit_cell mở bình thường
        # Đóng cell editor trước khi mở dialog Quick Map
        try:
            tr.sheet.MT.hide_text_editor()
        except Exception:
            pass
        self._thdm_quick_map(sec, data_idx)

    def _thdm_quick_map(self, sec, data_idx):
        """Feature 2+3: Mở batch fuzzy dialog cho tất cả lỗi lookup của dòng,
        áp kết quả, rồi hỏi batch-fill cho các dòng cùng ItemName."""
        fails = self._thdm_val_lk_fails.get((sec, data_idx))
        if not fails:
            return
        sec_rows = [r for r in self._thdm_preview_data if r.get('_section') == sec]
        row = sec_rows[data_idx] if data_idx < len(sec_rows) else None
        if row is None:
            return

        # Tạm bật collect mode để _lookup_generic build candidates (không show popup)
        _saved_collect = getattr(self, '_fuzzy_collect_mode', False)
        _saved_pending = list(getattr(self, '_fuzzy_pending', []))
        _saved_ctx     = dict(getattr(self, '_fuzzy_ctx', {}))
        self._fuzzy_collect_mode = True
        self._fuzzy_pending      = []
        for sql_col, raw_val, cache_key, lk_rec in fails:
            cache = self._thdm_validate_caches.get(sec, {}).get(cache_key, [])
            if not cache:
                continue
            kl = lk_rec.get('kieu_lookup', 'fuzzy_code')
            self._fuzzy_ctx = {
                'section': sec,
                'field'  : lk_rec.get('ten_excel') or sql_col,
                'row_idx': data_idx + 1,
            }
            self._lookup_generic(str(raw_val), cache, kl, _cache_key=cache_key)
        pending = list(self._fuzzy_pending)
        self._fuzzy_collect_mode = _saved_collect
        self._fuzzy_pending      = _saved_pending
        self._fuzzy_ctx          = _saved_ctx

        if not pending:
            return

        resolutions = self._show_batch_fuzzy_dialog(pending)
        if not resolutions:
            return

        fixed_any = False
        for sql_col, raw_val, cache_key, lk_rec in fails:
            res_key    = (str(raw_val), cache_key)
            if res_key not in resolutions:
                continue
            chosen_val = resolutions[res_key]
            if chosen_val is None:
                continue   # user bỏ qua → không ghi đè

            # Ghi nhận mapping; giữ raw code trong preview_data để lookup lại đúng
            if not hasattr(self, '_fuzzy_resolutions'):
                self._fuzzy_resolutions = {}
            self._fuzzy_resolutions[(str(raw_val), cache_key)] = chosen_val

            # Xóa msg lỗi tương ứng
            err_key = (sec, data_idx)
            if err_key in self._thdm_val_errors:
                old_msg = f"Mã '{raw_val}' không tồn tại trong DB"
                self._thdm_val_errors[err_key] = [
                    m for m in self._thdm_val_errors[err_key] if m != old_msg]
                if not self._thdm_val_errors[err_key]:
                    del self._thdm_val_errors[err_key]
                    if err_key in self._thdm_val_lk_fails:
                        del self._thdm_val_lk_fails[err_key]
            fixed_any = True

            # Batch auto-fill dòng cùng ItemName
            item_name_col = None
            for lk in self._thdm_lk_recs.get(sec, []):
                if lk.get('sql_col') and 'name' in lk['sql_col'].lower():
                    item_name_col = lk['sql_col']
                    break
            cur_name  = row.get(item_name_col) if item_name_col else None
            same_rows = []
            if cur_name:
                for other_idx, other_row in enumerate(sec_rows):
                    if other_idx == data_idx:
                        continue
                    if other_row.get(item_name_col) == cur_name:
                        if any(f[0] == sql_col
                               for f in self._thdm_val_lk_fails.get((sec, other_idx), [])):
                            same_rows.append(other_idx)
            if same_rows:
                confirm = self._ask_msg(
                    "Áp dụng hàng loạt?",
                    f"Tìm thấy {len(same_rows)} dòng khác có cùng \"{cur_name}\" "
                    f"và cùng lỗi cột {lk_rec.get('ten_excel') or sql_col}.\n"
                    "Áp dụng giá trị vừa chọn cho tất cả?")
                if confirm:
                    for other_idx in same_rows:
                        ek = (sec, other_idx)
                        # Đọc raw value gốc trước khi ghi đè
                        other_raw = next(
                            (f[1] for f in self._thdm_val_lk_fails.get(ek, [])
                             if f[0] == sql_col),
                            raw_val)
                        self._fuzzy_resolutions[(str(other_raw), cache_key)] = chosen_val
                        if ek in self._thdm_val_errors:
                            old_msg2 = f"Mã '{other_raw}' không tồn tại trong DB"
                            self._thdm_val_errors[ek] = [
                                m for m in self._thdm_val_errors[ek]
                                if m != old_msg2]
                            if not self._thdm_val_errors[ek]:
                                del self._thdm_val_errors[ek]
                                if ek in self._thdm_val_lk_fails:
                                    self._thdm_val_lk_fails[ek] = [
                                        f for f in self._thdm_val_lk_fails[ek]
                                        if f[0] != sql_col]
                                    if not self._thdm_val_lk_fails[ek]:
                                        del self._thdm_val_lk_fails[ek]

        if fixed_any:
            self._thdm_apply_filter(self._thdm_filter_mode)
            self._thdm_update_error_labels()

    def _thdm_toggle_filter(self, value):
        """Feature 4: CTkSegmentedButton callback."""
        if "lỗi" in value.lower() or "⚠" in value:
            self._thdm_filter_mode = "errors_only"
        else:
            self._thdm_filter_mode = "all"
        self._thdm_apply_filter(self._thdm_filter_mode)

    def _thdm_apply_filter(self, mode="all"):
        """Re-render bảng với filter mode đang chọn."""
        visible_only = (mode == "errors_only")
        self._thdm_fill_section_trees(
            error_map=self._thdm_val_errors,
            visible_only=visible_only)
        # Force synchronous flush — SheetTable dùng after_idle (deferred); flush
        # ngay tại đây để tksheet.set_sheet_data được gọi trước khi trả event loop.
        for _tr in self._thdm_sec_trees.values():
            try:
                _tr._flush()
            except Exception:
                pass

    def _thdm_update_filter_btns(self):
        """Feature 4: Cập nhật label CTkSegmentedButton (Tất cả N / ⚠ Dòng lỗi N)."""
        n_all = len(self._thdm_preview_data)
        n_err = len(self._thdm_val_errors)
        try:
            self._thdm_filter_seg.configure(
                values=[f"Tất cả ({n_all})", f"⚠ Dòng lỗi ({n_err})"])
            # Giữ selection theo filter_mode hiện tại
            if self._thdm_filter_mode == "errors_only":
                self._thdm_filter_seg.set(f"⚠ Dòng lỗi ({n_err})")
            else:
                self._thdm_filter_seg.set(f"Tất cả ({n_all})")
        except Exception:
            pass

    def _thdm_update_error_labels(self):
        """Cập nhật lbl_thdm_result + trạng thái nút Insert sau Quick Map."""
        n_err = len(self._thdm_val_errors)
        self._thdm_update_filter_btns()
        if n_err == 0:
            self._thdm_validated = True
            self.btn_thdm_insert.configure(state="normal")
            self.lbl_thdm_result.config(
                text="✅  Tất cả lỗi đã được sửa — sẵn sàng tạo THDM", fg=C["green"])
        else:
            self.lbl_thdm_result.config(
                text=f"❌  Còn {n_err} dòng lỗi", fg=C["red"])

    def _thdm_col_display(self, d, ridx):
        """Giá trị hiển thị cho cột không phải Excel (resolve nhẹ, không cần DB).
        Giống logic tab Header — giá trị thật vẫn sinh khi INSERT."""
        nguon = d.get('nguon_dl') or ''
        mac   = d.get('mac_dinh') or ''
        if nguon == 'CoDinh':
            return '' if mac.upper() in ('NULL', 'EMPTY', '') else mac
        if nguon == 'UILookup':
            v = {
                'creator':    self._thdm_creator_user_id or self._current_creator_user_id,
                'product_id': self._thdm_selected_product_id,
                'order_id':   self._thdm_selected_order_id,
                'period_id':  self._thdm_selected_period_id,
            }.get(mac)
            return '' if v in (None, '') else str(v)
        if nguon == 'HeThong':
            mu = mac.upper()
            if mu == 'NOW':          return self._thdm_now_str
            if mac == 'BizDocId_SO': return str(self._thdm_selected_order_id or '')
            if mu == 'AUTO_INC':     return str(ridx + 1)
            return '(tự sinh)'
        if nguon in ('SP', 'TinhToan', 'Lookup', 'MucLookup'):
            return '(khi INSERT)'
        if nguon == 'Excel':
            return '(từ Excel)'
        return ''

    # ── THDM: Kiểm tra dữ liệu trước INSERT (giống Validate bên BOM) ───────────
    def _thdm_validate(self):
        if not self._thdm_preview_data:
            self._show_msg("Chưa có dữ liệu",
                "Hãy bấm 📊 Tổng hợp trước khi Kiểm tra.", kind="info")
            return
        self.btn_thdm_validate.configure(state="disabled", text="⏳  Đang kiểm tra...")
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(text="⏳  Đang kiểm tra dữ liệu...", fg=self._text_fg)
        self._fuzzy_collect_mode = True
        self._fuzzy_pending      = []
        self._fuzzy_ctx          = {}
        threading.Thread(target=self._thdm_validate_worker, daemon=True).start()

    def _thdm_validate_worker(self):
        # Reset lỗi cũ và cache — đảm bảo ô đã sửa tay được re-validate từ đầu
        self._thdm_val_errors      = {}
        self._thdm_val_lk_fails    = {}
        self._thdm_validate_caches = {}

        # FIX 1: Xóa fuzzy resolutions cũ — tránh kết quả tra cứu cũ che khuất mã
        # vừa được bổ sung vào DB. Các resolve fuzzy sẽ được thực hiện lại với
        # cache mới nhất từ DB trong lần validate này.
        if hasattr(self, '_fuzzy_resolutions'):
            self._fuzzy_resolutions.clear()

        conn = None
        try:
            mapping = load_mapping()
            cfg     = mapping.get('_CONFIG', {})
            child_sections = [s for s, i in cfg.items()
                              if i.get('parent_section') == 'THDM_HEADER']

            # FIX 2: Refresh các ô TRỐNG trong _thdm_preview_data từ file Excel.
            # Khi user bổ sung Mã VTTB vào file Excel rồi bấm Kiểm tra, dữ liệu
            # trong _thdm_preview_data vẫn là giá trị cũ (None/rỗng) từ lần Tổng
            # hợp trước. Re-parse file Excel và cập nhật các ô đang trống — ô đã
            # có giá trị (kể cả edit thủ công trên UI) được giữ nguyên.
            import os as _os
            _xls_path = getattr(self, '_thdm_excel_path', None)
            if _xls_path and _os.path.exists(_xls_path):
                try:
                    _wb_c, _wb_l, _wb_h = _thdm_open_workbook(_xls_path)
                    _sn = _thdm_find_thvt_sheet(_wb_c)
                    if _sn:
                        for _sec in child_sections:
                            _sm   = mapping.get(_sec, [])
                            _scfg = cfg.get(_sec, {})
                            _exp  = _scfg.get('expand_muc', False)
                            _rf   = _scfg.get('row_filter') or ''
                            _fresh = _thdm_parse_thvt_sheet(
                                _wb_c, _wb_l, _wb_h, _sn, _sm, expand_muc=_exp)
                            if _rf:
                                _fresh = _thdm_apply_row_filter(_fresh, _rf)
                            _sec_rows = [r for r in self._thdm_preview_data
                                         if r.get('_section') == _sec]
                            for _old, _new in zip(_sec_rows, _fresh):
                                for _k, _v in _new.items():
                                    if _k == '_section':
                                        continue
                                    _old_v    = _old.get(_k)
                                    _old_emp  = (_old_v is None
                                                 or str(_old_v).strip() == '')
                                    _new_emp  = (_v is None
                                                 or (isinstance(_v, str)
                                                     and _v.strip() == ''))
                                    if _old_emp and not _new_emp:
                                        _old[_k] = (_v.strip()
                                                     if isinstance(_v, str) else _v)
                    _wb_c.close()
                    if _wb_l: _wb_l.close()
                    if _wb_h: _wb_h.close()
                except Exception:
                    pass  # re-parse thất bại → tiếp tục với data hiện tại

            conn        = self._get_db_conn()
            errors      = {}   # (section, idx) → [msg,...]
            lk_fails    = {}   # (section, idx) → [(sql_col, raw, key, lk_rec),...]
            all_caches  = {}   # {sec: {cache_key: entries}}
            all_lk_recs = {}   # {sec: [lk_rec,...]}
            for sec in child_sections:
                sec_map = mapping.get(sec, [])
                req_excel = [r for r in sec_map
                    if str(r.get('bat_buoc', '')).strip() in ('1', '1.0')
                    and r.get('nguon_dl') == 'Excel' and r.get('sql_col')]
                lk_recs = [r for r in sec_map
                    if r.get('nguon_dl') == 'Excel' and r.get('kieu_lookup')
                    and r.get('bang_master') and r.get('truong_lay_ve')
                    and r.get('truong_so_sanh')]
                caches = {}
                for r in lk_recs:
                    key = (r['bang_master'], r.get('dieu_kien_master', ''),
                           r['truong_so_sanh'], r['truong_lay_ve'])
                    if key not in caches:
                        try:
                            caches[key] = self._build_cache_generic(conn, *key)
                        except Exception as _ce:
                            # FIX 3: Log rõ khi không build được cache — giúp
                            # chẩn đoán tại sao lookup thất bại (quyền, tên bảng...).
                            self._log('cache',
                                      f"Build cache {r['bang_master']}",
                                      0, 'Warn', str(_ce), 'warn')
                            caches[key] = []
                all_caches[sec]  = caches
                all_lk_recs[sec] = lk_recs
                sec_rows = [row for row in self._thdm_preview_data
                            if row.get('_section') == sec]
                for idx, row in enumerate(sec_rows):
                    msgs = []
                    for rr in req_excel:
                        v = row.get(rr['sql_col'])
                        if v is None or str(v).strip() == '':
                            msgs.append(f"Thiếu {rr.get('ten_excel') or rr['sql_col']}")
                    for rr in lk_recs:
                        raw = row.get(rr['sql_col'])
                        if raw is None or str(raw).strip() == '':
                            continue   # ô trống đã bắt ở check bắt buộc
                        key = (rr['bang_master'], rr.get('dieu_kien_master', ''),
                               rr['truong_so_sanh'], rr['truong_lay_ve'])
                        # Set context cho collect mode để _fuzzy_pending có đủ thông tin
                        self._fuzzy_ctx = {
                            'section': sec,
                            'field'  : rr.get('ten_excel') or rr['sql_col'],
                            'row_idx': idx + 1,
                        }
                        found, _ = self._lookup_generic(
                            raw, caches.get(key, []), rr.get('kieu_lookup', 'exact'),
                            _cache_key=key)
                        if found is None:
                            msgs.append(f"Mã '{raw}' không tồn tại trong DB")
                            lk_fails.setdefault((sec, idx), []).append(
                                (rr['sql_col'], raw, key, rr))
                    if msgs:
                        errors[(sec, idx)] = msgs
            self._thdm_validate_caches = all_caches
            self._thdm_lk_recs         = all_lk_recs
            self._thdm_val_lk_fails    = lk_fails
            self.after(0, lambda e=errors: self._thdm_validate_done(e, None))
        except Exception as e:
            self.after(0, lambda err=str(e): self._thdm_validate_done(None, err))
        finally:
            self._fuzzy_collect_mode = False
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _thdm_validate_done(self, errors, err):
        self.btn_thdm_validate.configure(state="normal", text="🔍  ④ Kiểm tra")
        if err is not None:
            self.lbl_thdm_result.config(text="❌  Lỗi kiểm tra", fg=C["red"])
            self._show_msg("Lỗi kiểm tra", err, kind="error")
            return

        # Hiện batch fuzzy popup cho tất cả lỗi tra cứu mã
        pending = getattr(self, '_fuzzy_pending', [])
        if pending and errors:
            resolutions = self._show_batch_fuzzy_dialog(pending)
            if resolutions:
                self._apply_validate_resolutions(resolutions, errors)

        self._thdm_val_errors = errors
        self._thdm_validated  = True
        n_err = len(errors)
        self._thdm_fill_section_trees(errors)
        self._thdm_update_filter_btns()
        if n_err == 0:
            self.btn_thdm_insert.configure(state="normal")
            self.lbl_thdm_result.config(
                text="✅  Kiểm tra đạt — sẵn sàng tạo THDM", fg=C["green"])
            return
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(
            text=f"❌  {n_err} dòng lỗi — sửa lại rồi bấm Kiểm tra",
            fg=C["red"])
        # Nhảy tới tab section có lỗi đầu tiên
        sec_first = next(iter(errors))[0]
        for _ti, _s in enumerate(self._thdm_sec_trees.keys(), start=1):
            if _s == sec_first:
                try:
                    self._thdm_result_nb.select(_ti)
                except Exception:
                    pass
                break
        # Liệt kê tối đa 15 lỗi còn lại (sau khi đã resolve qua batch popup)
        lbls = self.mapping.get('_CONFIG', {})
        lines = []
        for (sec, idx), msgs in list(errors.items())[:15]:
            _lbl = lbls.get(sec, {}).get('label') or sec
            lines.append(f"• [{_lbl}] dòng {idx + 1}: {'; '.join(msgs)}")
        more = f"\n… và {n_err - 15} dòng lỗi khác." if n_err > 15 else ""
        self._show_msg("Kiểm tra: còn lỗi",
            f"Còn {n_err} dòng lỗi (đã tô đỏ trong bảng):\n\n"
            + "\n".join(lines) + more, kind="warning")

    def _apply_validate_resolutions(self, resolutions, errors):
        """Áp kết quả batch fuzzy dialog vào TOÀN BỘ _thdm_preview_data,
        xóa lỗi tương ứng khỏi errors và lk_fails."""
        if not resolutions:
            return
        lk_fails = self._thdm_val_lk_fails

        # Bước 1: Từ lk_fails, build bảng (sec, sql_col, raw_str) → (chosen_val, key, rr)
        # dùng để match tất cả dòng trong preview_data (không chỉ dòng trong lk_fails).
        apply_map = {}   # (sec, sql_col, raw_str) → (chosen_val, key, rr)
        for (sec, idx), fails in lk_fails.items():
            for sql_col, raw, key, rr in fails:
                chosen_val = resolutions.get((str(raw), key))
                if chosen_val is not None:
                    apply_map[(sec, sql_col, str(raw))] = (chosen_val, key, rr)

        if not apply_map:
            return

        # Bước 2: Ghi nhận mapping vào _fuzzy_resolutions cho TẤT CẢ dòng khớp giá trị
        # preview_data KHÔNG bị ghi đè — raw code giữ nguyên để lookup lại đúng
        if not hasattr(self, '_fuzzy_resolutions'):
            self._fuzzy_resolutions = {}
        fixed_keys = set()  # (sec, sql_col, raw_str) đã được áp
        for row in self._thdm_preview_data:
            sec = row.get('_section', '')
            for (map_sec, sql_col, raw_str), (chosen_val, _key, _rr) in apply_map.items():
                if map_sec != sec:
                    continue
                if str(row.get(sql_col, '')) == raw_str:
                    self._fuzzy_resolutions[(raw_str, _key)] = chosen_val
                    fixed_keys.add((sec, sql_col, raw_str))

        # Bước 3: Dọn lk_fails và errors cho tất cả entry đã được fix
        for (sec, idx), fails in list(lk_fails.items()):
            remaining = []
            for sql_col, raw, key, rr in fails:
                if (sec, sql_col, str(raw)) in fixed_keys:
                    # Xóa msg lỗi tương ứng
                    if (sec, idx) in errors:
                        old_msg = f"Mã '{raw}' không tồn tại trong DB"
                        errors[(sec, idx)] = [
                            m for m in errors[(sec, idx)] if m != old_msg]
                        if not errors[(sec, idx)]:
                            del errors[(sec, idx)]
                else:
                    remaining.append((sql_col, raw, key, rr))
            if remaining:
                lk_fails[(sec, idx)] = remaining
            else:
                del lk_fails[(sec, idx)]
        # Cache đã dùng để resolve → xóa để validate lần sau build lại từ DB
        self._thdm_validate_caches.clear()

    def _thdm_on_cell_edited(self, event, sec):
        """NHỊP 1: Sync ô vừa sửa → preview_data và bỏ tô đỏ ngay lập tức."""
        try:
            vis_row = event.row
            col_idx = event.column
            new_val = event.value
        except (AttributeError, TypeError):
            return
        if vis_row == 0:
            return  # dòng sql_names header

        # ── Resolve data_idx (hoạt động đúng kể cả khi đang lọc ⚠ Dòng lỗi) ──
        vis_indices = self._thdm_sec_visible.get(sec, [])
        data_idx = vis_indices[vis_row - 1] if vis_row - 1 < len(vis_indices) else None
        if data_idx is None:
            return

        # ── Resolve sql_col ──
        col_defs = self._thdm_sec_col_defs.get(sec, [])
        if col_idx >= len(col_defs):
            return
        sql_col = col_defs[col_idx].get('sql_col')
        if not sql_col:
            return

        # ── Cập nhật _thdm_preview_data ──
        sec_rows = [r for r in self._thdm_preview_data if r.get('_section') == sec]
        if data_idx >= len(sec_rows):
            return
        sec_rows[data_idx][sql_col] = new_val

        # ── Immediate dehighlight: bỏ tô đỏ dòng vừa sửa ngay lập tức ──
        err_key = (sec, data_idx)
        if err_key in self._thdm_val_errors:
            self._thdm_val_errors.pop(err_key, None)
            self._thdm_val_lk_fails.pop(err_key, None)
            if self._thdm_filter_mode == "errors_only":
                # Khi đang lọc ⚠: re-render để dòng vừa sửa biến khỏi view lỗi
                self._thdm_apply_filter("errors_only")
            else:
                # Chỉ đổi tag + rehighlight — không reload data vào tksheet
                tr = self._thdm_sec_trees.get(sec)
                if tr is not None and vis_row < len(tr._rows):
                    iid = tr._rows[vis_row]
                    tr._row_tags[iid] = ("row_normal" if data_idx % 2 == 0
                                         else "row_alt")
                    tr.rehighlight()
            self._thdm_update_filter_btns()

        # Data đã thay đổi → cần Kiểm tra lại trước khi Insert
        self._thdm_validated = False
        try:
            self.btn_thdm_insert.configure(state="disabled")
        except Exception:
            pass
        try:
            self.lbl_thdm_err_bar.configure(text="")
        except Exception:
            pass

    def _thdm_on_paste(self, event, sec):
        """Sync Ctrl+V paste → _thdm_preview_data và SheetTable._data.
        Gọi bởi extra_bindings ctrl_v sau khi tksheet hoàn tất paste nội bộ."""
        try:
            cells = event.get("cells", {}).get("table", {})
        except (AttributeError, TypeError):
            return
        if not cells:
            return

        vis_indices = self._thdm_sec_visible.get(sec, [])
        col_defs    = self._thdm_sec_col_defs.get(sec, [])
        sec_rows    = [r for r in self._thdm_preview_data if r.get('_section') == sec]
        tr          = self._thdm_sec_trees.get(sec)

        changed_data_idxs = set()
        for (datarn, datacn) in cells.keys():
            # event_data["cells"]["table"] lưu giá trị CŨ (trước paste, dùng cho undo).
            # Giá trị MỚI đã được tksheet ghi vào MT.data → đọc từ đó.
            if datarn == 0:
                continue  # hàng sql_names (header phụ)
            vis_pos = datarn - 1
            if vis_pos >= len(vis_indices):
                continue
            data_idx = vis_indices[vis_pos]
            if data_idx >= len(sec_rows):
                continue
            if datacn >= len(col_defs):
                continue
            sql_col = col_defs[datacn].get('sql_col')
            if not sql_col:
                continue
            # Đọc giá trị MỚI từ tksheet MT.data (đã được paste ghi vào)
            try:
                new_val = tr.sheet.MT.data[datarn][datacn]
            except (IndexError, AttributeError, TypeError):
                continue
            # Ghi vào _thdm_preview_data (nguồn sự thật)
            sec_rows[data_idx][sql_col] = new_val
            # Đồng bộ SheetTable._data để _flush() sau này không ghi đè lại
            if datarn < len(tr._rows):
                iid = tr._rows[datarn]
                row_vals = tr._data.get(iid)
                if row_vals is not None and datacn < len(row_vals):
                    row_vals[datacn] = "" if new_val is None else str(new_val)
            changed_data_idxs.add(data_idx)

        if not changed_data_idxs:
            return

        # Xóa highlight lỗi cho các dòng vừa chỉnh sửa
        any_err_cleared = False
        for data_idx in changed_data_idxs:
            err_key = (sec, data_idx)
            if err_key in self._thdm_val_errors:
                self._thdm_val_errors.pop(err_key, None)
                self._thdm_val_lk_fails.pop(err_key, None)
                any_err_cleared = True

        if any_err_cleared:
            if self._thdm_filter_mode == "errors_only":
                self._thdm_apply_filter("errors_only")
            else:
                if tr is not None:
                    for data_idx in changed_data_idxs:
                        try:
                            vis_pos = vis_indices.index(data_idx)
                            iid = tr._rows[vis_pos + 1]  # +1 vì row 0 là header
                            tr._row_tags[iid] = ("row_normal" if data_idx % 2 == 0
                                                 else "row_alt")
                        except (ValueError, IndexError):
                            pass
                    tr.rehighlight()
            self._thdm_update_filter_btns()

        self._thdm_validated = False
        try:
            self.btn_thdm_insert.configure(state="disabled")
        except Exception:
            pass
        try:
            self.lbl_thdm_err_bar.configure(text="")
        except Exception:
            pass

    # ── THDM: Export / View SQL / Insert ──────────────────────────────────────

    def _thdm_export_excel(self):
        if not self._thdm_preview_data:
            return
        import tkinter.filedialog as fd
        path = fd.asksaveasfilename(
            title="Xuất Excel THDM",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="THDM_Export.xlsx")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            # Mỗi section detail 1 sheet riêng (THDM Vật tư / NVL Bổ sung...)
            for sec, col_defs in self._thdm_sec_col_defs.items():
                sec_rows = [r for r in self._thdm_preview_data
                            if r.get('_section') == sec]
                lbl = (self.mapping.get('_CONFIG', {}).get(sec, {}).get('label')
                       or sec)
                ws = wb.create_sheet(title=lbl[:31])
                # Hàng 1 = Ten_Excel (header xanh), Hàng 2 = SQL_Column
                for c, d in enumerate(col_defs, 1):
                    cell = ws.cell(row=1, column=c, value=d['ten_excel'])
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor="1D9E75")
                    cell.alignment = Alignment(horizontal="center")
                    scell = ws.cell(row=2, column=c, value=d['sql_col'])
                    scell.font      = Font(italic=True, color="9CDCFE")
                    scell.alignment = Alignment(horizontal="center")
                # Data rows — theo sql_col từ mapping (bắt đầu hàng 3)
                for i, r in enumerate(sec_rows, 3):
                    for c, d in enumerate(col_defs, 1):
                        ws.cell(row=i, column=c, value=r.get(d['sql_col']))
                # Column widths — convert pixel → char width (~7px per char)
                for c, d in enumerate(col_defs, 1):
                    char_w = max(round(d['width'] / 7), 8)
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(c)].width = char_w
            wb.save(path)
            self._show_export_success("Xuất thành công", f"Đã lưu:\n{path}", path)
        except Exception as e:
            self._show_msg("Lỗi xuất Excel", str(e))

    # ── THDM: Resolve row THVT theo mapping ───────────────────────────────────
    def _resolve_thdm_thvt_row(self, excel_row, biz_doc_id, sort_order, thvt_map, now):
        """
        Resolve 1 dòng THVT từ mapping THDM_THVT.
        Thin wrapper gọi _resolve_row_mapping() dùng chung với BOM.

        excel_row  : dict {sql_col: value} từ _thdm_preview_data
        biz_doc_id : UNIQUEIDENTIFIER vừa tạo cho phiếu THDM (hoặc placeholder)
        sort_order : thứ tự dòng (int, 1-based)
        thvt_map   : mapping.get('THDM_THVT', [])
        now        : datetime đồng nhất cho cả batch
        Returns    : dict {sql_col: value}
        """
        # BizDocId_SO = BizDocId string (ví dụ "11034510FO") — nvarchar trong bảng đích
        _parent = {'BizDocId_SO': self._thdm_selected_order_id}
        ctx = {
            'now':           now,
            'builtin_order': sort_order,
            'doc_id':        biz_doc_id,
            'parent_row':    _parent,
            'parent_fields': {'BizDocId_SO'},
            'ui_values': {
                'creator':    self._thdm_creator_user_id or self._current_creator_user_id,
                'product_id': self._thdm_selected_product_id,
                'order_id':   self._thdm_selected_order_id,
                'period_id':  self._thdm_selected_period_id,
            },
            'skip_nguon': {'SP', 'TinhToan'},
        }
        result = _resolve_row_mapping(
            thvt_map, ctx,
            get_excel_val=lambda rec: excel_row.get(rec['sql_col'])
        )
        result['ProductId'] = None   # child rows không có ProductId (chỉ parent cần)
        return result

    def _thdm_view_sql(self):
        """Hiện SQL INSERT sẽ thực hiện khi nhấn INSERT vào DB."""
        dt = THEMES[ctk.get_appearance_mode()]
        if not self._thdm_preview_data:
            return
        from datetime import datetime as _dt
        order_id   = self._thdm_selected_order_id or "NULL"
        bom_ids    = list(self._thdm_checked_ids)
        bom_list   = ",".join(str(i) for i in bom_ids)
        mapping    = load_mapping()
        cfg        = mapping.get('_CONFIG', {})
        header_map = mapping.get('THDM_HEADER', [])
        sp_config  = mapping.get('_SP_CONFIG', [])
        child_sections = [
            sec for sec, info in cfg.items()
            if info.get('parent_section') == 'THDM_HEADER'
        ]
        now        = _dt.now()

        ui_vals = {
            'product_id': self._thdm_selected_product_id,
            'order_id':   self._thdm_selected_order_id,   # BizDocId string
            'creator':    self._thdm_creator_user_id or self._current_creator_user_id,
            'period_id':  self._thdm_selected_period_id,
        }

        def _fmt(v):
            if v is None:                                return "NULL"
            if isinstance(v, str) and v.startswith('@'): return v
            if isinstance(v, _dt):                       return f"'{v:%Y-%m-%d %H:%M:%S}'"
            if isinstance(v, str):                       return f"N'{v.replace(chr(39), chr(39)*2)}'"
            if isinstance(v, float):                     return f"{v:.4f}"
            return str(v)

        # ── Resolve THDM_HEADER non-SP fields (CoDinh / HeThong / UILookup) ──
        hdr_out = {}
        for rec in header_map:
            sql_col  = rec['sql_col']
            nguon_dl = rec.get('nguon_dl', '')
            mac_dinh = rec.get('mac_dinh', '')
            kieu     = rec.get('kieu_dl', '').lower()
            if nguon_dl == 'CoDinh':
                if str(mac_dinh).upper() == 'NULL':
                    hdr_out[sql_col] = None
                else:
                    try:
                        if kieu in ('int', 'bigint'):
                            hdr_out[sql_col] = int(float(mac_dinh))
                        elif kieu in ('numeric', 'float', 'decimal'):
                            hdr_out[sql_col] = float(mac_dinh)
                        else:
                            hdr_out[sql_col] = mac_dinh
                    except (ValueError, TypeError):
                        hdr_out[sql_col] = mac_dinh
            elif nguon_dl == 'HeThong':
                hdr_out[sql_col] = now if mac_dinh == 'NOW' else None
            elif nguon_dl == 'UILookup':
                hdr_out[sql_col] = ui_vals.get(mac_dinh)
            # SP fields → placeholder (không chạy SP thật để tránh tiêu sequence)

        # EmployeeId: UILookup mac_dinh='creator' → UserId; override với EmployeeId thật
        _emp_id = getattr(self, '_thdm_creator_employee_id', None)
        if _emp_id is not None and 'EmployeeId' in hdr_out:
            hdr_out['EmployeeId'] = _emp_id

        hdr_out['BOMIdList'] = bom_list

        # SP fields: hiện dưới dạng comment placeholder
        sp_thdm_hdr = [s for s in sp_config
                       if s.get('section') == 'THDM_HEADER'
                       and s.get('isactive') == '1'
                       and s.get('sp_name', '').lower() != 'lookup']
        for sp_rec in sp_thdm_hdr:
            sql_col = sp_rec.get('sql_column') or sp_rec.get('sql_col', '')
            out_flds = [f.strip() for f in sp_rec.get('outputfields', '').split(',') if f.strip()]
            target = out_flds[0] if out_flds else sql_col
            if target:
                hdr_out[target] = f"/*{sp_rec.get('sp_name','SP')}*/"

        # Loại SP-lookup intermediates (Ws_Id) + IsDraftData khỏi INSERT
        _sp_temp_header = {
            s['sql_column'] for s in sp_config
            if s.get('section') == 'THDM_HEADER'
            and s.get('sp_name', '').lower() == 'lookup'
            and s.get('isactive') == '1'
        }
        _hdr_skip   = _sp_temp_header | {'IsDraftData'}
        valid_hcols = {r['sql_col'] for r in header_map} | {'BOMIdList'}
        ins_hcols   = [c for c in hdr_out if c in valid_hcols and c not in _hdr_skip]

        view_hdr = mapping.get('_CONFIG', {}).get('THDM_HEADER', {}).get(
            'view_insert', 'vB30BizDocDemand_Edit')

        # ── Build cache cho Excel lookup (tất cả child sections) ────────────
        _all_lk_recs_by_sec = {}
        _all_view_caches    = {}
        try:
            _vc = self._get_db_conn()
            for _sec in child_sections:
                _sm  = mapping.get(_sec, [])
                _lks = [r for r in _sm
                        if r.get('nguon_dl') == 'Excel'
                        and r.get('kieu_lookup') and r.get('bang_master')
                        and r.get('truong_lay_ve')]
                _all_lk_recs_by_sec[_sec] = _lks
                if _lks:
                    _all_view_caches.update(self._build_all_caches(_vc, _lks))
            _vc.close()
        except Exception:
            pass

        # ── Build SQL text ────────────────────────────────────────────────────
        hdr_cols_str = ', '.join(ins_hcols)
        hdr_vals_str = ', '.join(
            hdr_out[c] if isinstance(hdr_out[c], str) and hdr_out[c].startswith('/*')
            else _fmt(hdr_out[c])
            for c in ins_hcols
        )

        lines = [
            "-- ===== THDM INSERT preview (mapping-driven) =====",
            f"-- Dự án     (ProductId):  {self._thdm_selected_product_id}",
            f"-- Đơn hàng  (BizDocId_SO): {order_id}",
            f"-- BOM chọn  ({len(bom_ids)} BOM): {bom_list}",
            f"-- Người lập (CreatedBy):  {ui_vals['creator']}",
            f"-- Đợt       (PeriodId):   {self._thdm_selected_period_id}",
            "-- NOTE: /*SP_NAME*/ = giá trị do SP sinh ra khi INSERT thật",
            "",
            f"-- [1] Tạo phiếu THDM (parent) — {len(ins_hcols)} cột từ THDM_HEADER mapping",
            f"DECLARE @BizDocId VARCHAR(24) = /*{sp_thdm_hdr[0]['sp_name'] if sp_thdm_hdr else 'SP'}*/NULL;",
            f"INSERT INTO {view_hdr}",
            f"    ({hdr_cols_str})",
            "VALUES",
            f"    ({hdr_vals_str});",
            "",
            f"-- [2] Insert {len(self._thdm_preview_data)} dòng vật tư (child — tất cả sections)",
        ]

        # Group rows by section để build mỗi INSERT block riêng
        total_rows = self._thdm_preview_data
        for sec in child_sections:
            sec_map    = mapping.get(sec, [])
            view_sec   = cfg.get(sec, {}).get('view_insert', 'B30BizDocDemandBOM')
            lk_recs    = _all_lk_recs_by_sec.get(sec, [])
            sec_rows   = [r for r in total_rows if r.get('_section') == sec]
            if not sec_rows:
                continue

            # Resolve sample row để lấy cols
            sample = self._resolve_thdm_thvt_row(sec_rows[0], "@BizDocId", 1, sec_map, now)
            cols   = list(sample.keys())

            lines += [
                "",
                f"-- Section {sec} ({len(sec_rows)} dòng)",
                f"INSERT INTO {view_sec} ({', '.join(cols)})",
                "VALUES",
            ]
            for idx, r in enumerate(sec_rows):
                resolved = self._resolve_thdm_thvt_row(r, "@BizDocId", idx + 1, sec_map, now)
                for _lrec in lk_recs:
                    _sc  = _lrec.get('sql_col', '')
                    _kl  = _lrec.get('kieu_lookup', '')
                    _bm  = _lrec.get('bang_master', '')
                    _dk  = _lrec.get('dieu_kien_master', '')
                    _ss  = _lrec.get('truong_so_sanh', '')
                    _lv  = _lrec.get('truong_lay_ve', '')
                    _raw = resolved.get(_sc)
                    _ck  = _all_view_caches.get((_bm, _dk, _ss, _lv), [])
                    if _raw is not None and _ck:
                        _found, _ = self._lookup_generic(_raw, _ck, _kl, _no_popup=True)
                        resolved[_sc] = _found
                comma    = "," if idx < len(sec_rows) - 1 else ";"
                vals_str = ", ".join(_fmt(resolved.get(c)) for c in cols)
                lines.append(f"    ({vals_str}){comma}")

        sql_text = "\n".join(lines)

        dlg = ctk.CTkToplevel(self)
        dlg.title("SQL preview — THDM")
        dlg.geometry("900x580")
        dlg.grab_set()
        txt = tk.Text(dlg, bg=dt["bg_main"], fg="#D4D4D4",
            font=("Consolas", 10), wrap="none",
            insertbackground="#D4D4D4")
        txt.pack(fill=tk.BOTH, expand=True, padx=PAD_SM, pady=PAD_SM)
        txt.insert("1.0", sql_text)
        txt.configure(state="disabled")
        ctk.CTkButton(dlg, text="📋  Copy",
            command=lambda: (self.clipboard_clear(), self.clipboard_append(sql_text)),
            width=80, height=28).pack(pady=(0,8))

    def _thdm_insert_db(self):
        if not self._thdm_preview_data:
            self._show_msg("Không có dữ liệu",
                "Vui lòng tổng hợp dữ liệu từ Excel trước.", kind="info")
            return
        # Phải Kiểm tra đạt trước (giống luồng Import bên BOM)
        if not self._thdm_validated:
            self._show_msg("Chưa kiểm tra",
                "Hãy bấm ④ Kiểm tra trước khi tạo THDM.", kind="warning")
            return
        if self._thdm_val_errors:
            self._show_msg("Còn lỗi",
                f"Còn {len(self._thdm_val_errors)} dòng lỗi (tô đỏ trong bảng). "
                "Sửa file Excel/mapping rồi Tổng hợp và Kiểm tra lại.", kind="error")
            return
        # Bắt buộc chọn đủ 4 lookup trước khi INSERT
        _missing = []
        if not self._thdm_selected_product_id:
            _missing.append("Dự án")
        if not self._thdm_selected_order_id:
            _missing.append("Đơn hàng")
        if not (self._thdm_creator_user_id or self._current_creator_user_id):
            _missing.append("Nhân viên")
        if not self._thdm_selected_period_id:
            _missing.append("Đợt")
        if _missing:
            self._show_msg("Thiếu thông tin",
                "Vui lòng chọn đầy đủ trước khi Tạo THDM:\n\n"
                + "\n".join(f"  •  {f}" for f in _missing),
                kind="warning")
            return
        # Bước ④ đã xử lý hết fuzzy — kiểm tra còn dòng lỗi lookup không
        n_lk_fails = len(self._thdm_val_lk_fails)
        if n_lk_fails > 0:
            self._show_msg(
                "Còn lỗi mapping",
                f"Vẫn còn {n_lk_fails} dòng chưa tra được mã trong DB.\n"
                "Vui lòng sửa hết lỗi (bấm ④ Kiểm tra lại) trước khi Tạo THDM.",
                kind="warning")
            return
        n    = len(self._thdm_preview_data)
        boms = len(self._thdm_checked_ids)
        ok   = self._ask_msg("Xác nhận tạo THDM",
            f"Sẽ tạo 1 phiếu THDM mới với:\n"
            f"  • {boms} BOM được chọn\n"
            f"  • {n:,} dòng vật tư từ Excel\n\n"
            f"Tiếp tục?")
        if not ok:
            return
        # Bỏ bước prescan: bước ④ đã resolve hết → chạy thẳng INSERT
        # _fuzzy_batch_done=True suppresses popup cho mọi case còn sót
        self._fuzzy_resolutions = {}
        self._fuzzy_batch_done  = True
        self.btn_thdm_insert.configure(state="disabled", text="⏳  Đang tạo phiếu...")
        self._loading_dlg = self._make_loading_popup(
            "Đang tạo phiếu THDM...\nVui lòng không đóng cửa sổ.")
        threading.Thread(target=self._run_thdm_insert_bg, daemon=True).start()

    def _thdm_prescan_worker(self):
        """Background: quét _thdm_preview_data để thu thập fuzzy cases trước khi INSERT."""
        conn = None
        try:
            mapping = load_mapping()
            cfg = mapping.get('_CONFIG', {})
            child_sections = [sec for sec, info in cfg.items()
                              if info.get('parent_section') == 'THDM_HEADER']
            conn = self._get_db_conn()
            rows = self._thdm_preview_data
            for sec in child_sections:
                sec_map = mapping.get(sec, [])
                _fuzzy_recs = [
                    rec for rec in sec_map
                    if rec.get('nguon_dl') == 'Excel'
                    and rec.get('kieu_lookup', '') in ('fuzzy_code', 'fuzzy_name')
                    and rec.get('bang_master', '')
                    and rec.get('truong_so_sanh', '')
                    and rec.get('truong_lay_ve', '')
                ]
                if not _fuzzy_recs:
                    continue
                _caches = self._build_all_caches(conn, _fuzzy_recs)
                sec_rows = [r for r in rows if r.get('_section') == sec]
                for idx, row in enumerate(sec_rows):
                    for _lrec in _fuzzy_recs:
                        _bm  = _lrec.get('bang_master', '')
                        _dk  = _lrec.get('dieu_kien_master', '')
                        _ss  = _lrec.get('truong_so_sanh', '')
                        _lv  = _lrec.get('truong_lay_ve', '')
                        _sc  = _lrec.get('sql_col', '')
                        _kl  = _lrec.get('kieu_lookup', '')
                        _raw = row.get(_sc)
                        if _raw is None:
                            continue
                        _ck  = (_bm, _dk, _ss, _lv)
                        _cache = _caches.get(_ck, [])
                        if not _cache:
                            continue
                        self._fuzzy_ctx = {'section': sec, 'field': _sc, 'row_idx': idx}
                        self._lookup_generic(_raw, _cache, _kl, _cache_key=_ck)
        except Exception:
            pass
        finally:
            self._fuzzy_collect_mode = False
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
        self.after(0, self._after_thdm_prescan)

    def _after_thdm_prescan(self):
        """Main thread: sau prescan — hiện batch fuzzy dialog (nếu có) rồi bắt đầu INSERT."""
        try:
            if getattr(self, '_thdm_prescan_dlg', None):
                self._thdm_prescan_dlg.destroy()
                self._thdm_prescan_dlg = None
        except Exception:
            pass

        if getattr(self, '_fuzzy_pending', None):
            _resolved = self._show_batch_fuzzy_dialog(self._fuzzy_pending)
            self._fuzzy_resolutions = _resolved if _resolved is not None else {}
        else:
            self._fuzzy_resolutions = {}
        self._fuzzy_batch_done = True

        self._loading_dlg = self._make_loading_popup(
            "Đang tạo phiếu THDM...\nVui lòng không đóng cửa sổ.")
        threading.Thread(target=self._run_thdm_insert_bg, daemon=True).start()

    def _run_thdm_insert_bg(self):
        conn = None
        try:
            from datetime import datetime as _dt
            conn       = self._get_db_conn()
            cur        = conn.cursor()
            order_id   = self._thdm_selected_order_id
            bom_ids    = list(self._thdm_checked_ids)
            rows       = self._thdm_preview_data
            bom_list   = ",".join(str(i) for i in bom_ids)
            mapping    = load_mapping()
            cfg        = mapping.get('_CONFIG', {})
            header_map = mapping.get('THDM_HEADER', [])
            sp_config  = mapping.get('_SP_CONFIG', [])
            child_sections = [
                sec for sec, info in cfg.items()
                if info.get('parent_section') == 'THDM_HEADER'
            ]
            now        = _dt.now()

            # UI values cho UILookup branch
            ui_vals = {
                'product_id': self._thdm_selected_product_id,
                'order_id':   self._thdm_selected_order_id,   # BizDocId string "11034510FO"
                'creator':    self._thdm_creator_user_id or self._current_creator_user_id,
                'period_id':  self._thdm_selected_period_id,
            }
            # ── 1) Resolve THDM_HEADER từ mapping ──────────────────────────
            row_out = {}
            for rec in header_map:
                sql_col  = rec['sql_col']
                nguon_dl = rec.get('nguon_dl', '')
                mac_dinh = rec.get('mac_dinh', '')
                kieu     = rec.get('kieu_dl', '').lower()

                if nguon_dl == 'CoDinh':
                    if mac_dinh.upper() == 'NULL':
                        row_out[sql_col] = None
                    else:
                        try:
                            if kieu in ('int', 'bigint'):
                                row_out[sql_col] = int(float(mac_dinh))
                            elif kieu in ('numeric', 'float', 'decimal'):
                                row_out[sql_col] = float(mac_dinh)
                            else:
                                row_out[sql_col] = mac_dinh
                        except (ValueError, TypeError):
                            row_out[sql_col] = mac_dinh
                elif nguon_dl == 'HeThong':
                    if mac_dinh == 'NOW':
                        row_out[sql_col] = now
                elif nguon_dl == 'UILookup':
                    row_out[sql_col] = ui_vals.get(mac_dinh)
                # SP fields resolved in step 2

            # EmployeeId: UILookup mac_dinh='creator' trả về UserId nhưng cột cần EmployeeId (B20Employee.Id)
            _emp_id = getattr(self, '_thdm_creator_employee_id', None)
            if _emp_id is not None and 'EmployeeId' in row_out:
                row_out['EmployeeId'] = _emp_id

            # BOMIdList: trường đặc biệt không có trong mapping
            row_out['BOMIdList'] = bom_list

            # ── 2) Run _SP_CONFIG cho THDM_HEADER theo thứ tự ──────────────
            # Dùng _call_sp / _sp_lookup (cùng BOM) để xử lý đúng parameterize
            sp_thdm = [s for s in sp_config
                       if s.get('section') == 'THDM_HEADER'
                       and s.get('isactive') == '1']
            for sp_rec in sp_thdm:
                sql_col  = sp_rec.get('sql_column', '')
                sp_name  = sp_rec.get('sp_name', '')
                params_s = sp_rec.get('params', '')
                fallback = sp_rec.get('fallback', '')
                out_flds = [f.strip() for f in sp_rec.get('outputfields', '').split(',') if f.strip()]
                try:
                    if sp_name.lower() == 'lookup':
                        val = self._sp_lookup(conn, params_s, row_out)
                        row_out[sql_col] = val if val is not None else (int(fallback) if fallback else None)
                    else:
                        r = self._call_sp(conn, sp_name, params_s, row_out)
                        if out_flds:
                            row_out[out_flds[0]] = r
                        elif sql_col:
                            row_out[sql_col] = r if r is not None else (fallback or None)
                except Exception as e_sp:
                    if fallback and sql_col:
                        row_out[sql_col] = fallback
                    else:
                        raise RuntimeError(f"SP '{sp_name}' thất bại và không có fallback: {e_sp}") from e_sp

            # ── 3) INSERT parent — dynamic từ row_out ──────────────────────
            view_hdr   = mapping.get('_CONFIG', {}).get('THDM_HEADER', {}).get(
                'view_insert', 'vB30BizDocDemand_Edit')
            # Loại bỏ SP-lookup intermediates (Ws_Id...) — dùng làm param SP
            # nhưng không phải column của view → dùng cùng pattern BOM _sp_temp_fields
            _sp_temp_header = {
                s['sql_column'] for s in sp_config
                if s.get('section') == 'THDM_HEADER'
                and s.get('sp_name', '').lower() == 'lookup'
                and s.get('isactive') == '1'
            }
            # IsDraftData: không tồn tại trong B30BizDocDemand → exclude
            _thdm_hdr_skip = _sp_temp_header | {'IsDraftData'}
            valid_cols = {r['sql_col'] for r in header_map} | {'BOMIdList'}
            ins_cols   = [c for c in row_out if c in valid_cols and c not in _thdm_hdr_skip]
            ins_vals   = [row_out[c] for c in ins_cols]
            ph         = ', '.join('?' * len(ins_cols))
            cur.execute(
                f"INSERT INTO {view_hdr} ({', '.join(ins_cols)}) VALUES ({ph})",
                ins_vals)

            biz_doc_id = row_out.get('BizDocId')

            # ── 4) INSERT child sections — generic loop ────────────────────────
            _insert_count = 0
            for sec in child_sections:
                sec_cfg    = cfg[sec]
                sec_map    = mapping.get(sec, [])
                expand_muc = sec_cfg.get('expand_muc', False)
                view_sec   = sec_cfg.get('view_insert', 'B30BizDocDemandBOM')

                # Rows thuộc section này
                sec_rows = [r for r in rows if r.get('_section') == sec]
                if not sec_rows:
                    continue

                # SP_HOOK BeforeInsert cho section này
                sec_hooks_before = [
                    h for h in mapping.get('SP_HOOK', [])
                    if h.get('section', '') == sec
                    and h.get('event', '').lower() == 'beforeinsert'
                    and h.get('isactive', '') == '1'
                ]

                # Build cache cho Excel lookup fields (ItemId, CustomerId, ...)
                _lk_recs = [
                    rec for rec in sec_map
                    if rec.get('nguon_dl') == 'Excel'
                    and rec.get('kieu_lookup', '')
                    and rec.get('bang_master', '')
                    and rec.get('truong_lay_ve', '')
                ]
                _caches = self._build_all_caches(conn, _lk_recs) if _lk_recs else {}

                # MucLookup + DetailRowId_SO chỉ cho sections có expand_muc=True
                _detail_so_cache = []
                _detail_so_col   = 'DetailRowId_SO'
                if expand_muc:
                    _detail_so_rec = next(
                        (r for r in sec_map
                         if r.get('nguon_dl') == 'MucLookup'
                         and r.get('bang_master')
                         and r.get('truong_so_sanh')
                         and r.get('truong_lay_ve')),
                        None
                    )
                    if _detail_so_rec:
                        try:
                            _detail_so_cache = self._build_cache_generic(
                                conn,
                                _detail_so_rec['bang_master'],
                                _detail_so_rec.get('dieu_kien_master', ''),
                                _detail_so_rec['truong_so_sanh'],
                                _detail_so_rec['truong_lay_ve'],
                            )
                        except Exception:
                            pass
                        _detail_so_col = _detail_so_rec['sql_col']

                for idx, r in enumerate(sec_rows, 1):
                    try:
                        resolved = self._resolve_thdm_thvt_row(
                            r, biz_doc_id, idx, sec_map, now)
                        if not resolved:
                            continue

                        # Convert Excel code → DB Id qua lookup (ItemId, ...)
                        for _lrec in _lk_recs:
                            _sc    = _lrec.get('sql_col', '')
                            _kl    = _lrec.get('kieu_lookup', '')
                            _bm    = _lrec.get('bang_master', '')
                            _dk    = _lrec.get('dieu_kien_master', '')
                            _ss    = _lrec.get('truong_so_sanh', '')
                            _lv    = _lrec.get('truong_lay_ve', '')
                            _raw   = resolved.get(_sc)
                            _cache = _caches.get((_bm, _dk, _ss, _lv), [])
                            if _raw is not None and _cache:
                                _found, _ = self._lookup_generic(
                                    _raw, _cache, _kl, _cache_key=(_bm, _dk, _ss, _lv))
                                resolved[_sc] = _found

                        # Compound lookup DetailRowId_SO (chỉ expand_muc sections)
                        if _detail_so_cache:
                            _muc_raw = resolved.get(_detail_so_col)
                            if _muc_raw is not None:
                                _muc_key  = str(_muc_raw).replace('Muc ', '').strip()
                                _compound = (_muc_key, str(order_id))
                                _found_id, _ = self._lookup_generic(
                                    _compound, _detail_so_cache, 'exact')
                                resolved[_detail_so_col] = _found_id
                            # Factory-only / lookup fail → '' thay NULL
                            if resolved.get(_detail_so_col) is None:
                                resolved[_detail_so_col] = ''

                        # SP_HOOK BeforeInsert
                        if sec_hooks_before:
                            _run_row_sp_hooks(conn, sec_hooks_before, resolved,
                                              log_fn=None)

                        cols_c = list(resolved.keys())
                        vals_c = [resolved[c] for c in cols_c]
                        ph_c   = ", ".join("?" * len(cols_c))
                        cur.execute(
                            f"INSERT INTO {view_sec} ({', '.join(cols_c)}) VALUES ({ph_c})",
                            vals_c)
                        _insert_count += 1
                        if _insert_count % 50 == 0:
                            self.after(0, lambda n=_insert_count, t=len(rows):
                                self._update_loading_msg(
                                    f"Đang tạo phiếu THDM...\n"
                                    f"Đã insert {n:,}/{t:,} dòng vật tư"))
                    except Exception as _row_err:
                        raise RuntimeError(
                            f"[{sec}] Lỗi tại dòng #{idx} "
                            f"({r.get('ItemName') or r.get('ItemId') or '?'}): {_row_err}"
                        ) from _row_err

            conn.commit()
            self.after(0, lambda n=_insert_count: self._thdm_insert_done(n, None))
        except Exception as e:
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
            self.after(0, lambda err=e: self._thdm_insert_done(0, str(err)))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _thdm_insert_done(self, n_rows, error):
        try:
            if self._loading_dlg:
                self._loading_dlg.destroy()
                self._loading_dlg = None
        except Exception:
            pass
        self.btn_thdm_insert.configure(
            state="normal", text="💾  ⑤ Tạo THDM")
        _fname = os.path.basename(self._thdm_excel_path or '')
        if error:
            self._db_log('THDM', _fname, self._thdm_selected_order_id, 0, 'LOI', error)
            self._show_msg("Lỗi tạo THDM", error)
            return
        self._db_log('THDM', _fname, self._thdm_selected_order_id, n_rows, 'OK',
                     f"Đã tạo phiếu THDM với {n_rows:,} dòng vật tư")
        self._show_msg("Thành công",
            f"✅  Đã tạo phiếu THDM với {n_rows:,} dòng vật tư.", kind="info")
        self._log(
            f"Đơn hàng {self._thdm_selected_order_id}",
            "THDM", str(n_rows), "OK",
            f"Đã tạo phiếu THDM với {n_rows:,} dòng vật tư",
            "ok")

    # ─ Tab 4: Lịch sử Log ──────────────────────────────────────────────────────────────────────
    def _build_log_panel(self):
        dt = THEMES[ctk.get_appearance_mode()]
        self._log_expanded = False
        self.log_panel = ctk.CTkFrame(self, corner_radius=0, height=40)
        self.log_panel.grid(row=2, column=0, sticky="ew")
        self.log_panel.grid_propagate(False)
        self.log_panel.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(self.log_panel, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=PAD_SM, pady=6)
        hdr.grid_columnconfigure(0, weight=0)   # title — fixed
        hdr.grid_columnconfigure(1, weight=1)   # error label — fills remaining space

        ctk.CTkLabel(hdr, text="📋  Lịch sử Import",
            font=ctk.CTkFont(*FONT_MD_B)).grid(
            row=0, column=0, sticky="w")

        # Statusbar lỗi THDM — cập nhật bởi _thdm_on_row_select khi click dòng bảng
        self.lbl_thdm_err_bar = ctk.CTkLabel(
            hdr, text="",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("#DC2626", "#FF6B6B"),
            anchor="w", fg_color="transparent")
        self.lbl_thdm_err_bar.grid(row=0, column=1, sticky="ew", padx=(12, 4))

        self.btn_log_db = ctk.CTkButton(hdr, text="☁  Log DB",
            command=self._load_db_log,
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#007ACC"),
            border_color=("#3B82F6","#007ACC"),
            hover_color=("#DBEAFE","#1E3A5F"),
            font=ctk.CTkFont(*FONT_MD),
            width=90, height=24, corner_radius=6)
        self.btn_log_db.grid(row=0, column=2, padx=(4, 0))
        Tooltip(self.btn_log_db,
                lambda: "Tải 200 dòng log import gần nhất từ database"
                        " (bảng BOMTool_ImportLog)")
        ctk.CTkButton(hdr, text="🗑  Xóa",
            command=self._clear_log,
            fg_color="transparent", border_width=1,
            text_color="gray", border_color=("gray50","gray50"),
            hover_color=("gray90","gray25"),
            font=ctk.CTkFont(*FONT_MD),
            width=70, height=24, corner_radius=6).grid(
            row=0, column=3, padx=(4, 4))
        self.btn_log_toggle = ctk.CTkButton(hdr,
            text="▲  Mở log",
            command=self._toggle_log,
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#007ACC"),
            border_color=("#3B82F6","#007ACC"),
            hover_color=("#DBEAFE","#1E3A5F"),
            font=ctk.CTkFont(*FONT_MD_B),
            width=90, height=24, corner_radius=6)
        self.btn_log_toggle.grid(row=0, column=4)

        self.log_content = tk.Frame(self.log_panel, bg=dt["bg_main"], height=200)

        LOG_COLS = ["Thời gian","Tên file","Sheet","Tổng dòng","Trạng thái","Chi tiết"]
        if _HAS_TKSHEET:
            self.log_tree = SheetTable(self.log_content, columns=LOG_COLS)
            self.log_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(self.log_content, orient=tk.VERTICAL)
            self.log_tree = ttk.Treeview(self.log_content, style="BOM.Treeview",
                show="headings", columns=LOG_COLS, yscrollcommand=vsb.set, height=7)
            vsb.configure(command=self.log_tree.yview)
            self.log_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
        for col, w in zip(LOG_COLS, [155,250,80,80,100,500]):
            _la = guess_col_align(col)
            self.log_tree.heading(col, text=col, anchor=_la)
            self.log_tree.column(col, width=w, anchor=_la, stretch=(col=="Chi tiết"))
        self.log_content.rowconfigure(0, weight=1)
        self.log_content.columnconfigure(0, weight=1)
        self.log_tree.tag_configure("ok",   foreground=dt["log_ok"])
        self.log_tree.tag_configure("warn", foreground=dt["log_warn"])
        self.log_tree.tag_configure("err",  foreground=dt["log_err"])
        self._bind_copy(self.log_tree, "Log")

    def _toggle_log(self):
        if self._log_expanded:
            self.log_content.grid_forget()
            self.log_panel.configure(height=40)
            self.btn_log_toggle.configure(text="▲  Mở log")
            self._log_expanded = False
        else:
            self.log_content.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=PAD_XS, pady=(0,4))
            self.log_panel.configure(height=230)
            self.btn_log_toggle.configure(text="▼  Đóng log")
            self._log_expanded = True

    # ─ Helpers ───────────────────────────────────────────────────────────────────────────────────
    def _log(self, fname, sheet, rows, status, detail, tag):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_tree.insert("", 0,
            values=(ts, fname, sheet, rows, status, detail), tags=(tag,))
        if tag in ('error', 'warn'):
            try:
                _lp = os.path.join(os.path.dirname(sys.executable)
                                   if getattr(sys, 'frozen', False)
                                   else os.path.dirname(os.path.abspath(__file__)),
                                   'boho_import.log')
                with open(_lp, 'a', encoding='utf-8') as _lf:
                    _lf.write(f"[{ts}] [{tag.upper()}] {sheet} | {status} | {detail}\n")
            except Exception:
                pass

    def _clear_log(self):
        self.log_tree.delete(*self.log_tree.get_children())

    # ── DB Log: bảng BOMTool_ImportLog, nằm ở DB riêng [BOMTool] (không
    # chung DB với Bravo) — tạo bằng sql/create_[BOMTool].sql ──────────────
    DB_LOG_DATABASE = '[BOMTool]'
    DB_LOG_TABLE = f'{DB_LOG_DATABASE}.dbo.BOMTool_ImportLog'

    def _db_log(self, action, fname, ref_id, n_rows, status, detail):
        """Ghi 1 dòng log import vào bảng DB (chạy nền).
        Bảng chưa được tạo trên DB → bỏ qua êm, không làm phiền user."""
        def _worker():
            try:
                import getpass
                import socket
                conn = self._get_db_conn(timeout_sec=5)
                cur  = conn.cursor()
                cur.execute(
                    f"INSERT INTO {self.DB_LOG_TABLE} "
                    "(Computer, LoginUser, Creator, Action, FileName, RefId, "
                    "TotalRows, Status, Detail, AppVersion) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (socket.gethostname()[:100],
                     getpass.getuser()[:100],
                     str(getattr(self, '_current_creator_user_id', '') or '')[:100],
                     action,
                     (fname or '')[:255],
                     str(ref_id or '')[:50],
                     int(n_rows or 0),
                     status,
                     (detail or '')[:4000],
                     APP_VERSION))
                conn.commit()
                conn.close()
            except Exception:
                pass
        threading.Thread(target=_worker, daemon=True).start()

    def _load_db_log(self):
        self.btn_log_db.configure(state="disabled", text="⏳ Đang tải...")
        threading.Thread(target=self._load_db_log_worker, daemon=True).start()

    def _load_db_log_worker(self):
        try:
            conn = self._get_db_conn(timeout_sec=5)
            cur  = conn.cursor()
            cur.execute(
                f"SELECT TOP 200 LogTime, FileName, Action, TotalRows, Status, "
                f"Detail, Computer, LoginUser "
                f"FROM {self.DB_LOG_TABLE} ORDER BY Id DESC")
            rows = cur.fetchall()
            conn.close()
            self.after(0, lambda: self._load_db_log_done(rows, None))
        except Exception as e:
            self.after(0, lambda err=str(e): self._load_db_log_done(None, err))

    def _load_db_log_done(self, rows, error):
        self.btn_log_db.configure(state="normal", text="☁  Log DB")
        if error is not None:
            if (self.DB_LOG_DATABASE in error or 'Invalid object name' in error
                    or 'does not exist' in error or 'Cannot open database' in error):
                self._show_msg("Bảng log chưa có trên DB",
                    f"Database/bảng {self.DB_LOG_TABLE} chưa được tạo.\n\n"
                    "Gửi file sql\\create_[BOMTool].sql cho khách hàng chạy,\n"
                    "sau đó bấm lại nút này là xem được log.", kind="info")
            else:
                self._show_msg("Lỗi tải log DB", error, kind="error")
            return
        self._clear_log()
        for r in rows:
            ts     = r[0].strftime("%Y-%m-%d %H:%M:%S") if r[0] else ''
            status = (r[4] or '').strip()
            tag    = 'ok' if status.upper() == 'OK' else 'err'
            detail = f"{r[5] or ''}  [{r[6] or ''} / {r[7] or ''}]"
            self.log_tree.insert("", "end",
                values=(ts, r[1] or '', r[2] or '', r[3] if r[3] is not None else '',
                        status, detail),
                tags=(tag,))
        if not self._log_expanded:
            self._toggle_log()

    def _set_status(self, text, color=None):
        self.lbl_status.config(text=text, fg=(color or ("gray40","gray60")))

    def _get_password_for_file(self, filepath):
        """
        Hiện password dialog trên main thread, chờ kết quả từ worker thread.
        Trả về: password str hoặc None nếu hủy.
        """
        event    = threading.Event()
        result   = [None]

        def show_on_main():
            result[0] = _ask_excel_password(filepath)
            event.set()

        self.after(0, show_on_main)
        event.wait()
        return result[0]

    def _open_excel_with_password(self, filepath):
        """
        Mở file Excel, tự động hỏi mật khẩu nếu file bị mã hóa.
        Gọi từ worker thread.
        Trả về: (decrypted_bytes hoặc None, error_msg hoặc None)
        - None, None  → file không có password, dùng filepath bình thường
        - bytes, None → file có password, đã giải mã thành công
        - None, msg   → lỗi (hủy hoặc sai password)
        """
        if not _is_encrypted_excel(filepath):
            return None, None  # không cần decrypt

        # Nếu không có msoffcrypto, không thể giải mã — báo người dùng cài
        if not _HAS_MSOFFCRYPTO:
            return None, (
                "File Excel có mật khẩu nhưng thư viện giải mã chưa được cài.\n\n"
                "Chạy lệnh sau trong terminal rồi khởi động lại tool:\n"
                "    pip install msoffcrypto-tool"
            )

        # Hỏi mật khẩu (tối đa 3 lần)
        for attempt in range(3):
            password = self._get_password_for_file(filepath)
            if password is None:
                return None, "cancelled"
            try:
                decrypted = _decrypt_excel(filepath, password)
                return decrypted, None
            except Exception:
                if attempt < 2:
                    event2 = threading.Event()
                    _att = attempt + 1
                    def _warn(a=_att):
                        self._show_msg(
                            "Sai mật khẩu",
                            f"Mật khẩu không đúng (lần {a}/3).\nVui lòng thử lại.",
                            'warning')
                        event2.set()
                    self.after(0, _warn)
                    event2.wait()

        return None, "Sai mật khẩu 3 lần. Không thể mở file."

    def _load_db_config(self):
        """Đọc db_config.json. Tạo template nếu chưa có."""
        import json
        if not os.path.exists(DB_CONFIG_FILE):
            template = {
                "server":   "",
                "database": "",
                "username": "",
                "password": "",
                "driver":   "ODBC Driver 17 for SQL Server"
            }
            try:
                with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(template, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            return None
        try:
            import json as _json
            with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = _json.load(f)
            if cfg.get("server") and cfg.get("database"):
                return cfg
            return None
        except Exception:
            return None

    # ─ Export validate report ────────────────────────────────────────────────────────────────────────────
    def _export_validate_report(self):
        if not self.val_errors:
            self._show_msg("Chưa validate", "Hãy chạy Validate trước.", 'warning')
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.basename(self._current_file)
        fname_clean = re.sub(r'[\\/*?:\[\]]', '_', fname.rsplit(".", 1)[0])
        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="BOM_Validate_" + fname_clean + "_" + ts + ".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not out_path:
            return

        wb  = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Tong hop"

        def hdr_cell(ws, r, c, val, bg="1E3A5F"):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return cell

        for j, h in enumerate(["Bảng / Sheet", "Tổng lỗi", "Tổng cảnh báo",
                                "Trạng thái", "Chi tiết đầu tiên"], 1):
            hdr_cell(ws0, 1, j, h)
        ws0.column_dimensions["A"].width = 38
        ws0.column_dimensions["B"].width = 12
        ws0.column_dimensions["C"].width = 16
        ws0.column_dimensions["D"].width = 18
        ws0.column_dimensions["E"].width = 60

        for i, (label, errs) in enumerate(self.val_errors.items(), 2):
            n_e  = len([e for e in errs if e["severity"] == "error"])
            n_w  = len([e for e in errs if e["severity"] == "warning"])
            first = errs[0]["message"] if errs else "—"
            status = "OK" if n_e == 0 else (str(n_e) + " loi")
            ws0.cell(row=i, column=1, value=label)
            c2 = ws0.cell(row=i, column=2, value=n_e)
            c2.font = Font(color="C62828" if n_e else "000000", bold=bool(n_e))
            c3 = ws0.cell(row=i, column=3, value=n_w)
            c3.font = Font(color="E65100" if n_w else "000000")
            ws0.cell(row=i, column=4, value=status)
            ws0.cell(row=i, column=5, value=first)
            if i % 2 == 0:
                for c in range(1, 6):
                    ws0.cell(row=i, column=c).fill = PatternFill("solid", fgColor="F8FAFC")

        for label, errs in self.val_errors.items():
            if not errs:
                continue
            safe = re.sub(r'[\\/*?:\[\]]', '_', label)[:31]
            ws2  = wb.create_sheet(safe)
            for j, h in enumerate(["Dòng", "Cột / Trường", "Thông báo lỗi", "Mức độ"], 1):
                hdr_cell(ws2, 1, j, h)
            ws2.column_dimensions["A"].width = 10
            ws2.column_dimensions["B"].width = 28
            ws2.column_dimensions["C"].width = 70
            ws2.column_dimensions["D"].width = 14
            for i, err in enumerate(errs, 2):
                ws2.cell(row=i, column=1, value=str(err.get("row", "")))
                ws2.cell(row=i, column=2, value=str(err.get("col", "")))
                ws2.cell(row=i, column=3, value=err.get("message", ""))
                sc = ws2.cell(row=i, column=4, value=err.get("severity", ""))
                if err.get("severity") == "error":
                    sc.font = Font(color="C62828", bold=True)
                elif err.get("severity") == "warning":
                    sc.font = Font(color="E65100")

        wb.save(out_path)
        self._log(fname, "Report", "—", "OK - Da xuat",
                  os.path.basename(out_path), "ok")
        self._show_export_success("Xuất báo cáo", "Đã lưu:\n" + out_path, out_path)

    # ─ Batch nhieu file ───────────────────────────────────────────────────────────────────────────────────
    def _open_multiple_files(self):
        paths = filedialog.askopenfilenames(
            title="Chọn nhiều file BOM Excel",
            filetypes=[("Excel BOM", "*.xlsm *.xlsx"), ("Tất cả", "*.*")])
        if not paths:
            return
        self.btn_open.config(state=tk.DISABLED)
        self.btn_batch.config(state=tk.DISABLED)
        self._set_status("Đang xử lý " + str(len(paths)) + " file...", C["yellow"])

        def worker():
            mk = build_meta_keys_from_mapping(self.mapping)
            results = []
            for path in paths:
                try:
                    decrypted, err = self._open_excel_with_password(path)
                    if err == "cancelled":
                        results.append({
                            "path": path,
                            "error": "Bỏ qua — file có mật khẩu và người dùng đã hủy",
                            "n_err": 1, "n_wrn": 0
                        })
                        continue
                    if err:
                        results.append({
                            "path": path, "error": err,
                            "n_err": 1, "n_wrn": 0
                        })
                        continue

                    tables, meta, skipped, warns = parse_bom_file(
                        path, meta_keys=mk, _decrypted_bytes=decrypted)
                    val_errors = validate_layer1(tables, meta, self.mapping)
                    n_err, n_wrn = count_errors(val_errors)
                    results.append({
                        "path": path, "tables": tables, "meta": meta,
                        "val_errors": val_errors,
                        "n_err": n_err, "n_wrn": n_wrn,
                        "skipped": skipped
                    })
                except Exception as ex:
                    results.append({
                        "path": path, "error": str(ex),
                        "n_err": 1, "n_wrn": 0
                    })
            self.after(0, lambda: self._on_batch_done(results))

        threading.Thread(target=worker, daemon=True).start()

    def _on_batch_done(self, results):
        n_ok = sum(1 for r in results if r.get("n_err", 1) == 0)
        for r in results:
            fname = os.path.basename(r["path"])
            if "error" in r:
                self._log(fname, "Batch", "—", "Lỗi đọc file", r["error"], "err")
            else:
                total = sum(len(t["df"]) for t in r["tables"].values())
                tag   = "ok" if r["n_err"] == 0 else "err"
                detail = str(r["n_err"]) + " loi, " + str(r["n_wrn"]) + " canh bao"
                status = "OK" if r["n_err"] == 0 else (str(r["n_err"]) + " loi")
                self._log(fname, "Batch", total, status, detail, tag)

        last_ok = next((r for r in reversed(results) if "tables" in r), None)
        if last_ok:
            self.tables        = last_ok["tables"]
            self.global_meta   = last_ok["meta"]
            self.val_errors    = last_ok["val_errors"]
            self._current_file = last_ok["path"]
            self._on_parsed(last_ok.get("skipped", []), [])
            n_err, n_wrn = count_errors(last_ok["val_errors"])
            self.listbox.delete(0, tk.END)
            for name in self.tables:
                errs = last_ok["val_errors"].get(name, [])
                ne   = len([e for e in errs if e["severity"] == "error"])
                nw   = len([e for e in errs if e["severity"] == "warning"])
                icon = (" E" + str(ne)) if ne else ((" W" + str(nw)) if nw else " OK")
                self.listbox.insert(tk.END, name + icon)
            self.btn_report.config(state=tk.NORMAL)
            self.btn_import.config(state=tk.NORMAL if n_err == 0 else tk.DISABLED)

        self._toggle_log() if not self._log_expanded else None
        self.btn_open.config(state=tk.NORMAL)
        self.btn_batch.config(state=tk.NORMAL)
        status_txt = ("OK: " + str(n_ok) + "/" + str(len(results)) + " file")
        self._set_status(status_txt, C["green"] if n_ok == len(results) else C["yellow"])
        self._show_msg("Batch hoàn tất",
            f"Đã xử lý {len(results)} file\n"
            f"OK: {n_ok}  |  Lỗi: {len(results)-n_ok}\n\n"
            "Xem chi tiết ở tab Lịch sử Log", 'info')

    # ─ Selection ───────────────────────────────────────────────────────────────────────────────────
    def _on_select(self, _event):
        sel = self.listbox.curselection()
        if not sel:
            return

        raw        = self.listbox.get(sel[0])
        # Exact match (trước validate, không có icon) hoặc key + ' ' + icon (sau validate).
        # KHÔNG dùng raw.startswith(k) thô vì "BOM Phần III".startswith("BOM Phần II") = True.
        table_name = next(
            (k for k in self.tables if raw == k or raw.startswith(k + ' ')),
            None,
        )
        if not table_name:
            return

        tbl  = self.tables[table_name]
        df   = tbl["df"]
        errs = self.val_errors.get(table_name, [])

        n_rows = len(df)
        n_cols = len(df.columns)
        self.lbl_table.config(
            text="Bảng: " + table_name + "   —   " + str(n_rows) + " dòng × " + str(n_cols) + " cột",
            fg=self._text_fg)

        self.tree.delete(*self.tree.get_children())
        if df.empty or "Lỗi" in df.columns:
            self._update_warn_panel(errs, {}, [])
            self.tree["columns"] = ("msg",)
            self.tree.heading("msg", text="Thông báo")
            self.tree.column("msg", width=600)
            if not df.empty:
                self.tree.insert("", tk.END, values=(df.iloc[0]["Lỗi"],))
            return

        cols = list(df.columns)
        self.tree["columns"] = cols
        self.tree["displaycolumns"] = "#all"   # reset trước khi filter
        import tkinter.font as tkfont
        _hfont = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        _dfont = tkfont.Font(family="Segoe UI", size=12)
        _PADDING = 28
        _MIN_W   = 60
        _MAX_W   = 300  # giới hạn tối đa tránh cột quá rộng

        # Tính width từ heading + sample 60 dòng đầu dữ liệu
        col_w = {}
        for col in cols:
            display = col.replace("_", " ")
            _ca = guess_col_align(col)
            self.tree.heading(col, text=display, anchor=_ca)
            h_w = _hfont.measure(display) + _PADDING
            try:
                sample = df[col].dropna().astype(str).head(60)
                d_w = int(sample.map(lambda s: _dfont.measure(s)).max()) + _PADDING if len(sample) else 0
            except Exception:
                d_w = 0
            col_w[col] = min(max(h_w, d_w, _MIN_W), _MAX_W)
            self.tree.column(col, width=col_w[col], anchor=_ca, stretch=False)

        # ─ Hàng SQL column names (hàng đầu cố định) ──────────────────────────
        section = tbl.get("type", "")
        sql_to_excel = {}
        sql_row = []
        if section and self.mapping:
            for _rec in self.mapping.get(section, []):
                if _rec.get("sql_col") and _rec.get("ten_excel"):
                    sql_to_excel[_rec["sql_col"]] = _rec["ten_excel"]
            rev_map = build_reverse_map(self.mapping, section)
            for col in cols:
                sql_col = match_col_to_sql(_norm_vn(col), rev_map)
                sql_row.append(sql_col if sql_col else "—")
            # Mở rộng cột nếu tên SQL dài hơn nội dung đã tính
            for col, sv in zip(cols, sql_row):
                s_w = _dfont.measure(str(sv)) + _PADDING
                if s_w > col_w[col]:
                    col_w[col] = min(s_w, _MAX_W)
                    self.tree.column(col, width=col_w[col])
            self.tree.insert("", 0, values=sql_row, tags=("sql_names",))
            # Ẩn cột chưa mapped (sql_row = "—") khỏi tree display
            visible = [c for c, sv in zip(cols, sql_row) if sv != "—"]
            self.tree["displaycolumns"] = visible if visible else cols

        # ─ Data rows ───────────────────────────────────────────────────────────────────────────────────
        if hasattr(self, "_empty_frame"):
            self._empty_frame.place_forget()

        # Build cell-level error maps: {row_0based: set(excel_col_name)}
        error_rows  = set()
        warn_rows   = set()
        error_cells = {}   # row_0based → set of col names có lỗi
        warn_cells  = {}
        for e in errs:
            r = e.get("row")
            if not isinstance(r, int):
                continue
            r0 = r - 1
            sev = e.get("severity", "error")
            # Map SQL col → Excel col name trong df
            sql_col   = e.get("col", "")
            excel_col = sql_to_excel.get(sql_col, sql_col)
            matched   = next((c for c in cols
                              if c == excel_col or _norm_vn(c) == _norm_vn(excel_col)), None)
            if sev == "error":
                error_rows.add(r0)
                if matched:
                    error_cells.setdefault(r0, set()).add(matched)
            else:
                warn_rows.add(r0)
                if matched:
                    warn_cells.setdefault(r0, set()).add(matched)

        # Cập nhật warning panel
        self._update_warn_panel(errs, error_cells, cols)

        # Lưu map iid → tree iid để panel có thể jump
        self._warn_iid_to_tree_iid = {}
        tree_iids = []  # sẽ gán sau khi insert

        for _idx, (_, row) in enumerate(df.iterrows()):
            vals = []
            for col, v in zip(cols, row):
                s = str(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
                if _idx in error_cells and col in error_cells[_idx]:
                    s = "⚠ " + s if s else "⚠ —"
                elif _idx in warn_cells and col in warn_cells[_idx]:
                    s = "△ " + s if s else "△ —"
                vals.append(s)
            if _idx in error_rows:
                row_tag = "err_row"
            elif _idx in warn_rows:
                row_tag = "warn_row"
            else:
                row_tag = "evenrow" if _idx % 2 == 0 else "oddrow"
            iid = self.tree.insert("", tk.END, values=vals, tags=(row_tag,))
            tree_iids.append((_idx, iid))

        # Gán mapping warn_panel_iid → main tree iid (sau khi insert xong)
        # _update_warn_panel đã lưu row_0based trong warn_tree items
        for w_iid in self._warn_err_tree.get_children():
            r0 = self._warn_err_tree.set(w_iid, "rownum")
            # rownum được lưu dạng "Dòng N" — extract số
            try:
                r0_int = int(r0.replace("Dòng", "").strip()) - 1
                matched_iid = next((iid for idx, iid in tree_iids if idx == r0_int), None)
                if matched_iid:
                    self._warn_iid_to_tree_iid[w_iid] = matched_iid
            except (ValueError, StopIteration):
                pass

    # ─ Warning panel helpers ──────────────────────────────────────────────────────────────────────────────

    def _update_warn_panel(self, errs, error_cells, cols):
        """Populate warning panel từ danh sách errs sau khi _show_table xong."""
        self._warn_err_tree.delete(*self._warn_err_tree.get_children())
        self._warn_iid_to_tree_iid = {}

        err_list  = [e for e in errs if e.get("severity") == "error"   and isinstance(e.get("row"), int)]
        warn_list = [e for e in errs if e.get("severity") == "warning" and isinstance(e.get("row"), int)]
        # Thêm các lỗi không có row (header-level)
        hdr_errs  = [e for e in errs if e.get("severity") == "error"   and not isinstance(e.get("row"), int)]

        n_err  = len(err_list) + len(hdr_errs)
        n_warn = len(warn_list)

        if n_err == 0 and n_warn == 0:
            self._warn_outer.grid_remove()
            return

        # Hiển thị panel (grid_remove giữ cấu hình, grid() khôi phục)
        self._warn_outer.grid()

        badge_txt = (str(n_err) + " lỗi") + (" / " + str(n_warn) + " cảnh báo" if n_warn else "")
        self._warn_badge.config(text=badge_txt)

        first_msg = (err_list or hdr_errs or warn_list)[0].get("message", "")
        if len(first_msg) > 80:
            first_msg = first_msg[:77] + "..."
        extra = (n_err + n_warn - 1)
        suffix = ("  +" + str(extra) + " mục khác") if extra > 0 else ""
        self._warn_hdr_lbl.config(text=first_msg + suffix)

        # Điền danh sách lỗi vào warn_tree
        all_items = [(e, "err_item") for e in hdr_errs + err_list] + \
                    [(e, "warn_item") for e in warn_list]
        for e, tag in all_items:
            r = e.get("row")
            row_disp  = ("Dòng " + str(r)) if isinstance(r, int) else "Header"
            field_disp = e.get("col", "")
            msg_disp   = e.get("message", "")
            self._warn_err_tree.insert("", tk.END,
                values=(row_disp, field_disp, msg_disp), tags=(tag,))

        # Số dòng hiển thị: tối đa 5, tối thiểu 2
        n_items = len(all_items)
        self._warn_err_tree.config(height=min(max(n_items, 2), 5))

        # Đảm bảo list hiển thị theo trạng thái toggle
        if self._warn_expanded:
            self._warn_list_outer.pack(fill=tk.X)
        else:
            self._warn_list_outer.pack_forget()

    def _toggle_warn_panel(self):
        """Thu gọn / mở rộng phần danh sách lỗi."""
        self._warn_expanded = not self._warn_expanded
        if self._warn_expanded:
            self._warn_list_outer.pack(fill=tk.X)
            self._warn_toggle.config(text="▾")
        else:
            self._warn_list_outer.pack_forget()
            self._warn_toggle.config(text="▸")

    def _on_warn_item_click(self, event):
        """Click vào 1 dòng trong warning panel → scroll + select dòng đó trong main tree."""
        sel = self._warn_err_tree.selection()
        if not sel:
            return
        w_iid = sel[0]
        tree_iid = self._warn_iid_to_tree_iid.get(w_iid)
        if tree_iid and self.tree.exists(tree_iid):
            self.tree.selection_set(tree_iid)
            self.tree.see(tree_iid)
            self.tree.focus(tree_iid)

    # ─ Actions ────────────────────────────────────────────────────────────────────────────────────────────
    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Chọn file BOM Excel",
            filetypes=[("Excel BOM", "*.xlsm *.xlsx"), ("Tất cả", "*.*")])
        if not path:
            return

        self._set_status("Đang bóc tách dữ liệu...", C["yellow"])
        self.btn_open.config(state=tk.DISABLED)
        self.val_errors = {}
        self.update()

        def worker():
            try:
                # Kiểm tra + giải mã nếu file có password
                decrypted, err = self._open_excel_with_password(path)
                if err == "cancelled":
                    self.after(0, lambda: self._set_status("Đã hủy — file có mật khẩu", C["yellow"]))
                    return
                if err:
                    self.after(0, lambda: self._show_msg("Lỗi mật khẩu", err, 'error'))
                    self.after(0, lambda: self._set_status("Lỗi mật khẩu", C["red"]))
                    return

                mk     = build_meta_keys_from_mapping(self.mapping)
                tables, meta, skipped, warns = parse_bom_file(
                    path, meta_keys=mk, _decrypted_bytes=decrypted)
                self.tables        = tables
                self.global_meta   = meta
                self._current_file = path
                self.after(0, lambda: self._on_parsed(skipped, warns))
            except PermissionError:
                msg = "File đang được mở bởi ứng dụng khác (Excel?).\nĐóng file lại rồi thử lại."
                self.after(0, lambda m=msg: self._show_msg("Không thể đọc file", m, 'error'))
                self.after(0, lambda: self._set_status("Lỗi: file đang bị khóa", C["red"]))
            except Exception as ex:
                msg = str(ex)
                self.after(0, lambda m=msg: self._show_msg("Lỗi đọc file", m, 'error'))
                self.after(0, lambda: self._set_status("Lỗi đọc file", C["red"]))
            finally:
                self.after(0, lambda: self.btn_open.config(state=tk.NORMAL))

        threading.Thread(target=worker, daemon=True).start()

    def _on_parsed(self, skipped, warns):
        fname = os.path.basename(self._current_file)
        if skipped:
            self._skipped_sheets = skipped
            skip_info = f"  |  Bỏ qua: {len(skipped)} sheet"
        else:
            self._skipped_sheets = []
            skip_info = ""
        self._set_status("OK: " + fname + skip_info, C["green"])

        total = sum(len(t["df"]) for t in self.tables.values())
        self._log(fname, "—", total, "Da parse",
                  ("Bo qua " + str(len(skipped)) + " sheet") if skipped else "OK",
                  "ok" if not warns else "warn")

        self.listbox.delete(0, tk.END)
        for name in self.tables:
            self.listbox.insert(tk.END, name)

        self.btn_validate.config(state=tk.NORMAL)
        self.btn_import.config(state=tk.DISABLED)
        self.btn_report.config(state=tk.DISABLED)

        if self.tables:
            self.listbox.selection_set(0)
            self._on_select(None)

    def _run_validate(self):
        if not self.tables:
            return
        self.mapping = load_mapping()   # reload mỗi lần validate
        self.val_errors = validate_layer1(self.tables, self.global_meta, self.mapping)
        n_err, n_wrn    = count_errors(self.val_errors)

        sel = self.listbox.curselection()
        self.listbox.delete(0, tk.END)
        for name in self.tables:
            errs = self.val_errors.get(name, [])
            ne   = len([e for e in errs if e["severity"] == "error"])
            nw   = len([e for e in errs if e["severity"] == "warning"])
            icon = (" E" + str(ne)) if ne else ((" W" + str(nw)) if nw else " OK")
            self.listbox.insert(tk.END, name + icon)

        if sel:
            self.listbox.selection_set(sel[0])
        self._on_select(None)

        fname = os.path.basename(self._current_file)
        self.btn_report.config(state=tk.NORMAL)

        # Bat ky sheet nao co loi deu chan Import
        header_errs = len([
            e for e in self.val_errors.get("[H] Phiếu Header", [])
            if e.get("severity") == "error"
        ])

        if n_err == 0:
            self.btn_import.config(state=tk.NORMAL)
            self._set_status("Validate OK — " + str(n_wrn) + " canh bao", C["green"])
            self._log(fname, "Validate", "—", "OK", str(n_wrn) + " canh bao", "ok")
            self._show_msg("Validate",
                "Dữ liệu hợp lệ!\n" + str(n_wrn) + " cảnh báo (không chặn import)", 'info')
        else:
            self.btn_import.config(state=tk.DISABLED)
            self._set_status(str(n_err) + " lỗi  |  " + str(n_wrn) + " cảnh báo", C["red"])
            self._log(fname, "Validate", "—", str(n_err) + " loi",
                      str(n_err) + " loi, " + str(n_wrn) + " canh bao", "err")
            if header_errs > 0:
                self._show_msg("Validate",
                    "HEADER có " + str(header_errs) + " lỗi — cần sửa trước khi import.\n\n"
                    "Tổng: " + str(n_err) + " lỗi / " + str(n_wrn) + " cảnh báo.\n\n"
                    "Vui lòng mở file Excel gốc, sửa các dòng bị đánh dấu đỏ,\n"
                    "lưu lại và thực hiện lại từ Bước 1.", 'warning')
            else:
                sub_errs = n_err - header_errs
                self._show_msg("Validate",
                    "Header OK — nhưng có " + str(sub_errs) + " lỗi dữ liệu ở BOM sheets.\n\n"
                    "Import bị khóa cho đến khi sửa hết lỗi.\n\n"
                    "Vui lòng mở file Excel gốc, sửa các dòng bị đánh dấu đỏ,\n"
                    "lưu lại và thực hiện lại từ Bước 1.", 'warning')

    # ── DB helpers ───────────────────────────────────────────────────────────
    def _get_db_conn(self, timeout_sec: int = 5):
        """
        Trả về pyodbc connection hoặc raise Exception với thông báo rõ ràng.
        timeout_sec: giây chờ kết nối — mặc định 5s để phát hiện VPN down nhanh.
        """
        import pyodbc
        cfg = self.db_cfg
        if not cfg:
            raise Exception("Chưa có cấu hình DB (db_config.json)")
        base = (
            f"DRIVER={{{cfg['driver']}}};"
            f"SERVER={cfg['server']};"
            f"DATABASE={cfg['database']};"
            "TrustServerCertificate=yes;"
        )
        if cfg.get("trusted_connection", "").lower() == "yes":
            conn_str = base + "Trusted_Connection=yes;"
        else:
            conn_str = base + f"UID={cfg['username']};PWD={cfg['password']};"

        # Ưu tiên timeout từ config; timeout_sec là fallback khi config không có
        timeout = int(cfg.get('timeout', timeout_sec))
        try:
            return pyodbc.connect(conn_str, timeout=timeout)
        except Exception as e:
            err = str(e)
            # Phân biệt lỗi timeout / network để gợi ý VPN
            if any(k in err.lower() for k in (
                'timeout', 'timed out', 'network', 'server', 'tcp',
                'named pipes', 'sql server', '08001', 'hn011',
            )):
                raise ConnectionError(
                    f"⚠️  Không kết nối được DB sau {timeout}s.\n\n"
                    f"Kiểm tra:\n"
                    f"  • VPN đã kết nối chưa?\n"
                    f"  • Server {cfg.get('server')} có ping được không?\n\n"
                    f"Chi tiết: {err}"
                ) from e
            raise

    # ── Generic SP caller ────────────────────────────────────────────────────
    def _call_sp(self, conn, sp_name, params_str, row):
        """
        Goi SP generic tu SP_CONFIG.
        params_str: "@Key=FixedValue|@Key2={FieldName}"
          {FieldName} lay tu row[FieldName], con lai la literal.
        Tra ve gia tri cot dau tien dong dau tien, hoac None.
        """
        parts = [p.strip() for p in params_str.split('|') if p.strip()]
        param_clauses, param_values = [], []
        for part in parts:
            if '=' in part:
                key, val = part.split('=', 1)
                key = key.strip().lstrip('@')
                val = val.strip()
                if val.startswith('{') and val.endswith('}'):
                    val = row.get(val[1:-1])
                elif val == '':
                    continue  # literal rỗng → SP dùng default
                # Chuẩn hóa kiểu: NaN → None, float nguyên → int
                import math as _m
                if isinstance(val, float):
                    if _m.isnan(val):
                        val = None
                    elif val == int(val):
                        val = int(val)
                # None → literal NULL (tránh pyodbc infer sai type)
                if val is None:
                    param_clauses.append(f'@{key}=NULL')
                else:
                    param_clauses.append(f'@{key}=?')
                    param_values.append(val)
        sql = f"EXEC {sp_name} " + ", ".join(param_clauses)
        cur = conn.cursor()
        cur.execute(sql, param_values)
        r = cur.fetchone()
        return r[0] if r else None

    def _sp_lookup(self, conn, params_str, row):
        """
        Kieu SP dac biet 'lookup': query bang master lay 1 gia tri.
        params_str: "Bang=TableName|Where=Col={FieldName}|Lay=ReturnCol"
        """
        cfg = {}
        for part in params_str.split('|'):
            if '=' in part:
                k, v = part.split('=', 1)
                cfg[k.strip()] = v.strip()
        table = cfg.get('Bang', '')
        where = cfg.get('Where', '')
        ret   = cfg.get('Lay', '')
        if not (table and where and ret):
            return None
        where_r = re.sub(r'\{(\w+)\}',
                         lambda m: str(row.get(m.group(1), '') or ''), where)
        if '=' in where_r:
            w_col, w_val = where_r.split('=', 1)
            # Không bracket table có dấu chấm (cross-db: B10_Boho.dbo.B00Branch)
            tbl_ref = table if '.' in table else f'[{table}]'
            sql = (f"SELECT TOP 1 [{ret.strip()}] FROM {tbl_ref}"
                   f" WITH(NOLOCK) WHERE [{w_col.strip()}]=?")
            try:
                cur = conn.cursor()
                cur.execute(sql, w_val.strip())
                r = cur.fetchone()
                if not r:
                    return None
                val = r[0]
                # Normalize numeric: float/str '1.0' → '1' (tránh lỗi string concat trong SP)
                import math as _m
                if isinstance(val, float):
                    val = None if _m.isnan(val) else (int(val) if val == int(val) else val)
                elif isinstance(val, str):
                    try:
                        f = float(val)
                        if f == int(f):
                            val = str(int(f))
                    except (ValueError, TypeError):
                        pass
                return val
            except Exception as e:
                self._log('', '_sp_lookup', 0, 'Warn', str(e), 'warn')
        return None

    # ── Generic validators runner ─────────────────────────────────────────────
    def _run_validators(self, conn, row):
        """
        Doc sheet VALIDATORS, chay tung validator IsActive=CO.
        Tra True = tiep tuc, False = huy.
        """
        for v in self.mapping.get('VALIDATORS', []):
            if v.get('isactive', '') != '1':
                continue
            sql      = v.get('sql', '').strip()
            params_s = v.get('params', '').strip()
            warn_msg = v.get('warningmessage', '').strip()
            name     = v.get('validatorname', '')
            if not sql:
                continue
            params = [row.get(p.strip()) for p in params_s.split(',') if p.strip()]
            try:
                cur = conn.cursor()
                cur.execute(sql, params)
                first = cur.fetchone()
                has_vio = False
                result_display = ''
                if first is not None:
                    fv = first[0]
                    if not (isinstance(fv, (int, float)) and fv == 0) and fv is not None:
                        has_vio = True
                        rest = cur.fetchmany(4)
                        result_display = ', '.join(str(r[0]) for r in [first] + list(rest))
                if has_vio:
                    msg = warn_msg.replace('{result}', result_display)
                    for k, val in row.items():
                        msg = msg.replace('{' + k + '}', str(val or ''))
                    if not self._ask_msg(f"Canh bao: {name}",
                                               msg + "\n\nVan tiep tuc insert?"):
                        return False
            except Exception as e:
                self._log('', f'Validator {name}', 0, 'Warn', str(e), 'warn')
        return True

    # ── Item lookup — 3 tiers ────────────────────────────────────────────────
    def _norm_for_match(self, s):
        """Normalize de so sanh: bo dau, lowercase, chi giu alphanum."""
        import unicodedata as _ud
        s = str(s).lower().strip().replace('đ', 'd')
        nfkd = _ud.normalize('NFKD', s)
        s = ''.join(c for c in nfkd if not _ud.combining(c))
        return re.sub(r'[^a-z0-9]', '', s)

    def _show_suggest_dialog(self, item_code, candidates):
        """
        Hien popup top candidates de user chon.
        candidates: [(score, item_dict), ...]
        Tra ve item_id duoc chon hoac None.
        """
        _dt = THEMES.get(ctk.get_appearance_mode(), THEMES["Dark"])
        dlg = tk.Toplevel(self)
        dlg.title("Không tìm thấy mã chính xác — chọn tương đương")
        dlg.geometry("640x280")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=_dt["bg_main"])

        tk.Label(dlg, text=f'Ma Excel: "{item_code}"',
                 bg=_dt["bg_main"], fg=_dt["text_main"],
                 font=FONT_MD_B).pack(pady=(12, 2))
        tk.Label(dlg, text="Chọn mã tương đương hoặc bỏ qua:",
                 bg=_dt["bg_main"], fg=_dt["text_muted"], font=FONT_MD).pack()

        tree = ttk.Treeview(
            dlg, columns=("score", "code", "name"),
            show="headings", height=5, style="BOM.Treeview"
        )
        tree.heading("score", text="Do khop")
        tree.heading("code",  text="Mã B20Item")
        tree.heading("name",  text="Ten")
        tree.column("score", width=80,  anchor="center")
        tree.column("code",  width=160)
        tree.column("name",  width=360)
        for score, c in candidates:
            tree.insert("", tk.END,
                        values=(f"{score:.0f}%", c['code'], c['name']),
                        iid=str(c['id']))
        tree.pack(fill=tk.BOTH, expand=True, padx=PAD_MD, pady=PAD_SM)

        result = [None]

        def on_choose():
            sel = tree.selection()
            if sel:
                result[0] = int(sel[0])
                dlg.destroy()
            else:
                self._show_msg("Chưa chọn", "Hãy click chọn 1 dòng.", 'warning')

        def on_skip():
            dlg.destroy()

        tree.bind("<Double-Button-1>", lambda e: on_choose())
        tree.bind("<Return>", lambda e: on_choose())

        bf = ctk.CTkFrame(dlg, fg_color=_dt["bg_main"])
        bf.pack(pady=6)
        ctk.CTkButton(bf, text="Chọn", width=100, command=on_choose,
                      fg_color=C["green"], hover_color="#3D8B40",
                      text_color="white").pack(side=tk.LEFT, padx=6)
        ctk.CTkButton(bf, text="Bỏ qua (ItemId=NULL)", width=180, command=on_skip,
                      fg_color=_dt["bg_card"], hover_color=_dt["border"],
                      text_color=_dt["text_muted"]).pack(side=tk.LEFT, padx=6)

        dlg.wait_window()
        return result[0]

    def _show_batch_fuzzy_dialog(self, pending):
        """
        Dialog batch: hiển thị TẤT CẢ fuzzy cases cần xác nhận trong 1 lần.
        pending: list of {
            'val': str, 'candidates': [(score, item_dict), ...],
            'cache_key': tuple, 'section': str, 'field': str, 'row_idx': int|None
        }
        Returns: {(val, cache_key): chosen_id or None}
        """
        # Deduplicate theo (val, cache_key) — giữ context section/row đầu tiên gặp
        seen_keys = {}
        unique = []
        for item in pending:
            k = (item['val'], item.get('cache_key'))
            if k not in seen_keys:
                seen_keys[k] = True
                unique.append(item)

        if not unique:
            return {}

        _cfg = self.mapping.get('_CONFIG', {})
        SECTION_DISPLAY = {sec: (info.get('label') or sec) for sec, info in _cfg.items()}
        SECTION_DISPLAY.setdefault('HEADER', 'HEADER')

        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        n = len(unique)
        dlg.title(f"Xác nhận mapping fuzzy — {n} mã cần xử lý")
        dlg.geometry("860x520")
        dlg.resizable(True, True)
        dlg.transient(self)
        dlg.grab_set()
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        dlg.bind("<Escape>", lambda e: dlg.destroy())

        # ── Header info ───────────────────────────────────────────────────────
        top = ctk.CTkFrame(dlg, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(14, 4))
        ctk.CTkLabel(
            top,
            text=f"Tìm thấy {n} mã không khớp chính xác trong file Excel.",
            font=ctk.CTkFont(*FONT_LABEL),
        ).pack(anchor="w")
        ctk.CTkLabel(
            top,
            text="Chọn mã B20 tương đương cho từng dòng. Bỏ qua → giá trị sẽ là NULL.",
            font=ctk.CTkFont(*FONT_BODY),
            text_color=("gray45", "gray65"),
        ).pack(anchor="w", pady=(2, 0))

        # ── Column headers ────────────────────────────────────────────────────
        hdr_frame = ctk.CTkFrame(dlg, fg_color=("gray88", "gray22"), corner_radius=6)
        hdr_frame.pack(fill="x", padx=PAD_MD, pady=(6, 2))
        for col, (txt, w, anchor) in enumerate([
            ("Section · Trường", 200, "w"),
            ("Mã Excel (không khớp)", 190, "w"),
            ("Chọn mã B20 tương đương", 0, "w"),
        ]):
            ctk.CTkLabel(
                hdr_frame, text=txt,
                font=ctk.CTkFont(*FONT_BODY_B),
                width=w, anchor=anchor,
            ).grid(row=0, column=col, padx=(10, 4), pady=6, sticky="w")
        hdr_frame.grid_columnconfigure(2, weight=1)

        # ── Scrollable rows ───────────────────────────────────────────────────
        sf = ctk.CTkScrollableFrame(dlg, corner_radius=0)
        sf.pack(fill="both", expand=True, padx=PAD_MD, pady=0)
        sf.grid_columnconfigure(2, weight=1)

        choices = {}  # (val, cache_key) → (StringVar, {label: id})

        for idx, item in enumerate(unique):
            k = (item['val'], item.get('cache_key'))
            section_raw = item.get('section', '')
            section_lbl = SECTION_DISPLAY.get(section_raw, section_raw)
            row_idx     = item.get('row_idx')
            field_name  = item.get('field', '')
            row_part    = (f" / dòng {row_idx}" if row_idx else "")
            field_part  = (f" · {field_name}" if field_name else "")
            ctx_text    = f"{section_lbl}{row_part}{field_part}"

            bg = ("gray96", "gray17") if idx % 2 == 0 else ("gray92", "gray20")
            row_f = ctk.CTkFrame(sf, fg_color=bg, corner_radius=4)
            row_f.pack(fill="x", pady=1)
            row_f.grid_columnconfigure(2, weight=1)

            ctk.CTkLabel(
                row_f, text=ctx_text, width=200, anchor="w",
                font=ctk.CTkFont(*FONT_BODY),
                text_color=("gray40", "gray70"),
            ).grid(row=0, column=0, padx=(10, 4), pady=7, sticky="w")

            ctk.CTkLabel(
                row_f, text=item['val'], width=170, anchor="w",
                font=ctk.CTkFont(*FONT_BODY_B),
            ).grid(row=0, column=1, padx=PAD_XS, pady=7, sticky="w")

            # Build dropdown options
            option_labels = []
            id_map = {}
            for score, c in item.get('candidates', []):
                lbl = f"{score:.0f}% — {c['code']} — {c.get('name', '')[:45]}"
                option_labels.append(lbl)
                id_map[lbl] = c['id']
            skip_lbl = "⊘ Bỏ qua (NULL)"
            option_labels.append(skip_lbl)
            id_map[skip_lbl] = None

            var = ctk.StringVar(value=option_labels[0])
            choices[k] = (var, id_map)

            ctk.CTkOptionMenu(
                row_f, values=option_labels, variable=var,
                font=ctk.CTkFont(*FONT_BODY), anchor="w",
                fg_color=("#FFFFFF", "#252526"),
                text_color=("#0F172A", "#E8E8E8"),
                button_color=("#E2E8F0", "#1B2A4A"),
                button_hover_color=("#CBD5E1", "#253A60"),
                dropdown_fg_color=("#FFFFFF", "#252526"),
                dropdown_text_color=("#0F172A", "#E8E8E8"),
                dropdown_hover_color=("#BFDBFE", "#0066CC"),
            ).grid(row=0, column=2, padx=(4, 10), pady=5, sticky="ew")

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=PAD_MD)

        result = [None]

        def on_confirm():
            res = {}
            for k, (var, id_map) in choices.items():
                res[k] = id_map.get(var.get())
            result[0] = res
            dlg.destroy()

        def on_skip_all():
            result[0] = {k: None for k in choices}
            dlg.destroy()

        ctk.CTkButton(
            btn_frame, text="✓ Xác nhận và tiếp tục import",
            command=on_confirm, width=220,
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            btn_frame, text="Bỏ qua tất cả (NULL)",
            command=on_skip_all, width=180,
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            border_color=("gray50", "gray55"),
            hover_color=("gray90", "gray25"),
        ).pack(side="right")

        # Center dialog on parent window
        dlg.update_idletasks()
        pw, ph = self.winfo_width(), self.winfo_height()
        px, py = self.winfo_rootx(), self.winfo_rooty()
        mx = px + (pw - 860) // 2
        my = py + (ph - 520) // 2
        dlg.geometry(f"860x520+{mx}+{my}")
        dlg.deiconify()
        dlg.wait_window()

        return result[0] if result[0] is not None else {}

    # ── Generic lookup engine ─────────────────────────────────────────────────

    def _build_cache_generic(self, conn, bang_master, dieu_kien, truong_ss, truong_lv):
        """
        Build cache chung cho bất kỳ bảng master nào.
        truong_ss hỗ trợ nhiều trường phân cách bằng dấu phẩy: "Code,Name"
        Returns: list of dict {ss_list, lv, ht}
          ss_list = [val_field1, val_field2, ...] — thứ tự ưu tiên lookup
          lv      = giá trị lấy về (Truong_Lay_Ve)
          ht      = giá trị hiển thị (field đầu tiên)
        """
        ss_fields = [f.strip() for f in re.split(r'[|,]', truong_ss) if f.strip()]
        where     = f"WHERE {dieu_kien}" if dieu_kien else ""
        cur       = conn.cursor()
        col_sel   = ', '.join(f'[{f}]' for f in ss_fields)
        try:
            sql = f"SELECT {col_sel}, [{truong_lv}] FROM [{bang_master}] WITH(NOLOCK) {where}"
            cur.execute(sql)
            rows = cur.fetchall()
            n = len(ss_fields)
            return [
                {
                    'ss_list': [r[i] for i in range(n)],
                    'lv'     : r[n],
                    'ht'     : r[0],   # field đầu tiên dùng làm display
                }
                for r in rows
            ]
        except Exception:
            # Fallback: chỉ lấy field đầu
            sql = f"SELECT [{ss_fields[0]}], [{truong_lv}] FROM [{bang_master}] WITH(NOLOCK) {where}"
            cur.execute(sql)
            return [{'ss_list': [r[0]], 'lv': r[1], 'ht': r[0]} for r in cur.fetchall()]

    def _lookup_generic(self, value, cache, kieu_lookup, nguong_fuzzy=92, _cache_key=None, _no_popup=False):
        """
        Engine lookup dùng chung.
        cache: list of dict {ss_list, lv, ht}  — từ _build_cache_generic
          ss_list: list các giá trị so sánh theo thứ tự ưu tiên (hỗ trợ multi-field "Code,Name")
        kieu_lookup: fuzzy_code | fuzzy_name | exact | exact_code | exact_name | validate_only
        Returns: (result_value, match_type_str)
        """
        if not cache:
            return None, 'none'

        # ── Compound AND key: value là tuple (val1, val2, ...) ──────────────────
        if isinstance(value, tuple):
            def _cv(v):
                """Float nguyên (1.0) → '1', tránh '1.0' != '1'."""
                if v is None: return ''
                if isinstance(v, float):
                    return str(int(v)) if v == int(v) else str(v).strip()
                return str(v).strip()
            val_parts  = [_cv(v) for v in value]
            norm_parts = [self._norm_for_match(v) for v in val_parts]
            n = len(val_parts)
            def _and_exact(e):
                ss = [_cv(v) for v in e.get('ss_list', [])]
                return len(ss) >= n and all(ss[i] == val_parts[i] for i in range(n))
            def _and_norm(e):
                ss = [_cv(v) for v in e.get('ss_list', [])]
                return len(ss) >= n and all(self._norm_for_match(ss[i]) == norm_parts[i] for i in range(n))
            for e in cache:
                if _and_exact(e): return e['lv'], 'exact'
            for e in cache:
                if _and_norm(e):  return e['lv'], 'normalize'
            return None, 'none'   # compound key không dùng fuzzy
        # ─────────────────────────────────────────────────────────────────────────

        if not value:
            return None, 'none'

        val_str  = str(value).strip()
        norm_val = self._norm_for_match(val_str)

        def _ss_vals(e):
            """Trả về list các giá trị so sánh của entry (tương thích cả ss_list lẫn ss cũ)."""
            if 'ss_list' in e:
                return [str(v or '').strip() for v in e['ss_list']]
            return [str(e.get('ss') or '').strip()]

        # ── Tier 1: exact — so sánh từng field theo thứ tự ──────────────────
        for e in cache:
            for sv in _ss_vals(e):
                if sv == val_str:
                    return e['lv'], 'exact'

        # ── Tier 2: normalize — tương tự, theo thứ tự ───────────────────────
        for e in cache:
            for sv in _ss_vals(e):
                if self._norm_for_match(sv) == norm_val:
                    return e['lv'], 'normalize'

        # validate_only: check tồn tại → giữ value gốc nếu không tìm thấy
        if kieu_lookup == 'validate_only':
            return val_str, 'not_found'

        # exact (không fuzzy): dừng ở đây
        if kieu_lookup in ('exact', 'exact_code', 'exact_name'):
            return None, 'none'

        # ── Tier 3: fuzzy — score max trên tất cả ss fields ─────────────────
        try:
            from rapidfuzz import fuzz as _fuzz
        except ImportError:
            return None, 'none'

        def _score_entry(e):
            """Tính fuzzy score tốt nhất trên toàn bộ ss_list."""
            best = 0.0
            for sv in _ss_vals(e):
                cv_norm = self._norm_for_match(sv)
                if not cv_norm:
                    continue
                _lp = (min(len(norm_val), len(cv_norm))
                       / max(len(norm_val), len(cv_norm), 1))
                if kieu_lookup == 'fuzzy_name':
                    # sqrt(_lp) thay vì _lp: giảm deflation khi hai chuỗi khác độ dài nhiều
                    s = max(_fuzz.partial_ratio(norm_val, cv_norm) * (_lp ** 0.5),
                            _fuzz.ratio(norm_val, cv_norm))
                    # contains → floor 75: vào popup nhưng không bao giờ auto-map
                    if cv_norm and (norm_val in cv_norm or cv_norm in norm_val):
                        s = max(s, 75.0)
                else:   # fuzzy_code (default)
                    s = max(_fuzz.ratio(norm_val, cv_norm),
                            _fuzz.partial_ratio(norm_val, cv_norm) * _lp)
                if s > best:
                    best = s
            return best

        _min_score = 20 if kieu_lookup == 'fuzzy_name' else 40
        scored = [(s, e) for e in cache
                  if (s := _score_entry(e)) >= _min_score]

        if not scored:
            return None, 'none'

        scored.sort(key=lambda x: -x[0])

        # Popup cho user chọn (top 3) — hiển thị field đầu tiên làm "code"
        candidates = [
            (s, {
                'id'  : e['lv'],
                'code': _ss_vals(e)[0],
                'name': str(e.get('ht') or _ss_vals(e)[0] or ''),
            })
            for s, e in scored[:3]
        ]

        # ── Batch mode: dùng kết quả đã resolve sẵn hoặc collect ────────────
        res_key = (val_str, _cache_key)
        if hasattr(self, '_fuzzy_resolutions') and res_key in self._fuzzy_resolutions:
            chosen = self._fuzzy_resolutions[res_key]
            return chosen, ('fuzzy_user' if chosen else 'none')
        if getattr(self, '_fuzzy_collect_mode', False):
            entry = dict(getattr(self, '_fuzzy_ctx', {}))
            entry.update({'val': val_str, 'candidates': candidates, 'cache_key': _cache_key})
            if not hasattr(self, '_fuzzy_pending'):
                self._fuzzy_pending = []
            self._fuzzy_pending.append(entry)
            return None, 'fuzzy_pending'
        # Nếu user đã xác nhận batch dialog → suppress popup cũ, trả về NULL
        if getattr(self, '_fuzzy_batch_done', False):
            return None, 'none'
        # Caller yêu cầu không bật popup (ví dụ: validate worker) → trả về không tìm thấy
        if _no_popup:
            return None, 'none'
        # ─────────────────────────────────────────────────────────────────────

        chosen = self._show_suggest_dialog(val_str, candidates)
        return chosen, ('fuzzy_user' if chosen else 'none')

    # ── Mapping-driven cache builder ──────────────────────────────────────────
    def _build_all_caches(self, conn, mapping_records):
        """
        Scan mapping → pre-build cache cho mọi trường có lookup.
        Cache key = (bang_master, dieu_kien, truong_so_sanh, truong_lay_ve)
        """
        caches = {}
        seen   = set()
        skip_kl = {'', 'sp_rowid', 'sp_version', 'sp_code', 'lookup'}
        for rec in mapping_records:
            kl = rec.get('kieu_lookup', '')
            bm = rec.get('bang_master', '')
            dk = rec.get('dieu_kien_master', '')
            ss = rec.get('truong_so_sanh', '')
            lv = rec.get('truong_lay_ve',  '')
            if kl in skip_kl or not bm or not ss or not lv:
                continue
            cache_key = (bm, dk, ss, lv)
            if cache_key in seen:
                continue
            seen.add(cache_key)
            try:
                caches[cache_key] = self._build_cache_generic(conn, bm, dk, ss, lv)
                self._log('cache', f'Build cache {bm}', len(caches[cache_key]),
                          'OK', f'{ss}->{lv} WHERE {dk}', 'ok')
            except Exception as e:
                self._log('cache', f'Build cache {bm}', 0, 'Warn', str(e), 'warn')
                caches[cache_key] = []
        return caches

    def _resolve_header_field(self, rec, conn, meta, norm_meta, caches, now, row_out=None):
        """
        Resolve 1 trường từ mapping record (dùng cho HEADER).
        Trả về (value, match_type_str)
        """
        nguon    = rec['nguon_dl']
        mac_dinh = rec['mac_dinh']
        kl       = rec.get('kieu_lookup', '')
        bm       = rec.get('bang_master', '')
        dk       = rec.get('dieu_kien_master', '')
        ten_excel = rec.get('ten_excel', '')

        # ── CoDinh: đọc Mac_dinh, tự-cast kiểu ─────────────────────────────
        if nguon == 'CoDinh':
            if mac_dinh.upper() == 'NULL':
                return None, 'codinh'
            kieu_dl_cd = rec.get('kieu_dl', '').lower()
            if mac_dinh in ('', 'EMPTY'):
                # date/datetime không thể là empty string → NULL
                if kieu_dl_cd in ('date', 'datetime'):
                    return None, 'codinh'
                return '', 'codinh'
            try:    return int(mac_dinh),   'codinh'
            except (ValueError, TypeError): pass
            try:    return float(mac_dinh), 'codinh'
            except (ValueError, TypeError): pass
            return mac_dinh, 'codinh'

        # ── HeThong: system-generated ────────────────────────────────────────
        elif nguon == 'HeThong':
            if mac_dinh == 'NOW':
                return now, 'hethong'
            return None, 'hethong'   # ModifiedBy → Bravo fill khi import thật

        # ── UILookup: giá trị lấy từ combo trên giao diện ───────────────────
        elif nguon == 'UILookup':
            if mac_dinh == 'creator':
                return self._current_creator_user_id, 'ui_lookup'
            elif mac_dinh == 'product_id':
                return self._thdm_selected_product_id, 'ui_lookup'
            elif mac_dinh == 'order_id':
                return self._thdm_selected_order_id, 'ui_lookup'
            elif mac_dinh == 'period_id':
                return self._thdm_selected_period_id, 'ui_lookup'
            return None, 'ui_lookup'

        # ── SP / TinhToan: resolve trong pass 2 ──────────────────────────
        elif nguon in ('SP', 'TinhToan'):
            return None, nguon.lower()

        # ── Excel: đọc meta, lookup master nếu cần ───────────────────────────
        else:
            raw = None
            if ten_excel:
                if '|' in ten_excel:
                    # Multi-field: mỗi phần tách bởi | → build tuple để AND lookup
                    # @FieldName → lấy từ row_out (field đã resolve trước đó trong cùng header)
                    _parts, _any = [], False
                    for _te in ten_excel.split('|'):
                        _te = _te.strip()
                        if _te.startswith('@'):
                            _fn = _te[1:]
                            _v = (row_out or {}).get(_fn)
                        else:
                            _v = meta.get(_te)
                            if _v is None:
                                _v = norm_meta.get(_norm_vn(_te))
                        _parts.append(str(_v or '').strip())
                        if _v is not None:
                            _any = True
                    raw = tuple(_parts) if _any else None
                else:
                    # Single field: giữ nguyên logic cũ
                    # Dùng explicit None check thay vì "or" để không bỏ sót giá trị 0
                    _v = meta.get(ten_excel)
                    if _v is None:
                        _v = norm_meta.get(_norm_vn(ten_excel))
                    raw = _v
            if raw is None:
                if mac_dinh == 'EMPTY':
                    raw = ''
                elif mac_dinh not in ('', 'NULL'):
                    raw = mac_dinh

            kieu_dl = rec.get('kieu_dl', '').lower()

            if not kl:
                # ── Kieu_DL transform chỉ dùng cho direct value (không lookup) ──
                if raw is not None:
                    if kieu_dl in ('date', 'datetime'):
                        if isinstance(raw, datetime.datetime):
                            raw = raw.date()
                        elif isinstance(raw, datetime.date):
                            pass   # already date
                        elif isinstance(raw, str):
                            s = raw.strip()
                            if not s or s.lower() in ('null', 'none', 'nan'):
                                # Empty/NULL string cho date field → NULL
                                raw = None
                            else:
                                parsed = None
                                for fmt in _DATE_FMTS:
                                    try:
                                        parsed = datetime.datetime.strptime(s, fmt).date()
                                        break
                                    except (ValueError, TypeError):
                                        pass
                                # Nếu parse thất bại → NULL thay vì để lại string
                                # (tránh SQL Server nhận '' → 1900-01-01)
                                raw = parsed
                        elif isinstance(raw, (int, float)):
                            import math as _math2
                            if _math2.isnan(raw) if isinstance(raw, float) else False:
                                raw = None
                    elif kieu_dl in ('number', 'decimal', 'float', 'numeric'):
                        qty_str = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        try:    raw = float(qty_str) if qty_str else None
                        except (ValueError, TypeError): raw = None
                    elif kieu_dl == 'int':
                        try:
                            qty_str = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                            raw = int(float(qty_str)) if qty_str else None
                        except (ValueError, TypeError): raw = None
                return raw, 'excel_direct'

            # ── Generic lookup engine ─────────────────────────────────────
            ss = rec.get('truong_so_sanh', '')
            lv = rec.get('truong_lay_ve',  '')
            if ss and lv:
                cache_key = (bm, dk, ss, lv)
                cache = caches.get(cache_key, [])
                nguong = int(rec.get('nguong_fuzzy', 0) or 0) or 92
                self._fuzzy_ctx = {
                    'section': 'HEADER',
                    'field': rec.get('sql_col', ''),
                    'row_idx': None,
                }
                return self._lookup_generic(raw, cache, kl, nguong, _cache_key=cache_key)

            return raw, 'excel_direct'

    # ── BOM Detail helpers (BOM2 / BOM3 / BOM4) ──────────────────────────────

    def _build_bom_detail_caches(self, conn, mapping_recs):
        """Build lookup caches cho một section BOM detail (tương tự _build_all_caches)."""
        caches = {}
        for rec in mapping_recs:
            bm  = _nan_str(rec.get('bang_master', ''))
            dk  = _nan_str(rec.get('dieu_kien_master', ''))
            ss  = _nan_str(rec.get('truong_so_sanh', ''))
            lv  = _nan_str(rec.get('truong_lay_ve', ''))
            kl  = _nan_str(rec.get('kieu_lookup', ''))
            if not (bm and ss and lv and kl):
                continue
            key = (bm, dk, ss, lv)
            if key not in caches:
                try:
                    caches[key] = self._build_cache_generic(conn, bm, dk, ss, lv)
                except Exception as e:
                    self._log('', f'detail_cache({bm})', 0, 'Warn', str(e), 'warn')
        return caches

    def _resolve_detail_row(self, mapping_recs, df_row, parent_row, conn,
                            detail_caches, now, builtin_order, bom_detail_type,
                            sp_cfgs=None, bom_section=''):
        """
        Resolve một hàng DataFrame (BOM2/3/4) → dict {sql_col: value}.
        - mapping_recs : list record từ mapping sheet BOM2/3/4
        - df_row       : pandas Series (1 hàng dữ liệu)
        - parent_row   : dict B20BOM header đã resolve (để copy BranchCode, EffectiveDate, ...)
        - builtin_order: số thứ tự hàng (1, 2, 3...)
        - bom_detail_type: 2 / 3 / 4
        """
        import datetime as _dt
        import math as _math

        # Các field copy từ B20BOM parent
        PARENT_FIELDS = {
            'BranchCode', 'EffectiveDate', 'FinishedDate',
            'ItemId0', 'DetailRowId_SO', 'ParentBizDocId', 'ProductionProcessId',
        }

        # ── Pass 1a: CoDinh / HeThong / UILookup / SP-defer ─────────────
        # Dùng shared _resolve_row_mapping() — đồng nhất với THDM pipeline.
        # BOMDetailType inject vào parent_row để HeThong tự pick đúng.
        _bom_parent = dict(parent_row)
        _bom_parent['BOMDetailType'] = bom_detail_type
        _ctx = {
            'now':           now,
            'builtin_order': builtin_order,
            'doc_id':        '__BOMID__',   # placeholder thay thật khi generate SQL
            'parent_row':    _bom_parent,
            'parent_fields': PARENT_FIELDS | {'BOMDetailType'},
            'ui_values': {
                'creator':    self._current_creator_user_id,
            },
            'skip_nguon': {'SP', 'TinhToan'},
        }
        _non_excel = [r for r in mapping_recs
                      if (r.get('nguon_dl') or 'Excel') != 'Excel']
        row_out = _resolve_row_mapping(_non_excel, _ctx, get_excel_val=None)

        # ── Pass 1b: Excel fields (df_row lookup + master lookup) ─────────
        for rec in mapping_recs:
            sql_col = rec.get('sql_col', '')
            nguon   = rec.get('nguon_dl', 'Excel')
            mac     = _nan_str(rec.get('mac_dinh', ''))
            ten     = _nan_str(rec.get('ten_excel', ''))
            kl      = _nan_str(rec.get('kieu_lookup', ''))
            kd      = _nan_str(rec.get('kieu_dl', ''))
            bm      = _nan_str(rec.get('bang_master', ''))
            dk      = _nan_str(rec.get('dieu_kien_master', ''))
            ss      = _nan_str(rec.get('truong_so_sanh', ''))
            lv      = _nan_str(rec.get('truong_lay_ve', ''))

            if nguon != 'Excel':
                continue   # đã xử lý bởi _resolve_row_mapping() ở trên

            # ── Excel (df_row lookup + master lookup) ─────────────────────
            raw = None
            if ten:
                if '|' in ten:
                    # Multi-field compound key: "Col1|Col2" hoặc "Col1|@ParentField"
                    # @FieldName → lấy từ parent_row (header đã resolve)
                    _parts, _any = [], False
                    for _t in ten.split('|'):
                        _t = _t.strip()
                        if _t.startswith('@'):
                            _pv = parent_row.get(_t[1:])
                            _parts.append(str(_pv or '').strip())
                            if _pv is not None: _any = True
                        else:
                            _pv, _nt = None, _norm_vn(_t)
                            for _c in df_row.index:
                                _cs = str(_c).strip()
                                if _cs == _t or _norm_vn(_cs) == _nt:
                                    _v2 = df_row[_c]
                                    if _v2 is not None and not (isinstance(_v2, float) and _math.isnan(_v2)):
                                        _pv = _v2
                                    break
                            _parts.append(str(_pv or '').strip())
                            if _pv is not None: _any = True
                    raw = tuple(_parts) if _any else None
                else:
                    _matched_col = None
                    _norm_ten = _norm_vn(ten)
                    # Pass 1: exact hoặc norm-exact match
                    for col in df_row.index:
                        col_s = str(col).strip()
                        if col_s == ten or _norm_vn(col_s) == _norm_ten:
                            v = df_row[col]
                            _matched_col = col
                            if v is None or (isinstance(v, float) and _math.isnan(v)):
                                raw = None
                            else:
                                raw = v
                            break
                    # Pass 2: suffix match — xử lý cột bị merge header cha
                    # Ví dụ: "SLg_Tên_Vật_Tư" (norm: slgtenvattu) match "Tên vật tư" (norm: tenvattu)
                    if _matched_col is None and _norm_ten:
                        for col in df_row.index:
                            col_s = str(col).strip()
                            norm_col = _norm_vn(col_s)
                            if norm_col.endswith(_norm_ten) and len(norm_col) > len(_norm_ten):
                                v = df_row[col]
                                _matched_col = col
                                if v is None or (isinstance(v, float) and _math.isnan(v)):
                                    raw = None
                                else:
                                    raw = v
                                break

            _raw_empty = (raw is None
                          or (isinstance(raw, float) and _math.isnan(raw))
                          or (isinstance(raw, str) and raw.strip() == ''))
            if _raw_empty:
                raw = None
                _mac_upper = mac.upper()
                if _mac_upper == 'EMPTY':
                    raw = ''
                elif _mac_upper == 'NOW':
                    raw = now
                elif _mac_upper == 'CREATOR':
                    raw = self._current_creator_user_id
                elif mac and mac not in ('', 'NULL'):
                    try:    raw = int(mac)
                    except (ValueError, TypeError):
                        try:    raw = float(mac)
                        except (ValueError, TypeError): raw = mac

            # Type coercion (chỉ khi không lookup)
            if raw is not None and not kl:
                if kd == 'date':
                    if isinstance(raw, _dt.datetime):
                        raw = raw.date()
                    elif isinstance(raw, str):
                        for fmt in _DATE_FMTS:
                            try: raw = _dt.datetime.strptime(raw, fmt).date(); break
                            except (ValueError, TypeError): pass
                elif kd in ('number', 'decimal', 'float', 'numeric'):
                    try:
                        qs = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        raw = float(qs) if qs else None
                    except (ValueError, TypeError): raw = None
                elif kd == 'int':
                    try:
                        qs = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        raw = int(float(qs)) if qs else None
                    except (ValueError, TypeError): raw = None

            # Lookup master nếu có (cùng engine với HEADER)
            if kl and ss and lv:
                cache_key = (bm, dk, ss, lv)
                cache = detail_caches.get(cache_key, [])
                nguong = int(rec.get('nguong_fuzzy', 0) or 0) or 92
                self._fuzzy_ctx = {
                    'section': bom_section,
                    'field': sql_col,
                    'row_idx': builtin_order,
                }
                raw, _ = self._lookup_generic(raw, cache, kl, nguong, _cache_key=cache_key)

            row_out[sql_col] = raw

        # ── Pass 2: SP fields (row_out đã có đủ context để truyền vào params) ──
        if sp_cfgs:
            # Xác định field nào là "secondary output" của multi-output SP
            _claimed = set()
            for rec2 in mapping_recs:
                if rec2.get('nguon_dl') != 'SP':
                    continue
                s_col = rec2.get('sql_col', '')
                cfg2  = sp_cfgs.get(s_col)
                if cfg2:
                    out_fs = [f.strip() for f in cfg2.get('outputfields', '').split(',') if f.strip()]
                    if len(out_fs) > 1:
                        _claimed.update(out_fs[1:])   # primary = out_fs[0], rest = secondary

            _done_multi = set()
            for rec2 in mapping_recs:
                if rec2.get('nguon_dl') != 'SP':
                    continue
                s_col = rec2.get('sql_col', '')
                if s_col in _claimed or s_col in _done_multi:
                    continue
                cfg2 = sp_cfgs.get(s_col)
                if not cfg2:
                    continue
                try:
                    sp_name2  = cfg2.get('sp_name', '').strip()
                    params_s2 = cfg2.get('params', '').strip()
                    fallback2 = cfg2.get('fallback', '').strip() or None
                    out_fs2   = [f.strip() for f in cfg2.get('outputfields', '').split(',') if f.strip()]
                    if sp_name2.lower() == 'lookup':
                        r2 = self._sp_lookup(conn, params_s2, row_out)
                        row_out[s_col] = r2 if r2 is not None else fallback2
                    elif out_fs2:
                        # Multi-output SP: build params → execute → distribute kết quả
                        pc_h, pv_h = [], []
                        for ph in params_s2.split('|'):
                            ph = ph.strip()
                            if '=' not in ph: continue
                            kh, vh = ph.split('=', 1)
                            kh = kh.strip().lstrip('@'); vh = vh.strip()
                            if vh.startswith('{') and vh.endswith('}'):
                                vh = row_out.get(vh[1:-1])
                            if vh is None:
                                pc_h.append(f'@{kh}=NULL')
                            else:
                                pc_h.append(f'@{kh}=?')
                                pv_h.append(vh)
                        cur2 = conn.cursor()
                        cur2.execute(f'EXEC {sp_name2} ' + ', '.join(pc_h), pv_h)
                        sp_row2 = cur2.fetchone()
                        if sp_row2 and cur2.description:
                            sp_res2 = dict(zip([d[0] for d in cur2.description], sp_row2))
                            for f2 in out_fs2:
                                row_out[f2] = sp_res2.get(f2)
                        _done_multi.add(s_col)
                    else:
                        r2 = self._call_sp(conn, sp_name2, params_s2, row_out)
                        row_out[s_col] = r2 if r2 is not None else fallback2
                except Exception as e2:
                    self._log('', f'SP_detail {s_col}', 0, 'Warn', str(e2), 'warn')

        return row_out

    def _generate_bom_details(self, conn, bom_id, parent_row, tables, mapping, now, export_only):
        """
        Tạo INSERT INTO B20BOMDetail cho BOM2, BOM3, BOM4.
        3 pha:
          1. Resolve tất cả rows của mỗi section (per-row BeforeInsert hooks)
          2. BeforeInsertBatch hooks (gọi SP 1 lần / nhóm SP)
          3. INSERT / export SQL tất cả rows đã resolve
        - export_only=True  → trả về list[str] câu SQL
        - export_only=False → execute trực tiếp, trả về số hàng insert
        """
        import math as _math

        # Derive từ _CONFIG + DETAIL mapping (BOMDetailType CoDinh)
        SECTION_TO_TYPE = {}
        for _sec, _cfg in mapping.get('_CONFIG', {}).items():
            if 'BOMDetail' not in _cfg.get('view_insert', ''):
                continue
            for _r in mapping.get(_sec, []):
                if _r.get('sql_col') == 'BOMDetailType':
                    try:
                        SECTION_TO_TYPE[_sec] = int(float(_r.get('mac_dinh', 0)))
                    except (ValueError, TypeError):
                        pass
                    break
        DB_TABLE        = 'B20BOMDetail'
        NVARCHAR_TYPES  = {'nvarchar', 'nchar', 'ntext'}
        sql_lines       = []
        total_rows      = 0


        def _progress(msg, done=None, total=None):
            def _ui(m=msg, d=done, t=total):
                self._update_loading_msg("Đang import dữ liệu vào BRAVO...\n" + m)
                if d is not None and t:
                    self._update_loading_progress(d / t)
            self.after(0, _ui)

        # ── Pha 1: resolve tất cả sections ───────────────────────────────────
        # section_data[section] = {
        #   'resolved'        : [row_vals, ...],
        #   'col_kieu'        : {sql_col: kieu_dl},
        #   'bom_detail_type' : int,
        # }
        section_data = {}

        for label, tbl in tables.items():
            section = tbl.get('type')
            if section not in SECTION_TO_TYPE:
                continue
            df = tbl.get('df')
            if df is None or df.empty or 'Lỗi' in df.columns:
                continue

            bom_detail_type = SECTION_TO_TYPE[section]
            detail_recs     = mapping.get(section, [])
            if not detail_recs:
                continue

            _n_df_rows = len(df) if df is not None else 0
            _progress(f"Đang xử lý: {label} ({_n_df_rows} dòng)...")
            detail_caches = (getattr(self, '_ps_bom_caches', {}).get(section)
                           or self._build_bom_detail_caches(conn, detail_recs))

            sp_cfgs_section = {
                s['sql_column']: s
                for s in mapping.get('SP_CONFIG', [])
                if s.get('isactive', '') == '1'
                and s.get('section', '') in ('', section)
            }

            # Per-row BeforeInsert hooks
            sp_hooks_before = [
                h for h in mapping.get('SP_HOOK', [])
                if h.get('section', '') == section
                and h.get('event', '').lower() == 'beforeinsert'
            ]

            col_kieu = {r['sql_col']: r.get('kieu_dl', '').lower() for r in detail_recs}

            stt_col  = next((c for c in df.columns if _norm_vn(str(c)) == 'stt'), None)
            ff_fields = [
                (r['ten_excel'], r['sql_col'])
                for r in detail_recs
                if r.get('fill_forward') == '1' and r.get('ten_excel')
            ]
            current_ff   = {}
            builtin_order = 0
            resolved_rows = []


            for _, df_row in df.iterrows():
                if df_row.isna().all():
                    continue

                def _get_stt(raw):
                    """str(0 or '') = '' — xử lý integer 0 đúng."""
                    if raw in (0, 0.0): return "0"
                    if raw is None or raw == '': return ''
                    if isinstance(raw, float) and _math.isnan(raw): return ''
                    return str(raw).strip()
                if stt_col:
                    stt_v = _get_stt(df_row.get(stt_col))
                else:
                    # Fallback: tìm cột tên 'STT' trực tiếp trong DataFrame
                    stt_v = ''
                    for _col in df_row.index:
                        if str(_col).upper() == 'STT':
                            stt_v = _get_stt(df_row.get(_col))
                            break
                is_section_header = bool(stt_v and SECTION_STT_PATTERN.match(stt_v))

                # Safety: dòng có STT là chữ cái (vd 'I') nhưng có kích thước vật lý
                # → là sub-group GIAO RỜI (MODULE MD1...), không phải section header thật
                if is_section_header:
                    _hd_gen = any(
                        v not in (None, '', 0, '0', 0.0)
                        and not (isinstance(v, float) and _math.isnan(v))
                        for col, v in df_row.items()
                        if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(col)))
                    )
                    if _hd_gen:
                        is_section_header = False  # có dims → data row (GIAO RỜI/sub-group), không phải section header

                if is_section_header:
                    for (ten_excel, sql_col_ff) in ff_fields:
                        raw = df_row.get(ten_excel)
                        if raw is not None and not (isinstance(raw, float) and _math.isnan(raw)) \
                                and str(raw).strip() not in ('', 'nan'):
                            current_ff[sql_col_ff] = str(raw).strip()
                    continue

                builtin_order += 1
                row_vals = self._resolve_detail_row(
                    detail_recs, df_row, parent_row, conn,
                    detail_caches, now, builtin_order, bom_detail_type,
                    sp_cfgs=sp_cfgs_section, bom_section=section
                )

                # Fill-Forward
                for (ten_excel, sql_col_ff) in ff_fields:
                    if sql_col_ff in current_ff:
                        row_vals[sql_col_ff] = current_ff[sql_col_ff]

                # Per-row SP_HOOK BeforeInsert — dùng shared runner
                _run_row_sp_hooks(
                    conn, sp_hooks_before, row_vals,
                    log_fn=lambda msg: self._log('', msg, 0, 'Warn', msg, 'warn'),
                )

                resolved_rows.append(row_vals)

            section_data[section] = {
                'resolved'       : resolved_rows,
                'col_kieu'       : col_kieu,
                'bom_detail_type': bom_detail_type,
            }

        # ── Pha 2: BeforeInsertBatch hooks (gọi SP theo nhóm SP_Name) ─────────
        if not export_only and section_data:
            batch_hooks = [
                h for h in mapping.get('SP_HOOK', [])
                if h.get('event', '').lower() == 'beforeinsertbatch'
                and h.get('section', '') in section_data
                and h.get('isactive', '') == '1'
            ]
            # Gom theo sp_name để gọi mỗi SP 1 lần duy nhất
            by_sp = {}
            for h in batch_hooks:
                sp = h['sp_name']
                by_sp.setdefault(sp, []).append(h)
            for sp_name, hooks in by_sp.items():
                _n_rows = sum(
                    len(section_data.get(h.get('section',''), {}).get('resolved', []))
                    for h in hooks
                )
                _progress(f"Đang xử lý batch ({_n_rows} dòng)...")
                self._run_batch_hook_grouped(conn, sp_name, hooks, section_data, parent_row)

        # ── Pha 3: INSERT / export SQL ────────────────────────────────────────
        _total_all = sum(len(d['resolved']) for d in section_data.values())
        # U8: chuyển sang determinate mode ngay khi biết tổng số dòng
        self.after(0, lambda t=_total_all: self._switch_loading_to_determinate(t))
        _progress(f"Đang ghi {_total_all:,} dòng vào BRAVO...", done=0, total=_total_all)
        for section, data in section_data.items():
            resolved_rows   = data['resolved']
            col_kieu        = data['col_kieu']
            bom_detail_type = data['bom_detail_type']

            if export_only:
                sql_lines.append(f'\n-- {section} (BOMDetailType={bom_detail_type})\n')

            def _sv(v, col=None):
                """Render giá trị → SQL literal (dùng trong export)."""
                import datetime as _dt2
                if v is None:
                    return 'NULL'
                if v == '__BOMID__':
                    return str(bom_id) if isinstance(bom_id, int) else bom_id
                if isinstance(v, bool):   return '1' if v else '0'
                _col_type   = col_kieu.get(col, '').lower() if col else ''
                _is_str_col = _col_type in ('varchar','nvarchar','char','nchar','text','ntext')
                if isinstance(v, float):
                    if _math.isnan(v):   return 'NULL'
                    if _is_str_col:
                        s  = str(int(v)) if v == int(v) else str(v)
                        px = 'N' if _col_type in NVARCHAR_TYPES else ''
                        return px + "'" + s + "'"
                    return str(v)
                if isinstance(v, int):
                    if _is_str_col:
                        px = 'N' if _col_type in NVARCHAR_TYPES else ''
                        return px + "'" + str(v) + "'"
                    return str(v)
                if isinstance(v, _dt2.datetime):
                    return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'"
                if isinstance(v, _dt2.date):
                    return "'" + v.strftime('%Y-%m-%d') + "'"
                s = str(v).strip()
                if s.lower() in ('nan', 'none', ''):  return 'NULL'
                px = 'N' if col_kieu.get(col, '') in NVARCHAR_TYPES else ''
                return px + "'" + s.replace("'", "''") + "'"

            for row_vals in resolved_rows:
                cols = list(row_vals.keys())
                if export_only:
                    col_list = ', '.join(f'[{c}]' for c in cols)
                    val_list = ', '.join(_sv(row_vals[c], c) for c in cols)
                    sql_lines.append(
                        f'INSERT INTO {DB_TABLE} ({col_list}) VALUES ({val_list});\n'
                    )
                else:
                    exec_vals = [
                        bom_id if row_vals[c] == '__BOMID__' else row_vals[c]
                        for c in cols
                    ]
                    sql_exec = (
                        f"INSERT INTO {DB_TABLE} ("
                        + ', '.join(f'[{c}]' for c in cols)
                        + ') VALUES (' + ', '.join(['?'] * len(cols)) + ')'
                    )
                    cur = conn.cursor()
                    cur.execute(sql_exec, exec_vals)
                total_rows += 1
                if not export_only and total_rows % 100 == 0:
                    _progress(f"Đang ghi: {total_rows:,}/{_total_all:,} dòng...",
                              done=total_rows, total=_total_all)

        _progress(f"Hoàn tất: {total_rows:,} dòng chi tiết")
        return sql_lines if export_only else total_rows

    def _run_batch_hook_grouped(self, conn, sp_name, hooks, section_data, parent_row):
        """
        Gọi SP 1 lần với nhiều XML params (mỗi hook = 1 section = 1 XML param).
        SP trả về N result sets theo thứ tự hooks; mỗi result set chứa Id→fields
        để map ngược lại vào resolved_rows (in-place).

        Config SP_HOOK:
          XmlParam  : tên param XML của SP   (vd: @_B20BOMDetail)
          XmlTag    : tên thẻ XML mỗi dòng   (vd: Detail2)
          XmlFields : fields đưa vào XML      (vd: Id,ItemType,ItemName,ItemId,Name,Unit)
          Condition : EMPTY(field) hoặc rỗng
          Params    : scalar params từ parent  (vd: @_ItemId0={ItemId0}|@_BranchCode={BranchCode})
          OutputFields: fields lấy từ SP kết quả (vd: ItemId)
        """
        EMPTY_XML = '<NewDataSet />'

        # Build per-hook: batch_idx, xml_string, output_fields
        hook_meta = []       # [{batch_idx, h_outs}, ...]
        xml_args  = {}       # {xml_param: xml_string}
        pc_scalar, pv_scalar = [], []   # scalar params (từ hook đầu tiên)

        for hook in hooks:
            section    = hook.get('section', '')
            cond       = hook.get('condition', '').strip()
            xml_param  = hook.get('xmlparam', '').strip()
            xml_tag    = hook.get('xmltag', '').strip()
            xml_fields = [f.strip() for f in hook.get('xmlfields', '').split(',') if f.strip()]
            h_outs     = [f.strip() for f in hook.get('outputfields', '').split(',') if f.strip()]

            if not (xml_param and xml_tag and xml_fields):
                hook_meta.append({'batch_idx': {}, 'h_outs': h_outs})
                xml_args[xml_param or f'_empty_{len(xml_args)}'] = EMPTY_XML
                continue

            rows = section_data.get(section, {}).get('resolved', [])

            # Lọc rows theo condition
            batch_idx = {}
            for idx, row_vals in enumerate(rows):
                should_run = True
                if cond.startswith('EMPTY(') and cond.endswith(')'):
                    cond_field = cond[6:-1].strip()
                    should_run = not row_vals.get(cond_field)
                elif cond.startswith('NOTEMPTY(') and cond.endswith(')'):
                    cond_field = cond[9:-1].strip()
                    should_run = bool(row_vals.get(cond_field))
                else:
                    cond_field = ''
                if should_run:
                    batch_idx[idx] = row_vals

            hook_meta.append({'batch_idx': batch_idx, 'h_outs': h_outs})

            # Build XML (Id = idx+1 để map ngược)
            def _xml_escape(v):
                if v is None: return ''
                return (str(v)
                        .replace('&', '&amp;')
                        .replace('<', '&lt;')
                        .replace('>', '&gt;')
                        .replace('"', '&quot;'))

            if batch_idx:
                # SP dùng OPENXML flag=1 (attribute-centric)
                # → phải dùng attribute XML: <Detail2 Id="1" ItemType="A" ... />
                parts = ['<NewDataSet>']
                for idx, row_vals in batch_idx.items():
                    attrs = [f'Id="{idx + 1}"']
                    for field in xml_fields:
                        if field == 'Id':
                            continue
                        attrs.append(f'{field}="{_xml_escape(row_vals.get(field))}"')
                    parts.append(f'  <{xml_tag} {" ".join(attrs)} />')
                parts.append('</NewDataSet>')
                xml_args[xml_param] = '\n'.join(parts)
            else:
                xml_args[xml_param] = EMPTY_XML

            # Scalar params: chỉ build 1 lần từ hook đầu tiên có params
            if not pc_scalar:
                for ph in hook.get('params', '').split('|'):
                    ph = ph.strip()
                    if '=' not in ph:
                        continue
                    k, v = ph.split('=', 1)
                    k = k.strip().lstrip('@'); v = v.strip()
                    if v.startswith('{') and v.endswith('}'):
                        v = parent_row.get(v[1:-1])
                    if v is None:
                        pc_scalar.append(f'@{k}=NULL')
                    else:
                        pc_scalar.append(f'@{k}=?')
                        pv_scalar.append(v)

        if not xml_args:
            return

        # Build final EXEC call: tất cả params (kể cả XML) dùng ? placeholder
        # SP khai báo @_B20BOMDetail XML nên SQL Server tự convert NVARCHAR→XML
        pc_final = pc_scalar.copy()
        pv_final = pv_scalar.copy()
        for xml_param, xml_str in xml_args.items():
            pc_final.append(f'{xml_param}=?')
            pv_final.append(xml_str)

        exec_sql = f'EXEC {sp_name} ' + ', '.join(pc_final)

        try:
            cur = conn.cursor()
            # Giới hạn lock wait 120s — tránh treo vô tận nếu SP bị block
            cur.execute("SET LOCK_TIMEOUT 120000")
            cur.execute(exec_sql, pv_final)

            # Đọc result sets theo thứ tự hooks
            rs_idx   = 0
            mapped   = 0   # tổng số fields được map ngược lại
            while True:
                if rs_idx < len(hook_meta) and cur.description:
                    meta      = hook_meta[rs_idx]
                    batch_idx = meta['batch_idx']
                    h_outs    = meta['h_outs']
                    desc      = [d[0] for d in cur.description]
                    rs_rows   = cur.fetchall()
                    for sp_row in rs_rows:
                        sp_res = dict(zip(desc, sp_row))
                        sp_id  = sp_res.get('Id')
                        if sp_id is not None:
                            orig_idx = int(sp_id) - 1
                            if orig_idx in batch_idx:
                                for fout in h_outs:
                                    if fout in sp_res and sp_res[fout] is not None:
                                        batch_idx[orig_idx][fout] = sp_res[fout]
                                        mapped += 1

                if not cur.nextset():
                    break
                rs_idx += 1

        except Exception as e:
            err_msg = str(e)
            # Phân biệt lỗi permission để thông báo rõ hơn
            is_permission = (
                'EXECUTE permission' in err_msg
                or 'permission was denied' in err_msg.lower()
                or ('42000' in err_msg and 'permission' in err_msg.lower())
            )
            is_truncated = 'truncated' in err_msg.lower() or '2628' in err_msg or '8152' in err_msg
            if is_permission:
                friendly = (
                    f"Không có quyền EXECUTE stored procedure '{sp_name}'.\n"
                    f"Liên hệ DBA để cấp quyền:\n"
                    f"  GRANT EXECUTE ON {sp_name} TO [<login>]\n\n"
                    f"Chi tiết: {err_msg}"
                )
            elif is_truncated:
                import re as _re
                m = _re.search(r"column '([^']+)'.*Truncated value: '([^']+)'", err_msg)
                if m:
                    col, val = m.group(1), m.group(2)
                    friendly = (
                        f"Dữ liệu quá dài cho cột '{col}'.\n"
                        f"Giá trị bị cắt bớt: '{val}' ({len(val)} ký tự)\n"
                        f"→ Liên hệ DBA mở rộng độ dài cột hoặc rút ngắn mã.\n\n"
                        f"Chi tiết: {err_msg}"
                    )
                else:
                    friendly = f"Dữ liệu quá dài cho cột trong DB.\nChi tiết: {err_msg}"
            else:
                friendly = err_msg
            self._log('', f'SP_HOOK_BATCH {sp_name}', 0, 'Error', friendly, 'error')
            raise RuntimeError(friendly) from e


    # ── Custom dialog helpers (thay thế tkinter messagebox) ──────────────────
    def _show_msg(self, title, msg, kind='info'):
        """Custom CTk dialog: kind = 'info' | 'warning' | 'error'."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        # Màu icon theo loại
        _colors = {'error': '#E05252', 'warning': '#E0A030', 'info': '#4A9ECC'}
        _icons  = {'error': '✕', 'warning': '⚠', 'info': 'ℹ'}
        _c = _colors.get(kind, _colors['info'])
        _i = _icons.get(kind, _icons['info'])
        # Icon circle
        icon_frm = ctk.CTkFrame(dlg, fg_color=_c, corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text=_i, font=ctk.CTkFont(*FONT_TITLE),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont(*FONT_LABEL)).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont(*FONT_BODY),
                     wraplength=320, justify="center").pack(pady=(0, 16), padx=24)
        ctk.CTkButton(dlg, text="OK", width=100,
                      command=dlg.destroy).pack(pady=(0, 20))
        dlg.bind("<Return>", lambda e: dlg.destroy())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        dlg.update_idletasks()
        w, h = 380, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.focus_set()
        dlg.wait_window()

    def _toast(self, msg, duration=2500, kind="info"):
        """U9: Non-blocking toast overlay. kind = 'info' | 'success' | 'warning' | 'error'."""
        _bg = {'info': '#4A9ECC', 'success': '#3A8A50', 'warning': '#C07828', 'error': '#B84040'}
        bg = _bg.get(kind, _bg['info'])
        t = ctk.CTkToplevel(self)
        t.overrideredirect(True)
        t.attributes('-topmost', True)
        t.attributes('-alpha', 0.92)
        frm = ctk.CTkFrame(t, fg_color=bg, corner_radius=10)
        frm.pack(fill="both", expand=True, padx=0, pady=0)
        ctk.CTkLabel(frm, text=msg, font=ctk.CTkFont(*FONT_BODY),
                     text_color="white", wraplength=320).pack(padx=PAD_MD, pady=PAD_SM)
        t.update_idletasks()
        w = t.winfo_reqwidth() + 24
        h = t.winfo_reqheight()
        sw, sh = t.winfo_screenwidth(), t.winfo_screenheight()
        t.geometry(f"{w}x{h}+{(sw - w) // 2}+{sh - h - 60}")
        t.after(duration, t.destroy)

    def _show_export_success(self, title, msg, path):
        """Dialog xuất thành công với nút Mở file và Đóng."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        icon_frm = ctk.CTkFrame(dlg, fg_color='#4A9ECC', corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text='✓', font=ctk.CTkFont(*FONT_TITLE),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont(*FONT_LABEL)).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont(*FONT_BODY),
                     wraplength=340, justify="center").pack(pady=(0, 16), padx=24)
        btn_frm = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frm.pack(pady=(0, 20))
        def _open():
            dlg.destroy()
            try:
                os.startfile(path)
            except Exception as _e:
                pass
        ctk.CTkButton(btn_frm, text="📂  Mở file", width=110, command=_open).pack(side="left", padx=PAD_SM)
        ctk.CTkButton(btn_frm, text="Đóng", width=90, fg_color="gray40",
                      hover_color="gray30", command=dlg.destroy).pack(side="left", padx=PAD_SM)
        dlg.bind("<Return>", lambda e: dlg.destroy())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        dlg.update_idletasks()
        w, h = 400, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.focus_set()
        dlg.wait_window()

    def _ask_msg(self, title, msg):
        """Custom CTk yes/no dialog. Trả về True nếu Yes."""
        _result = [False]
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        icon_frm = ctk.CTkFrame(dlg, fg_color='#4A9ECC', corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text='?', font=ctk.CTkFont(*FONT_TITLE),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont(*FONT_LABEL)).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont(*FONT_BODY),
                     wraplength=320, justify="center").pack(pady=(0, 16), padx=24)
        btn_frm = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frm.pack(pady=(0, 20))
        def _yes():
            _result[0] = True
            dlg.destroy()
        ctk.CTkButton(btn_frm, text="Có", width=90, command=_yes).pack(side="left", padx=PAD_SM)
        ctk.CTkButton(btn_frm, text="Không", width=90, fg_color="gray40",
                      hover_color="gray30", command=dlg.destroy).pack(side="left", padx=PAD_SM)
        dlg.bind("<Return>", lambda e: _yes())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        dlg.update_idletasks()
        w, h = 380, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.focus_set()
        dlg.wait_window()
        return _result[0]

    def _make_loading_popup(self, msg="Đang import dữ liệu vào BRAVO...\nVui lòng không đóng cửa sổ.", grab=True):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Đang xử lý")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)   # chặn đóng cửa sổ
        lbl = ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont(*FONT_MD),
                     wraplength=300, justify="center")
        lbl.pack(pady=(22, 10), padx=PAD_LG)
        self._loading_lbl = lbl   # giữ ref để update từ background thread
        pb = ctk.CTkProgressBar(dlg, mode="indeterminate", width=300)
        pb.pack(padx=PAD_LG, pady=(0, 22))
        pb.start()
        self._loading_pb = pb     # U8: ref cho determinate mode khi biết _total_all
        # Canh giữa màn hình
        dlg.update_idletasks()
        w, h = 340, 160
        sw = dlg.winfo_screenwidth()
        sh = dlg.winfo_screenheight()
        x  = (sw - w) // 2
        y  = (sh - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")
        if grab:
            dlg.grab_set()
        else:
            dlg.attributes('-topmost', True)
            dlg.lift()
            dlg.focus_force()
        dlg.update()
        return dlg

    def _update_loading_msg(self, msg):
        """Cập nhật text loading dialog (phải gọi từ main thread qua self.after)."""
        lbl = getattr(self, '_loading_lbl', None)
        if lbl:
            try:
                lbl.configure(text=msg)
            except Exception:
                pass

    def _switch_loading_to_determinate(self, total):
        """U8: Chuyển progress bar sang determinate mode khi biết tổng số dòng."""
        pb = getattr(self, '_loading_pb', None)
        if pb:
            try:
                pb.stop()
                pb.configure(mode="determinate")
                pb.set(0.0)
            except Exception:
                pass

    def _update_loading_progress(self, fraction):
        """U8: Cập nhật giá trị progress bar (0.0–1.0). Gọi từ main thread."""
        pb = getattr(self, '_loading_pb', None)
        if pb:
            try:
                pb.set(min(max(fraction, 0.0), 1.0))
            except Exception:
                pass

    # ── Import entry point (button command) ───────────────────────────────────
    def _start_import(self):
        """
        Chạy trên main thread:
          1. Pre-checks nhanh
          2. Kết nối DB, build row, resolve fields, SP, validators
          3. Nếu export_only → _run_export_sql (main thread, cần file dialog)
          4. Nếu INSERT thật → hiện loading popup + thread _run_insert_bg
        """
        # U10: re-entrancy guard — chặn double-click khi import đang chạy
        if getattr(self, '_import_running', False):
            return
        self._import_running = True
        self.btn_import.configure(state="disabled")

        import datetime
        self.mapping = load_mapping()
        fname = os.path.basename(self._current_file)

        # ── 1. Pre-checks ────────────────────────────────────────────────────
        if self.val_errors is None:
            self._show_msg("Chưa Validate", "Hãy bấm Validate trước khi Import.", 'warning')
            return
        header_errors = any(
            r.get("severity") == "error"
            for r in self.val_errors.get("[H] Phiếu Header", [])
        )
        if header_errors:
            self._show_msg(
                "Lỗi HEADER",
                "Dữ liệu Header còn lỗi. Kiểm tra lại trước khi Import."
            , 'error')
            return
        if not self._current_creator_user_id:
            self._show_msg(
                "Chưa chọn Nhân viên",
                "Vui lòng chọn Nhân viên trước khi Import.",
                'warning')
            return

        # ── 2. Kết nối DB ────────────────────────────────────────────────────
        self._set_status("⏳  Đang kết nối DB...", C["yellow"])
        self.update()   # force render status ngay lập tức
        try:
            conn = self._get_db_conn()
            self._set_status("🔗  Đã kết nối", C["green"])
        except ConnectionError as e:
            self._set_status("❌  Kết nối thất bại", C["red"])
            self._log(fname, "Import", 0, "Lỗi kết nối", str(e), "error")
            self._show_msg("Lỗi kết nối DB", str(e))
            return
        except Exception as e:
            self._set_status("❌  Lỗi DB", C["red"])
            self._log(fname, "Import", 0, "Lỗi", str(e), "error")
            self._show_msg("Lỗi DB", str(e))
            return

        conn.autocommit = False

        self._set_status("⏳  Đang phân tích dữ liệu...", C["yellow"])
        _scan_dlg = self._make_loading_popup(
            "⏳  Đang phân tích file...\nVui lòng chờ.", grab=False)

        self._fuzzy_collect_mode = True
        self._fuzzy_pending      = []
        self._fuzzy_ctx          = {}
        self._fuzzy_resolutions  = {}
        self._fuzzy_batch_done   = False
        self._ps_bom_caches      = {}

        def _prescan_worker():
            import math as _math
            import datetime as _dt
            try:
                _now       = _dt.datetime.now()
                _meta      = self.global_meta
                _norm_meta = {_norm_vn(k): v for k, v in _meta.items()}
                _hmap = self.mapping.get('HEADER', [])
                self._ps_header_caches = self._build_all_caches(conn, _hmap)
                for _r in _hmap:
                    if _r.get('nguon_dl') in ('SP', 'TinhToan', 'HeThong', 'CoDinh'):
                        continue
                    if _r.get('kieu_lookup', '') not in ('fuzzy_code', 'fuzzy_name'):
                        continue
                    _cache_key_h = (_r.get('bang_master',''), _r.get('dieu_kien_master',''),
                                    _r.get('truong_so_sanh',''), _r.get('truong_lay_ve',''))
                    _cache_h = self._ps_header_caches.get(_cache_key_h, [])
                    self._fuzzy_ctx = {
                        'section': 'HEADER', 'field': _r.get('sql_col', ''), 'row_idx': None}
                    self._resolve_header_field(
                        _r, conn, _meta, _norm_meta, self._ps_header_caches, _now)
                # Sections có parent_section=HEADER (BOM detail sections)
                _PS_SECT = {
                    s for s, cfg in self.mapping.get('_CONFIG', {}).items()
                    if cfg.get('parent_section') == 'HEADER'
                }
                for _lbl, _tbl in self.tables.items():
                    _sec = _tbl.get('type')
                    if _sec not in _PS_SECT:
                        continue
                    _df = _tbl.get('df')
                    if _df is None or _df.empty or 'Lỗi' in _df.columns:
                        continue
                    _all_recs = self.mapping.get(_sec, [])
                    _fuzzy_recs = [
                        r for r in _all_recs
                        if _nan_str(r.get('kieu_lookup', '')) in ('fuzzy_code', 'fuzzy_name')
                        and _nan_str(r.get('nguon_dl', '')) not in ('CoDinh', 'HeThong', 'SP', 'TinhToan')
                        and _nan_str(r.get('truong_so_sanh', ''))
                        and _nan_str(r.get('truong_lay_ve', ''))
                    ]
                    if not _fuzzy_recs:
                        continue
                    _dcaches = self._build_bom_detail_caches(conn, _all_recs)
                    self._ps_bom_caches[_sec] = _dcaches
                    _stt_col = next((c for c in _df.columns if _norm_vn(str(c)) == 'stt'), None)
                    _order = 0
                    for _, _drow in _df.iterrows():
                        if _drow.isna().all():
                            continue
                        _stt_v = str(_drow.get(_stt_col, '') or '').strip() if _stt_col else ''
                        if _stt_v and SECTION_STT_PATTERN.match(_stt_v):
                            continue
                        _order += 1
                        for _rec in _fuzzy_recs:
                            _ten    = _nan_str(_rec.get('ten_excel', ''))
                            _kl     = _nan_str(_rec.get('kieu_lookup', ''))
                            _bm     = _nan_str(_rec.get('bang_master', ''))
                            _dk     = _nan_str(_rec.get('dieu_kien_master', ''))
                            _ss     = _nan_str(_rec.get('truong_so_sanh', ''))
                            _lv     = _nan_str(_rec.get('truong_lay_ve', ''))
                            _scol   = _rec.get('sql_col', '')
                            _nguong = int(_rec.get('nguong_fuzzy', 0) or 0) or 92
                            _raw = None
                            if _ten:
                                _norm_ten = _norm_vn(_ten)
                                for _col in _drow.index:
                                    _cs = str(_col).strip()
                                    if _cs == _ten or _norm_vn(_cs) == _norm_ten:
                                        _v = _drow[_col]
                                        if _v is None or (isinstance(_v, float) and _math.isnan(_v)):
                                            break
                                        _raw = _v
                                        break
                            if _raw is None:
                                continue
                            _cache_key = (_bm, _dk, _ss, _lv)
                            _cache = _dcaches.get(_cache_key, [])
                            self._fuzzy_ctx = {
                                'section': _sec, 'field': _scol, 'row_idx': _order}
                            self._lookup_generic(_raw, _cache, _kl, _nguong, _cache_key=_cache_key)
            except Exception as _pe:
                import traceback as _tb
                _err_msg = _tb.format_exc()
                self.after(0, lambda m=_err_msg: self._log(
                    fname, 'Prescan', 0, 'Error', m[:500], 'error'))
            finally:
                self._fuzzy_collect_mode = False
            self.after(0, _after_prescan)

        def _after_prescan():
            # ── Phase 1 (Main Thread): UI only ───────────────────────────────
            try: _scan_dlg.destroy()
            except Exception: pass
            self._set_status("", None)

            if self._fuzzy_pending:
                _resolved = self._show_batch_fuzzy_dialog(self._fuzzy_pending)
                self._fuzzy_resolutions = _resolved if _resolved is not None else {}
            else:
                self._fuzzy_resolutions = {}
            self._fuzzy_batch_done   = True
            self._fuzzy_collect_mode = False

            # Loading popup không modal — UI vẫn responsive trong khi resolve header
            self._loading_dlg = self._make_loading_popup(
                "Đang chuẩn bị dữ liệu...\nVui lòng chờ.", grab=False)
            threading.Thread(target=_header_resolve_bg, daemon=True).start()

        def _header_resolve_bg():
            # ── Phase 2 (Background Thread): toàn bộ DB work — KHÔNG GỌI UI TRỰC TIẾP ──
            import datetime as _dt
            import traceback as _tb
            _result = {
                'ok': False, 'error': None,
                'row': {}, 'lookup_log': {},
                'sp_errors': [],        # list[(sql_col, err_str)]
                'null_required': [],    # list[rec_dict]
                'violations': [],       # list[(vname, msg_str)]
                'validator_warns': [],  # list[(vname, exc_str)] — exception khi chạy SQL
                'export_only': False,
                'now': _dt.datetime.now(),
                'cols': [], 'sql_stmt': '', 'table_name': '', 'header_map': [],
            }
            try:
                now        = _result['now']
                meta       = self.global_meta
                norm_meta  = {_norm_vn(k): v for k, v in meta.items()}
                header_map = self.mapping.get('HEADER', [])
                _result['header_map'] = header_map

                # ── Giải quyết trường không-SP ───────────────────────────────
                self.after(0, lambda: self._update_loading_msg(
                    "Đang chuẩn bị dữ liệu...\nĐang giải quyết các trường header..."))
                row, lookup_log = {}, {}
                for rec in header_map:
                    if rec['nguon_dl'] == 'SP':
                        continue
                    sql_col = rec['sql_col']
                    val, mt = self._resolve_header_field(
                        rec, conn, meta, norm_meta,
                        self._ps_header_caches, now, row_out=row)
                    row[sql_col] = val
                    if mt not in ('codinh', 'hethong', 'excel_direct', 'sp',
                                  'tinhtoan', 'passthrough', 'fuzzy_pending'):
                        lookup_log[sql_col] = mt

                # ── Giải quyết trường SP ─────────────────────────────────────
                self.after(0, lambda: self._update_loading_msg(
                    "Đang chuẩn bị dữ liệu...\nĐang gọi Stored Procedure..."))
                sp_cfgs = {
                    s['sql_column']: s
                    for s in self.mapping.get('SP_CONFIG', [])
                    if s.get('isactive', '') == '1'
                    and s.get('section', '') in ('', 'HEADER')
                }
                for rec in header_map:
                    if rec['nguon_dl'] != 'SP':
                        continue
                    sql_col = rec['sql_col']
                    sp_cfg  = sp_cfgs.get(sql_col)
                    if not sp_cfg:
                        row.setdefault(sql_col, None)
                        continue
                    try:
                        sp_name  = sp_cfg.get('sp_name', '').strip()
                        params_s = sp_cfg.get('params', '').strip()
                        fallback = sp_cfg.get('fallback', '').strip() or None
                        if sp_name.lower() == 'lookup':
                            result_sp = self._sp_lookup(conn, params_s, row)
                        else:
                            result_sp = self._call_sp(conn, sp_name, params_s, row)
                        row[sql_col] = result_sp if result_sp is not None else fallback
                    except Exception as e:
                        _result['sp_errors'].append((sql_col, str(e)))
                        row[sql_col] = sp_cfg.get('fallback') or None

                # ── Kiểm tra trường bắt buộc NULL ────────────────────────────
                _result['null_required'] = [
                    rec for rec in header_map
                    if rec.get('bat_buoc', '').rstrip('0').rstrip('.') == '1'
                    and rec.get('nguon_dl') != 'SP'
                    and row.get(rec['sql_col']) is None
                ]

                # ── Build SQL INSERT header ───────────────────────────────────
                table_name = (self.db_cfg.get('table_name', 'B20BOM')
                              if self.db_cfg else 'B20BOM')
                _sp_temp_fields = {
                    s['sql_column'] for s in self.mapping.get('SP_CONFIG', [])
                    if s.get('sp_name', '').lower() == 'lookup'
                    and s.get('section', '') in ('', 'HEADER')
                }
                cols     = [c for c in row.keys() if c not in _sp_temp_fields]
                sql_stmt = (
                    "INSERT INTO " + table_name + " ("
                    + ", ".join("[" + c + "]" for c in cols)
                    + ") OUTPUT INSERTED.Id VALUES ("
                    + ", ".join(["?"] * len(cols)) + ")"
                )

                # ── Chạy validators SQL: thu thập vi phạm, không show dialog ─
                export_only = bool(
                    getattr(self, 'var_export_sql', None) and self.var_export_sql.get())
                _result['export_only'] = export_only
                if not export_only:
                    self.after(0, lambda: self._update_loading_msg(
                        "Đang chuẩn bị dữ liệu...\nĐang kiểm tra điều kiện import..."))
                    for _v in self.mapping.get('VALIDATORS', []):
                        if _v.get('isactive', '') != '1':
                            continue
                        _vsql   = _v.get('sql', '').strip()
                        _vparam = _v.get('params', '').strip()
                        _warn_m = _v.get('warningmessage', '').strip()
                        _vname  = _v.get('validatorname', '')
                        if not _vsql:
                            continue
                        _params = [row.get(p.strip()) for p in _vparam.split(',') if p.strip()]
                        try:
                            _cur = conn.cursor()
                            _cur.execute(_vsql, _params)
                            _first = _cur.fetchone()
                            if _first is not None:
                                _fv = _first[0]
                                if not (isinstance(_fv, (int, float)) and _fv == 0) and _fv is not None:
                                    _rest = _cur.fetchmany(4)
                                    _disp = ', '.join(str(r[0]) for r in [_first] + list(_rest))
                                    _msg  = _warn_m.replace('{result}', _disp)
                                    for _k, _v2 in row.items():
                                        _msg = _msg.replace('{' + _k + '}', str(_v2 or ''))
                                    _result['violations'].append((_vname, _msg))
                        except Exception as _ve:
                            _result['validator_warns'].append((_vname, str(_ve)))

                _result.update({
                    'ok': True, 'row': row, 'lookup_log': lookup_log,
                    'cols': cols, 'sql_stmt': sql_stmt, 'table_name': table_name,
                })
            except Exception:
                _result['error'] = _tb.format_exc()

            self.after(0, lambda r=_result: _on_header_resolved(r))

        def _on_header_resolved(result):
            # ── Phase 3 (Main Thread): hiển thị dialogs + khởi động insert ───
            def _abort():
                """Đóng loading popup và đóng connection."""
                # U10: clear re-entrancy flag khi abort
                self._import_running = False
                _n_err, _ = count_errors(self.val_errors or {})
                self.btn_import.configure(state="normal" if _n_err == 0 else "disabled")
                try:
                    if getattr(self, '_loading_dlg', None):
                        self._loading_dlg.destroy()
                        self._loading_dlg = None
                except Exception:
                    pass
                try:
                    conn.autocommit = True
                    conn.close()
                except Exception:
                    pass

            # Lỗi không mong muốn từ Phase 2
            if result.get('error'):
                _abort()
                self._log('', 'prescan', 0, 'Error', result['error'], 'error')
                self._show_msg(
                    "Lỗi chuẩn bị dữ liệu",
                    "Đã xảy ra lỗi khi đọc dữ liệu từ file Excel.\n"
                    "Chi tiết kỹ thuật đã được ghi vào log panel phía dưới.",
                    'error')
                return

            row        = result['row']
            lookup_log = result['lookup_log']
            now        = result['now']
            cols       = result['cols']
            sql_stmt   = result['sql_stmt']
            table_name = result['table_name']
            header_map = result['header_map']

            # Hiển thị lỗi SP (cảnh báo, không dừng)
            for _sc, _es in result.get('sp_errors', []):
                self._log(fname, f'SP {_sc}', 0, 'Error', _es, 'error')
                self._show_msg(f'SP Error — {_sc}',
                               f'Lỗi khi gọi SP cho [{_sc}]:\n{_es}', 'warning')

            # Ghi log exception của từng validator SQL
            for _vn, _ve in result.get('validator_warns', []):
                self._log('', f'Validator {_vn}', 0, 'Warn', _ve, 'warn')

            # Xử lý trường bắt buộc NULL
            _null_req = result.get('null_required', [])
            if _null_req:
                _names = ', '.join(
                    f"[{r['sql_col']}]" + (f" ({r['ten_excel']})" if r.get('ten_excel') else '')
                    for r in _null_req
                )
                _go = self._ask_msg(
                    "Thiếu dữ liệu bắt buộc",
                    f"Các trường sau bắt buộc nhưng không lấy được giá trị "
                    f"(lookup thất bại hoặc để trống):\n\n{_names}\n\n"
                    f"Tiếp tục import với các trường này để NULL không?")
                if not _go:
                    _abort()
                    return

            # Xử lý vi phạm Validators — tổng hợp 1 lần duy nhất thay vì per-validator
            _violations = result.get('violations', [])
            if _violations:
                _vio_text = '\n\n'.join(
                    f"• {_n}:\n  {_m}" for _n, _m in _violations
                )
                if not self._ask_msg(
                    f"Cảnh báo Validator ({len(_violations)} mục)",
                    f"Phát hiện {len(_violations)} cảnh báo:\n\n{_vio_text}"
                    f"\n\nVẫn tiếp tục import không?"
                ):
                    _abort()
                    return

            # Nhánh export SQL — chỉ đóng popup, giữ connection cho _run_export_sql
            export_only = result.get('export_only', False)
            if export_only:
                try:
                    if getattr(self, '_loading_dlg', None):
                        self._loading_dlg.destroy()
                        self._loading_dlg = None
                except Exception:
                    pass
                _eo_args = (conn, row, cols, sql_stmt, header_map, fname, now)
                self.after(0, lambda a=_eo_args: self._run_export_sql(*a))
                return

            # Chuyển loading popup sang chế độ modal cho insert phase
            self._update_loading_msg(
                "Đang import dữ liệu vào BRAVO...\nVui lòng không đóng cửa sổ.")
            try:
                self._loading_dlg.grab_set()
                self._loading_dlg.lift()
            except Exception:
                self._loading_dlg = self._make_loading_popup()

            ctx = {
                'conn': conn, 'row': row, 'cols': cols, 'sql': sql_stmt,
                'lookup_log': lookup_log, 'fname': fname, 'now': now,
                'table_name': table_name, 'header_map': header_map,
            }
            threading.Thread(target=self._run_insert_bg, args=(ctx,), daemon=True).start()

        threading.Thread(target=_prescan_worker, daemon=True).start()

    def _run_insert_bg(self, ctx):
        """
        Background thread — INSERT B20BOM + B20BOMDetail trong 1 transaction.
        Không gọi bất kỳ UI nào trực tiếp; kết quả post qua self.after().
        """
        conn       = ctx['conn']
        row        = ctx['row']
        cols       = ctx['cols']
        sql        = ctx['sql']
        fname      = ctx['fname']
        now        = ctx['now']
        table_name = ctx['table_name']
        lookup_log = ctx['lookup_log']
        try:
            cur = conn.cursor()
            cur.execute(sql, [row[c] for c in cols])
            new_id = cur.fetchone()[0]
            _detail_count = self._generate_bom_details(
                conn, new_id, row, self.tables, self.mapping, now, export_only=False
            )
            conn.commit()
            result = {
                'ok': True,
                'new_id': new_id,
                'detail_count': _detail_count,
                'row': row,
                'fname': fname,
                'lookup_log': lookup_log,
                'table_name': table_name,
            }
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            result = {'ok': False, 'error': str(e), 'fname': fname}
        finally:
            try:
                conn.autocommit = True
                conn.close()
            except Exception:
                pass
        self.after(0, lambda: self._finish_import(result))

    def _finish_import(self, result):
        # U10: clear re-entrancy flag and restore btn_import state
        self._import_running = False
        _n_err, _ = count_errors(self.val_errors or {})
        self.btn_import.configure(state="normal" if _n_err == 0 else "disabled")

        try:
            if self._loading_dlg:
                self._loading_dlg.destroy()
                self._loading_dlg = None
        except Exception:
            pass
        fname = result.get('fname', '')
        if result.get('ok'):
            new_id = result['new_id']
            n_det  = result['detail_count']
            self._set_status(f"\u2705  Import thành công — BOM ID {new_id}, {n_det} dòng chi tiết", C["green"])
            self._log(fname, "Import", str(n_det), "OK", f"BOM ID={new_id}, {n_det} dòng chi tiết", "ok")
            self._db_log('BOM', fname, new_id, n_det, 'OK',
                         f"BOM ID={new_id}, {n_det} dòng chi tiết")
            self._last_import_bom_id = new_id
            self._last_import_fname  = fname
            self.btn_undo_import.configure(state="normal")
            self._show_msg("Import thành công", f"BOM đã được tạo thành công!\nBOM ID: {new_id}\nSố dòng chi tiết: {n_det:,}", kind="info")
        else:
            err = result.get('error', 'Unknown error')
            self._set_status("\u274c  Import thất bại", C["red"])
            self._log(fname, "Import", "0", "Lỗi", err, "error")
            self._db_log('BOM', fname, '', 0, 'LOI', err)
            self._show_msg("Import thất bại", f"Lỗi: {err}", kind="error")


    # ── Hoàn tác import (gọi SP usp_BOMTool_DeleteBOM — khách hàng deploy) ────
    def _undo_last_import(self):
        bom_id = self._last_import_bom_id
        if bom_id is None:
            return
        if not self._ask_msg("Hoàn tác import",
                f"Xóa BOM vừa import khỏi BRAVO?\n\n"
                f"BOM ID: {bom_id}\n"
                f"File: {self._last_import_fname}\n\n"
                f"Thao tác này KHÔNG thể tự hoàn tác lại."):
            return
        self.btn_undo_import.configure(state="disabled", text="⏳  Đang xóa...")
        threading.Thread(target=self._undo_worker, args=(bom_id,),
                         daemon=True).start()

    def _undo_worker(self, bom_id):
        conn = None
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute("EXEC dbo.usp_BOMTool_DeleteBOM ?", (bom_id,))
            row   = cur.fetchone()
            n_det = row[1] if row else 0
            conn.commit()
            self.after(0, lambda n=n_det: self._undo_done(bom_id, n, None))
        except Exception as e:
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
            self.after(0, lambda err=str(e): self._undo_done(bom_id, 0, err))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _undo_done(self, bom_id, n_det, error):
        self.btn_undo_import.configure(text="↩  Hoàn tác import")
        if error:
            self.btn_undo_import.configure(state="normal")
            if 'Could not find stored procedure' in error \
                    or 'usp_BOMTool_DeleteBOM' in error:
                self._show_msg("Chưa có SP trên DB",
                    "Stored procedure usp_BOMTool_DeleteBOM chưa được tạo "
                    "trên database.\n\n"
                    "Gửi file sql\\create_usp_bomtool_deletebom.sql cho khách hàng "
                    "chạy, sau đó bấm lại nút này.", kind="info")
            else:
                self._show_msg("Lỗi hoàn tác", error, kind="error")
            return
        self._last_import_bom_id = None
        self.btn_undo_import.configure(state="disabled")
        self._set_status(f"↩  Đã xóa BOM ID {bom_id} ({n_det} dòng chi tiết)",
                         C["green"])
        self._log(self._last_import_fname, "Undo", str(n_det), "OK",
                  f"Đã xóa BOM ID={bom_id}", "warn")
        self._db_log('UNDO', self._last_import_fname, bom_id, n_det, 'OK',
                     f"Đã xóa BOM ID={bom_id} ({n_det} dòng chi tiết)")
        self._show_msg("Đã hoàn tác",
            f"BOM ID {bom_id} và {n_det:,} dòng detail đã được xóa khỏi BRAVO.",
            kind="info")


if __name__ == "__main__":
    app = BOMToolApp()
    app.mainloop()



