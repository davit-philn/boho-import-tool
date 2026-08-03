"""
services/bom_parser.py — Parse BOM Excel, THDM sheets, shared row resolver.
"""
import re
import os
import sys
import datetime
import unicodedata
import io
import math
import logging
import pandas as pd
from openpyxl import load_workbook

try:
    import msoffcrypto
    _HAS_MSOFFCRYPTO = True
except ImportError:
    _HAS_MSOFFCRYPTO = False

try:                          # A1: guard — bom_parser chạy được không cần GUI
    import tkinter as tk
    _HAS_TK = True
except ImportError:
    tk = None
    _HAS_TK = False

from services.utils import _norm_vn, _nan_str, guess_col_align
from services.mapping_loader import (
    load_mapping, build_reverse_map, match_col_to_sql,
    build_meta_keys_from_mapping, _load_section_rows,
    MAPPING_FILE, BASE_DIR,
    SECTION_STT_PATTERN, NUMERIC_STT_PATTERN, _ROMAN_SIMPLE_RE,
    HEADER_ANCHORS,
)
_log = logging.getLogger(__name__)

def _is_roman_numeral(s: str) -> bool:
    """True nếu s là số La Mã thực dùng trong BOM (I–XXXIX, chỉ dùng I/V/X).
    Loại trừ chữ cái đơn A–Z làm section header (A,B,C,D không khớp pattern này)."""
    s = s.rstrip('.')
    m = _ROMAN_SIMPLE_RE.match(s)
    return bool(m) and bool(m.group(0).rstrip('.'))
FOOTER_KEYWORDS   = ["Tổng cộng", "Kiểm duyệt", "Người Lập", "Người lập", "*Ghi chú", "Ghi chú:",
                     "Ngày tháng năm", "Ngày   tháng", "duyệt", "Ghi chú cuối"]
# Pattern nhận diện dòng ký tên: STT trống + cột số có text dài
_SIGNER_COL_RE    = re.compile(r"(DAI|DAY|RONG|Width|Length|Thickness)", re.IGNORECASE)

def _load_sheet_config(mapping_path=None):
    """Đọc sheet _CONFIG từ mapping file → trả list config dict cho các section BOM*.

    Single Source of Truth — không hardcode fallback.
    _CONFIG columns (R4 trở đi, sau 3 dòng header):
      [0] Section  [1] Label  [2] View_Insert  [3] SheetNameContains
      [4] SheetNameExclude  [5] DataStartRow  ...

    Chỉ lấy BOM* — HEADER/THDM_* xử lý riêng.
    Trả về [] nếu không đọc được (caller tự xử lý).
    """
    path = mapping_path or MAPPING_FILE
    try:
        import openpyxl as _opx
        _wb = _opx.load_workbook(path, read_only=True, data_only=True)
        if '_CONFIG' not in _wb.sheetnames:
            return []
        cfg = []
        for row in _wb['_CONFIG'].iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            section = str(row[0]).strip()
            if not section.upper().startswith('BOM'):
                continue
            label     = str(row[1]).strip() if row[1] else section
            contains  = [x.strip().upper() for x in str(row[3] or '').split(',') if x.strip()]
            excludes  = [x.strip().upper() for x in str(row[4] or '').split(',') if x.strip()]
            start_row = str(row[5]).strip() if row[5] else 'AUTO'
            if contains:
                cfg.append({"section": section, "label": label,
                             "contains": contains, "excludes": excludes,
                             "start_row": start_row})
        _wb.close()
        return cfg
    except Exception:
        return []

def _detect_sheet_type(name, sheet_config=None):
    """Match tên sheet Excel với config dùng regex word-boundary matching.

    Dùng \\b (word boundary) để phân biệt chính xác:
      'II' không khớp trong 'III'   ('I' tiếp theo phá vỡ boundary)
      'V'  không khớp trong 'IV'    ('I' trước V là word-char → không có boundary)

    Normalize underscore → space trước khi match vì trong Python regex,
    '_' là word-char nên 'BOM_IV' sẽ KHÔNG có \\b trước 'I' nếu không normalize.

    Xử lý đúng tất cả dạng separator: space, hyphen, underscore, dot, ngoặc...
    Phải truyền sheet_config từ caller. Nếu không → [] → UNKNOWN.
    """
    cfg = sheet_config or []
    normalized = name.upper().replace('_', ' ')
    for item in cfg:
        contains_match = all(
            re.search(r'\b' + re.escape(kw) + r'\b', normalized)
            for kw in item["contains"]
        )
        excludes_hit = any(
            re.search(r'\b' + re.escape(kw) + r'\b', normalized)
            for kw in item["excludes"]
        )
        if contains_match and not excludes_hit:
            return item["section"]
    return "UNKNOWN"

def _resolve_formula(val):
    """Chuyển formula string (data_only=False) thành giá trị thực.
    - ="some text"  → "some text"
    - =A1, =CONCAT(...) → None  (không resolve được, sẽ fallback sang cached)
    - Giá trị thường → giữ nguyên
    """
    if val is None:
        return None
    if isinstance(val, str) and val.startswith('='):
        m = re.match(r'^="(.*)"$', val, re.DOTALL)
        if m:
            return m.group(1)
        # Công thức phức tạp không resolve được
        return None
    return val

def _merge_meta_rows(live_rows, cached_rows):
    """Gộp live rows (data_only=False) và cached rows (data_only=True).
    Ưu tiên live value đã resolve; fallback sang cached khi formula phức tạp.
    """
    merged = []
    for i in range(max(len(live_rows), len(cached_rows))):
        live_row   = live_rows[i]   if i < len(live_rows)   else ()
        cached_row = cached_rows[i] if i < len(cached_rows) else ()
        n = max(len(live_row), len(cached_row))
        row = []
        for j in range(n):
            live_val   = live_row[j]   if j < len(live_row)   else None
            cached_val = cached_row[j] if j < len(cached_row) else None
            resolved   = _resolve_formula(live_val)
            row.append(resolved if resolved is not None else cached_val)
        merged.append(tuple(row))
    return merged

def _extract_meta(rows, meta_keys=None):
    """
    Scan 15 dòng đầu, tìm label → lấy giá trị ô kế bên.
    meta_keys: {label: regex_pattern} — dynamic từ mapping.

    Hai guard chống false-positive:
    1. Label cell phải ngắn (<=40 ký tự) — loại instruction text dài như
       'Nhập lại số lượng SP theo đơn hàng...' khỏi bị nhận là label.
    2. Quét giá trị tối đa 8 cột sang phải — ngăn bắt ô từ bảng tham chiếu
       bên phải sheet (cách xa 20+ cột).
    """
    keys = meta_keys if meta_keys else {}
    meta = {}
    for row in rows:
        for key, pat in keys.items():
            if key in meta:
                continue
            for i, cell in enumerate(row):
                cell_str = str(cell).strip() if cell else ''
                if cell_str and len(cell_str) <= 40 and re.search(pat, cell_str, re.IGNORECASE):
                    for j in range(i + 1, min(i + 9, len(row))):
                        if row[j] is not None and str(row[j]).strip():
                            meta[key] = str(row[j]).strip()
                            break
                    break
    return meta

def _find_header_row(rows):
    for i, row in enumerate(rows):
        texts = [str(c).strip() for c in row if c is not None]
        hits  = sum(1 for a in HEADER_ANCHORS if any(a.lower() in t.lower() for t in texts))
        if hits >= 2:
            return i
    return None

def _norm_col(s):
    return re.sub(r"[\s\n]+", "_", str(s).strip()).rstrip("_") if s else ""

def _build_headers(r1, r2):
    """Gộp 2 dòng header (merged cell) thành tên cột duy nhất.
    Forward-fill r1 để xử lý merged cell: khi r1[i]=None nhưng r2[i] có giá trị,
    nghĩa là ô r1 bị merge → kế thừa giá trị parent từ cột trước.
    """
    n = max(len(r1), len(r2))
    r1_f = list(r1) + [None] * max(0, n - len(r1))
    r2_f = list(r2) + [None] * max(0, n - len(r2)) if r2 else [None] * n

    # Forward-fill r1 cho merged cells: r1[i]=None nhưng r2[i] có giá trị
    last_r1 = None
    for i in range(n):
        if r1_f[i] is not None:
            last_r1 = r1_f[i]
        elif r2_f[i] is not None and last_r1 is not None:
            r1_f[i] = last_r1

    headers = []
    for i in range(n):
        v1 = _norm_col(r1_f[i]) if r1_f[i] is not None else ""
        v2 = _norm_col(r2_f[i]) if r2_f[i] is not None else ""
        if v1 and v2:   headers.append(f"{v1}_{v2}")
        elif v1:        headers.append(v1)
        elif v2:        headers.append(v2)
        else:           headers.append(f"COL_{i}")
    # Deduplicate
    seen, result = {}, []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result

def _is_data_row(rd, key_cols):
    """True nếu dòng có ít nhất 1 giá trị thực trong các cột key."""
    for k in key_cols:
        v = rd.get(k)
        if v is not None and str(v).strip() not in ('', '0', 'None', 'nan'):
            return True
    return False

# ─── Shared Excel parse pipeline (BOM + THDM) ───────────────────────────────

def _build_excel_col_map(headers, mapping_recs, col_aliases=None):
    """
    Build {col_index → sql_col} từ headers và mapping records.
    Dùng cho openpyxl-based section parsing (THDM; BOM tương lai).

    col_aliases: dict {normalized_header → canonical} để chuẩn hoá tên cột Excel
                 (vd: 'mavttb' → 'mavattu')
    """
    aliases    = col_aliases or {}
    excel_recs = [r for r in mapping_recs
                  if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')]
    col_map    = {}
    for i, h in enumerate(headers):
        h_vn       = _norm_vn(str(h).replace('_', ' '))
        h_canonical = aliases.get(h_vn, h_vn)
        for rec in excel_recs:
            ten_vn = _norm_vn(rec['ten_excel'])
            if (h_vn == ten_vn
                    or ten_vn == h_canonical
                    or (ten_vn and ten_vn in h_vn)
                    or (h_canonical and ten_vn in h_canonical)
                    or (h_vn and h_vn in ten_vn)):
                col_map[i] = rec['sql_col']
                break
    return col_map


def _parse_section_excel_rows(all_rows, headers, col_map, mapping_recs, stt_idx,
                               skip_no_item=True, item_cols=('ItemId', 'ItemName'),
                               ff_always_override=False,
                               extra_col_extractor=None,
                               roman_as_hdr=False):
    """
    Pipeline chung parse Excel rows cho 1 section — dùng cho cả BOM và THDM.

    Xử lý đầy đủ:
    - Bỏ dòng trống
    - Footer detection (FOOTER_KEYWORDS)
    - Section header detection (SECTION_STT_PATTERN + GIAO RỜI / sub-group check)
    - Fill_Forward: capture từ section header → apply cho data rows kế tiếp
    - Bỏ dòng không có identity (skip_no_item=True)

    all_rows    : list of value tuples (openpyxl rows)
    headers     : list[str] — tên cột đã build (kể cả 2-row merge)
    col_map     : {col_index → sql_col}   ← từ _build_excel_col_map()
    mapping_recs: list of mapping records cho section này
    stt_idx     : index cột STT trong headers (None nếu không tìm thấy)

    Trả về: list of {sql_col: value} dicts đã apply Fill_Forward.
    """
    import math as _m

    ff_fields   = [r['sql_col'] for r in mapping_recs
                   if r.get('fill_forward') == '1' and r.get('sql_col')]
    dim_indices = [
        i for i, h in enumerate(headers)
        if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(h)))
    ]

    current_ff = {}
    footer_hit = False
    rows_out   = []

    for row in all_rows:
        # Bỏ dòng hoàn toàn trống
        if all(c is None for c in row):
            continue
        if footer_hit:
            break

        padded = list(row) + [None] * max(0, len(headers) - len(row))

        # Footer detection
        row_text = ' '.join(str(v) for v in padded if v is not None)
        if any(kw.lower() in row_text.lower() for kw in FOOTER_KEYWORDS):
            footer_hit = True
            continue

        # STT value
        stt_val = padded[stt_idx] if stt_idx is not None and stt_idx < len(padded) else None
        stt_str = str(stt_val).strip() if stt_val is not None else ''

        is_hdr = bool(stt_str and SECTION_STT_PATTERN.match(stt_str))

        # GIAO RỜI / sub-group: có kích thước vật lý → data row, không phải section header
        if is_hdr and dim_indices:
            if any(
                padded[i] not in (None, '', 0, 0.0)
                and not (isinstance(padded[i], float) and _m.isnan(padded[i]))
                for i in dim_indices if i < len(padded)
            ):
                is_hdr = False

        # Build row dict {sql_col: value}
        rd = {sql_col: padded[ci]
              for ci, sql_col in col_map.items()
              if ci < len(padded)}

        if is_hdr:
            if not _is_roman_numeral(stt_str) or roman_as_hdr:
                # Chữ cái (A, B, C...) → capture Fill_Forward, bỏ qua không INSERT
                for sql_col in ff_fields:
                    v = rd.get(sql_col)
                    if v is not None and str(v).strip() not in ('', 'nan'):
                        current_ff[sql_col] = v
                continue
            # Số La Mã (I, II, III...) → fall-through, xử lý như data row

        # Apply Fill_Forward vào data row
        for sql_col in ff_fields:
            if sql_col in current_ff:
                if ff_always_override or not rd.get(sql_col):
                    rd[sql_col] = current_ff[sql_col]

        # Extra columns (e.g., CHI TIẾT MỤC data cho THDM transpose)
        if extra_col_extractor is not None:
            rd.update(extra_col_extractor(padded))

        # Bỏ dòng không có identity
        if skip_no_item:
            _id_v  = rd.get(item_cols[0]) if item_cols else None
            _nm_v  = rd.get(item_cols[1]) if len(item_cols) > 1 else None
            _has_id = bool(_id_v is not None and str(_id_v).strip() not in ('', 'nan'))
            _has_nm = bool(_nm_v is not None and str(_nm_v).strip() not in ('', 'nan'))
            if not _has_id and not _has_nm:
                continue
            # Dòng tiêu đề nhóm: ItemId rỗng + ItemName toàn HOA → bỏ qua
            # VD: "VẬT TƯ CHÍNH", "NGUYÊN VẬT LIỆU CHÍNH"
            # Ngoại lệ: STT số > 0 → data row thực (vd: M1_KÍNH THỦY THEO MẪU), không bỏ qua
            if not _has_id and _has_nm:
                _nm_s = str(_nm_v).strip()
                if _nm_s.upper() == _nm_s and any(c.isalpha() for c in _nm_s):
                    if not (stt_str and NUMERIC_STT_PATTERN.match(stt_str)):
                        continue

        rows_out.append(rd)

    return rows_out


def _run_row_sp_hooks(conn, hooks, row_vals, log_fn=None):
    """
    Chạy SP_HOOK event='beforeinsert' cho 1 data row.
    Cập nhật row_vals in-place với output fields từ SP.
    Dùng chung cho BOM và THDM.

    conn    : pyodbc connection
    hooks   : list of hook dicts (SP_HOOK mapping, event='beforeinsert', đã filter theo section)
    row_vals: dict {sql_col: value} — sẽ bị cập nhật in-place
    log_fn  : optional callable(msg: str) để log warning
    """
    for hook in hooks:
        cond       = hook.get('condition', '').strip()
        should_run = True
        if cond.startswith('EMPTY(') and cond.endswith(')'):
            should_run = not row_vals.get(cond[6:-1].strip())
        elif cond.startswith('NOTEMPTY(') and cond.endswith(')'):
            should_run = bool(row_vals.get(cond[9:-1].strip()))
        if not should_run:
            continue
        try:
            h_sp   = hook.get('sp_name', '').strip()
            h_par  = hook.get('params', '').strip()
            h_outs = [f.strip() for f in hook.get('outputfields', '').split(',') if f.strip()]
            pc, pv = [], []
            for ph in h_par.split('|'):
                ph = ph.strip()
                if '=' not in ph:
                    continue
                k, v = ph.split('=', 1)
                k = k.strip().lstrip('@')
                v = v.strip()
                if v.startswith('{') and v.endswith('}'):
                    v = row_vals.get(v[1:-1])
                if v is None:
                    pc.append(f'@{k}=NULL')
                else:
                    pc.append(f'@{k}=?')
                    pv.append(v)
            cur = conn.cursor()
            cur.execute(f'EXEC {h_sp} ' + ', '.join(pc), pv)
            hk_row = cur.fetchone()
            if hk_row and cur.description:
                hk_res = dict(zip([d[0] for d in cur.description], hk_row))
                for fh in h_outs:
                    if fh in hk_res:
                        row_vals[fh] = hk_res[fh]
        except Exception as eh:
            if log_fn:
                log_fn(f'SP_HOOK {hook.get("sp_name")} warn: {eh}')

def _parse_sheet(ws, sheet_name, live_meta_rows=None, meta_keys=None, hidden_rows=None, hidden_cols=None, sheet_config=None):
    # Bỏ qua các row/col bị ẩn (hidden) trong Excel
    # hidden_rows: set of 1-based row indices
    # hidden_cols: set of 0-based col indices
    _hr = hidden_rows or set()
    _hc = hidden_cols or set()

    def _filter_cols(row):
        if not _hc:
            return row
        return tuple(v for i, v in enumerate(row) if i not in _hc)

    all_rows = [
        _filter_cols(vals)
        for row_idx, vals in enumerate(ws.iter_rows(values_only=True), start=1)
        if row_idx not in _hr
    ]
    if not all_rows:
        return None

    stype = _detect_sheet_type(sheet_name, sheet_config=sheet_config)
    if stype == "UNKNOWN":
        return None

    # Dùng live_meta_rows (data_only=False) để đọc đúng giá trị formula cells
    meta_rows = live_meta_rows if live_meta_rows is not None else all_rows[:15]
    meta = _extract_meta(meta_rows, meta_keys=meta_keys)
    hidx = _find_header_row(all_rows[:15])

    if hidx is None:
        return {"sheet_name": sheet_name, "sheet_type": stype,
                "metadata": meta, "df": pd.DataFrame()}

    r1 = all_rows[hidx]
    r2 = all_rows[hidx + 1] if hidx + 1 < len(all_rows) else ()

    # Kiểm tra r2 có phải dòng sub-header không
    has_sub = bool(r2) and r2[0] is None and any(
        c is not None and str(c).strip() not in ("", "0")
        for c in (r2[4:] if r2 else []))

    if has_sub:
        headers    = _build_headers(r1, r2)
        data_start = hidx + 2
    else:
        headers    = [_norm_col(c) if c else f"COL_{i}" for i, c in enumerate(r1)]
        data_start = hidx + 1

    key_cols = [h for h in headers if any(k in h for k in
        ("Tên", "Mã", "STT", "Name", "Code", "Item"))]

    rows_out = []
    _footer_seen = False   # Khi gặp footer (Kiểm duyệt / Người Lập...) → bỏ tất cả dòng phía sau
    for row in all_rows[data_start:]:
        if all(c is None for c in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        rd     = dict(zip(headers, padded[:len(headers)]))

        _raw_stt = rd.get("STT")
        stt = ("0" if _raw_stt in (0, 0.0)
               else "" if _raw_stt is None or _raw_stt == ""
               else str(_raw_stt).strip())

        # Section header: STT là chữ cái đơn hoặc chữ.số (A, B, C, F, B.1, B.2...)
        # Số La Mã (I, II, III...) được giữ lại như data row.
        # Áp dụng thống nhất cho BOM2, BOM3, BOM4 — không cần check SLg
        # → giữ lại cho Fill_Forward capture, không INSERT
        if SECTION_STT_PATTERN.match(stt):
            # Nếu dòng có kích thước vật lý (DÀY/RỘNG/DÀI) thì là GIAO RỜI/sub-group (vd MODULE MD1 STT='I')
            # → không phải section header thật → không update Fill-Forward
            _has_dims = any(
                rd.get(col) not in (None, '', 0, '0', 0.0)
                for col in rd
                if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(col)))
            )
            if not _has_dims:
                if _is_roman_numeral(stt):
                    pass  # Số La Mã → fall-through, xử lý như data row
                else:
                    rows_out.append(rd)   # Chữ cái → giữ lại cho Fill_Forward, không INSERT
                    continue
            # Có kích thước → fall-through: xử lý như data row bình thường
        # Bỏ dòng trống STT=0 (Phần III có nhiều dòng này)
        if stt == "0":
            continue
        # Bỏ dòng footer (case-insensitive) — và đặt cờ để bỏ qua tất cả dòng phía sau
        # (bao gồm dòng tên người ký xuất hiện sau "Kiểm duyệt / Người Lập")
        row_text = " ".join(str(v) for v in rd.values() if v)
        if any(kw.lower() in row_text.lower() for kw in FOOTER_KEYWORDS):
            _footer_seen = True
            continue
        if _footer_seen:
            continue
        # Bỏ dòng có STT không hợp lệ: không phải số (1, 2, 1.1...) và không phải La Mã
        # → ví dụ tên người ký nằm trong cột STT: "Đỗ Hồng Thái", "Nguyễn Tân Hồng"
        if stt and not NUMERIC_STT_PATTERN.match(stt) and not _is_roman_numeral(stt):
            continue
        # Bỏ dòng ký tên: STT trống + (cột số có text phi số) HOẶC (Số lượng có text phi số)
        # Cần check cả SLg vì BOM3 không có cột DÀI/Width nên _SIGNER_COL_RE không match
        if not stt:
            def _is_non_numeric(v):
                try: float(str(v).replace(',', '.')); return False
                except (ValueError, TypeError): return True
            _dim_text = any(_is_non_numeric(val) for col, val in rd.items()
                            if val and _SIGNER_COL_RE.search(col))
            _slg_h_sig = next((h for h in rd
                               if _norm_vn(str(h)) in ('slg', 'soluong', 'soluongsp')), None)
            _slg_v_sig = rd.get(_slg_h_sig) if _slg_h_sig else None
            _slg_text  = (_slg_v_sig is not None
                          and str(_slg_v_sig).strip() not in ('', 'nan')
                          and _is_non_numeric(_slg_v_sig))
            if _dim_text or _slg_text:
                continue
        # Bỏ dòng không có dữ liệu key
        if not _is_data_row(rd, key_cols):
            continue

        rows_out.append(rd)

    df = pd.DataFrame(rows_out, columns=headers) if rows_out else pd.DataFrame(columns=headers)

    # Bỏ cột toàn None hoặc tên COL_
    drop = [c for c in df.columns if c.startswith("COL_") or df[c].isna().all()]
    df   = df.drop(columns=drop, errors="ignore")

    return {"sheet_name": sheet_name, "sheet_type": stype, "metadata": meta, "df": df}

def _ask_excel_password(filename):
    """
    Hiện dialog nhỏ hỏi mật khẩu Excel.
    Trả về: password (str) hoặc None nếu người dùng bấm Hủy.
    """
    result = [None]
    dialog = tk.Toplevel()
    dialog.title("File được bảo vệ bằng mật khẩu")
    dialog.resizable(False, False)
    dialog.grab_set()
    dialog.focus_force()

    # Center dialog
    dialog.update_idletasks()
    w, h = 400, 180
    x = (dialog.winfo_screenwidth()  - w) // 2
    y = (dialog.winfo_screenheight() - h) // 2
    dialog.geometry(f"{w}x{h}+{x}+{y}")
    dialog.configure(bg="#1e1e1e")

    tk.Label(dialog, text="🔒  File Excel có mật khẩu",
             font=("Segoe UI", 11, "bold"),
             bg="#1e1e1e", fg="#F1F5F9").pack(pady=(16, 2))
    tk.Label(dialog, text=os.path.basename(filename),
             font=("Segoe UI", 9), bg="#1e1e1e", fg="#94A3B8").pack()
    tk.Label(dialog, text="Nhập mật khẩu để mở file:",
             font=("Segoe UI", 9), bg="#1e1e1e", fg="#CBD5E1").pack(pady=(10, 4))

    entry = tk.Entry(dialog, show="*", font=("Segoe UI", 10),
                     bg="#334155", fg="white", insertbackground="white",
                     relief="flat", width=32)
    entry.pack(ipady=5)
    entry.focus_set()

    err_lbl = tk.Label(dialog, text="", font=("Segoe UI", 8),
                       bg="#1e1e1e", fg="#F87171")
    err_lbl.pack()

    def on_ok(event=None):
        pw = entry.get()
        if not pw:
            err_lbl.config(text="Vui lòng nhập mật khẩu.")
            return
        result[0] = pw
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    btn_frame = tk.Frame(dialog, bg="#1e1e1e")
    btn_frame.pack(pady=(4, 0))
    tk.Button(btn_frame, text="Hủy", width=10, command=on_cancel,
              bg="#334155", fg="white", relief="flat",
              activebackground="#475569", cursor="hand2").pack(side="left", padx=6)
    tk.Button(btn_frame, text="✓  Xác nhận", width=12, command=on_ok,
              bg="#2563eb", fg="white", relief="flat",
              activebackground="#1d4ed8", cursor="hand2").pack(side="left", padx=6)

    entry.bind("<Return>", on_ok)
    dialog.bind("<Escape>", lambda e: on_cancel())
    dialog.wait_window()
    return result[0]


def _decrypt_excel(filepath, password):
    """
    Giải mã file Excel có mật khẩu bằng msoffcrypto.
    Trả về BytesIO chứa nội dung đã giải mã, hoặc raise Exception nếu sai mật khẩu.
    """
    decrypted = io.BytesIO()
    with open(filepath, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def _is_encrypted_excel(filepath):
    """
    Kiểm tra xem file có bị mã hóa không.
    Method 1: msoffcrypto (nếu đã install).
    Method 2: fallback — thử mở bằng zipfile, nếu fail → có thể bị mã hóa.
    """
    # Method 1: msoffcrypto chính xác hơn
    if _HAS_MSOFFCRYPTO:
        try:
            with open(filepath, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                return office_file.is_encrypted()
        except Exception:
            pass

    # Method 2: fallback — file xlsx là ZIP, nếu không mở được → encrypted
    import zipfile
    try:
        with zipfile.ZipFile(filepath, 'r'):
            return False  # mở được ZIP → không có password
    except zipfile.BadZipFile:
        return True   # không phải ZIP → khả năng cao bị mã hóa
    except Exception:
        return False


def parse_bom_file(filepath, meta_keys=None, _decrypted_bytes=None):
    """
    Parse file BOM Excel.
    meta_keys: {label: regex} từ build_meta_keys_from_mapping() — None → dùng META_KEYS.
    _decrypted_bytes: BytesIO đã giải mã (truyền vào nếu file có mật khẩu).
    Trả về: (tables_dict, global_meta, skipped_sheets, warnings)
    tables_dict = { label: {df, type, warnings} }
    """
    # Đọc sheet config tươi mỗi lần parse — phản ánh đúng _CONFIG hiện tại
    # trong mapping file (thêm BOM6, BOM7... → tự nhận, không restart app).
    sheet_cfg = _load_sheet_config()

    # Nguồn dữ liệu: file gốc hoặc BytesIO đã giải mã
    src1 = _decrypted_bytes if _decrypted_bytes is not None else filepath

    wb = load_workbook(src1, read_only=True, data_only=True)

    # Mở lần 2 không data_only để đọc đúng giá trị ô có công thức
    try:
        if _decrypted_bytes is not None:
            _decrypted_bytes.seek(0)
        src2 = _decrypted_bytes if _decrypted_bytes is not None else filepath
        wb_live = load_workbook(src2, read_only=True, data_only=False)
    except Exception:
        wb_live = None

    # Mở lần 3 (không read_only) chỉ để lấy thông tin row bị ẩn
    _hidden_rows_map = {}
    _hidden_cols_map = {}   # sheet_name → set of 0-based col indices bị ẩn
    try:
        if _decrypted_bytes is not None:
            _decrypted_bytes.seek(0)
        src3 = _decrypted_bytes if _decrypted_bytes is not None else filepath
        _wb_hid = load_workbook(src3, read_only=False, data_only=True)
        from openpyxl.utils import column_index_from_string
        for _sname in _wb_hid.sheetnames:
            _ws_hid = _wb_hid[_sname]
            _hidden_rows_map[_sname] = {
                r for r, dim in _ws_hid.row_dimensions.items() if dim.hidden
            }
            _hidden_cols_map[_sname] = {
                column_index_from_string(c) - 1   # chuyển về 0-based
                for c, dim in _ws_hid.column_dimensions.items() if dim.hidden
            }
        _wb_hid.close()
    except Exception:
        _hidden_rows_map = {}
        _hidden_cols_map = {}

    parsed   = []
    skipped  = []
    warnings = []
    global_meta = {}

    for name in wb.sheetnames:
        # Bỏ qua sheet bị ẩn (hidden / veryHidden)
        if wb[name].sheet_state != 'visible':
            continue
        # Lấy 15 dòng đầu từ wb_live (formula-aware) và wb (cached) rồi merge
        live_meta_rows = None
        if wb_live and name in wb_live.sheetnames:
            ws_live = wb_live[name]
            live_raw    = []
            cached_raw  = []
            for i, row in enumerate(ws_live.iter_rows(values_only=True)):
                if i >= 15:
                    break
                live_raw.append(row)
            cached_raw = list(wb[name].iter_rows(min_row=1, max_row=15, values_only=True))
            live_meta_rows = _merge_meta_rows(live_raw, cached_raw)

        res = _parse_sheet(wb[name], name, live_meta_rows=live_meta_rows, meta_keys=meta_keys, hidden_rows=_hidden_rows_map.get(name), hidden_cols=_hidden_cols_map.get(name), sheet_config=sheet_cfg)
        if res:
            parsed.append(res)
            global_meta.update(res["metadata"])
        else:
            skipped.append(name)

    if wb_live:
        wb_live.close()
    wb.close()

    tables = {}

    # Bảng H: Header
    if global_meta:
        tables["[H] Phiếu Header"] = {
            "df": pd.DataFrame([global_meta]), "type": "HEADER", "warnings": []}
    else:
        w = "Không tìm thấy thông tin đầu phiếu (Dự án, Đơn hàng...)"
        tables["[H] Phiếu Header"] = {
            "df": pd.DataFrame([{"Lỗi": w}]), "type": "HEADER", "warnings": [w]}
        warnings.append(w)

    # Bảng BOM — đọc từ sheet_cfg (fresh từ mapping file mỗi lần parse)
    if not sheet_cfg:
        w = "⚠ Không đọc được cấu hình BOM từ mapping file (_CONFIG). Không nhận dạng được section BOM nào."
        warnings.append(w)

    for cfg_item in sheet_cfg:
        section   = cfg_item["section"]
        label     = cfg_item["label"]
        found = next((s for s in parsed if s["sheet_type"] == section), None)
        if found:
            tables[label] = {"df": found["df"], "type": section, "warnings": []}
        else:
            w = f"Không tìm thấy sheet {section}"
            tables[label] = {
                "df": pd.DataFrame([{"Lỗi": w}]), "type": section, "warnings": [w]}
            warnings.append(w)

    return tables, global_meta, skipped, warnings

# ─────────────────────────────────────────────────────────────────────────────
# 3b. THDM EXCEL PARSER  (giống BOM: detect header, map cột theo Ten_Excel)
# ─────────────────────────────────────────────────────────────────────────────

# Fallback anchors khi mapping chưa load — chỉ dùng nếu thdm_map rỗng
_THDM_HEADER_ANCHORS_FALLBACK = [
    "mã vật tư", "tên vật tư", "đvt", "quy cách", "mã vt",
    "ten vat tu", "ma vat tu", "don vi tinh",
    "mã vttb", "tên vt/tb", "ma vttb", "ten vt",
]

# Alias: normalize tên cột viết tắt → norm của Ten_Excel trong mapping
# Vẫn cần cho bước col_map building (khớp header Excel → sql_col)
_THDM_COL_ALIASES = {
    'mavttb':      'mavattu',
    'tenvttb':     'tenvattu',
    'mavt':        'mavattu',
    'tenvt':       'tenvattu',
    'tenvttbmavt': 'tenvattu',
}

def _thdm_find_thvt_sheet(wb):
    """Tìm sheet THVT trong workbook.
    Ưu tiên exact 'TH VT', sau đó flexible contains (normalize tiếng Việt)."""
    if "TH VT" in wb.sheetnames:
        return "TH VT"
    _candidates = ["THVT", "TH_VT", "TONG HOP VAT TU", "TỔNG HỢP VẬT TƯ", "THVAT TU"]
    for sname in wb.sheetnames:
        n = _norm_vn(sname)
        if any(_norm_vn(c) in n for c in _candidates):
            return sname
    return None

def _thdm_find_header_row(all_rows, anchors=None, max_scan=25):
    """
    Tìm dòng header trong sheet THVT.
    anchors: list norm string từ mapping Ten_Excel. Nếu None → dùng fallback.
    """
    _anchors = anchors if anchors else _THDM_HEADER_ANCHORS_FALLBACK
    for i, row in enumerate(all_rows[:max_scan]):
        texts = [_norm_vn(str(c)) for c in row if c is not None]
        hits  = sum(1 for a in _anchors
                    if any(a in t or t in a for t in texts))
        if hits >= 2:
            return i
    return None

def _norm_muc_key(v):
    """Chuẩn hoá Mục number: '1.0'→'1', '1.10'→'1.1', 'Muc 1'→'1'."""
    if v is None:
        return None
    s = str(v).strip().replace('Muc ', '').replace('MUC ', '').strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else f'{f:g}'
    except (ValueError, TypeError):
        return s


def _thdm_load_bom_qty_dict(conn, bom_ids):
    """
    Query B20BOMDetail (BOMDetailType IN 2,3) cho các BOM đã chọn.
    Kết quả: dict {(item_code_str, muc_no_str): dinh_muc}
    - item_code_str : B20Item.Code  (VD: 'NVL9-000464')
    - muc_no_str    : BuiltinOrder0 đã normalize (VD: '1', '1.1')

    ĐỊNH MỨC = SUM(Quantity9 + QuantityFactory + QuantityConstruction + QuantitySubcontractor)
    Lý do: BOMDetailType=2 (CT) dùng Quantity9, BOMDetailType=3 (TP) dùng QuantityFactory/Construction/Subcontractor.
    Cộng tất cả an toàn vì các cột còn lại = NULL (→ 0) theo từng loại.
    """
    if not bom_ids:
        return {}
    ph  = ', '.join('?' * len(bom_ids))
    sql = f"""
        SELECT i.Code,
               dt.BuiltinOrder0,
               SUM(
                   ISNULL(bd.Quantity9, 0)
                   + ISNULL(bd.QuantityFactory, 0)
                   + ISNULL(bd.QuantityConstruction, 0)
                   + ISNULL(bd.QuantitySubcontractor, 0)
               ) AS DinhMuc
        FROM   B20BOMDetail          bd
        JOIN   B20BOM                bom ON bom.Id              = bd.BOMId
        JOIN   B30BizDocDetailSO     dt  ON dt.DetailRowId_SO   = bom.DetailRowId_SO
        JOIN   B20Item               i   ON i.Id                = bd.ItemId
        WHERE  bd.BOMId   IN ({ph})
          AND  bd.BOMDetailType IN (2, 3)
          AND  bom.DetailRowId_SO IS NOT NULL
          AND  bom.DetailRowId_SO != ''
        GROUP BY i.Code, dt.BuiltinOrder0
    """
    cur = conn.cursor()
    cur.execute(sql, list(bom_ids))
    result = {}
    for row in cur.fetchall():
        code     = str(row[0]).strip() if row[0] is not None else None
        muc      = _norm_muc_key(row[1])
        dinh_muc = float(row[2]) if row[2] is not None else 0.0  # Decimal → float
        if code and muc:
            result[(code, muc)] = dinh_muc
    return result


def _thdm_expand_muc_rows(rows, bom_qty_dict=None):
    """
    Transpose CHI TIẾT MỤC: 1 Excel row → N DB rows (1 per Mục có CT hoặc TP > 0).

    Logic (khi có bom_qty_dict):
    - Quantity9 (ĐỊNH MỨC P.KT) lấy từ B20BOMDetail SUM(Qty9) theo (ItemCode, MụcNo)
    - QuantityFactory = Quantity9_DB - CT - TP  (có thể âm nếu CT+TP > ĐỊNH MỨC)
    - Nếu không match trong bom_qty_dict → Quantity9 = None, QuantityFactory = None

    Logic fallback (bom_qty_dict=None — chế độ cũ):
    - NHÀ MÁY = ĐỊNH MỨC - tổng CT - tổng TP
    - QuantityFactory phân bổ theo tỷ lệ (CT_X+TP_X) / tổng(CT+TP)

    Mỗi parsed row phải có:
        _muc_ct: {muc_id: ct_value}   — từ CHI TIẾT GIAO CÔNG TRÌNH
        _muc_tp: {muc_id: tp_value}   — từ CHI TIẾT GIAO THẦU PHỤ
    """
    def _fnum(v):
        """Convert Excel cell value (str/int/float/None/Decimal) → float."""
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    expanded = []
    for row in rows:
        muc_ct   = row.pop('_muc_ct', {}) or {}
        muc_tp   = row.pop('_muc_tp', {}) or {}
        all_mucs = sorted(set(muc_ct) | set(muc_tp))

        if not all_mucs:
            # Không có Mục nào → factory-only, giữ nguyên
            expanded.append(row)
        else:
            # ItemCode từ Excel (Code string, VD: 'NVL9-000464')
            item_code = str(row.get('ItemId') or '').strip()

            # Fallback ratio (dùng khi bom_qty_dict=None)
            dinh_muc   = _fnum(row.get('Quantity9') or row.get('Quantity'))
            total_ct   = sum(_fnum(muc_ct.get(m)) for m in all_mucs)
            total_tp   = sum(_fnum(muc_tp.get(m)) for m in all_mucs)
            nha_may    = dinh_muc - total_ct - total_tp
            total_ctpt = total_ct + total_tp
            fac_distributed = 0.0

            for idx, muc_id in enumerate(all_mucs):
                ct_val  = _fnum(muc_ct.get(muc_id))
                tp_val  = _fnum(muc_tp.get(muc_id))
                is_last = (idx == len(all_mucs) - 1)

                if bom_qty_dict is not None:
                    # ── Chế độ BOM lookup ──────────────────────────────────────
                    muc_key = _norm_muc_key(muc_id)
                    qty9_db = bom_qty_dict.get((item_code, muc_key))
                    if qty9_db is not None:
                        fac_val  = qty9_db - ct_val - tp_val   # cho phép âm
                        qty_muc  = qty9_db
                    else:
                        fac_val  = None
                        qty_muc  = None
                else:
                    # ── Fallback: phân bổ tỷ lệ (chế độ cũ) ──────────────────
                    if is_last:
                        fac_val = max(0, (nha_may or 0) - fac_distributed)
                    elif total_ctpt > 0:
                        fac_val = max(0, round((nha_may or 0) * (ct_val + tp_val) / total_ctpt))
                    else:
                        fac_val = 0
                    fac_distributed += fac_val
                    qty_muc = ct_val + tp_val + fac_val

                new_row = dict(row)
                new_row['DetailRowId_SO']        = muc_id
                new_row['QuantityConstruction']  = ct_val
                new_row['QuantitySubcontractor'] = tp_val
                new_row['QuantityFactory']       = fac_val
                new_row['Quantity']              = qty_muc
                new_row['Quantity9']             = qty_muc
                expanded.append(new_row)
    return expanded


def _thdm_parse_thvt_sheet(wb_cached, wb_live, wb_hid, sheet_name, thdm_map,
                            expand_muc=True, bom_qty_dict=None):
    """
    Parse sheet TH VT giống BOM:
    - Detect header động (không hardcode dòng 12)
    - Map cột theo Ten_Excel từ thdm_map
    - Xử lý formula, hidden rows/cols
    - expand_muc=True: expand 1 row → N rows per Mục (THDM_THVT)
    - expand_muc=False: giữ nguyên rows, không expand (THDM_NVL...)
    Trả về: list of dict {sql_col: value}
    """
    from openpyxl.utils import column_index_from_string

    # Hidden rows / cols
    _hr, _hc = set(), set()
    if wb_hid and sheet_name in wb_hid.sheetnames:
        _ws_h = wb_hid[sheet_name]
        _hr   = {r for r, d in _ws_h.row_dimensions.items()    if d.hidden}
        _hc   = {column_index_from_string(c) - 1
                 for c, d in _ws_h.column_dimensions.items() if d.hidden}

    def _fcols(row):
        return tuple(v for i, v in enumerate(row) if i not in _hc) if _hc else row

    # Đọc cached (data_only=True)
    ws_c    = wb_cached[sheet_name]
    c_rows  = [_fcols(vals)
               for idx, vals in enumerate(ws_c.iter_rows(values_only=True), 1)
               if idx not in _hr]

    # Merge với live (formula resolution)
    if wb_live and sheet_name in wb_live.sheetnames:
        ws_l   = wb_live[sheet_name]
        l_rows = [_fcols(vals)
                  for idx, vals in enumerate(ws_l.iter_rows(values_only=True), 1)
                  if idx not in _hr]
        all_rows = _merge_meta_rows(l_rows, c_rows)
    else:
        all_rows = c_rows

    if not all_rows:
        return []

    # Build anchors từ mapping Ten_Excel (mapping-driven, như BOM)
    mapping_anchors = [
        _norm_vn(r['ten_excel'])
        for r in thdm_map
        if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')
    ]

    # Tìm dòng header
    hidx = _thdm_find_header_row(all_rows, anchors=mapping_anchors or None)
    if hidx is None:
        # Gợi ý: liệt kê các tên cột tìm thấy ở vùng đầu để dễ debug
        found_hdrs = []
        for row in all_rows[:10]:
            row_texts = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if row_texts:
                found_hdrs.append(" | ".join(row_texts[:8]))
        hint = "\n".join(found_hdrs[:5]) if found_hdrs else "(sheet trống)"
        expected = ", ".join(r['ten_excel'] for r in thdm_map
                             if r.get('nguon_dl') == 'Excel' and r.get('ten_excel'))[:120]
        raise ValueError(
            f"Không tìm thấy dòng header trong sheet '{sheet_name}'.\n\n"
            f"Cột mapping cần khớp: {expected}\n\n"
            f"10 dòng đầu tìm thấy:\n{hint}"
        )

    # Build headers (hỗ trợ 2-dòng merged header giống BOM)
    r1 = all_rows[hidx]
    r2 = all_rows[hidx + 1] if hidx + 1 < len(all_rows) else ()
    has_sub = bool(r2) and r2[0] is None and any(
        c is not None and str(c).strip() not in ("", "0") for c in (r2[4:] if r2 else []))
    if has_sub:
        headers    = _build_headers(r1, r2)
        data_start = hidx + 2
    else:
        headers    = [_norm_col(c) if c else f"COL_{i}" for i, c in enumerate(r1)]
        data_start = hidx + 1

    # Build col_map: {col_index → sql_col} — dùng shared helper + _THDM_COL_ALIASES
    col_map = _build_excel_col_map(headers, thdm_map, col_aliases=_THDM_COL_ALIASES)

    # Tìm vị trí cột STT
    stt_idx = next(
        (i for i, h in enumerate(headers)
         if _norm_vn(h.replace('_', ' ')).startswith('stt')),
        None)

    # ── Detect CHI TIẾT GIAO CT / TP Mục columns từ raw header rows ─────────
    r1_raw = all_rows[hidx]
    r2_raw = all_rows[hidx + 1] if has_sub else ()

    def _detect_chi_tiet_cols(r1, r2, kw_norm):
        """Scan r1/r2 tìm super-header chứa kw_norm → {col_idx: muc_id}."""
        kw_norm  = _norm_vn(kw_norm)   # FIX: normalize keyword (strip spaces/diacritics)
        muc_cols = {}
        in_sect  = False
        n        = max(len(r1), len(r2) if r2 else 0)
        r1p      = list(r1) + [None] * max(0, n - len(r1))
        r2p      = list(r2) + [None] * max(0, n - len(r2)) if r2 else [None] * n
        for i in range(n):
            v1 = r1p[i]
            v2 = r2p[i]
            if v1 is not None and str(v1).strip():
                if kw_norm in _norm_vn(str(v1).lower()):
                    in_sect = True
                elif in_sect:
                    break   # next non-empty super-header → kết thúc section
            if in_sect and v2 and str(v2).strip().lower().startswith('muc '):
                muc_id = str(v2).strip()[4:].strip()   # "Muc 1.1" → "1.1"
                if muc_id:
                    muc_cols[i] = muc_id
        return muc_cols

    ct_muc_cols = _detect_chi_tiet_cols(r1_raw, r2_raw, 'giao cong trinh')
    tp_muc_cols = _detect_chi_tiet_cols(r1_raw, r2_raw, 'giao thau phu')

    # ── Build extra_col_extractor cho _parse_section_excel_rows ──────────────
    if ct_muc_cols or tp_muc_cols:
        def _muc_extractor(padded):
            ct = {mid: padded[ci] for ci, mid in ct_muc_cols.items()
                  if ci < len(padded) and padded[ci] and padded[ci] != 0}
            tp = {mid: padded[ci] for ci, mid in tp_muc_cols.items()
                  if ci < len(padded) and padded[ci] and padded[ci] != 0}
            return {'_muc_ct': ct, '_muc_tp': tp}
    else:
        _muc_extractor = None

    # Parse rows — shared pipeline (empty filter, footer, section header, Fill_Forward, no-item filter)
    rows = _parse_section_excel_rows(
        all_rows[data_start:], headers, col_map, thdm_map, stt_idx,
        skip_no_item=True, item_cols=('ItemId', 'ItemName'),
        ff_always_override=True,   # ItemType luôn lấy section letter (A, D, E1...) từ header
        extra_col_extractor=_muc_extractor,
        roman_as_hdr=True,         # I, II... trong TH VT là tên section, không phải số La Mã
    )

    # Fix col_map collision: "Định mức P.KT" map vào Quantity9 trước (first-match)
    # → Quantity không được map → copy Quantity9 → Quantity cho mỗi row
    for r in rows:
        if r.get('Quantity') is None and r.get('Quantity9') is not None:
            r['Quantity'] = r['Quantity9']

    # ── Expand Mục rows: 1 Excel row → N DB rows (CHI TIẾT GIAO CT/TP) ──────
    if expand_muc and (ct_muc_cols or tp_muc_cols):
        rows = _thdm_expand_muc_rows(rows, bom_qty_dict=bom_qty_dict)

    return rows


def _thdm_open_workbook(path):
    """Mở file THDM 3 lần (cached / live / hid) giống parse_bom_excel."""
    from openpyxl import load_workbook
    wb_c = load_workbook(path, read_only=True,  data_only=True)
    try:
        wb_l = load_workbook(path, read_only=True,  data_only=False)
    except Exception:
        wb_l = None
    try:
        wb_h = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        wb_h = None
    return wb_c, wb_l, wb_h


def _thdm_apply_row_filter(rows, filter_str):
    """
    Lọc rows theo expression đơn giản: 'ColName>0', 'ColName>=0.01', v.v.
    Chỉ hỗ trợ: >=, <=, !=, >, <, == (1 điều kiện).
    """
    import re as _re_flt
    m = _re_flt.match(r'^(\w+)\s*(>=|<=|!=|>|<|==)\s*(.+)$', filter_str.strip())
    if not m:
        return rows
    col, op, val_str = m.group(1), m.group(2), m.group(3).strip()
    try:
        val = float(val_str)
    except ValueError:
        val = val_str

    def _check(row):
        rv = row.get(col)
        if rv is None:
            return False
        try:
            rv = float(rv)
        except (TypeError, ValueError):
            pass
        if op == '>':  return rv > val
        if op == '<':  return rv < val
        if op == '>=': return rv >= val
        if op == '<=': return rv <= val
        if op == '==': return rv == val
        if op == '!=': return rv != val
        return True

    return [r for r in rows if _check(r)]


def _thdm_get_detail_col_defs(section_map):
    """Col defs cho tab detail THDM — hiển thị TẤT CẢ cột của section.
    - Cột định danh = SQL_Column (duy nhất) → header text = Ten_Excel,
      hàng phụ (sql_names) = SQL_Column.
    - Ưu tiên: cột có Ten_Excel xếp lên trước, giữ nguyên thứ tự mapping
      trong từng nhóm; cột không có Ten_Excel xếp sau.
    """
    with_te, without_te = [], []
    for r in section_map:
        sql_col = (r.get('sql_col') or '').strip()
        if not sql_col:
            continue
        ten_excel = (r.get('ten_excel') or '').strip()
        nguon     = r.get('nguon_dl') or ''
        kieu      = str(r.get('kieu_dl') or '').lower()
        is_num    = any(t in kieu for t in ('decimal', 'float', 'int', 'money', 'numeric'))
        try:
            do_dai = int(float(r.get('do_dai') or 0))
        except (ValueError, TypeError):
            do_dai = 0
        anchor = 'e' if is_num else guess_col_align(ten_excel or sql_col)
        if is_num:        px = 130
        elif do_dai <= 20:  px = 120
        elif do_dai <= 50:  px = 150
        elif do_dai <= 100: px = 180
        else:               px = 200
        # cột hệ thống (không Excel) hẹp hơn cho đỡ chiếm chỗ
        if not ten_excel:
            px = min(px, 120)
        d = {
            'sql_col':   sql_col,
            'ten_excel': ten_excel,
            'nguon_dl':  nguon,
            'mac_dinh':  (r.get('mac_dinh') or '').strip(),
            'width':     px,
            'anchor':    anchor,
            'stretch':   bool(ten_excel),
        }
        (with_te if ten_excel else without_te).append(d)
    return with_te + without_te


# ─────────────────────────────────────────────────────────────────────────────
# 3c. SHARED ROW RESOLVER — dùng chung BOM và THDM
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_row_mapping(mapping_recs, ctx, get_excel_val=None):
    """
    Resolve một row từ mapping records theo NguonDL — dùng chung BOM và THDM.

    mapping_recs : list of mapping record (HEADER/DETAIL section)
    ctx          : dict chứa tất cả nguồn dữ liệu:
        now           : datetime  — dùng cho HeThong NOW
        builtin_order : int       — AUTO_INC sort order
        doc_id        : str/uuid  — BOMId / BizDocId placeholder
        parent_row    : dict      — resolved parent row (HEADER / THDM_HEADER)
        parent_fields : set       — sql_col copy từ parent (HeThong)
        ui_values     : dict      — {'creator': ..., 'product_id': ..., 'order_id': ...}
        skip_nguon    : set       — skip hoàn toàn (mặc định: {'SP', 'TinhToan'})
    get_excel_val : callable(rec) → value
        THDM: lambda rec: excel_row.get(rec['sql_col'])
        BOM:  existing complex lookup — truyền None để skip Excel fields
    Returns: dict {sql_col: value}
    """
    now           = ctx.get('now')
    parent_row    = ctx.get('parent_row') or {}
    parent_fields = ctx.get('parent_fields') or set()
    doc_id        = ctx.get('doc_id')
    builtin_order = ctx.get('builtin_order', 1)
    ui_values     = ctx.get('ui_values') or {}
    skip_nguon    = ctx.get('skip_nguon', {'SP', 'TinhToan'})

    row_out = {}
    for rec in mapping_recs:
        sql_col = (rec.get('sql_col') or '').strip()
        if not sql_col:
            continue
        nguon   = (rec.get('nguon_dl') or 'Excel').strip()
        mac     = str(rec.get('mac_dinh') or '').strip()
        kieu_dl = str(rec.get('kieu_dl') or '').lower()

        # ── Skip ─────────────────────────────────────────────────────────────
        if nguon in skip_nguon:
            row_out.setdefault(sql_col, None)
            continue

        # ── Excel ─────────────────────────────────────────────────────────────
        if nguon == 'Excel':
            if get_excel_val is not None:
                val = get_excel_val(rec)
                _is_empty = (val is None
                             or (isinstance(val, float) and math.isnan(val))
                             or (isinstance(val, str) and val.strip() == ''))
                if _is_empty:
                    val = None
                    mac_upper = mac.upper()
                    if mac_upper == 'EMPTY':
                        val = None if kieu_dl in ('date', 'datetime') else ''
                    elif mac_upper == 'NOW':
                        val = ctx.get('now')
                    elif mac_upper == 'CREATOR':
                        val = (ctx.get('ui_values') or {}).get('creator')
                    elif mac and mac_upper not in ('NULL', 'NONE', ''):
                        try:    val = int(mac)
                        except (ValueError, TypeError):
                            try:    val = float(mac)
                            except (ValueError, TypeError): val = mac
                row_out[sql_col] = val
            continue

        # ── CoDinh ────────────────────────────────────────────────────────────
        if nguon == 'CoDinh':
            if not mac or mac.upper() == 'NULL':
                row_out[sql_col] = None
            elif mac.upper() == 'EMPTY':
                row_out[sql_col] = None if kieu_dl in ('date', 'datetime') else ''
            else:
                try:    row_out[sql_col] = int(mac)
                except (ValueError, TypeError):
                    try:    row_out[sql_col] = float(mac)
                    except (ValueError, TypeError): row_out[sql_col] = mac
            continue

        # ── HeThong ───────────────────────────────────────────────────────────
        if nguon == 'HeThong':
            if mac == 'NOW' or sql_col in ('CreatedAt', 'ModifiedAt'):
                row_out[sql_col] = now
            elif mac == 'AUTO_INC':
                row_out[sql_col] = builtin_order
            elif sql_col in ('BOMId', 'BizDocId') or mac == 'NEWID':
                row_out[sql_col] = doc_id
            elif sql_col in parent_fields:
                row_out[sql_col] = parent_row.get(sql_col)
            elif mac and mac in parent_row:
                # mac_dinh = tên field của parent → copy giá trị
                row_out[sql_col] = parent_row[mac]
            elif mac and mac not in ('NULL', 'EMPTY', 'NOW', 'AUTO_INC', 'NEWID'):
                row_out[sql_col] = mac   # giá trị cố định (fallback)
            else:
                row_out[sql_col] = None
            continue

        # ── UILookup ──────────────────────────────────────────────────────────
        if nguon == 'UILookup':
            row_out[sql_col] = ui_values.get(mac)
            continue

        # ── MucLookup: đọc raw Muc key từ Excel row (đã set bởi _thdm_expand_muc_rows)
        # Giá trị ban đầu = muc_id ('1', '1.1'...), sẽ được replace bằng DB id
        # trong _thdm_insert_worker sau khi compound lookup (BuiltinOrder0|BizDocId)
        if nguon == 'MucLookup':
            if get_excel_val is not None:
                val = get_excel_val(rec)
                row_out[sql_col] = val
            else:
                row_out[sql_col] = None
            continue

    return row_out



