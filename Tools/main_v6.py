"""
main_v6.py — BOHO BOM2BRAVO
UI layer: CustomTkinter + tksheet
Version: xem APP_VERSION bên dưới (nguồn version duy nhất — title bar,
log DB và bộ build đều theo số này; khi phát hành nhớ cập nhật
VERSION trong build.bat + AppVersion trong setup.iss cho khớp).
"""
APP_VERSION = "2.1"

# Tên 3 tab chính — nguồn duy nhất, dùng ở mọi chỗ thay vì literal string /
# substring match. Trước đây _on_tab_changed() check `"THDM" in tab_name`,
# khi đổi TAB_THDM sang "Tổng hợp định mức" thì chuỗi không còn chứa "THDM"
# nữa nên auto-reload khi switch tab bị vô hiệu âm thầm — đổi tên tab từ giờ
# chỉ cần sửa đúng 1 dòng dưới đây.
TAB_IMPORT  = "📥  Import BOM"
TAB_THDM    = "📊  Tổng hợp định mức"
TAB_CATALOG = "🌳  Tra cứu Danh mục"

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
try:
    import tksheet
    _HAS_TKSHEET = True
except ImportError:
    _HAS_TKSHEET = False
import customtkinter as ctk
import re, os, sys, threading, datetime, unicodedata, io
from openpyxl import load_workbook
try:
    import msoffcrypto
    _HAS_MSOFFCRYPTO = True
except ImportError:
    _HAS_MSOFFCRYPTO = False

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


# ── Compat wrappers: business logic gọi .config(fg=, bg=) như tkinter thường ─
class CLabel(ctk.CTkLabel):
    def config(self, **kw):
        mp = {}
        for k, v in kw.items():
            if k in ("fg", "foreground"):   mp["text_color"] = v
            elif k in ("bg", "background"): mp["fg_color"]   = v
            else:                           mp[k]            = v
        super().configure(**mp)

class CButton(ctk.CTkButton):
    def config(self, **kw):
        mp = {}
        for k, v in kw.items():
            if k in ("fg", "foreground"):   mp["text_color"] = v
            elif k in ("bg", "background"): mp["fg_color"]   = v
            else:                           mp[k]            = v
        super().configure(**mp)


# ── SearchCombo: Entry + popup Listbox có ô tìm kiếm (kiểu BRAVO) ─────────────
class _SearchCombo(ctk.CTkFrame):
    """Searchable dropdown: hiển thị Entry, click/nhập mở popup Listbox có search."""

    _BG_DARK  = "#2B2B2B"
    _BG_ENTRY = "#3C3C3C"
    _FG       = "#DCDCDC"
    _SEL_BG   = "#1D4ED8"
    _SEL_FG   = "#FFFFFF"
    _FONT     = ("Segoe UI", 11)

    # Combo đang mở popup (class-level) — mở combo mới thì đóng combo cũ
    _open_combo = None

    def __init__(self, master, values=None, width=200, height=32,
                 font=None, command=None, placeholder="—", state="normal",
                 popup_width=None, **kw):
        super().__init__(master, fg_color="transparent",
                         width=width, height=height, **kw)
        self.pack_propagate(False)
        self._width      = width
        self._height     = height
        self._popup_width = popup_width  # None = auto (max widget_width, 480)
        self._font       = font or ctk.CTkFont("Segoe UI", 11)
        self._command    = command
        self._placeholder = placeholder
        self._all_values: list[str] = list(values or [])
        self._popup: tk.Toplevel | None = None
        self._listbox: tk.Listbox | None = None
        self._pop_search_var = tk.StringVar()
        self._state  = state          # "normal" | "disabled"
        self._value  = tk.StringVar(value=placeholder)

        # ── Display entry (read-only, click → popup) ──
        self._entry = ctk.CTkEntry(
            self, textvariable=self._value,
            font=self._font, height=height,
            corner_radius=6, width=width - 30,
            state="readonly" if state == "normal" else "disabled")
        self._entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._entry.bind("<Button-1>", self._toggle_popup)

        # ── Arrow button ──
        self._btn = ctk.CTkButton(
            self, text="▾", width=28, height=height,
            corner_radius=6,
            fg_color=("gray78", "gray28"),
            hover_color=("gray68", "gray38"),
            font=ctk.CTkFont("Segoe UI", 10),
            command=self._toggle_popup,
            state=state)
        self._btn.pack(side=tk.LEFT)

    # ── Public API (compatible với CTkComboBox) ───────────────────────────────
    def get(self) -> str:
        return self._value.get()

    def set(self, value: str):
        self._value.set(value)

    def configure(self, **kw):
        if "values" in kw:
            self._all_values = list(kw.pop("values"))
        if "state" in kw:
            self._state = kw.pop("state")
            entry_state = "readonly" if self._state == "normal" else "disabled"
            self._entry.configure(state=entry_state)
            self._btn.configure(state=self._state)
        # CTkFrame không nhận các kw lạ, bỏ qua
        allowed = {k: v for k, v in kw.items()
                   if k in ("fg_color", "border_color", "border_width",
                             "corner_radius", "width", "height")}
        if allowed:
            super().configure(**allowed)

    # ── Popup ─────────────────────────────────────────────────────────────────
    def _toggle_popup(self, _event=None):
        if self._state == "disabled":
            return
        if self._popup and self._popup.winfo_exists():
            self._close_popup()
        else:
            self._open_popup()

    def _open_popup(self):
        if self._popup and self._popup.winfo_exists():
            return

        # Đóng popup của combo khác đang mở (tránh 2 dropdown chèn nhau)
        other = _SearchCombo._open_combo
        if other is not None and other is not self:
            try:
                other._close_popup()
            except Exception:
                pass
        _SearchCombo._open_combo = self

        self.update_idletasks()
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height() + 2
        w = self._popup_width if self._popup_width else max(self.winfo_width(), self._width, 600)

        pop = tk.Toplevel(self)
        pop.overrideredirect(True)
        pop.configure(bg=self._BG_DARK)
        pop.geometry(f"{w}x280+{x}+{y}")
        pop.lift()
        pop.attributes("-topmost", True)
        self._popup = pop

        # Search entry
        self._pop_search_var.set("")
        search_entry = tk.Entry(
            pop, textvariable=self._pop_search_var,
            bg=self._BG_ENTRY, fg=self._FG,
            font=self._FONT, insertbackground=self._FG,
            relief="flat", bd=6)
        search_entry.pack(fill=tk.X, padx=4, pady=(4, 2))
        search_entry.focus_set()
        self._pop_search_var.trace_add("write", self._on_search_change)

        # Listbox + scrollbar
        frm = tk.Frame(pop, bg=self._BG_DARK)
        frm.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))
        vsb = tk.Scrollbar(frm, orient=tk.VERTICAL, bg=self._BG_DARK)
        lb  = tk.Listbox(
            frm, bg=self._BG_DARK, fg=self._FG,
            selectbackground=self._SEL_BG, selectforeground=self._SEL_FG,
            font=self._FONT, relief="flat", bd=0,
            activestyle="none",
            yscrollcommand=vsb.set)
        vsb.configure(command=lb.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._listbox = lb

        self._fill_listbox(self._all_values)

        lb.bind("<<ListboxSelect>>", self._on_listbox_select)
        lb.bind("<Return>",          self._on_listbox_select)
        lb.bind("<Escape>",          lambda e: self._close_popup())
        search_entry.bind("<Return>",  self._on_listbox_select)
        search_entry.bind("<Escape>",  lambda e: self._close_popup())
        search_entry.bind("<Down>",    lambda e: lb.focus_set())
        lb.bind("<Up>", lambda e: search_entry.focus_set()
                if lb.curselection() and lb.curselection()[0] == 0 else None)

        # Đóng khi click ngoài hoặc app mất focus (kể cả switch sang app khác)
        pop.bind("<FocusOut>", self._on_popup_focus_out)
        self._poll_focus()

    def _poll_focus(self):
        """Poll 250ms — đóng popup nếu main window mất focus (switch sang app khác)."""
        if not (self._popup and self._popup.winfo_exists()):
            return
        try:
            has_focus = self.winfo_toplevel().focus_displayof() is not None
        except Exception:
            has_focus = True  # nếu không xác định được thì giữ nguyên
        if not has_focus:
            self._close_popup()
            return
        self.after(250, self._poll_focus)

    def _on_search_change(self, *_):
        kw = self._pop_search_var.get().strip().lower()
        if kw:
            filtered = [v for v in self._all_values if kw in v.lower()]
        else:
            filtered = self._all_values
        self._fill_listbox(filtered)

    def _fill_listbox(self, values: list[str]):
        if not (self._listbox and self._popup and self._popup.winfo_exists()):
            return
        self._listbox.delete(0, tk.END)
        for v in values:
            self._listbox.insert(tk.END, v)

    def _on_listbox_select(self, event=None):
        if not self._listbox:
            return
        sel = self._listbox.curselection()
        if not sel:
            if self._listbox.size() > 0:
                idx = 0
            else:
                return
        else:
            idx = sel[0]
        value = self._listbox.get(idx)
        self._value.set(value)
        self._close_popup()
        if self._command:
            self._command(value)

    def _on_popup_focus_out(self, event=None):
        # Delay nhỏ để click vào listbox kịp register trước khi đóng
        self.after(200, self._maybe_close)

    def _maybe_close(self):
        if not (self._popup and self._popup.winfo_exists()):
            return
        try:
            fw = self._popup.focus_get()
        except Exception:
            fw = None
        if fw is None:
            self._close_popup()

    def _close_popup(self):
        if self._popup:
            try:
                self._popup.destroy()
            except Exception:
                pass
            self._popup  = None
            self._listbox = None
        if _SearchCombo._open_combo is self:
            _SearchCombo._open_combo = None


# ── Tooltip ────────────────────────────────────────────────────────────────────
class Tooltip:
    def __init__(self, widget, get_text):
        self._w, self._get, self._tip = widget, get_text, None
        widget.bind("<Enter>",       self._show, add="+")
        widget.bind("<Leave>",       self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")

    def _show(self, _=None):
        text = self._get()
        if not text: return
        self._hide()
        x = self._w.winfo_rootx()
        y = self._w.winfo_rooty() + self._w.winfo_height() + 6
        self._tip = tk.Toplevel(self._w)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        tk.Label(self._tip, text=text,
                 bg="#252526", fg="#9CDCFE",
                 font=("Segoe UI", 12),
                 relief="flat", padx=10, pady=5,
                 wraplength=700, justify="left").pack()

    def _hide(self, _=None):
        if self._tip:
            try: self._tip.destroy()
            except: pass
            self._tip = None


# ── SheetTable: tksheet với API tối giản của ttk.Treeview ────────────────────
class SheetTable:
    """Bảng dữ liệu dùng tksheet (grid kiểu Excel) nhưng giữ API tối giản của
    ttk.Treeview show='headings' — code đổ dữ liệu cũ (insert / delete /
    heading / column / tag_configure / see / selection_set / set / exists)
    chạy được không đổi. Render được gộp qua after_idle nên insert hàng
    nghìn dòng trong vòng lặp không bị chậm.
    Chỉ khởi tạo khi _HAS_TKSHEET=True; ngược lại dùng ttk.Treeview như cũ."""

    _ANCHOR_MAP = {"w": "w", "e": "e", "center": "center", "c": "center"}

    def __init__(self, master, columns=(), **_kw):
        self.sheet = tksheet.Sheet(
            master,
            headers=[str(c) for c in columns],
            data=[],
            theme="dark",
            show_row_index=False,
            show_top_left=False,
            row_height=24,
            header_height=30,
            show_horizontal_grid=True,
            show_vertical_grid=True,
            font=("Segoe UI", 10, "normal"),
            header_font=("Segoe UI", 10, "bold"),
        )
        self.sheet.enable_bindings(
            "single_select", "drag_select", "row_select", "column_select",
            "column_width_resize", "arrowkeys", "copy",
            "right_click_popup_menu",
        )
        self._columns      = [str(c) for c in columns]
        self._headings     = {c: c for c in self._columns}
        self._col_widths   = {}    # col → width (áp trong _flush)
        self._col_aligns   = {}    # col → 'w'/'e'/'center'
        self._applied_cols = list(self._columns)   # bộ cột đã áp vào sheet
        self._style_dirty  = False # có width/align mới chưa áp
        self._tags      = {}     # tag → {'bg':…, 'fg':…}
        self._rows      = []     # thứ tự iid
        self._data      = {}     # iid → list[str]
        self._row_tags  = {}     # iid → tag đầu tiên
        self._center_tag = None  # tag của các dòng cần căn giữa (vd sql_names)
        self._next_iid  = 0
        self._pending   = False

    def center_rows_by_tag(self, tag):
        """Căn giữa mọi dòng mang tag này (vd hàng phụ SQL_Column),
        không ảnh hưởng căn lề cột của các dòng dữ liệu khác."""
        self._center_tag = tag
        self._schedule()

    # ── Geometry manager / event delegation ──────────────────────────────────
    def grid(self, **kw):        self.sheet.grid(**kw)
    def pack(self, **kw):        self.sheet.pack(**kw)
    def place(self, **kw):       self.sheet.place(**kw)
    def bind(self, *a, **k):     self.sheet.MT.bind(*a, **k)
    def after(self, *a, **k):    return self.sheet.after(*a, **k)
    def after_cancel(self, *a):  return self.sheet.after_cancel(*a)

    # ── API tương thích Treeview ──────────────────────────────────────────────
    def __setitem__(self, key, value):
        if key == "columns":
            self._columns    = [str(c) for c in value]
            self._headings   = {c: c for c in self._columns}
            self._col_widths = {}
            self._col_aligns = {}
            self._schedule()

    def __getitem__(self, key):
        if key == "columns":
            return tuple(self._columns)
        raise KeyError(key)

    def heading(self, col, text=None, anchor=None, **_kw):
        if text is not None and col in self._headings:
            self._headings[col] = text
            self._schedule()

    def column(self, col, width=None, anchor=None, **_kw):
        if col not in self._columns:
            return
        if width is not None:
            self._col_widths[col] = int(width)
            self._style_dirty = True
        if anchor is not None:
            self._col_aligns[col] = self._ANCHOR_MAP.get(str(anchor), "w")
            self._style_dirty = True
        self._schedule()

    def tag_configure(self, tag, background=None, foreground=None, **_kw):
        self._tags[tag] = {"bg": background, "fg": foreground}

    def insert(self, _parent, index, values=(), tags=(), **_kw):
        iid = f"r{self._next_iid}"
        self._next_iid += 1
        vals = ["" if v is None else str(v) for v in values]
        if index in (0, "0"):
            self._rows.insert(0, iid)
        else:
            self._rows.append(iid)
        self._data[iid] = vals
        if tags:
            self._row_tags[iid] = (tags[0] if isinstance(tags, (list, tuple))
                                   else str(tags))
        self._schedule()
        return iid

    def delete(self, *iids):
        if not iids:
            return
        drop = set(iids)
        self._rows = [i for i in self._rows if i not in drop]
        for i in drop:
            self._data.pop(i, None)
            self._row_tags.pop(i, None)
        self._schedule()

    def get_children(self, _item=None):
        return tuple(self._rows)

    def exists(self, iid):
        return iid in self._data

    def set(self, iid, col):
        vals = self._data.get(iid)
        if vals is None or col not in self._columns:
            return ""
        idx = self._columns.index(col)
        return vals[idx] if idx < len(vals) else ""

    def item(self, iid, option=None, **_kw):
        vals = self._data.get(iid, [])
        if option == "values":
            return tuple(vals)
        return {"values": tuple(vals)}

    def selection_set(self, iid):
        self._flush()
        if iid in self._data:
            try:
                self.sheet.select_row(self._rows.index(iid))
            except Exception:
                pass

    def selection(self):
        self._flush()
        try:
            return tuple(self._rows[r]
                         for r in sorted(self.sheet.get_selected_rows())
                         if r < len(self._rows))
        except Exception:
            return ()

    def see(self, iid):
        self._flush()
        if iid in self._data:
            try:
                self.sheet.see(row=self._rows.index(iid), column=0)
            except Exception:
                pass

    def focus(self, _iid=None):
        return None

    def config(self, **_kw):
        pass
    configure = config

    def yview(self, *_a):
        pass

    def xview(self, *_a):
        pass

    # ── Render gộp ────────────────────────────────────────────────────────────
    def _schedule(self):
        if not self._pending:
            self._pending = True
            self.sheet.after_idle(self._flush)

    def _flush(self):
        if not self._pending:
            return
        self._pending = False
        ncols = len(self._columns)
        data  = []
        for iid in self._rows:
            vals = self._data[iid]
            # Pad/cắt cho khớp số cột — tksheet tạo cột theo độ dài row
            if len(vals) < ncols:
                vals = vals + [""] * (ncols - len(vals))
            elif len(vals) > ncols:
                vals = vals[:ncols]
            data.append(vals)
        # Bộ cột thay đổi (VD: Raw Excel gán cột động sau khi đọc file)
        # → phải reset col positions, nếu không tksheet giữ layout cột cũ.
        # So thêm với total_columns() thực tế: bảng được gán cột lúc CHƯA có
        # dữ liệu (VD: tab "Đã xử lý") thì sheet đang có 0 cột dù
        # _applied_cols đã đúng — vẫn phải reset.
        cols_changed = self._applied_cols != self._columns
        try:
            if self.sheet.total_columns() != ncols:
                cols_changed = True
        except Exception:
            pass
        try:
            self.sheet.set_sheet_data(data,
                                      reset_col_positions=cols_changed,
                                      redraw=False)
        except Exception:
            self.sheet.set_sheet_data(data)
        try:
            self.sheet.headers([self._headings.get(c, c)
                                for c in self._columns])
        except Exception:
            pass
        # Lưới an toàn: set_sheet_data/headers đôi khi không co hết cột khi
        # chuyển từ bảng nhiều cột → ít cột (để lại cột trống O,P,Q...).
        # Xoá thẳng các cột dư phía sau.
        if ncols > 0:
            try:
                extra = self.sheet.total_columns() - ncols
                if extra > 0:
                    self.sheet.del_columns(
                        list(range(ncols, ncols + extra)),
                        redraw=False)
            except Exception:
                pass
        if cols_changed or self._style_dirty:
            for col, w in self._col_widths.items():
                if col in self._columns:
                    try:
                        self.sheet.column_width(
                            column=self._columns.index(col), width=w)
                    except Exception:
                        pass
            for col, al in self._col_aligns.items():
                if col in self._columns:
                    try:
                        self.sheet.align_columns(
                            columns=[self._columns.index(col)], align=al)
                    except Exception:
                        pass
            self._style_dirty = False
        self._applied_cols = list(self._columns)
        try:
            self.sheet.dehighlight_all()
        except Exception:
            pass
        by_tag = {}
        for idx, iid in enumerate(self._rows):
            t = self._row_tags.get(iid)
            if t in self._tags:
                by_tag.setdefault(t, []).append(idx)
        for t, idxs in by_tag.items():
            cfg = self._tags[t]
            if not (cfg.get("bg") or cfg.get("fg")):
                continue
            try:
                self.sheet.highlight_rows(rows=idxs, bg=cfg.get("bg"),
                                          fg=cfg.get("fg"), redraw=False)
            except Exception:
                pass
        # Căn giữa các dòng mang center-tag (vd hàng phụ SQL_Column) — ghi đè
        # căn lề cột, không đụng các dòng dữ liệu khác
        if self._center_tag is not None:
            c_idx = [i for i, iid in enumerate(self._rows)
                     if self._row_tags.get(iid) == self._center_tag]
            if c_idx:
                try:
                    self.sheet.align_rows(rows=c_idx, align='center', redraw=False)
                except Exception:
                    pass
        try:
            self.sheet.refresh()
        except Exception:
            pass


def get_base_dir():
    if getattr(sys, 'frozen', False):          # Đang chạy là .exe (PyInstaller)
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR     = get_base_dir()
CONFIG_DIR   = os.path.join(BASE_DIR, "config")
MAPPING_FILE   = os.path.join(CONFIG_DIR, "CK_Mapping_v5.xlsx")
DB_CONFIG_FILE = os.path.join(CONFIG_DIR, "db_config.json")

# UserId mặc định dùng cho CreatedBy khi nhân viên được chọn không có tài
# khoản tương ứng trong B00UserList (EmployeeId không map được sang Id user).
DEFAULT_CREATOR_USER_ID = 823

os.makedirs(CONFIG_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# 2. MAPPING LOADER — đọc CK_Mapping_v2.xlsx (format mới, mỗi sheet = 1 section)
# ─────────────────────────────────────────────────────────────────────────────

def _norm_vn(s):
    """Chuẩn hóa tiếng Việt: bỏ dấu, lowercase, giữ chữ/số.
    'đ' (U+0111) không decompose bằng NFKD — xử lý riêng."""
    s = str(s).lower().strip()
    s = s.replace('đ', 'd')
    nfkd = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r'[^a-z0-9]', '', s)

# Cột mapping chuẩn (14 cột, không có STT)
_MAPPING_ENG_COLS = [
    "SQL_Column", "Ten_Excel", "Kieu_DL", "Do_dai",
    "Bat_buoc", "Mac_dinh", "Nguon_DL",
    "Bang_Master", "Dieu_kien_Master", "Kieu_Lookup",
    "Truong_So_Sanh", "Truong_Lay_Ve", "Ghi_chu",
    "Fill_Forward",
]
# HEADER/DETAIL sheet có thêm cột Section ở đầu (15 cột)
_SECTION_ENG_COLS = ["Section"] + _MAPPING_ENG_COLS


def _nan_str(v, default=''):
    """Chuyển pandas NaN / None thành chuỗi rỗng."""
    import math as _math
    if v is None:
        return default
    if isinstance(v, float) and _math.isnan(v):
        return default
    s = str(v).strip()
    return default if s.lower() == 'nan' else s


# ─── Helper readers cho CK_Mapping_v5 (7 sheets) ───────────────────────────

def _load_config(xl):
    """Đọc _CONFIG → {section: {label, view_insert, sheet_contains, ...}}"""
    if '_CONFIG' not in xl.sheet_names:
        return {}
    df = pd.read_excel(xl, sheet_name='_CONFIG', header=None, skiprows=3)
    COLS = ['Section', 'Label', 'View_Insert', 'SheetNameContains',
            'SheetNameExclude', 'DataStartRow', 'ParentSection',
            'ExpandMuc', 'RowFilter']
    for i, c in enumerate(COLS):
        if i < len(df.columns):
            df.rename(columns={df.columns[i]: c}, inplace=True)
    result = {}
    for _, row in df.iterrows():
        sec = _nan_str(row.get('Section'))
        if not sec:
            continue
        _em = _nan_str(row.get('ExpandMuc'))   # NaN/None/empty → ''
        try:
            _expand_muc = int(float(_em)) != 0 if _em else False
        except (ValueError, TypeError):
            _expand_muc = False
        result[sec] = {
            'label':          _nan_str(row.get('Label')),
            'view_insert':    _nan_str(row.get('View_Insert')),
            'sheet_contains': _nan_str(row.get('SheetNameContains')),
            'sheet_exclude':  _nan_str(row.get('SheetNameExclude')),
            'data_start_row': _nan_str(row.get('DataStartRow')),
            'parent_section': _nan_str(row.get('ParentSection')),
            'expand_muc':     _expand_muc,
            'row_filter':     _nan_str(row.get('RowFilter')),
        }
    return result


def _load_config_rows(xl, sheet_name, col_names):
    """Generic reader: sheet dạng config (skiprows=3), trả về list[dict]."""
    if sheet_name not in xl.sheet_names:
        return []
    df = pd.read_excel(xl, sheet_name=sheet_name, header=None, skiprows=3)
    for i, c in enumerate(col_names):
        if i < len(df.columns):
            df.rename(columns={df.columns[i]: c}, inplace=True)
    records = []
    for _, row in df.iterrows():
        rec = {c.lower(): _nan_str(row.get(c)) for c in col_names if c in df.columns}
        if any(rec.values()):
            records.append(rec)
    return records


def _load_section_rows(xl, sheet_name, section):
    """Đọc HEADER hoặc DETAIL sheet, filter theo cột Section."""
    if sheet_name not in xl.sheet_names:
        return []
    df = pd.read_excel(xl, sheet_name=sheet_name, header=None, skiprows=3)
    for i, c in enumerate(_SECTION_ENG_COLS):
        if i < len(df.columns):
            df.rename(columns={df.columns[i]: c}, inplace=True)
    if 'Section' not in df.columns:
        return []
    df = df[df['Section'].astype(str).str.strip() == section]
    records = []
    for _, row in df.iterrows():
        sql_col = _nan_str(row.get('SQL_Column'))
        if not sql_col or sql_col == 'SQL_Column':
            continue
        records.append({
            "sql_col":          sql_col,
            "ten_excel":        _nan_str(row.get("Ten_Excel")),
            "kieu_dl":          _nan_str(row.get("Kieu_DL")),
            "do_dai":           _nan_str(row.get("Do_dai")),
            "bat_buoc":         _nan_str(row.get("Bat_buoc")),
            "mac_dinh":         _nan_str(row.get("Mac_dinh")),
            "nguon_dl":         _nan_str(row.get("Nguon_DL"), "Excel"),
            "bang_master":      _nan_str(row.get("Bang_Master")),
            "dieu_kien_master": _nan_str(row.get("Dieu_kien_Master")),
            "kieu_lookup":      _nan_str(row.get("Kieu_Lookup")),
            "truong_so_sanh":   _nan_str(row.get("Truong_So_Sanh")),
            "truong_lay_ve":    _nan_str(row.get("Truong_Lay_Ve")),
            "ghi_chu":          _nan_str(row.get("Ghi_chu")),
            "fill_forward":     "1" if str(_nan_str(row.get("Fill_Forward"))).split('.')[0] == '1' else "",
        })
    return records


def load_mapping(mapping_path=MAPPING_FILE):
    """
    Đọc CK_Mapping_v5.xlsx — 7 sheets mới.
    Trả về dict với keys:
      _CONFIG  → {section: metadata}
      HEADER, BOM2, BOM3, BOM4, THDM_HEADER, THDM_THVT → list[record]
      _SP_CONFIG / SP_CONFIG (alias)  → list[record]
      SP_HOOK, VALIDATORS             → list[record]
    Trả {} nếu file không tồn tại.
    """
    if not os.path.exists(mapping_path):
        return {}
    result = {}
    try:
        xl = pd.ExcelFile(mapping_path, engine='openpyxl')

        # 1) _CONFIG — registry tất cả sections
        config = _load_config(xl)
        result['_CONFIG'] = config

        # 2) HEADER/DETAIL sections — derive từ _CONFIG
        for section, cfg in config.items():
            sheet_name = 'HEADER' if not cfg.get('parent_section') else 'DETAIL'
            result[section] = _load_section_rows(xl, sheet_name, section)

        # 3) _SP_CONFIG (gộp BOM + THDM, section rõ ràng)
        SP_COLS = ['Section', 'SQL_Column', 'SP_Name', 'Params',
                   'OutputFields', 'Fallback', 'IsActive']
        sp_rows = _load_config_rows(xl, '_SP_CONFIG', SP_COLS)
        for rec in sp_rows:  # normalize fallback '1.0' → '1'
            fb = rec.get('fallback', '')
            try:
                f = float(fb)
                if f == int(f):
                    rec['fallback'] = str(int(f))
            except (ValueError, TypeError):
                pass
        result['_SP_CONFIG'] = sp_rows
        result['SP_CONFIG']  = sp_rows   # backward compat

        # 4) SP_HOOK
        HOOK_COLS = ['Section', 'Event', 'SP_Name', 'Condition', 'OutputFields',
                     'Params', 'IsActive', 'XmlParam', 'XmlTag', 'XmlFields']
        result['SP_HOOK'] = [
            r for r in _load_config_rows(xl, 'SP_HOOK', HOOK_COLS)
            if r.get('isactive') == '1'
        ]

        # 5) VALIDATORS
        VAL_COLS = ['Section', 'ValidatorName', 'SQL', 'Params',
                    'WarningMessage', 'WarnOnly', 'IsActive']
        result['VALIDATORS'] = [
            r for r in _load_config_rows(xl, 'VALIDATORS', VAL_COLS)
            if r.get('validatorname')
        ]

    except Exception as e:
        print(f"[Mapping] Lỗi đọc file mapping: {e}")
    return result


# ── VALIDATORS compat: old code dùng VAL_COLS không có Section/WarnOnly ────
# Không cần thay đổi — các record đã có key 'section' và 'warnonly' thêm vào
# downstream code đọc 'validatorname','sql','params','warningmessage','isactive'
# sẽ tự bỏ qua key lạ. WarnOnly check thêm khi cần.


# ── CATALOG_FILTER — đọc trực tiếp qua openpyxl (giữ nguyên, chỉ đổi path) ─
# _catalog_read_filter_ids dùng MAPPING_FILE (đã update line 272)





def build_reverse_map(mapping, section='BOM2'):
    """
    {norm_ten_excel: sql_col} — chỉ với các trường Nguon_DL=Excel.
    Lưu cả key đầy đủ lẫn key bỏ phần trong ngoặc.
    """
    reverse = {}
    for r in mapping.get(section, []):
        if r["nguon_dl"] != "Excel":
            continue
        ten = r["ten_excel"]
        sql = r["sql_col"]
        norm_full  = _norm_vn(ten)
        norm_short = _norm_vn(re.sub(r'\(.*?\)', '', ten))
        if norm_full:  reverse[norm_full]  = sql
        if norm_short: reverse[norm_short] = sql
    return reverse

def match_col_to_sql(excel_col_norm, reverse_map):
    """Khớp cột Excel (đã norm) → SQL column. None nếu không tìm thấy."""
    if excel_col_norm in reverse_map:
        return reverse_map[excel_col_norm]
    for key, sql_col in reverse_map.items():
        if key and excel_col_norm and (
            excel_col_norm.startswith(key[:6]) or key.startswith(excel_col_norm[:6])
        ):
            return sql_col
    return None

# ─────────────────────────────────────────────────────────────────────────────
# 3. PARSER — cải tiến từ main_v1.py
# ─────────────────────────────────────────────────────────────────────────────

# META_KEYS — fallback khi chưa load mapping.
# Khi mapping đã load, dùng build_meta_keys_from_mapping() thay thế.
META_KEYS = {
    "Dự án":        r"Dự án",
    "Đơn hàng":     r"Đơn hàng",
    "Tên sản phẩm": r"Tên [Ss]ản phẩm|Sản phẩm",
    "Mã sản phẩm":  r"Mã sản phẩm|Mã SP",
    "Số lượng":     r"Số lượng",
    "Mục số":       r"Mục số",
    "Hoàn thiện":   r"Hoàn thiện",
    "Kích thước":   r"Kích thước",
    "Kết cấu":      r"Kết cấu",
    "Vật liệu":     r"Vật liệu",
}

def build_meta_keys_from_mapping(mapping):
    """
    Build META_KEYS động từ HEADER mapping.
    Mỗi trường Nguon_DL=Excel có Ten_Excel → dùng làm key + regex pattern.
    Fallback sang META_KEYS cứng nếu mapping trống.
    """
    keys = {}
    for rec in mapping.get('HEADER', []):
        if rec.get('nguon_dl') != 'Excel':
            continue
        ten = rec.get('ten_excel', '').strip()
        if not ten:
            continue
        if '|' in ten:
            # Multi-field: đăng ký từng phần riêng (bỏ @FieldName vì đó là ref parent_row)
            for part in ten.split('|'):
                part = part.strip()
                if part and not part.startswith('@'):
                    keys[part] = re.escape(part)
        else:
            keys[ten] = re.escape(ten)
    return keys if keys else META_KEYS

HEADER_ANCHORS    = ["MÃ SP", "STT", "Tên chi tiết", "Tên Vật Tư", "Mã chi tiết", "Mã vật tư"]
SECTION_STT_PATTERN = re.compile(r"^[A-Za-z][A-Za-z0-9]*(\.[0-9]+)?\.?$")  # A, B, E1, E2, I., II., III., A.1...
NUMERIC_STT_PATTERN = re.compile(r"^\d+(\.\d+)*$")               # 1, 2, 1.1, 2.3...
_ROMAN_SIMPLE_RE = re.compile(
    r'^(X{0,3})(IX|IV|V?I{0,3})\.?$', re.IGNORECASE)

def _is_roman_numeral(s: str) -> bool:
    """True nếu s là số La Mã thực dùng trong BOM (I–XXXIX, chỉ dùng I/V/X).
    Loại trừ chữ cái đơn A–Z làm section header (A,B,C,D không khớp pattern này)."""
    s = s.rstrip('.')
    m = _ROMAN_SIMPLE_RE.match(s)
    return bool(m) and bool(m.group(0).rstrip('.'))
FOOTER_KEYWORDS   = ["Tổng cộng", "Kiểm duyệt", "Người Lập", "Người lập", "*Ghi chú", "Ghi chú:",
                     "Ngày tháng năm", "Ngày   tháng", "duyệt", "Ghi chú cuối"]
# Pattern nhận diện dòng ký tên: STT trống + cột số có text dài
_SIGNER_COL_RE    = re.compile(r"(DAI|DAY|RONG|Width|Length|Thickness)", re.IGNORECASE)

def _load_sheet_config(mapping_path=None):
    """Đọc sheet _CONFIG từ mapping file → trả list config dict cho các section BOM*.

    Đây là SINGLE SOURCE OF TRUTH cho section detection.
    Không có hardcode fallback — nếu _CONFIG thiếu/lỗi, trả về [] và caller
    tự xử lý (parse_bom_file sẽ thêm warning rõ ràng cho user).

    _CONFIG columns (R4 trở đi, sau 3 dòng header):
      [0] Section  [1] Label  [2] View_Insert  [3] SheetNameContains
      [4] SheetNameExclude  [5] DataStartRow  [6] ParentSection ...

    Chỉ lấy các dòng BOM* — HEADER và THDM_* được xử lý riêng.
    """
    path = mapping_path or MAPPING_FILE
    try:
        import openpyxl as _opx
        _wb = _opx.load_workbook(path, read_only=True, data_only=True)
        if '_CONFIG' not in _wb.sheetnames:
            return []
        cfg = []
        for row in _wb['_CONFIG'].iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            section = str(row[0]).strip()
            if not section.upper().startswith('BOM'):
                continue
            label     = str(row[1]).strip() if row[1] else section
            contains  = [x.strip().upper() for x in str(row[3] or '').split(',') if x.strip()]
            excludes  = [x.strip().upper() for x in str(row[4] or '').split(',') if x.strip()]
            start_row = str(row[5]).strip() if row[5] else 'AUTO'
            if contains:
                cfg.append({"section": section, "label": label,
                             "contains": contains, "excludes": excludes,
                             "start_row": start_row})
        _wb.close()
        return cfg
    except Exception:
        return []

def _detect_sheet_type(name, sheet_config=None):
    """Match tên sheet Excel với config — hoàn toàn data-driven.

    sheet_config phải được truyền từ caller (parse_bom_file → _parse_sheet).
    Nếu không truyền → không match sheet nào → trả về 'UNKNOWN'.
    Không dùng global cache để tránh stale data.
    """
    cfg = sheet_config or []
    n = name.upper()
    for item in cfg:
        if all(c in n for c in item["contains"]) and \
           not any(e in n for e in item["excludes"]):
            return item["section"]
    return "UNKNOWN"

def _resolve_formula(val):
    """Chuyển formula string (data_only=False) thành giá trị thực.
    - ="some text"  → "some text"
    - =A1, =CONCAT(...) → None  (không resolve được, sẽ fallback sang cached)
    - Giá trị thường → giữ nguyên
    """
    if val is None:
        return None
    if isinstance(val, str) and val.startswith('='):
        m = re.match(r'^="(.*)"$', val, re.DOTALL)
        if m:
            return m.group(1)
        # Công thức phức tạp không resolve được
        return None
    return val

def _merge_meta_rows(live_rows, cached_rows):
    """Gộp live rows (data_only=False) và cached rows (data_only=True).
    Ưu tiên live value đã resolve; fallback sang cached khi formula phức tạp.
    """
    merged = []
    for i in range(max(len(live_rows), len(cached_rows))):
        live_row   = live_rows[i]   if i < len(live_rows)   else ()
        cached_row = cached_rows[i] if i < len(cached_rows) else ()
        n = max(len(live_row), len(cached_row))
        row = []
        for j in range(n):
            live_val   = live_row[j]   if j < len(live_row)   else None
            cached_val = cached_row[j] if j < len(cached_row) else None
            resolved   = _resolve_formula(live_val)
            row.append(resolved if resolved is not None else cached_val)
        merged.append(tuple(row))
    return merged

def _extract_meta(rows, meta_keys=None):
    """
    Scan 15 dòng đầu, tìm label → lấy giá trị ô kế bên.
    meta_keys: {label: regex_pattern} — dynamic từ mapping hoặc fallback META_KEYS.
    """
    keys = meta_keys if meta_keys else META_KEYS
    meta = {}
    for row in rows:
        for key, pat in keys.items():
            if key in meta:
                continue
            for i, cell in enumerate(row):
                if cell and re.search(pat, str(cell), re.IGNORECASE):
                    for j in range(i + 1, len(row)):
                        if row[j] is not None and str(row[j]).strip():
                            meta[key] = str(row[j]).strip()
                            break
                    break
    return meta

def _find_header_row(rows):
    for i, row in enumerate(rows):
        texts = [str(c).strip() for c in row if c is not None]
        hits  = sum(1 for a in HEADER_ANCHORS if any(a.lower() in t.lower() for t in texts))
        if hits >= 2:
            return i
    return None

def _norm_col(s):
    return re.sub(r"[\s\n]+", "_", str(s).strip()).rstrip("_") if s else ""

def _build_headers(r1, r2):
    """Gộp 2 dòng header (merged cell) thành tên cột duy nhất.
    Forward-fill r1 để xử lý merged cell: khi r1[i]=None nhưng r2[i] có giá trị,
    nghĩa là ô r1 bị merge → kế thừa giá trị parent từ cột trước.
    """
    n = max(len(r1), len(r2))
    r1_f = list(r1) + [None] * max(0, n - len(r1))
    r2_f = list(r2) + [None] * max(0, n - len(r2)) if r2 else [None] * n

    # Forward-fill r1 cho merged cells: r1[i]=None nhưng r2[i] có giá trị
    last_r1 = None
    for i in range(n):
        if r1_f[i] is not None:
            last_r1 = r1_f[i]
        elif r2_f[i] is not None and last_r1 is not None:
            r1_f[i] = last_r1

    headers = []
    for i in range(n):
        v1 = _norm_col(r1_f[i]) if r1_f[i] is not None else ""
        v2 = _norm_col(r2_f[i]) if r2_f[i] is not None else ""
        if v1 and v2:   headers.append(f"{v1}_{v2}")
        elif v1:        headers.append(v1)
        elif v2:        headers.append(v2)
        else:           headers.append(f"COL_{i}")
    # Deduplicate
    seen, result = {}, []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result

def _is_data_row(rd, key_cols):
    """True nếu dòng có ít nhất 1 giá trị thực trong các cột key."""
    for k in key_cols:
        v = rd.get(k)
        if v is not None and str(v).strip() not in ('', '0', 'None', 'nan'):
            return True
    return False

# ─── Shared Excel parse pipeline (BOM + THDM) ───────────────────────────────

def _build_excel_col_map(headers, mapping_recs, col_aliases=None):
    """
    Build {col_index → sql_col} từ headers và mapping records.
    Dùng cho openpyxl-based section parsing (THDM; BOM tương lai).

    col_aliases: dict {normalized_header → canonical} để chuẩn hoá tên cột Excel
                 (vd: 'mavttb' → 'mavattu')
    """
    aliases    = col_aliases or {}
    excel_recs = [r for r in mapping_recs
                  if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')]
    col_map    = {}
    for i, h in enumerate(headers):
        h_vn       = _norm_vn(str(h).replace('_', ' '))
        h_canonical = aliases.get(h_vn, h_vn)
        for rec in excel_recs:
            ten_vn = _norm_vn(rec['ten_excel'])
            if (h_vn == ten_vn
                    or ten_vn == h_canonical
                    or (ten_vn and ten_vn in h_vn)
                    or (h_canonical and ten_vn in h_canonical)
                    or (h_vn and h_vn in ten_vn)):
                col_map[i] = rec['sql_col']
                break
    return col_map


def _parse_section_excel_rows(all_rows, headers, col_map, mapping_recs, stt_idx,
                               skip_no_item=True, item_cols=('ItemId', 'ItemName'),
                               ff_always_override=False,
                               extra_col_extractor=None):
    """
    Pipeline chung parse Excel rows cho 1 section — dùng cho cả BOM và THDM.

    Xử lý đầy đủ:
    - Bỏ dòng trống
    - Footer detection (FOOTER_KEYWORDS)
    - Section header detection (SECTION_STT_PATTERN + GIAO RỜI / sub-group check)
    - Fill_Forward: capture từ section header → apply cho data rows kế tiếp
    - Bỏ dòng không có identity (skip_no_item=True)

    all_rows    : list of value tuples (openpyxl rows)
    headers     : list[str] — tên cột đã build (kể cả 2-row merge)
    col_map     : {col_index → sql_col}   ← từ _build_excel_col_map()
    mapping_recs: list of mapping records cho section này
    stt_idx     : index cột STT trong headers (None nếu không tìm thấy)

    Trả về: list of {sql_col: value} dicts đã apply Fill_Forward.
    """
    import math as _m

    ff_fields   = [r['sql_col'] for r in mapping_recs
                   if r.get('fill_forward') == '1' and r.get('sql_col')]
    dim_indices = [
        i for i, h in enumerate(headers)
        if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(h)))
    ]

    current_ff = {}
    footer_hit = False
    rows_out   = []

    for row in all_rows:
        # Bỏ dòng hoàn toàn trống
        if all(c is None for c in row):
            continue
        if footer_hit:
            break

        padded = list(row) + [None] * max(0, len(headers) - len(row))

        # Footer detection
        row_text = ' '.join(str(v) for v in padded if v is not None)
        if any(kw.lower() in row_text.lower() for kw in FOOTER_KEYWORDS):
            footer_hit = True
            continue

        # STT value
        stt_val = padded[stt_idx] if stt_idx is not None and stt_idx < len(padded) else None
        stt_str = str(stt_val).strip() if stt_val is not None else ''

        is_hdr = bool(stt_str and SECTION_STT_PATTERN.match(stt_str))

        # GIAO RỜI / sub-group: có kích thước vật lý → data row, không phải section header
        if is_hdr and dim_indices:
            if any(
                padded[i] not in (None, '', 0, 0.0)
                and not (isinstance(padded[i], float) and _m.isnan(padded[i]))
                for i in dim_indices if i < len(padded)
            ):
                is_hdr = False

        # Build row dict {sql_col: value}
        rd = {sql_col: padded[ci]
              for ci, sql_col in col_map.items()
              if ci < len(padded)}

        if is_hdr:
            if not _is_roman_numeral(stt_str):
                # Chữ cái (A, B, C...) → capture Fill_Forward, bỏ qua không INSERT
                for sql_col in ff_fields:
                    v = rd.get(sql_col)
                    if v is not None and str(v).strip() not in ('', 'nan'):
                        current_ff[sql_col] = v
                continue
            # Số La Mã (I, II, III...) → fall-through, xử lý như data row

        # Apply Fill_Forward vào data row
        for sql_col in ff_fields:
            if sql_col in current_ff:
                if ff_always_override or not rd.get(sql_col):
                    rd[sql_col] = current_ff[sql_col]

        # Extra columns (e.g., CHI TIẾT MỤC data cho THDM transpose)
        if extra_col_extractor is not None:
            rd.update(extra_col_extractor(padded))

        # Bỏ dòng không có identity
        if skip_no_item and not any(rd.get(c) for c in item_cols):
            continue

        rows_out.append(rd)

    return rows_out


def _run_row_sp_hooks(conn, hooks, row_vals, log_fn=None):
    """
    Chạy SP_HOOK event='beforeinsert' cho 1 data row.
    Cập nhật row_vals in-place với output fields từ SP.
    Dùng chung cho BOM và THDM.

    conn    : pyodbc connection
    hooks   : list of hook dicts (SP_HOOK mapping, event='beforeinsert', đã filter theo section)
    row_vals: dict {sql_col: value} — sẽ bị cập nhật in-place
    log_fn  : optional callable(msg: str) để log warning
    """
    for hook in hooks:
        cond       = hook.get('condition', '').strip()
        should_run = True
        if cond.startswith('EMPTY(') and cond.endswith(')'):
            should_run = not row_vals.get(cond[6:-1].strip())
        elif cond.startswith('NOTEMPTY(') and cond.endswith(')'):
            should_run = bool(row_vals.get(cond[9:-1].strip()))
        if not should_run:
            continue
        try:
            h_sp   = hook.get('sp_name', '').strip()
            h_par  = hook.get('params', '').strip()
            h_outs = [f.strip() for f in hook.get('outputfields', '').split(',') if f.strip()]
            pc, pv = [], []
            for ph in h_par.split('|'):
                ph = ph.strip()
                if '=' not in ph:
                    continue
                k, v = ph.split('=', 1)
                k = k.strip().lstrip('@')
                v = v.strip()
                if v.startswith('{') and v.endswith('}'):
                    v = row_vals.get(v[1:-1])
                if v is None:
                    pc.append(f'@{k}=NULL')
                else:
                    pc.append(f'@{k}=?')
                    pv.append(v)
            cur = conn.cursor()
            cur.execute(f'EXEC {h_sp} ' + ', '.join(pc), pv)
            hk_row = cur.fetchone()
            if hk_row and cur.description:
                hk_res = dict(zip([d[0] for d in cur.description], hk_row))
                for fh in h_outs:
                    if fh in hk_res:
                        row_vals[fh] = hk_res[fh]
        except Exception as eh:
            if log_fn:
                log_fn(f'SP_HOOK {hook.get("sp_name")} warn: {eh}')

def _parse_sheet(ws, sheet_name, live_meta_rows=None, meta_keys=None, hidden_rows=None, hidden_cols=None, sheet_config=None):
    # Bỏ qua các row/col bị ẩn (hidden) trong Excel
    # hidden_rows: set of 1-based row indices
    # hidden_cols: set of 0-based col indices
    _hr = hidden_rows or set()
    _hc = hidden_cols or set()

    def _filter_cols(row):
        if not _hc:
            return row
        return tuple(v for i, v in enumerate(row) if i not in _hc)

    all_rows = [
        _filter_cols(vals)
        for row_idx, vals in enumerate(ws.iter_rows(values_only=True), start=1)
        if row_idx not in _hr
    ]
    if not all_rows:
        return None

    stype = _detect_sheet_type(sheet_name, sheet_config=sheet_config)
    if stype == "UNKNOWN":
        return None

    # Dùng live_meta_rows (data_only=False) để đọc đúng giá trị formula cells
    meta_rows = live_meta_rows if live_meta_rows is not None else all_rows[:15]
    meta = _extract_meta(meta_rows, meta_keys=meta_keys)
    hidx = _find_header_row(all_rows[:15])

    if hidx is None:
        return {"sheet_name": sheet_name, "sheet_type": stype,
                "metadata": meta, "df": pd.DataFrame()}

    r1 = all_rows[hidx]
    r2 = all_rows[hidx + 1] if hidx + 1 < len(all_rows) else ()

    # Kiểm tra r2 có phải dòng sub-header không
    has_sub = bool(r2) and r2[0] is None and any(
        c is not None and str(c).strip() not in ("", "0")
        for c in (r2[4:] if r2 else []))

    if has_sub:
        headers    = _build_headers(r1, r2)
        data_start = hidx + 2
    else:
        headers    = [_norm_col(c) if c else f"COL_{i}" for i, c in enumerate(r1)]
        data_start = hidx + 1

    key_cols = [h for h in headers if any(k in h for k in
        ("Tên", "Mã", "STT", "Name", "Code", "Item"))]

    rows_out = []
    _footer_seen = False   # Khi gặp footer (Kiểm duyệt / Người Lập...) → bỏ tất cả dòng phía sau
    for row in all_rows[data_start:]:
        if all(c is None for c in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        rd     = dict(zip(headers, padded[:len(headers)]))

        _raw_stt = rd.get("STT")
        stt = ("0" if _raw_stt in (0, 0.0)
               else "" if _raw_stt is None or _raw_stt == ""
               else str(_raw_stt).strip())

        # Section header: STT là chữ cái đơn hoặc chữ.số (A, B, C, F, B.1, B.2...)
        # Số La Mã (I, II, III...) được giữ lại như data row.
        # Áp dụng thống nhất cho BOM2, BOM3, BOM4 — không cần check SLg
        # → giữ lại cho Fill_Forward capture, không INSERT
        if SECTION_STT_PATTERN.match(stt):
            # Nếu dòng có kích thước vật lý (DÀY/RỘNG/DÀI) thì là GIAO RỜI/sub-group (vd MODULE MD1 STT='I')
            # → không phải section header thật → không update Fill-Forward
            _has_dims = any(
                rd.get(col) not in (None, '', 0, '0', 0.0)
                for col in rd
                if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(col)))
            )
            if not _has_dims:
                if _is_roman_numeral(stt):
                    pass  # Số La Mã → fall-through, xử lý như data row
                else:
                    rows_out.append(rd)   # Chữ cái → giữ lại cho Fill_Forward, không INSERT
                    continue
            # Có kích thước → fall-through: xử lý như data row bình thường
        # Bỏ dòng trống STT=0 (Phần III có nhiều dòng này)
        if stt == "0":
            continue
        # Bỏ dòng footer (case-insensitive) — và đặt cờ để bỏ qua tất cả dòng phía sau
        # (bao gồm dòng tên người ký xuất hiện sau "Kiểm duyệt / Người Lập")
        row_text = " ".join(str(v) for v in rd.values() if v)
        if any(kw.lower() in row_text.lower() for kw in FOOTER_KEYWORDS):
            _footer_seen = True
            continue
        if _footer_seen:
            continue
        # Bỏ dòng có STT không hợp lệ: không phải số (1, 2, 1.1...) và không phải La Mã
        # → ví dụ tên người ký nằm trong cột STT: "Đỗ Hồng Thái", "Nguyễn Tân Hồng"
        if stt and not NUMERIC_STT_PATTERN.match(stt) and not _is_roman_numeral(stt):
            continue
        # Bỏ dòng ký tên: STT trống + (cột số có text phi số) HOẶC (Số lượng có text phi số)
        # Cần check cả SLg vì BOM3 không có cột DÀI/Width nên _SIGNER_COL_RE không match
        if not stt:
            def _is_non_numeric(v):
                try: float(str(v).replace(',', '.')); return False
                except: return True
            _dim_text = any(_is_non_numeric(val) for col, val in rd.items()
                            if val and _SIGNER_COL_RE.search(col))
            _slg_h_sig = next((h for h in rd
                               if _norm_vn(str(h)) in ('slg', 'soluong', 'soluongsp')), None)
            _slg_v_sig = rd.get(_slg_h_sig) if _slg_h_sig else None
            _slg_text  = (_slg_v_sig is not None
                          and str(_slg_v_sig).strip() not in ('', 'nan')
                          and _is_non_numeric(_slg_v_sig))
            if _dim_text or _slg_text:
                continue
        # Bỏ dòng không có dữ liệu key
        if not _is_data_row(rd, key_cols):
            continue

        rows_out.append(rd)

    df = pd.DataFrame(rows_out, columns=headers) if rows_out else pd.DataFrame(columns=headers)

    # Bỏ cột toàn None hoặc tên COL_
    drop = [c for c in df.columns if c.startswith("COL_") or df[c].isna().all()]
    df   = df.drop(columns=drop, errors="ignore")

    return {"sheet_name": sheet_name, "sheet_type": stype, "metadata": meta, "df": df}

def _ask_excel_password(filename):
    """
    Hiện dialog nhỏ hỏi mật khẩu Excel.
    Trả về: password (str) hoặc None nếu người dùng bấm Hủy.
    """
    result = [None]
    dialog = tk.Toplevel()
    dialog.title("File được bảo vệ bằng mật khẩu")
    dialog.resizable(False, False)
    dialog.grab_set()
    dialog.focus_force()

    # Center dialog
    dialog.update_idletasks()
    w, h = 400, 180
    x = (dialog.winfo_screenwidth()  - w) // 2
    y = (dialog.winfo_screenheight() - h) // 2
    dialog.geometry(f"{w}x{h}+{x}+{y}")
    dialog.configure(bg="#1e1e1e")

    tk.Label(dialog, text="🔒  File Excel có mật khẩu",
             font=("Segoe UI", 11, "bold"),
             bg="#1e1e1e", fg="#F1F5F9").pack(pady=(16, 2))
    tk.Label(dialog, text=os.path.basename(filename),
             font=("Segoe UI", 9), bg="#1e1e1e", fg="#94A3B8").pack()
    tk.Label(dialog, text="Nhập mật khẩu để mở file:",
             font=("Segoe UI", 9), bg="#1e1e1e", fg="#CBD5E1").pack(pady=(10, 4))

    entry = tk.Entry(dialog, show="*", font=("Segoe UI", 10),
                     bg="#334155", fg="white", insertbackground="white",
                     relief="flat", width=32)
    entry.pack(ipady=5)
    entry.focus_set()

    err_lbl = tk.Label(dialog, text="", font=("Segoe UI", 8),
                       bg="#1e1e1e", fg="#F87171")
    err_lbl.pack()

    def on_ok(event=None):
        pw = entry.get()
        if not pw:
            err_lbl.config(text="Vui lòng nhập mật khẩu.")
            return
        result[0] = pw
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    btn_frame = tk.Frame(dialog, bg="#1e1e1e")
    btn_frame.pack(pady=(4, 0))
    tk.Button(btn_frame, text="Hủy", width=10, command=on_cancel,
              bg="#334155", fg="white", relief="flat",
              activebackground="#475569", cursor="hand2").pack(side="left", padx=6)
    tk.Button(btn_frame, text="✓  Xác nhận", width=12, command=on_ok,
              bg="#2563eb", fg="white", relief="flat",
              activebackground="#1d4ed8", cursor="hand2").pack(side="left", padx=6)

    entry.bind("<Return>", on_ok)
    dialog.wait_window()
    return result[0]


def _decrypt_excel(filepath, password):
    """
    Giải mã file Excel có mật khẩu bằng msoffcrypto.
    Trả về BytesIO chứa nội dung đã giải mã, hoặc raise Exception nếu sai mật khẩu.
    """
    decrypted = io.BytesIO()
    with open(filepath, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def _is_encrypted_excel(filepath):
    """
    Kiểm tra xem file có bị mã hóa không.
    Method 1: msoffcrypto (nếu đã install).
    Method 2: fallback — thử mở bằng zipfile, nếu fail → có thể bị mã hóa.
    """
    # Method 1: msoffcrypto chính xác hơn
    if _HAS_MSOFFCRYPTO:
        try:
            with open(filepath, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                return office_file.is_encrypted()
        except Exception:
            pass

    # Method 2: fallback — file xlsx là ZIP, nếu không mở được → encrypted
    import zipfile
    try:
        with zipfile.ZipFile(filepath, 'r'):
            return False  # mở được ZIP → không có password
    except zipfile.BadZipFile:
        return True   # không phải ZIP → khả năng cao bị mã hóa
    except Exception:
        return False


def parse_bom_file(filepath, meta_keys=None, _decrypted_bytes=None):
    """
    Parse file BOM Excel.
    meta_keys: {label: regex} từ build_meta_keys_from_mapping() — None → dùng META_KEYS.
    _decrypted_bytes: BytesIO đã giải mã (truyền vào nếu file có mật khẩu).
    Trả về: (tables_dict, global_meta, skipped_sheets, warnings)
    tables_dict = { label: {df, type, warnings} }
    """
    # Load sheet config fresh mỗi lần parse — đọc lại từ _CONFIG trong mapping file
    # để label/section phản ánh đúng mapping file hiện tại, không cần restart app.
    sheet_cfg = _load_sheet_config()
    if not sheet_cfg:
        # _CONFIG thiếu hoặc không đọc được → không detect được section BOM nào
        import warnings as _w
        _w.warn(
            "BOM section config không tải được từ mapping file (_CONFIG sheet). "
            "Không thể nhận dạng sheet BOM trong file Excel.",
            RuntimeWarning, stacklevel=2,
        )

    # Nguồn dữ liệu: file gốc hoặc BytesIO đã giải mã
    src1 = _decrypted_bytes if _decrypted_bytes is not None else filepath

    wb = load_workbook(src1, read_only=True, data_only=True)

    # Mở lần 2 không data_only để đọc đúng giá trị ô có công thức
    try:
        if _decrypted_bytes is not None:
            _decrypted_bytes.seek(0)
        src2 = _decrypted_bytes if _decrypted_bytes is not None else filepath
        wb_live = load_workbook(src2, read_only=True, data_only=False)
    except Exception:
        wb_live = None

    # Mở lần 3 (không read_only) chỉ để lấy thông tin row bị ẩn
    _hidden_rows_map = {}
    _hidden_cols_map = {}   # sheet_name → set of 0-based col indices bị ẩn
    try:
        if _decrypted_bytes is not None:
            _decrypted_bytes.seek(0)
        src3 = _decrypted_bytes if _decrypted_bytes is not None else filepath
        _wb_hid = load_workbook(src3, read_only=False, data_only=True)
        from openpyxl.utils import column_index_from_string
        for _sname in _wb_hid.sheetnames:
            _ws_hid = _wb_hid[_sname]
            _hidden_rows_map[_sname] = {
                r for r, dim in _ws_hid.row_dimensions.items() if dim.hidden
            }
            _hidden_cols_map[_sname] = {
                column_index_from_string(c) - 1   # chuyển về 0-based
                for c, dim in _ws_hid.column_dimensions.items() if dim.hidden
            }
        _wb_hid.close()
    except Exception:
        _hidden_rows_map = {}
        _hidden_cols_map = {}

    parsed   = []
    skipped  = []
    warnings = []
    global_meta = {}

    if not sheet_cfg:
        warnings.append(
            "⚠ Không đọc được cấu hình BOM từ mapping file (_CONFIG sheet). "
            "Không có section BOM nào được nhận dạng trong file này."
        )

    for name in wb.sheetnames:
        # Bỏ qua sheet bị ẩn (hidden / veryHidden)
        if wb[name].sheet_state != 'visible':
            continue
        # Lấy 15 dòng đầu từ wb_live (formula-aware) và wb (cached) rồi merge
        live_meta_rows = None
        if wb_live and name in wb_live.sheetnames:
            ws_live = wb_live[name]
            live_raw    = []
            cached_raw  = []
            for i, row in enumerate(ws_live.iter_rows(values_only=True)):
                if i >= 15:
                    break
                live_raw.append(row)
            cached_raw = list(wb[name].iter_rows(min_row=1, max_row=15, values_only=True))
            live_meta_rows = _merge_meta_rows(live_raw, cached_raw)

        res = _parse_sheet(wb[name], name, live_meta_rows=live_meta_rows, meta_keys=meta_keys, hidden_rows=_hidden_rows_map.get(name), hidden_cols=_hidden_cols_map.get(name), sheet_config=sheet_cfg)
        if res:
            parsed.append(res)
            global_meta.update(res["metadata"])
        else:
            skipped.append(name)

    if wb_live:
        wb_live.close()
    wb.close()

    tables = {}

    # Bảng H: Header
    if global_meta:
        tables["[H] Phiếu Header"] = {
            "df": pd.DataFrame([global_meta]), "type": "HEADER", "warnings": []}
    else:
        w = "Không tìm thấy thông tin đầu phiếu (Dự án, Đơn hàng...)"
        tables["[H] Phiếu Header"] = {
            "df": pd.DataFrame([{"Lỗi": w}]), "type": "HEADER", "warnings": [w]}
        warnings.append(w)

    # Bảng BOM — đọc từ sheet_cfg (load fresh từ mapping, không dùng global cache)
    for cfg_item in sheet_cfg:
        section   = cfg_item["section"]   # 'BOM2', 'BOM3', 'BOM4', 'BOM5'...
        label     = cfg_item["label"]     # '[2] BOM Phần II...'
        found = next((s for s in parsed if s["sheet_type"] == section), None)
        if found:
            tables[label] = {"df": found["df"], "type": section, "warnings": []}
        else:
            w = f"Không tìm thấy sheet {section}"
            tables[label] = {
                "df": pd.DataFrame([{"Lỗi": w}]), "type": section, "warnings": [w]}
            warnings.append(w)

    return tables, global_meta, skipped, warnings

# ─────────────────────────────────────────────────────────────────────────────
# 3b. THDM EXCEL PARSER  (giống BOM: detect header, map cột theo Ten_Excel)
# ─────────────────────────────────────────────────────────────────────────────

# Fallback anchors khi mapping chưa load — chỉ dùng nếu thdm_map rỗng
_THDM_HEADER_ANCHORS_FALLBACK = [
    "mã vật tư", "tên vật tư", "đvt", "quy cách", "mã vt",
    "ten vat tu", "ma vat tu", "don vi tinh",
    "mã vttb", "tên vt/tb", "ma vttb", "ten vt",
]

# Alias: normalize tên cột viết tắt → norm của Ten_Excel trong mapping
# Vẫn cần cho bước col_map building (khớp header Excel → sql_col)
_THDM_COL_ALIASES = {
    'mavttb':      'mavattu',
    'tenvttb':     'tenvattu',
    'mavt':        'mavattu',
    'tenvt':       'tenvattu',
    'tenvttbmavt': 'tenvattu',
}

def _thdm_find_thvt_sheet(wb):
    """Tìm sheet THVT trong workbook.
    Ưu tiên exact 'TH VT', sau đó flexible contains (normalize tiếng Việt)."""
    if "TH VT" in wb.sheetnames:
        return "TH VT"
    _candidates = ["THVT", "TH_VT", "TONG HOP VAT TU", "TỔNG HỢP VẬT TƯ", "THVAT TU"]
    for sname in wb.sheetnames:
        n = _norm_vn(sname)
        if any(_norm_vn(c) in n for c in _candidates):
            return sname
    return None

def _thdm_find_header_row(all_rows, anchors=None, max_scan=25):
    """
    Tìm dòng header trong sheet THVT.
    anchors: list norm string từ mapping Ten_Excel. Nếu None → dùng fallback.
    """
    _anchors = anchors if anchors else _THDM_HEADER_ANCHORS_FALLBACK
    for i, row in enumerate(all_rows[:max_scan]):
        texts = [_norm_vn(str(c)) for c in row if c is not None]
        hits  = sum(1 for a in _anchors
                    if any(a in t or t in a for t in texts))
        if hits >= 2:
            return i
    return None

def _norm_muc_key(v):
    """Chuẩn hoá Mục number: '1.0'→'1', '1.10'→'1.1', 'Muc 1'→'1'."""
    if v is None:
        return None
    s = str(v).strip().replace('Muc ', '').replace('MUC ', '').strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else f'{f:g}'
    except (ValueError, TypeError):
        return s


def _thdm_load_bom_qty_dict(conn, bom_ids):
    """
    Query B20BOMDetail (BOMDetailType IN 2,3) cho các BOM đã chọn.
    Kết quả: dict {(item_code_str, muc_no_str): dinh_muc}
    - item_code_str : B20Item.Code  (VD: 'NVL9-000464')
    - muc_no_str    : BuiltinOrder0 đã normalize (VD: '1', '1.1')

    ĐỊNH MỨC = SUM(Quantity9 + QuantityFactory + QuantityConstruction + QuantitySubcontractor)
    Lý do: BOMDetailType=2 (CT) dùng Quantity9, BOMDetailType=3 (TP) dùng QuantityFactory/Construction/Subcontractor.
    Cộng tất cả an toàn vì các cột còn lại = NULL (→ 0) theo từng loại.
    """
    if not bom_ids:
        return {}
    ph  = ', '.join('?' * len(bom_ids))
    sql = f"""
        SELECT i.Code,
               dt.BuiltinOrder0,
               SUM(
                   ISNULL(bd.Quantity9, 0)
                   + ISNULL(bd.QuantityFactory, 0)
                   + ISNULL(bd.QuantityConstruction, 0)
                   + ISNULL(bd.QuantitySubcontractor, 0)
               ) AS DinhMuc
        FROM   B20BOMDetail          bd
        JOIN   B20BOM                bom ON bom.Id              = bd.BOMId
        JOIN   B30BizDocDetailSO     dt  ON dt.DetailRowId_SO   = bom.DetailRowId_SO
        JOIN   B20Item               i   ON i.Id                = bd.ItemId
        WHERE  bd.BOMId   IN ({ph})
          AND  bd.BOMDetailType IN (2, 3)
          AND  bom.DetailRowId_SO IS NOT NULL
          AND  bom.DetailRowId_SO != ''
        GROUP BY i.Code, dt.BuiltinOrder0
    """
    cur = conn.cursor()
    cur.execute(sql, list(bom_ids))
    result = {}
    for row in cur.fetchall():
        code     = str(row[0]).strip() if row[0] is not None else None
        muc      = _norm_muc_key(row[1])
        dinh_muc = float(row[2]) if row[2] is not None else 0.0  # Decimal → float
        if code and muc:
            result[(code, muc)] = dinh_muc
    return result


def _thdm_expand_muc_rows(rows, bom_qty_dict=None):
    """
    Transpose CHI TIẾT MỤC: 1 Excel row → N DB rows (1 per Mục có CT hoặc TP > 0).

    Logic (khi có bom_qty_dict):
    - Quantity9 (ĐỊNH MỨC P.KT) lấy từ B20BOMDetail SUM(Qty9) theo (ItemCode, MụcNo)
    - QuantityFactory = Quantity9_DB - CT - TP  (có thể âm nếu CT+TP > ĐỊNH MỨC)
    - Nếu không match trong bom_qty_dict → Quantity9 = None, QuantityFactory = None

    Logic fallback (bom_qty_dict=None — chế độ cũ):
    - NHÀ MÁY = ĐỊNH MỨC - tổng CT - tổng TP
    - QuantityFactory phân bổ theo tỷ lệ (CT_X+TP_X) / tổng(CT+TP)

    Mỗi parsed row phải có:
        _muc_ct: {muc_id: ct_value}   — từ CHI TIẾT GIAO CÔNG TRÌNH
        _muc_tp: {muc_id: tp_value}   — từ CHI TIẾT GIAO THẦU PHỤ
    """
    def _fnum(v):
        """Convert Excel cell value (str/int/float/None/Decimal) → float."""
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    expanded = []
    for row in rows:
        muc_ct   = row.pop('_muc_ct', {}) or {}
        muc_tp   = row.pop('_muc_tp', {}) or {}
        all_mucs = sorted(set(muc_ct) | set(muc_tp))

        if not all_mucs:
            # Không có Mục nào → factory-only, giữ nguyên
            expanded.append(row)
        else:
            # ItemCode từ Excel (Code string, VD: 'NVL9-000464')
            item_code = str(row.get('ItemId') or '').strip()

            # Fallback ratio (dùng khi bom_qty_dict=None)
            dinh_muc   = _fnum(row.get('Quantity9') or row.get('Quantity'))
            total_ct   = sum(_fnum(muc_ct.get(m)) for m in all_mucs)
            total_tp   = sum(_fnum(muc_tp.get(m)) for m in all_mucs)
            nha_may    = dinh_muc - total_ct - total_tp
            total_ctpt = total_ct + total_tp
            fac_distributed = 0.0

            for idx, muc_id in enumerate(all_mucs):
                ct_val  = _fnum(muc_ct.get(muc_id))
                tp_val  = _fnum(muc_tp.get(muc_id))
                is_last = (idx == len(all_mucs) - 1)

                if bom_qty_dict is not None:
                    # ── Chế độ BOM lookup ──────────────────────────────────────
                    muc_key = _norm_muc_key(muc_id)
                    qty9_db = bom_qty_dict.get((item_code, muc_key))
                    if qty9_db is not None:
                        fac_val  = qty9_db - ct_val - tp_val   # cho phép âm
                        qty_muc  = qty9_db
                    else:
                        fac_val  = None
                        qty_muc  = None
                else:
                    # ── Fallback: phân bổ tỷ lệ (chế độ cũ) ──────────────────
                    if is_last:
                        fac_val = max(0, (nha_may or 0) - fac_distributed)
                    elif total_ctpt > 0:
                        fac_val = max(0, round((nha_may or 0) * (ct_val + tp_val) / total_ctpt))
                    else:
                        fac_val = 0
                    fac_distributed += fac_val
                    qty_muc = ct_val + tp_val + fac_val

                new_row = dict(row)
                new_row['DetailRowId_SO']        = muc_id
                new_row['QuantityConstruction']  = ct_val
                new_row['QuantitySubcontractor'] = tp_val
                new_row['QuantityFactory']       = fac_val
                new_row['Quantity']              = qty_muc
                new_row['Quantity9']             = qty_muc
                expanded.append(new_row)
    return expanded


def _thdm_parse_thvt_sheet(wb_cached, wb_live, wb_hid, sheet_name, thdm_map,
                            expand_muc=True, bom_qty_dict=None):
    """
    Parse sheet TH VT giống BOM:
    - Detect header động (không hardcode dòng 12)
    - Map cột theo Ten_Excel từ thdm_map
    - Xử lý formula, hidden rows/cols
    - expand_muc=True: expand 1 row → N rows per Mục (THDM_THVT)
    - expand_muc=False: giữ nguyên rows, không expand (THDM_NVL...)
    Trả về: list of dict {sql_col: value}
    """
    from openpyxl.utils import column_index_from_string

    # Hidden rows / cols
    _hr, _hc = set(), set()
    if wb_hid and sheet_name in wb_hid.sheetnames:
        _ws_h = wb_hid[sheet_name]
        _hr   = {r for r, d in _ws_h.row_dimensions.items()    if d.hidden}
        _hc   = {column_index_from_string(c) - 1
                 for c, d in _ws_h.column_dimensions.items() if d.hidden}

    def _fcols(row):
        return tuple(v for i, v in enumerate(row) if i not in _hc) if _hc else row

    # Đọc cached (data_only=True)
    ws_c    = wb_cached[sheet_name]
    c_rows  = [_fcols(vals)
               for idx, vals in enumerate(ws_c.iter_rows(values_only=True), 1)
               if idx not in _hr]

    # Merge với live (formula resolution)
    if wb_live and sheet_name in wb_live.sheetnames:
        ws_l   = wb_live[sheet_name]
        l_rows = [_fcols(vals)
                  for idx, vals in enumerate(ws_l.iter_rows(values_only=True), 1)
                  if idx not in _hr]
        all_rows = _merge_meta_rows(l_rows, c_rows)
    else:
        all_rows = c_rows

    if not all_rows:
        return []

    # Build anchors từ mapping Ten_Excel (mapping-driven, như BOM)
    mapping_anchors = [
        _norm_vn(r['ten_excel'])
        for r in thdm_map
        if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')
    ]

    # Tìm dòng header
    hidx = _thdm_find_header_row(all_rows, anchors=mapping_anchors or None)
    if hidx is None:
        # Gợi ý: liệt kê các tên cột tìm thấy ở vùng đầu để dễ debug
        found_hdrs = []
        for row in all_rows[:10]:
            row_texts = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if row_texts:
                found_hdrs.append(" | ".join(row_texts[:8]))
        hint = "\n".join(found_hdrs[:5]) if found_hdrs else "(sheet trống)"
        expected = ", ".join(r['ten_excel'] for r in thdm_map
                             if r.get('nguon_dl') == 'Excel' and r.get('ten_excel'))[:120]
        raise ValueError(
            f"Không tìm thấy dòng header trong sheet '{sheet_name}'.\n\n"
            f"Cột mapping cần khớp: {expected}\n\n"
            f"10 dòng đầu tìm thấy:\n{hint}"
        )

    # Build headers (hỗ trợ 2-dòng merged header giống BOM)
    r1 = all_rows[hidx]
    r2 = all_rows[hidx + 1] if hidx + 1 < len(all_rows) else ()
    has_sub = bool(r2) and r2[0] is None and any(
        c is not None and str(c).strip() not in ("", "0") for c in (r2[4:] if r2 else []))
    if has_sub:
        headers    = _build_headers(r1, r2)
        data_start = hidx + 2
    else:
        headers    = [_norm_col(c) if c else f"COL_{i}" for i, c in enumerate(r1)]
        data_start = hidx + 1

    # Build col_map: {col_index → sql_col} — dùng shared helper + _THDM_COL_ALIASES
    col_map = _build_excel_col_map(headers, thdm_map, col_aliases=_THDM_COL_ALIASES)

    # Tìm vị trí cột STT
    stt_idx = next(
        (i for i, h in enumerate(headers)
         if _norm_vn(h.replace('_', ' ')).startswith('stt')),
        None)

    # ── Detect CHI TIẾT GIAO CT / TP Mục columns từ raw header rows ─────────
    r1_raw = all_rows[hidx]
    r2_raw = all_rows[hidx + 1] if has_sub else ()

    def _detect_chi_tiet_cols(r1, r2, kw_norm):
        """Scan r1/r2 tìm super-header chứa kw_norm → {col_idx: muc_id}."""
        kw_norm  = _norm_vn(kw_norm)   # FIX: normalize keyword (strip spaces/diacritics)
        muc_cols = {}
        in_sect  = False
        n        = max(len(r1), len(r2) if r2 else 0)
        r1p      = list(r1) + [None] * max(0, n - len(r1))
        r2p      = list(r2) + [None] * max(0, n - len(r2)) if r2 else [None] * n
        for i in range(n):
            v1 = r1p[i]
            v2 = r2p[i]
            if v1 is not None and str(v1).strip():
                if kw_norm in _norm_vn(str(v1).lower()):
                    in_sect = True
                elif in_sect:
                    break   # next non-empty super-header → kết thúc section
            if in_sect and v2 and str(v2).strip().lower().startswith('muc '):
                muc_id = str(v2).strip()[4:].strip()   # "Muc 1.1" → "1.1"
                if muc_id:
                    muc_cols[i] = muc_id
        return muc_cols

    ct_muc_cols = _detect_chi_tiet_cols(r1_raw, r2_raw, 'giao cong trinh')
    tp_muc_cols = _detect_chi_tiet_cols(r1_raw, r2_raw, 'giao thau phu')

    # ── Build extra_col_extractor cho _parse_section_excel_rows ──────────────
    if ct_muc_cols or tp_muc_cols:
        def _muc_extractor(padded):
            ct = {mid: padded[ci] for ci, mid in ct_muc_cols.items()
                  if ci < len(padded) and padded[ci] and padded[ci] != 0}
            tp = {mid: padded[ci] for ci, mid in tp_muc_cols.items()
                  if ci < len(padded) and padded[ci] and padded[ci] != 0}
            return {'_muc_ct': ct, '_muc_tp': tp}
    else:
        _muc_extractor = None

    # Parse rows — shared pipeline (empty filter, footer, section header, Fill_Forward, no-item filter)
    rows = _parse_section_excel_rows(
        all_rows[data_start:], headers, col_map, thdm_map, stt_idx,
        skip_no_item=True, item_cols=('ItemId', 'ItemName'),
        ff_always_override=True,   # ItemType luôn lấy section letter (A, D, E1...) từ header
        extra_col_extractor=_muc_extractor,
    )

    # Fix col_map collision: "Định mức P.KT" map vào Quantity9 trước (first-match)
    # → Quantity không được map → copy Quantity9 → Quantity cho mỗi row
    for r in rows:
        if r.get('Quantity') is None and r.get('Quantity9') is not None:
            r['Quantity'] = r['Quantity9']

    # ── Expand Mục rows: 1 Excel row → N DB rows (CHI TIẾT GIAO CT/TP) ──────
    if expand_muc and (ct_muc_cols or tp_muc_cols):
        rows = _thdm_expand_muc_rows(rows, bom_qty_dict=bom_qty_dict)

    return rows


def _thdm_open_workbook(path):
    """Mở file THDM 3 lần (cached / live / hid) giống parse_bom_excel."""
    from openpyxl import load_workbook
    wb_c = load_workbook(path, read_only=True,  data_only=True)
    try:
        wb_l = load_workbook(path, read_only=True,  data_only=False)
    except Exception:
        wb_l = None
    try:
        wb_h = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        wb_h = None
    return wb_c, wb_l, wb_h


def _thdm_apply_row_filter(rows, filter_str):
    """
    Lọc rows theo expression đơn giản: 'ColName>0', 'ColName>=0.01', v.v.
    Chỉ hỗ trợ: >=, <=, !=, >, <, == (1 điều kiện).
    """
    import re as _re_flt
    m = _re_flt.match(r'^(\w+)\s*(>=|<=|!=|>|<|==)\s*(.+)$', filter_str.strip())
    if not m:
        return rows
    col, op, val_str = m.group(1), m.group(2), m.group(3).strip()
    try:
        val = float(val_str)
    except ValueError:
        val = val_str

    def _check(row):
        rv = row.get(col)
        if rv is None:
            return False
        try:
            rv = float(rv)
        except (TypeError, ValueError):
            pass
        if op == '>':  return rv > val
        if op == '<':  return rv < val
        if op == '>=': return rv >= val
        if op == '<=': return rv <= val
        if op == '==': return rv == val
        if op == '!=': return rv != val
        return True

    return [r for r in rows if _check(r)]


def _thdm_get_detail_col_defs(section_map):
    """Col defs cho tab detail THDM — hiển thị TẤT CẢ cột của section.
    - Cột định danh = SQL_Column (duy nhất) → header text = Ten_Excel,
      hàng phụ (sql_names) = SQL_Column.
    - Ưu tiên: cột có Ten_Excel xếp lên trước, giữ nguyên thứ tự mapping
      trong từng nhóm; cột không có Ten_Excel xếp sau.
    """
    with_te, without_te = [], []
    for r in section_map:
        sql_col = (r.get('sql_col') or '').strip()
        if not sql_col:
            continue
        ten_excel = (r.get('ten_excel') or '').strip()
        nguon     = r.get('nguon_dl') or ''
        kieu      = str(r.get('kieu_dl') or '').lower()
        is_num    = any(t in kieu for t in ('decimal', 'float', 'int', 'money', 'numeric'))
        try:
            do_dai = int(float(r.get('do_dai') or 0))
        except (ValueError, TypeError):
            do_dai = 0
        anchor = 'e' if is_num else 'w'
        if is_num:        px = 130
        elif do_dai <= 20:  px = 120
        elif do_dai <= 50:  px = 150
        elif do_dai <= 100: px = 180
        else:               px = 200
        # cột hệ thống (không Excel) hẹp hơn cho đỡ chiếm chỗ
        if not ten_excel:
            px = min(px, 120)
        d = {
            'sql_col':   sql_col,
            'ten_excel': ten_excel,
            'nguon_dl':  nguon,
            'mac_dinh':  (r.get('mac_dinh') or '').strip(),
            'width':     px,
            'anchor':    anchor,
            'stretch':   bool(ten_excel),
        }
        (with_te if ten_excel else without_te).append(d)
    return with_te + without_te


# ─────────────────────────────────────────────────────────────────────────────
# 3c. SHARED ROW RESOLVER — dùng chung BOM và THDM
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_row_mapping(mapping_recs, ctx, get_excel_val=None):
    """
    Resolve một row từ mapping records theo NguonDL — dùng chung BOM và THDM.

    mapping_recs : list of mapping record (HEADER/DETAIL section)
    ctx          : dict chứa tất cả nguồn dữ liệu:
        now           : datetime  — dùng cho HeThong NOW
        builtin_order : int       — AUTO_INC sort order
        doc_id        : str/uuid  — BOMId / BizDocId placeholder
        parent_row    : dict      — resolved parent row (HEADER / THDM_HEADER)
        parent_fields : set       — sql_col copy từ parent (HeThong)
        ui_values     : dict      — {'creator': ..., 'product_id': ..., 'order_id': ...}
        skip_nguon    : set       — skip hoàn toàn (mặc định: {'SP', 'TinhToan'})
    get_excel_val : callable(rec) → value
        THDM: lambda rec: excel_row.get(rec['sql_col'])
        BOM:  existing complex lookup — truyền None để skip Excel fields
    Returns: dict {sql_col: value}
    """
    now           = ctx.get('now')
    parent_row    = ctx.get('parent_row') or {}
    parent_fields = ctx.get('parent_fields') or set()
    doc_id        = ctx.get('doc_id')
    builtin_order = ctx.get('builtin_order', 1)
    ui_values     = ctx.get('ui_values') or {}
    skip_nguon    = ctx.get('skip_nguon', {'SP', 'TinhToan'})

    row_out = {}
    for rec in mapping_recs:
        sql_col = (rec.get('sql_col') or '').strip()
        if not sql_col:
            continue
        nguon   = (rec.get('nguon_dl') or 'Excel').strip()
        mac     = str(rec.get('mac_dinh') or '').strip()
        kieu_dl = str(rec.get('kieu_dl') or '').lower()

        # ── Skip ─────────────────────────────────────────────────────────────
        if nguon in skip_nguon:
            row_out.setdefault(sql_col, None)
            continue

        # ── Excel ─────────────────────────────────────────────────────────────
        if nguon == 'Excel':
            if get_excel_val is not None:
                val = get_excel_val(rec)
                if val is None:
                    if mac.upper() == 'EMPTY':
                        # EMPTY → '' (không để NULL, DB có thể NOT NULL)
                        val = None if kieu_dl in ('date', 'datetime') else ''
                    elif mac and mac.upper() not in ('NULL', 'NONE', ''):
                        # Fallback sang mac_dinh nếu có giá trị cụ thể
                        try:    val = int(mac)
                        except:
                            try:    val = float(mac)
                            except: val = mac
                row_out[sql_col] = val
            continue

        # ── CoDinh ────────────────────────────────────────────────────────────
        if nguon == 'CoDinh':
            if not mac or mac.upper() == 'NULL':
                row_out[sql_col] = None
            elif mac.upper() == 'EMPTY':
                row_out[sql_col] = None if kieu_dl in ('date', 'datetime') else ''
            else:
                try:    row_out[sql_col] = int(mac)
                except:
                    try:    row_out[sql_col] = float(mac)
                    except: row_out[sql_col] = mac
            continue

        # ── HeThong ───────────────────────────────────────────────────────────
        if nguon == 'HeThong':
            if mac == 'NOW' or sql_col in ('CreatedAt', 'ModifiedAt'):
                row_out[sql_col] = now
            elif mac == 'AUTO_INC':
                row_out[sql_col] = builtin_order
            elif sql_col in ('BOMId', 'BizDocId') or mac == 'NEWID':
                row_out[sql_col] = doc_id
            elif sql_col in parent_fields:
                row_out[sql_col] = parent_row.get(sql_col)
            elif mac and mac in parent_row:
                # mac_dinh = tên field của parent → copy giá trị
                row_out[sql_col] = parent_row[mac]
            elif mac and mac not in ('NULL', 'EMPTY', 'NOW', 'AUTO_INC', 'NEWID'):
                row_out[sql_col] = mac   # giá trị cố định (fallback)
            else:
                row_out[sql_col] = None
            continue

        # ── UILookup ──────────────────────────────────────────────────────────
        if nguon == 'UILookup':
            row_out[sql_col] = ui_values.get(mac)
            continue

        # ── MucLookup: đọc raw Muc key từ Excel row (đã set bởi _thdm_expand_muc_rows)
        # Giá trị ban đầu = muc_id ('1', '1.1'...), sẽ được replace bằng DB id
        # trong _thdm_insert_worker sau khi compound lookup (BuiltinOrder0|BizDocId)
        if nguon == 'MucLookup':
            if get_excel_val is not None:
                val = get_excel_val(rec)
                row_out[sql_col] = val
            else:
                row_out[sql_col] = None
            continue

    return row_out


# ─────────────────────────────────────────────────────────────────────────────
# 4. VALIDATE LAYER 1
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_HEADER_FIELDS = ["Mã sản phẩm", "Dự án", "Đơn hàng"]

NUMERIC_COL_PATTERNS = [
    r"(Dày|Rộng|Dài|Thickness|Width|Length)",
    r"(Số lượng|Quantity|SLg|SLSP)",
    r"(Khối lượng|Weight|WeightCalc)",
    r"(Chiều dài|EdgeLength)",
    r"(Diện tích|Area|FinishArea|PrimerArea)",
    r"(Số cạnh|EdgeCount)",
    r"(Số mặt|FinishSide|PrimerSide)",
]
NUMERIC_RE = re.compile("|".join(NUMERIC_COL_PATTERNS), re.IGNORECASE)

_DATE_FMTS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y"]

def _try_parse_date(s, fmt):
    try:
        datetime.datetime.strptime(s, fmt)
        return True
    except (ValueError, TypeError):
        return False

def validate_layer1(tables, global_meta, mapping=None):
    """
    Validate cấu trúc dữ liệu.
    Khi có mapping, kiểm tra thêm: trường bắt buộc, datetime format, nvarchar length.
    Trả về: { table_label: [ {row, col, message, severity} ] }
    """
    result = {}

    # ── Header — lấy danh sách bắt buộc từ mapping (Bat_buoc=CÓ, Nguon_DL=Excel) ───
    h_errors = []
    norm_meta_h = {_norm_vn(k): v for k, v in global_meta.items()}
    if mapping and mapping.get('HEADER'):
        required_recs = [
            r for r in mapping['HEADER']
            if r.get('bat_buoc', '').rstrip('0').rstrip('.') == '1'
            and r.get('nguon_dl', '') == 'Excel'
        ]
        for rec in required_recs:
            ten = rec.get('ten_excel', '').strip()
            if not ten:
                continue
            val = (global_meta.get(ten)
                   or norm_meta_h.get(_norm_vn(ten))
                   or '')
            if not str(val).strip():
                h_errors.append({
                    "row": "Header", "col": ten,
                    "message": f"Trường bắt buộc '{ten}' đang trống",
                    "severity": "error"
                })
    else:
        # fallback khi chưa load mapping
        for field in REQUIRED_HEADER_FIELDS:
            val = global_meta.get(field, '')
            if not str(val).strip():
                h_errors.append({
                    "row": "Header", "col": field,
                    "message": f"Trường bắt buộc '{field}' đang trống",
                    "severity": "error"
                })
    result["[H] Phiếu Header"] = h_errors

    # ── Detail tables ─────────────────────────────────────────────────────────
    for label, tbl in tables.items():
        if label == "[H] Phiếu Header":
            continue
        df      = tbl["df"]
        section = tbl["type"]   # BOM2 / BOM3 / BOM4
        errors  = []

        if df.empty or "Lỗi" in df.columns:
            result[label] = errors
            continue

        # Xây col_info: {excel_col → record từ mapping}
        col_info = {}
        if mapping:
            rev_map  = build_reverse_map(mapping, section)
            sec_recs = {r["sql_col"]: r for r in mapping.get(section, [])}
            for col in df.columns:
                sql_col = match_col_to_sql(_norm_vn(col), rev_map)
                if sql_col and sql_col in sec_recs:
                    col_info[col] = sec_recs[sql_col]

        # ── Check 1: Trường bắt buộc từ mapping ──────────────────────────
        if mapping:
            for rec in mapping.get(section, []):
                if rec["bat_buoc"].rstrip("0").rstrip(".") != "1" or rec["nguon_dl"] != "Excel":
                    continue
                matched_col = next(
                    (c for c, info in col_info.items()
                     if info["sql_col"] == rec["sql_col"]), None)
                if matched_col is None:
                    errors.append({
                        "row": "—", "col": rec["sql_col"],
                        "message": (f"Thiếu cột bắt buộc: '{rec['ten_excel']}'"
                                    f"  (SQL: {rec['sql_col']})"),
                        "severity": "error"
                    })
                    continue
                for idx, val in enumerate(df[matched_col]):
                    v = "" if (val is None or
                               (isinstance(val, float) and pd.isna(val))) \
                        else str(val).strip()
                    if not v:
                        errors.append({
                            "row": idx + 1, "col": matched_col,
                            "message": (f"Dòng {idx+1} | '{matched_col}'"
                                        f": trường bắt buộc đang trống"),
                            "severity": "error"
                        })

        # ── Check 2: Kiểu số ──────────────────────────────────────────────
        # Loại trừ cột có chứa "Tên"/"Name" (text) hoặc "Lọc"/"Loc" (flag IN/OUT)
        _NUMERIC_EXCLUDE_RE = re.compile(
            r"(Tên|Name|Lọc|Loc|_IN\b|_OUT\b)", re.IGNORECASE)
        # Giá trị placeholder (dấu gạch ngang) → coi như NULL, bỏ qua
        _PLACEHOLDER_RE = re.compile(r"^-+$")
        for col in df.columns:
            if not NUMERIC_RE.search(col):
                continue
            if _NUMERIC_EXCLUDE_RE.search(col):
                continue   # cột flag/text ẩn sau tên có keyword số
            for idx, val in enumerate(df[col]):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                val_str = str(val).strip()
                if not val_str or _PLACEHOLDER_RE.match(val_str):
                    continue   # trống hoặc dấu gạch (--, -, ---) → bỏ qua
                try:
                    float(val_str.replace(",", "."))
                except (ValueError, TypeError):
                    errors.append({
                        "row": idx + 1, "col": col,
                        "message": (f"Dòng {idx+1} | '{col}'"
                                    f": '{val}' không phải số"),
                        "severity": "error"
                    })

        # ── Check 3: Kiểu datetime ────────────────────────────────────────
        for col, info in col_info.items():
            if info["kieu_dl"] != "datetime":
                continue
            for idx, val in enumerate(df[col]):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                if isinstance(val, (datetime.datetime, datetime.date)):
                    continue   # openpyxl đã parse → OK
                val_str = str(val).strip()
                if not val_str:
                    continue
                if not any(_try_parse_date(val_str, f) for f in _DATE_FMTS):
                    errors.append({
                        "row": idx + 1, "col": col,
                        "message": (f"Dòng {idx+1} | '{col}'"
                                    f": '{val_str}' không đúng định dạng ngày"
                                    f" (dd/mm/yyyy)"),
                        "severity": "warning"
                    })

        # ── Check 4: char/varchar/nvarchar length ─────────────────────────
        for col, info in col_info.items():
            if info["kieu_dl"] not in ("nvarchar", "varchar", "char", "nchar") \
                    or not info.get("do_dai"):
                continue
            try:
                max_len = int(info["do_dai"])
            except (ValueError, TypeError):
                continue
            for idx, val in enumerate(df[col]):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                v_str = str(val)
                if len(v_str) > max_len:
                    errors.append({
                        "row": idx + 1, "col": col,
                        "message": (f"Dòng {idx+1} | '{col}'"
                                    f": {len(v_str)} ký tự > max {max_len}"),
                        "severity": "warning"
                    })

        result[label] = errors

    return result

def count_errors(val_errors):
    n_err = sum(len([e for e in v if e["severity"] == "error"])  for v in val_errors.values())
    n_wrn = sum(len([e for e in v if e["severity"] == "warning"]) for v in val_errors.values())
    return n_err, n_wrn

# ─────────────────────────────────────────────────────────────────────────────
# 5. GUI
# ─────────────────────────────────────────────────────────────────────────────

C = {
    "bg":     "#0F172A",
    "panel":  "#1E293B",
    "border": "#334155",
    "accent": "#38BDF8",
    "green":  "#10B981",
    "yellow": "#F59E0B",
    "red":    "#EF4444",
    "text":   "#F1F5F9",
    "muted":  "#64748B",
    "white":  "#FFFFFF",
}

class BOMToolApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # An ngay - tranh flash den khi khoi dong
        self.title(f"BOHO IMPORT BOM/THDM v{APP_VERSION}")
        self.geometry("1360x820")
        # Set app icon
        try:
            if getattr(sys, 'frozen', False):
                base = sys._MEIPASS
            else:
                base = os.path.dirname(os.path.abspath(__file__))
            ico_path = os.path.join(base, "icon.ico")
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
        except Exception:
            pass

        self.tables        = {}
        self.global_meta   = {}
        self.val_errors    = {}
        self.mapping       = load_mapping()
        self._current_file = ""
        self._batch_files  = []
        self._loading_dlg  = None
        self._last_import_bom_id = None   # BOM Id của lần import gần nhất (cho Hoàn tác)
        self._last_import_fname  = ''
        self.db_cfg        = self._load_db_config()
        self._skipped_sheets = []
        self._current_creator_code = None   # UserCode được chọn ở combobox Người lập
        self._thdm_creator_code    = None   # Creator riêng cho THDM tab

        ctk.set_widget_scaling(1.0)
        self._setup_styles()
        self._build_ui()

        # Hiện startup overlay, show window, rồi maximize sau 100ms
        self._show_startup_overlay()
        self.deiconify()
        self.update()
        self.after(100, lambda: self.wm_state('zoomed'))

    def _setup_styles(self):
        """TTK styles cho Treeview — CTk tự lo phần còn lại."""
        mode = ctk.get_appearance_mode()  # "Dark" or "Light"
        is_dark = mode == "Dark"

        tv_bg   = "#1E1E1E" if is_dark else "#FFFFFF"
        tv_panel= "#252526" if is_dark else "#F3F4F6"
        tv_text = "#CCCCCC" if is_dark else "#1E293B"
        tv_head = "#2D2D30" if is_dark else "#E5E7EB"
        tv_acc  = "#007ACC" if is_dark else "#3B82F6"
        tv_sel  = "#37373D" if is_dark else "#BFDBFE"
        # Màu đường kẻ ngăn cách giữa các ô/row
        tv_sep  = "#3C3C3C" if is_dark else "#CBD5E1"

        s = ttk.Style()
        s.theme_use("clam")
        s.configure("BOM.Treeview",
            rowheight=33, font=("Segoe UI", 13),
            background=tv_sep, foreground=tv_text,
            fieldbackground=tv_sep, borderwidth=0, indent=12)
        s.configure("BOM.Treeview.Heading",
            font=("Segoe UI", 12, "bold"),
            background=tv_head, foreground=tv_acc,
            borderwidth=1, relief="solid")
        # Đường kẻ dọc giữa các ô dữ liệu (clam theme cell element)
        s.configure("BOM.Treeview.Cell",
            padding=1,
            borderwidth=1,
            relief="solid")
        s.map("BOM.Treeview",
            background=[("selected", tv_sel)],
            foreground=[("selected", tv_text)])

        self._tv_bg    = tv_bg
        self._tv_panel = tv_panel
        self._tv_text  = tv_text
        self._tv_sel   = tv_sel

    # ── CTk appearance callbacks (từ sidebar OptionMenu) ──────────────────────
    def _setup_scrollbar_style(self):
        """Scrollbar mỏng, phẳng kiểu VS Code."""
        mode = ctk.get_appearance_mode()
        is_dark = mode == "Dark"
        sb_bg     = "#3E3E42" if is_dark else "#C8C8C8"
        sb_trough = "#1E1E1E" if is_dark else "#F0F0F0"
        sb_active = "#686868" if is_dark else "#A0A0A0"
        sb_arrow  = "#858585" if is_dark else "#787878"
        s = ttk.Style()
        for orient in ("Vertical", "Horizontal"):
            name = f"{orient}.TScrollbar"
            s.configure(name,
                gripcount=0,
                relief="flat",
                background=sb_bg,
                darkcolor=sb_trough,
                lightcolor=sb_trough,
                troughcolor=sb_trough,
                bordercolor=sb_trough,
                arrowcolor=sb_arrow,
                arrowsize=12,
                width=10 if orient == "Vertical" else 10)
            s.map(name,
                background=[("active", sb_active), ("pressed", sb_active)],
                arrowcolor=[("active", "#CCCCCC" if is_dark else "#555555")])
        # Slim scrollbar variant (no arrows) for main content
        for orient in ("Vertical", "Horizontal"):
            name = f"Slim.{orient}.TScrollbar"
            s.configure(name,
                gripcount=0,
                relief="flat",
                background=sb_bg,
                darkcolor=sb_trough,
                lightcolor=sb_trough,
                troughcolor=sb_trough,
                bordercolor=sb_trough,
                arrowcolor=sb_trough,   # arrows same color = hidden
                arrowsize=0,
                width=8 if orient == "Vertical" else 8)
            s.map(name,
                background=[("active", sb_active)])

    def _open_settings(self):
        """Popup Cài đặt: Giao diện + Tỷ lệ hiển thị — xuất hiện ngay dưới nút 🎨."""
        # Toggle: bấm lại thì đóng
        try:
            if hasattr(self, '_settings_dlg') and self._settings_dlg.winfo_exists():
                self._settings_dlg.destroy()
                return
        except Exception:
            pass

        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        dlg.title("Cài đặt")
        dlg.geometry("270x130")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        self._settings_dlg = dlg

        # (Đã bỏ lựa chọn Giao diện Dark/Light — app thiết kế dark-only,
        #  chọn Light sẽ ra giao diện nửa trắng nửa đen vì hàng chục widget
        #  tk/tksheet hardcode màu tối. Muốn hỗ trợ Light thật sự phải
        #  refactor toàn bộ hệ màu — để backlog.)

        ctk.CTkLabel(dlg, text="Tỷ lệ hiển thị:",
            font=ctk.CTkFont("Segoe UI", 12), anchor="w").pack(
            fill="x", padx=16, pady=(14, 2))
        sm = ctk.CTkOptionMenu(dlg, values=["80%", "90%", "100%", "110%", "120%"],
            command=self.change_scaling_event,
            font=ctk.CTkFont("Segoe UI", 12))
        sm.pack(fill="x", padx=16)
        sm.set(getattr(self, '_current_scaling_pct', '100%'))

        # Neo popup góc phải màn hình, ngay dưới nút 🎨
        def _place_and_show():
            try:
                sw = self.winfo_screenwidth()
                by = self._btn_settings.winfo_rooty() + self._btn_settings.winfo_height() + 4
                dlg.geometry(f"+{sw - 285}+{by}")
            except Exception:
                dlg.geometry(f"+{self.winfo_screenwidth() - 285}+{55}")
            dlg.deiconify()
            dlg.lift()

        self.after(10, _place_and_show)

    def change_scaling_event(self, value: str):
        scale = int(value.replace("%", "")) / 100
        ctk.set_widget_scaling(scale)
        self._current_scaling_pct = value
        # ctk.set_widget_scaling() → CTkTabview._set_scaling() → _configure_grid()
        # tự reset thanh tab về sticky="ns" (căn giữa) — áp lại canh trái.
        try:
            self.nb._segmented_button.grid_configure(sticky="w", padx=8, pady=(4, 0))
        except Exception:
            pass

    # ── Startup overlay ─────────────────────────────────────────────────────────

    def _show_startup_overlay(self):
        """Overlay kiểm tra kết nối DB khi khởi động. Che toàn bộ main UI."""
        BG = "#181818"
        overlay = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        overlay.place(x=0, y=0, relwidth=1, relheight=1)
        overlay.lift()
        self._startup_overlay = overlay

        # Khung nội dung căn giữa
        center = ctk.CTkFrame(overlay, fg_color="transparent")
        center.place(relx=0.5, rely=0.46, anchor="center")

        # Logo circle
        logo_wrap = ctk.CTkFrame(center, fg_color="#0F172A",
                                 width=88, height=88, corner_radius=44)
        logo_wrap.pack(pady=(0, 22))
        logo_wrap.pack_propagate(False)
        ctk.CTkLabel(logo_wrap, text="B",
                     font=ctk.CTkFont("Segoe UI", 42, "bold"),
                     text_color="#2563EB").place(relx=0.5, rely=0.48, anchor="center")

        # Tên app
        ctk.CTkLabel(center, text="BOHO IMPORT BOM/THDM",
                     font=ctk.CTkFont("Segoe UI", 24, "bold"),
                     text_color="#F0F0F0").pack()
        ctk.CTkLabel(center, text="v1  —  BOHO",
                     font=ctk.CTkFont("Segoe UI", 12),
                     text_color="#555555").pack(pady=(3, 22))

        # Separator
        ctk.CTkFrame(center, fg_color="#2A2A2A", height=1, width=340).pack(pady=(0, 22))

        # Thông tin kết nối
        cfg = self.db_cfg or {}
        server   = cfg.get("server",   "Chưa cấu hình")
        database = cfg.get("database", "—")
        db_card = ctk.CTkFrame(center, fg_color="#222222", corner_radius=10)
        db_card.pack(pady=(0, 26), ipadx=24, ipady=4)
        ctk.CTkLabel(db_card, text=f"🖧  {server}",
                     font=ctk.CTkFont("Segoe UI", 13),
                     text_color="#CCCCCC").pack(padx=28, pady=(12, 3))
        ctk.CTkLabel(db_card, text=f"📋  {database}",
                     font=ctk.CTkFont("Segoe UI", 12),
                     text_color="#777777").pack(padx=28, pady=(0, 12))

        # Status label
        self._startup_status_lbl = ctk.CTkLabel(
            center, text="⏳  Đang kiểm tra kết nối...",
            font=ctk.CTkFont("Segoe UI", 13),
            text_color="#888888")
        self._startup_status_lbl.pack()

        # Button frame (ẩn, chỉ hiện khi có lỗi)
        self._startup_btn_frame = ctk.CTkFrame(center, fg_color="transparent")

        self.update()
        threading.Thread(target=self._startup_check_worker, daemon=True).start()

    def _startup_check_worker(self):
        """Background thread: thử kết nối DB. Delay tối thiểu 0.8s để overlay hiện đủ lâu."""
        import time
        t0 = time.time()
        if not self.db_cfg:
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda: self._startup_check_done(False,
                "Chưa có cấu hình DB (db_config.json)."))
            return
        try:
            conn = self._get_db_conn(timeout_sec=15)
            conn.close()
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda: self._startup_check_done(True, None))
        except Exception as e:
            elapsed = time.time() - t0
            time.sleep(max(0, 0.8 - elapsed))
            self.after(0, lambda err=e: self._startup_check_done(False, str(err)))

    def _startup_check_done(self, ok, error, attempt=1):
        """Main thread: cập nhật UI startup sau khi có kết quả kiểm tra."""
        if ok:
            self._startup_status_lbl.configure(
                text="✅  Kết nối thành công — Đang mở ứng dụng...",
                text_color="#4EC9B0")
            self.after(1200, self._dismiss_startup_overlay)
        else:
            if attempt < 3:
                # Tự động thử lại tối đa 3 lần sau 2s — cho server/VPN kịp ổn định
                self._startup_status_lbl.configure(
                    text=f"⏳  Đang thử lại ({attempt}/3)...",
                    text_color="#888888")
                def _retry_worker(att=attempt):
                    import time
                    time.sleep(2.0)
                    try:
                        conn = self._get_db_conn(timeout_sec=15)
                        conn.close()
                        self.after(0, lambda: self._startup_check_done(True, None, att + 1))
                    except Exception as ex:
                        self.after(0, lambda e=ex: self._startup_check_done(False, str(e), att + 1))
                threading.Thread(target=_retry_worker, daemon=True).start()
            else:
                # Sau 3 lần thất bại → hiện nút cho user
                self._startup_status_lbl.configure(
                    text="❌  Không kết nối được DB",
                    text_color="#F87171")
                bf = self._startup_btn_frame
                bf.pack(pady=(18, 0))
                # Hiện chi tiết lỗi để dễ debug
                err_short = str(error)[:200] if error else "Không rõ nguyên nhân"
                ctk.CTkLabel(bf,
                    text=err_short,
                    font=ctk.CTkFont("Segoe UI", 10),
                    text_color="#888888",
                    wraplength=420,
                    justify="left").pack(pady=(0, 8))
                ctk.CTkLabel(bf,
                    text="⚠️  Kiểm tra VPN hoặc cấu hình DB rồi thử lại",
                    font=ctk.CTkFont("Segoe UI", 11),
                    text_color="#555555").pack(pady=(0, 14))
                row = ctk.CTkFrame(bf, fg_color="transparent")
                row.pack()
                ctk.CTkButton(row, text="🔄  Thử lại",
                    command=self._startup_retry,
                    fg_color=("#3B82F6", "#007ACC"),
                    hover_color=("#2563EB", "#005F99"),
                    width=120, height=36, corner_radius=8).pack(side="left", padx=8)
                ctk.CTkButton(row, text="Bỏ qua →",
                    command=self._dismiss_startup_overlay,
                    fg_color="transparent", border_width=1,
                    border_color="#444444", text_color="#888888",
                    hover_color="#2A2A2A",
                    width=100, height=36, corner_radius=8).pack(side="left", padx=8)

    def _startup_retry(self):
        """Xóa nút, reset status, thử kết nối lại."""
        for w in self._startup_btn_frame.winfo_children():
            w.destroy()
        self._startup_btn_frame.pack_forget()
        self._startup_status_lbl.configure(
            text="⏳  Đang thử lại...", text_color="#888888")
        threading.Thread(target=self._startup_check_worker, daemon=True).start()

    def _dismiss_startup_overlay(self):
        """Đóng overlay, để lộ main UI."""
        try:
            self._startup_overlay.destroy()
        except Exception:
            pass
        # Auto-load Người lập + Dự án THDM — cố ý đợi tới khi overlay kiểm
        # tra kết nối khởi động đã có kết quả (thành công hoặc user bấm Bỏ
        # qua), thay vì bắn theo timer cố định 600ms/800ms độc lập với kết
        # quả kiểm tra (như trước đây). Race condition cũ: nếu VPN vừa mới
        # kết nối lúc app khởi động, 2 lệnh timer cũ fire trước khi VPN kịp
        # lên, báo lỗi và KHÔNG tự thử lại — dù startup overlay retry sau
        # đó thành công, 2 combo này vẫn kẹt ở trạng thái lỗi cũ.
        if not getattr(self, '_auto_loaded_after_startup', False):
            self._auto_loaded_after_startup = True
            self.after(200, self._load_creator_combo)
            self.after(400, self._thdm_load_products)


    def _build_ui(self):
        # ── Root grid ────────────────────────────────────────────────────────
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.var_export_sql = tk.BooleanVar(value=False)

        # ── Top bar ──────────────────────────────────────────────────────────
        topbar = ctk.CTkFrame(self, corner_radius=0, height=46,
                              fg_color=("gray83", "gray13"))
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(1, weight=1)

        # Trái: App name
        ctk.CTkLabel(topbar,
            text="🏭  BOHO  BOM / THDM",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=("#16A34A", "#4EC9B0")).grid(
            row=0, column=0, padx=(14, 16), pady=12)

        # Giữa: DB info
        if self.db_cfg:
            srv = self.db_cfg.get("server", "?")
            db  = self.db_cfg.get("database", "?")
            db_txt   = f"🗄  {srv} / {db}"
            db_color = ("gray40", "gray55")
        else:
            db_txt   = "⚠️  Chưa có db_config.json"
            db_color = ("#D97706", "#D7BA7D")
        ctk.CTkLabel(topbar, text=db_txt,
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=db_color).grid(row=0, column=1, sticky="w")

        # Phải: status + mapping + settings
        right_bar = ctk.CTkFrame(topbar, fg_color="transparent")
        right_bar.grid(row=0, column=2, padx=(0, 10))

        self.lbl_status = CLabel(right_bar, text="Sẵn sàng...",
            font=ctk.CTkFont("Segoe UI", 11, slant="italic"),
            text_color=("gray40", "gray55"),
            fg_color="transparent")
        self.lbl_status.pack(side=tk.LEFT, padx=(0, 10))
        Tooltip(self.lbl_status,
            lambda: ("Bỏ qua: " + ", ".join(self._skipped_sheets))
                    if self._skipped_sheets else "")

        mapping_ok = bool(self.mapping)
        ctk.CTkButton(right_bar, text="⚙️  Mapping",
            command=self._open_mapping_window,
            fg_color="transparent", border_width=1,
            text_color=("#16A34A","#4ADE80") if mapping_ok else ("#D97706","#D7BA7D"),
            hover_color=("gray90","gray25"),
            font=ctk.CTkFont("Segoe UI", 11),
            width=110, height=28, corner_radius=6).pack(side=tk.LEFT, padx=(0, 4))

        self._btn_settings = ctk.CTkButton(right_bar, text="🎨",
            command=self._open_settings,
            fg_color="transparent", border_width=1,
            text_color=("gray50","gray60"),
            hover_color=("gray90","gray25"),
            font=ctk.CTkFont("Segoe UI", 11),
            width=34, height=28, corner_radius=6)
        self._btn_settings.pack(side=tk.LEFT)

        # ── Tabview ──────────────────────────────────────────────────────────
        self.nb = ctk.CTkTabview(self,
            fg_color=("gray95", "gray14"),
            segmented_button_fg_color=("gray80", "gray17"),
            segmented_button_selected_color=("#D97706", "#B45309"),
            segmented_button_selected_hover_color=("#B45309", "#92400E"),
            segmented_button_unselected_color=("gray80", "gray17"),
            segmented_button_unselected_hover_color=("gray75", "gray22"),
            text_color=("gray20", "gray80"),
            border_width=0, corner_radius=0)
        self.nb.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        self.nb.add(TAB_IMPORT)
        self.nb.add(TAB_THDM)
        self.nb.add(TAB_CATALOG)
        try:
            self.nb._segmented_button.configure(
                font=ctk.CTkFont("Segoe UI", 13, "bold"), height=40)
            self.nb._segmented_button.grid_configure(sticky="w", padx=8, pady=(4, 0))
        except Exception:
            pass

        self._build_tab_import()
        self._build_tab_catalog()
        self._build_tab_thdm()
        self._build_log_panel()

        # Auto-load Người lập + Dự án THDM giờ được kích từ
        # _dismiss_startup_overlay() (chỉ chạy sau khi biết kết quả kiểm
        # tra kết nối khởi động — xem ghi chú tại đó).
        # Reload khi user switch sang tab THDM (bắt sự kiện tab change)
        self.nb.configure(command=self._on_tab_changed)

    def _build_tab_import(self):
        tab = self.nb.tab(TAB_IMPORT)

        # ── Action bar ngang ─────────────────────────────────────────────────
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=48,
                           corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        def _ab_btn(text, cmd, fg, bg, hover, border=0, state="normal",
                    text_color_disabled=None):
            kw = dict(font=ctk.CTkFont("Segoe UI", 11, "bold"),
                fg_color=bg, text_color=fg,
                hover_color=hover, border_width=border,
                corner_radius=6, height=32, state=state)
            if border:
                kw["border_color"] = fg
            if text_color_disabled:
                # CTk mặc định chữ nút disabled dùng gray60 — trên nền màu
                # đặc (xanh lá) bị mờ khó đọc. Nút nào truyền vào sẽ dùng
                # màu tương phản rõ hơn thay vì xám mặc định.
                kw["text_color_disabled"] = text_color_disabled
            b = CButton(bar, text=text, command=cmd, **kw)
            b.pack(side=tk.LEFT, padx=(0, 4), pady=8)
            return b

        # Padding trái
        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_open = _ab_btn("📂  ①  Chọn file Excel", self._open_file,
            fg=("#FFFFFF","#FFFFFF"), bg=("#D97706","#B45309"),
            hover=("#B45309","#92400E"))

        # Batch — tạo nhưng không pack (ẩn mặc định)
        self.btn_batch = CButton(bar,
            text="📚  Batch nhiều file", command=self._open_multiple_files,
            font=ctk.CTkFont("Segoe UI", 11, "bold"),
            fg_color=("#D97706","#B45309"), text_color="#FFFFFF",
            hover_color=("#B45309","#92400E"),
            corner_radius=6, height=32)

        # Separator
        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=10)

        self.btn_validate = _ab_btn("🔍  ②  Kiểm tra", self._run_validate,
            fg=("#3B82F6","#60A5FA"), bg="transparent",
            hover=("#DBEAFE","#1E3A5F"), border=1, state="disabled")
        self.btn_report = _ab_btn("📋  Báo cáo lỗi", self._export_validate_report,
            fg=("#3B82F6","#60A5FA"), bg="transparent",
            hover=("#DBEAFE","#1E3A5F"), border=1, state="disabled")

        # Separator
        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=10)

        # Separator
        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=10)

        # Nhân viên label
        ctk.CTkLabel(bar, text="Nhân viên:",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray55"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0, 4))

        self.cmb_creator = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=200, height=32,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._on_creator_change,
            placeholder="— Chọn nhân viên —")
        self.cmb_creator.pack(side=tk.LEFT, padx=(0, 6), pady=8)

        # Separator
        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=6, pady=10)

        self.btn_import = _ab_btn("🚀  ③  Import vào BRAVO", self._start_import,
            fg=("#FFFFFF","#FFFFFF"), bg=("#16A34A","#16A34A"),
            hover=("#15803D","#15803D"), state="disabled",
            text_color_disabled="#1A1A1A")

        self.btn_undo_import = _ab_btn("↩  Hoàn tác import", self._undo_last_import,
            fg=("#F87171","#F87171"), bg="transparent",
            hover=("#FEE2E2","#3F1D1D"), border=1, state="disabled")
        Tooltip(self.btn_undo_import,
                lambda: "Xóa BOM vừa import khỏi BRAVO (gọi SP usp_BOMTool_DeleteBOM"
                        " — khách hàng cần deploy SP trước)")

        # ── Body: sheet list (trái) | treeview (phải) ────────────────────────
        body = tk.Frame(tab, bg="#1E1E1E")
        body.pack(fill=tk.BOTH, expand=True)

        # Sheet list
        sb_frame = tk.Frame(body, bg="#161616", width=170)
        sb_frame.pack(side=tk.LEFT, fill=tk.Y)
        sb_frame.pack_propagate(False)

        tk.Label(sb_frame, text="BẢNG DỮ LIỆU",
            bg="#161616", fg="#555555",
            font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=12, pady=(8, 4))

        lb_wrap = tk.Frame(sb_frame, bg="#161616")
        lb_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self.listbox = tk.Listbox(lb_wrap,
            bg="#161616", fg="#CCCCCC",
            selectbackground="#37373D", selectforeground="#CCCCCC",
            font=("Segoe UI", 11),
            relief=tk.FLAT, bd=0, highlightthickness=1,
            highlightcolor="#007ACC", highlightbackground="#2A2A2A",
            activestyle="none")
        self.listbox.pack(fill=tk.BOTH, expand=True)
        self.listbox.bind("<<ListboxSelect>>", self._on_select)

        # Separator
        tk.Frame(body, bg="#2A2A2A", width=1).pack(side=tk.LEFT, fill=tk.Y)

        # Content phải — dùng grid để _tf_outer luôn giữ expand (fix horizontal scrollbar)
        right = ctk.CTkFrame(body, fg_color="transparent")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8, pady=8)
        right.rowconfigure(0, weight=0)   # lbl_table
        right.rowconfigure(1, weight=0)   # _warn_outer (ẩn/hiện)
        right.rowconfigure(2, weight=1)   # _tf_outer — luôn chiếm phần còn lại
        right.columnconfigure(0, weight=1)

        self.lbl_table = CLabel(right,
            text="— Chưa có dữ liệu —",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=("gray50","gray60"),
            fg_color="transparent", anchor="w")
        self.lbl_table.grid(row=0, column=0, sticky="ew", pady=(0, 2))

        # ── Warning panel (thay thế lbl_errors) ─────────────────────────────
        self._warn_outer = tk.Frame(right, bg="#1E1E1E")
        # Đặt vào grid row=1, ẩn ngay bằng grid_remove() (giữ cấu hình grid)
        self._warn_outer.grid(row=1, column=0, sticky="ew")
        self._warn_outer.grid_remove()   # ẩn khi chưa có lỗi

        # header: badge + text tóm tắt + toggle
        _wh = tk.Frame(self._warn_outer, bg="#251515", pady=2)
        _wh.pack(fill=tk.X)

        self._warn_badge = tk.Label(_wh, text="", bg="#7B1A1A", fg="#FFCCCC",
            font=("Segoe UI", 10, "bold"), padx=8, pady=0, cursor="hand2")
        self._warn_badge.pack(side=tk.LEFT, padx=(4, 6))

        self._warn_hdr_lbl = tk.Label(_wh, text="", bg="#251515", fg="#CC8888",
            font=("Segoe UI", 10), anchor="w")
        self._warn_hdr_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self._warn_expanded = True
        self._warn_toggle = tk.Label(_wh, text="▾", bg="#251515", fg="#666666",
            font=("Segoe UI", 11), cursor="hand2", padx=6)
        self._warn_toggle.pack(side=tk.RIGHT)
        self._warn_toggle.bind("<Button-1>", lambda e: self._toggle_warn_panel())
        self._warn_badge.bind("<Button-1>",  lambda e: self._toggle_warn_panel())

        # list area (collapsible)
        self._warn_list_outer = tk.Frame(self._warn_outer, bg="#1A1212")
        self._warn_list_outer.pack(fill=tk.X)

        self._warn_err_tree = ttk.Treeview(
            self._warn_list_outer,
            columns=("rownum", "field", "msg"),
            show="", height=4, selectmode="browse")
        self._warn_err_tree.column("rownum", width=70,  minwidth=60,  anchor="w", stretch=False)
        self._warn_err_tree.column("field",  width=90,  minwidth=70,  anchor="w", stretch=False)
        self._warn_err_tree.column("msg",    width=500, minwidth=200, anchor="w", stretch=True)

        _wsb = ttk.Scrollbar(self._warn_list_outer, orient="vertical",
            command=self._warn_err_tree.yview)
        self._warn_err_tree.configure(yscrollcommand=_wsb.set)
        self._warn_err_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
        _wsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._warn_err_tree.tag_configure("err_item",  foreground="#F87171", background="#1E1212")
        self._warn_err_tree.tag_configure("warn_item", foreground="#FCD34D", background="#1E1C12")
        self._warn_err_tree.bind("<ButtonRelease-1>", self._on_warn_item_click)

        # map: warn_tree iid → main tree iid (populated trong _show_table)
        self._warn_iid_to_tree_iid = {}

        # Empty state overlay
        self._tf_outer = tk.Frame(right, bg="#1E1E1E")
        self._tf_outer.grid(row=2, column=0, sticky="nsew", pady=(4, 0))

        # Nền #000000 — KHÔNG dùng #1E1E1E (màu nền chuẩn của app) vì khung
        # này nằm đè lên canvas tksheet (SheetTable theme="dark" có
        # table_bg=#000000, đen tuyền chứ không phải #1E1E1E) — lệch màu sẽ
        # hiện thành 1 khung xám nổi bật giữa nền đen, nhìn rất lạc quẻ.
        _EMPTY_BG = "#000000"
        # fill toàn bộ _tf_outer để che hoàn toàn tksheet bên dưới
        self._empty_frame = tk.Frame(self._tf_outer, bg=_EMPTY_BG)
        self._empty_frame.place(x=0, y=0, relwidth=1, relheight=1)
        # sub-frame chứa nội dung, căn giữa bên trong _empty_frame
        _ec = tk.Frame(self._empty_frame, bg=_EMPTY_BG)
        _ec.place(relx=0.5, rely=0.45, anchor="center")
        tk.Label(_ec, text="📂", bg=_EMPTY_BG, fg="#3E3E42",
                 font=("Segoe UI", 48)).pack()
        tk.Label(_ec,
                 text="①  Chọn file Excel ở sidebar để bắt đầu",
                 bg=_EMPTY_BG, fg="#555555",
                 font=("Segoe UI", 13)).pack(pady=(4, 0))
        tk.Label(_ec,
                 text="Chọn Nhân viên  →  ②  Kiểm tra  →  ③  Import vào BRAVO",
                 bg=_EMPTY_BG, fg="#3E3E42",
                 font=("Segoe UI", 11)).pack(pady=(2, 0))

        tf = self._tf_outer
        if _HAS_TKSHEET:
            self.tree = SheetTable(tf)
            self.tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
            hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
            self.tree = ttk.Treeview(tf, style="BOM.Treeview", show="headings",
                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.configure(command=self.tree.yview)
            hsb.configure(command=self.tree.xview)
            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)

        # self.tree được tạo SAU _empty_frame nên theo thứ tự chồng lớp mặc
        # định của Tk, nó nằm TRÊN _empty_frame và che khuất hoàn toàn màn
        # hình hướng dẫn (dù _empty_frame vẫn còn đó, chỉ là bị đè). Nâng nó
        # lên trên cùng — khi có dữ liệu, nơi khác đã gọi place_forget() để
        # ẩn hẳn empty_frame nên không xung đột.
        self._empty_frame.lift()

        self.tree.tag_configure("sql_names",
            background="#2D2D30", foreground="#9CDCFE",
            font=("Segoe UI", 12, "italic"))
        self.tree.tag_configure("oddrow",  background="#1E1E1E")
        self.tree.tag_configure("evenrow", background="#2A2D2E")
        self.tree.tag_configure("err_row",  background="#261818")
        self.tree.tag_configure("warn_row", background="#232000")
        # Căn giữa riêng hàng phụ SQL (giống THDM); dữ liệu vẫn căn trái
        if hasattr(self.tree, 'center_rows_by_tag'):
            self.tree.center_rows_by_tag("sql_names")

        # Cell tooltip
        self._tree_tip = None
        def _motion(event):
            item = self.tree.identify_row(event.y)
            col  = self.tree.identify_column(event.x)
            if not item or not col:
                _hide(); return
            idx = int(col[1:]) - 1
            vals = self.tree.item(item, "values")
            text = str(vals[idx]) if idx < len(vals) else ""
            if not text or text == "—": _hide(); return
            if self._tree_tip:
                try: self._tree_tip.destroy()
                except: pass
            tip = tk.Toplevel(self)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{event.x_root+16}+{event.y_root+12}")
            tk.Label(tip, text=text, bg="#252526", fg="#CCCCCC",
                     font=("Segoe UI", 12), relief="solid", bd=1,
                     padx=8, pady=4).pack()
            self._tree_tip = tip
        def _hide(*_):
            if getattr(self, "_tree_tip", None):
                try: self._tree_tip.destroy()
                except: pass
                self._tree_tip = None
        if isinstance(self.tree, ttk.Treeview):
            # Tooltip + copy thủ công chỉ cần cho Treeview;
            # SheetTable (tksheet) có sẵn copy Ctrl+C / chuột phải.
            self.tree.bind("<Motion>", _motion)
            self.tree.bind("<Leave>",  _hide)
            self._bind_copy(self.tree, "Data")

    # ── Copy helper ──────────────────────────────────────────────────────────
    def _bind_copy(self, tree, title=""):
        """
        Gắn copy cho bất kỳ ttk.Treeview nào:
          - Click trái vào ô  → copy giá trị ô đó + tooltip "Đã copy"
          - Ctrl+C            → copy dòng đang select (tab-separated, kèm header)
          - Ctrl+A            → select all
          - Right-click       → context menu
        """
        if not isinstance(tree, ttk.Treeview):
            return   # SheetTable (tksheet) đã có copy tích hợp
        _tip_job  = [None]   # after() job id
        _tip_wnd  = [None]   # toplevel tooltip window

        # ── Tooltip "Đã copy" ────────────────────────────────────────────────
        def _show_copied_tip(x_root, y_root, text):
            # Huỷ tooltip cũ nếu có
            if _tip_wnd[0]:
                try: _tip_wnd[0].destroy()
                except Exception: pass
            if _tip_job[0]:
                tree.after_cancel(_tip_job[0])

            tip = tk.Toplevel(tree)
            tip.wm_overrideredirect(True)          # không có title bar
            tip.wm_attributes("-topmost", True)
            tip.configure(bg="#094771")

            # Cắt text dài
            display = text if len(text) <= 60 else text[:57] + "..."
            lbl = tk.Label(tip, text=f"✓ Đã copy: {display}",
                           bg="#094771", fg="white",
                           font=("Segoe UI", 10), padx=8, pady=4)
            lbl.pack()

            tip.update_idletasks()
            tip.wm_geometry(f"+{x_root + 12}+{y_root - 30}")
            _tip_wnd[0] = tip

            # Tự đóng sau 1.4 giây
            def _destroy():
                try: tip.destroy()
                except Exception: pass
                _tip_wnd[0] = None
            _tip_job[0] = tree.after(1400, _destroy)

        # ── Click trái vào ô ────────────────────────────────────────────────
        def _on_click(event):
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)   # "#1", "#2", ...
            if not row_id or not col_id:
                return
            try:
                col_idx = int(col_id.lstrip("#")) - 1
                vals = tree.item(row_id, "values")
                if col_idx < 0 or col_idx >= len(vals):
                    return
                cell_val = str(vals[col_idx])
                tree.clipboard_clear()
                tree.clipboard_append(cell_val)
                _show_copied_tip(event.x_root, event.y_root, cell_val)
                # Cell highlight — flash ngắn rồi tự ẩn (tránh để vết khi scroll/click khác)
                try:
                    bbox = tree.bbox(row_id, col_id)
                    if bbox:
                        x, y, w, h = bbox
                        if not hasattr(tree, '_cell_hl'):
                            tree._cell_hl    = tk.Frame(tree,
                                highlightbackground="#007ACC",
                                highlightthickness=2, bd=0, bg="")
                            tree._cell_hl_job = [None]
                        # Hủy job ẩn cũ nếu có
                        if tree._cell_hl_job[0]:
                            tree.after_cancel(tree._cell_hl_job[0])
                        tree._cell_hl.place(x=x, y=y, width=w, height=h)
                        tree._cell_hl.lift()
                        # Tự ẩn sau 500ms
                        def _hide_hl(hl=tree._cell_hl, job=tree._cell_hl_job):
                            try: hl.place_forget()
                            except Exception: pass
                            job[0] = None
                        tree._cell_hl_job[0] = tree.after(500, _hide_hl)
                except Exception:
                    pass
            except Exception:
                pass

        # ── Copy dòng được chọn ─────────────────────────────────────────────
        def _rows_to_text(iids):
            cols = tree["columns"]
            lines = ["\t".join(str(tree.heading(c)["text"]) for c in cols)]
            for iid in iids:
                vals = tree.item(iid, "values")
                lines.append("\t".join(str(v) for v in vals))
            return "\n".join(lines)

        def _copy_selected(event=None):
            sel = tree.selection()
            if not sel:
                return
            text = _rows_to_text(sel)
            tree.clipboard_clear()
            tree.clipboard_append(text)

        def _copy_all(event=None):
            all_iids = tree.get_children()
            if not all_iids:
                return
            text = _rows_to_text(all_iids)
            tree.clipboard_clear()
            tree.clipboard_append(text)

        def _select_all(event=None):
            tree.selection_set(tree.get_children())
            return "break"

        # ── Right-click menu ─────────────────────────────────────────────────
        def _show_menu(event):
            menu = tk.Menu(tree, tearoff=0,
                           bg="#2D2D30", fg="#CCCCCC",
                           activebackground="#094771", activeforeground="white",
                           relief="flat", bd=0)
            menu.add_command(label="Copy ô đang click        (click trái)", command=lambda: None)
            menu.add_separator()
            menu.add_command(label="Copy dòng được chọn  Ctrl+C", command=_copy_selected)
            menu.add_command(label="Copy tất cả             Ctrl+A → Ctrl+C", command=_copy_all)
            menu.add_separator()
            menu.add_command(label="Chọn tất cả              Ctrl+A", command=_select_all)
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        tree.bind("<Button-1>",  _on_click)
        tree.bind("<Control-c>", _copy_selected)
        tree.bind("<Control-C>", _copy_selected)
        tree.bind("<Control-a>", _select_all)
        tree.bind("<Control-A>", _select_all)
        tree.bind("<Button-3>",  _show_menu)

    def _open_mapping_window(self):
        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        dlg.title("Mapping Config")
        dlg.geometry("1200x680")
        dlg.transient(self)
        dlg.grab_set()
        self._build_mapping_content(dlg)
        dlg.update_idletasks()
        mx = self.winfo_rootx() + (self.winfo_width()  - dlg.winfo_width())  // 2
        my = self.winfo_rooty() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"1200x680+{mx}+{my}")
        dlg.deiconify()
        dlg.focus_force()

    def _build_mapping_content(self, parent):
        bar = ctk.CTkFrame(parent, corner_radius=0, height=48)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.pack(fill=tk.X, padx=12, pady=8)

        ctk.CTkLabel(inner, text="Phần:",
            font=ctk.CTkFont("Segoe UI", 12, "bold")).pack(side=tk.LEFT, padx=(0,8))

        _sect_names = list(self.mapping.get('_CONFIG', {}).keys())
        self._map_section_var = tk.StringVar(
            value=_sect_names[0] if _sect_names else "")
        section_cb = ttk.Combobox(inner, textvariable=self._map_section_var,
            values=_sect_names, state="readonly", width=22,
            font=("Segoe UI", 12))
        section_cb.pack(side=tk.LEFT)
        section_cb.bind("<<ComboboxSelected>>", lambda _: self._refresh_mapping_tab())

        for txt, color in [("✅ OK","#4EC9B0"), ("⚠️ Thiếu","#D7BA7D"),
                            ("❓ Cần xác nhận","#F97316"), ("🔧 System","gray")]:
            ctk.CTkLabel(inner, text=txt,
                font=ctk.CTkFont("Segoe UI", 12),
                text_color=color).pack(side=tk.RIGHT, padx=6)

        MAP_COLS   = ["SQL Column","Tên trên Excel","Kiểu DL","Bắt buộc",
                      "Nguồn DL","Bảng Master","Điều kiện","Kiểu Lookup","Ghi chú"]
        MAP_WIDTHS = [170,220,90,80,90,130,220,130,280]

        tf = tk.Frame(parent, bg="#1E1E1E")
        tf.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        if _HAS_TKSHEET:
            self.map_tree = SheetTable(tf, columns=MAP_COLS)
            self.map_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL)
            hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL)
            self.map_tree = ttk.Treeview(tf, style="BOM.Treeview", show="headings",
                columns=MAP_COLS, yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.configure(command=self.map_tree.yview)
            hsb.configure(command=self.map_tree.xview)
            self.map_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)
        for col, w in zip(MAP_COLS, MAP_WIDTHS):
            self.map_tree.heading(col, text=col)
            self.map_tree.column(col, width=w, anchor="w", stretch=(col=="Ghi chú"))
        self.map_tree.tag_configure("sys",  foreground="gray")
        self.map_tree.tag_configure("miss", foreground="#D7BA7D")
        self.map_tree.tag_configure("conf", foreground="#F97316")
        self.map_tree.tag_configure("ok",   foreground="#CCCCCC")
        self._bind_copy(self.map_tree, "Mapping")

        self._map_info_lbl = tk.Label(parent, text="",
            bg="#1E1E1E", fg="gray", font=("Segoe UI", 12), anchor="w")
        self._map_info_lbl.pack(fill=tk.X, padx=10, pady=(0,4))
        self._refresh_mapping_tab()

    def _refresh_mapping_tab(self):
        section = self._map_section_var.get()
        records = self.mapping.get(section, [])

        self.map_tree.delete(*self.map_tree.get_children())
        n_ok = n_miss = n_conf = n_sys = 0

        for r in records:
            nguon  = r["nguon_dl"]
            kl     = r.get("kieu_lookup", "")
            kieu   = r["kieu_dl"] + (f"({r['do_dai']})" if r["do_dai"] else "")

            # Assign tag theo Nguon_DL
            if nguon in ("HeThong", "CoDinh"):
                tag = "sys";  n_sys  += 1
            elif nguon == "TinhToan":
                tag = "conf"; n_conf += 1
            elif nguon == "SP":
                tag = "miss"; n_miss += 1
            else:
                tag = "ok";   n_ok   += 1

            vals = (r["sql_col"], r["ten_excel"], kieu, r["bat_buoc"],
                    nguon, r["bang_master"], r["dieu_kien_master"],
                    kl, r["ghi_chu"])
            self.map_tree.insert("", tk.END, values=vals, tags=(tag,))

        total = len(records)
        self._map_info_lbl.config(
            text=("Section: " + section + "  |  Tong: " + str(total) + " cot"
                  + "  |  OK: " + str(n_ok)
                  + "  |  Thieu info: " + str(n_miss)
                  + "  |  Can xac nhan: " + str(n_conf)
                  + "  |  System: " + str(n_sys)))

    # ─ Tab 3: Tra cứu Danh mục ──────────────────────────────────────────────────────
    def _build_tab_catalog(self):
        tab = self.nb.tab(TAB_CATALOG)

        # ── Action bar ngang ─────────────────────────────────────────────────
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=48,
                           corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_catalog_load = ctk.CTkButton(bar, text="🔄  Tải dữ liệu",
            command=self._catalog_load,
            font=ctk.CTkFont("Segoe UI", 11, "bold"),
            fg_color=("#1D4ED8","#1D4ED8"),
            hover_color=("#1E40AF","#1E40AF"),
            height=32, corner_radius=6)
        self.btn_catalog_load.pack(side=tk.LEFT, padx=(0,4), pady=8)

        self.btn_catalog_export = CButton(bar, text="📤  Export Excel",
            command=self._catalog_export_excel,
            font=ctk.CTkFont("Segoe UI", 11, "bold"),
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#60A5FA"),
            hover_color=("#DBEAFE","#1E3A5F"),
            height=32, corner_radius=6, state="disabled")
        self.btn_catalog_export.pack(side=tk.LEFT, padx=(0,8), pady=8)

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=28).pack(side=tk.LEFT, padx=(0,8), pady=10)

        self.catalog_search_var = tk.StringVar()
        self.catalog_search_var.trace_add("write", lambda *_: self._catalog_filter())
        ctk.CTkEntry(bar, textvariable=self.catalog_search_var,
            font=ctk.CTkFont("Segoe UI", 11),
            placeholder_text="🔍  Tìm mã / tên...",
            width=220, height=32, corner_radius=6).pack(side=tk.LEFT, pady=8)

        self.lbl_catalog_status = CLabel(bar, text="—  Chưa tải",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_catalog_status.pack(side=tk.RIGHT, padx=(0,12))

        # ── Treeview ─────────────────────────────────────────────────────────
        self._catalog_all_rows = []
        tf = tk.Frame(tab, bg="#1E1E1E")
        tf.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0,8))
        COLS = ("Mã Vật Tư","Tên Vật Tư","DVT")
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
        self.catalog_tree = ttk.Treeview(tf, style="BOM.Treeview", show="tree headings",
            columns=COLS, yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.configure(command=self.catalog_tree.yview)
        hsb.configure(command=self.catalog_tree.xview)
        # Cột cây (#0) — mũi tên expand/collapse + icon folder.
        # Phải đủ rộng cho node cấp 2–3 (mỗi cấp thụt lề 12px + mũi tên + icon 16px),
        # hẹp quá icon bị ép dính vào cột Mã Vật Tư.
        self.catalog_tree.column("#0",         width=72,  minwidth=36,  stretch=False)
        self.catalog_tree.heading("#0",        text="")
        self.catalog_tree.heading("Mã Vật Tư",  text="Mã Vật Tư",  anchor="w")
        self.catalog_tree.heading("Tên Vật Tư", text="Tên Vật Tư", anchor="w")
        self.catalog_tree.heading("DVT",        text="DVT",        anchor="center")
        self.catalog_tree.column("Mã Vật Tư",  width=170, anchor="w",      stretch=False)
        self.catalog_tree.column("Tên Vật Tư", width=500, anchor="w",      stretch=True)
        self.catalog_tree.column("DVT",        width=70,  anchor="center", stretch=False)
        # Style: nhóm = teal bold, nhóm inactive = xám italic, lá = màu text thường
        self.catalog_tree.tag_configure("group", foreground="#4EC9B0",
                                        font=("Segoe UI", 10, "bold"))
        self.catalog_tree.tag_configure("group_inactive", foreground="#5A7070",
                                        font=("Segoe UI", 10, "italic"))
        self.catalog_tree.tag_configure("leaf",  foreground="#CCCCCC")
        self.catalog_tree.tag_configure("orphan_group", foreground="#7A7A7A",
                                        font=("Segoe UI", 10, "italic"))
        self.catalog_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)
        self.catalog_tree.bind("<Double-1>", self._catalog_copy_code)
        self.catalog_tree.bind("<<TreeviewOpen>>",  self._catalog_on_expand)
        self.catalog_tree.bind("<<TreeviewClose>>", self._catalog_on_collapse)
        self._bind_copy(self.catalog_tree, "Catalog")
        self._init_catalog_icons()

    # ── Catalog helper methods ──────────────────────────────────────────────────────────────────

    def _init_catalog_icons(self):
        """Tạo icon 16×16 cho catalog tree bằng tk.PhotoImage thuần —
        KHÔNG cần PIL (bản cũ dùng PIL nên máy thiếu Pillow là mất icon)."""
        self._icons_loaded = False
        try:
            def _img(rects):
                im = tk.PhotoImage(width=16, height=16)
                for (x0, y0, x1, y1, color) in rects:
                    im.put(color, to=(x0, y0, x1 + 1, y1 + 1))
                return im

            # Folder đóng (amber đậm)
            self._icon_folder_closed = _img([
                (0, 2, 5, 4,  "#C8881A"),   # tab góc trên-trái
                (0, 5, 15, 13, "#9A6010"),  # viền
                (1, 6, 14, 12, "#E8A828"),  # thân
                (1, 6, 14, 6,  "#F0C060"),  # viền trên sáng
            ])
            # Folder mở (amber sáng)
            self._icon_folder_open = _img([
                (0, 2, 5, 4,  "#E0A030"),
                (0, 5, 15, 13, "#9A6010"),
                (1, 6, 14, 12, "#F0C050"),
                (1, 6, 14, 6,  "#F8D878"),
            ])
            # Chồng tài liệu (nhóm "Mã lẻ không thuộc nhóm")
            self._icon_orphan = _img([
                (3, 3, 13, 13, "#555555"),   # bóng sau
                (1, 1, 11, 12, "#888888"),   # thân
                (3, 3, 9, 3,   "#BBBBBB"),   # dòng kẻ
                (3, 6, 9, 6,   "#BBBBBB"),
                (3, 9, 9, 9,   "#BBBBBB"),
            ])
            self._icons_loaded = True
        except Exception:
            pass  # môi trường lạ → chạy bình thường không icon

    def _catalog_on_collapse(self, event=None):
        """Đổi icon thư mục từ mở → đóng khi user collapse node."""
        if not getattr(self, '_icons_loaded', False):
            return
        iid = self.catalog_tree.focus()
        if not iid or iid == "_orphan_":
            return
        try:
            int(iid)
        except ValueError:
            return
        row = self._catalog_id_to_row.get(int(iid))
        if row and row['is_group']:
            self.catalog_tree.item(iid, image=self._icon_folder_closed)

    def _catalog_read_filter_ids(self):
        """
        Đọc sheet CATALOG_FILTER từ CK_Mapping_v2.xlsx.
        Trả về list[int] các RootId được khai báo.
        Nếu sheet không có hoặc rỗng → trả về [] (không lọc = load tất cả).
        Dòng bắt đầu bằng '#' bị bỏ qua (comment).
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True, data_only=True)
            if 'CATALOG_FILTER' not in wb.sheetnames:
                return []
            ws = wb['CATALOG_FILTER']
            ids = []
            for row in ws.iter_rows(min_row=4, values_only=True):  # skip title/header rows
                val = row[0] if row else None
                if val is None:
                    continue
                s = str(val).strip()
                if not s or s.startswith('#'):
                    continue
                try:
                    ids.append(int(float(s)))
                except (ValueError, TypeError):
                    pass
            return ids
        except Exception:
            return []

    def _catalog_load(self):
        """Trigger: disable nút, hiện loading, bắt thread worker."""
        self.lbl_catalog_status.config(text="⏳  Đang tải danh mục...", fg=C["yellow"])
        try:
            self.btn_catalog_load.configure(state="disabled", text="⏳  Đang tải...")
        except Exception:
            pass
        self.update_idletasks()
        threading.Thread(target=self._catalog_load_worker, daemon=True).start()

    def _catalog_load_worker(self):
        """Background thread: kết nối DB, chạy CTE, trả về data qua after()."""
        try:
            filter_ids = self._catalog_read_filter_ids()
            conn = self._get_db_conn()
            cur  = conn.cursor()

            try:
                cur.execute("SELECT TOP 1 Unit FROM B20Item")
                has_unit = True
            except Exception:
                has_unit = False

            unit_anchor    = "Unit"    if has_unit else "N'' AS Unit"
            unit_recursive = "b.Unit"  if has_unit else "N''"

            sort_anchor    = "CAST(RIGHT(REPLICATE('0',10)+CAST(Id AS VARCHAR(10)),10) AS VARCHAR(MAX))"
            sort_recursive = "t.SortPath+'/'+RIGHT(REPLICATE('0',10)+CAST(b.Id AS VARCHAR(10)),10)"

            if filter_ids:
                if len(filter_ids) == 1:
                    sql = (
                        "WITH CatalogTree AS ("
                        f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                        f" {sort_anchor} AS SortPath"
                        " FROM B20Item"
                        " WHERE ParentId=? AND (IsActive=1 OR IsGroup=1)"
                        " UNION ALL"
                        f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                        " t.Lvl+1,"
                        f" {sort_recursive}"
                        " FROM B20Item b"
                        " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                        " WHERE b.IsActive=1 OR b.IsGroup=1"
                        ")"
                        " SELECT Id,ParentId,IsGroup,Code,Name,Unit,IsActive,Lvl"
                        " FROM CatalogTree ORDER BY SortPath OPTION(MAXRECURSION 0)"
                    )
                    cur.execute(sql, filter_ids)
                else:
                    placeholders = ','.join('?' * len(filter_ids))
                    sql = (
                        "WITH CatalogTree AS ("
                        f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                        f" {sort_anchor} AS SortPath"
                        " FROM B20Item"
                        f" WHERE Id IN ({placeholders}) AND (IsActive=1 OR IsGroup=1)"
                        " UNION ALL"
                        f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                        " t.Lvl+1,"
                        f" {sort_recursive}"
                        " FROM B20Item b"
                        " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                        " WHERE b.IsActive=1 OR b.IsGroup=1"
                        ")"
                        " SELECT Id,ParentId,IsGroup,Code,Name,Unit,IsActive,Lvl"
                        " FROM CatalogTree ORDER BY SortPath OPTION(MAXRECURSION 0)"
                    )
                    cur.execute(sql, filter_ids)
            else:
                sql = (
                    "WITH CatalogTree AS ("
                    f" SELECT Id, ParentId, IsGroup, Code, Name, {unit_anchor}, IsActive, 0 AS Lvl,"
                    f" {sort_anchor} AS SortPath"
                    " FROM B20Item"
                    " WHERE (ParentId IS NULL OR ParentId=0) AND (IsActive=1 OR IsGroup=1)"
                    " UNION ALL"
                    f" SELECT b.Id, b.ParentId, b.IsGroup, b.Code, b.Name, {unit_recursive}, b.IsActive,"
                    " t.Lvl+1,"
                    f" {sort_recursive}"
                    " FROM B20Item b"
                    " INNER JOIN CatalogTree t ON b.ParentId=t.Id"
                    " WHERE b.IsActive=1 OR b.IsGroup=1"
                    ")"
                    " SELECT Id,ParentId,IsGroup,Code,Name,Unit,IsActive,Lvl"
                    " FROM CatalogTree ORDER BY SortPath OPTION(MAXRECURSION 0)"
                )
                cur.execute(sql)

            rows = cur.fetchall()
            conn.close()

            all_rows = [
                {
                    'id'       : r[0],
                    'parent_id': r[1] or 0,
                    'is_group' : bool(r[2]),
                    'code'     : (r[3] or '').strip(),
                    'name'     : (r[4] or '').strip(),
                    'unit'     : (r[5] or '').strip(),
                    'active'   : bool(r[6]),
                    'lvl'      : r[7],
                }
                for r in rows
            ]
            self.after(0, lambda d=all_rows, f=filter_ids: self._catalog_load_done(d, f, None))
        except Exception as e:
            self.after(0, lambda err=e: self._catalog_load_done(None, [], err))

    def _catalog_load_done(self, all_rows, filter_ids, error):
        """Main thread callback: cập nhật UI sau khi worker hoàn tất."""
        try:
            self.btn_catalog_load.configure(state="normal", text="🔄  Tải dữ liệu")
        except Exception:
            pass

        if error:
            if isinstance(error, ConnectionError):
                self.lbl_catalog_status.config(text="❌  VPN/mạng lỗi", fg=C["red"])
                self._show_msg("Lỗi kết nối DB", str(error))
            else:
                self.lbl_catalog_status.config(text=f"❌  Lỗi: {error}", fg=C["red"])
                self._show_msg("Lỗi tải danh mục", str(error))
            return

        self._catalog_all_rows = all_rows
        self._catalog_filter()
        total = sum(1 for r in all_rows if r['active'])
        filter_note = f"  (lọc {len(filter_ids)} nhóm)" if filter_ids else ""
        self.lbl_catalog_status.config(
            text=f"✅  {total:,} vật tư{filter_note}", fg=C["green"])
        try:
            self.btn_catalog_export.configure(state="normal")
        except Exception:
            pass
    def _catalog_filter(self):
        """
        Không có keyword → dựng cây lazy (cha trước con, theo SortPath từ SQL).
        Có keyword   → flat list filter theo code/name.
        """
        if not hasattr(self, '_catalog_all_rows'):
            return
        q    = self.catalog_search_var.get().strip().lower() if hasattr(self, 'catalog_search_var') else ''
        tree = self.catalog_tree
        tree.delete(*tree.get_children())
        total = sum(1 for r in self._catalog_all_rows if r['active'])

        if q:
            # Flat search: hiện active nodes khớp keyword
            shown = 0
            for row in self._catalog_all_rows:
                if not row.get('active', True):
                    continue
                if q not in row['code'].lower() and q not in row['name'].lower():
                    continue
                tag = "group" if row['is_group'] else "leaf"
                _fic = {}
                if row['is_group'] and getattr(self, '_icons_loaded', False):
                    _fic = {"image": self._icon_folder_closed}
                tree.insert("", "end",
                            values=(row['code'], row['name'], row['unit']),
                            tags=(tag,), **_fic)
                shown += 1
                if shown >= 3000:
                    break
            self.lbl_catalog_status.config(
                text=f"\U0001f50d  {shown:,} / {total:,} kết quả", fg=C["text"])
        else:
            # Tree view: data SQL đã sort đúng thứ tự → build cây 1 lần
            id_set = {row['id'] for row in self._catalog_all_rows}
            children_map: dict = {}
            for row in self._catalog_all_rows:
                pid = row['parent_id']
                if pid and pid in id_set:
                    children_map.setdefault(pid, []).append(row)
            # Nhóm (IsGroup) luôn hiện trước mã lẻ trong mỗi cấp
            def _sort_key(r):
                return (0 if r['is_group'] else 1, r['code'])
            for lst in children_map.values():
                lst.sort(key=_sort_key)
            self._catalog_children_map = children_map
            self._catalog_id_to_row    = {row['id']: row for row in self._catalog_all_rows}

            # Bottom-up leaf count: SQL sort parent-trước-con
            # → reversed() đảm bảo con được xử lý trước cha
            _leaf_count: dict = {}
            for row in reversed(self._catalog_all_rows):
                if not row['is_group']:
                    _leaf_count[row['id']] = 1
                else:
                    _leaf_count[row['id']] = sum(
                        _leaf_count.get(c['id'], 0)
                        for c in children_map.get(row['id'], [])
                    )
            self._catalog_leaf_count = _leaf_count

            # Root = node có parent_id không nằm trong id_set
            _roots = [r for r in self._catalog_all_rows
                      if not r['parent_id'] or r['parent_id'] not in id_set]
            _roots.sort(key=_sort_key)

            orphan_leaves = []
            group_count   = 0
            for row in _roots:
                if not row['is_group']:
                    # Mã lẻ không thuộc nhóm nào → gom lại cuối danh sách
                    orphan_leaves.append(row)
                    continue
                group_count += 1
                iid   = str(row['id'])
                cnt   = _leaf_count.get(row['id'], 0)
                name_disp = f"{row['name']}  ({cnt:,})" if cnt else row['name']
                tag   = "group" if row.get('active', True) else "group_inactive"
                _icon_kw = {"image": self._icon_folder_closed} if getattr(self, '_icons_loaded', False) else {}
                tree.insert("", "end", iid=iid, text="",
                            values=(row['code'], name_disp, row['unit']),
                            open=False, tags=(tag,), **_icon_kw)
                if row['id'] in children_map:
                    tree.insert(iid, "end", iid=f"_ph_{iid}",
                                values=("", "\u2026", ""), tags=())

            # Nhóm ảo cuối cùng: gom mã lẻ không thuộc nhóm nào
            self._catalog_orphan_leaves = orphan_leaves
            if orphan_leaves:
                _orphan_icon_kw = {"image": self._icon_orphan} if getattr(self, '_icons_loaded', False) else {}
                tree.insert("", "end", iid="_orphan_", text="",
                            values=("—",
                                    f"Mã lẻ (không thuộc nhóm)  ({len(orphan_leaves):,})",
                                    ""),
                            open=False, tags=("orphan_group",), **_orphan_icon_kw)
                tree.insert("_orphan_", "end", iid="_ph__orphan_",
                            values=("", "\u2026", ""), tags=())

            n_leaves = sum(1 for r in self._catalog_all_rows if not r['is_group'])
            self.lbl_catalog_status.config(
                text=f"\U0001f4c2  {group_count} nhóm  \u2022  {n_leaves:,} vật tư",
                fg=C["text"])

    def _catalog_on_expand(self, event=None):
        """Lazy-load children khi user click mở rộng một node trong catalog tree."""
        iid = self.catalog_tree.focus()
        if not iid:
            return
        children = self.catalog_tree.get_children(iid)
        # Chỉ xử lý nếu child đầu tiên là placeholder
        if not children or not str(children[0]).startswith("_ph_"):
            return
        # Xóa placeholder, insert real children
        self.catalog_tree.delete(children[0])

        # Swap icon parent: đóng → mở
        if getattr(self, '_icons_loaded', False) and iid != "_orphan_":
            try:
                int(iid)
                self.catalog_tree.item(iid, image=self._icon_folder_open)
            except ValueError:
                pass

        # Xử lý nhóm ảo "Mã lẻ không thuộc nhóm"
        if iid == "_orphan_":
            for leaf_row in getattr(self, '_catalog_orphan_leaves', []):
                leaf_iid = str(leaf_row['id'])
                self.catalog_tree.insert("_orphan_", "end", iid=leaf_iid, text="",
                    values=(leaf_row['code'], leaf_row['name'], leaf_row['unit']),
                    open=False, tags=("leaf",))
            return

        try:
            row_id = int(iid)
        except ValueError:
            return
        cmap = getattr(self, '_catalog_children_map', {})
        for child_row in cmap.get(row_id, []):
            child_iid = str(child_row['id'])
            if child_row['is_group']:
                tag = "group" if child_row.get('active', True) else "group_inactive"
                _cicon = {"image": self._icon_folder_closed} if getattr(self, '_icons_loaded', False) else {}
            else:
                tag = "leaf"
                _cicon = {}
            self.catalog_tree.insert(iid, "end", iid=child_iid, text="",
                values=(child_row['code'], child_row['name'], child_row['unit']),
                open=False, tags=(tag,), **_cicon)
            # Nếu child cũng có con → thêm placeholder
            if child_row['id'] in cmap:
                self.catalog_tree.insert(child_iid, "end",
                    iid=f"_ph_{child_iid}",
                    values=("", "\u2026", ""), tags=())

    def _catalog_export_excel(self):
        """Xuất toàn bộ danh mục ra file .xlsx."""
        if not hasattr(self, '_catalog_all_rows') or not self._catalog_all_rows:
            self._show_msg("Chưa có dữ liệu", "Hãy bấm 'Tải dữ liệu' trước.", 'warning')
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="DanhMucVatTu.xlsx")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "B20Item"
            ws.append(["Mã Vật Tư", "Tên Vật Tư", "ĐVT"])
            ws.row_dimensions[1].height = 20
            hdr_fill = PatternFill("solid", fgColor="2B7A78")
            hdr_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for row in self._catalog_all_rows:
                ws.append([row['code'], row['name'], row['unit']])
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 12
            wb.save(path)
            self._show_export_success(
                "Xuất Excel thành công",
                f"Đã xuất {len(self._catalog_all_rows):,} vật tư\n\u2192 {path}",
                path)
        except Exception as e:
            self._show_msg("Lỗi xuất Excel", str(e))

    def _catalog_copy_code(self, event):
        """Double-click dòng -> copy Mã Vật Tư vào clipboard."""
        sel = self.catalog_tree.selection()
        if not sel:
            return
        code = self.catalog_tree.item(sel[0], "values")[0]
        self.clipboard_clear()
        self.clipboard_append(code)
        self.lbl_catalog_status.config(
            text=f"📋  Đã copy: {code}", fg=C["accent"])

    # ─ Tab 3: Tổng hợp THDM ───────────────────────────────────────────────────────────────────

    def _on_tab_changed(self):
        """Khi switch sang tab THDM mà chưa có dữ liệu → tự tải.
        Luôn đóng tất cả _SearchCombo popup để tránh ghost dropdown khi chuyển tab."""
        # Đóng tất cả popup đang mở
        for attr in ('cmb_creator', 'cmb_thdm_product', 'cmb_thdm_order',
                     'cmb_thdm_creator', 'cmb_thdm_period'):
            combo = getattr(self, attr, None)
            if combo is not None:
                try:
                    combo._close_popup()
                except Exception:
                    pass
        try:
            tab_name = self.nb.get()
        except Exception:
            return
        # So khớp CHÍNH XÁC với TAB_THDM (không dùng substring "THDM in ..."
        # — dễ vỡ mỗi khi đổi tên tab, xem ghi chú tại TAB_THDM đầu file)
        if tab_name == TAB_THDM and not self._thdm_product_map:
            self._thdm_load_products()

    def _build_tab_thdm(self):
        tab = self.nb.tab(TAB_THDM)

        # ── Action bar ngang ─────────────────────────────────────────────────
        bar = ctk.CTkFrame(tab, fg_color=("gray88","gray16"), height=56,
                           corner_radius=0)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        ctk.CTkFrame(bar, fg_color="transparent", width=8).pack(side=tk.LEFT)

        self.btn_thdm_load = ctk.CTkButton(bar, text="🔄  ① Tải dữ liệu",
            command=self._thdm_load_products,
            font=ctk.CTkFont("Segoe UI", 11, "bold"),
            fg_color=("#1D4ED8","#1D4ED8"),
            hover_color=("#1E40AF","#1E40AF"),
            height=32, corner_radius=6)
        self.btn_thdm_load.pack(side=tk.LEFT, padx=(0,10), pady=12)

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,10), pady=10)

        # Dự án
        CLabel(bar, text="Dự án:",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_product = _SearchCombo(bar,
            values=["— Chọn dự án —"],
            width=230, height=32,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._thdm_on_product_change,
            placeholder="— Chọn dự án —")
        self.cmb_thdm_product.set("— Chọn dự án —")
        self.cmb_thdm_product.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,10), pady=10)

        # Đơn hàng
        CLabel(bar, text="Đơn hàng:",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_order = _SearchCombo(bar,
            values=["— Chọn đơn hàng —"],
            width=240, height=32,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._thdm_on_order_change,
            placeholder="— Chọn đơn hàng —",
            state="disabled")
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self.cmb_thdm_order.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        CLabel(bar, text="Nhân viên:",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_creator = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=200, height=32,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._on_thdm_creator_change,
            placeholder="— Chọn nhân viên —")
        self.cmb_thdm_creator.set("— Đang tải... —")
        self.cmb_thdm_creator.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        CLabel(bar, text="Đợt:",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray65"),
            fg_color="transparent").pack(side=tk.LEFT, padx=(0,4))
        self.cmb_thdm_period = _SearchCombo(bar,
            values=["— Đang tải... —"],
            width=180, height=32,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._on_thdm_period_change,
            placeholder="— Chọn đợt —")
        self.cmb_thdm_period.set("— Đang tải... —")
        self.cmb_thdm_period.pack(side=tk.LEFT, padx=(0,10))

        ctk.CTkFrame(bar, fg_color=("gray65","gray30"),
                     width=1, height=36).pack(side=tk.LEFT, padx=(0,8), pady=10)

        self.lbl_thdm_status = CLabel(bar, text="—  Chưa tải",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_status.pack(side=tk.LEFT)

        # ── Main area: left (BOM list) | right (preview) ─────────────────────
        main = tk.Frame(tab, bg="#1E1E1E")
        main.pack(fill=tk.BOTH, expand=True)
        main.columnconfigure(0, weight=35, uniform="thdm")
        main.columnconfigure(1, weight=65, uniform="thdm")
        main.rowconfigure(0, weight=1)

        # ── LEFT: BOM list ────────────────────────────────────────────────────
        lf = tk.Frame(main, bg="#161616")
        lf.grid(row=0, column=0, sticky="nsew")
        lf.rowconfigure(1, weight=1)
        lf.columnconfigure(0, weight=1)

        # Filter
        fe = ctk.CTkFrame(lf, fg_color=("gray85","gray20"), height=32, corner_radius=0)
        fe.grid(row=0, column=0, sticky="ew")
        fe.pack_propagate(False)
        self.thdm_search_var = tk.StringVar()
        self.thdm_search_var.trace_add("write", lambda *_: self._thdm_filter_bom())
        ctk.CTkEntry(fe, textvariable=self.thdm_search_var,
            font=ctk.CTkFont("Segoe UI", 11),
            placeholder_text="🔍  Tìm BOM...",
            height=30, corner_radius=0, border_width=0,
            fg_color="transparent").pack(fill=tk.X, padx=6, pady=1)

        # BOM treeview
        btf = tk.Frame(lf, bg="#161616")
        btf.grid(row=1, column=0, sticky="nsew")
        btf.rowconfigure(0, weight=1)
        btf.columnconfigure(0, weight=1)

        # ── tksheet (Excel-like grid) ─────────────────────────────────────────
        if _HAS_TKSHEET:
            self.thdm_bom_sheet = tksheet.Sheet(
                btf,
                headers=["✓", "Mục số", "Mã BOM", "Tên SP", "Version"],
                data=[],
                theme="dark",
                show_row_index=False,
                show_top_left=False,
                row_height=24,
                header_height=30,
                show_horizontal_grid=True,
                show_vertical_grid=True,
                font=("Segoe UI", 10, "normal"),
                header_font=("Segoe UI", 10, "bold"),
            )
            self.thdm_bom_sheet.enable_bindings(
                "single_select", "drag_select", "row_select",
                "column_width_resize", "arrowkeys",
            )
            # Bind trên MT (main table canvas) — hoạt động độc lập với tksheet version
            self.thdm_bom_sheet.MT.bind(
                "<ButtonRelease-1>",
                lambda e: self.after(40, self._thdm_process_sheet_click)
            )
            self.thdm_bom_sheet.grid(row=0, column=0, columnspan=2, sticky="nsew")
            self.thdm_bom_sheet.column_width(column=0, width=30)
            self.thdm_bom_sheet.column_width(column=1, width=70)
            self.thdm_bom_sheet.column_width(column=2, width=120)
            self.thdm_bom_sheet.column_width(column=3, width=200)
            self.thdm_bom_sheet.column_width(column=4, width=70)
            self.thdm_bom_tree = None   # alias — fallback path sử dụng
            # Cột "Tên SP" tự giãn lấp hết chiều ngang khi resize cửa sổ
            self._thdm_bom_stretch_job = None
            def _bom_sheet_on_resize(_e=None):
                if self._thdm_bom_stretch_job:
                    try:
                        self.after_cancel(self._thdm_bom_stretch_job)
                    except Exception:
                        pass
                self._thdm_bom_stretch_job = self.after(
                    100, self._thdm_bom_autostretch)
            self.thdm_bom_sheet.bind("<Configure>", _bom_sheet_on_resize)
        else:
            # Fallback: ttk.Treeview khi tksheet chưa cài
            BOM_COLS = ("✓", "Mục số", "Mã BOM", "Tên SP", "Version")
            vsb_b = ttk.Scrollbar(btf, orient=tk.VERTICAL, style="Slim.Vertical.TScrollbar")
            self.thdm_bom_tree = ttk.Treeview(btf, style="BOM.Treeview", show="headings",
                columns=BOM_COLS, yscrollcommand=vsb_b.set, selectmode="extended")
            vsb_b.configure(command=self.thdm_bom_tree.yview)
            self.thdm_bom_tree.heading("✓", text="✓", anchor="center")
            self.thdm_bom_tree.column("✓", width=26, minwidth=26, anchor="center", stretch=False)
            for _bc, _ba, _bw, _bs in [
                ("Mục số", "w", 70, False), ("Mã BOM", "w", 120, False),
                ("Tên SP",  "w", 200, True),  ("Version", "w", 70, False),
            ]:
                self.thdm_bom_tree.heading(_bc, text=_bc, anchor=_ba)
                self.thdm_bom_tree.column(_bc, width=_bw, minwidth=50, anchor=_ba, stretch=_bs)
            self.thdm_bom_tree.tag_configure("checked",   foreground="#4EC9B0")
            self.thdm_bom_tree.tag_configure("unchecked", foreground="#AAAAAA")
            self.thdm_bom_tree.grid(row=0, column=0, sticky="nsew")
            vsb_b.grid(row=0, column=1, sticky="ns")
            self.thdm_bom_tree.bind("<ButtonRelease-1>", self._thdm_on_bom_click)
            self.thdm_bom_sheet = None

        # Bottom bar left
        lb = ctk.CTkFrame(lf, fg_color=("gray85","gray18"), height=36, corner_radius=0)
        lb.grid(row=2, column=0, sticky="ew")
        lb.pack_propagate(False)

        self.btn_thdm_sel_all = ctk.CTkButton(lb, text="☑ Tất cả",
            command=lambda: self._thdm_toggle_all(True),
            font=ctk.CTkFont("Segoe UI", 10),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            hover_color=("gray80","gray25"),
            width=80, height=26, corner_radius=5)
        self.btn_thdm_sel_all.pack(side=tk.LEFT, padx=(6,2), pady=5)

        self.btn_thdm_desel_all = ctk.CTkButton(lb, text="☐ Bỏ",
            command=lambda: self._thdm_toggle_all(False),
            font=ctk.CTkFont("Segoe UI", 10),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            hover_color=("gray80","gray25"),
            width=60, height=26, corner_radius=5)
        self.btn_thdm_desel_all.pack(side=tk.LEFT, padx=(0,6), pady=5)

        self.lbl_thdm_sel_count = CLabel(lb, text="0 BOM chọn",
            font=ctk.CTkFont("Segoe UI", 10),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_sel_count.pack(side=tk.LEFT)

        # Xanh dương (hành động xử lý) — xanh lá dành riêng cho nút INSERT
        self.btn_thdm_aggregate = ctk.CTkButton(lb, text="📊  ③ Tổng hợp",
            command=self._thdm_aggregate,
            font=ctk.CTkFont("Segoe UI", 11, "bold"),
            fg_color=("#2563EB","#1D4ED8"),
            hover_color=("#1D4ED8","#1E40AF"),
            height=26, corner_radius=5, state="disabled")
        self.btn_thdm_aggregate.pack(side=tk.RIGHT, padx=6, pady=5)

        # Separator
        tk.Frame(main, bg="#2A2A2A", width=1).grid(
            row=0, column=0, sticky="nse")

        # ── RIGHT: File picker + Preview ─────────────────────────────────────
        rf = tk.Frame(main, bg="#1E1E1E")
        rf.grid(row=0, column=1, sticky="nsew", padx=(4,0))
        rf.rowconfigure(1, weight=1)
        rf.columnconfigure(0, weight=1)

        # File picker row
        fp = ctk.CTkFrame(rf, fg_color=("gray85","gray20"), height=42, corner_radius=0)
        fp.grid(row=0, column=0, sticky="ew", pady=(0, 2))
        fp.pack_propagate(False)

        self.btn_thdm_pick_excel = ctk.CTkButton(fp, text="📂  ② Chọn file Excel THDM",
            command=self._thdm_pick_excel,
            font=ctk.CTkFont("Segoe UI", 11),
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray65"),
            hover_color=("gray80","gray25"),
            height=28, corner_radius=5)
        self.btn_thdm_pick_excel.pack(side=tk.LEFT, padx=(6,8), pady=7)

        ctk.CTkFrame(fp, fg_color=("gray65","gray30"),
                     width=1, height=24).pack(side=tk.LEFT, padx=(0,8), pady=9)

        self.lbl_thdm_excel_path = CLabel(fp, text="Chưa chọn file",
            font=ctk.CTkFont("Segoe UI", 10),
            text_color=("gray40","gray55"),
            fg_color="transparent")
        self.lbl_thdm_excel_path.pack(side=tk.LEFT)

        # ── Content area — switches between pre-preview and result ────────────
        ca = tk.Frame(rf, bg="#1E1E1E")
        ca.grid(row=1, column=0, sticky="nsew")
        ca.rowconfigure(0, weight=1)
        ca.columnconfigure(0, weight=1)
        self._thdm_content_area = ca

        # ── PRE-PREVIEW frame (hiện sau khi chọn file, trước khi Tổng hợp) ──
        pref = tk.Frame(ca, bg="#1E1E1E")
        pref.grid(row=0, column=0, sticky="nsew")
        pref.rowconfigure(1, weight=1)
        pref.columnconfigure(0, weight=1)
        self._thdm_pre_frame = pref

        self._lbl_thdm_pre_status = tk.Label(pref,
            text="⏳  Đang đọc file Excel...",
            font=("Segoe UI", 11), fg="#888888", bg="#1E1E1E", anchor="w")
        self._lbl_thdm_pre_status.grid(row=0, column=0, sticky="ew", padx=12, pady=(8, 4))

        self._thdm_pre_nb = ttk.Notebook(pref)
        self._thdm_pre_nb.grid(row=1, column=0, sticky="nsew", padx=4, pady=(0, 4))

        # Tab 1 — Raw Excel
        raw_tab = tk.Frame(self._thdm_pre_nb, bg="#2D2D2D")
        self._thdm_pre_nb.add(raw_tab, text="  📄 Raw Excel  ")
        raw_tab.rowconfigure(0, weight=1)
        raw_tab.columnconfigure(0, weight=1)
        if _HAS_TKSHEET:
            self.thdm_raw_tree = SheetTable(raw_tab, columns=("_ph",))
            self.thdm_raw_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb_r = ttk.Scrollbar(raw_tab, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
            hsb_r = ttk.Scrollbar(raw_tab, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
            self.thdm_raw_tree = ttk.Treeview(raw_tab, style="BOM.Treeview", show="headings",
                columns=("_ph",), yscrollcommand=vsb_r.set, xscrollcommand=hsb_r.set)
            vsb_r.configure(command=self.thdm_raw_tree.yview)
            hsb_r.configure(command=self.thdm_raw_tree.xview)
            self.thdm_raw_tree.grid(row=0, column=0, sticky="nsew")
            vsb_r.grid(row=0, column=1, sticky="ns")
            hsb_r.grid(row=1, column=0, sticky="ew")

        # (Tab "Đã xử lý" đã bỏ hẳn — chỉ xem Raw trước khi Tổng hợp,
        #  dữ liệu đã xử lý xem ở màn hình kết quả đa tab sau khi Tổng hợp.)

        pref.grid_remove()   # ẩn cho đến khi file được chọn

        # ── RESULT frame (hiển thị sau khi Tổng hợp) ─────────────────────────
        ptf = tk.Frame(ca, bg="#2D2D2D", bd=1, relief="flat")
        ptf.grid(row=0, column=0, sticky="nsew")
        ptf.rowconfigure(0, weight=1)
        ptf.columnconfigure(0, weight=1)
        self._thdm_result_frame = ptf

        # ── Notebook kết quả: Đầu phiếu + mỗi section detail 1 tab (giống BOM)
        self._thdm_result_nb = ttk.Notebook(ptf)
        self._thdm_result_nb.grid(row=0, column=0, sticky="nsew")

        def _mk_result_grid(title, columns):
            tabf = tk.Frame(self._thdm_result_nb, bg="#2D2D2D")
            self._thdm_result_nb.add(tabf, text=title)
            tabf.rowconfigure(0, weight=1)
            tabf.columnconfigure(0, weight=1)
            if _HAS_TKSHEET:
                tr = SheetTable(tabf, columns=columns)
                tr.grid(row=0, column=0, sticky="nsew")
            else:
                _v = ttk.Scrollbar(tabf, orient=tk.VERTICAL,   style="Slim.Vertical.TScrollbar")
                _h = ttk.Scrollbar(tabf, orient=tk.HORIZONTAL, style="Slim.Horizontal.TScrollbar")
                tr = ttk.Treeview(tabf, style="BOM.Treeview", show="headings",
                    columns=columns, yscrollcommand=_v.set, xscrollcommand=_h.set)
                _v.configure(command=tr.yview)
                _h.configure(command=tr.xview)
                tr.grid(row=0, column=0, sticky="nsew")
                _v.grid(row=0, column=1, sticky="ns")
                _h.grid(row=1, column=0, sticky="ew")
            tr.tag_configure("row_normal", foreground="#CCCCCC")
            tr.tag_configure("row_alt",    foreground="#AAAAAA")
            tr.tag_configure("err_row",  background="#3A1414", foreground="#F87171")
            tr.tag_configure("warn_row", background="#2A2410", foreground="#D7BA7D")
            tr.tag_configure("sql_names", background="#2D2D30", foreground="#9CDCFE")
            return tr

        # Tab 0 — Header (THDM_HEADER, resolve display-only)
        # Hiển thị NGANG giống BOM: mỗi field 1 cột, cột gán động khi fill
        self._thdm_header_tree = _mk_result_grid("  📋 Header  ", ())

        # Tab 1..n — mỗi child section của THDM_HEADER một tab
        self._thdm_sec_trees    = {}   # sec → tree
        self._thdm_sec_col_defs = {}   # sec → col defs
        for _sec, _info in self.mapping.get('_CONFIG', {}).items():
            if _info.get('parent_section') != 'THDM_HEADER':
                continue
            # Cột định danh = SQL_Column (duy nhất), header text = Ten_Excel
            _defs = _thdm_get_detail_col_defs(self.mapping.get(_sec, []))
            _cols = tuple(d['sql_col'] for d in _defs)
            _lbl  = _info.get('label') or _sec
            _tr   = _mk_result_grid(f"  {_lbl}  ", _cols)
            for d in _defs:
                # header rỗng → tksheet hiện chữ cái cột (P,Q,R...);
                # cột hệ thống không có Ten_Excel thì dùng SQL_Column làm header
                _htxt = d['ten_excel'] or d['sql_col']
                _tr.heading(d['sql_col'], text=_htxt, anchor='center')
                _tr.column(d['sql_col'], width=d['width'], anchor=d['anchor'],
                           stretch=d['stretch'], minwidth=50)
            # Căn giữa riêng hàng phụ SQL_Column (2 dòng header nhìn cho gọn),
            # dữ liệu vẫn giữ căn lề theo kiểu cột
            if hasattr(_tr, 'center_rows_by_tag'):
                _tr.center_rows_by_tag("sql_names")
            self._thdm_sec_trees[_sec]    = _tr
            self._thdm_sec_col_defs[_sec] = _defs

        # Alias tương thích code cũ: tree chính = section đầu tiên (THDM_THVT)
        self.thdm_preview_tree = (next(iter(self._thdm_sec_trees.values()), None)
                                  or self._thdm_header_tree)

        # Action bar bottom
        ab = ctk.CTkFrame(rf, fg_color="transparent", height=38)
        ab.grid(row=2, column=0, sticky="ew", pady=(4, 0))
        ab.pack_propagate(False)

        # Tạm ẩn "Xuất Excel" / "Xem SQL" theo yêu cầu — user không cần,
        # gây rối vì không nằm trong luồng ①→⑤. Widget vẫn tạo (không .pack())
        # để các chỗ .configure(state=...) khác trong code không lỗi;
        # muốn hiện lại chỉ cần thêm .pack(...) như 2 nút bên dưới.
        self.btn_thdm_export_xl = CButton(ab, text="📤  Xuất Excel",
            command=self._thdm_export_excel,
            font=ctk.CTkFont("Segoe UI", 12),
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            hover_color=("gray90", "gray25"),
            width=120, height=30, corner_radius=6, state="disabled")

        self.btn_thdm_view_sql = CButton(ab, text="📋  Xem SQL",
            command=self._thdm_view_sql,
            font=ctk.CTkFont("Segoe UI", 12),
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            hover_color=("gray90", "gray25"),
            width=100, height=30, corner_radius=6, state="disabled")

        self.btn_thdm_validate = CButton(ab, text="🔍  ④ Kiểm tra",
            command=self._thdm_validate,
            font=ctk.CTkFont("Segoe UI", 12, "bold"),
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6", "#60A5FA"),
            hover_color=("#DBEAFE", "#1E3A5F"),
            width=110, height=30, corner_radius=6, state="disabled")
        self.btn_thdm_validate.pack(side=tk.LEFT, padx=(0, 4))
        Tooltip(self.btn_thdm_validate,
                lambda: "Soát dòng thiếu Mã VTTB hoặc mã không có trong danh mục"
                        " — phải đạt 0 lỗi mới INSERT được")

        # INSERT — màu xanh lá, đồng nhất với nút Import bên tab BOM
        self.btn_thdm_insert = CButton(ab, text="💾  ⑤ Tạo THDM",
            command=self._thdm_insert_db,
            font=ctk.CTkFont("Segoe UI", 12, "bold"),
            fg_color=("#16A34A", "#16A34A"),
            hover_color=("#15803D", "#15803D"),
            text_color="#FFFFFF",
            # CTk mặc định chữ nút disabled dùng gray60 — trên nền xanh lá
            # bị mờ khó đọc. Đổi sang chữ đen cho tương phản rõ khi disabled.
            text_color_disabled="#1A1A1A",
            width=210, height=30, corner_radius=6, state="disabled")
        self.btn_thdm_insert.pack(side=tk.RIGHT)

        # Label kết quả tổng hợp/kiểm tra — đặt ở action bar dưới (đủ chỗ,
        # tránh tràn như khi để trên thanh lookup)
        self.lbl_thdm_result = CLabel(ab, text="",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray40", "gray55"),
            fg_color="transparent")
        self.lbl_thdm_result.pack(side=tk.LEFT, padx=(10, 6))

        # ── Empty state (hiện khi chưa chọn gì, giống tab Import BOM) ────────
        self._thdm_result_frame.grid_remove()
        _ef = tk.Frame(ca, bg="#1E1E1E")
        _ef.place(relx=0.5, rely=0.45, anchor="center")
        tk.Label(_ef, text="📋", bg="#1E1E1E", fg="#3E3E42",
                 font=("Segoe UI", 48)).pack()
        tk.Label(_ef, text="①  Tải dữ liệu và chọn Dự án / Đơn hàng / Nhân viên / Đợt",
                 bg="#1E1E1E", fg="#555555", font=("Segoe UI", 13)).pack(pady=(4, 0))
        tk.Label(_ef, text="Tick chọn BOM ở danh sách bên trái  →  ②  Chọn file Excel THDM",
                 bg="#1E1E1E", fg="#3E3E42", font=("Segoe UI", 11)).pack(pady=(2, 0))
        tk.Label(_ef, text="③ Tổng hợp  →  ④ Kiểm tra  →  ⑤ Tạo THDM",
                 bg="#1E1E1E", fg="#3E3E42", font=("Segoe UI", 11)).pack(pady=(2, 0))
        self._thdm_empty_frame = _ef

        # Internal state
        self._thdm_all_bom_rows      = []    # rows BOM từ DB
        self._thdm_checked_ids       = set() # BOM id đang checked
        self._thdm_preview_data      = []    # rows parsed từ Excel
        self._thdm_excel_path        = None  # đường dẫn file Excel đã chọn
        self._thdm_product_map       = {}    # ProductId → Name
        self._thdm_order_map         = {}    # BizDocId → DocNo
        self._thdm_period_map        = {}    # PeriodId → Name
        self._thdm_selected_product_id  = None
        self._thdm_selected_order_id    = None
        self._thdm_selected_period_id   = None
        self._thdm_sheet_row_ids        = []    # row index → BOM Id (tksheet)
        self._thdm_period_active_map    = {}    # display → PeriodId (filtered by order)
        self._thdm_valid_builtin_orders = None  # set of BuiltinOrder0 cho đợt đang chọn
        self._thdm_validated            = False # đã Kiểm tra đạt chưa
        self._thdm_val_errors           = {}    # (section, idx) → [msg,...]

    # ── THDM: B00Lookup helper ─────────────────────────────────────────────────

    @staticmethod
    def _bravo_view_to_table(name: str) -> str:
        """Chuyển tên view BRAVO sang tên bảng gốc.

        Quy tắc BRAVO: v{TênBảng}_{HậuTố...}
          vB20Product_Edit      → B20Product
          vB30BizDocSO_Edit1    → B30BizDocSO
          vB20BOM_Edit          → B20BOM
          B20Product            → B20Product  (đã là bảng, giữ nguyên)
        """
        import re
        if name and re.match(r'^v[A-Z]', name):
            base = name[1:]                    # bỏ chữ 'v' đầu
            base = re.sub(r'_.*$', '', base)   # bỏ _HậuTố
            return base
        return name

    @staticmethod
    def _b00lookup_safe_disp_col(cur, table: str, disp_col: str,
                                 sort_str: str = '', extra_str: str = '') -> str:
        """Tìm cột display an toàn trong bảng gốc.

        DisplayMember trong B00Lookup đôi khi là cột computed chỉ có trong view
        (VD: DocInfo, ItemInfo). Hàm này kiểm tra cột bằng SELECT TOP 0 rồi
        fallback theo thứ tự:
          1. disp_col gốc
          2. Các cột common BRAVO hay dùng làm display
          3. Các cột trong Sort (lấy tên, bỏ ASC/DESC)
          4. Các cột trong ExtraMemberList
        """
        def col_exists(col_name: str) -> bool:
            try:
                cur.execute(f"SELECT TOP 0 [{col_name}] FROM {table}")
                return True
            except Exception:
                return False

        if col_exists(disp_col):
            return disp_col

        # Fallback 1: cột display phổ biến trong BRAVO
        for fb in ('DocNo', 'Code', 'Name', 'Description'):
            if col_exists(fb):
                return fb

        # Fallback 2: lấy tên cột từ Sort (VD: "DocDate DESC,DocNo DESC")
        import re
        for token in re.split(r'[,\s]+', sort_str or ''):
            token = token.strip()
            if token.upper() in ('ASC', 'DESC', ''):
                continue
            if col_exists(token):
                return token

        # Fallback 3: ExtraMemberList
        for token in (extra_str or '').split(','):
            token = token.strip()
            if token and col_exists(token):
                return token

        return disp_col  # trả về gốc, để query tự báo lỗi

    def _b00lookup_build_query(self, lookup_key, extra_where=None, extra_params=None):
        """Đọc B00Lookup → build câu SELECT động trên bảng gốc (không qua view).

        BRAVO rules được áp dụng:
        - HiddenValueMember ưu tiên hơn ValueMember (là id thực).
        - DisplayMember đôi khi là cột computed chỉ có trong view → tự động
          fallback sang cột thực gần nhất (DocNo, Name, Code, ...).
        - FilterKey dùng format BRAVO riêng → caller truyền extra_where thay thế.
        """
        conn = self._get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT LookupTable, ValueMember, DisplayMember,
                   HiddenValueMember, Sort, ExtraMemberList
            FROM   B10_Boho.dbo.B00Lookup
            WHERE  LookupKey = ?
        """, (lookup_key,))
        row = cur.fetchone()
        if not row:
            conn.close()
            raise ValueError(
                f"Không tìm thấy LookupKey='{lookup_key}' trong B10_Boho.dbo.B00Lookup")

        raw_table, val_col, disp_col, hidden_val, sort_str, extra_str = row

        # Ưu tiên HiddenValueMember (id thực); ValueMember đôi khi rỗng
        actual_val  = (hidden_val or '').strip() or (val_col  or '').strip()
        actual_disp = (disp_col  or '').strip()

        if not actual_val:
            conn.close()
            raise ValueError(
                f"B00Lookup '{lookup_key}': ValueMember và HiddenValueMember đều rỗng")
        if not raw_table:
            conn.close()
            raise ValueError(f"B00Lookup '{lookup_key}': LookupTable rỗng")

        # Chuyển view name → bảng gốc
        table = self._bravo_view_to_table(raw_table)

        # === TỰ ĐỘNG CHECK QUYỀN / TỒN TẠI ĐỂ FALLBACK DB ===
        try:
            cur.execute(f"SELECT TOP 0 1 FROM {table}")
        except Exception as e:
            err_msg = str(e).lower()
            # 229: Denied Permission, 208: Invalid Object Name
            if "229" in err_msg or "208" in err_msg or "denied" in err_msg:
                table = f"B10_Boho.dbo.{table}"
        # ====================================================

        # DisplayMember có thể là cột computed của view → tìm cột an toàn
        safe_disp = self._b00lookup_safe_disp_col(
            cur, table, actual_disp, sort_str or '', extra_str or '')

        conn.close()

        where  = f" WHERE {extra_where}" if extra_where else ""
        params = list(extra_params or [])
        sql = (f"SELECT {actual_val}, {safe_disp} "
               f"FROM {table}{where} "
               f"ORDER BY 2")
        return sql, params

    # ── BOM: Load Người lập ───────────────────────────────────────────────────

    def _load_creator_combo(self):
        """Load danh sách UserList từ B00Lookup vào cmb_creator (background thread)."""
        threading.Thread(target=self._load_creator_worker, daemon=True).start()

    def _load_creator_worker(self, _retry=0):
        try:
            sql, params = self._b00lookup_build_query(
                "Employee",
                extra_where="IsGroup = 0 AND IsActive = 1",
            )
            # Inject cột Code vào SELECT để hiển thị "Name — Code"
            # sql dạng: "SELECT Id, Name FROM <table> WHERE ... ORDER BY 2"
            sql = sql.replace(" FROM ", ", Code FROM ", 1)
            conn = self._get_db_conn()
            cur  = conn.cursor()
            try:
                cur.execute(sql, params)
                rows = cur.fetchall()   # [(EmployeeId, Name, Code), ...]
            except Exception:
                # Fallback: Code column không tồn tại → dùng 2 cột
                sql_fb = sql.replace(", Code FROM ", " FROM ", 1)
                cur.execute(sql_fb, params)
                rows = [(r[0], r[1], None) for r in cur.fetchall()]

            # Map EmployeeId → UserId thật (B00UserList.Id) — CreatedBy phải là
            # UserId, không phải EmployeeId. Nhân viên chưa có tài khoản sẽ
            # fallback về DEFAULT_CREATOR_USER_ID khi build _creator_map.
            emp_to_user = {}
            try:
                cur.execute(
                    "SELECT EmployeeId, Id FROM B00UserList WHERE EmployeeId IS NOT NULL")
                emp_to_user = {emp_id: user_id for emp_id, user_id in cur.fetchall()}
            except Exception as e:
                err_msg = str(e).lower()
                if "229" in err_msg or "208" in err_msg or "denied" in err_msg:
                    try:
                        cur.execute(
                            "SELECT EmployeeId, Id FROM B10_Boho.dbo.B00UserList "
                            "WHERE EmployeeId IS NOT NULL")
                        emp_to_user = {emp_id: user_id for emp_id, user_id in cur.fetchall()}
                    except Exception:
                        emp_to_user = {}

            conn.close()
            self.after(0, lambda d=rows, m=emp_to_user: self._load_creator_done(d, None, m))
        except Exception as e:
            _MAX_RETRY = 10
            if _retry < _MAX_RETRY:
                # Auto-retry sau 5 giây (chờ VPN kết nối)
                self.after(5000, lambda r=_retry: threading.Thread(
                    target=self._load_creator_worker,
                    args=(r + 1,), daemon=True).start())
            else:
                self.after(0, lambda err=e: self._load_creator_done([], str(err), {}))

    def _load_creator_done(self, rows, err, emp_to_user=None):
        if err:
            _err_val = ["⚠  Lỗi tải danh sách — click để thử lại"]
            self.cmb_creator.configure(values=_err_val)
            self.cmb_creator.set(_err_val[0])
            if hasattr(self, 'cmb_thdm_creator'):
                self.cmb_thdm_creator.configure(values=_err_val)
                self.cmb_thdm_creator.set(_err_val[0])
            return
        # rows = [(EmployeeId, Name, Code), ...]
        emp_to_user = emp_to_user or {}
        # Hiển thị "Code | Name", map display → UserId thật (B00UserList.Id),
        # fallback DEFAULT_CREATOR_USER_ID nếu nhân viên chưa có tài khoản.
        display_list = [
            f"{code}  |  {name}" if code else str(name)
            for _, name, code in rows
        ]
        self._creator_map = {
            disp: emp_to_user.get(emp_id, DEFAULT_CREATOR_USER_ID)
            for (emp_id, _, __), disp in zip(rows, display_list)
        }
        self.cmb_creator.configure(values=display_list)
        self.cmb_creator.set("— Chọn nhân viên —")
        self._current_creator_code = None
        if hasattr(self, 'cmb_thdm_creator'):
            self.cmb_thdm_creator.configure(values=display_list)
            self.cmb_thdm_creator.set("— Chọn nhân viên —")
            self._thdm_creator_code = None

    def _on_creator_change(self, selected_name):
        if selected_name and ("Lỗi" in selected_name or "Đang tải" in selected_name):
            self._load_creator_combo()
            return
        self._current_creator_code = getattr(self, '_creator_map', {}).get(selected_name)
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if selected_name and "|" in selected_name:
            self.cmb_creator.set(selected_name.split("|")[0].strip())

    def _on_thdm_creator_change(self, selected_name):
        if selected_name and ("Lỗi" in selected_name or "Đang tải" in selected_name):
            self._load_creator_combo()
            return
        self._thdm_creator_code = getattr(self, '_creator_map', {}).get(selected_name)
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if selected_name and "|" in selected_name:
            self.cmb_thdm_creator.set(selected_name.split("|")[0].strip())

    # ── THDM: Load Dự án ──────────────────────────────────────────────────────

    def _thdm_load_products(self):
        self.btn_thdm_load.configure(state="disabled", text="⏳  Đang tải...")
        self.lbl_thdm_status.config(text="⏳  Đang tải dự án...", fg=C["text"])
        threading.Thread(target=self._thdm_load_products_worker, daemon=True).start()

    def _thdm_load_products_worker(self):
        try:
            sql, params = self._b00lookup_build_query("Product1",
                extra_where="IsActive = 1")
            # Inject Code vào SELECT để hiển thị "Code | Name [Id]"
            sql = sql.replace(" FROM ", ", Code FROM ", 1)
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute(sql, params)
            rows = cur.fetchall()   # (Id, Name, Code)
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_products_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_products_done([], str(err)))

    def _thdm_load_products_done(self, rows, error):
        self.btn_thdm_load.configure(state="normal", text="🔄  ① Tải dữ liệu")
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg="#F87171")
            return
        # Hiển thị "Code  |  Name  [Id]", map display → Id
        def _prod_disp(r):
            row_id, name, code = r[0], r[1] or "", r[2] or ""
            return f"{code}  |  {name}  [{row_id}]"
        display_list = [_prod_disp(r) for r in rows]
        self._thdm_product_map = {disp: r[0] for r, disp in zip(rows, display_list)}
        values = ["— Chọn dự án —"] + display_list
        self.cmb_thdm_product.configure(values=values)
        self.cmb_thdm_product.set("— Chọn dự án —")
        self.cmb_thdm_order.configure(state="disabled", values=["— Chọn đơn hàng —"])
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self._thdm_selected_product_id = None
        self._thdm_selected_order_id   = None
        self._thdm_all_bom_rows = []
        self._thdm_checked_ids  = set()
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} dự án — chọn dự án để tiếp tục",
            fg=C["green"])

    def _thdm_on_product_change(self, value=None):
        selected = self.cmb_thdm_product.get()
        product_id = self._thdm_product_map.get(selected)
        if product_id is None:
            self.cmb_thdm_order.configure(state="disabled",
                                          values=["— Chọn đơn hàng —"])
            self.cmb_thdm_order.set("— Chọn đơn hàng —")
            self._thdm_selected_product_id = None
            return
        # Sau khi chọn chỉ hiển thị Code (phần trước |)
        if "|" in selected:
            self.cmb_thdm_product.set(selected.split("|")[0].strip())
        self._thdm_selected_product_id = product_id
        self._thdm_load_orders(product_id)

    # ── THDM: Load Đơn hàng ───────────────────────────────────────────────────

    def _thdm_load_orders(self, product_id):
        self.cmb_thdm_order.configure(state="disabled")
        self.lbl_thdm_status.config(text="⏳  Đang tải đơn hàng...", fg=C["text"])
        threading.Thread(
            target=self._thdm_load_orders_worker,
            args=(product_id,), daemon=True).start()

    def _thdm_load_orders_worker(self, product_id):
        try:
            # Custom SQL: hiển thị DocNo2 | DocDate | Description
            # BizDocId là FK dùng trong BOM.ParentBizDocId (không phải Id integer)
            sql = """
                SELECT BizDocId,
                       ISNULL(DocNo2, DocNo)                        AS DocNo2,
                       CONVERT(varchar(10), DocDate, 103)           AS DocDate,
                       ISNULL(Description, '')                      AS Descr
                FROM   B30BizDocSO
                WHERE  DocCode = 'FO'
                  AND  ProductId = ?
                  AND  IsActive  = 1
                ORDER  BY DocDate DESC, DocNo2
            """
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute(sql, [product_id])
            rows = cur.fetchall()   # (Id, DocNo2, DocDate, Descr)
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_orders_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_orders_done([], str(err)))

    def _thdm_load_orders_done(self, rows, error):
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg="#F87171")
            return
        # Hiển thị "DocNo2  |  DocDate  |  Description", map display → Id
        def _ord_disp(r):
            return f"{r[1]}  |  {r[2]}  |  {r[3]}"
        display_list = [_ord_disp(r) for r in rows]
        self._thdm_order_map = {disp: r[0] for r, disp in zip(rows, display_list)}
        values = ["— Chọn đơn hàng —"] + display_list
        self.cmb_thdm_order.configure(state="normal", values=values)
        self.cmb_thdm_order.set("— Chọn đơn hàng —")
        self._thdm_selected_order_id = None
        self._thdm_all_bom_rows = []
        self._thdm_checked_ids  = set()
        # Reset Đợt combo khi đơn hàng list thay đổi
        self.cmb_thdm_period.configure(state="disabled", values=["— Chọn đợt —"])
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        self._thdm_period_active_map    = {}
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} đơn hàng",
            fg=C["green"])

    def _thdm_on_order_change(self, value=None):
        selected = self.cmb_thdm_order.get()
        order_id = self._thdm_order_map.get(selected)
        self._thdm_selected_order_id = order_id
        # Reset Đợt và BOM filter khi đổi đơn hàng
        self.cmb_thdm_period.configure(state="disabled", values=["— Chọn đợt —"])
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        self._thdm_period_active_map    = {}
        if order_id is None:
            return
        # Sau khi chọn chỉ hiển thị DocNo2 (phần trước |)
        if "|" in selected:
            self.cmb_thdm_order.set(selected.split("|")[0].strip())
        self._thdm_load_bom_list()
        self._thdm_load_period_by_order(order_id)   # Task 10

    # ── THDM: Load Đợt (Period) — Task 10 ────────────────────────────────────
    # Đợt được lọc theo đơn hàng đang chọn qua B30BizDocDetailFactory

    def _thdm_load_period_by_order(self, order_id):
        """Bắt đầu load danh sách Đợt cho đơn hàng (background)."""
        self.cmb_thdm_period.configure(state="disabled", values=["⏳  Đang tải đợt..."])
        self.cmb_thdm_period.set("⏳  Đang tải đợt...")
        self._thdm_selected_period_id   = None
        self._thdm_valid_builtin_orders = None
        threading.Thread(
            target=self._thdm_period_by_order_worker,
            args=(order_id,), daemon=True).start()

    def _thdm_period_by_order_worker(self, order_id):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            # B00Lookup[Period]: LookupTable=B20Period, HiddenValueMember=Id, DisplayMember=Name
            cur.execute("""
                SELECT DISTINCT f.PeriodId,
                       ISNULL(p.Name, CAST(f.PeriodId AS NVARCHAR(50)))
                FROM   B30BizDocDetailFactory AS f
                LEFT JOIN B20Period AS p ON p.Id = f.PeriodId
                WHERE  f.BizDocId = ?
                  AND  f.PeriodId IS NOT NULL
                ORDER BY ISNULL(p.Name, CAST(f.PeriodId AS NVARCHAR(50)))
            """, [order_id])
            rows = cur.fetchall()   # [(PeriodId, DisplayName), ...]
            conn.close()
            self.after(0, lambda d=rows: self._thdm_period_by_order_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_period_by_order_done([], str(err)))

    def _thdm_period_by_order_done(self, rows, error):
        if error:
            self.cmb_thdm_period.configure(
                state="normal", values=["⚠  Lỗi tải đợt"])
            self.cmb_thdm_period.set("⚠  Lỗi tải đợt")
            # Hiện lỗi thật ra status để debug
            self.lbl_thdm_status.config(text=f"❌  Đợt: {error}", fg="#F87171")
            return
        if not rows:
            self.cmb_thdm_period.configure(values=["— Không có đợt —"])
            self.cmb_thdm_period.set("— Không có đợt —")
            return
        self._thdm_period_active_map = {r[1]: r[0] for r in rows}  # Name → PeriodId
        values = ["— Chọn đợt —"] + [r[1] for r in rows]
        self.cmb_thdm_period.configure(state="normal", values=values)
        self.cmb_thdm_period.set("— Chọn đợt —")
        self._thdm_selected_period_id = None

    def _on_thdm_period_change(self, selected_name=None):
        selected = self.cmb_thdm_period.get()
        self._thdm_selected_period_id = self._thdm_period_active_map.get(selected)
        if self._thdm_selected_period_id is None:
            self._thdm_valid_builtin_orders = None
            self._thdm_filter_bom()
            return
        # Task 11: load BuiltinOrder0 cho đợt này → filter BOM grid
        self._thdm_load_builtin_orders_for_period()

    # ── THDM: Filter BOM theo BuiltinOrder0 — Task 11 ─────────────────────────

    def _thdm_load_builtin_orders_for_period(self):
        """Query DISTINCT BuiltinOrder0 cho (order, period) → filter BOM grid."""
        order_id  = self._thdm_selected_order_id
        period_id = self._thdm_selected_period_id
        if order_id is None or period_id is None:
            return
        threading.Thread(
            target=self._thdm_builtin_orders_worker,
            args=(order_id, period_id), daemon=True).start()

    def _thdm_builtin_orders_worker(self, order_id, period_id):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute("""
                SELECT DISTINCT BuiltinOrder0
                FROM   B30BizDocDetailFactory
                WHERE  BizDocId    = ?
                  AND  PeriodId    = ?
                  AND  BuiltinOrder0 IS NOT NULL
            """, [order_id, period_id])
            valid = {str(r[0]) for r in cur.fetchall()}
            conn.close()
            self.after(0, lambda v=valid: self._thdm_builtin_orders_done(v, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_builtin_orders_done(set(), str(err)))

    def _thdm_builtin_orders_done(self, valid_orders, error):
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg="#F87171")
            return
        self._thdm_valid_builtin_orders = valid_orders
        self._thdm_filter_bom()

    # ── THDM: File picker Excel ────────────────────────────────────────────────

    def _thdm_pick_excel(self):
        import tkinter.filedialog as fd, os
        path = fd.askopenfilename(
            title="Chọn file Excel THDM",
            filetypes=[("Excel", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")])
        if not path:
            return
        self._thdm_excel_path = path
        self.lbl_thdm_excel_path.config(text=os.path.basename(path))
        # Reset preview
        self._thdm_clear_preview_trees()
        self._thdm_preview_data = []
        self._thdm_validated  = False
        self._thdm_val_errors = {}
        self.btn_thdm_export_xl.configure(state="disabled")
        self.btn_thdm_view_sql.configure(state="disabled")
        self.btn_thdm_validate.configure(state="disabled")
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(text="")
        self._thdm_update_ui()
        self._thdm_start_pre_preview(path)

    # ── THDM: Pre-preview (đọc Excel ngay sau khi chọn file) ──────────────────

    def _thdm_start_pre_preview(self, path):
        """Hiện pre-preview frame + bắt đầu đọc file (background)."""
        if hasattr(self, '_thdm_empty_frame'):
            self._thdm_empty_frame.place_forget()
        self._thdm_result_frame.grid_remove()
        self._thdm_pre_frame.grid()
        self._thdm_pre_nb.grid_remove()
        self._lbl_thdm_pre_status.config(text="⏳  Đang đọc file Excel...")
        threading.Thread(
            target=self._thdm_pre_preview_worker, args=(path,), daemon=True).start()

    def _thdm_pre_preview_worker(self, path):
        try:
            mapping         = load_mapping()
            thdm_map        = mapping.get('THDM_THVT', [])
            mapping_anchors = [
                _norm_vn(r['ten_excel'])
                for r in thdm_map
                if r.get('nguon_dl') == 'Excel' and r.get('ten_excel')
            ]

            # Mở workbook giống aggregate_worker (1 lần, tái dùng cho parse)
            wb_c, wb_l, wb_h = _thdm_open_workbook(path)
            sheet_name = _thdm_find_thvt_sheet(wb_c)
            if not sheet_name:
                raise ValueError(
                    "Không tìm thấy sheet 'TH VT'.\n"
                    "Tên hợp lệ: 'TH VT', 'THVT', 'TH_VT', 'Tổng hợp vật tư'.")

            # Đọc raw rows từ wb_c
            ws_c     = wb_c[sheet_name]
            raw_base = [tuple(c.value for c in row) for row in ws_c.iter_rows()]

            # Áp merge info từ wb_h (read_only=False → có merged_cells)
            # Mục đích: fill giá trị sang các ô bị merge để hiển thị như Excel
            merge_map = {}   # (row_0idx, col_0idx) → value
            if wb_h and sheet_name in wb_h.sheetnames:
                ws_h = wb_h[sheet_name]
                for mr in ws_h.merged_cells.ranges:
                    top_val = ws_h.cell(mr.min_row, mr.min_col).value
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            if not (r == mr.min_row and c == mr.min_col):
                                merge_map[(r - 1, c - 1)] = top_val

            if merge_map:
                raw_rows = []
                for ri, row in enumerate(raw_base):
                    lst = list(row)
                    for ci in range(len(lst)):
                        if (ri, ci) in merge_map:
                            lst[ci] = merge_map[(ri, ci)]
                    raw_rows.append(tuple(lst))
            else:
                raw_rows = raw_base

            # Dùng raw_base (trước merge) để detect hidx
            # — merge fill có thể làm row CHI TIẾT trông giống header hơn
            hidx = _thdm_find_header_row(raw_base, anchors=mapping_anchors or None)

            # Full parse (giống _thdm_aggregate_worker)
            rows_raw = _thdm_parse_thvt_sheet(wb_c, wb_l, wb_h, sheet_name, thdm_map)
            wb_c.close()
            if wb_l: wb_l.close()
            if wb_h: wb_h.close()

            rows_out = []
            for rd in rows_raw:
                row_norm = {}
                for k, v in rd.items():
                    row_norm[k] = v.strip() if isinstance(v, str) else v
                rows_out.append(row_norm)

            self.after(0, lambda rr=raw_rows, rb=raw_base, hi=hidx, po=rows_out:
                       self._thdm_pre_preview_done(rr, rb, hi, po, None))
        except Exception as exc:
            self.after(0, lambda e=exc:
                       self._thdm_pre_preview_done([], [], None, [], str(e)))

    def _thdm_pre_preview_done(self, raw_rows, raw_base, hidx, parsed_rows, error):
        if error:
            self._lbl_thdm_pre_status.config(text=f"❌  {error}")
            self._thdm_pre_nb.grid()   # hiện notebook dù lỗi (tabs rỗng)
            return

        MAX_RAW = 500

        # ── Raw Excel tab ─────────────────────────────────────────────────────
        # raw_rows = merged version (để hiển thị body),
        # nhưng col_names lấy từ raw_rows[hidx] (row STT/MãVT...)
        # + fallback sang raw_rows[hidx+1] (row Muc 1, Muc 1.2...) cho ô trống
        if hidx is not None and hidx < len(raw_rows):
            # Dùng raw_base (TRƯỚC merge) để build col_names
            # — merge có thể span nhiều row → raw_rows[hidx] bị fill "CHI TIẾT..." sai
            hrow_b  = raw_base[hidx]     if hidx < len(raw_base)     else ()
            hrow2_b = raw_base[hidx + 1] if hidx + 1 < len(raw_base) else ()
            n       = max(len(hrow_b), len(hrow2_b))
            col_names, seen = [], {}
            for i in range(n):
                v1 = hrow_b[i]  if i < len(hrow_b)  else None
                v2 = hrow2_b[i] if i < len(hrow2_b) else None
                # Ưu tiên v2 (Muc 1, NHÀ MÁY...) vì cụ thể hơn
                # v1 có thể là super-header chung ("CHI TIẾT GIAO CT") — chỉ dùng khi v2 trống
                nm = str(v2).strip() if v2 is not None and str(v2).strip() else ""
                if not nm:
                    nm = str(v1).strip() if v1 is not None and str(v1).strip() else ""
                if not nm:
                    nm = f"_C{i}"
                if nm in seen:
                    seen[nm] += 1
                    nm = f"{nm}_{seen[nm]}"
                else:
                    seen[nm] = 0
                col_names.append(nm)
            # Body: bắt đầu sau 2 dòng header (col headings đã hiển thị đủ thông tin)
            display_rows = raw_rows[hidx + 2:]
        else:
            n_cols       = max((len(r) for r in raw_rows[:5]), default=10)
            col_names    = [f"C{i}" for i in range(n_cols)]
            display_rows = raw_rows

        import datetime as _dt

        def _raw_fmt(v):
            """Format giá trị raw cell: datetime → chuỗi ngày, số → str, None → ''."""
            if v is None:
                return ""
            if isinstance(v, (_dt.datetime, _dt.date)):
                return v.strftime("%d/%m/%Y")
            return str(v)

        raw_tree = self.thdm_raw_tree
        for iid in raw_tree.get_children():
            raw_tree.delete(iid)
        raw_tree["columns"] = tuple(col_names)
        for nm in col_names:
            raw_tree.heading(nm, text=nm, anchor='w')
            raw_tree.column(nm, width=90, anchor='w', minwidth=50, stretch=False)
        raw_tree.tag_configure("row_header", foreground="#60A5FA", background="#1E3A5F")
        raw_tree.tag_configure("row_normal", foreground="#CCCCCC")
        raw_tree.tag_configure("row_alt",    foreground="#888888")
        raw_tree.tag_configure("row_blank",  foreground="#444444")

        for i, row in enumerate(display_rows[:MAX_RAW]):
            padded = list(row) + [None] * max(0, len(col_names) - len(row))
            vals   = [_raw_fmt(v) for v in padded[:len(col_names)]]
            is_blank = all(v == "" for v in vals)
            if is_blank:
                tag = "row_blank"
            else:
                tag = "row_normal" if i % 2 == 0 else "row_alt"
            raw_tree.insert("", "end", values=tuple(vals), tags=(tag,))

        # ── Parsed data (tab "Đã xử lý" đã bỏ — chỉ giữ data cho Tổng hợp) ────
        self._thdm_preview_data = parsed_rows   # pre-populate để Tổng hợp không cần parse lại

        # Hiện notebook + cập nhật status
        self._thdm_pre_nb.grid()
        n_total  = len(raw_rows)
        n_parsed = len(parsed_rows)
        suffix   = f" (hiển thị {MAX_RAW})" if len(display_rows) > MAX_RAW else ""
        self._lbl_thdm_pre_status.config(
            text=(f"✅  {n_total:,} dòng Raw{suffix}  ·  {n_parsed:,} dòng đã parse"
                  f"  —  Kiểm tra xong thì nhấn 📊 Tổng hợp"))

    # ── THDM: Load BOM list ────────────────────────────────────────────────────

    def _thdm_load_bom_list(self):
        self.lbl_thdm_status.config(text="⏳  Đang tải danh sách BOM...", fg=C["text"])
        threading.Thread(target=self._thdm_load_bom_list_worker, daemon=True).start()

    def _thdm_load_bom_list_worker(self):
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            where_extra = ""
            params = []
            if self._thdm_selected_order_id is not None:
                where_extra = "  AND  BOM.ParentBizDocId = ?"
                params = [self._thdm_selected_order_id]
            # CTE: chỉ lấy Version mới nhất của mỗi BOM.Code (task 7)
            cur.execute(f"""
                WITH RankedBOM AS (
                    SELECT BOM.Id,
                           BOM.Code,
                           BOM.Description,
                           CONVERT(varchar(10), BOM.DocDate, 103) AS DocDate,
                           BOM.IsActive,
                           Prd.Name          AS ProductName1,
                           Dt0.BuiltinOrder0,
                           Dt.DocNo          AS DocInforFO,
                           BOM.Version,
                           ROW_NUMBER() OVER (
                               PARTITION BY BOM.Code
                               ORDER BY     ISNULL(TRY_CAST(BOM.Version AS FLOAT), -1) DESC,
                                            BOM.DocDate DESC
                           ) AS _rn
                    FROM   B20BOM AS BOM
                    LEFT JOIN B20Product        AS Prd ON Prd.Id             = BOM.ProductId1
                    LEFT JOIN B30BizDocDetailSO AS Dt0 ON Dt0.DetailRowId_SO = BOM.DetailRowId_SO
                    LEFT JOIN B30BizDocSO       AS Dt  ON Dt.BizDocId        = BOM.ParentBizDocId
                    WHERE  BOM.IsGroup  = 0
                      AND  BOM.IsActive = 1
                      {where_extra}
                )
                SELECT Id, Code, Description, DocDate, IsActive,
                       ProductName1, BuiltinOrder0, DocInforFO, Version
                FROM   RankedBOM
                WHERE  _rn = 1
                ORDER  BY TRY_CAST(BuiltinOrder0 AS FLOAT), Code
            """, params)
            cols = [c[0] for c in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            conn.close()
            self.after(0, lambda d=rows: self._thdm_load_bom_list_done(d, None))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_load_bom_list_done([], str(err)))

    def _thdm_load_bom_list_done(self, rows, error):
        if error:
            self.lbl_thdm_status.config(text=f"❌  {error}", fg="#F87171")
            return
        self._thdm_all_bom_rows = rows
        self._thdm_checked_ids  = set()
        self._thdm_filter_bom()
        self.lbl_thdm_status.config(
            text=f"✅  {len(rows):,} BOM — chọn BOM rồi chọn file Excel",
            fg=C["green"])

    def _thdm_filter_bom(self):
        kw = self.thdm_search_var.get().strip().lower()
        valid_orders = self._thdm_valid_builtin_orders  # None = không filter

        if _HAS_TKSHEET and self.thdm_bom_sheet is not None:
            # ── tksheet path ─────────────────────────────────────────────────
            data = []
            self._thdm_sheet_row_ids = []
            for r in self._thdm_all_bom_rows:
                code      = r.get("Code",         "") or ""
                item_name = r.get("ProductName1", "") or ""
                version   = str(r.get("Version",  "") or "")
                muc_so    = str(r.get("BuiltinOrder0", "") or "")
                if valid_orders is not None and muc_so not in valid_orders:
                    continue
                if kw and kw not in code.lower() \
                       and kw not in item_name.lower() \
                       and kw not in version.lower():
                    continue
                chk = "☑" if r["Id"] in self._thdm_checked_ids else "☐"
                data.append([chk, muc_so, code, item_name, version])
                self._thdm_sheet_row_ids.append(r["Id"])
            self.thdm_bom_sheet.set_sheet_data(data, reset_col_positions=False)
            self._thdm_apply_sheet_highlights()
        else:
            # ── Fallback: ttk.Treeview path ──────────────────────────────────
            tree = self.thdm_bom_tree
            for iid in tree.get_children():
                tree.delete(iid)
            for r in self._thdm_all_bom_rows:
                code      = r.get("Code",         "") or ""
                item_name = r.get("ProductName1", "") or ""
                version   = str(r.get("Version",  "") or "")
                muc_so    = str(r.get("BuiltinOrder0", "") or "")
                if valid_orders is not None and muc_so not in valid_orders:
                    continue
                if kw and kw not in code.lower() \
                       and kw not in item_name.lower() \
                       and kw not in version.lower():
                    continue
                rid = str(r["Id"])
                chk = "☑" if r["Id"] in self._thdm_checked_ids else "☐"
                tag = "checked" if r["Id"] in self._thdm_checked_ids else "unchecked"
                tree.insert("", "end", iid=rid,
                    values=(chk, muc_so, code, item_name, version),
                    tags=(tag,))

    def _thdm_bom_autostretch(self):
        """Giãn cột 'Tên SP' (index 3) lấp hết chiều ngang trống của BOM list."""
        self._thdm_bom_stretch_job = None
        sh = getattr(self, 'thdm_bom_sheet', None)
        if not (_HAS_TKSHEET and sh is not None):
            return
        try:
            total  = sh.MT.winfo_width()
            others = sum(int(sh.column_width(column=c)) for c in (0, 1, 2, 4))
            w = total - others - 6
            if w > 200:
                sh.column_width(column=3, width=w)
        except Exception:
            pass

    def _thdm_apply_sheet_highlights(self):
        """Tô màu các row đang được check trong tksheet."""
        self.thdm_bom_sheet.dehighlight_all()
        checked_rows = [
            i for i, rid in enumerate(self._thdm_sheet_row_ids)
            if rid in self._thdm_checked_ids
        ]
        if checked_rows:
            self.thdm_bom_sheet.highlight_rows(
                rows=checked_rows, bg="#1e3a5f", fg="#90c8f8", redraw=True)
        else:
            self.thdm_bom_sheet.refresh()

    def _thdm_on_bom_click(self, event=None):
        if _HAS_TKSHEET and self.thdm_bom_sheet is not None:
            # tksheet path — được gọi từ fallback Treeview nếu sheet=None
            pass   # xử lý ở _thdm_process_sheet_click (gọi qua MT.bind)
        else:
            # ── Fallback: ttk.Treeview path ──────────────────────────────────
            iid = self.thdm_bom_tree.identify_row(event.y)
            if not iid:
                return
            iid_to_id = {str(r["Id"]): r["Id"] for r in self._thdm_all_bom_rows}
            clicked_id = iid_to_id.get(iid)
            if clicked_id is None:
                return
            new_checked  = clicked_id not in self._thdm_checked_ids
            selected_iids = set(self.thdm_bom_tree.selection())
            selected_iids.add(iid)
            for s_iid in selected_iids:
                s_id = iid_to_id.get(s_iid)
                if s_id is None:
                    continue
                if new_checked:
                    self._thdm_checked_ids.add(s_id)
                else:
                    self._thdm_checked_ids.discard(s_id)
                chk  = "☑" if new_checked else "☐"
                tag  = "checked" if new_checked else "unchecked"
                vals = list(self.thdm_bom_tree.item(s_iid, "values"))
                vals[0] = chk
                self.thdm_bom_tree.item(s_iid, values=vals, tags=(tag,))
            self._thdm_update_ui()

    def _thdm_process_sheet_click(self):
        """Xử lý toggle checkbox sau khi tksheet update selection (gọi qua after(40))."""
        if not (_HAS_TKSHEET and self.thdm_bom_sheet is not None):
            return
        sheet = self.thdm_bom_sheet
        selected_rows = []

        # Method 1: get_selected_rows — works khi row header được click
        try:
            r = sheet.get_selected_rows()
            selected_rows = sorted(r)
        except Exception:
            pass

        # Method 2: get_selected_cells — works khi enable_bindings("row_select") chọn cell
        if not selected_rows:
            try:
                cells = sheet.get_selected_cells()
                selected_rows = sorted({c[0] for c in cells})
            except Exception:
                pass

        # Method 3: get_currently_selected — fallback cuối
        if not selected_rows:
            try:
                sel = sheet.get_currently_selected()
                if sel is not None:
                    if hasattr(sel, 'row'):
                        selected_rows = [sel.row]
                    elif isinstance(sel, (list, tuple)) and len(sel) > 0:
                        first = sel[0]
                        selected_rows = [first[0] if isinstance(first, (list, tuple)) else first]
            except Exception:
                pass

        if not selected_rows:
            return
        row_idx = selected_rows[0]
        if row_idx >= len(self._thdm_sheet_row_ids):
            return

        clicked_id  = self._thdm_sheet_row_ids[row_idx]
        new_checked = clicked_id not in self._thdm_checked_ids

        for r in selected_rows:
            if r >= len(self._thdm_sheet_row_ids):
                continue
            row_id = self._thdm_sheet_row_ids[r]
            if new_checked:
                self._thdm_checked_ids.add(row_id)
            else:
                self._thdm_checked_ids.discard(row_id)
            chk = "☑" if new_checked else "☐"
            try:
                self.thdm_bom_sheet.set_cell_data(r, 0, chk, redraw=False)
            except Exception:
                try:
                    self.thdm_bom_sheet.set_cell_data(r, 0, chk)
                except Exception:
                    pass

        self._thdm_apply_sheet_highlights()
        self._thdm_update_ui()

    def _thdm_toggle_all(self, select: bool):
        for r in self._thdm_all_bom_rows:
            if select:
                self._thdm_checked_ids.add(r["Id"])
            else:
                self._thdm_checked_ids.discard(r["Id"])
        self._thdm_filter_bom()
        self._thdm_update_ui()

    def _thdm_update_ui(self):
        n = len(self._thdm_checked_ids)
        self.lbl_thdm_sel_count.config(
            text=f"{n} BOM được chọn",
            fg=C["accent"] if n else C["text"])
        state = "normal" if n else "disabled"
        self.btn_thdm_aggregate.configure(state=state)

    # ── THDM: Đọc Excel và preview ────────────────────────────────────────────

    def _thdm_aggregate(self):
        if not self._thdm_checked_ids:
            self._show_msg("Chưa chọn BOM",
                "Vui lòng chọn ít nhất 1 BOM trong danh sách.", kind="info")
            return
        if not self._thdm_excel_path:
            self._show_msg("Chưa chọn file Excel",
                "Vui lòng nhấn '📂 Chọn file Excel THDM' để chọn file.", kind="info")
            return
        self.btn_thdm_aggregate.configure(state="disabled", text="⏳  Đang đọc Excel...")
        self.btn_thdm_export_xl.configure(state="disabled")
        self.btn_thdm_view_sql.configure(state="disabled")
        self.btn_thdm_validate.configure(state="disabled")
        self.btn_thdm_insert.configure(state="disabled")
        self._thdm_clear_preview_trees()
        path = self._thdm_excel_path
        threading.Thread(target=self._thdm_aggregate_worker, args=(path,), daemon=True).start()

    def _thdm_aggregate_worker(self, path):
        try:
            mapping   = load_mapping()
            cfg       = mapping.get('_CONFIG', {})

            # Tìm tất cả child sections của THDM_HEADER (theo thứ tự khai báo trong _CONFIG)
            child_sections = [
                sec for sec, info in cfg.items()
                if info.get('parent_section') == 'THDM_HEADER'
            ]
            if not child_sections:
                raise ValueError("Không tìm thấy child section nào của THDM_HEADER trong _CONFIG.")

            # ── Query B20BOMDetail cho các BOM đã chọn ─────────────────────────
            bom_qty_dict = None
            bom_qty_warn = None
            if self._thdm_checked_ids:
                try:
                    conn_bom = self._get_db_conn()
                    bom_qty_dict = _thdm_load_bom_qty_dict(conn_bom, self._thdm_checked_ids)
                    conn_bom.close()
                except Exception as e_bom:
                    bom_qty_warn = str(e_bom)
                    bom_qty_dict = None

            # Mở file 3 lần (cached / live formula / hidden detection) — dùng chung
            wb_c, wb_l, wb_h = _thdm_open_workbook(path)

            # Detect sheet TH VT — tất cả THDM child sections đều parse sheet này
            sheet_name = _thdm_find_thvt_sheet(wb_c)
            if not sheet_name:
                raise ValueError(
                    "Không tìm thấy sheet 'TH VT' trong file Excel.\n"
                    "Tên sheet hợp lệ: 'TH VT', 'THVT', 'TH_VT', 'Tổng hợp vật tư'.")

            rows_out = []
            for sec in child_sections:
                sec_cfg    = cfg[sec]
                sec_map    = mapping.get(sec, [])
                expand_muc = sec_cfg.get('expand_muc', False)
                row_filter = sec_cfg.get('row_filter') or ''

                rows_raw = _thdm_parse_thvt_sheet(
                    wb_c, wb_l, wb_h, sheet_name, sec_map,
                    expand_muc=expand_muc,
                    bom_qty_dict=bom_qty_dict if expand_muc else None,
                )

                # Áp dụng RowFilter nếu có (vd: 'Quantity>0')
                if row_filter:
                    rows_raw = _thdm_apply_row_filter(rows_raw, row_filter)

                # Chuẩn hoá + gán tag _section cho từng row
                for rd in rows_raw:
                    row_norm = {'_section': sec}
                    for k, v in rd.items():
                        row_norm[k] = v.strip() if isinstance(v, str) else v
                    rows_out.append(row_norm)

            wb_c.close()
            if wb_l: wb_l.close()
            if wb_h: wb_h.close()

            self.after(0, lambda d=rows_out, w=bom_qty_warn: self._thdm_aggregate_done(d, None, bom_warn=w))
        except Exception as e:
            self.after(0, lambda err=e: self._thdm_aggregate_done([], str(err)))

    def _thdm_clear_preview_trees(self):
        """Xóa dữ liệu tab Đầu phiếu + tất cả tab section detail."""
        for tr in [self._thdm_header_tree] + list(self._thdm_sec_trees.values()):
            ch = tr.get_children()
            if ch:
                tr.delete(*ch)

    def _thdm_fill_header_preview(self):
        """Đổ preview THDM_HEADER — CÙNG cấu trúc với tab detail:
        header = Ten_Excel (rỗng → SQL_Column), hàng phụ xanh = SQL_Column,
        hàng giá trị = resolve nhẹ (display-only, không cần DB). Cột có
        Ten_Excel xếp trước. Giá trị SP/tự sinh là placeholder, thật khi INSERT."""
        tr = self._thdm_header_tree
        ch = tr.get_children()
        if ch:
            tr.delete(*ch)
        if not getattr(self, '_thdm_now_str', None):
            self._thdm_now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        defs = _thdm_get_detail_col_defs((self.mapping or {}).get('THDM_HEADER', []))
        self._thdm_header_col_defs = defs

        import tkinter.font as _tkf
        _hfont = _tkf.Font(family="Segoe UI", size=12, weight="bold")
        _font  = _tkf.Font(family="Segoe UI", size=11)
        tr["columns"] = tuple(d['sql_col'] for d in defs)
        for d in defs:
            htxt = d['ten_excel'] or d['sql_col']
            val  = self._thdm_col_display(d, 0)
            tr.heading(d['sql_col'], text=htxt, anchor='center')
            w = max(_hfont.measure(htxt), _font.measure(d['sql_col']),
                    _font.measure(str(val))) + 24
            # Cột căn trái; chỉ 2 dòng header (tiêu đề + hàng phụ SQL) căn giữa
            tr.column(d['sql_col'], width=max(min(w, 320), 80),
                      anchor='w', stretch=False)
        # Hàng phụ (xanh) = SQL_Column ; hàng giá trị = resolved
        tr.insert("", "end", values=tuple(d['sql_col'] for d in defs),
                  tags=("sql_names",))
        tr.insert("", "end",
                  values=tuple(self._thdm_col_display(d, 0) for d in defs),
                  tags=("row_normal",))
        # Căn giữa riêng hàng phụ SQL_Column (dòng giá trị vẫn căn trái)
        if hasattr(tr, 'center_rows_by_tag'):
            tr.center_rows_by_tag("sql_names")

    def _thdm_aggregate_done(self, rows, error, bom_warn=None):
        self.btn_thdm_aggregate.configure(state="normal", text="📊  ③ Tổng hợp")
        # Switch từ pre-preview → result frame
        if hasattr(self, '_thdm_pre_frame'):
            self._thdm_pre_frame.grid_remove()
            self._thdm_result_frame.grid()
        if error:
            self._show_msg("Lỗi đọc Excel", str(error))
            return
        if bom_warn:
            self._show_msg("Cảnh báo BOM",
                f"Không load được định mức từ B20BOMDetail:\n{bom_warn}\n\n"
                "QuantityFactory sẽ dùng tỉ lệ phân bổ thay vì ĐỊNH MỨC thực tế.",
                kind="warning")
        self._thdm_preview_data = rows
        # Tổng hợp mới → reset trạng thái kiểm tra
        self._thdm_validated  = False
        self._thdm_val_errors = {}
        _sec_counts = self._thdm_fill_section_trees()

        # Tab Đầu phiếu
        self._thdm_fill_header_preview()

        if rows:
            self.btn_thdm_export_xl.configure(state="normal")
            self.btn_thdm_view_sql.configure(state="normal")
            self.btn_thdm_validate.configure(state="normal")
            # INSERT chỉ bật sau khi Kiểm tra đạt (giống luồng BOM)
            self.btn_thdm_insert.configure(state="disabled")
            _detail = "  ·  ".join(
                f"{(self.mapping.get('_CONFIG', {}).get(s, {}).get('label') or s)}: {n:,}"
                for s, n in _sec_counts.items())
            self.lbl_thdm_result.config(
                text=f"📊  {len(rows):,} dòng ({_detail}) — bấm ④ Kiểm tra",
                fg=C["yellow"])

    def _thdm_fill_section_trees(self, error_map=None):
        """Đổ dữ liệu từng section vào tab riêng. error_map (optional):
        {(section, idx_trong_section): [msg,...]} → tô đỏ dòng lỗi.
        Trả về dict {section: số_dòng}."""
        error_map = error_map or {}
        self._thdm_clear_preview_trees()

        def _fmt(v, is_num=False):
            if v is None:
                return ""
            if is_num:
                try:
                    return f"{float(v):,.4f}"
                except (TypeError, ValueError):
                    pass
            return str(v) if v != "" else ""

        self._thdm_now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = self._thdm_preview_data
        counts = {}
        for _tab_i, (sec, tr) in enumerate(self._thdm_sec_trees.items(), start=1):
            defs     = self._thdm_sec_col_defs.get(sec, [])
            sec_rows = [r for r in rows if r.get('_section') == sec]
            counts[sec] = len(sec_rows)
            n_err_sec = 0
            # Hàng phụ (xanh) hiển thị SQL_Column tương ứng mỗi cột
            tr.insert("", "end",
                      values=tuple(d['sql_col'] for d in defs),
                      tags=("sql_names",))
            for i, r in enumerate(sec_rows):
                if (sec, i) in error_map:
                    tag = "err_row"
                    n_err_sec += 1
                else:
                    tag = "row_normal" if i % 2 == 0 else "row_alt"
                vals = []
                for d in defs:
                    v = r.get(d['sql_col'])
                    # Cột không phải Excel & chưa có sẵn trong row → hiện giá
                    # trị sẽ resolve khi INSERT (CoDinh/UILookup/HeThong...)
                    if v is None and (d.get('nguon_dl') or 'Excel') != 'Excel':
                        vals.append(self._thdm_col_display(d, i))
                    else:
                        vals.append(_fmt(v, is_num=d.get('anchor') == 'e'))
                tr.insert("", "end", values=tuple(vals), tags=(tag,))
            lbl = (self.mapping.get('_CONFIG', {}).get(sec, {}).get('label')
                   or sec)
            _cnt = (f"{len(sec_rows):,}" if not n_err_sec
                    else f"{len(sec_rows):,}, ⚠{n_err_sec}")
            try:
                self._thdm_result_nb.tab(_tab_i, text=f"  {lbl} ({_cnt})  ")
            except Exception:
                pass
        return counts

    def _thdm_col_display(self, d, ridx):
        """Giá trị hiển thị cho cột không phải Excel (resolve nhẹ, không cần DB).
        Giống logic tab Header — giá trị thật vẫn sinh khi INSERT."""
        nguon = d.get('nguon_dl') or ''
        mac   = d.get('mac_dinh') or ''
        if nguon == 'CoDinh':
            return '' if mac.upper() in ('NULL', 'EMPTY', '') else mac
        if nguon == 'UILookup':
            v = {
                'creator':    self._thdm_creator_code or self._current_creator_code,
                'product_id': self._thdm_selected_product_id,
                'order_id':   self._thdm_selected_order_id,
                'period_id':  self._thdm_selected_period_id,
            }.get(mac)
            return '' if v in (None, '') else str(v)
        if nguon == 'HeThong':
            mu = mac.upper()
            if mu == 'NOW':          return self._thdm_now_str
            if mac == 'BizDocId_SO': return str(self._thdm_selected_order_id or '')
            if mu == 'AUTO_INC':     return str(ridx + 1)
            return '(tự sinh)'
        if nguon in ('SP', 'TinhToan', 'Lookup', 'MucLookup'):
            return '(khi INSERT)'
        if nguon == 'Excel':
            return '(từ Excel)'
        return ''

    # ── THDM: Kiểm tra dữ liệu trước INSERT (giống Validate bên BOM) ───────────
    def _thdm_validate(self):
        if not self._thdm_preview_data:
            self._show_msg("Chưa có dữ liệu",
                "Hãy bấm 📊 Tổng hợp trước khi Kiểm tra.", kind="info")
            return
        self.btn_thdm_validate.configure(state="disabled", text="⏳  Đang kiểm tra...")
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(text="⏳  Đang kiểm tra dữ liệu...", fg=C["text"])
        threading.Thread(target=self._thdm_validate_worker, daemon=True).start()

    def _thdm_validate_worker(self):
        conn = None
        try:
            mapping = load_mapping()
            cfg     = mapping.get('_CONFIG', {})
            child_sections = [s for s, i in cfg.items()
                              if i.get('parent_section') == 'THDM_HEADER']
            conn   = self._get_db_conn()
            errors = {}   # (section, idx) → [msg,...]
            for sec in child_sections:
                sec_map = mapping.get(sec, [])
                req_excel = [r for r in sec_map
                    if str(r.get('bat_buoc', '')).strip() in ('1', '1.0')
                    and r.get('nguon_dl') == 'Excel' and r.get('sql_col')]
                lk_recs = [r for r in sec_map
                    if r.get('nguon_dl') == 'Excel' and r.get('kieu_lookup')
                    and r.get('bang_master') and r.get('truong_lay_ve')
                    and r.get('truong_so_sanh')]
                caches = {}
                for r in lk_recs:
                    key = (r['bang_master'], r.get('dieu_kien_master', ''),
                           r['truong_so_sanh'], r['truong_lay_ve'])
                    if key not in caches:
                        try:
                            caches[key] = self._build_cache_generic(conn, *key)
                        except Exception:
                            caches[key] = []
                sec_rows = [row for row in self._thdm_preview_data
                            if row.get('_section') == sec]
                for idx, row in enumerate(sec_rows):
                    msgs = []
                    for rr in req_excel:
                        v = row.get(rr['sql_col'])
                        if v is None or str(v).strip() == '':
                            msgs.append(f"Thiếu {rr.get('ten_excel') or rr['sql_col']}")
                    for rr in lk_recs:
                        raw = row.get(rr['sql_col'])
                        if raw is None or str(raw).strip() == '':
                            continue   # ô trống đã bắt ở check bắt buộc
                        key = (rr['bang_master'], rr.get('dieu_kien_master', ''),
                               rr['truong_so_sanh'], rr['truong_lay_ve'])
                        found, _ = self._lookup_generic(
                            raw, caches.get(key, []), rr.get('kieu_lookup', 'exact'))
                        if found is None:
                            msgs.append(
                                f"Không tìm thấy {rr.get('ten_excel') or rr['sql_col']} "
                                f"'{raw}' trong {rr['bang_master']}")
                    if msgs:
                        errors[(sec, idx)] = msgs
            self.after(0, lambda e=errors: self._thdm_validate_done(e, None))
        except Exception as e:
            self.after(0, lambda err=str(e): self._thdm_validate_done(None, err))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _thdm_validate_done(self, errors, err):
        self.btn_thdm_validate.configure(state="normal", text="🔍  ④ Kiểm tra")
        if err is not None:
            self.lbl_thdm_result.config(text="❌  Lỗi kiểm tra", fg=C["red"])
            self._show_msg("Lỗi kiểm tra", err, kind="error")
            return
        self._thdm_val_errors = errors
        self._thdm_validated  = True
        n_err = len(errors)
        self._thdm_fill_section_trees(errors)
        if n_err == 0:
            self.btn_thdm_insert.configure(state="normal")
            self.lbl_thdm_result.config(
                text="✅  Kiểm tra đạt — sẵn sàng tạo THDM", fg=C["green"])
            return
        self.btn_thdm_insert.configure(state="disabled")
        self.lbl_thdm_result.config(
            text=f"❌  {n_err} dòng lỗi — sửa file Excel/mapping rồi Tổng hợp lại",
            fg=C["red"])
        # Nhảy tới tab section có lỗi đầu tiên
        sec_first = next(iter(errors))[0]
        for _ti, _s in enumerate(self._thdm_sec_trees.keys(), start=1):
            if _s == sec_first:
                try:
                    self._thdm_result_nb.select(_ti)
                except Exception:
                    pass
                break
        # Liệt kê tối đa 15 lỗi đầu
        lbls = self.mapping.get('_CONFIG', {})
        lines = []
        for (sec, idx), msgs in list(errors.items())[:15]:
            _lbl = lbls.get(sec, {}).get('label') or sec
            lines.append(f"• [{_lbl}] dòng {idx + 1}: {'; '.join(msgs)}")
        more = f"\n… và {n_err - 15} dòng lỗi khác." if n_err > 15 else ""
        self._show_msg("Kiểm tra: có lỗi",
            f"Phát hiện {n_err} dòng lỗi (đã tô đỏ trong bảng):\n\n"
            + "\n".join(lines) + more, kind="warning")

    # ── THDM: Export / View SQL / Insert ──────────────────────────────────────

    def _thdm_export_excel(self):
        if not self._thdm_preview_data:
            return
        import tkinter.filedialog as fd
        path = fd.asksaveasfilename(
            title="Xuất Excel THDM",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="THDM_Export.xlsx")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            # Mỗi section detail 1 sheet riêng (THDM Vật tư / NVL Bổ sung...)
            for sec, col_defs in self._thdm_sec_col_defs.items():
                sec_rows = [r for r in self._thdm_preview_data
                            if r.get('_section') == sec]
                lbl = (self.mapping.get('_CONFIG', {}).get(sec, {}).get('label')
                       or sec)
                ws = wb.create_sheet(title=lbl[:31])
                # Hàng 1 = Ten_Excel (header xanh), Hàng 2 = SQL_Column
                for c, d in enumerate(col_defs, 1):
                    cell = ws.cell(row=1, column=c, value=d['ten_excel'])
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor="1D9E75")
                    cell.alignment = Alignment(horizontal="center")
                    scell = ws.cell(row=2, column=c, value=d['sql_col'])
                    scell.font      = Font(italic=True, color="9CDCFE")
                    scell.alignment = Alignment(horizontal="center")
                # Data rows — theo sql_col từ mapping (bắt đầu hàng 3)
                for i, r in enumerate(sec_rows, 3):
                    for c, d in enumerate(col_defs, 1):
                        ws.cell(row=i, column=c, value=r.get(d['sql_col']))
                # Column widths — convert pixel → char width (~7px per char)
                for c, d in enumerate(col_defs, 1):
                    char_w = max(round(d['width'] / 7), 8)
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(c)].width = char_w
            wb.save(path)
            self._show_export_success("Xuất thành công", f"Đã lưu:\n{path}", path)
        except Exception as e:
            self._show_msg("Lỗi xuất Excel", str(e))

    # ── THDM: Resolve row THVT theo mapping ───────────────────────────────────
    def _resolve_thdm_thvt_row(self, excel_row, biz_doc_id, sort_order, thvt_map, now):
        """
        Resolve 1 dòng THVT từ mapping THDM_THVT.
        Thin wrapper gọi _resolve_row_mapping() dùng chung với BOM.

        excel_row  : dict {sql_col: value} từ _thdm_preview_data
        biz_doc_id : UNIQUEIDENTIFIER vừa tạo cho phiếu THDM (hoặc placeholder)
        sort_order : thứ tự dòng (int, 1-based)
        thvt_map   : mapping.get('THDM_THVT', [])
        now        : datetime đồng nhất cho cả batch
        Returns    : dict {sql_col: value}
        """
        # BizDocId_SO = order_id đang chọn — đặt trong parent_row để
        # HeThong mac_dinh='BizDocId_SO' tự copy đúng
        _parent = {'BizDocId_SO': self._thdm_selected_order_id}
        ctx = {
            'now':           now,
            'builtin_order': sort_order,
            'doc_id':        biz_doc_id,
            'parent_row':    _parent,
            'parent_fields': {'BizDocId_SO'},
            'ui_values': {
                'creator':    self._thdm_creator_code or self._current_creator_code,
                'product_id': self._thdm_selected_product_id,
                'order_id':   self._thdm_selected_order_id,
                'period_id':  self._thdm_selected_period_id,
            },
            'skip_nguon': {'SP', 'TinhToan'},
        }
        result = _resolve_row_mapping(
            thvt_map, ctx,
            get_excel_val=lambda rec: excel_row.get(rec['sql_col'])
        )
        result['ProductId'] = None   # child rows không có ProductId (chỉ parent cần)
        return result

    def _thdm_view_sql(self):
        """Hiện SQL INSERT sẽ thực hiện khi nhấn INSERT vào DB."""
        if not self._thdm_preview_data:
            return
        from datetime import datetime as _dt
        order_id   = self._thdm_selected_order_id or "NULL"
        bom_ids    = list(self._thdm_checked_ids)
        bom_list   = ",".join(str(i) for i in bom_ids)
        mapping    = load_mapping()
        cfg        = mapping.get('_CONFIG', {})
        header_map = mapping.get('THDM_HEADER', [])
        sp_config  = mapping.get('_SP_CONFIG', [])
        child_sections = [
            sec for sec, info in cfg.items()
            if info.get('parent_section') == 'THDM_HEADER'
        ]
        now        = _dt.now()

        ui_vals = {
            'product_id': self._thdm_selected_product_id,
            'order_id':   self._thdm_selected_order_id,
            'creator':    self._thdm_creator_code or self._current_creator_code,
            'period_id':  self._thdm_selected_period_id,
        }

        def _fmt(v):
            if v is None:                                return "NULL"
            if isinstance(v, str) and v.startswith('@'): return v
            if isinstance(v, _dt):                       return f"'{v:%Y-%m-%d %H:%M:%S}'"
            if isinstance(v, str):                       return f"N'{v.replace(chr(39), chr(39)*2)}'"
            if isinstance(v, float):                     return f"{v:.4f}"
            return str(v)

        # ── Resolve THDM_HEADER non-SP fields (CoDinh / HeThong / UILookup) ──
        hdr_out = {}
        for rec in header_map:
            sql_col  = rec['sql_col']
            nguon_dl = rec.get('nguon_dl', '')
            mac_dinh = rec.get('mac_dinh', '')
            kieu     = rec.get('kieu_dl', '').lower()
            if nguon_dl == 'CoDinh':
                if str(mac_dinh).upper() == 'NULL':
                    hdr_out[sql_col] = None
                else:
                    try:
                        if kieu in ('int', 'bigint'):
                            hdr_out[sql_col] = int(float(mac_dinh))
                        elif kieu in ('numeric', 'float', 'decimal'):
                            hdr_out[sql_col] = float(mac_dinh)
                        else:
                            hdr_out[sql_col] = mac_dinh
                    except (ValueError, TypeError):
                        hdr_out[sql_col] = mac_dinh
            elif nguon_dl == 'HeThong':
                hdr_out[sql_col] = now if mac_dinh == 'NOW' else None
            elif nguon_dl == 'UILookup':
                hdr_out[sql_col] = ui_vals.get(mac_dinh)
            # SP fields → placeholder (không chạy SP thật để tránh tiêu sequence)

        hdr_out['BOMIdList'] = bom_list

        # SP fields: hiện dưới dạng comment placeholder
        sp_thdm_hdr = [s for s in sp_config
                       if s.get('section') == 'THDM_HEADER'
                       and s.get('isactive') == '1'
                       and s.get('sp_name', '').lower() != 'lookup']
        for sp_rec in sp_thdm_hdr:
            sql_col = sp_rec.get('sql_column') or sp_rec.get('sql_col', '')
            out_flds = [f.strip() for f in sp_rec.get('outputfields', '').split(',') if f.strip()]
            target = out_flds[0] if out_flds else sql_col
            if target:
                hdr_out[target] = f"/*{sp_rec.get('sp_name','SP')}*/"

        # Loại SP-lookup intermediates (Ws_Id) + IsDraftData khỏi INSERT
        _sp_temp_header = {
            s['sql_column'] for s in sp_config
            if s.get('section') == 'THDM_HEADER'
            and s.get('sp_name', '').lower() == 'lookup'
            and s.get('isactive') == '1'
        }
        _hdr_skip   = _sp_temp_header | {'IsDraftData'}
        valid_hcols = {r['sql_col'] for r in header_map} | {'BOMIdList'}
        ins_hcols   = [c for c in hdr_out if c in valid_hcols and c not in _hdr_skip]

        view_hdr = mapping.get('_CONFIG', {}).get('THDM_HEADER', {}).get(
            'view_insert', 'vB30BizDocDemand_Edit')

        # ── Build cache cho Excel lookup (tất cả child sections) ────────────
        _all_lk_recs_by_sec = {}
        _all_view_caches    = {}
        try:
            _vc = self._get_db_conn()
            for _sec in child_sections:
                _sm  = mapping.get(_sec, [])
                _lks = [r for r in _sm
                        if r.get('nguon_dl') == 'Excel'
                        and r.get('kieu_lookup') and r.get('bang_master')
                        and r.get('truong_lay_ve')]
                _all_lk_recs_by_sec[_sec] = _lks
                if _lks:
                    _all_view_caches.update(self._build_all_caches(_vc, _lks))
            _vc.close()
        except Exception:
            pass

        # ── Build SQL text ────────────────────────────────────────────────────
        hdr_cols_str = ', '.join(ins_hcols)
        hdr_vals_str = ', '.join(
            hdr_out[c] if isinstance(hdr_out[c], str) and hdr_out[c].startswith('/*')
            else _fmt(hdr_out[c])
            for c in ins_hcols
        )

        lines = [
            f"-- ===== THDM INSERT preview (mapping-driven) =====",
            f"-- Dự án     (ProductId):  {self._thdm_selected_product_id}",
            f"-- Đơn hàng  (BizDocId_SO): {order_id}",
            f"-- BOM chọn  ({len(bom_ids)} BOM): {bom_list}",
            f"-- Người lập (CreatedBy):  {ui_vals['creator']}",
            f"-- Đợt       (PeriodId):   {self._thdm_selected_period_id}",
            f"-- NOTE: /*SP_NAME*/ = giá trị do SP sinh ra khi INSERT thật",
            f"",
            f"-- [1] Tạo phiếu THDM (parent) — {len(ins_hcols)} cột từ THDM_HEADER mapping",
            f"DECLARE @BizDocId BIGINT = /*{sp_thdm_hdr[0]['sp_name'] if sp_thdm_hdr else 'SP'}*/NULL;",
            f"INSERT INTO {view_hdr}",
            f"    ({hdr_cols_str})",
            f"VALUES",
            f"    ({hdr_vals_str});",
            f"",
            f"-- [2] Insert {len(self._thdm_preview_data)} dòng vật tư (child — tất cả sections)",
        ]

        # Group rows by section để build mỗi INSERT block riêng
        total_rows = self._thdm_preview_data
        for sec in child_sections:
            sec_map    = mapping.get(sec, [])
            view_sec   = cfg.get(sec, {}).get('view_insert', 'B30BizDocDemandBOM')
            lk_recs    = _all_lk_recs_by_sec.get(sec, [])
            sec_rows   = [r for r in total_rows if r.get('_section') == sec]
            if not sec_rows:
                continue

            # Resolve sample row để lấy cols
            sample = self._resolve_thdm_thvt_row(sec_rows[0], "@BizDocId", 1, sec_map, now)
            cols   = list(sample.keys())

            lines += [
                f"",
                f"-- Section {sec} ({len(sec_rows)} dòng)",
                f"INSERT INTO {view_sec} ({', '.join(cols)})",
                f"VALUES",
            ]
            for idx, r in enumerate(sec_rows):
                resolved = self._resolve_thdm_thvt_row(r, "@BizDocId", idx + 1, sec_map, now)
                for _lrec in lk_recs:
                    _sc  = _lrec.get('sql_col', '')
                    _kl  = _lrec.get('kieu_lookup', '')
                    _bm  = _lrec.get('bang_master', '')
                    _dk  = _lrec.get('dieu_kien_master', '')
                    _ss  = _lrec.get('truong_so_sanh', '')
                    _lv  = _lrec.get('truong_lay_ve', '')
                    _raw = resolved.get(_sc)
                    _ck  = _all_view_caches.get((_bm, _dk, _ss, _lv), [])
                    if _raw is not None and _ck:
                        _found, _ = self._lookup_generic(_raw, _ck, _kl)
                        resolved[_sc] = _found
                comma    = "," if idx < len(sec_rows) - 1 else ";"
                vals_str = ", ".join(_fmt(resolved.get(c)) for c in cols)
                lines.append(f"    ({vals_str}){comma}")

        sql_text = "\n".join(lines)

        dlg = ctk.CTkToplevel(self)
        dlg.title("SQL preview — THDM")
        dlg.geometry("900x580")
        dlg.grab_set()
        txt = tk.Text(dlg, bg="#1E1E1E", fg="#D4D4D4",
            font=("Consolas", 10), wrap="none",
            insertbackground="#D4D4D4")
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        txt.insert("1.0", sql_text)
        txt.configure(state="disabled")
        ctk.CTkButton(dlg, text="📋  Copy",
            command=lambda: (self.clipboard_clear(), self.clipboard_append(sql_text)),
            width=80, height=28).pack(pady=(0,8))

    def _thdm_insert_db(self):
        if not self._thdm_preview_data:
            self._show_msg("Không có dữ liệu",
                "Vui lòng tổng hợp dữ liệu từ Excel trước.", kind="info")
            return
        # Phải Kiểm tra đạt trước (giống luồng Import bên BOM)
        if not self._thdm_validated:
            self._show_msg("Chưa kiểm tra",
                "Hãy bấm ④ Kiểm tra trước khi tạo THDM.", kind="warning")
            return
        if self._thdm_val_errors:
            self._show_msg("Còn lỗi",
                f"Còn {len(self._thdm_val_errors)} dòng lỗi (tô đỏ trong bảng). "
                "Sửa file Excel/mapping rồi Tổng hợp và Kiểm tra lại.", kind="error")
            return
        # Bắt buộc chọn đủ 4 lookup trước khi INSERT
        _missing = []
        if not self._thdm_selected_product_id:
            _missing.append("Dự án")
        if not self._thdm_selected_order_id:
            _missing.append("Đơn hàng")
        if not (self._thdm_creator_code or self._current_creator_code):
            _missing.append("Nhân viên")
        if not self._thdm_selected_period_id:
            _missing.append("Đợt")
        if _missing:
            self._show_msg("Thiếu thông tin",
                "Vui lòng chọn đầy đủ trước khi Tạo THDM:\n\n"
                + "\n".join(f"  •  {f}" for f in _missing),
                kind="warning")
            return
        import tkinter.messagebox as mb
        n    = len(self._thdm_preview_data)
        boms = len(self._thdm_checked_ids)
        ok   = mb.askyesno("Xác nhận tạo THDM",
            f"Sẽ tạo 1 phiếu THDM mới với:\n"
            f"  • {boms} BOM được chọn\n"
            f"  • {n:,} dòng vật tư từ Excel\n\n"
            f"Tiếp tục?")
        if not ok:
            return
        self.btn_thdm_insert.configure(state="disabled", text="⏳  Đang tạo THDM...")
        self._loading_dlg = self._make_loading_popup(
            "Đang tạo phiếu THDM...\nVui lòng không đóng cửa sổ.")
        threading.Thread(target=self._thdm_insert_worker, daemon=True).start()

    def _thdm_insert_worker(self):
        conn = None
        try:
            from datetime import datetime as _dt
            conn       = self._get_db_conn()
            cur        = conn.cursor()
            order_id   = self._thdm_selected_order_id
            bom_ids    = list(self._thdm_checked_ids)
            rows       = self._thdm_preview_data
            bom_list   = ",".join(str(i) for i in bom_ids)
            mapping    = load_mapping()
            cfg        = mapping.get('_CONFIG', {})
            header_map = mapping.get('THDM_HEADER', [])
            sp_config  = mapping.get('_SP_CONFIG', [])
            child_sections = [
                sec for sec, info in cfg.items()
                if info.get('parent_section') == 'THDM_HEADER'
            ]
            now        = _dt.now()

            # UI values cho UILookup branch
            ui_vals = {
                'product_id': self._thdm_selected_product_id,
                'order_id':   order_id,
                'creator':    self._thdm_creator_code or self._current_creator_code,
                'period_id':  self._thdm_selected_period_id,
            }
            # ── 1) Resolve THDM_HEADER từ mapping ──────────────────────────
            row_out = {}
            for rec in header_map:
                sql_col  = rec['sql_col']
                nguon_dl = rec.get('nguon_dl', '')
                mac_dinh = rec.get('mac_dinh', '')
                kieu     = rec.get('kieu_dl', '').lower()

                if nguon_dl == 'CoDinh':
                    if mac_dinh.upper() == 'NULL':
                        row_out[sql_col] = None
                    else:
                        try:
                            if kieu in ('int', 'bigint'):
                                row_out[sql_col] = int(float(mac_dinh))
                            elif kieu in ('numeric', 'float', 'decimal'):
                                row_out[sql_col] = float(mac_dinh)
                            else:
                                row_out[sql_col] = mac_dinh
                        except (ValueError, TypeError):
                            row_out[sql_col] = mac_dinh
                elif nguon_dl == 'HeThong':
                    if mac_dinh == 'NOW':
                        row_out[sql_col] = now
                elif nguon_dl == 'UILookup':
                    row_out[sql_col] = ui_vals.get(mac_dinh)
                # SP fields resolved in step 2

            # BOMIdList: trường đặc biệt không có trong mapping
            row_out['BOMIdList'] = bom_list

            # ── 2) Run _SP_CONFIG cho THDM_HEADER theo thứ tự ──────────────
            # Dùng _call_sp / _sp_lookup (cùng BOM) để xử lý đúng parameterize
            sp_thdm = [s for s in sp_config
                       if s.get('section') == 'THDM_HEADER'
                       and s.get('isactive') == '1']
            for sp_rec in sp_thdm:
                sql_col  = sp_rec.get('sql_column', '')
                sp_name  = sp_rec.get('sp_name', '')
                params_s = sp_rec.get('params', '')
                fallback = sp_rec.get('fallback', '')
                out_flds = [f.strip() for f in sp_rec.get('outputfields', '').split(',') if f.strip()]
                try:
                    if sp_name.lower() == 'lookup':
                        val = self._sp_lookup(conn, params_s, row_out)
                        row_out[sql_col] = val if val is not None else (int(fallback) if fallback else None)
                    else:
                        r = self._call_sp(conn, sp_name, params_s, row_out)
                        if out_flds:
                            row_out[out_flds[0]] = r
                        elif sql_col:
                            row_out[sql_col] = r if r is not None else (fallback or None)
                except Exception as e_sp:
                    if fallback and sql_col:
                        row_out[sql_col] = fallback
                    else:
                        raise RuntimeError(f"SP '{sp_name}' thất bại và không có fallback: {e_sp}") from e_sp

            # ── 3) INSERT parent — dynamic từ row_out ──────────────────────
            view_hdr   = mapping.get('_CONFIG', {}).get('THDM_HEADER', {}).get(
                'view_insert', 'vB30BizDocDemand_Edit')
            # Loại bỏ SP-lookup intermediates (Ws_Id...) — dùng làm param SP
            # nhưng không phải column của view → dùng cùng pattern BOM _sp_temp_fields
            _sp_temp_header = {
                s['sql_column'] for s in sp_config
                if s.get('section') == 'THDM_HEADER'
                and s.get('sp_name', '').lower() == 'lookup'
                and s.get('isactive') == '1'
            }
            # IsDraftData: column không tồn tại trong vB30BizDocDemand_Edit
            # (chỉ có trong vB20BOM_Edit) → exclude khỏi THDM INSERT
            _thdm_hdr_skip = _sp_temp_header | {'IsDraftData'}
            valid_cols = {r['sql_col'] for r in header_map} | {'BOMIdList'}
            ins_cols   = [c for c in row_out if c in valid_cols and c not in _thdm_hdr_skip]
            ins_vals   = [row_out[c] for c in ins_cols]
            ph         = ', '.join('?' * len(ins_cols))
            cur.execute(
                f"INSERT INTO {view_hdr} ({', '.join(ins_cols)}) VALUES ({ph})",
                ins_vals)

            biz_doc_id = row_out.get('BizDocId')

            # ── 4) INSERT child sections — generic loop ────────────────────────
            _insert_count = 0
            for sec in child_sections:
                sec_cfg    = cfg[sec]
                sec_map    = mapping.get(sec, [])
                expand_muc = sec_cfg.get('expand_muc', False)
                view_sec   = sec_cfg.get('view_insert', 'B30BizDocDemandBOM')

                # Rows thuộc section này
                sec_rows = [r for r in rows if r.get('_section') == sec]
                if not sec_rows:
                    continue

                # SP_HOOK BeforeInsert cho section này
                sec_hooks_before = [
                    h for h in mapping.get('SP_HOOK', [])
                    if h.get('section', '') == sec
                    and h.get('event', '').lower() == 'beforeinsert'
                    and h.get('isactive', '') == '1'
                ]

                # Build cache cho Excel lookup fields (ItemId, CustomerId, ...)
                _lk_recs = [
                    rec for rec in sec_map
                    if rec.get('nguon_dl') == 'Excel'
                    and rec.get('kieu_lookup', '')
                    and rec.get('bang_master', '')
                    and rec.get('truong_lay_ve', '')
                ]
                _caches = self._build_all_caches(conn, _lk_recs) if _lk_recs else {}

                # MucLookup + DetailRowId_SO chỉ cho sections có expand_muc=True
                _detail_so_cache = []
                _detail_so_col   = 'DetailRowId_SO'
                if expand_muc:
                    _detail_so_rec = next(
                        (r for r in sec_map
                         if r.get('nguon_dl') == 'MucLookup'
                         and r.get('bang_master')
                         and r.get('truong_so_sanh')
                         and r.get('truong_lay_ve')),
                        None
                    )
                    if _detail_so_rec:
                        try:
                            _detail_so_cache = self._build_cache_generic(
                                conn,
                                _detail_so_rec['bang_master'],
                                _detail_so_rec.get('dieu_kien_master', ''),
                                _detail_so_rec['truong_so_sanh'],
                                _detail_so_rec['truong_lay_ve'],
                            )
                        except Exception as _e:
                            print(f'[THDM] Build cache DetailRowId_SO ({sec}): {_e}')
                        _detail_so_col = _detail_so_rec['sql_col']

                for idx, r in enumerate(sec_rows, 1):
                    try:
                        resolved = self._resolve_thdm_thvt_row(
                            r, biz_doc_id, idx, sec_map, now)
                        if not resolved:
                            continue

                        # Convert Excel code → DB Id qua lookup (ItemId, ...)
                        for _lrec in _lk_recs:
                            _sc    = _lrec.get('sql_col', '')
                            _kl    = _lrec.get('kieu_lookup', '')
                            _bm    = _lrec.get('bang_master', '')
                            _dk    = _lrec.get('dieu_kien_master', '')
                            _ss    = _lrec.get('truong_so_sanh', '')
                            _lv    = _lrec.get('truong_lay_ve', '')
                            _raw   = resolved.get(_sc)
                            _cache = _caches.get((_bm, _dk, _ss, _lv), [])
                            if _raw is not None and _cache:
                                _found, _ = self._lookup_generic(_raw, _cache, _kl)
                                resolved[_sc] = _found

                        # Compound lookup DetailRowId_SO (chỉ expand_muc sections)
                        if _detail_so_cache:
                            _muc_raw = resolved.get(_detail_so_col)
                            if _muc_raw is not None:
                                _muc_key  = str(_muc_raw).replace('Muc ', '').strip()
                                _compound = (_muc_key, str(order_id))
                                _found_id, _ = self._lookup_generic(
                                    _compound, _detail_so_cache, 'exact')
                                resolved[_detail_so_col] = _found_id
                            # Factory-only / lookup fail → '' thay NULL
                            if resolved.get(_detail_so_col) is None:
                                resolved[_detail_so_col] = ''

                        # SP_HOOK BeforeInsert
                        if sec_hooks_before:
                            _run_row_sp_hooks(conn, sec_hooks_before, resolved,
                                              log_fn=None)

                        cols_c = list(resolved.keys())
                        vals_c = [resolved[c] for c in cols_c]
                        ph_c   = ", ".join("?" * len(cols_c))
                        cur.execute(
                            f"INSERT INTO {view_sec} ({', '.join(cols_c)}) VALUES ({ph_c})",
                            vals_c)
                        _insert_count += 1
                        if _insert_count % 50 == 0:
                            self.after(0, lambda n=_insert_count, t=len(rows):
                                self._update_loading_msg(
                                    f"Đang tạo phiếu THDM...\n"
                                    f"Đã insert {n:,}/{t:,} dòng vật tư"))
                    except Exception as _row_err:
                        raise RuntimeError(
                            f"[{sec}] Lỗi tại dòng #{idx} "
                            f"({r.get('ItemName') or r.get('ItemId') or '?'}): {_row_err}"
                        ) from _row_err

            conn.commit()
            self.after(0, lambda n=_insert_count: self._thdm_insert_done(n, None))
        except Exception as e:
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
            self.after(0, lambda err=e: self._thdm_insert_done(0, str(err)))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _thdm_insert_done(self, n_rows, error):
        try:
            if self._loading_dlg:
                self._loading_dlg.destroy()
                self._loading_dlg = None
        except Exception:
            pass
        self.btn_thdm_insert.configure(
            state="normal", text="💾  ⑤ Tạo THDM")
        _fname = os.path.basename(self._thdm_excel_path or '')
        if error:
            self._db_log('THDM', _fname, self._thdm_selected_order_id, 0, 'LOI', error)
            self._show_msg("Lỗi tạo THDM", error)
            return
        self._db_log('THDM', _fname, self._thdm_selected_order_id, n_rows, 'OK',
                     f"Đã tạo phiếu THDM với {n_rows:,} dòng vật tư")
        self._show_msg("Thành công",
            f"✅  Đã tạo phiếu THDM với {n_rows:,} dòng vật tư.", kind="info")
        self._log(
            f"Đơn hàng {self._thdm_selected_order_id}",
            "THDM", str(n_rows), "OK",
            f"Đã tạo phiếu THDM với {n_rows:,} dòng vật tư",
            "ok")

    # ─ Tab 4: Lịch sử Log ──────────────────────────────────────────────────────────────────────
    def _build_log_panel(self):
        self._log_expanded = False
        self.log_panel = ctk.CTkFrame(self, corner_radius=0, height=40)
        self.log_panel.grid(row=2, column=0, sticky="ew")
        self.log_panel.grid_propagate(False)
        self.log_panel.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(self.log_panel, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=8, pady=6)
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text="📋  Lịch sử Import",
            font=ctk.CTkFont("Segoe UI", 12, "bold")).grid(
            row=0, column=0, sticky="w")
        self.btn_log_db = ctk.CTkButton(hdr, text="☁  Log DB",
            command=self._load_db_log,
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#007ACC"),
            hover_color=("#DBEAFE","#1E3A5F"),
            font=ctk.CTkFont("Segoe UI", 12),
            width=90, height=24, corner_radius=6)
        self.btn_log_db.grid(row=0, column=1, padx=(4, 0))
        Tooltip(self.btn_log_db,
                lambda: "Tải 200 dòng log import gần nhất từ database"
                        " (bảng BOMTool_ImportLog)")
        ctk.CTkButton(hdr, text="🗑  Xóa",
            command=self._clear_log,
            fg_color="transparent", border_width=1,
            text_color="gray", hover_color=("gray90","gray25"),
            font=ctk.CTkFont("Segoe UI", 12),
            width=70, height=24, corner_radius=6).grid(
            row=0, column=2, padx=(4, 4))
        self.btn_log_toggle = ctk.CTkButton(hdr,
            text="▲  Mở log",
            command=self._toggle_log,
            fg_color="transparent", border_width=1,
            text_color=("#3B82F6","#007ACC"),
            hover_color=("#DBEAFE","#1E3A5F"),
            font=ctk.CTkFont("Segoe UI", 12, "bold"),
            width=90, height=24, corner_radius=6)
        self.btn_log_toggle.grid(row=0, column=3)

        self.log_content = tk.Frame(self.log_panel, bg="#1E1E1E", height=200)

        LOG_COLS = ["Thời gian","Tên file","Sheet","Tổng dòng","Trạng thái","Chi tiết"]
        if _HAS_TKSHEET:
            self.log_tree = SheetTable(self.log_content, columns=LOG_COLS)
            self.log_tree.grid(row=0, column=0, sticky="nsew")
        else:
            vsb = ttk.Scrollbar(self.log_content, orient=tk.VERTICAL)
            self.log_tree = ttk.Treeview(self.log_content, style="BOM.Treeview",
                show="headings", columns=LOG_COLS, yscrollcommand=vsb.set, height=7)
            vsb.configure(command=self.log_tree.yview)
            self.log_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
        for col, w in zip(LOG_COLS, [155,250,80,80,100,500]):
            self.log_tree.heading(col, text=col)
            self.log_tree.column(col, width=w, anchor="w", stretch=(col=="Chi tiết"))
        self.log_content.rowconfigure(0, weight=1)
        self.log_content.columnconfigure(0, weight=1)
        self.log_tree.tag_configure("ok",   foreground="#4EC9B0")
        self.log_tree.tag_configure("warn", foreground="#D7BA7D")
        self.log_tree.tag_configure("err",  foreground="#F48771")
        self._bind_copy(self.log_tree, "Log")

    def _toggle_log(self):
        if self._log_expanded:
            self.log_content.grid_forget()
            self.log_panel.configure(height=40)
            self.btn_log_toggle.configure(text="▲  Mở log")
            self._log_expanded = False
        else:
            self.log_content.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=4, pady=(0,4))
            self.log_panel.configure(height=230)
            self.btn_log_toggle.configure(text="▼  Đóng log")
            self._log_expanded = True

    # ─ Helpers ───────────────────────────────────────────────────────────────────────────────────
    def _log(self, fname, sheet, rows, status, detail, tag):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_tree.insert("", 0,
            values=(ts, fname, sheet, rows, status, detail), tags=(tag,))
        if tag in ('error', 'warn'):
            try:
                _lp = os.path.join(os.path.dirname(sys.executable)
                                   if getattr(sys, 'frozen', False)
                                   else os.path.dirname(os.path.abspath(__file__)),
                                   'boho_import.log')
                with open(_lp, 'a', encoding='utf-8') as _lf:
                    _lf.write(f"[{ts}] [{tag.upper()}] {sheet} | {status} | {detail}\n")
            except Exception:
                pass

    def _clear_log(self):
        self.log_tree.delete(*self.log_tree.get_children())

    # ── DB Log: bảng BOMTool_ImportLog, nằm ở DB riêng [BOMTool] (không
    # chung DB với Bravo) — tạo bằng sql/create_[BOMTool].sql ──────────────
    DB_LOG_DATABASE = '[BOMTool]'
    DB_LOG_TABLE = f'{DB_LOG_DATABASE}.dbo.BOMTool_ImportLog'

    def _db_log(self, action, fname, ref_id, n_rows, status, detail):
        """Ghi 1 dòng log import vào bảng DB (chạy nền).
        Bảng chưa được tạo trên DB → bỏ qua êm, không làm phiền user."""
        def _worker():
            try:
                import getpass, socket
                conn = self._get_db_conn(timeout_sec=5)
                cur  = conn.cursor()
                cur.execute(
                    f"INSERT INTO {self.DB_LOG_TABLE} "
                    "(Computer, LoginUser, Creator, Action, FileName, RefId, "
                    "TotalRows, Status, Detail, AppVersion) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (socket.gethostname()[:100],
                     getpass.getuser()[:100],
                     str(getattr(self, '_current_creator_code', '') or '')[:100],
                     action,
                     (fname or '')[:255],
                     str(ref_id or '')[:50],
                     int(n_rows or 0),
                     status,
                     (detail or '')[:4000],
                     APP_VERSION))
                conn.commit()
                conn.close()
            except Exception as e:
                print(f"[DBLog] Bỏ qua ghi log DB: {e}")
        threading.Thread(target=_worker, daemon=True).start()

    def _load_db_log(self):
        self.btn_log_db.configure(state="disabled", text="⏳ Đang tải...")
        threading.Thread(target=self._load_db_log_worker, daemon=True).start()

    def _load_db_log_worker(self):
        try:
            conn = self._get_db_conn(timeout_sec=5)
            cur  = conn.cursor()
            cur.execute(
                f"SELECT TOP 200 LogTime, FileName, Action, TotalRows, Status, "
                f"Detail, Computer, LoginUser "
                f"FROM {self.DB_LOG_TABLE} ORDER BY Id DESC")
            rows = cur.fetchall()
            conn.close()
            self.after(0, lambda: self._load_db_log_done(rows, None))
        except Exception as e:
            self.after(0, lambda err=str(e): self._load_db_log_done(None, err))

    def _load_db_log_done(self, rows, error):
        self.btn_log_db.configure(state="normal", text="☁  Log DB")
        if error is not None:
            if (self.DB_LOG_DATABASE in error or 'Invalid object name' in error
                    or 'does not exist' in error or 'Cannot open database' in error):
                self._show_msg("Bảng log chưa có trên DB",
                    f"Database/bảng {self.DB_LOG_TABLE} chưa được tạo.\n\n"
                    "Gửi file sql\\create_[BOMTool].sql cho khách hàng chạy,\n"
                    "sau đó bấm lại nút này là xem được log.", kind="info")
            else:
                self._show_msg("Lỗi tải log DB", error, kind="error")
            return
        self._clear_log()
        for r in rows:
            ts     = r[0].strftime("%Y-%m-%d %H:%M:%S") if r[0] else ''
            status = (r[4] or '').strip()
            tag    = 'ok' if status.upper() == 'OK' else 'err'
            detail = f"{r[5] or ''}  [{r[6] or ''} / {r[7] or ''}]"
            self.log_tree.insert("", "end",
                values=(ts, r[1] or '', r[2] or '', r[3] if r[3] is not None else '',
                        status, detail),
                tags=(tag,))
        if not self._log_expanded:
            self._toggle_log()

    def _set_status(self, text, color=None):
        self.lbl_status.config(text=text, fg=(color or ("gray40","gray60")))

    def _get_password_for_file(self, filepath):
        """
        Hiện password dialog trên main thread, chờ kết quả từ worker thread.
        Trả về: password str hoặc None nếu hủy.
        """
        event    = threading.Event()
        result   = [None]

        def show_on_main():
            result[0] = _ask_excel_password(filepath)
            event.set()

        self.after(0, show_on_main)
        event.wait()
        return result[0]

    def _open_excel_with_password(self, filepath):
        """
        Mở file Excel, tự động hỏi mật khẩu nếu file bị mã hóa.
        Gọi từ worker thread.
        Trả về: (decrypted_bytes hoặc None, error_msg hoặc None)
        - None, None  → file không có password, dùng filepath bình thường
        - bytes, None → file có password, đã giải mã thành công
        - None, msg   → lỗi (hủy hoặc sai password)
        """
        if not _is_encrypted_excel(filepath):
            return None, None  # không cần decrypt

        # Nếu không có msoffcrypto, không thể giải mã — báo người dùng cài
        if not _HAS_MSOFFCRYPTO:
            return None, (
                "File Excel có mật khẩu nhưng thư viện giải mã chưa được cài.\n\n"
                "Chạy lệnh sau trong terminal rồi khởi động lại tool:\n"
                "    pip install msoffcrypto-tool"
            )

        # Hỏi mật khẩu (tối đa 3 lần)
        for attempt in range(3):
            password = self._get_password_for_file(filepath)
            if password is None:
                return None, "cancelled"
            try:
                decrypted = _decrypt_excel(filepath, password)
                return decrypted, None
            except Exception:
                if attempt < 2:
                    event2 = threading.Event()
                    _att = attempt + 1
                    def _warn(a=_att):
                        messagebox.showwarning(
                            "Sai mật khẩu",
                            f"Mật khẩu không đúng (lần {a}/3).\nVui lòng thử lại.")
                        event2.set()
                    self.after(0, _warn)
                    event2.wait()

        return None, "Sai mật khẩu 3 lần. Không thể mở file."

    def _load_db_config(self):
        """Đọc db_config.json. Tạo template nếu chưa có."""
        import json
        if not os.path.exists(DB_CONFIG_FILE):
            template = {
                "server":   "",
                "database": "",
                "username": "",
                "password": "",
                "driver":   "ODBC Driver 17 for SQL Server"
            }
            try:
                with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(template, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            return None
        try:
            import json as _json
            with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = _json.load(f)
            if cfg.get("server") and cfg.get("database"):
                return cfg
            return None
        except Exception:
            return None

    # ─ Export validate report ────────────────────────────────────────────────────────────────────────────
    def _export_validate_report(self):
        if not self.val_errors:
            self._show_msg("Chưa validate", "Hãy chạy Validate trước.", 'warning')
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.basename(self._current_file)
        fname_clean = re.sub(r'[\\/*?:\[\]]', '_', fname.rsplit(".", 1)[0])
        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="BOM_Validate_" + fname_clean + "_" + ts + ".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not out_path:
            return

        wb  = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Tong hop"

        def hdr_cell(ws, r, c, val, bg="1E3A5F"):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return cell

        for j, h in enumerate(["Bảng / Sheet", "Tổng lỗi", "Tổng cảnh báo",
                                "Trạng thái", "Chi tiết đầu tiên"], 1):
            hdr_cell(ws0, 1, j, h)
        ws0.column_dimensions["A"].width = 38
        ws0.column_dimensions["B"].width = 12
        ws0.column_dimensions["C"].width = 16
        ws0.column_dimensions["D"].width = 18
        ws0.column_dimensions["E"].width = 60

        for i, (label, errs) in enumerate(self.val_errors.items(), 2):
            n_e  = len([e for e in errs if e["severity"] == "error"])
            n_w  = len([e for e in errs if e["severity"] == "warning"])
            first = errs[0]["message"] if errs else "—"
            status = "OK" if n_e == 0 else (str(n_e) + " loi")
            ws0.cell(row=i, column=1, value=label)
            c2 = ws0.cell(row=i, column=2, value=n_e)
            c2.font = Font(color="C62828" if n_e else "000000", bold=bool(n_e))
            c3 = ws0.cell(row=i, column=3, value=n_w)
            c3.font = Font(color="E65100" if n_w else "000000")
            ws0.cell(row=i, column=4, value=status)
            ws0.cell(row=i, column=5, value=first)
            if i % 2 == 0:
                for c in range(1, 6):
                    ws0.cell(row=i, column=c).fill = PatternFill("solid", fgColor="F8FAFC")

        for label, errs in self.val_errors.items():
            if not errs:
                continue
            safe = re.sub(r'[\\/*?:\[\]]', '_', label)[:31]
            ws2  = wb.create_sheet(safe)
            for j, h in enumerate(["Dòng", "Cột / Trường", "Thông báo lỗi", "Mức độ"], 1):
                hdr_cell(ws2, 1, j, h)
            ws2.column_dimensions["A"].width = 10
            ws2.column_dimensions["B"].width = 28
            ws2.column_dimensions["C"].width = 70
            ws2.column_dimensions["D"].width = 14
            for i, err in enumerate(errs, 2):
                ws2.cell(row=i, column=1, value=str(err.get("row", "")))
                ws2.cell(row=i, column=2, value=str(err.get("col", "")))
                ws2.cell(row=i, column=3, value=err.get("message", ""))
                sc = ws2.cell(row=i, column=4, value=err.get("severity", ""))
                if err.get("severity") == "error":
                    sc.font = Font(color="C62828", bold=True)
                elif err.get("severity") == "warning":
                    sc.font = Font(color="E65100")

        wb.save(out_path)
        self._log(fname, "Report", "—", "OK - Da xuat",
                  os.path.basename(out_path), "ok")
        self._show_export_success("Xuất báo cáo", "Đã lưu:\n" + out_path, out_path)

    # ─ Batch nhieu file ───────────────────────────────────────────────────────────────────────────────────
    def _open_multiple_files(self):
        paths = filedialog.askopenfilenames(
            title="Chọn nhiều file BOM Excel",
            filetypes=[("Excel BOM", "*.xlsm *.xlsx"), ("Tất cả", "*.*")])
        if not paths:
            return
        self.btn_open.config(state=tk.DISABLED)
        self.btn_batch.config(state=tk.DISABLED)
        self._set_status("Đang xử lý " + str(len(paths)) + " file...", C["yellow"])

        def worker():
            mk = build_meta_keys_from_mapping(self.mapping)
            results = []
            for path in paths:
                try:
                    decrypted, err = self._open_excel_with_password(path)
                    if err == "cancelled":
                        results.append({
                            "path": path,
                            "error": "Bỏ qua — file có mật khẩu và người dùng đã hủy",
                            "n_err": 1, "n_wrn": 0
                        })
                        continue
                    if err:
                        results.append({
                            "path": path, "error": err,
                            "n_err": 1, "n_wrn": 0
                        })
                        continue

                    tables, meta, skipped, warns = parse_bom_file(
                        path, meta_keys=mk, _decrypted_bytes=decrypted)
                    val_errors = validate_layer1(tables, meta, self.mapping)
                    n_err, n_wrn = count_errors(val_errors)
                    results.append({
                        "path": path, "tables": tables, "meta": meta,
                        "val_errors": val_errors,
                        "n_err": n_err, "n_wrn": n_wrn,
                        "skipped": skipped
                    })
                except Exception as ex:
                    results.append({
                        "path": path, "error": str(ex),
                        "n_err": 1, "n_wrn": 0
                    })
            self.after(0, lambda: self._on_batch_done(results))

        threading.Thread(target=worker, daemon=True).start()

    def _on_batch_done(self, results):
        n_ok = sum(1 for r in results if r.get("n_err", 1) == 0)
        for r in results:
            fname = os.path.basename(r["path"])
            if "error" in r:
                self._log(fname, "Batch", "—", "Lỗi đọc file", r["error"], "err")
            else:
                total = sum(len(t["df"]) for t in r["tables"].values())
                tag   = "ok" if r["n_err"] == 0 else "err"
                detail = str(r["n_err"]) + " loi, " + str(r["n_wrn"]) + " canh bao"
                status = "OK" if r["n_err"] == 0 else (str(r["n_err"]) + " loi")
                self._log(fname, "Batch", total, status, detail, tag)

        last_ok = next((r for r in reversed(results) if "tables" in r), None)
        if last_ok:
            self.tables        = last_ok["tables"]
            self.global_meta   = last_ok["meta"]
            self.val_errors    = last_ok["val_errors"]
            self._current_file = last_ok["path"]
            self._on_parsed(last_ok.get("skipped", []), [])
            n_err, n_wrn = count_errors(last_ok["val_errors"])
            self.listbox.delete(0, tk.END)
            for name in self.tables:
                errs = last_ok["val_errors"].get(name, [])
                ne   = len([e for e in errs if e["severity"] == "error"])
                nw   = len([e for e in errs if e["severity"] == "warning"])
                icon = (" E" + str(ne)) if ne else ((" W" + str(nw)) if nw else " OK")
                self.listbox.insert(tk.END, name + icon)
            self.btn_report.config(state=tk.NORMAL)
            self.btn_import.config(state=tk.NORMAL if n_err == 0 else tk.DISABLED)

        self._toggle_log() if not self._log_expanded else None
        self.btn_open.config(state=tk.NORMAL)
        self.btn_batch.config(state=tk.NORMAL)
        status_txt = ("OK: " + str(n_ok) + "/" + str(len(results)) + " file")
        self._set_status(status_txt, C["green"] if n_ok == len(results) else C["yellow"])
        self._show_msg("Batch hoàn tất",
            f"Đã xử lý {len(results)} file\n"
            f"OK: {n_ok}  |  Lỗi: {len(results)-n_ok}\n\n"
            "Xem chi tiết ở tab Lịch sử Log", 'info')

    # ─ Selection ───────────────────────────────────────────────────────────────────────────────────
    def _on_select(self, _event):
        sel = self.listbox.curselection()
        if not sel:
            return

        raw        = self.listbox.get(sel[0])
        table_name = next((k for k in self.tables if raw.startswith(k)), None)
        if not table_name:
            return

        tbl  = self.tables[table_name]
        df   = tbl["df"]
        errs = self.val_errors.get(table_name, [])
        err_msgs = [e["message"] for e in errs if e["severity"] == "error"]

        n_rows = len(df)
        n_cols = len(df.columns)
        self.lbl_table.config(
            text="Bảng: " + table_name + "   —   " + str(n_rows) + " dòng × " + str(n_cols) + " cột",
            fg=C["text"])

        self.tree.delete(*self.tree.get_children())
        if df.empty or "Lỗi" in df.columns:
            self._update_warn_panel(errs, {}, [])
            self.tree["columns"] = ("msg",)
            self.tree.heading("msg", text="Thông báo")
            self.tree.column("msg", width=600)
            if not df.empty:
                self.tree.insert("", tk.END, values=(df.iloc[0]["Lỗi"],))
            return

        cols = list(df.columns)
        self.tree["columns"] = cols
        import tkinter.font as tkfont
        _hfont = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        _dfont = tkfont.Font(family="Segoe UI", size=12)
        _PADDING = 28
        _MIN_W   = 60
        _MAX_W   = 300  # giới hạn tối đa tránh cột quá rộng

        # Tính width từ heading + sample 60 dòng đầu dữ liệu
        col_w = {}
        for col in cols:
            display = col.replace("_", " ")
            self.tree.heading(col, text=display, anchor="center")
            h_w = _hfont.measure(display) + _PADDING
            try:
                sample = df[col].dropna().astype(str).head(60)
                d_w = int(sample.map(lambda s: _dfont.measure(s)).max()) + _PADDING if len(sample) else 0
            except Exception:
                d_w = 0
            col_w[col] = min(max(h_w, d_w, _MIN_W), _MAX_W)
            self.tree.column(col, width=col_w[col], anchor="center", stretch=False)

        # ─ Hàng SQL column names (hàng đầu cố định) ──────────────────────────
        section = tbl.get("type", "")
        sql_to_excel = {}
        sql_row = []
        if section and self.mapping:
            for _rec in self.mapping.get(section, []):
                if _rec.get("sql_col") and _rec.get("ten_excel"):
                    sql_to_excel[_rec["sql_col"]] = _rec["ten_excel"]
            rev_map = build_reverse_map(self.mapping, section)
            for col in cols:
                sql_col = match_col_to_sql(_norm_vn(col), rev_map)
                sql_row.append(sql_col if sql_col else "—")
            # Mở rộng cột nếu tên SQL dài hơn nội dung đã tính
            for col, sv in zip(cols, sql_row):
                s_w = _dfont.measure(str(sv)) + _PADDING
                if s_w > col_w[col]:
                    col_w[col] = min(s_w, _MAX_W)
                    self.tree.column(col, width=col_w[col])
            self.tree.insert("", 0, values=sql_row, tags=("sql_names",))

        # ─ Data rows ───────────────────────────────────────────────────────────────────────────────────
        if hasattr(self, "_empty_frame"):
            self._empty_frame.place_forget()

        # Build cell-level error maps: {row_0based: set(excel_col_name)}
        error_rows  = set()
        warn_rows   = set()
        error_cells = {}   # row_0based → set of col names có lỗi
        warn_cells  = {}
        for e in errs:
            r = e.get("row")
            if not isinstance(r, int):
                continue
            r0 = r - 1
            sev = e.get("severity", "error")
            # Map SQL col → Excel col name trong df
            sql_col   = e.get("col", "")
            excel_col = sql_to_excel.get(sql_col, sql_col)
            matched   = next((c for c in cols
                              if c == excel_col or _norm_vn(c) == _norm_vn(excel_col)), None)
            if sev == "error":
                error_rows.add(r0)
                if matched:
                    error_cells.setdefault(r0, set()).add(matched)
            else:
                warn_rows.add(r0)
                if matched:
                    warn_cells.setdefault(r0, set()).add(matched)

        # Cập nhật warning panel
        self._update_warn_panel(errs, error_cells, cols)

        # Lưu map iid → tree iid để panel có thể jump
        self._warn_iid_to_tree_iid = {}
        tree_iids = []  # sẽ gán sau khi insert

        for _idx, (_, row) in enumerate(df.iterrows()):
            vals = []
            for col, v in zip(cols, row):
                s = str(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
                if _idx in error_cells and col in error_cells[_idx]:
                    s = "⚠ " + s if s else "⚠ —"
                elif _idx in warn_cells and col in warn_cells[_idx]:
                    s = "△ " + s if s else "△ —"
                vals.append(s)
            if _idx in error_rows:
                row_tag = "err_row"
            elif _idx in warn_rows:
                row_tag = "warn_row"
            else:
                row_tag = "evenrow" if _idx % 2 == 0 else "oddrow"
            iid = self.tree.insert("", tk.END, values=vals, tags=(row_tag,))
            tree_iids.append((_idx, iid))

        # Gán mapping warn_panel_iid → main tree iid (sau khi insert xong)
        # _update_warn_panel đã lưu row_0based trong warn_tree items
        for w_iid in self._warn_err_tree.get_children():
            r0 = self._warn_err_tree.set(w_iid, "rownum")
            # rownum được lưu dạng "Dòng N" — extract số
            try:
                r0_int = int(r0.replace("Dòng", "").strip()) - 1
                matched_iid = next((iid for idx, iid in tree_iids if idx == r0_int), None)
                if matched_iid:
                    self._warn_iid_to_tree_iid[w_iid] = matched_iid
            except (ValueError, StopIteration):
                pass

    # ─ Warning panel helpers ──────────────────────────────────────────────────────────────────────────────

    def _update_warn_panel(self, errs, error_cells, cols):
        """Populate warning panel từ danh sách errs sau khi _show_table xong."""
        self._warn_err_tree.delete(*self._warn_err_tree.get_children())
        self._warn_iid_to_tree_iid = {}

        err_list  = [e for e in errs if e.get("severity") == "error"   and isinstance(e.get("row"), int)]
        warn_list = [e for e in errs if e.get("severity") == "warning" and isinstance(e.get("row"), int)]
        # Thêm các lỗi không có row (header-level)
        hdr_errs  = [e for e in errs if e.get("severity") == "error"   and not isinstance(e.get("row"), int)]

        n_err  = len(err_list) + len(hdr_errs)
        n_warn = len(warn_list)

        if n_err == 0 and n_warn == 0:
            self._warn_outer.grid_remove()
            return

        # Hiển thị panel (grid_remove giữ cấu hình, grid() khôi phục)
        self._warn_outer.grid()

        badge_txt = (str(n_err) + " lỗi") + (" / " + str(n_warn) + " cảnh báo" if n_warn else "")
        self._warn_badge.config(text=badge_txt)

        first_msg = (err_list or hdr_errs or warn_list)[0].get("message", "")
        if len(first_msg) > 80:
            first_msg = first_msg[:77] + "..."
        extra = (n_err + n_warn - 1)
        suffix = ("  +" + str(extra) + " mục khác") if extra > 0 else ""
        self._warn_hdr_lbl.config(text=first_msg + suffix)

        # Điền danh sách lỗi vào warn_tree
        all_items = [(e, "err_item") for e in hdr_errs + err_list] + \
                    [(e, "warn_item") for e in warn_list]
        for e, tag in all_items:
            r = e.get("row")
            row_disp  = ("Dòng " + str(r)) if isinstance(r, int) else "Header"
            field_disp = e.get("col", "")
            msg_disp   = e.get("message", "")
            self._warn_err_tree.insert("", tk.END,
                values=(row_disp, field_disp, msg_disp), tags=(tag,))

        # Số dòng hiển thị: tối đa 5, tối thiểu 2
        n_items = len(all_items)
        self._warn_err_tree.config(height=min(max(n_items, 2), 5))

        # Đảm bảo list hiển thị theo trạng thái toggle
        if self._warn_expanded:
            self._warn_list_outer.pack(fill=tk.X)
        else:
            self._warn_list_outer.pack_forget()

    def _toggle_warn_panel(self):
        """Thu gọn / mở rộng phần danh sách lỗi."""
        self._warn_expanded = not self._warn_expanded
        if self._warn_expanded:
            self._warn_list_outer.pack(fill=tk.X)
            self._warn_toggle.config(text="▾")
        else:
            self._warn_list_outer.pack_forget()
            self._warn_toggle.config(text="▸")

    def _on_warn_item_click(self, event):
        """Click vào 1 dòng trong warning panel → scroll + select dòng đó trong main tree."""
        sel = self._warn_err_tree.selection()
        if not sel:
            return
        w_iid = sel[0]
        tree_iid = self._warn_iid_to_tree_iid.get(w_iid)
        if tree_iid and self.tree.exists(tree_iid):
            self.tree.selection_set(tree_iid)
            self.tree.see(tree_iid)
            self.tree.focus(tree_iid)

    # ─ Actions ────────────────────────────────────────────────────────────────────────────────────────────
    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Chọn file BOM Excel",
            filetypes=[("Excel BOM", "*.xlsm *.xlsx"), ("Tất cả", "*.*")])
        if not path:
            return

        self._set_status("Đang bóc tách dữ liệu...", C["yellow"])
        self.btn_open.config(state=tk.DISABLED)
        self.val_errors = {}
        self.update()

        def worker():
            try:
                # Kiểm tra + giải mã nếu file có password
                decrypted, err = self._open_excel_with_password(path)
                if err == "cancelled":
                    self.after(0, lambda: self._set_status("Đã hủy — file có mật khẩu", C["yellow"]))
                    return
                if err:
                    self.after(0, lambda: self._show_msg("Lỗi mật khẩu", err, 'error'))
                    self.after(0, lambda: self._set_status("Lỗi mật khẩu", C["red"]))
                    return

                mk     = build_meta_keys_from_mapping(self.mapping)
                tables, meta, skipped, warns = parse_bom_file(
                    path, meta_keys=mk, _decrypted_bytes=decrypted)
                self.tables        = tables
                self.global_meta   = meta
                self._current_file = path
                self.after(0, lambda: self._on_parsed(skipped, warns))
            except Exception as ex:
                msg = str(ex)
                self.after(0, lambda: self._show_msg("Lỗi đọc file", msg, 'error'))
                self.after(0, lambda: self._set_status("Lỗi đọc file", C["red"]))
            finally:
                self.after(0, lambda: self.btn_open.config(state=tk.NORMAL))

        threading.Thread(target=worker, daemon=True).start()

    def _on_parsed(self, skipped, warns):
        fname = os.path.basename(self._current_file)
        if skipped:
            self._skipped_sheets = skipped
            skip_info = f"  |  Bỏ qua: {len(skipped)} sheet"
        else:
            self._skipped_sheets = []
            skip_info = ""
        self._set_status("OK: " + fname + skip_info, C["green"])

        total = sum(len(t["df"]) for t in self.tables.values())
        self._log(fname, "—", total, "Da parse",
                  ("Bo qua " + str(len(skipped)) + " sheet") if skipped else "OK",
                  "ok" if not warns else "warn")

        self.listbox.delete(0, tk.END)
        for name in self.tables:
            self.listbox.insert(tk.END, name)

        self.btn_validate.config(state=tk.NORMAL)
        self.btn_import.config(state=tk.DISABLED)
        self.btn_report.config(state=tk.DISABLED)

        if self.tables:
            self.listbox.selection_set(0)
            self._on_select(None)

    def _run_validate(self):
        if not self.tables:
            return
        self.mapping = load_mapping()   # reload mỗi lần validate
        self.val_errors = validate_layer1(self.tables, self.global_meta, self.mapping)
        n_err, n_wrn    = count_errors(self.val_errors)

        sel = self.listbox.curselection()
        self.listbox.delete(0, tk.END)
        for name in self.tables:
            errs = self.val_errors.get(name, [])
            ne   = len([e for e in errs if e["severity"] == "error"])
            nw   = len([e for e in errs if e["severity"] == "warning"])
            icon = (" E" + str(ne)) if ne else ((" W" + str(nw)) if nw else " OK")
            self.listbox.insert(tk.END, name + icon)

        if sel:
            self.listbox.selection_set(sel[0])
        self._on_select(None)

        fname = os.path.basename(self._current_file)
        self.btn_report.config(state=tk.NORMAL)

        # Bat ky sheet nao co loi deu chan Import
        header_errs = len([
            e for e in self.val_errors.get("[H] Phiếu Header", [])
            if e.get("severity") == "error"
        ])

        if n_err == 0:
            self.btn_import.config(state=tk.NORMAL)
            self._set_status("Validate OK — " + str(n_wrn) + " canh bao", C["green"])
            self._log(fname, "Validate", "—", "OK", str(n_wrn) + " canh bao", "ok")
            self._show_msg("Validate",
                "Dữ liệu hợp lệ!\n" + str(n_wrn) + " cảnh báo (không chặn import)", 'info')
        else:
            self.btn_import.config(state=tk.DISABLED)
            self._set_status(str(n_err) + " lỗi  |  " + str(n_wrn) + " cảnh báo", C["red"])
            self._log(fname, "Validate", "—", str(n_err) + " loi",
                      str(n_err) + " loi, " + str(n_wrn) + " canh bao", "err")
            if header_errs > 0:
                self._show_msg("Validate",
                    "HEADER có " + str(header_errs) + " lỗi — cần sửa trước khi import.\n\n"
                    "Tổng: " + str(n_err) + " lỗi / " + str(n_wrn) + " cảnh báo.\n\n"
                    "Vui lòng mở file Excel gốc, sửa các dòng bị đánh dấu đỏ,\n"
                    "lưu lại và thực hiện lại từ Bước 1.", 'warning')
            else:
                sub_errs = n_err - header_errs
                self._show_msg("Validate",
                    "Header OK — nhưng có " + str(sub_errs) + " lỗi dữ liệu ở BOM sheets.\n\n"
                    "Import bị khóa cho đến khi sửa hết lỗi.\n\n"
                    "Vui lòng mở file Excel gốc, sửa các dòng bị đánh dấu đỏ,\n"
                    "lưu lại và thực hiện lại từ Bước 1.", 'warning')

    # ── DB helpers ───────────────────────────────────────────────────────────
    def _get_db_conn(self, timeout_sec: int = 5):
        """
        Trả về pyodbc connection hoặc raise Exception với thông báo rõ ràng.
        timeout_sec: giây chờ kết nối — mặc định 5s để phát hiện VPN down nhanh.
        """
        import pyodbc
        cfg = self.db_cfg
        if not cfg:
            raise Exception("Chưa có cấu hình DB (db_config.json)")
        base = (
            f"DRIVER={{{cfg['driver']}}};"
            f"SERVER={cfg['server']};"
            f"DATABASE={cfg['database']};"
            "TrustServerCertificate=yes;"
        )
        if cfg.get("trusted_connection", "").lower() == "yes":
            conn_str = base + "Trusted_Connection=yes;"
        else:
            conn_str = base + f"UID={cfg['username']};PWD={cfg['password']};"

        # Ưu tiên timeout từ config; timeout_sec là fallback khi config không có
        timeout = int(cfg.get('timeout', timeout_sec))
        try:
            return pyodbc.connect(conn_str, timeout=timeout)
        except Exception as e:
            err = str(e)
            # Phân biệt lỗi timeout / network để gợi ý VPN
            if any(k in err.lower() for k in (
                'timeout', 'timed out', 'network', 'server', 'tcp',
                'named pipes', 'sql server', '08001', 'hn011',
            )):
                raise ConnectionError(
                    f"⚠️  Không kết nối được DB sau {timeout}s.\n\n"
                    f"Kiểm tra:\n"
                    f"  • VPN đã kết nối chưa?\n"
                    f"  • Server {cfg.get('server')} có ping được không?\n\n"
                    f"Chi tiết: {err}"
                ) from e
            raise

    # ── Generic SP caller ────────────────────────────────────────────────────
    def _call_sp(self, conn, sp_name, params_str, row):
        """
        Goi SP generic tu SP_CONFIG.
        params_str: "@Key=FixedValue|@Key2={FieldName}"
          {FieldName} lay tu row[FieldName], con lai la literal.
        Tra ve gia tri cot dau tien dong dau tien, hoac None.
        """
        parts = [p.strip() for p in params_str.split('|') if p.strip()]
        param_clauses, param_values = [], []
        for part in parts:
            if '=' in part:
                key, val = part.split('=', 1)
                key = key.strip().lstrip('@')
                val = val.strip()
                if val.startswith('{') and val.endswith('}'):
                    val = row.get(val[1:-1])
                elif val == '':
                    continue  # literal rỗng → SP dùng default
                # Chuẩn hóa kiểu: NaN → None, float nguyên → int
                import math as _m
                if isinstance(val, float):
                    if _m.isnan(val):
                        val = None
                    elif val == int(val):
                        val = int(val)
                # None → literal NULL (tránh pyodbc infer sai type)
                if val is None:
                    param_clauses.append(f'@{key}=NULL')
                else:
                    param_clauses.append(f'@{key}=?')
                    param_values.append(val)
        sql = f"EXEC {sp_name} " + ", ".join(param_clauses)
        cur = conn.cursor()
        cur.execute(sql, param_values)
        r = cur.fetchone()
        return r[0] if r else None

    def _sp_lookup(self, conn, params_str, row):
        """
        Kieu SP dac biet 'lookup': query bang master lay 1 gia tri.
        params_str: "Bang=TableName|Where=Col={FieldName}|Lay=ReturnCol"
        """
        cfg = {}
        for part in params_str.split('|'):
            if '=' in part:
                k, v = part.split('=', 1)
                cfg[k.strip()] = v.strip()
        table = cfg.get('Bang', '')
        where = cfg.get('Where', '')
        ret   = cfg.get('Lay', '')
        if not (table and where and ret):
            return None
        where_r = re.sub(r'\{(\w+)\}',
                         lambda m: str(row.get(m.group(1), '') or ''), where)
        if '=' in where_r:
            w_col, w_val = where_r.split('=', 1)
            # Không bracket table có dấu chấm (cross-db: B10_Boho.dbo.B00Branch)
            tbl_ref = table if '.' in table else f'[{table}]'
            sql = (f"SELECT TOP 1 [{ret.strip()}] FROM {tbl_ref}"
                   f" WITH(NOLOCK) WHERE [{w_col.strip()}]=?")
            try:
                cur = conn.cursor()
                cur.execute(sql, w_val.strip())
                r = cur.fetchone()
                if not r:
                    return None
                val = r[0]
                # Normalize numeric: float/str '1.0' → '1' (tránh lỗi string concat trong SP)
                import math as _m
                if isinstance(val, float):
                    val = None if _m.isnan(val) else (int(val) if val == int(val) else val)
                elif isinstance(val, str):
                    try:
                        f = float(val)
                        if f == int(f):
                            val = str(int(f))
                    except (ValueError, TypeError):
                        pass
                return val
            except Exception as e:
                self._log('', '_sp_lookup', 0, 'Warn', str(e), 'warn')
        return None

    # ── Generic validators runner ─────────────────────────────────────────────
    def _run_validators(self, conn, row):
        """
        Doc sheet VALIDATORS, chay tung validator IsActive=CO.
        Tra True = tiep tuc, False = huy.
        """
        for v in self.mapping.get('VALIDATORS', []):
            if v.get('isactive', '') != '1':
                continue
            sql      = v.get('sql', '').strip()
            params_s = v.get('params', '').strip()
            warn_msg = v.get('warningmessage', '').strip()
            name     = v.get('validatorname', '')
            if not sql:
                continue
            params = [row.get(p.strip()) for p in params_s.split(',') if p.strip()]
            try:
                cur = conn.cursor()
                cur.execute(sql, params)
                first = cur.fetchone()
                has_vio = False
                result_display = ''
                if first is not None:
                    fv = first[0]
                    if not (isinstance(fv, (int, float)) and fv == 0) and fv is not None:
                        has_vio = True
                        rest = cur.fetchmany(4)
                        result_display = ', '.join(str(r[0]) for r in [first] + list(rest))
                if has_vio:
                    msg = warn_msg.replace('{result}', result_display)
                    for k, val in row.items():
                        msg = msg.replace('{' + k + '}', str(val or ''))
                    if not self._ask_msg(f"Canh bao: {name}",
                                               msg + "\n\nVan tiep tuc insert?"):
                        return False
            except Exception as e:
                self._log('', f'Validator {name}', 0, 'Warn', str(e), 'warn')
        return True

    # ── Item lookup — 3 tiers ────────────────────────────────────────────────
    def _norm_for_match(self, s):
        """Normalize de so sanh: bo dau, lowercase, chi giu alphanum."""
        import unicodedata as _ud
        s = str(s).lower().strip().replace('đ', 'd')
        nfkd = _ud.normalize('NFKD', s)
        s = ''.join(c for c in nfkd if not _ud.combining(c))
        return re.sub(r'[^a-z0-9]', '', s)

    def _show_suggest_dialog(self, item_code, candidates):
        """
        Hien popup top candidates de user chon.
        candidates: [(score, item_dict), ...]
        Tra ve item_id duoc chon hoac None.
        """
        dlg = tk.Toplevel(self)
        dlg.title("Không tìm thấy mã chính xác — chọn tương đương")
        dlg.geometry("640x280")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=C["bg"])

        tk.Label(dlg, text=f'Ma Excel: "{item_code}"',
                 bg=C["bg"], fg=C["text"],
                 font=("Segoe UI", 12, "bold")).pack(pady=(12, 2))
        tk.Label(dlg, text="Chọn mã tương đương hoặc bỏ qua:",
                 bg=C["bg"], fg=C["muted"], font=("Segoe UI", 12)).pack()

        tree = ttk.Treeview(
            dlg, columns=("score", "code", "name"),
            show="headings", height=5, style="BOM.Treeview"
        )
        tree.heading("score", text="Do khop")
        tree.heading("code",  text="Mã B20Item")
        tree.heading("name",  text="Ten")
        tree.column("score", width=80,  anchor="center")
        tree.column("code",  width=160)
        tree.column("name",  width=360)
        for score, c in candidates:
            tree.insert("", tk.END,
                        values=(f"{score:.0f}%", c['code'], c['name']),
                        iid=str(c['id']))
        tree.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        result = [None]

        def on_choose():
            sel = tree.selection()
            if sel:
                result[0] = int(sel[0])
                dlg.destroy()
            else:
                self._show_msg("Chưa chọn", "Hãy click chọn 1 dòng.", 'warning')

        def on_skip():
            dlg.destroy()

        bf = tk.Frame(dlg, bg=C["bg"])
        bf.pack(pady=6)
        tk.Button(bf, text="Chọn", width=12, command=on_choose,
                  bg=C["green"], fg="white", relief=tk.FLAT).pack(side=tk.LEFT, padx=6)
        tk.Button(bf, text="Bo qua (ItemId0=NULL)", width=22, command=on_skip,
                  bg=C["panel"], fg=C["muted"], relief=tk.FLAT).pack(side=tk.LEFT, padx=6)

        dlg.wait_window()
        return result[0]

    def _show_batch_fuzzy_dialog(self, pending):
        """
        Dialog batch: hiển thị TẤT CẢ fuzzy cases cần xác nhận trong 1 lần.
        pending: list of {
            'val': str, 'candidates': [(score, item_dict), ...],
            'cache_key': tuple, 'section': str, 'field': str, 'row_idx': int|None
        }
        Returns: {(val, cache_key): chosen_id or None}
        """
        # Deduplicate theo (val, cache_key) — giữ context section/row đầu tiên gặp
        seen_keys = {}
        unique = []
        for item in pending:
            k = (item['val'], item.get('cache_key'))
            if k not in seen_keys:
                seen_keys[k] = True
                unique.append(item)

        if not unique:
            return {}

        SECTION_DISPLAY = {
            'HEADER': 'HEADER', 'BOM2': 'BOM II', 'BOM3': 'BOM III', 'BOM4': 'BOM IV',
        }

        dlg = ctk.CTkToplevel(self)
        dlg.withdraw()
        n = len(unique)
        dlg.title(f"Xác nhận mapping fuzzy — {n} mã cần xử lý")
        dlg.geometry("860x520")
        dlg.resizable(True, True)
        dlg.transient(self)
        dlg.grab_set()
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)

        # ── Header info ───────────────────────────────────────────────────────
        top = ctk.CTkFrame(dlg, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(14, 4))
        ctk.CTkLabel(
            top,
            text=f"Tìm thấy {n} mã không khớp chính xác trong file Excel.",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
        ).pack(anchor="w")
        ctk.CTkLabel(
            top,
            text="Chọn mã B20 tương đương cho từng dòng. Bỏ qua → giá trị sẽ là NULL.",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=("gray45", "gray65"),
        ).pack(anchor="w", pady=(2, 0))

        # ── Column headers ────────────────────────────────────────────────────
        hdr_frame = ctk.CTkFrame(dlg, fg_color=("gray88", "gray22"), corner_radius=6)
        hdr_frame.pack(fill="x", padx=12, pady=(6, 2))
        for col, (txt, w, anchor) in enumerate([
            ("Section · Trường", 200, "w"),
            ("Mã Excel (không khớp)", 190, "w"),
            ("Chọn mã B20 tương đương", 0, "w"),
        ]):
            ctk.CTkLabel(
                hdr_frame, text=txt,
                font=ctk.CTkFont("Segoe UI", 11, "bold"),
                width=w, anchor=anchor,
            ).grid(row=0, column=col, padx=(10, 4), pady=6, sticky="w")
        hdr_frame.grid_columnconfigure(2, weight=1)

        # ── Scrollable rows ───────────────────────────────────────────────────
        sf = ctk.CTkScrollableFrame(dlg, corner_radius=0)
        sf.pack(fill="both", expand=True, padx=12, pady=0)
        sf.grid_columnconfigure(2, weight=1)

        choices = {}  # (val, cache_key) → (StringVar, {label: id})

        for idx, item in enumerate(unique):
            k = (item['val'], item.get('cache_key'))
            section_raw = item.get('section', '')
            section_lbl = SECTION_DISPLAY.get(section_raw, section_raw)
            row_idx     = item.get('row_idx')
            field_name  = item.get('field', '')
            row_part    = (f" / dòng {row_idx}" if row_idx else "")
            field_part  = (f" · {field_name}" if field_name else "")
            ctx_text    = f"{section_lbl}{row_part}{field_part}"

            bg = ("gray96", "gray17") if idx % 2 == 0 else ("gray92", "gray20")
            row_f = ctk.CTkFrame(sf, fg_color=bg, corner_radius=4)
            row_f.pack(fill="x", pady=1)
            row_f.grid_columnconfigure(2, weight=1)

            ctk.CTkLabel(
                row_f, text=ctx_text, width=200, anchor="w",
                font=ctk.CTkFont("Segoe UI", 11),
                text_color=("gray40", "gray70"),
            ).grid(row=0, column=0, padx=(10, 4), pady=7, sticky="w")

            ctk.CTkLabel(
                row_f, text=item['val'], width=170, anchor="w",
                font=ctk.CTkFont("Segoe UI", 11, "bold"),
            ).grid(row=0, column=1, padx=4, pady=7, sticky="w")

            # Build dropdown options
            option_labels = []
            id_map = {}
            for score, c in item.get('candidates', []):
                lbl = f"{score:.0f}% — {c['code']} — {c.get('name', '')[:45]}"
                option_labels.append(lbl)
                id_map[lbl] = c['id']
            skip_lbl = "⊘ Bỏ qua (NULL)"
            option_labels.append(skip_lbl)
            id_map[skip_lbl] = None

            var = ctk.StringVar(value=option_labels[0])
            choices[k] = (var, id_map)

            ctk.CTkOptionMenu(
                row_f, values=option_labels, variable=var,
                font=ctk.CTkFont("Segoe UI", 11), anchor="w",
            ).grid(row=0, column=2, padx=(4, 10), pady=5, sticky="ew")

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=12)

        result = [None]

        def on_confirm():
            res = {}
            for k, (var, id_map) in choices.items():
                res[k] = id_map.get(var.get())
            result[0] = res
            dlg.destroy()

        def on_skip_all():
            result[0] = {k: None for k in choices}
            dlg.destroy()

        ctk.CTkButton(
            btn_frame, text="✓ Xác nhận và tiếp tục import",
            command=on_confirm, width=220,
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            btn_frame, text="Bỏ qua tất cả (NULL)",
            command=on_skip_all, width=180,
            fg_color="transparent", border_width=1,
            text_color=("gray50", "gray60"),
            hover_color=("gray90", "gray25"),
        ).pack(side="right")

        # Center dialog on parent window
        dlg.update_idletasks()
        pw, ph = self.winfo_width(), self.winfo_height()
        px, py = self.winfo_rootx(), self.winfo_rooty()
        mx = px + (pw - 860) // 2
        my = py + (ph - 520) // 2
        dlg.geometry(f"860x520+{mx}+{my}")
        dlg.deiconify()
        dlg.wait_window()

        return result[0] if result[0] is not None else {}

    # ── Generic lookup engine ─────────────────────────────────────────────────

    def _build_cache_generic(self, conn, bang_master, dieu_kien, truong_ss, truong_lv):
        """
        Build cache chung cho bất kỳ bảng master nào.
        truong_ss hỗ trợ nhiều trường phân cách bằng dấu phẩy: "Code,Name"
        Returns: list of dict {ss_list, lv, ht}
          ss_list = [val_field1, val_field2, ...] — thứ tự ưu tiên lookup
          lv      = giá trị lấy về (Truong_Lay_Ve)
          ht      = giá trị hiển thị (field đầu tiên)
        """
        ss_fields = [f.strip() for f in re.split(r'[|,]', truong_ss) if f.strip()]
        where     = f"WHERE {dieu_kien}" if dieu_kien else ""
        cur       = conn.cursor()
        col_sel   = ', '.join(f'[{f}]' for f in ss_fields)
        try:
            sql = f"SELECT {col_sel}, [{truong_lv}] FROM [{bang_master}] WITH(NOLOCK) {where}"
            cur.execute(sql)
            rows = cur.fetchall()
            n = len(ss_fields)
            return [
                {
                    'ss_list': [r[i] for i in range(n)],
                    'lv'     : r[n],
                    'ht'     : r[0],   # field đầu tiên dùng làm display
                }
                for r in rows
            ]
        except Exception:
            # Fallback: chỉ lấy field đầu
            sql = f"SELECT [{ss_fields[0]}], [{truong_lv}] FROM [{bang_master}] WITH(NOLOCK) {where}"
            cur.execute(sql)
            return [{'ss_list': [r[0]], 'lv': r[1], 'ht': r[0]} for r in cur.fetchall()]

    def _lookup_generic(self, value, cache, kieu_lookup, nguong_fuzzy=92, _cache_key=None):
        """
        Engine lookup dùng chung.
        cache: list of dict {ss_list, lv, ht}  — từ _build_cache_generic
          ss_list: list các giá trị so sánh theo thứ tự ưu tiên (hỗ trợ multi-field "Code,Name")
        kieu_lookup: fuzzy_code | fuzzy_name | exact | exact_code | exact_name | validate_only
        Returns: (result_value, match_type_str)
        """
        if not cache:
            return None, 'none'

        # ── Compound AND key: value là tuple (val1, val2, ...) ──────────────────
        if isinstance(value, tuple):
            def _cv(v):
                """Float nguyên (1.0) → '1', tránh '1.0' != '1'."""
                if v is None: return ''
                if isinstance(v, float):
                    return str(int(v)) if v == int(v) else str(v).strip()
                return str(v).strip()
            val_parts  = [_cv(v) for v in value]
            norm_parts = [self._norm_for_match(v) for v in val_parts]
            n = len(val_parts)
            def _and_exact(e):
                ss = [_cv(v) for v in e.get('ss_list', [])]
                return len(ss) >= n and all(ss[i] == val_parts[i] for i in range(n))
            def _and_norm(e):
                ss = [_cv(v) for v in e.get('ss_list', [])]
                return len(ss) >= n and all(self._norm_for_match(ss[i]) == norm_parts[i] for i in range(n))
            for e in cache:
                if _and_exact(e): return e['lv'], 'exact'
            for e in cache:
                if _and_norm(e):  return e['lv'], 'normalize'
            return None, 'none'   # compound key không dùng fuzzy
        # ─────────────────────────────────────────────────────────────────────────

        if not value:
            return None, 'none'

        val_str  = str(value).strip()
        norm_val = self._norm_for_match(val_str)

        def _ss_vals(e):
            """Trả về list các giá trị so sánh của entry (tương thích cả ss_list lẫn ss cũ)."""
            if 'ss_list' in e:
                return [str(v or '').strip() for v in e['ss_list']]
            return [str(e.get('ss') or '').strip()]

        # ── Tier 1: exact — so sánh từng field theo thứ tự ──────────────────
        for e in cache:
            for sv in _ss_vals(e):
                if sv == val_str:
                    return e['lv'], 'exact'

        # ── Tier 2: normalize — tương tự, theo thứ tự ───────────────────────
        for e in cache:
            for sv in _ss_vals(e):
                if self._norm_for_match(sv) == norm_val:
                    return e['lv'], 'normalize'

        # validate_only: check tồn tại → giữ value gốc nếu không tìm thấy
        if kieu_lookup == 'validate_only':
            return val_str, 'not_found'

        # exact (không fuzzy): dừng ở đây
        if kieu_lookup in ('exact', 'exact_code', 'exact_name'):
            return None, 'none'

        # ── Tier 3: fuzzy — score max trên tất cả ss fields ─────────────────
        try:
            from rapidfuzz import fuzz as _fuzz
        except ImportError:
            return None, 'none'

        def _score_entry(e):
            """Tính fuzzy score tốt nhất trên toàn bộ ss_list."""
            best = 0.0
            for sv in _ss_vals(e):
                cv_norm = self._norm_for_match(sv)
                if not cv_norm:
                    continue
                _lp = (min(len(norm_val), len(cv_norm))
                       / max(len(norm_val), len(cv_norm), 1))
                if kieu_lookup == 'fuzzy_name':
                    s = max(_fuzz.partial_ratio(norm_val, cv_norm) * _lp,
                            _fuzz.ratio(norm_val, cv_norm))
                else:   # fuzzy_code (default)
                    s = max(_fuzz.ratio(norm_val, cv_norm),
                            _fuzz.partial_ratio(norm_val, cv_norm) * _lp)
                if s > best:
                    best = s
            return best

        if kieu_lookup == 'fuzzy_name':
            # Tier 2b: contains check trên tất cả fields
            for e in cache:
                for sv in _ss_vals(e):
                    cv_norm = self._norm_for_match(sv)
                    if cv_norm and (norm_val in cv_norm or cv_norm in norm_val):
                        return e['lv'], 'contains'

        _min_score = 20 if kieu_lookup == 'fuzzy_name' else 40
        scored = [(s, e) for e in cache
                  if (s := _score_entry(e)) >= _min_score]

        if not scored:
            return None, 'none'

        scored.sort(key=lambda x: -x[0])
        top_score, top_entry = scored[0]

        if top_score >= nguong_fuzzy:
            return top_entry['lv'], 'fuzzy_auto'

        # Popup cho user chọn (top 3) — hiển thị field đầu tiên làm "code"
        candidates = [
            (s, {
                'id'  : e['lv'],
                'code': _ss_vals(e)[0],
                'name': str(e.get('ht') or _ss_vals(e)[0] or ''),
            })
            for s, e in scored[:3]
        ]

        # ── Batch mode: collect hoặc dùng kết quả đã resolve sẵn ─────────────
        res_key = (val_str, _cache_key)
        if getattr(self, '_fuzzy_collect_mode', False):
            entry = dict(getattr(self, '_fuzzy_ctx', {}))
            entry.update({'val': val_str, 'candidates': candidates, 'cache_key': _cache_key})
            if not hasattr(self, '_fuzzy_pending'):
                self._fuzzy_pending = []
            self._fuzzy_pending.append(entry)
            return None, 'fuzzy_pending'
        if hasattr(self, '_fuzzy_resolutions') and res_key in self._fuzzy_resolutions:
            chosen = self._fuzzy_resolutions[res_key]
            return chosen, ('fuzzy_user' if chosen else 'none')
        # Nếu user đã xác nhận batch dialog → suppress popup cũ, trả về NULL
        if getattr(self, '_fuzzy_batch_done', False):
            return None, 'none'
        # ─────────────────────────────────────────────────────────────────────

        chosen = self._show_suggest_dialog(val_str, candidates)
        return chosen, ('fuzzy_user' if chosen else 'none')

    # ── Mapping-driven cache builder ──────────────────────────────────────────
    def _build_all_caches(self, conn, mapping_records):
        """
        Scan mapping → pre-build cache cho mọi trường có lookup.
        Cache key = (bang_master, dieu_kien, truong_so_sanh, truong_lay_ve)
        """
        caches = {}
        seen   = set()
        skip_kl = {'', 'sp_rowid', 'sp_version', 'sp_code', 'lookup'}
        for rec in mapping_records:
            kl = rec.get('kieu_lookup', '')
            bm = rec.get('bang_master', '')
            dk = rec.get('dieu_kien_master', '')
            ss = rec.get('truong_so_sanh', '')
            lv = rec.get('truong_lay_ve',  '')
            if kl in skip_kl or not bm or not ss or not lv:
                continue
            cache_key = (bm, dk, ss, lv)
            if cache_key in seen:
                continue
            seen.add(cache_key)
            try:
                caches[cache_key] = self._build_cache_generic(conn, bm, dk, ss, lv)
                self._log('cache', f'Build cache {bm}', len(caches[cache_key]),
                          'OK', f'{ss}->{lv} WHERE {dk}', 'ok')
            except Exception as e:
                self._log('cache', f'Build cache {bm}', 0, 'Warn', str(e), 'warn')
                caches[cache_key] = []
        return caches

    def _resolve_header_field(self, rec, conn, meta, norm_meta, caches, now, row_out=None):
        """
        Resolve 1 trường từ mapping record (dùng cho HEADER).
        Trả về (value, match_type_str)
        """
        nguon    = rec['nguon_dl']
        mac_dinh = rec['mac_dinh']
        kl       = rec.get('kieu_lookup', '')
        bm       = rec.get('bang_master', '')
        dk       = rec.get('dieu_kien_master', '')
        ten_excel = rec.get('ten_excel', '')

        # ── CoDinh: đọc Mac_dinh, tự-cast kiểu ─────────────────────────────
        if nguon == 'CoDinh':
            if mac_dinh.upper() == 'NULL':
                return None, 'codinh'
            kieu_dl_cd = rec.get('kieu_dl', '').lower()
            if mac_dinh in ('', 'EMPTY'):
                # date/datetime không thể là empty string → NULL
                if kieu_dl_cd in ('date', 'datetime'):
                    return None, 'codinh'
                return '', 'codinh'
            try:    return int(mac_dinh),   'codinh'
            except: pass
            try:    return float(mac_dinh), 'codinh'
            except: pass
            return mac_dinh, 'codinh'

        # ── HeThong: system-generated ────────────────────────────────────────
        elif nguon == 'HeThong':
            if mac_dinh == 'NOW':
                return now, 'hethong'
            return None, 'hethong'   # ModifiedBy → Bravo fill khi import thật

        # ── UILookup: giá trị lấy từ combo trên giao diện ───────────────────
        elif nguon == 'UILookup':
            if mac_dinh == 'creator':
                return self._current_creator_code, 'ui_lookup'
            elif mac_dinh == 'product_id':
                return self._thdm_selected_product_id, 'ui_lookup'
            elif mac_dinh == 'order_id':
                return self._thdm_selected_order_id, 'ui_lookup'
            elif mac_dinh == 'period_id':
                return self._thdm_selected_period_id, 'ui_lookup'
            return None, 'ui_lookup'

        # ── SP / TinhToan: resolve trong pass 2 ──────────────────────────
        elif nguon in ('SP', 'TinhToan'):
            return None, nguon.lower()

        # ── Excel: đọc meta, lookup master nếu cần ───────────────────────────
        else:
            raw = None
            if ten_excel:
                if '|' in ten_excel:
                    # Multi-field: mỗi phần tách bởi | → build tuple để AND lookup
                    # @FieldName → lấy từ row_out (field đã resolve trước đó trong cùng header)
                    _parts, _any = [], False
                    for _te in ten_excel.split('|'):
                        _te = _te.strip()
                        if _te.startswith('@'):
                            _fn = _te[1:]
                            _v = (row_out or {}).get(_fn)
                        else:
                            _v = meta.get(_te)
                            if _v is None:
                                _v = norm_meta.get(_norm_vn(_te))
                        _parts.append(str(_v or '').strip())
                        if _v is not None:
                            _any = True
                    raw = tuple(_parts) if _any else None
                else:
                    # Single field: giữ nguyên logic cũ
                    # Dùng explicit None check thay vì "or" để không bỏ sót giá trị 0
                    _v = meta.get(ten_excel)
                    if _v is None:
                        _v = norm_meta.get(_norm_vn(ten_excel))
                    raw = _v
            if raw is None:
                if mac_dinh == 'EMPTY':
                    raw = ''
                elif mac_dinh not in ('', 'NULL'):
                    raw = mac_dinh

            kieu_dl = rec.get('kieu_dl', '').lower()

            if not kl:
                # ── Kieu_DL transform chỉ dùng cho direct value (không lookup) ──
                if raw is not None:
                    if kieu_dl in ('date', 'datetime'):
                        if isinstance(raw, datetime.datetime):
                            raw = raw.date()
                        elif isinstance(raw, datetime.date):
                            pass   # already date
                        elif isinstance(raw, str):
                            s = raw.strip()
                            if not s or s.lower() in ('null', 'none', 'nan'):
                                # Empty/NULL string cho date field → NULL
                                raw = None
                            else:
                                parsed = None
                                for fmt in _DATE_FMTS:
                                    try:
                                        parsed = datetime.datetime.strptime(s, fmt).date()
                                        break
                                    except (ValueError, TypeError):
                                        pass
                                # Nếu parse thất bại → NULL thay vì để lại string
                                # (tránh SQL Server nhận '' → 1900-01-01)
                                raw = parsed
                        elif isinstance(raw, (int, float)):
                            import math as _math2
                            if _math2.isnan(raw) if isinstance(raw, float) else False:
                                raw = None
                    elif kieu_dl in ('number', 'decimal', 'float', 'numeric'):
                        qty_str = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        try:    raw = float(qty_str) if qty_str else None
                        except: raw = None
                    elif kieu_dl == 'int':
                        try:
                            qty_str = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                            raw = int(float(qty_str)) if qty_str else None
                        except: raw = None
                return raw, 'excel_direct'

            # ── Generic lookup engine ─────────────────────────────────────
            ss = rec.get('truong_so_sanh', '')
            lv = rec.get('truong_lay_ve',  '')
            if ss and lv:
                cache_key = (bm, dk, ss, lv)
                cache = caches.get(cache_key, [])
                nguong = int(rec.get('nguong_fuzzy', 0) or 0) or 92
                self._fuzzy_ctx = {
                    'section': 'HEADER',
                    'field': rec.get('sql_col', ''),
                    'row_idx': None,
                }
                return self._lookup_generic(raw, cache, kl, nguong, _cache_key=cache_key)

            return raw, 'excel_direct'

    # ── BOM Detail helpers (BOM2 / BOM3 / BOM4) ──────────────────────────────

    def _build_bom_detail_caches(self, conn, mapping_recs):
        """Build lookup caches cho một section BOM detail (tương tự _build_all_caches)."""
        caches = {}
        for rec in mapping_recs:
            bm  = _nan_str(rec.get('bang_master', ''))
            dk  = _nan_str(rec.get('dieu_kien_master', ''))
            ss  = _nan_str(rec.get('truong_so_sanh', ''))
            lv  = _nan_str(rec.get('truong_lay_ve', ''))
            kl  = _nan_str(rec.get('kieu_lookup', ''))
            if not (bm and ss and lv and kl):
                continue
            key = (bm, dk, ss, lv)
            if key not in caches:
                try:
                    caches[key] = self._build_cache_generic(conn, bm, dk, ss, lv)
                except Exception as e:
                    self._log('', f'detail_cache({bm})', 0, 'Warn', str(e), 'warn')
        return caches

    def _resolve_detail_row(self, mapping_recs, df_row, parent_row, conn,
                            detail_caches, now, builtin_order, bom_detail_type,
                            sp_cfgs=None, bom_section=''):
        """
        Resolve một hàng DataFrame (BOM2/3/4) → dict {sql_col: value}.
        - mapping_recs : list record từ mapping sheet BOM2/3/4
        - df_row       : pandas Series (1 hàng dữ liệu)
        - parent_row   : dict B20BOM header đã resolve (để copy BranchCode, EffectiveDate, ...)
        - builtin_order: số thứ tự hàng (1, 2, 3...)
        - bom_detail_type: 2 / 3 / 4
        """
        import datetime as _dt, math as _math

        # Các field copy từ B20BOM parent
        PARENT_FIELDS = {
            'BranchCode', 'EffectiveDate', 'FinishedDate',
            'ItemId0', 'DetailRowId_SO', 'ParentBizDocId', 'ProductionProcessId',
        }

        # ── Pass 1a: CoDinh / HeThong / UILookup / SP-defer ─────────────
        # Dùng shared _resolve_row_mapping() — đồng nhất với THDM pipeline.
        # BOMDetailType inject vào parent_row để HeThong tự pick đúng.
        _bom_parent = dict(parent_row)
        _bom_parent['BOMDetailType'] = bom_detail_type
        _ctx = {
            'now':           now,
            'builtin_order': builtin_order,
            'doc_id':        '__BOMID__',   # placeholder thay thật khi generate SQL
            'parent_row':    _bom_parent,
            'parent_fields': PARENT_FIELDS | {'BOMDetailType'},
            'ui_values': {
                'creator':    self._current_creator_code,
            },
            'skip_nguon': {'SP', 'TinhToan'},
        }
        _non_excel = [r for r in mapping_recs
                      if (r.get('nguon_dl') or 'Excel') != 'Excel']
        row_out = _resolve_row_mapping(_non_excel, _ctx, get_excel_val=None)

        # ── Pass 1b: Excel fields (df_row lookup + master lookup) ─────────
        for rec in mapping_recs:
            sql_col = rec.get('sql_col', '')
            nguon   = rec.get('nguon_dl', 'Excel')
            mac     = _nan_str(rec.get('mac_dinh', ''))
            ten     = _nan_str(rec.get('ten_excel', ''))
            kl      = _nan_str(rec.get('kieu_lookup', ''))
            kd      = _nan_str(rec.get('kieu_dl', ''))
            bm      = _nan_str(rec.get('bang_master', ''))
            dk      = _nan_str(rec.get('dieu_kien_master', ''))
            ss      = _nan_str(rec.get('truong_so_sanh', ''))
            lv      = _nan_str(rec.get('truong_lay_ve', ''))

            if nguon != 'Excel':
                continue   # đã xử lý bởi _resolve_row_mapping() ở trên

            # ── Excel (df_row lookup + master lookup) ─────────────────────
            raw = None
            if ten:
                if '|' in ten:
                    # Multi-field compound key: "Col1|Col2" hoặc "Col1|@ParentField"
                    # @FieldName → lấy từ parent_row (header đã resolve)
                    _parts, _any = [], False
                    for _t in ten.split('|'):
                        _t = _t.strip()
                        if _t.startswith('@'):
                            _pv = parent_row.get(_t[1:])
                            _parts.append(str(_pv or '').strip())
                            if _pv is not None: _any = True
                        else:
                            _pv, _nt = None, _norm_vn(_t)
                            for _c in df_row.index:
                                _cs = str(_c).strip()
                                if _cs == _t or _norm_vn(_cs) == _nt:
                                    _v2 = df_row[_c]
                                    if _v2 is not None and not (isinstance(_v2, float) and _math.isnan(_v2)):
                                        _pv = _v2
                                    break
                            _parts.append(str(_pv or '').strip())
                            if _pv is not None: _any = True
                    raw = tuple(_parts) if _any else None
                else:
                    _matched_col = None
                    _norm_ten = _norm_vn(ten)
                    # Pass 1: exact hoặc norm-exact match
                    for col in df_row.index:
                        col_s = str(col).strip()
                        if col_s == ten or _norm_vn(col_s) == _norm_ten:
                            v = df_row[col]
                            _matched_col = col
                            if v is None or (isinstance(v, float) and _math.isnan(v)):
                                raw = None
                            else:
                                raw = v
                            break
                    # Pass 2: suffix match — xử lý cột bị merge header cha
                    # Ví dụ: "SLg_Tên_Vật_Tư" (norm: slgtenvattu) match "Tên vật tư" (norm: tenvattu)
                    if _matched_col is None and _norm_ten:
                        for col in df_row.index:
                            col_s = str(col).strip()
                            norm_col = _norm_vn(col_s)
                            if norm_col.endswith(_norm_ten) and len(norm_col) > len(_norm_ten):
                                v = df_row[col]
                                _matched_col = col
                                if v is None or (isinstance(v, float) and _math.isnan(v)):
                                    raw = None
                                else:
                                    raw = v
                                break

            if raw is None:
                if mac == 'EMPTY':
                    raw = ''
                elif mac and mac not in ('', 'NULL'):
                    try:    raw = int(mac)
                    except:
                        try:    raw = float(mac)
                        except: raw = mac
                else:
                    raw = None

            # Type coercion (chỉ khi không lookup)
            if raw is not None and not kl:
                if kd == 'date':
                    if isinstance(raw, _dt.datetime):
                        raw = raw.date()
                    elif isinstance(raw, str):
                        for fmt in _DATE_FMTS:
                            try: raw = _dt.datetime.strptime(raw, fmt).date(); break
                            except: pass
                elif kd in ('number', 'decimal', 'float', 'numeric'):
                    try:
                        qs = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        raw = float(qs) if qs else None
                    except: raw = None
                elif kd == 'int':
                    try:
                        qs = re.sub(r'[^\d,\.]', '', str(raw)).replace(',', '.')
                        raw = int(float(qs)) if qs else None
                    except: raw = None

            # Lookup master nếu có (cùng engine với HEADER)
            if kl and ss and lv:
                cache_key = (bm, dk, ss, lv)
                cache = detail_caches.get(cache_key, [])
                nguong = int(rec.get('nguong_fuzzy', 0) or 0) or 92
                self._fuzzy_ctx = {
                    'section': bom_section,
                    'field': sql_col,
                    'row_idx': builtin_order,
                }
                raw, _ = self._lookup_generic(raw, cache, kl, nguong, _cache_key=cache_key)

            row_out[sql_col] = raw

        # ── Pass 2: SP fields (row_out đã có đủ context để truyền vào params) ──
        if sp_cfgs:
            # Xác định field nào là "secondary output" của multi-output SP
            _claimed = set()
            for rec2 in mapping_recs:
                if rec2.get('nguon_dl') != 'SP':
                    continue
                s_col = rec2.get('sql_col', '')
                cfg2  = sp_cfgs.get(s_col)
                if cfg2:
                    out_fs = [f.strip() for f in cfg2.get('outputfields', '').split(',') if f.strip()]
                    if len(out_fs) > 1:
                        _claimed.update(out_fs[1:])   # primary = out_fs[0], rest = secondary

            _done_multi = set()
            for rec2 in mapping_recs:
                if rec2.get('nguon_dl') != 'SP':
                    continue
                s_col = rec2.get('sql_col', '')
                if s_col in _claimed or s_col in _done_multi:
                    continue
                cfg2 = sp_cfgs.get(s_col)
                if not cfg2:
                    continue
                try:
                    sp_name2  = cfg2.get('sp_name', '').strip()
                    params_s2 = cfg2.get('params', '').strip()
                    fallback2 = cfg2.get('fallback', '').strip() or None
                    out_fs2   = [f.strip() for f in cfg2.get('outputfields', '').split(',') if f.strip()]
                    if sp_name2.lower() == 'lookup':
                        r2 = self._sp_lookup(conn, params_s2, row_out)
                        row_out[s_col] = r2 if r2 is not None else fallback2
                    elif out_fs2:
                        # Multi-output SP: build params → execute → distribute kết quả
                        pc_h, pv_h = [], []
                        for ph in params_s2.split('|'):
                            ph = ph.strip()
                            if '=' not in ph: continue
                            kh, vh = ph.split('=', 1)
                            kh = kh.strip().lstrip('@'); vh = vh.strip()
                            if vh.startswith('{') and vh.endswith('}'):
                                vh = row_out.get(vh[1:-1])
                            if vh is None:
                                pc_h.append(f'@{kh}=NULL')
                            else:
                                pc_h.append(f'@{kh}=?')
                                pv_h.append(vh)
                        cur2 = conn.cursor()
                        cur2.execute(f'EXEC {sp_name2} ' + ', '.join(pc_h), pv_h)
                        sp_row2 = cur2.fetchone()
                        if sp_row2 and cur2.description:
                            sp_res2 = dict(zip([d[0] for d in cur2.description], sp_row2))
                            for f2 in out_fs2:
                                row_out[f2] = sp_res2.get(f2)
                        _done_multi.add(s_col)
                    else:
                        r2 = self._call_sp(conn, sp_name2, params_s2, row_out)
                        row_out[s_col] = r2 if r2 is not None else fallback2
                except Exception as e2:
                    self._log('', f'SP_detail {s_col}', 0, 'Warn', str(e2), 'warn')

        return row_out

    def _generate_bom_details(self, conn, bom_id, parent_row, tables, mapping, now, export_only):
        """
        Tạo INSERT INTO B20BOMDetail cho BOM2, BOM3, BOM4.
        3 pha:
          1. Resolve tất cả rows của mỗi section (per-row BeforeInsert hooks)
          2. BeforeInsertBatch hooks (gọi SP 1 lần / nhóm SP)
          3. INSERT / export SQL tất cả rows đã resolve
        - export_only=True  → trả về list[str] câu SQL
        - export_only=False → execute trực tiếp, trả về số hàng insert
        """
        import math as _math

        # Derive từ _CONFIG + DETAIL mapping (BOMDetailType CoDinh)
        SECTION_TO_TYPE = {}
        for _sec, _cfg in mapping.get('_CONFIG', {}).items():
            if 'BOMDetail' not in _cfg.get('view_insert', ''):
                continue
            for _r in mapping.get(_sec, []):
                if _r.get('sql_col') == 'BOMDetailType':
                    try:
                        SECTION_TO_TYPE[_sec] = int(float(_r.get('mac_dinh', 0)))
                    except (ValueError, TypeError):
                        pass
                    break
        DB_TABLE        = 'B20BOMDetail'
        NVARCHAR_TYPES  = {'nvarchar', 'nchar', 'ntext'}
        sql_lines       = []
        total_rows      = 0


        def _progress(msg):
            self.after(0, lambda m=msg: self._update_loading_msg(
                "Đang import dữ liệu vào BRAVO...\n" + m))

        # ── Pha 1: resolve tất cả sections ───────────────────────────────────
        # section_data[section] = {
        #   'resolved'        : [row_vals, ...],
        #   'col_kieu'        : {sql_col: kieu_dl},
        #   'bom_detail_type' : int,
        # }
        section_data = {}
        _progress("Phase 1: Resolve rows BOM detail...")

        for label, tbl in tables.items():
            section = tbl.get('type')
            if section not in SECTION_TO_TYPE:
                continue
            df = tbl.get('df')
            if df is None or df.empty or 'Lỗi' in df.columns:
                continue

            bom_detail_type = SECTION_TO_TYPE[section]
            detail_recs     = mapping.get(section, [])
            if not detail_recs:
                continue

            _n_df_rows = len(df) if df is not None else 0
            _progress(f"Phase 1: {section} — {_n_df_rows} dòng Excel, build cache...")
            detail_caches = (getattr(self, '_ps_bom_caches', {}).get(section)
                           or self._build_bom_detail_caches(conn, detail_recs))

            sp_cfgs_section = {
                s['sql_column']: s
                for s in mapping.get('SP_CONFIG', [])
                if s.get('isactive', '') == '1'
                and s.get('section', '') in ('', section)
            }

            # Per-row BeforeInsert hooks
            sp_hooks_before = [
                h for h in mapping.get('SP_HOOK', [])
                if h.get('section', '') == section
                and h.get('event', '').lower() == 'beforeinsert'
            ]

            col_kieu = {r['sql_col']: r.get('kieu_dl', '').lower() for r in detail_recs}

            stt_col  = next((c for c in df.columns if _norm_vn(str(c)) == 'stt'), None)
            ff_fields = [
                (r['ten_excel'], r['sql_col'])
                for r in detail_recs
                if r.get('fill_forward') == '1' and r.get('ten_excel')
            ]
            current_ff   = {}
            builtin_order = 0
            resolved_rows = []


            for _, df_row in df.iterrows():
                if df_row.isna().all():
                    continue

                def _get_stt(raw):
                    """str(0 or '') = '' — xử lý integer 0 đúng."""
                    if raw in (0, 0.0): return "0"
                    if raw is None or raw == '': return ''
                    if isinstance(raw, float) and _math.isnan(raw): return ''
                    return str(raw).strip()
                if stt_col:
                    stt_v = _get_stt(df_row.get(stt_col))
                else:
                    # Fallback: tìm cột tên 'STT' trực tiếp trong DataFrame
                    stt_v = ''
                    for _col in df_row.index:
                        if str(_col).upper() == 'STT':
                            stt_v = _get_stt(df_row.get(_col))
                            break
                is_section_header = bool(stt_v and SECTION_STT_PATTERN.match(stt_v))

                # Safety: dòng có STT là chữ cái (vd 'I') nhưng có kích thước vật lý
                # → là sub-group GIAO RỜI (MODULE MD1...), không phải section header thật
                if is_section_header:
                    _hd_gen = any(
                        v not in (None, '', 0, '0', 0.0)
                        and not (isinstance(v, float) and _math.isnan(v))
                        for col, v in df_row.items()
                        if re.search(r'(dai|day|rong|width|length)', _norm_vn(str(col)))
                    )
                    if _hd_gen:
                        is_section_header = False  # có dims → data row (GIAO RỜI/sub-group), không phải section header

                if is_section_header:
                    for (ten_excel, sql_col_ff) in ff_fields:
                        raw = df_row.get(ten_excel)
                        if raw is not None and not (isinstance(raw, float) and _math.isnan(raw)) \
                                and str(raw).strip() not in ('', 'nan'):
                            current_ff[sql_col_ff] = str(raw).strip()
                    continue

                builtin_order += 1
                row_vals = self._resolve_detail_row(
                    detail_recs, df_row, parent_row, conn,
                    detail_caches, now, builtin_order, bom_detail_type,
                    sp_cfgs=sp_cfgs_section, bom_section=section
                )

                # Fill-Forward
                for (ten_excel, sql_col_ff) in ff_fields:
                    if sql_col_ff in current_ff:
                        row_vals[sql_col_ff] = current_ff[sql_col_ff]

                # Per-row SP_HOOK BeforeInsert — dùng shared runner
                _run_row_sp_hooks(
                    conn, sp_hooks_before, row_vals,
                    log_fn=lambda msg: self._log('', msg, 0, 'Warn', msg, 'warn'),
                )

                resolved_rows.append(row_vals)

            _progress(f"Phase 1: {section} — resolved {len(resolved_rows)} data rows")
            section_data[section] = {
                'resolved'       : resolved_rows,
                'col_kieu'       : col_kieu,
                'bom_detail_type': bom_detail_type,
            }

        # ── Pha 2: BeforeInsertBatch hooks (gọi SP theo nhóm SP_Name) ─────────
        if not export_only and section_data:
            batch_hooks = [
                h for h in mapping.get('SP_HOOK', [])
                if h.get('event', '').lower() == 'beforeinsertbatch'
                and h.get('section', '') in section_data
                and h.get('isactive', '') == '1'
            ]
            # Gom theo sp_name để gọi mỗi SP 1 lần duy nhất
            by_sp = {}
            for h in batch_hooks:
                sp = h['sp_name']
                by_sp.setdefault(sp, []).append(h)
            for sp_name, hooks in by_sp.items():
                _n_rows = sum(
                    len(section_data.get(h.get('section',''), {}).get('resolved', []))
                    for h in hooks
                )
                _progress(f"Phase 2: SP_HOOK BeforeInsertBatch — {sp_name} ({_n_rows} rows)...")
                self._run_batch_hook_grouped(conn, sp_name, hooks, section_data, parent_row)
                _progress(f"Phase 2: SP_HOOK xong — {sp_name}")

        # ── Pha 3: INSERT / export SQL ────────────────────────────────────────
        _total_all = sum(len(d['resolved']) for d in section_data.values())
        _progress(f"Phase 3: INSERT vào {DB_TABLE} (0/{_total_all:,} dòng)...")
        for section, data in section_data.items():
            resolved_rows   = data['resolved']
            col_kieu        = data['col_kieu']
            bom_detail_type = data['bom_detail_type']

            if export_only:
                sql_lines.append(f'\n-- {section} (BOMDetailType={bom_detail_type})\n')

            def _sv(v, col=None):
                """Render giá trị → SQL literal (dùng trong export)."""
                import datetime as _dt2
                if v is None:
                    return 'NULL'
                if v == '__BOMID__':
                    return str(bom_id) if isinstance(bom_id, int) else bom_id
                if isinstance(v, bool):   return '1' if v else '0'
                _col_type   = col_kieu.get(col, '').lower() if col else ''
                _is_str_col = _col_type in ('varchar','nvarchar','char','nchar','text','ntext')
                if isinstance(v, float):
                    if _math.isnan(v):   return 'NULL'
                    if _is_str_col:
                        s  = str(int(v)) if v == int(v) else str(v)
                        px = 'N' if _col_type in NVARCHAR_TYPES else ''
                        return px + "'" + s + "'"
                    return str(v)
                if isinstance(v, int):
                    if _is_str_col:
                        px = 'N' if _col_type in NVARCHAR_TYPES else ''
                        return px + "'" + str(v) + "'"
                    return str(v)
                if isinstance(v, _dt2.datetime):
                    return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'"
                if isinstance(v, _dt2.date):
                    return "'" + v.strftime('%Y-%m-%d') + "'"
                s = str(v).strip()
                if s.lower() in ('nan', 'none', ''):  return 'NULL'
                px = 'N' if col_kieu.get(col, '') in NVARCHAR_TYPES else ''
                return px + "'" + s.replace("'", "''") + "'"

            for row_vals in resolved_rows:
                cols = list(row_vals.keys())
                if export_only:
                    col_list = ', '.join(f'[{c}]' for c in cols)
                    val_list = ', '.join(_sv(row_vals[c], c) for c in cols)
                    sql_lines.append(
                        f'INSERT INTO {DB_TABLE} ({col_list}) VALUES ({val_list});\n'
                    )
                else:
                    exec_vals = [
                        bom_id if row_vals[c] == '__BOMID__' else row_vals[c]
                        for c in cols
                    ]
                    sql_exec = (
                        f"INSERT INTO {DB_TABLE} ("
                        + ', '.join(f'[{c}]' for c in cols)
                        + ') VALUES (' + ', '.join(['?'] * len(cols)) + ')'
                    )
                    cur = conn.cursor()
                    cur.execute(sql_exec, exec_vals)
                total_rows += 1
                if not export_only and total_rows % 100 == 0:
                    _progress(f"Phase 3: INSERT vào {DB_TABLE} "
                              f"({total_rows:,}/{_total_all:,} dòng)...")

        _progress(f"Phase 3: INSERT xong — {total_rows} dòng chi tiết")
        return sql_lines if export_only else total_rows

    def _run_batch_hook_grouped(self, conn, sp_name, hooks, section_data, parent_row):
        """
        Gọi SP 1 lần với nhiều XML params (mỗi hook = 1 section = 1 XML param).
        SP trả về N result sets theo thứ tự hooks; mỗi result set chứa Id→fields
        để map ngược lại vào resolved_rows (in-place).

        Config SP_HOOK:
          XmlParam  : tên param XML của SP   (vd: @_B20BOMDetail)
          XmlTag    : tên thẻ XML mỗi dòng   (vd: Detail2)
          XmlFields : fields đưa vào XML      (vd: Id,ItemType,ItemName,ItemId,Name,Unit)
          Condition : EMPTY(field) hoặc rỗng
          Params    : scalar params từ parent  (vd: @_ItemId0={ItemId0}|@_BranchCode={BranchCode})
          OutputFields: fields lấy từ SP kết quả (vd: ItemId)
        """
        EMPTY_XML = '<NewDataSet />'

        # Build per-hook: batch_idx, xml_string, output_fields
        hook_meta = []       # [{batch_idx, h_outs}, ...]
        xml_args  = {}       # {xml_param: xml_string}
        pc_scalar, pv_scalar = [], []   # scalar params (từ hook đầu tiên)

        for hook in hooks:
            section    = hook.get('section', '')
            cond       = hook.get('condition', '').strip()
            xml_param  = hook.get('xmlparam', '').strip()
            xml_tag    = hook.get('xmltag', '').strip()
            xml_fields = [f.strip() for f in hook.get('xmlfields', '').split(',') if f.strip()]
            h_outs     = [f.strip() for f in hook.get('outputfields', '').split(',') if f.strip()]

            if not (xml_param and xml_tag and xml_fields):
                hook_meta.append({'batch_idx': {}, 'h_outs': h_outs})
                xml_args[xml_param or f'_empty_{len(xml_args)}'] = EMPTY_XML
                continue

            rows = section_data.get(section, {}).get('resolved', [])

            # Lọc rows theo condition
            batch_idx = {}
            for idx, row_vals in enumerate(rows):
                should_run = True
                if cond.startswith('EMPTY(') and cond.endswith(')'):
                    cond_field = cond[6:-1].strip()
                    should_run = not row_vals.get(cond_field)
                elif cond.startswith('NOTEMPTY(') and cond.endswith(')'):
                    cond_field = cond[9:-1].strip()
                    should_run = bool(row_vals.get(cond_field))
                else:
                    cond_field = ''
                if should_run:
                    batch_idx[idx] = row_vals

            hook_meta.append({'batch_idx': batch_idx, 'h_outs': h_outs})

            # Build XML (Id = idx+1 để map ngược)
            def _xml_escape(v):
                if v is None: return ''
                return (str(v)
                        .replace('&', '&amp;')
                        .replace('<', '&lt;')
                        .replace('>', '&gt;')
                        .replace('"', '&quot;'))

            if batch_idx:
                # SP dùng OPENXML flag=1 (attribute-centric)
                # → phải dùng attribute XML: <Detail2 Id="1" ItemType="A" ... />
                parts = ['<NewDataSet>']
                for idx, row_vals in batch_idx.items():
                    attrs = [f'Id="{idx + 1}"']
                    for field in xml_fields:
                        if field == 'Id':
                            continue
                        attrs.append(f'{field}="{_xml_escape(row_vals.get(field))}"')
                    parts.append(f'  <{xml_tag} {" ".join(attrs)} />')
                parts.append('</NewDataSet>')
                xml_args[xml_param] = '\n'.join(parts)
            else:
                xml_args[xml_param] = EMPTY_XML

            # Scalar params: chỉ build 1 lần từ hook đầu tiên có params
            if not pc_scalar:
                for ph in hook.get('params', '').split('|'):
                    ph = ph.strip()
                    if '=' not in ph:
                        continue
                    k, v = ph.split('=', 1)
                    k = k.strip().lstrip('@'); v = v.strip()
                    if v.startswith('{') and v.endswith('}'):
                        v = parent_row.get(v[1:-1])
                    if v is None:
                        pc_scalar.append(f'@{k}=NULL')
                    else:
                        pc_scalar.append(f'@{k}=?')
                        pv_scalar.append(v)

        if not xml_args:
            return

        # Build final EXEC call: tất cả params (kể cả XML) dùng ? placeholder
        # SP khai báo @_B20BOMDetail XML nên SQL Server tự convert NVARCHAR→XML
        pc_final = pc_scalar.copy()
        pv_final = pv_scalar.copy()
        for xml_param, xml_str in xml_args.items():
            pc_final.append(f'{xml_param}=?')
            pv_final.append(xml_str)

        exec_sql = f'EXEC {sp_name} ' + ', '.join(pc_final)

        try:
            cur = conn.cursor()
            # Giới hạn lock wait 120s — tránh treo vô tận nếu SP bị block
            cur.execute("SET LOCK_TIMEOUT 120000")
            cur.execute(exec_sql, pv_final)

            # Đọc result sets theo thứ tự hooks
            rs_idx   = 0
            mapped   = 0   # tổng số fields được map ngược lại
            while True:
                if rs_idx < len(hook_meta) and cur.description:
                    meta      = hook_meta[rs_idx]
                    batch_idx = meta['batch_idx']
                    h_outs    = meta['h_outs']
                    desc      = [d[0] for d in cur.description]
                    rs_rows   = cur.fetchall()
                    for sp_row in rs_rows:
                        sp_res = dict(zip(desc, sp_row))
                        sp_id  = sp_res.get('Id')
                        if sp_id is not None:
                            orig_idx = int(sp_id) - 1
                            if orig_idx in batch_idx:
                                for fout in h_outs:
                                    if fout in sp_res and sp_res[fout] is not None:
                                        batch_idx[orig_idx][fout] = sp_res[fout]
                                        mapped += 1

                if not cur.nextset():
                    break
                rs_idx += 1

        except Exception as e:
            err_msg = str(e)
            # Phân biệt lỗi permission để thông báo rõ hơn
            is_permission = (
                'EXECUTE permission' in err_msg
                or 'permission was denied' in err_msg.lower()
                or ('42000' in err_msg and 'permission' in err_msg.lower())
            )
            is_truncated = 'truncated' in err_msg.lower() or '2628' in err_msg or '8152' in err_msg
            if is_permission:
                friendly = (
                    f"Không có quyền EXECUTE stored procedure '{sp_name}'.\n"
                    f"Liên hệ DBA để cấp quyền:\n"
                    f"  GRANT EXECUTE ON {sp_name} TO [<login>]\n\n"
                    f"Chi tiết: {err_msg}"
                )
            elif is_truncated:
                import re as _re
                m = _re.search(r"column '([^']+)'.*Truncated value: '([^']+)'", err_msg)
                if m:
                    col, val = m.group(1), m.group(2)
                    friendly = (
                        f"Dữ liệu quá dài cho cột '{col}'.\n"
                        f"Giá trị bị cắt bớt: '{val}' ({len(val)} ký tự)\n"
                        f"→ Liên hệ DBA mở rộng độ dài cột hoặc rút ngắn mã.\n\n"
                        f"Chi tiết: {err_msg}"
                    )
                else:
                    friendly = f"Dữ liệu quá dài cho cột trong DB.\nChi tiết: {err_msg}"
            else:
                friendly = err_msg
            self._log('', f'SP_HOOK_BATCH {sp_name}', 0, 'Error', friendly, 'error')
            raise RuntimeError(friendly) from e


    # ── Custom dialog helpers (thay thế tkinter messagebox) ──────────────────
    def _show_msg(self, title, msg, kind='info'):
        """Custom CTk dialog: kind = 'info' | 'warning' | 'error'."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        # Màu icon theo loại
        _colors = {'error': '#E05252', 'warning': '#E0A030', 'info': '#4A9ECC'}
        _icons  = {'error': '✕', 'warning': '⚠', 'info': 'ℹ'}
        _c = _colors.get(kind, _colors['info'])
        _i = _icons.get(kind, _icons['info'])
        # Icon circle
        icon_frm = ctk.CTkFrame(dlg, fg_color=_c, corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text=_i, font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont("Segoe UI", 13, "bold")).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont("Segoe UI", 11),
                     wraplength=320, justify="center").pack(pady=(0, 16), padx=24)
        ctk.CTkButton(dlg, text="OK", width=100,
                      command=dlg.destroy).pack(pady=(0, 20))
        dlg.update_idletasks()
        w, h = 380, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.wait_window()

    def _show_export_success(self, title, msg, path):
        """Dialog xuất thành công với nút Mở file và Đóng."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        icon_frm = ctk.CTkFrame(dlg, fg_color='#4A9ECC', corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text='✓', font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont("Segoe UI", 13, "bold")).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont("Segoe UI", 11),
                     wraplength=340, justify="center").pack(pady=(0, 16), padx=24)
        btn_frm = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frm.pack(pady=(0, 20))
        def _open():
            dlg.destroy()
            try:
                os.startfile(path)
            except Exception as _e:
                pass
        ctk.CTkButton(btn_frm, text="📂  Mở file", width=110, command=_open).pack(side="left", padx=8)
        ctk.CTkButton(btn_frm, text="Đóng", width=90, fg_color="gray40",
                      hover_color="gray30", command=dlg.destroy).pack(side="left", padx=8)
        dlg.update_idletasks()
        w, h = 400, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.wait_window()

    def _ask_msg(self, title, msg):
        """Custom CTk yes/no dialog. Trả về True nếu Yes."""
        _result = [False]
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        dlg.attributes('-topmost', True)
        dlg.lift()
        icon_frm = ctk.CTkFrame(dlg, fg_color='#4A9ECC', corner_radius=22, width=44, height=44)
        icon_frm.pack(pady=(22, 0))
        icon_frm.pack_propagate(False)
        ctk.CTkLabel(icon_frm, text='?', font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(dlg, text=title,
                     font=ctk.CTkFont("Segoe UI", 13, "bold")).pack(pady=(10, 2), padx=24)
        ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont("Segoe UI", 11),
                     wraplength=320, justify="center").pack(pady=(0, 16), padx=24)
        btn_frm = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frm.pack(pady=(0, 20))
        def _yes():
            _result[0] = True
            dlg.destroy()
        ctk.CTkButton(btn_frm, text="Có", width=90, command=_yes).pack(side="left", padx=8)
        ctk.CTkButton(btn_frm, text="Không", width=90, fg_color="gray40",
                      hover_color="gray30", command=dlg.destroy).pack(side="left", padx=8)
        dlg.update_idletasks()
        w, h = 380, dlg.winfo_reqheight() + 20
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.wait_window()
        return _result[0]

    def _make_loading_popup(self, msg="Đang import dữ liệu vào BRAVO...\nVui lòng không đóng cửa sổ.", grab=True):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Đang xử lý")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)   # chặn đóng cửa sổ
        lbl = ctk.CTkLabel(dlg, text=msg,
                     font=ctk.CTkFont("Segoe UI", 12),
                     wraplength=300, justify="center")
        lbl.pack(pady=(22, 10), padx=20)
        self._loading_lbl = lbl   # giữ ref để update từ background thread
        pb = ctk.CTkProgressBar(dlg, mode="indeterminate", width=300)
        pb.pack(padx=20, pady=(0, 22))
        pb.start()
        # Canh giữa màn hình
        dlg.update_idletasks()
        w, h = 340, 160
        sw = dlg.winfo_screenwidth()
        sh = dlg.winfo_screenheight()
        x  = (sw - w) // 2
        y  = (sh - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")
        if grab:
            dlg.grab_set()
        else:
            dlg.attributes('-topmost', True)
            dlg.lift()
            dlg.focus_force()
        dlg.update()
        return dlg

    def _update_loading_msg(self, msg):
        """Cập nhật text loading dialog (phải gọi từ main thread qua self.after)."""
        lbl = getattr(self, '_loading_lbl', None)
        if lbl:
            try:
                lbl.configure(text=msg)
            except Exception:
                pass

    # ── Import entry point (button command) ───────────────────────────────────
    def _start_import(self):
        """
        Chạy trên main thread:
          1. Pre-checks nhanh
          2. Kết nối DB, build row, resolve fields, SP, validators
          3. Nếu export_only → _run_export_sql (main thread, cần file dialog)
          4. Nếu INSERT thật → hiện loading popup + thread _run_insert_bg
        """
        import datetime
        self.mapping = load_mapping()
        fname = os.path.basename(self._current_file)

        # ── 1. Pre-checks ────────────────────────────────────────────────────
        if self.val_errors is None:
            self._show_msg("Chưa Validate", "Hãy bấm Validate trước khi Import.", 'warning')
            return
        header_errors = any(
            r.get("severity") == "error"
            for r in self.val_errors.get("[H] Phiếu Header", [])
        )
        if header_errors:
            self._show_msg(
                "Lỗi HEADER",
                "Dữ liệu Header còn lỗi. Kiểm tra lại trước khi Import."
            , 'error')
            return
        if not self._current_creator_code:
            self._show_msg(
                "Chưa chọn Nhân viên",
                "Vui lòng chọn Nhân viên trước khi Import.",
                'warning')
            return

        # ── 2. Kết nối DB ────────────────────────────────────────────────────
        self._set_status("⏳  Đang kết nối DB...", C["yellow"])
        self.update()   # force render status ngay lập tức
        try:
            conn = self._get_db_conn()
            self._set_status("🔗  Đã kết nối", C["green"])
        except ConnectionError as e:
            self._set_status("❌  Kết nối thất bại", C["red"])
            self._show_msg("Lỗi kết nối DB", str(e))
            self._log(fname, "Import", 0, "Lỗi kết nối", str(e), "error")
            return
        except Exception as e:
            self._set_status("❌  Lỗi DB", C["red"])
            self._show_msg("Lỗi DB", str(e))
            self._log(fname, "Import", 0, "Lỗi", str(e), "error")
            return

        conn.autocommit = False

        self._set_status("⏳  Đang phân tích dữ liệu...", C["yellow"])
        _scan_dlg = self._make_loading_popup(
            "⏳  Đang phân tích file...\nVui lòng chờ.", grab=False)

        self._fuzzy_collect_mode = True
        self._fuzzy_pending      = []
        self._fuzzy_ctx          = {}
        self._fuzzy_resolutions  = {}
        self._fuzzy_batch_done   = False
        self._ps_bom_caches      = {}

        def _prescan_worker():
            import math as _math, datetime as _dt
            try:
                _now       = _dt.datetime.now()
                _meta      = self.global_meta
                _norm_meta = {_norm_vn(k): v for k, v in _meta.items()}
                _hmap = self.mapping.get('HEADER', [])
                self._ps_header_caches = self._build_all_caches(conn, _hmap)
                for _r in _hmap:
                    if _r.get('nguon_dl') in ('SP', 'TinhToan', 'HeThong', 'CoDinh'):
                        continue
                    if _r.get('kieu_lookup', '') not in ('fuzzy_code', 'fuzzy_name'):
                        continue
                    _cache_key_h = (_r.get('bang_master',''), _r.get('dieu_kien_master',''),
                                    _r.get('truong_so_sanh',''), _r.get('truong_lay_ve',''))
                    _cache_h = self._ps_header_caches.get(_cache_key_h, [])
                    self._fuzzy_ctx = {
                        'section': 'HEADER', 'field': _r.get('sql_col', ''), 'row_idx': None}
                    self._resolve_header_field(
                        _r, conn, _meta, _norm_meta, self._ps_header_caches, _now)
                # Sections có parent_section=HEADER (BOM detail sections)
                _PS_SECT = {
                    s for s, cfg in self.mapping.get('_CONFIG', {}).items()
                    if cfg.get('parent_section') == 'HEADER'
                }
                for _lbl, _tbl in self.tables.items():
                    _sec = _tbl.get('type')
                    if _sec not in _PS_SECT:
                        continue
                    _df = _tbl.get('df')
                    if _df is None or _df.empty or 'Lỗi' in _df.columns:
                        continue
                    _all_recs = self.mapping.get(_sec, [])
                    _fuzzy_recs = [
                        r for r in _all_recs
                        if _nan_str(r.get('kieu_lookup', '')) in ('fuzzy_code', 'fuzzy_name')
                        and _nan_str(r.get('nguon_dl', '')) not in ('CoDinh', 'HeThong', 'SP', 'TinhToan')
                        and _nan_str(r.get('truong_so_sanh', ''))
                        and _nan_str(r.get('truong_lay_ve', ''))
                    ]
                    if not _fuzzy_recs:
                        continue
                    _dcaches = self._build_bom_detail_caches(conn, _all_recs)
                    self._ps_bom_caches[_sec] = _dcaches
                    _stt_col = next((c for c in _df.columns if _norm_vn(str(c)) == 'stt'), None)
                    _order = 0
                    for _, _drow in _df.iterrows():
                        if _drow.isna().all():
                            continue
                        _stt_v = str(_drow.get(_stt_col, '') or '').strip() if _stt_col else ''
                        if _stt_v and SECTION_STT_PATTERN.match(_stt_v):
                            continue
                        _order += 1
                        for _rec in _fuzzy_recs:
                            _ten    = _nan_str(_rec.get('ten_excel', ''))
                            _kl     = _nan_str(_rec.get('kieu_lookup', ''))
                            _bm     = _nan_str(_rec.get('bang_master', ''))
                            _dk     = _nan_str(_rec.get('dieu_kien_master', ''))
                            _ss     = _nan_str(_rec.get('truong_so_sanh', ''))
                            _lv     = _nan_str(_rec.get('truong_lay_ve', ''))
                            _scol   = _rec.get('sql_col', '')
                            _nguong = int(_rec.get('nguong_fuzzy', 0) or 0) or 92
                            _raw = None
                            if _ten:
                                _norm_ten = _norm_vn(_ten)
                                for _col in _drow.index:
                                    _cs = str(_col).strip()
                                    if _cs == _ten or _norm_vn(_cs) == _norm_ten:
                                        _v = _drow[_col]
                                        if _v is None or (isinstance(_v, float) and _math.isnan(_v)):
                                            break
                                        _raw = _v
                                        break
                            if _raw is None:
                                continue
                            _cache_key = (_bm, _dk, _ss, _lv)
                            _cache = _dcaches.get(_cache_key, [])
                            self._fuzzy_ctx = {
                                'section': _sec, 'field': _scol, 'row_idx': _order}
                            self._lookup_generic(_raw, _cache, _kl, _nguong, _cache_key=_cache_key)
            except Exception as _pe:
                import traceback as _tb
                _err_msg = _tb.format_exc()
                self.after(0, lambda m=_err_msg: self._log(
                    fname, 'Prescan', 0, 'Error', m[:500], 'error'))
            finally:
                self._fuzzy_collect_mode = False
            self.after(0, _after_prescan)

        def _after_prescan():
            import datetime as _dt
            try: _scan_dlg.destroy()
            except Exception: pass
            self._set_status("", None)

            if self._fuzzy_pending:
                _resolved = self._show_batch_fuzzy_dialog(self._fuzzy_pending)
                self._fuzzy_resolutions = _resolved if _resolved is not None else {}
            else:
                self._fuzzy_resolutions = {}
            self._fuzzy_batch_done   = True
            self._fuzzy_collect_mode = False   # cho phép import dùng _fuzzy_resolutions

            now       = _dt.datetime.now()
            meta      = self.global_meta
            norm_meta = {_norm_vn(k): v for k, v in meta.items()}
            header_map = self.mapping.get('HEADER', [])

            row, lookup_log = {}, {}
            for rec in header_map:
                if rec['nguon_dl'] == 'SP':
                    continue
                sql_col = rec['sql_col']
                val, mt = self._resolve_header_field(
                    rec, conn, meta, norm_meta, self._ps_header_caches, now, row_out=row)
                row[sql_col] = val
                if mt not in ('codinh', 'hethong', 'excel_direct', 'sp',
                              'tinhtoan', 'passthrough', 'fuzzy_pending'):
                    lookup_log[sql_col] = mt

            sp_cfgs = {
                s['sql_column']: s
                for s in self.mapping.get('SP_CONFIG', [])
                if s.get('isactive', '') == '1'
                and s.get('section', '') in ('', 'HEADER')
            }
            for rec in header_map:
                if rec['nguon_dl'] != 'SP':
                    continue
                sql_col = rec['sql_col']
                sp_cfg  = sp_cfgs.get(sql_col)
                if not sp_cfg:
                    row.setdefault(sql_col, None)
                    continue
                try:
                    sp_name  = sp_cfg.get('sp_name', '').strip()
                    params_s = sp_cfg.get('params', '').strip()
                    fallback = sp_cfg.get('fallback', '').strip() or None
                    if sp_name.lower() == 'lookup':
                        result_sp = self._sp_lookup(conn, params_s, row)
                    else:
                        result_sp = self._call_sp(conn, sp_name, params_s, row)
                    row[sql_col] = result_sp if result_sp is not None else fallback
                except Exception as e:
                    self._log(fname, f'SP {sql_col}', 0, 'Error', str(e), 'error')
                    self._show_msg(f'SP Error — {sql_col}',
                                           f'Lỗi khi gọi SP cho [{sql_col}]:\n{e}', 'warning')
                    row[sql_col] = sp_cfg.get('fallback') or None

            _null_required = [
                rec for rec in header_map
                if rec.get('bat_buoc', '').rstrip('0').rstrip('.') == '1'
                and rec.get('nguon_dl') != 'SP'
                and row.get(rec['sql_col']) is None
            ]
            if _null_required:
                _names = ', '.join(
                    f"[{r['sql_col']}]" + (f" ({r['ten_excel']})" if r.get('ten_excel') else '')
                    for r in _null_required
                )
                import tkinter.messagebox as _mb
                _go = _mb.askyesno(
                    "Thiếu dữ liệu bắt buộc",
                    f"Các trường sau bắt buộc nhưng không lấy được giá trị "
                    f"(lookup thất bại hoặc để trống):\n\n{_names}\n\n"
                    f"Tiếp tục import với các trường này để NULL không?",
                    icon="warning"
                )
                if not _go:
                    conn.autocommit = True
                    conn.close()
                    return

            table_name = (self.db_cfg.get('table_name', 'B20BOM') if self.db_cfg else 'B20BOM')
            _sp_temp_fields = {
                s['sql_column'] for s in self.mapping.get('SP_CONFIG', [])
                if s.get('sp_name', '').lower() == 'lookup'
                and s.get('section', '') in ('', 'HEADER')
            }
            cols     = [c for c in row.keys() if c not in _sp_temp_fields]
            sql_stmt = (
                "INSERT INTO " + table_name + " ("
                + ", ".join("[" + c + "]" for c in cols)
                + ") OUTPUT INSERTED.Id VALUES ("
                + ", ".join(["?"] * len(cols)) + ")"
            )
            export_only = getattr(self, "var_export_sql", None) and self.var_export_sql.get()
            if not export_only:
                if not self._run_validators(conn, row):
                    conn.autocommit = True
                    conn.close()
                    return
            if export_only:
                _eo_args = (conn, row, cols, sql_stmt, header_map, fname, now)
                self.after(0, lambda a=_eo_args: self._run_export_sql(*a))
                return
            self._loading_dlg = self._make_loading_popup()
            ctx = {
                'conn': conn, 'row': row, 'cols': cols, 'sql': sql_stmt,
                'lookup_log': lookup_log, 'fname': fname, 'now': now,
                'table_name': table_name, 'header_map': header_map,
            }
            threading.Thread(target=self._run_insert_bg, args=(ctx,), daemon=True).start()

        threading.Thread(target=_prescan_worker, daemon=True).start()

    def _run_insert_bg(self, ctx):
        """
        Background thread — INSERT B20BOM + B20BOMDetail trong 1 transaction.
        Không gọi bất kỳ UI nào trực tiếp; kết quả post qua self.after().
        """
        conn       = ctx['conn']
        row        = ctx['row']
        cols       = ctx['cols']
        sql        = ctx['sql']
        fname      = ctx['fname']
        now        = ctx['now']
        table_name = ctx['table_name']
        lookup_log = ctx['lookup_log']
        try:
            cur = conn.cursor()
            cur.execute(sql, [row[c] for c in cols])
            new_id = cur.fetchone()[0]
            _detail_count = self._generate_bom_details(
                conn, new_id, row, self.tables, self.mapping, now, export_only=False
            )
            conn.commit()
            result = {
                'ok': True,
                'new_id': new_id,
                'detail_count': _detail_count,
                'row': row,
                'fname': fname,
                'lookup_log': lookup_log,
                'table_name': table_name,
            }
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            result = {'ok': False, 'error': str(e), 'fname': fname}
        finally:
            try:
                conn.autocommit = True
                conn.close()
            except Exception:
                pass
        self.after(0, lambda: self._finish_import(result))

    def _finish_import(self, result):
        try:
            if self._loading_dlg:
                self._loading_dlg.destroy()
                self._loading_dlg = None
        except Exception:
            pass
        fname = result.get('fname', '')
        if result.get('ok'):
            new_id = result['new_id']
            n_det  = result['detail_count']
            self._set_status(f"\u2705  Import thành công — BOM ID {new_id}, {n_det} dòng chi tiết", C["green"])
            self._log(fname, "Import", str(n_det), "OK", f"BOM ID={new_id}, {n_det} dòng chi tiết", "ok")
            self._db_log('BOM', fname, new_id, n_det, 'OK',
                         f"BOM ID={new_id}, {n_det} dòng chi tiết")
            self._last_import_bom_id = new_id
            self._last_import_fname  = fname
            self.btn_undo_import.configure(state="normal")
            self._show_msg("Import thành công", f"BOM đã được tạo thành công!\nBOM ID: {new_id}\nSố dòng chi tiết: {n_det:,}", kind="info")
        else:
            err = result.get('error', 'Unknown error')
            self._set_status("\u274c  Import thất bại", C["red"])
            self._log(fname, "Import", "0", "Lỗi", err, "error")
            self._db_log('BOM', fname, '', 0, 'LOI', err)
            self._show_msg("Import thất bại", f"Lỗi: {err}", kind="error")


    # ── Hoàn tác import (gọi SP usp_BOMTool_DeleteBOM — khách hàng deploy) ────
    def _undo_last_import(self):
        bom_id = self._last_import_bom_id
        if bom_id is None:
            return
        if not self._ask_msg("Hoàn tác import",
                f"Xóa BOM vừa import khỏi BRAVO?\n\n"
                f"BOM ID: {bom_id}\n"
                f"File: {self._last_import_fname}\n\n"
                f"Thao tác này KHÔNG thể tự hoàn tác lại."):
            return
        self.btn_undo_import.configure(state="disabled", text="⏳  Đang xóa...")
        threading.Thread(target=self._undo_worker, args=(bom_id,),
                         daemon=True).start()

    def _undo_worker(self, bom_id):
        conn = None
        try:
            conn = self._get_db_conn()
            cur  = conn.cursor()
            cur.execute("EXEC dbo.usp_BOMTool_DeleteBOM ?", (bom_id,))
            row   = cur.fetchone()
            n_det = row[1] if row else 0
            conn.commit()
            self.after(0, lambda n=n_det: self._undo_done(bom_id, n, None))
        except Exception as e:
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
            self.after(0, lambda err=str(e): self._undo_done(bom_id, 0, err))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _undo_done(self, bom_id, n_det, error):
        self.btn_undo_import.configure(text="↩  Hoàn tác import")
        if error:
            self.btn_undo_import.configure(state="normal")
            if 'Could not find stored procedure' in error \
                    or 'usp_BOMTool_DeleteBOM' in error:
                self._show_msg("Chưa có SP trên DB",
                    "Stored procedure usp_BOMTool_DeleteBOM chưa được tạo "
                    "trên database.\n\n"
                    "Gửi file sql\\create_usp_bomtool_deletebom.sql cho khách hàng "
                    "chạy, sau đó bấm lại nút này.", kind="info")
            else:
                self._show_msg("Lỗi hoàn tác", error, kind="error")
            return
        self._last_import_bom_id = None
        self.btn_undo_import.configure(state="disabled")
        self._set_status(f"↩  Đã xóa BOM ID {bom_id} ({n_det} dòng chi tiết)",
                         C["green"])
        self._log(self._last_import_fname, "Undo", str(n_det), "OK",
                  f"Đã xóa BOM ID={bom_id}", "warn")
        self._db_log('UNDO', self._last_import_fname, bom_id, n_det, 'OK',
                     f"Đã xóa BOM ID={bom_id} ({n_det} dòng chi tiết)")
        self._show_msg("Đã hoàn tác",
            f"BOM ID {bom_id} và {n_det:,} dòng detail đã được xóa khỏi BRAVO.",
            kind="info")


if __name__ == "__main__":
    app = BOMToolApp()
    app.mainloop()
