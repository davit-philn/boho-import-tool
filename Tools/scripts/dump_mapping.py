"""
Chạy script này để dump toàn bộ CK_Mapping_v2.xlsx ra console.
    python Tools/dump_mapping.py
"""
import os, json
import openpyxl

MAPPING_FILE = os.path.join(os.path.dirname(__file__), 'config', 'CK_Mapping_v2.xlsx')

SECTION_COLS = ['STT','SQL_Column','Ten_Excel','Kieu_DL','Do_dai',
                'Bat_buoc','Mac_dinh','Nguon_DL','Bang_Master',
                'Dieu_kien_Master','Kieu_Lookup','Truong_So_Sanh','Truong_Lay_Ve','Ghi_chu']
SP_COLS   = ['SQL_Column','SP_Name','Params','Fallback','IsActive']
VAL_COLS  = ['ValidatorName','SQL','Params','WarningMessage','IsActive']

def read_sheet_rows(ws, header_row, col_names):
    rows = []
    for r in ws.iter_rows(min_row=header_row+1, values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in r]
        if not any(vals):
            continue
        n = min(len(vals), len(col_names))
        row = {col_names[i]: vals[i] for i in range(n)}
        if any(v for v in row.values()):
            rows.append(row)
    return rows

wb = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

SECTION_SHEETS = ['HEADER','BOM2','BOM3','BOM4']
result = {}

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"{'='*70}")
    print(f"Sheet: [{sn}]")
    print('='*70)

    if sn in SECTION_SHEETS:
        rows = read_sheet_rows(ws, 3, SECTION_COLS)
        for r in rows:
            flag = ''
            if r.get('Nguon_DL') == 'CoDinh' and not r.get('Mac_dinh'):
                flag = '  ⚠ CoDinh nhưng Mac_dinh TRỐNG'
            if r.get('Bat_buoc') == '1' and r.get('Nguon_DL') == 'Excel' and not r.get('Ten_Excel'):
                flag += '  ⚠ Bat_buoc nhưng Ten_Excel TRỐNG'
            print(f"  {r.get('SQL_Column',''):20s} | nguon={r.get('Nguon_DL',''):8s} | "
                  f"mac_dinh={r.get('Mac_dinh',''):10s} | ten_excel={r.get('Ten_Excel',''):15s} | "
                  f"kieu_dl={r.get('Kieu_DL',''):8s} | bat_buoc={r.get('Bat_buoc','')} | "
                  f"kl={r.get('Kieu_Lookup',''):12s}{flag}")

    elif sn == 'SP_CONFIG':
        rows = read_sheet_rows(ws, 2, SP_COLS)
        for r in rows:
            print(f"  {r.get('SQL_Column',''):10s} | sp={r.get('SP_Name',''):40s} | "
                  f"fallback={r.get('Fallback',''):5s} | isactive={r.get('IsActive','')}")
            print(f"    params: {r.get('Params','')}")

    elif sn == 'VALIDATORS':
        rows = read_sheet_rows(ws, 2, VAL_COLS)
        for r in rows:
            print(f"  {r.get('ValidatorName',''):20s} | isactive={r.get('IsActive','')} | "
                  f"params={r.get('Params','')}")

    else:
        print("  (sheet khác — bỏ qua)")

print("\nDone.")
