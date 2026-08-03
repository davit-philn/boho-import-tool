"""
patch_mapping.py — chạy 1 lần để:
  1. Thêm dòng Ws_Id (SP/lookup) vào HEADER trước dòng RowId
  2. Đổi Fallback số → text trong SP_CONFIG (tránh '1.0')
  3. Migrate IsActive: CÓ/KHÔNG → 1/0 trên toàn bộ SP_CONFIG + VALIDATORS
  4. Migrate Bat_buoc: CÓ/KHÔNG → 1/0 trên toàn bộ HEADER + BOM sheets

Chạy (đóng Excel trước):
    python Tools/patch_mapping.py
"""
import os
import sys
import unicodedata
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from copy import copy

MAPPING_FILE = os.path.join(os.path.dirname(__file__), 'config', 'CK_Mapping_v2.xlsx')

CO_VALUES  = {'CÓ', 'CO', 'YES', 'TRUE', 'X', '1', 'CÓ'}  # đều → '1'
KHONG_VALUES = {'KHÔNG', 'KHONG', 'NO', 'FALSE', '0', ''}   # đều → '0'

def _nfc(s):
    return unicodedata.normalize('NFC', str(s or '')).strip().upper()

def _to_01(val):
    """Đổi CÓ/KHÔNG (và mọi dạng) về '1' hoặc '0'."""
    n = _nfc(val)
    if n in {_nfc(v) for v in CO_VALUES}:
        return '1'
    return '0'

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def clone_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── 1. HEADER: thêm dòng Ws_Id trước RowId ───────────────────────────────────
def patch_header_add_wsid(ws):
    rowid_row = None
    for row in ws.iter_rows(min_row=4):
        if _nfc(row[1].value) == 'ROWID':
            rowid_row = row[0].row
            break

    if rowid_row is None:
        print("  [HEADER] Không tìm thấy dòng RowId, bỏ qua")
        return

    for row in ws.iter_rows(min_row=4):
        if _nfc(row[1].value) == 'WS_ID':
            print("  [HEADER] Ws_Id đã có, bỏ qua")
            return

    ws.insert_rows(rowid_row)

    ref_row = rowid_row + 1
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=ref_row, column=col)
        dst = ws.cell(row=rowid_row, column=col)
        clone_style(src, dst)
        dst.value = None

    col_map = {_nfc(cell.value): cell.column
               for cell in ws[3] if cell.value}

    def set_cell(col_name, value):
        c = col_map.get(_nfc(col_name))
        if c:
            ws.cell(row=rowid_row, column=c).value = value

    set_cell('SQL_Column',  'Ws_Id')
    set_cell('Kieu_DL',     'int')
    set_cell('Nguon_DL',    'SP')
    set_cell('Kieu_Lookup', 'lookup')

    print(f"  [HEADER] Thêm dòng Ws_Id tại row {rowid_row}")


# ── 2. SP_CONFIG: Fallback số → text, IsActive → 1/0 ─────────────────────────
def patch_sp_config(ws):
    # Tìm cột từ row 2
    col_map = {_nfc(cell.value): cell.column
               for cell in ws[2] if cell.value}
    fallback_col  = col_map.get('FALLBACK')
    isactive_col  = col_map.get('ISACTIVE')

    fixed_fb = fixed_ia = 0
    for row in ws.iter_rows(min_row=3):
        if fallback_col:
            cell = row[fallback_col - 1]
            if isinstance(cell.value, (int, float)):
                cell.value = str(int(cell.value))
                cell.number_format = '@'
                fixed_fb += 1
        if isactive_col:
            cell = row[isactive_col - 1]
            if cell.value is not None:
                new_val = _to_01(cell.value)
                if str(cell.value) != new_val:
                    cell.value = new_val
                    cell.number_format = '@'
                    fixed_ia += 1

    print(f"  [SP_CONFIG] Fallback: {fixed_fb} ô, IsActive: {fixed_ia} ô đã fix")


# ── 3. VALIDATORS: IsActive → 1/0 ────────────────────────────────────────────
def patch_validators(ws):
    col_map = {_nfc(cell.value): cell.column
               for cell in ws[2] if cell.value}
    isactive_col = col_map.get('ISACTIVE')
    if not isactive_col:
        print("  [VALIDATORS] Không tìm thấy cột IsActive")
        return

    fixed = 0
    for row in ws.iter_rows(min_row=3):
        cell = row[isactive_col - 1]
        if cell.value is not None:
            new_val = _to_01(cell.value)
            if str(cell.value) != new_val:
                cell.value = new_val
                cell.number_format = '@'
                fixed += 1
    print(f"  [VALIDATORS] IsActive: {fixed} ô đã fix")


# ── 4. HEADER + BOM sheets: Bat_buoc → 1/0 ───────────────────────────────────
def patch_batbuoc(ws):
    col_map = {_nfc(cell.value): cell.column
               for cell in ws[3] if cell.value}
    batbuoc_col = col_map.get('BAT_BUOC')
    if not batbuoc_col:
        print(f"  [{ws.title}] Không tìm thấy cột Bat_buoc")
        return

    fixed = 0
    for row in ws.iter_rows(min_row=4):
        cell = row[batbuoc_col - 1]
        if cell.value is not None and str(cell.value).strip():
            new_val = _to_01(cell.value)
            if str(cell.value) != new_val:
                cell.value = new_val
                cell.number_format = '@'
                fixed += 1
    print(f"  [{ws.title}] Bat_buoc: {fixed} ô đã fix")


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if not os.path.exists(MAPPING_FILE):
        print(f"Không tìm thấy: {MAPPING_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(MAPPING_FILE)

    print("=== 1. HEADER: thêm Ws_Id ===")
    if 'HEADER' in wb.sheetnames:
        patch_header_add_wsid(wb['HEADER'])
    else:
        print("  Sheet HEADER không tồn tại")

    print("=== 2. SP_CONFIG: fix Fallback + IsActive ===")
    if 'SP_CONFIG' in wb.sheetnames:
        patch_sp_config(wb['SP_CONFIG'])
    else:
        print("  Sheet SP_CONFIG không tồn tại")

    print("=== 3. VALIDATORS: fix IsActive ===")
    if 'VALIDATORS' in wb.sheetnames:
        patch_validators(wb['VALIDATORS'])
    else:
        print("  Sheet VALIDATORS không tồn tại")

    print("=== 4. Bat_buoc → 1/0 trên tất cả section sheets ===")
    for sn in ['HEADER', 'BOM2', 'BOM3', 'BOM4']:
        if sn in wb.sheetnames:
            patch_batbuoc(wb[sn])

    wb.save(MAPPING_FILE)
    print(f"\nĐã lưu: {MAPPING_FILE}")
    print("Kiểm tra lại trong Excel trước khi chạy tool.")
