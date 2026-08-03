"""
Script tạo/cập nhật CK_Mapping_v2.xlsx:
  - Thêm cột Truong_So_Sanh + Truong_Lay_Ve vào HEADER (nếu chưa có)
  - Thêm sheet SP_CONFIG (nếu chưa có)
  - Thêm sheet VALIDATORS (nếu chưa có)

Chạy 1 lần sau khi đóng Excel:
    python Tools/update_mapping_cols.py
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

MAPPING_FILE = os.path.join(os.path.dirname(__file__), 'config', 'CK_Mapping_v2.xlsx')

# Giá trị Truong_So_Sanh / Truong_Lay_Ve cho HEADER
HEADER_LOOKUP = {
    'TypeB':              ('Code',     'Code'),
    'ProductId1':         ('Name',     'Id'),
    'ParentBizDocId':     ('BizDocId', 'BizDocId'),
    'ItemId0':            ('Code',     'Id'),
    'SubcontractingCode': ('Code',     'Code'),
    'Ws_Id':              ('BranchCode', 'Ws_Id'),
}
BOM_LOOKUP = {}
INSERT_AT = 12   # chèn trước cột L (Ghi_chu)


# ── Helpers ────────────────────────────────────────────────────────────────────
def clone_style(src, dst):
    if src.has_style:
        dst.font      = copy(src.font)
        dst.fill      = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.border    = copy(src.border)
        dst.number_format = src.number_format


def header_fill():
    return PatternFill(fill_type='solid', fgColor='4472C4')


def header_font():
    return Font(bold=True, color='FFFFFF', size=10)


def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def write_header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = header_fill()
    c.font   = header_font()
    c.border = thin_border()
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return c


def write_data_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.border = thin_border()
    c.alignment = Alignment(vertical='center', wrap_text=True)
    return c


# ── Bước 1: Thêm Truong_So_Sanh / Truong_Lay_Ve vào section sheets ───────────
def process_section_sheet(ws, lookup_map):
    if ws.cell(row=3, column=INSERT_AT).value == 'Truong_So_Sanh':
        print(f'  [{ws.title}] cột đã có, bỏ qua')
        return

    ws.insert_cols(INSERT_AT, 2)

    c2a = ws.cell(row=2, column=INSERT_AT);     c2a.value = 'Trường so sánh'
    c2b = ws.cell(row=2, column=INSERT_AT + 1); c2b.value = 'Trường lấy về'
    c3a = ws.cell(row=3, column=INSERT_AT);     c3a.value = 'Truong_So_Sanh'
    c3b = ws.cell(row=3, column=INSERT_AT + 1); c3b.value = 'Truong_Lay_Ve'

    for row_h in (2, 3):
        src = ws.cell(row=row_h, column=INSERT_AT - 1)
        clone_style(src, ws.cell(row=row_h, column=INSERT_AT))
        clone_style(src, ws.cell(row=row_h, column=INSERT_AT + 1))

    for row_i in range(4, ws.max_row + 1):
        sql_col = ws.cell(row=row_i, column=2).value
        if not sql_col:
            continue
        ss, lv = lookup_map.get(str(sql_col).strip(), ('', ''))
        ws.cell(row=row_i, column=INSERT_AT    ).value = ss
        ws.cell(row=row_i, column=INSERT_AT + 1).value = lv

    col_l = get_column_letter(INSERT_AT)
    col_m = get_column_letter(INSERT_AT + 1)
    ws.column_dimensions[col_l].width = 16
    ws.column_dimensions[col_m].width = 14
    print(f'  [{ws.title}] Truong_So_Sanh/Truong_Lay_Ve OK')


# ── Bước 2: Tạo sheet SP_CONFIG ───────────────────────────────────────────────
SP_HEADERS_VN = ['SQL Column', 'Tên SP', 'Params', 'Fallback', 'Kích hoạt']
SP_HEADERS_EN = ['SQL_Column', 'SP_Name', 'Params', 'Fallback', 'IsActive']
SP_COL_WIDTHS = [16, 30, 60, 20, 10]

SP_DATA = [
    # SQL_Column, SP_Name, Params, Fallback, IsActive
    ['Ws_Id',   'lookup',
     'Bang=B00Branch|Where=BranchCode={BranchCode}|Lay=Ws_Id',
     '1', '1'],
    ['RowId',   'B10_Boho.dbo.usp_sys_CreateSttBySeq',
     '@_TableName=B20BOM|@_Ws_Id={Ws_Id}|@_Ext=BOM',
     '', '1'],
    ['Version', 'B10_Boho.dbo.usp_B20BOM_AutoVersion',
     '@_ItemId0={ItemId0}|@_ParentBizDocId={ParentBizDocId}|@_TypeB={TypeB}',
     '1', '1'],
    ['Code',    'usp_sys_AutoNewCode',
     '@Code={Code}|@TableName=B20BOM|@ColumnName=Code',
     '', '1'],
]

def create_sp_config_sheet(wb):
    if 'SP_CONFIG' in wb.sheetnames:
        print('  [SP_CONFIG] đã có, bỏ qua')
        return

    ws = wb.create_sheet('SP_CONFIG')
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # Row 1: VN headers
    for col, h in enumerate(SP_HEADERS_VN, 1):
        write_header_cell(ws, 1, col, h)
    # Row 2: EN keys
    for col, h in enumerate(SP_HEADERS_EN, 1):
        write_header_cell(ws, 2, col, h)
    # Data rows
    for r_i, row_data in enumerate(SP_DATA, 3):
        for col, val in enumerate(row_data, 1):
            write_data_cell(ws, r_i, col, val)
    # Column widths
    for col, w in enumerate(SP_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'
    print('  [SP_CONFIG] tạo mới OK')


# ── Bước 3: Tạo sheet VALIDATORS ─────────────────────────────────────────────
VAL_HEADERS_VN = ['Tên validator', 'SQL kiểm tra', 'Params (từ row)', 'Nội dung cảnh báo', 'Kích hoạt']
VAL_HEADERS_EN = ['ValidatorName', 'SQL',          'Params',          'WarningMessage',     'IsActive']
VAL_COL_WIDTHS = [20, 80, 35, 50, 10]

VAL_DATA = [
    ['CheckUniqueBOM',
     "SELECT TOP 5 Code FROM B20BOM WHERE ItemId0=? AND Version=? AND EffectiveDate=? AND IsActive=1",
     'ItemId0,Version,EffectiveDate',
     'BOM đã tồn tại với cùng ItemId0/Version/EffectiveDate: {result}',
     '1'],
    ['CodeUnique',
     "SELECT COUNT(1) FROM B20BOM WHERE Code=?",
     'Code',
     'Code "{Code}" đã tồn tại trong B20BOM',
     '1'],
]

def create_validators_sheet(wb):
    if 'VALIDATORS' in wb.sheetnames:
        print('  [VALIDATORS] đã có, bỏ qua')
        return

    ws = wb.create_sheet('VALIDATORS')
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    for col, h in enumerate(VAL_HEADERS_VN, 1):
        write_header_cell(ws, 1, col, h)
    for col, h in enumerate(VAL_HEADERS_EN, 1):
        write_header_cell(ws, 2, col, h)
    for r_i, row_data in enumerate(VAL_DATA, 3):
        for col, val in enumerate(row_data, 1):
            c = write_data_cell(ws, r_i, col, val)
            if col == 2:  # SQL column: wrap
                c.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[r_i].height = 45
    for col, w in enumerate(VAL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'
    print('  [VALIDATORS] tạo mới OK')


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if not os.path.exists(MAPPING_FILE):
        print(f'Không tìm thấy: {MAPPING_FILE}')
        sys.exit(1)

    wb = openpyxl.load_workbook(MAPPING_FILE)

    SECTION_SHEETS = ['HEADER', 'BOM2', 'BOM3', 'BOM4',
                      'THDM_HEADER', 'THDM_THVT', 'THDM_BOM2', 'THDM_BOM3']

    print('=== Bước 1: Thêm cột Truong_So_Sanh/Truong_Lay_Ve ===')
    for sn in wb.sheetnames:
        if sn in SECTION_SHEETS:
            lmap = HEADER_LOOKUP if sn == 'HEADER' else BOM_LOOKUP
            process_section_sheet(wb[sn], lmap)

    print('=== Bước 2: Tạo sheet SP_CONFIG ===')
    create_sp_config_sheet(wb)

    print('=== Bước 3: Tạo sheet VALIDATORS ===')
    create_validators_sheet(wb)

    wb.save(MAPPING_FILE)
    print(f'\nĐã lưu: {MAPPING_FILE}')
    print('Kiểm tra lại trong Excel trước khi chạy tool.')
